"""
期权趋势排名分析交易系统 by playbonze
基于期权的隐含波动率、希腊字母、成交量和资金流向进行排名分析
自动选择期权品种进行交易
"""

import asyncio
import logging
import requests
import json
import pandas as pd
import numpy as np
import math
import decimal
import random
import datetime
import time
from datetime import datetime as dt
import sys
import os
import re
import pytz
import threading
import matplotlib.pyplot as plt
import socket
import warnings
import traceback
from tqsdk import TqApi, TqAuth, TqAccount, TqSim, TqBacktest, TqKq, BacktestFinished, TqChan, TargetPosTask, TqNotify
from tqsdk.ta import ATR, BOLL, MACD, OPTION_GREEKS, OPTION_IMPV, BS_VALUE, OPTION_VALUE
from tqsdk.tafunc import ma, ema, ema2, sma, time_to_datetime, crossup, crossdown, time_to_str, get_his_volatility, get_delta, get_gamma, get_vega, get_theta, get_t
from typing import List, Callable
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# 添加外部模块路径以支持 UnifiedLogger
sys.path.append(os.path.join(os.path.dirname(__file__), "wisecoin-catboost"))
from pb_quant_seektop_common import UnifiedLogger

# 设置统一日志
logger = UnifiedLogger.setup_logger_auto(__file__)

# 为了兼容现有代码，创建别名
logger_print = logger

# 全局文件配置 - 期权版本
SYMBOL_EXCEL_FILE = "wisecoin-期权品种.xlsx"

# 期权筛选参数
OPTION_FILTER_CONFIG = {
    'exclude_exchanges': ['SSE', 'SZSE'],  # 排除的交易所（股票期权）
    'max_quote_num': 99999,                  # 需要获取的期权行情个数（默认100，设置较大值如99999则获取全部）
}

async def get_all_option_symbols():
    """获取所有期权合约"""
    try:
        # 获取所有期权合约
        option_list = await api.query_quotes(ins_class='OPTION', expired=False)
        
        # 过滤掉股票期权（只保留商品期权和股指期权）
        filtered_options = []
        for symbol in option_list:
            # 排除上交所和深交所的股票期权
            exchange = symbol.split('.')[0]
            if exchange not in OPTION_FILTER_CONFIG['exclude_exchanges']:
                filtered_options.append(symbol)
        
        filtered_options = sorted(filtered_options)
        logger.info(f"获取到{len(filtered_options)}个期权合约（已过滤股票期权）")
        
        # 获取详细的合约信息
        logger.info(f"正在获取 {len(filtered_options)} 个合约的详细信息...")
        
        # 分批获取合约信息，避免超时
        batch_size = 500
        all_symbol_info = []
        for i in range(0, len(filtered_options), batch_size):
            batch = filtered_options[i:i + batch_size]
            try:
                batch_df = await asyncio.wait_for(api.query_symbol_info(batch), 60)
                if not batch_df.empty:
                    all_symbol_info.append(batch_df)
                logger.info(f"已获取 {min(i + batch_size, len(filtered_options))}/{len(filtered_options)} 个合约信息")
            except Exception as e:
                logger.warning(f"获取批次 {i}-{i+batch_size} 失败: {e}")
        
        if not all_symbol_info:
            logger.error("未能获取到任何合约信息")
            return filtered_options
            
        symbol_info_df = pd.concat(all_symbol_info)
        
        # 按产品分组并保存到 Excel
        if not symbol_info_df.empty:
            logger.info(f"正在按产品分类导出到 {SYMBOL_EXCEL_FILE}...")
            
            # 使用 ExcelWriter 保存多个工作表
            with pd.ExcelWriter(SYMBOL_EXCEL_FILE, engine='openpyxl') as writer:
                # 确定哪个列包含标的信息，优先尝试 underlying_symbol
                mapping_col = 'underlying_symbol' if 'underlying_symbol' in symbol_info_df.columns else 'symbol'
                
                # 增加 'product' 列用于分组
                def get_product(row):
                    symbol = row[mapping_col]
                    if pd.isna(symbol):
                        # 如果没有 underlying_symbol，尝试从自身 symbol 提取
                        symbol = row.get('symbol', 'Unknown')
                        
                    # 格式如 SHFE.cu2401 -> SHFE.cu
                    if '.' in str(symbol):
                        parts = str(symbol).split('.')
                        exchange = parts[0]
                        # 移除数字和后续字符提取品种码
                        code_match = re.match(r'^[a-zA-Z]+', parts[1])
                        if code_match:
                            code = code_match.group(0)
                            return f"{exchange}.{code}"
                    return str(symbol)

                symbol_info_df['product'] = symbol_info_df.apply(get_product, axis=1)
                
                # 按 product 分组
                grouped = symbol_info_df.groupby('product')
                
                # 统计已导出的标的数量
                underlying_count = 0
                
                for product, group in grouped:
                    # Excel 工作表名称限制为 31 个字符
                    sheet_name = str(product).replace('.', '_')[-31:]
                    
                    # 按照行权价和标的排序
                    sort_cols = []
                    if 'underlying_symbol' in group.columns:
                        sort_cols.append('underlying_symbol')
                    if 'strike_price' in group.columns:
                        sort_cols.append('strike_price')
                    
                    if sort_cols:
                        group = group.sort_values(sort_cols)
                        
                    # 写入 Excel
                    group.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 优化格式：自动调整列宽
                    ws = writer.sheets[sheet_name]
                    for idx, col in enumerate(group.columns):
                        max_len = max(
                            group[col].astype(str).map(len).max() if not group[col].empty else 0,
                            len(str(col))
                        ) + 2
                        column_letter = get_column_letter(idx + 1)
                        ws.column_dimensions[column_letter].width = min(max_len, 50)
                    
                    underlying_count += 1
            
            logger.info(f"成功将 {len(symbol_info_df)} 个期权合约按 {underlying_count} 个产品分类导出到 {SYMBOL_EXCEL_FILE}")
        
        return filtered_options
    except Exception as e:
        logger.error(f"获取期权品种列表失败：{e}")
        import traceback
        logger.error(traceback.format_exc())
        return []


async def get_option_quotes_from_excel():
    """
    从 Excel 中读取期权品种并获取实时行情
    支持“断点续传”：识别 wisecoin-期权品种 有的合约，但行情 Excel 中尚未获取数据的部分
    """
    try:
        if not os.path.exists(SYMBOL_EXCEL_FILE):
            logger.error(f"未找到期权品种文件: {SYMBOL_EXCEL_FILE}")
            return
            
        logger.info(f"正在从 {SYMBOL_EXCEL_FILE} 读取期权品种并获取行情...")
        
        # 读取 Excel 的所有工作表
        xls = pd.ExcelFile(SYMBOL_EXCEL_FILE)
        QUOTE_EXCEL_FILE = "wisecoin-期权行情.xlsx"
        
        # 1. 收集所有需要获取行情的 symbols
        all_symbols_to_fetch = []
        sheet_df_map = {}
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            sheet_df_map[sheet_name] = df
            symbol_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
            if symbol_col not in df.columns:
                continue
            symbols = df[symbol_col].dropna().tolist()
            all_symbols_to_fetch.extend(symbols)
        
        if not all_symbols_to_fetch:
            logger.error("Excel 中没有找到任何期权合约")
            return
        
        # 去重并排序
        unique_symbols = sorted(list(set(all_symbols_to_fetch)))
        
        # 应用获取个数限制
        max_num = OPTION_FILTER_CONFIG.get('max_quote_num', 100)
        if len(unique_symbols) > max_num:
            logger.info(f"期权合约总数 {len(unique_symbols)} 超过限制 {max_num}，仅获取前 {max_num} 个")
            unique_symbols = unique_symbols[:max_num]
            
        # ---------------- 断点续传逻辑 ----------------
        all_quote_data = {}
        already_fetched_symbols = set()
        
        if os.path.exists(QUOTE_EXCEL_FILE):
            try:
                logger.info(f"检测到已存在的行情文件 {QUOTE_EXCEL_FILE}，尝试加载断点...")
                existing_xls = pd.ExcelFile(QUOTE_EXCEL_FILE)
                for s_name in existing_xls.sheet_names:
                    if s_name == "Summary" or s_name == "Progress": continue
                    existing_df = pd.read_excel(existing_xls, sheet_name=s_name)
                    if existing_df.empty: continue
                    
                    # 确定 symbol 列
                    s_col = 'symbol' if 'symbol' in existing_df.columns else ('instrument_id' if 'instrument_id' in existing_df.columns else None)
                    if not s_col: continue
                    
                    # 将已有数据加载到 all_quote_data
                    for _, row in existing_df.iterrows():
                        sym = row[s_col]
                        # 只有当 last_price 有效或者是之前同步过的才视为已获取
                        if not pd.isna(row.get('last_price')):
                            all_quote_data[sym] = row.to_dict()
                            already_fetched_symbols.add(sym)
                
                logger.info(f"成功恢复断点：已获取 {len(already_fetched_symbols)} 个合约，剩余 {max(0, len(unique_symbols) - len(already_fetched_symbols))} 个待同步")
            except Exception as ex:
                logger.warning(f"加载断点文件失败，将从头开始获取: {ex}")
        
        # 过滤掉已经获取过的 symbols
        symbols_to_fetch_now = [s for s in unique_symbols if s not in already_fetched_symbols]
        
        if not symbols_to_fetch_now:
            logger.info("🎉 所有计划内的期权行情已全部获取完成，无需续传。")
            if all_quote_data:
                pass
            return

        logger.info(f"共需同步 {len(unique_symbols)} 个期权合约，当前实际待同步 {len(symbols_to_fetch_now)} 个")
        
        # 2. 分批订阅并同步行情
        logger.info("开始同步行情数据 (参考 wisecoin_trade_01_data.py 的逐个 await 方式，每批 50 个订阅)...")
        
        batch_size = 200
        save_interval = 3000  # 每 500 个新获取的保存一次
        reset_interval = 3000  # 每 2000 个新获取的行情后重建 api
        processed_new_count = 0
        api_restart_count = 0
        
        def _rebuild_api_if_needed():
            global api
            nonlocal api_restart_count
            try:
                api_restart_count += 1
                logger.info(f"⚙️ 达到重建阈值，准备重建 TqApi，第 {api_restart_count} 次重建")
                try:
                    api.close()
                except Exception as close_ex:
                    logger.warning(f"关闭现有 api 失败: {close_ex}")
                # 按 RUN_MODE 重建，与主流程一致
                if RUN_MODE == 1:
                    backtest_start_dt, backtest_end_dt = (datetime.date.today() + datetime.timedelta(days=-1), datetime.date.today() + datetime.timedelta(days=0))
                    acc_sim = TqSim(init_balance=10000000)
                    api = TqApi(account=acc_sim, backtest=TqBacktest(start_dt=backtest_start_dt,
                                end_dt=backtest_end_dt), debug=False, web_gui=False, auth=TqAuth('playbonze', 'abC!@#123'))
                elif RUN_MODE == 2:
                    api = TqApi(TqKq(), debug=False, web_gui=False,
                                auth=TqAuth('huaying', 'bonze13'))
                elif RUN_MODE == 3:
                    api = TqApi(TqAccount('simnow', '207302', 'Bonze!0613'), debug=False,
                                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
                elif RUN_MODE == 4:
                    api = TqApi(TqAccount('B渤海期货', '98908572', 'bonze613'), debug=False,
                                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
                elif RUN_MODE == 5:
                    api = TqApi(TqAccount('H华安期货', '100919200', 'bonze613'), debug=False,
                                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
                elif RUN_MODE == 6:
                    api = TqApi(TqAccount('J金信期货', '80016087', 'bonze613'), debug=False,
                                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
                elif RUN_MODE == 7:
                    api = TqApi(TqAccount('D东吴期货', '526178061', 'bonze613'), debug=False,
                                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
                elif RUN_MODE == 8:
                    api = TqApi(TqAccount('H宏源期货', '901212925', 'bonze613'), debug=False,
                                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
                else:
                    api = TqApi(TqKq(), debug=False, web_gui=False,
                                auth=TqAuth('huaying', 'bonze13'))
                logger.info("✅ TqApi 重建完成，继续获取行情...")
            except Exception as rebuild_ex:
                logger.error(f"重建 TqApi 失败: {rebuild_ex}")
        
        for i in range(0, len(symbols_to_fetch_now), batch_size):
            batch = symbols_to_fetch_now[i:i + batch_size]
            try:
                # 直接使用 async 版本的 api.get_quote_list(batch) 获取并等待行情
                quotes = await api.get_quote_list(batch)
                
                # 处理获取到的行情
                for q in quotes:
                    symbol = q._path[-1]
                    # 整理并转换 quote 字段
                    q_dict = dict(q)
                    # 价格补偿逻辑
                    if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
                        if not pd.isna(q_dict.get('pre_close')):
                            q_dict['last_price'] = q_dict['pre_close']
                    
                    all_quote_data[symbol] = q_dict
                    processed_new_count += 1
                    already_fetched_symbols.add(symbol)
                    
                    # 每 500 个触发一次保存
                    if processed_new_count % save_interval == 0:
                        _save_quotes_to_excel(QUOTE_EXCEL_FILE, sheet_df_map, all_quote_data, len(already_fetched_symbols), len(unique_symbols))
                    
                    # 每 2000 个触发一次重建 api（先保存确保数据不丢）
                    if processed_new_count % reset_interval == 0:
                        _save_quotes_to_excel(QUOTE_EXCEL_FILE, sheet_df_map, all_quote_data, len(already_fetched_symbols), len(unique_symbols))
                        _rebuild_api_if_needed()
                
                # 即使没有达到 save_interval，在每个 batch 完成后若这是最后一批，也进行保存
                if i + batch_size >= len(symbols_to_fetch_now):
                    _save_quotes_to_excel(QUOTE_EXCEL_FILE, sheet_df_map, all_quote_data, len(already_fetched_symbols), len(unique_symbols))
                        
            except Exception as e:
                logger.error(f"同步批次 {i//batch_size + 1} 时发生错误: {e}")
            
            # 进度显示
            current_total = len(already_fetched_symbols)
            logger.info(f"  [行情同步] 进度: {current_total}/{len(unique_symbols)} ({current_total/len(unique_symbols):.1%})")

        logger.info(f"🚀 行情数据获取及保存完成: {QUOTE_EXCEL_FILE}")
        if all_quote_data:
            pass
            
    except Exception as e:
        logger.error(f"获取期权行情异常: {e}")
        logger.error(traceback.format_exc())

async def get_underlying_futures_quotes():
    """
    读取 wisecoin-期权行情.xlsx 中的标的，获取其对应的期货行情
    并保存为 wisecoin-期货行情.xlsx，字段与期权行情保持一致。
    【优化：严格对应所有字段，包括 query_symbol_info 的字段】
    """
    QUOTE_EXCEL_FILE = "wisecoin-期权行情.xlsx"
    FUTURES_EXCEL_FILE = "wisecoin-期货行情.xlsx"
    
    if not os.path.exists(QUOTE_EXCEL_FILE):
        logger.warning(f"未找到期权行情文件: {QUOTE_EXCEL_FILE}，无法获取期货行情。")
        return

    try:
        logger.info(f"正在读取 {QUOTE_EXCEL_FILE} 以获取标的期货列表和字段模板...")
        xls = pd.ExcelFile(QUOTE_EXCEL_FILE)
        all_underlyings = set()
        
        # 获取字段模板 (从第一个数据 Sheet 获取)
        template_columns = []
        for sheet_name in xls.sheet_names:
            if sheet_name in ["Summary", "Progress", "Summary_Stats"]: continue
            df_temp = pd.read_excel(xls, sheet_name=sheet_name, nrows=0)
            template_columns = df_temp.columns.tolist()
            break
            
        if not template_columns:
            logger.error("未能从期权行情文件中提取到字段模板")
            return

        # 遍历所有 Sheet 获取唯一的 underlying_symbol
        for sheet_name in xls.sheet_names:
            if sheet_name in ["Summary", "Progress", "Summary_Stats"]: continue
            df = pd.read_excel(xls, sheet_name=sheet_name)
            if 'underlying_symbol' in df.columns:
                symbols = df['underlying_symbol'].dropna().unique().tolist()
                for s in symbols:
                    if s and isinstance(s, str) and '.' in s:
                        all_underlyings.add(s)
        
        unique_underlyings = sorted(list(all_underlyings))
        
        if not unique_underlyings:
            logger.warning("未在期权行情中发现有效的标的期货合约。")
            return
            
        logger.info(f"获取到 {len(unique_underlyings)} 个标的期货合约，准备同步详细信息和行情...")
        
        # 1. 获取详细的合约信息 (query_symbol_info)
        batch_size = 500
        all_symbol_info_dfs = []
        for i in range(0, len(unique_underlyings), batch_size):
            batch = unique_underlyings[i:i + batch_size]
            try:
                info_df = await asyncio.wait_for(api.query_symbol_info(batch), 30)
                if not info_df.empty:
                    all_symbol_info_dfs.append(info_df)
            except Exception as e:
                logger.warning(f"获取期货合约信息批次 {i} 失败: {e}")
        
        if not all_symbol_info_dfs:
            logger.error("未能获取到任何期货合约详细信息")
            return
            
        full_info_df = pd.concat(all_symbol_info_dfs)
        
        # 2. 获取实时行情 (Quote List) - 分批获取
        batch_size = 200
        futures_quotes = {}
        for i in range(0, len(unique_underlyings), batch_size):
            batch = unique_underlyings[i:i + batch_size]
            try:
                quotes = await api.get_quote_list(batch)
                for q in quotes:
                    futures_quotes[q._path[-1]] = q
                logger.info(f"已同步期货行情: {min(i + batch_size, len(unique_underlyings))}/{len(unique_underlyings)}")
            except Exception as e:
                logger.warning(f"获取期货行情批次 {i} 失败: {e}")
        
        futures_final_data = []
        
        # --- 建立快速查找字典 (智能识别并拼接合约代码) ---
        info_lookup = {}
        for idx, row in full_info_df.iterrows():
            # 策略：优先检查 index，其次检查 symbol 列，最后拼接 exchange_id.instrument_id
            s_key = None
            
            # 1. 检查 index 是否是 symbol (带 '.' 的字符串)
            if isinstance(idx, str) and '.' in idx:
                s_key = idx
            
            # 2. 如果 index 不是，检查是否存在 'symbol' 列
            if not s_key:
                s_col = row.get('symbol')
                if not pd.isna(s_col) and s_col:
                    s_key = str(s_col)
            
            # 3. 最后手段：拼接 exchange_id 和 instrument_id
            if not s_key:
                inst = row.get('instrument_id')
                exch = row.get('exchange_id')
                if inst and exch:
                    s_key = str(inst) if '.' in str(inst) else f"{exch}.{inst}"
            
            if s_key:
                info_lookup[str(s_key).strip()] = row.to_dict()

        logger.info(f"合约信息汇总完成: {len(full_info_df)} 条记录, 匹配字典大小: {len(info_lookup)}")
        # 调试：打印一个匹配示例
        if unique_underlyings and info_lookup:
            logger.info(f"匹配测试: Excel标的='{unique_underlyings[0]}', 字典首个Key='{list(info_lookup.keys())[0]}'")

        for symbol in unique_underlyings:
            sym_str = str(symbol).strip()
            try:
                # 获取该合约的静态信息
                static_info = info_lookup.get(sym_str)
                
                if static_info is None:
                    # 尝试不区分大小写匹配
                    sym_upper = sym_str.upper()
                    for k, v in info_lookup.items():
                        if k.upper() == sym_upper:
                            static_info = v
                            break
                
                if static_info is None:
                    # 记录未匹配到的合约
                    logger.debug(f"合约 {sym_str} 未在字典中找到匹配静态信息，跳过")
                    continue
                
                # 获取动态行情
                q = futures_quotes.get(sym_str)
                if q is None:
                    continue
                    
                q_dict = dict(q)
                
                # 价格补偿逻辑
                if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
                    if not pd.isna(q_dict.get('pre_close')):
                        q_dict['last_price'] = q_dict['pre_close']
                
                # 合并静态信息和行情信息 (行情信息覆盖静态信息中的同名字段)
                combined = static_info.copy()
                combined.update(q_dict)
                
                # 增加 product 字段用于分组
                if '.' in symbol:
                    parts = str(symbol).split('.')
                    exchange = parts[0]
                    code_match = re.match(r'^[a-zA-Z]+', parts[1])
                    if code_match:
                        combined['product'] = f"{exchange}.{code_match.group(0)}"
                    else:
                        combined['product'] = str(symbol)
                else:
                    combined['product'] = "Unknown"
                
                # 按照模板对齐字段
                aligned_row = {}
                for col in template_columns:
                    aligned_row[col] = combined.get(col, np.nan)
                
                # 确保 product 字段存在 (即使模板中没有，分组也需要)
                aligned_row['product'] = combined['product']
                
                futures_final_data.append(aligned_row)
            except Exception as e:
                logger.warning(f"处理合约 {symbol} 数据失败: {e}")
            
        if not futures_final_data:
            logger.error(f"未能整理出任何期货行情数据。尝试匹配了 {len(unique_underlyings)} 个标的，匹配结果字典大小: {len(info_lookup)}")
            return
            
        futures_df = pd.DataFrame(futures_final_data)
        
        # 按产品分组导出到 Excel
        logger.info(f"正在将期货行情导出到 {FUTURES_EXCEL_FILE}...")
        with pd.ExcelWriter(FUTURES_EXCEL_FILE, engine='openpyxl') as writer:
            # 1. 汇总 Sheet
            # 在保存到 Excel 时，移除为了分组而添加但在模板中可能不存在的 extra 字段（如果有的话）
            summary_cols = [c for c in template_columns if c in futures_df.columns]
            if 'product' not in summary_cols and 'product' in futures_df.columns:
                summary_cols.append('product')
            
            futures_df[summary_cols].to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. 分产品 Sheet
            grouped = futures_df.groupby('product')
            for product, group in grouped:
                sheet_name = str(product).replace('.', '_')[-31:]
                # 保持列顺序与模板一致
                group[summary_cols].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 自动调整列宽
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(summary_cols):
                    try:
                        max_len = max(
                            group[col].astype(str).map(len).max() if not group[col].empty else 0,
                            len(str(col))
                        ) + 2
                    except:
                        max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)
                    
        logger.info(f"🚀 期货行情保存完成: {FUTURES_EXCEL_FILE}，共记录 {len(futures_df)} 个标的合约，字段严格对应期权行情。")

    except Exception as e:
        logger.error(f"获取期货行情异常: {e}")
        logger.error(traceback.format_exc())

    except Exception as e:
        logger.error(f"获取期货行情异常: {e}")
        logger.error(traceback.format_exc())

def _save_quotes_to_excel(file_path, sheet_df_map, all_quote_data, current_count, total_count):
    """内部辅助函数：将已获取的行情保存到 Excel"""
    try:
        logger.info(f"  [保存进度] 正在更新 Excel 数据 ({current_count}/{total_count})...")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheet_df_map.items():
                symbol_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
                if symbol_col not in df.columns:
                    continue
                
                # 合并原始字段和行情字段 (仅包含已获取到的 symbol)
                merged_rows = []
                for _, row in df.iterrows():
                    symbol = row[symbol_col]
                    if symbol in all_quote_data:
                        combined_data = row.to_dict()
                        # 行情数据转换
                        q_info = all_quote_data[symbol]
                        # 确保行情数据是字典
                        if not isinstance(q_info, dict):
                            # 如果是从 Excel 加载回来的，本来就是字典；如果是新获取的 Quote 对象转换的，也是字典
                            pass
                        combined_data.update(q_info)
                        merged_rows.append(combined_data)
                
                if not merged_rows:
                    continue
                
                final_df = pd.DataFrame(merged_rows)
                
                # 按照 underlying_symbol 、 strike_price 、 option_class (或 call_or_put) 排序
                sort_priority = ['underlying_symbol', 'strike_price', 'option_class', 'call_or_put']
                available_sort_keys = [k for k in sort_priority if k in final_df.columns]
                
                if available_sort_keys:
                    # 确保 strike_price 是数值类型以便正确排序
                    if 'strike_price' in final_df.columns:
                        final_df['strike_price'] = pd.to_numeric(final_df['strike_price'], errors='coerce')
                    final_df = final_df.sort_values(available_sort_keys)
                
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 自动调整格式
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(final_df.columns):
                    try:
                        max_len = max(final_df[col].astype(str).map(len).max() if not final_df[col].empty else 0, len(str(col))) + 2
                    except:
                        max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)
                
    except Exception as e:
        logger.error(f"保存 Excel 失败: {e}")
    except Exception as e:
        logger.error(f"阶段性保存 Excel 失败: {e}")

async def get_not_underlying_futures_quotes():
    """
    获取全市场非期权标的的期货行情
    1. 使用 api.query_quotes(ins_class='FUTURE', expired=False)
    2. 过滤掉 wisecoin-期权行情.xlsx 中已有的 underlying_symbol
    3. 形成 wisecoin-期货行情-无期权.xlsx
    """
    OPTION_QUOTE_FILE = "wisecoin-期权行情.xlsx"
    TARGET_FILE = "wisecoin-期货行情-无期权.xlsx"
    
    if not os.path.exists(OPTION_QUOTE_FILE):
        logger.warning(f"未找到期权行情文件: {OPTION_QUOTE_FILE}，无法过滤标的，将获取全市场期货。")
        underlying_symbols = set()
    else:
        try:
            logger.info(f"正在读取 {OPTION_QUOTE_FILE} 以获取已有的期权标的列表...")
            xls = pd.ExcelFile(OPTION_QUOTE_FILE)
            underlying_symbols = set()
            
            # 获取字段模板 (从第一个数据 Sheet 获取)
            template_columns = []
            for sheet_name in xls.sheet_names:
                if sheet_name in ["Summary", "Progress", "Summary_Stats"]: continue
                df_temp = pd.read_excel(xls, sheet_name=sheet_name, nrows=0)
                template_columns = df_temp.columns.tolist()
                break
                
            for sheet_name in xls.sheet_names:
                if sheet_name in ["Summary", "Progress", "Summary_Stats"]: continue
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'underlying_symbol' in df.columns:
                    symbols = df['underlying_symbol'].dropna().unique().tolist()
                    for s in symbols:
                        if s and isinstance(s, str):
                            underlying_symbols.add(s.strip())
            
            logger.info(f"获取到 {len(underlying_symbols)} 个期权标的合约。")
        except Exception as e:
            logger.error(f"读取期权行情文件失败: {e}")
            underlying_symbols = set()
            template_columns = []

    try:
        # 1. 获取全市场期货
        logger.info("正在获取全市场活跃期货列表...")
        all_futures = await api.query_quotes(ins_class='FUTURE', expired=False)
        
        # 2. 过滤掉期权标的
        not_underlying_futures = [s for s in all_futures if s.strip() not in underlying_symbols]
        logger.info(f"全市场期货共 {len(all_futures)} 个，其中非期权标的期货共 {len(not_underlying_futures)} 个")
        
        if not not_underlying_futures:
            logger.info("未发现任何非期权标的的期货合约。")
            return

        # 3. 获取详细信息 (query_symbol_info)
        batch_size = 500
        all_symbol_info_dfs = []
        for i in range(0, len(not_underlying_futures), batch_size):
            batch = not_underlying_futures[i:i + batch_size]
            try:
                info_df = await asyncio.wait_for(api.query_symbol_info(batch), 30)
                if not info_df.empty:
                    all_symbol_info_dfs.append(info_df)
                logger.info(f"已获取 {min(i + batch_size, len(not_underlying_futures))}/{len(not_underlying_futures)} 个期货信息")
            except Exception as e:
                logger.warning(f"获取期货信息批次 {i} 失败: {e}")
        
        if not all_symbol_info_dfs:
            logger.error("未能获取到任何期货合约详细信息")
            return
            
        full_info_df = pd.concat(all_symbol_info_dfs)
        
        # 4. 获取实时行情 (Quote)
        logger.info("开始同步行情数据 (分批订阅)...")
        api.get_quote_list(not_underlying_futures)
        
        futures_final_data = []
        
        # --- 建立快速查找字典 (智能识别并拼接合约代码) ---
        info_lookup = {}
        for idx, row in full_info_df.iterrows():
            # 策略：优先检查 index，其次检查 symbol 列，最后拼接 exchange_id.instrument_id
            s_key = None
            
            # 1. 检查 index 是否是 symbol (带 '.' 的字符串)
            if isinstance(idx, str) and '.' in idx:
                s_key = idx
            
            # 2. 如果 index 不是，检查是否存在 'symbol' 列
            if not s_key:
                s_col = row.get('symbol')
                if not pd.isna(s_col) and s_col:
                    s_key = str(s_col)
            
            # 3. 最后手段：拼接 exchange_id 和 instrument_id
            if not s_key:
                inst = row.get('instrument_id')
                exch = row.get('exchange_id')
                if inst and exch:
                    s_key = str(inst) if '.' in str(inst) else f"{exch}.{inst}"
            
            if s_key:
                info_lookup[str(s_key).strip()] = row.to_dict()

        logger.info(f"合约信息汇总完成: {len(full_info_df)} 条记录, 匹配字典大小: {len(info_lookup)}")
        if not_underlying_futures and info_lookup:
            logger.info(f"匹配测试: 计划获取='{not_underlying_futures[0]}', 字典首个Key='{list(info_lookup.keys())[0]}'")

        for symbol in not_underlying_futures:
            sym_str = str(symbol).strip()
            try:
                static_info = info_lookup.get(sym_str)
                
                if static_info is None:
                    # 尝试不区分大小写匹配
                    sym_upper = sym_str.upper()
                    for k, v in info_lookup.items():
                        if k.upper() == sym_upper:
                            static_info = v
                            break
                
                if static_info is None:
                    continue
                
                q = await api.get_quote(symbol)
                q_dict = dict(q)
                
                if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
                    if not pd.isna(q_dict.get('pre_close')):
                        q_dict['last_price'] = q_dict['pre_close']
                
                combined = static_info.copy()
                combined.update(q_dict)
                
                # 增加 product 字段用于分组
                if '.' in sym_str:
                    parts = sym_str.split('.')
                    exchange = parts[0]
                    code_match = re.match(r'^[a-zA-Z]+', parts[1])
                    if code_match:
                        combined['product'] = f"{exchange}.{code_match.group(0)}"
                    else:
                        combined['product'] = sym_str
                else:
                    combined['product'] = "Unknown"
                
                # 对齐字段
                aligned_row = {}
                # 如果没有模板，使用 combined 的所有键
                cols_to_use = template_columns if (template_columns and len(template_columns) > 0) else list(combined.keys())
                for col in cols_to_use:
                    aligned_row[col] = combined.get(col, np.nan)
                
                if 'product' not in aligned_row:
                    aligned_row['product'] = combined['product']
                
                futures_final_data.append(aligned_row)
            except Exception as e:
                logger.debug(f"处理合约 {symbol} 数据失败: {e}")

        if not futures_final_data:
            logger.error(f"未能整理出任何非期权标的期货行情数据。尝试匹配了 {len(not_underlying_futures)} 个合约，匹配字典大小: {len(info_lookup)}")
            if len(info_lookup) > 0:
                logger.error(f"字典前5个Key示例: {list(info_lookup.keys())[:5]}")
            return
            
        futures_df = pd.DataFrame(futures_final_data)
        
        # 5. 保存到 Excel
        logger.info(f"正在将非标的期货行情导出到 {TARGET_FILE}...")
        with pd.ExcelWriter(TARGET_FILE, engine='openpyxl') as writer:
            # Summary Sheet
            summary_cols = list(futures_df.columns)
            futures_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # 分产品 Sheet
            grouped = futures_df.groupby('product')
            for product, group in grouped:
                sheet_name = str(product).replace('.', '_')[-31:]
                group.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 调整列宽
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(group.columns):
                    try:
                        max_len = max(group[col].astype(str).map(len).max(), len(str(col))) + 2
                    except: max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)
                    
        logger.info(f"🚀 非标的期货行情保存完成: {TARGET_FILE}，共记录 {len(futures_df)} 个合约。")

    except Exception as e:
        logger.error(f"获取非标的期货行情异常: {e}")
        logger.error(traceback.format_exc())


# 运行模式
RUN_MODE = 2
if os.path.basename(__file__) in ['pb-quant-test-1.py', 'pb-quant-test-2.py']:
    RUN_MODE = 1
elif os.path.basename(__file__) in ['pb-quant-kq.py']:
    RUN_MODE = 2
elif os.path.basename(__file__) in ['pb-quant-bh.py']:
    RUN_MODE = 4
elif os.path.basename(__file__) in ['pb-quant-jx.py']:
    RUN_MODE = 6

if RUN_MODE == 1:
    backtest_start_dt, backtest_end_dt = (datetime.date.today() + datetime.timedelta(days=-1), datetime.date.today() + datetime.timedelta(days=0))
    acc_sim = TqSim(init_balance=10000000)
    api = TqApi(account=acc_sim, backtest=TqBacktest(start_dt=backtest_start_dt,
                end_dt=backtest_end_dt), debug=False, web_gui=False, auth=TqAuth('playbonze', 'abC!@#123'))
elif RUN_MODE == 2:
    api = TqApi(TqKq(), debug=False, web_gui=False,
                auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 3:
    api = TqApi(TqAccount('simnow', '207302', 'Bonze!0613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 4:
    api = TqApi(TqAccount('B渤海期货', '98908572', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 5:
    api = TqApi(TqAccount('H华安期货', '100919200', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 6:
    api = TqApi(TqAccount('J金信期货', '80016087', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 7:
    api = TqApi(TqAccount('D东吴期货', '526178061', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 8:
    api = TqApi(TqAccount('H宏源期货', '901212925', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))

logger.info('模式%d，期权策略开始运行。' % RUN_MODE)

# 期权策略参数
try:
    # 创建期权分析任务
    async def run_sequence():
        await get_all_option_symbols()
        await get_option_quotes_from_excel()
        await get_underlying_futures_quotes()
        await get_not_underlying_futures_quotes()
        logger.info("所有期权及标的期货、非标的期货信息处理完成。")
        return

    api.create_task(run_sequence())

    while True:
        api.wait_update()
        
except BacktestFinished:
    api.close()
    print('\n期权回测完成。')
except Exception as e:
    logger_print.info(f'{repr(e)}，line {sys._getframe().f_lineno}。')
    logger_print.info(traceback.format_exc())
