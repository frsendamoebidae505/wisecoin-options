#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WiseCoin 期货期权联动分析脚本。

等价于原 05wisecoin-futures-analyze.py，使用新架构模块实现。

功能:
1. 期货资金流向分析 - 沉淀资金、成交资金（保证金计算）
2. 期权PCR情绪分析 - Put/Call Ratio多空判断
3. 杠杆涨跌计算 - 实际涨跌幅 × 杠杆倍数
4. 货权联动信号检测 - 期货期权共振/背离检测
5. 最大痛点关联 - 期权Max Pain与期货价格距离

输出: wisecoin-货权联动.xlsx

Usage:
    python3 -m cli.futures_analyzer
"""
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

import numpy as np
import pandas as pd

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from common.logger import StructuredLogger
from core.futures_analyzer import (
    FuturesAnalyzer,
    FuturesAnalysisResult,
    LinkageAnalysisResult,
    TrendDirection,
    FlowSignal,
    ResonanceLevel,
)

# Excel 样式
try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 设置日志
logger = StructuredLogger("futures_analyzer")

# 默认文件路径
DEFAULT_SYMBOL_PARAMS_FILE = "wisecoin-symbol-params.json"
DEFAULT_FUTURE_QUOTE_FILE = "wisecoin-期货行情.xlsx"
DEFAULT_FUTURE_QUOTE_FILE_NO_OPT = "wisecoin-期货行情-无期权.xlsx"
DEFAULT_OPTION_RANKING_FILE = "wisecoin-期权排行.xlsx"
DEFAULT_OPTION_REFERENCE_FILE = "wisecoin-期权参考.xlsx"
DEFAULT_OUTPUT_FILE = "wisecoin-货权联动.xlsx"


def _extract_category_name(categories_field):
    """
    从 categories 字段提取板块名称

    Args:
        categories_field: 可能是字符串、列表或None

    Returns:
        str: 板块名称，如 '农副'、'软商'、'能化' 等
    """
    if pd.isna(categories_field):
        return '未分类'

    try:
        # 如果是字符串，尝试解析为JSON
        if isinstance(categories_field, str):
            import ast
            categories_field = ast.literal_eval(categories_field)

        # 如果是列表，取第一个元素
        if isinstance(categories_field, list) and len(categories_field) > 0:
            cat = categories_field[0]
            if isinstance(cat, dict) and 'name' in cat:
                return cat['name']

        # 如果是字典
        if isinstance(categories_field, dict) and 'name' in categories_field:
            return categories_field['name']

    except Exception:
        pass

    return '未分类'


class FuturesAnalysisRunner:
    """
    期货期权联动分析执行器。

    封装期货期权联动分析流程，生成多维度分析报告。
    """

    def __init__(
        self,
        symbol_params_file: str = DEFAULT_SYMBOL_PARAMS_FILE,
        future_quote_file: str = DEFAULT_FUTURE_QUOTE_FILE,
        future_quote_file_no_opt: str = DEFAULT_FUTURE_QUOTE_FILE_NO_OPT,
        option_ranking_file: str = DEFAULT_OPTION_RANKING_FILE,
        option_reference_file: str = DEFAULT_OPTION_REFERENCE_FILE,
        output_file: str = DEFAULT_OUTPUT_FILE,
    ):
        """
        初始化分析执行器。

        Args:
            symbol_params_file: 品种参数文件路径
            future_quote_file: 期货行情文件路径（有期权）
            future_quote_file_no_opt: 期货行情文件路径（无期权）
            option_ranking_file: 期权排行文件路径
            option_reference_file: 期权参考文件路径
            output_file: 输出文件路径
        """
        self.symbol_params_file = symbol_params_file
        self.future_quote_file = future_quote_file
        self.future_quote_file_no_opt = future_quote_file_no_opt
        self.option_ranking_file = option_ranking_file
        self.option_reference_file = option_reference_file
        self.output_file = output_file

        # 初始化分析器
        self.futures_analyzer = FuturesAnalyzer()

        # 加载品种参数
        self.symbol_params = self._load_symbol_params()

        # 缓存数据
        self._margin_ratios: Dict[str, float] = {}

    def _load_symbol_params(self) -> Dict:
        """加载品种参数配置"""
        if not os.path.exists(self.symbol_params_file):
            logger.warning(f"品种参数文件不存在: {self.symbol_params_file}")
            return {}

        try:
            with open(self.symbol_params_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载品种参数失败: {e}")
            return {}

    def _get_margin_ratio(self, product_code: str) -> float:
        """获取品种保证金率"""
        if product_code in self._margin_ratios:
            return self._margin_ratios[product_code]

        # 从配置中查找
        for exchange, products in self.symbol_params.items():
            if exchange.startswith('_'):
                continue
            if not isinstance(products, dict):
                continue
            for p_code, p_info in products.items():
                if isinstance(p_info, dict) and 'margin_ratio' in p_info:
                    self._margin_ratios[p_code.upper()] = p_info['margin_ratio']

        return self._margin_ratios.get(product_code.upper(), 0.10)

    def _load_futures_data(self) -> pd.DataFrame:
        """加载期货行情数据"""
        futures_df_list = []

        # 加载有期权的期货行情
        if os.path.exists(self.future_quote_file):
            try:
                xls = pd.ExcelFile(self.future_quote_file)
                if 'Summary' in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name='Summary')
                    futures_df_list.append(df)
                    logger.info(f"加载期货行情(有期权): {len(df)} 个合约")
            except Exception as e:
                logger.error(f"加载期货行情(有期权)失败: {e}")

        # 加载无期权的期货行情
        if os.path.exists(self.future_quote_file_no_opt):
            try:
                xls = pd.ExcelFile(self.future_quote_file_no_opt)
                sheet_name = 'Summary' if 'Summary' in xls.sheet_names else xls.sheet_names[0]
                df = pd.read_excel(xls, sheet_name=sheet_name)
                futures_df_list.append(df)
                logger.info(f"加载期货行情(无期权): {len(df)} 个合约")
            except Exception as e:
                logger.error(f"加载期货行情(无期权)失败: {e}")

        if not futures_df_list:
            logger.warning("未找到有效期货行情数据")
            return pd.DataFrame()

        return pd.concat(futures_df_list, ignore_index=True)

    def _load_option_ranking(self) -> Optional[pd.DataFrame]:
        """加载期权排行数据"""
        if not os.path.exists(self.option_ranking_file):
            logger.warning(f"期权排行文件不存在: {self.option_ranking_file}")
            return None

        try:
            xls = pd.ExcelFile(self.option_ranking_file)
            if '期权排行' in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name='期权排行')
                logger.info(f"加载期权排行: {len(df)} 个标的")
                return df
        except Exception as e:
            logger.warning(f"加载期权排行失败: {e}")

        return None

    def _analyze_futures(self, futures_df: pd.DataFrame) -> pd.DataFrame:
        """
        分析期货合约。

        Args:
            futures_df: 原始期货行情数据

        Returns:
            期货分析结果 DataFrame
        """
        futures_analysis = []

        for _, row in futures_df.iterrows():
            try:
                symbol = row.get('instrument_id') or row.get('symbol', '')
                if not symbol or pd.isna(symbol):
                    continue

                # 提取品种代码
                product_code = ''
                if '.' in str(symbol):
                    parts = str(symbol).split('.')
                    if len(parts) >= 2:
                        code_match = re.match(r'^([a-zA-Z]+)', parts[1])
                        if code_match:
                            product_code = code_match.group(1).upper()

                if not product_code:
                    product_code = re.sub(r'[^a-zA-Z]', '', str(symbol)).upper()

                # 获取保证金率
                margin_ratio = self._get_margin_ratio(product_code)
                leverage = 1.0 / margin_ratio if margin_ratio > 0 else 10.0

                # 价格数据
                last_price = row.get('last_price', 0) or row.get('settlement', 0) or 0
                pre_close = row.get('pre_close', 0) or row.get('pre_settlement', 0) or 0
                multiplier = row.get('volume_multiple', 1) or 1

                if pd.isna(last_price) or last_price <= 0:
                    last_price = pre_close
                if pd.isna(pre_close) or pre_close <= 0:
                    continue

                # 持仓量和成交量
                open_interest = row.get('open_interest', 0) or 0
                volume = row.get('volume', 0) or 0
                pre_oi = row.get('pre_open_interest', 0) or 0

                # 核心计算
                # 1. 实际沉淀资金（保证金口径，亿元）
                chendian = (open_interest * last_price * multiplier * margin_ratio) / 1e8

                # 2. 实际成交资金（保证金口径，亿元）
                chengjiao = (volume * last_price * multiplier * margin_ratio) / 1e8

                # 3. 杠杆涨跌
                price_change_pct = ((last_price - pre_close) / pre_close) * 100
                leverage_change = price_change_pct * leverage

                # 4. 持仓变化
                oi_change = open_interest - pre_oi if pre_oi > 0 else 0
                oi_change_pct = (oi_change / pre_oi * 100) if pre_oi > 0 else 0

                # 5. 换手率
                turnover = (volume / open_interest * 100) if open_interest > 0 else 0

                # 6. 资金流向判断
                if oi_change > 0 and price_change_pct > 0:
                    flow_direction = '增仓上涨'
                    flow_signal = 2
                elif oi_change > 0 and price_change_pct < 0:
                    flow_direction = '增仓下跌'
                    flow_signal = -2
                elif oi_change < 0 and price_change_pct > 0:
                    flow_direction = '减仓上涨'
                    flow_signal = 1
                elif oi_change < 0 and price_change_pct < 0:
                    flow_direction = '减仓下跌'
                    flow_signal = -1
                else:
                    flow_direction = '持平'
                    flow_signal = 0

                # 7. 趋势状态分类
                trend_state, trend_dir, trend_strength = self._classify_trend_state(
                    price_change_pct, oi_change_pct
                )

                futures_analysis.append({
                    '合约': symbol,
                    '品种代码': product_code,
                    '现价': round(last_price, 2),
                    '昨收': round(pre_close, 2),
                    '涨跌%': round(price_change_pct, 2),
                    '杠杆涨跌%': round(leverage_change, 2),
                    '杠杆倍数': round(leverage, 1),
                    '保证金率%': round(margin_ratio * 100, 1),
                    '持仓量': int(open_interest),
                    '成交量': int(volume),
                    '持仓变化': int(oi_change),
                    '持仓变化%': round(oi_change_pct, 2),
                    '沉淀资金(亿)': round(chendian, 4),
                    '成交资金(亿)': round(chengjiao, 4),
                    '换手率%': round(turnover, 2),
                    '资金流向': flow_direction,
                    '流向信号': flow_signal,
                    '合约乘数': int(multiplier),
                    '趋势状态': trend_state,
                    '趋势方向': trend_dir,
                    '趋势强度': trend_strength,
                })
            except Exception as e:
                logger.debug(f"分析期货合约时出错: {e}")
                continue

        return pd.DataFrame(futures_analysis)

    def _classify_trend_state(self, price_change_pct: float, oi_change_pct: float) -> tuple:
        """分类期货趋势状态"""
        # 趋势状态
        if price_change_pct > 2 and oi_change_pct > 5:
            state = '强势多头'
            direction = '多'
            strength = 3
        elif price_change_pct > 1 and oi_change_pct > 2:
            state = '温和多头'
            direction = '多'
            strength = 2
        elif price_change_pct > 0 and oi_change_pct > 0:
            state = '弱势多头'
            direction = '多'
            strength = 1
        elif price_change_pct < -2 and oi_change_pct > 5:
            state = '强势空头'
            direction = '空'
            strength = 3
        elif price_change_pct < -1 and oi_change_pct > 2:
            state = '温和空头'
            direction = '空'
            strength = 2
        elif price_change_pct < 0 and oi_change_pct > 0:
            state = '弱势空头'
            direction = '空'
            strength = 1
        elif price_change_pct > 0 and oi_change_pct < 0:
            state = '减仓反弹'
            direction = '多'
            strength = 1
        elif price_change_pct < 0 and oi_change_pct < 0:
            state = '减仓回调'
            direction = '空'
            strength = 1
        else:
            state = '震荡'
            direction = '中性'
            strength = 0

        return state, direction, strength

    def _analyze_linkage(
        self,
        futures_df: pd.DataFrame,
        option_df: Optional[pd.DataFrame]
    ) -> pd.DataFrame:
        """
        货权联动分析。

        Args:
            futures_df: 期货分析结果
            option_df: 期权排行数据

        Returns:
            货权联动分析结果
        """
        if option_df is None or option_df.empty:
            return pd.DataFrame()

        correlation_analysis = []

        for _, opt_row in option_df.iterrows():
            try:
                underlying = opt_row.get('标的合约', '')
                if not underlying:
                    continue

                # 匹配期货合约
                fut_match = futures_df[futures_df['合约'] == underlying]
                if fut_match.empty:
                    continue

                fut_row = fut_match.iloc[0]

                # 期权数据
                pcr = opt_row.get('PCR(持仓)', 1.0)
                pcr_vol = opt_row.get('PCR(成交)', 1.0)
                max_pain = opt_row.get('最大痛点', 0)
                call_oi = opt_row.get('CALL持仓', 0)
                put_oi = opt_row.get('PUT持仓', 0)
                call_change = opt_row.get('CALL变化', 0)
                put_change = opt_row.get('PUT变化', 0)
                opt_sentiment = opt_row.get('情绪倾向', 0)  # 期权情绪倾向

                # 期货数据
                fut_price = fut_row.get('现价', 0)
                price_change = fut_row.get('涨跌%', 0)
                leverage_change = fut_row.get('杠杆涨跌%', 0)
                oi_change_pct = fut_row.get('持仓变化%', 0)
                flow_signal = fut_row.get('流向信号', 0)
                trend_state = fut_row.get('趋势状态', '')
                trend_dir = fut_row.get('趋势方向', '')

                # 期权资金结构
                opt_structure, opt_signal = self._classify_option_structure(
                    pcr, call_change, put_change
                )

                # 共振评分
                resonance_score = self._calculate_resonance(
                    flow_signal, opt_signal, price_change, pcr
                )

                # 联动状态
                linkage_state = self._determine_linkage_state(
                    trend_state, trend_dir, opt_structure, pcr
                )

                # 策略建议
                strategy = self._suggest_strategy(linkage_state, resonance_score)

                # 期货沉淀和期权沉淀
                fut_chendian = fut_row.get('沉淀资金(亿)', 0)
                opt_chendian = opt_row.get('沉淀资金(亿)', 0)  # 期权排行中的沉淀资金

                correlation_analysis.append({
                    '标的合约': underlying,
                    '期货现价': fut_price,
                    '涨跌%': price_change,
                    '杠杆涨跌%': leverage_change,
                    '持仓变化%': oi_change_pct,
                    '期货流向': fut_row.get('资金流向', ''),
                    '期货趋势': trend_state,
                    '期权PCR': round(pcr, 4),
                    'PCR成交': round(pcr_vol, 4),
                    '最大痛点': max_pain,
                    '痛点距离%': round((max_pain - fut_price) / fut_price * 100, 2) if fut_price > 0 else 0,
                    '期权结构': opt_structure,
                    '情绪倾向': opt_sentiment,  # 新增期权情绪
                    '共振评分': resonance_score,
                    '联动状态': linkage_state,
                    '策略建议': strategy,
                    '期货沉淀(亿)': round(fut_chendian, 4),
                    '期权沉淀(亿)': round(opt_chendian, 4),
                    '沉淀资金(亿)': round(fut_chendian, 4),
                })
            except Exception as e:
                logger.debug(f"分析联动时出错: {e}")
                continue

        return pd.DataFrame(correlation_analysis)

    def _classify_option_structure(self, pcr: float, call_change: float, put_change: float) -> tuple:
        """分类期权资金结构"""
        if call_change > 0 and put_change > 0:
            if 0.8 <= pcr <= 1.2:
                return '双向增仓', 0
            elif pcr < 0.8:
                return '偏多增仓', 1
            else:
                return '偏空增仓', -1
        elif call_change > put_change:
            return 'CALL主导', 1
        elif put_change > call_change:
            return 'PUT主导', -1
        else:
            return '中性', 0

    def _calculate_resonance(
        self,
        flow_signal: int,
        opt_signal: int,
        price_change: float,
        pcr: float
    ) -> int:
        """计算共振评分"""
        # 基础分
        score = 0

        # 期货与期权信号同向
        if flow_signal > 0 and opt_signal > 0:
            score += 2
        elif flow_signal < 0 and opt_signal < 0:
            score += 2
        elif flow_signal * opt_signal < 0:
            score -= 1

        # 价格与PCR关系
        if price_change > 0 and pcr < 0.8:
            score += 1
        elif price_change < 0 and pcr > 1.2:
            score += 1

        return max(-3, min(5, score))

    def _determine_linkage_state(
        self,
        trend_state: str,
        trend_dir: str,
        opt_structure: str,
        pcr: float
    ) -> str:
        """确定联动状态"""
        if trend_dir == '多' and opt_structure in ['CALL主导', '偏多增仓']:
            return '多头共振'
        elif trend_dir == '空' and opt_structure in ['PUT主导', '偏空增仓']:
            return '空头共振'
        elif trend_dir == '多' and opt_structure in ['PUT主导', '偏空增仓']:
            return '多头背离'
        elif trend_dir == '空' and opt_structure in ['CALL主导', '偏多增仓']:
            return '空头背离'
        else:
            return '中性震荡'

    def _suggest_strategy(self, linkage_state: str, resonance_score: int) -> str:
        """策略建议"""
        if linkage_state == '多头共振' and resonance_score >= 2:
            return '看多'
        elif linkage_state == '空头共振' and resonance_score >= 2:
            return '看空'
        elif linkage_state == '多头背离':
            return '谨慎看多'
        elif linkage_state == '空头背离':
            return '谨慎看空'
        else:
            return '观望'

    def _generate_futures_market_summary(
        self,
        futures_df: pd.DataFrame,
        correlation_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        生成期货市场概览。

        Args:
            futures_df: 期货分析结果
            correlation_df: 货权联动分析结果

        Returns:
            市场概览 DataFrame
        """
        summary_data = []

        # 整体统计
        total_contracts = len(futures_df)
        total_chendian = futures_df['沉淀资金(亿)'].sum()
        total_chengjiao = futures_df['成交资金(亿)'].sum()

        # 涨跌统计
        up_count = len(futures_df[futures_df['杠杆涨跌%'] > 0])
        down_count = len(futures_df[futures_df['杠杆涨跌%'] < 0])
        flat_count = total_contracts - up_count - down_count

        avg_leverage_change = futures_df['杠杆涨跌%'].mean()
        max_leverage_up = futures_df['杠杆涨跌%'].max()
        max_leverage_down = futures_df['杠杆涨跌%'].min()

        # 资金流向统计
        bullish_count = len(futures_df[futures_df['流向信号'] > 0])
        bearish_count = len(futures_df[futures_df['流向信号'] < 0])
        neutral_count = len(futures_df[futures_df['流向信号'] == 0])

        summary_data = [
            {'指标': '期货合约总数', '数值': total_contracts, '说明': '分析的期货合约数量'},
            {'指标': '期货沉淀资金(亿)', '数值': round(total_chendian, 2), '说明': '持仓量×价格×乘数×保证金率'},
            {'指标': '期货成交资金(亿)', '数值': round(total_chengjiao, 2), '说明': '成交量×价格×乘数×保证金率'},
            {'指标': '上涨品种数', '数值': up_count, '说明': '杠杆涨跌>0的品种'},
            {'指标': '下跌品种数', '数值': down_count, '说明': '杠杆涨跌<0的品种'},
            {'指标': '平盘品种数', '数值': flat_count, '说明': '杠杆涨跌=0的品种'},
            {'指标': '平均杠杆涨跌%', '数值': round(avg_leverage_change, 2), '说明': '所有品种杠杆涨跌均值'},
            {'指标': '最大杠杆涨幅%', '数值': round(max_leverage_up, 2), '说明': '单日最大杠杆收益'},
            {'指标': '最大杠杆跌幅%', '数值': round(max_leverage_down, 2), '说明': '单日最大杠杆亏损'},
            {'指标': '做多信号品种', '数值': bullish_count, '说明': '增仓上涨或减仓下跌'},
            {'指标': '做空信号品种', '数值': bearish_count, '说明': '增仓下跌或减仓上涨'},
            {'指标': '中性品种', '数值': neutral_count, '说明': '无明显方向'},
        ]

        # 市场情绪判断
        if bullish_count > bearish_count * 1.5:
            market_sentiment = '市场整体偏多'
        elif bearish_count > bullish_count * 1.5:
            market_sentiment = '市场整体偏空'
        else:
            market_sentiment = '市场情绪中性'

        summary_data.append({
            '指标': '期货市场情绪',
            '数值': market_sentiment,
            '说明': f'多头{bullish_count}个 vs 空头{bearish_count}个'
        })

        # 货权联动统计
        if not correlation_df.empty:
            strong_resonance = len(correlation_df[correlation_df['共振评分'] >= 3])
            weak_resonance = len(correlation_df[correlation_df['共振评分'] >= 1])
            divergence_count = len(correlation_df[correlation_df['联动状态'].str.contains('背离')])

            summary_data.append({
                '指标': '货权联动品种',
                '数值': len(correlation_df),
                '说明': '同时有期货和期权数据的品种'
            })
            summary_data.append({
                '指标': '强共振',
                '数值': strong_resonance,
                '说明': '共振评分≥3，重点跟踪'
            })
            summary_data.append({
                '指标': '背离警示',
                '数值': divergence_count,
                '说明': '期货期权明显背离，风险提示'
            })

        # 资金Top5
        top5_capital = futures_df.nlargest(5, '沉淀资金(亿)')['合约'].tolist()
        summary_data.append({
            '指标': '资金TOP5',
            '数值': ', '.join(top5_capital),
            '说明': '沉淀资金最大的5个品种'
        })

        # 涨幅Top5
        top5_up = futures_df.nlargest(5, '杠杆涨跌%')['合约'].tolist()
        summary_data.append({
            '指标': '涨幅TOP5',
            '数值': ', '.join(top5_up),
            '说明': '杠杆涨幅最大的5个品种'
        })

        # 跌幅Top5
        top5_down = futures_df.nsmallest(5, '杠杆涨跌%')['合约'].tolist()
        summary_data.append({
            '指标': '跌幅TOP5',
            '数值': ', '.join(top5_down),
            '说明': '杠杆跌幅最大的5个品种'
        })

        return pd.DataFrame(summary_data)

    def _generate_product_analysis(
        self,
        futures_df: pd.DataFrame,
        correlation_df: pd.DataFrame,
        futures_raw_df: pd.DataFrame
    ) -> Dict[str, pd.DataFrame]:
        """
        按期货品种维度生成汇总分析。

        Args:
            futures_df: 期货分析结果
            correlation_df: 货权联动分析结果
            futures_raw_df: 原始期货行情数据

        Returns:
            {'期货品种': DataFrame}
        """
        product_sheets = {}

        try:
            # 从原始期货数据中提取 product_id
            if 'product_id' not in futures_df.columns:
                if 'product_id' in futures_raw_df.columns:
                    product_map = {}
                    for _, row in futures_raw_df.iterrows():
                        symbol = row.get('instrument_id') or row.get('symbol', '')
                        product_id = row.get('product_id', '')
                        if symbol and product_id:
                            product_map[symbol] = product_id
                    futures_df['product_id'] = futures_df['合约'].map(product_map)
                else:
                    # 从合约代码提取品种
                    futures_df['product_id'] = futures_df['品种代码']

            # 按 product_id 分组统计
            product_data = []
            grouped = futures_df.groupby('product_id')

            for product_id, group in grouped:
                if pd.isna(product_id) or product_id == '':
                    continue

                product_summary = {
                    '品种代码': product_id,
                    '合约数量': len(group),
                    '沉淀资金(亿)': round(group['沉淀资金(亿)'].sum(), 4),
                    '成交资金(亿)': round(group['成交资金(亿)'].sum(), 4),
                    '平均杠杆涨跌%': round(group['杠杆涨跌%'].mean(), 2),
                    '最大杠杆涨跌%': round(group['杠杆涨跌%'].max(), 2),
                    '最小杠杆涨跌%': round(group['杠杆涨跌%'].min(), 2),
                    '总持仓量': int(group['持仓量'].sum()),
                    '总成交量': int(group['成交量'].sum()),
                    '看多合约数': len(group[group['流向信号'] > 0]),
                    '看空合约数': len(group[group['流向信号'] < 0]),
                    '中性合约数': len(group[group['流向信号'] == 0]),
                }

                # 品种情绪判断
                if product_summary['合约数量'] > 0:
                    bullish_ratio = product_summary['看多合约数'] / product_summary['合约数量']
                    if bullish_ratio >= 0.6:
                        product_summary['品种情绪'] = '偏多'
                    elif bullish_ratio <= 0.4:
                        product_summary['品种情绪'] = '偏空'
                    else:
                        product_summary['品种情绪'] = '中性'

                # 检查是否有期权联动数据
                if not correlation_df.empty and '标的合约' in correlation_df.columns:
                    product_corr = correlation_df[correlation_df['标的合约'].str.contains(product_id, na=False, case=False)]
                    if not product_corr.empty:
                        product_summary['有期权联动'] = '是'
                        product_summary['期权PCR均值'] = round(product_corr['期权PCR'].mean(), 4)
                        if '最大痛点' in product_corr.columns:
                            product_summary['期权痛点均值'] = round(product_corr['最大痛点'].mean(), 2)
                        if '共振评分' in product_corr.columns:
                            product_summary['共振评分均值'] = round(product_corr['共振评分'].mean(), 2)
                    else:
                        product_summary['有期权联动'] = '否'

                product_data.append(product_summary)

            if product_data:
                product_df = pd.DataFrame(product_data)
                product_df = product_df.sort_values(
                    by=['沉淀资金(亿)', '成交资金(亿)'],
                    ascending=False
                ).reset_index(drop=True)
                product_df.insert(0, '排名', range(1, len(product_df) + 1))
                product_sheets['期货品种'] = product_df
                logger.info(f"生成品种维度汇总: {len(product_df)} 个品种")

        except Exception as e:
            logger.error(f"生成品种维度分析失败: {e}")

        return product_sheets

    def _generate_sector_analysis(
        self,
        futures_df: pd.DataFrame,
        correlation_df: pd.DataFrame,
        futures_raw_df: pd.DataFrame
    ) -> Dict[str, pd.DataFrame]:
        """
        按期货板块维度生成汇总分析。

        Args:
            futures_df: 期货分析结果
            correlation_df: 货权联动分析结果
            futures_raw_df: 原始期货行情数据

        Returns:
            {'期货板块': DataFrame}
        """
        sector_sheets = {}

        try:
            # 从原始期货数据中提取 categories
            if 'categories' not in futures_raw_df.columns:
                logger.warning("未找到 categories 字段，跳过板块维度分析")
                return sector_sheets

            # 建立合约到板块的映射
            sector_map = {}
            for _, row in futures_raw_df.iterrows():
                symbol = row.get('instrument_id') or row.get('symbol', '')
                categories = row.get('categories')
                if symbol:
                    sector_name = _extract_category_name(categories)
                    sector_map[symbol] = sector_name

            # 添加板块列
            futures_df['板块'] = futures_df['合约'].map(sector_map)

            # 按板块分组统计
            sector_data = []
            grouped = futures_df.groupby('板块')
            total_capital = futures_df['沉淀资金(亿)'].sum()

            for sector_name, group in grouped:
                if pd.isna(sector_name) or sector_name == '未分类':
                    continue

                sector_capital = group['沉淀资金(亿)'].sum()

                sector_summary = {
                    '板块名称': sector_name,
                    '品种数量': group['品种代码'].nunique(),
                    '合约数量': len(group),
                    '沉淀资金(亿)': round(sector_capital, 4),
                    '成交资金(亿)': round(group['成交资金(亿)'].sum(), 4),
                    '资金占比%': round(sector_capital / total_capital * 100, 2) if total_capital > 0 else 0,
                    '平均杠杆涨跌%': round(group['杠杆涨跌%'].mean(), 2),
                    '最大杠杆涨跌%': round(group['杠杆涨跌%'].max(), 2),
                    '最小杠杆涨跌%': round(group['杠杆涨跌%'].min(), 2),
                    '总持仓量': int(group['持仓量'].sum()),
                    '总成交量': int(group['成交量'].sum()),
                    '看多合约数': len(group[group['流向信号'] > 0]),
                    '看空合约数': len(group[group['流向信号'] < 0]),
                    '中性合约数': len(group[group['流向信号'] == 0]),
                }

                # 板块情绪判断
                if sector_summary['合约数量'] > 0:
                    bullish_ratio = sector_summary['看多合约数'] / sector_summary['合约数量']
                    if bullish_ratio >= 0.6:
                        sector_summary['板块情绪'] = '偏多'
                    elif bullish_ratio <= 0.4:
                        sector_summary['板块情绪'] = '偏空'
                    else:
                        sector_summary['板块情绪'] = '中性'

                # 板块内品种排行
                product_ranking = group.groupby('品种代码').agg({
                    '沉淀资金(亿)': 'sum',
                    '成交资金(亿)': 'sum',
                    '杠杆涨跌%': 'mean',
                    '持仓量': 'sum',
                    '成交量': 'sum',
                    '流向信号': lambda x: (x > 0).sum() - (x < 0).sum()
                }).reset_index()

                product_ranking.columns = ['品种代码', '沉淀资金(亿)', '成交资金(亿)',
                                           '平均杠杆涨跌%', '总持仓量', '总成交量', '多空信号']
                product_ranking = product_ranking.sort_values('沉淀资金(亿)', ascending=False)

                # TOP3品种
                top3_products = product_ranking.head(3)['品种代码'].tolist()
                sector_summary['品种TOP3'] = ' | '.join(top3_products)

                # 所有品种
                all_products = product_ranking['品种代码'].tolist()
                sector_summary['品种'] = ' / '.join(all_products)

                sector_data.append(sector_summary)

            if sector_data:
                sector_df = pd.DataFrame(sector_data)
                sector_df = sector_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
                sector_df.insert(0, '排名', range(1, len(sector_df) + 1))
                sector_sheets['期货板块'] = sector_df
                logger.info(f"生成板块维度汇总: {len(sector_df)} 个板块")

        except Exception as e:
            logger.error(f"生成板块维度分析失败: {e}")

        return sector_sheets

    def generate_reports(
        self,
        futures_df: pd.DataFrame,
        correlation_df: pd.DataFrame,
        futures_raw_df: pd.DataFrame = None
    ) -> Dict[str, pd.DataFrame]:
        """
        生成多维度分析报告。

        Args:
            futures_df: 期货分析结果
            correlation_df: 货权联动分析结果
            futures_raw_df: 原始期货行情数据（用于提取品种和板块信息）

        Returns:
            各分页的 DataFrame 字典
        """
        sheets = {}

        # 0. 期货市场概览
        if not futures_df.empty:
            market_summary = self._generate_futures_market_summary(futures_df, correlation_df)
            sheets['期货市场'] = market_summary

        # 1. 货权联动
        if not correlation_df.empty:
            corr_sorted = correlation_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            corr_sorted.insert(0, '排名', range(1, len(corr_sorted) + 1))
            # 添加沉淀资金合计列
            if '期货沉淀(亿)' in corr_sorted.columns and '期权沉淀(亿)' in corr_sorted.columns:
                corr_sorted['沉淀资金合计(亿)'] = corr_sorted['期货沉淀(亿)'] + corr_sorted['期权沉淀(亿)']
            elif '沉淀资金(亿)' in corr_sorted.columns:
                # 从期货数据补充期权沉淀
                corr_sorted['沉淀资金合计(亿)'] = corr_sorted['沉淀资金(亿)']
            sheets['货权联动'] = corr_sorted

        # 2. 期货品种维度分析
        if not futures_df.empty and futures_raw_df is not None:
            product_sheets = self._generate_product_analysis(futures_df, correlation_df, futures_raw_df)
            sheets.update(product_sheets)

        # 3. 期货板块维度分析
        if not futures_df.empty and futures_raw_df is not None:
            sector_sheets = self._generate_sector_analysis(futures_df, correlation_df, futures_raw_df)
            sheets.update(sector_sheets)

        # 4. 期货排行
        if not futures_df.empty:
            fut_comprehensive = futures_df.copy()
            fut_comprehensive['综合评分'] = (
                fut_comprehensive['沉淀资金(亿)'].rank(pct=True) * 30 +
                fut_comprehensive['成交资金(亿)'].rank(pct=True) * 30 +
                fut_comprehensive['流向信号'].abs().rank(pct=True) * 20 +
                fut_comprehensive['杠杆涨跌%'].abs().rank(pct=True) * 20
            )
            fut_comprehensive = fut_comprehensive.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            fut_comprehensive.insert(0, '排名', range(1, len(fut_comprehensive) + 1))
            sheets['期货排行'] = fut_comprehensive

            # 3. 期货涨跌
            leverage_df = futures_df[['合约', '品种代码', '现价', '涨跌%', '杠杆涨跌%',
                                      '杠杆倍数', '保证金率%', '资金流向']].copy()
            leverage_df = leverage_df.sort_values('杠杆涨跌%', ascending=False).reset_index(drop=True)
            leverage_df.insert(0, '排名', range(1, len(leverage_df) + 1))
            sheets['期货涨跌'] = leverage_df

            # 4. 期货看多
            bullish_df = futures_df[futures_df['流向信号'] > 0].copy()
            bullish_df = bullish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            if not bullish_df.empty:
                bullish_df.insert(0, '排名', range(1, len(bullish_df) + 1))
                bullish_df = bullish_df[['排名', '合约', '品种代码', '现价', '杠杆涨跌%',
                                          '持仓变化', '资金流向', '沉淀资金(亿)', '流向信号']]
                sheets['期货看多'] = bullish_df

            # 5. 期货看空
            bearish_df = futures_df[futures_df['流向信号'] < 0].copy()
            bearish_df = bearish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            if not bearish_df.empty:
                bearish_df.insert(0, '排名', range(1, len(bearish_df) + 1))
                bearish_df = bearish_df[['排名', '合约', '品种代码', '现价', '杠杆涨跌%',
                                          '持仓变化', '资金流向', '沉淀资金(亿)', '流向信号']]
                sheets['期货看空'] = bearish_df

            # 6. 期货资金
            capital_df = futures_df[['合约', '品种代码', '现价', '沉淀资金(亿)', '成交资金(亿)',
                                      '持仓量', '持仓变化', '资金流向']].copy()
            capital_df['资金合计(亿)'] = capital_df['沉淀资金(亿)'] + capital_df['成交资金(亿)']
            capital_df = capital_df.sort_values('资金合计(亿)', ascending=False).reset_index(drop=True)
            capital_df.insert(0, '排名', range(1, len(capital_df) + 1))
            sheets['期货资金'] = capital_df

        return sheets

    def save_to_excel(self, sheets: Dict[str, pd.DataFrame], output_file: str) -> bool:
        """
        保存到 Excel 文件。

        Args:
            sheets: 各分页数据
            output_file: 输出文件路径

        Returns:
            是否成功
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    if df.empty:
                        continue
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 应用样式
                if OPENPYXL_AVAILABLE:
                    self._apply_formatting(writer, sheets)

            logger.info(f"保存分析报告: {output_file}")
            return True
        except Exception as e:
            logger.error(f"保存文件失败: {e}")
            return False

    def _apply_formatting(self, writer, sheets: Dict[str, pd.DataFrame]):
        """应用Excel格式"""
        # 颜色定义
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        purple_fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

        # 表头样式 - 深蓝色
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name, df in sheets.items():
            if df.empty:
                continue

            ws = writer.sheets[sheet_name]

            # 表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 列宽调整 - 支持中文字符宽度计算
            for idx, col in enumerate(df.columns):
                header_width = self._calculate_display_width(col)
                content_width = df[col].astype(str).map(self._calculate_display_width).max() if len(df) > 0 else 0

                # 针对特定列微调
                if str(col) == '数值':
                    padding = 6
                elif str(col) in ['沉淀资金合计(亿)', '资金合计(亿)']:
                    padding = 0
                elif '百分比' in str(col) or '%' in str(col) or '评分' in str(col):
                    padding = 1
                else:
                    padding = 2

                max_len = max(header_width, content_width) + padding
                ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)

            # 条件格式化
            self._apply_conditional_formatting(ws, df, sheet_name, green_fill, red_fill, yellow_fill, blue_fill, purple_fill, gold_fill)

    def _calculate_display_width(self, s) -> int:
        """计算字符串的显示宽度（中文占2单位，英文占1单位）"""
        if s is None or pd.isna(s):
            return 0
        s = str(s)
        width = 0
        for char in s:
            if ord(char) > 127:  # 中文字符或全角符号
                width += 2
            else:
                width += 1
        return width

    def _apply_conditional_formatting(self, ws, df, sheet_name, green_fill, red_fill, yellow_fill, blue_fill, purple_fill, gold_fill):
        """应用条件格式化"""
        # 涨跌类列
        change_cols = ['涨跌%', '杠杆涨跌%', '持仓变化', '持仓变化%']
        # 信号类列
        signal_cols = ['流向信号', '共振评分']
        # 趋势状态列
        trend_state_cols = ['趋势状态', '期货状态']
        # 联动状态列
        linkage_cols = ['联动状态', '共振标签']
        # 强度分数列
        score_cols = ['联动总分']

        cols = list(df.columns)

        for r_idx, row in enumerate(df.itertuples(), start=2):
            for col_idx, col_name in enumerate(cols, start=1):
                try:
                    # 安全获取值
                    val = None
                    try:
                        val = getattr(row, col_name.replace(' ', '_').replace('(', '').replace(')', '').replace('%', '').replace('/', '_'), None)
                    except:
                        # 如果属性名不匹配，尝试通过索引获取
                        pass

                    if val is None:
                        # 尝试直接从df获取
                        try:
                            val = df.iloc[r_idx - 2, col_idx - 1]
                        except:
                            continue

                    cell = ws.cell(row=r_idx, column=col_idx)

                    # 涨跌着色
                    if col_name in change_cols:
                        try:
                            num_val = float(val)
                            if num_val > 0:
                                cell.fill = red_fill  # 上涨红色
                            elif num_val < 0:
                                cell.fill = green_fill  # 下跌绿色
                        except:
                            pass

                    # 信号着色
                    elif col_name in signal_cols:
                        try:
                            num_val = float(val)
                            if col_name == '共振评分':
                                if num_val >= 7:
                                    cell.fill = gold_fill  # 强共振 金色
                                elif num_val >= 5:
                                    cell.fill = red_fill   # 共振 红色
                                elif num_val >= 3:
                                    cell.fill = yellow_fill  # 中性
                                elif num_val < 1:
                                    cell.fill = purple_fill  # 背离
                            else:
                                if num_val >= 2:
                                    cell.fill = red_fill  # 强多
                                elif num_val >= 1:
                                    cell.fill = yellow_fill  # 偏多
                                elif num_val <= -2:
                                    cell.fill = green_fill  # 强空
                                elif num_val <= -1:
                                    cell.fill = blue_fill  # 偏空
                        except:
                            pass

                    # 趋势状态着色
                    elif col_name in trend_state_cols:
                        str_val = str(val)
                        if '多头强化' in str_val:
                            cell.fill = red_fill
                        elif '多头衰减' in str_val:
                            cell.fill = yellow_fill
                        elif '空头强化' in str_val:
                            cell.fill = green_fill
                        elif '空头衰减' in str_val:
                            cell.fill = blue_fill
                        elif '震荡' in str_val:
                            cell.fill = purple_fill

                    # 联动状态着色
                    elif col_name in linkage_cols:
                        str_val = str(val)
                        if '确认' in str_val or '强共振' in str_val or '加速' in str_val:
                            cell.fill = gold_fill
                        elif '警惕' in str_val or '背离' in str_val:
                            cell.fill = purple_fill
                        elif '机会' in str_val or '共振' in str_val:
                            cell.fill = red_fill
                        elif '信号' in str_val:
                            cell.fill = yellow_fill

                    # 期权结构着色
                    elif col_name == '期权结构':
                        str_val = str(val)
                        if '看多' in str_val:
                            cell.fill = red_fill
                        elif '看空' in str_val:
                            cell.fill = green_fill
                        elif '波动率' in str_val:
                            cell.fill = blue_fill

                    # 期权情绪着色
                    elif col_name == '期权情绪':
                        str_val = str(val)
                        if '狂热' in str_val:
                            cell.fill = red_fill
                        elif '恐慌' in str_val:
                            cell.fill = green_fill
                        elif '筑底' in str_val:
                            cell.fill = yellow_fill
                        elif '冲高' in str_val:
                            cell.fill = blue_fill

                    # 共振等级着色
                    elif col_name == '共振等级':
                        str_val = str(val)
                        if '⭐⭐⭐⭐' in str_val:
                            cell.fill = gold_fill
                        elif '⭐⭐⭐' in str_val:
                            cell.fill = red_fill
                        elif '⚠️' in str_val:
                            cell.fill = purple_fill

                    # 强度模型评分着色
                    elif col_name in score_cols:
                        try:
                            score = float(val)
                            if col_name == '联动总分':
                                if score >= 80:
                                    cell.fill = gold_fill
                                elif score >= 60:
                                    cell.fill = red_fill
                                elif score >= 40:
                                    cell.fill = yellow_fill
                                else:
                                    cell.fill = blue_fill
                        except:
                            pass

                except Exception:
                    continue

    def run(self) -> bool:
        """
        执行完整分析流程。

        Returns:
            是否成功
        """
        logger.info("=" * 60)
        logger.info("开始期货期权联动分析...")
        logger.info("=" * 60)

        # 1. 加载期货数据
        futures_raw_df = self._load_futures_data()
        if futures_raw_df.empty:
            logger.error("未找到有效期货行情数据")
            return False

        # 2. 分析期货
        futures_df = self._analyze_futures(futures_raw_df)
        if futures_df.empty:
            logger.error("期货分析结果为空")
            return False
        logger.info(f"完成 {len(futures_df)} 个期货合约分析")

        # 3. 加载期权排行
        option_df = self._load_option_ranking()

        # 4. 货权联动分析
        correlation_df = self._analyze_linkage(futures_df, option_df)
        if not correlation_df.empty:
            logger.info(f"完成 {len(correlation_df)} 个货权联动分析")

        # 5. 生成报告
        sheets = self.generate_reports(futures_df, correlation_df, futures_raw_df)

        # 6. 保存到 Excel
        success = self.save_to_excel(sheets, self.output_file)

        if success:
            logger.info("=" * 60)
            logger.info("期货期权联动分析完成!")
            logger.info(f"输出文件: {self.output_file}")
            logger.info(f"包含 {len(sheets)} 个分析维度")
            logger.info(f"期货合约: {len(futures_df)} 个")
            if not correlation_df.empty:
                logger.info(f"货权联动: {len(correlation_df)} 个")
            logger.info("=" * 60)

        return success

    def merge_analysis_reports(self) -> bool:
        """
        合并期权排行和货权联动到市场概览。

        Returns:
            是否成功
        """
        logger.info("开始合并生成 wisecoin-市场概览.xlsx...")

        target_file = "wisecoin-市场概览.xlsx"
        reports = {
            'FUTURE': self.output_file,  # wisecoin-货权联动.xlsx
            'OPTION': self.option_ranking_file  # wisecoin-期权排行.xlsx
        }

        # 工作表顺序
        sheet_order = [
            '期货市场', '期权市场', '货权联动', '期货品种', '期货板块',
            '期货排行', '期权排行', '期货涨跌', '期权PCR', '期权痛点',
            '期货看多', '期货看空', '期货资金', '期权资金', '方向型期权', '波动率型期权'
        ]

        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl 不可用，无法合并报告")
            return False

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            # 加载源工作簿
            wbs = {}
            for key, path in reports.items():
                if os.path.exists(path):
                    wbs[key] = openpyxl.load_workbook(path)
                    logger.info(f"加载源文件: {path}")
                else:
                    logger.warning(f"缺失源文件: {path}")

            if not wbs:
                logger.error("未找到任何源分析文件，无法合并")
                return False

            # 创建新工作簿
            new_wb = openpyxl.Workbook()
            if "Sheet" in new_wb.sheetnames:
                del new_wb["Sheet"]

            added_count = 0
            for sheet_name in sheet_order:
                source_ws = None
                for key, wb in wbs.items():
                    if sheet_name in wb.sheetnames:
                        source_ws = wb[sheet_name]
                        break

                if source_ws:
                    new_ws = new_wb.create_sheet(title=sheet_name)
                    self._copy_sheet(source_ws, new_ws)
                    added_count += 1

            new_wb.save(target_file)
            logger.info(f"市场概览合并成功: {target_file}, 共包含 {added_count} 个分页")
            return True

        except Exception as e:
            logger.error(f"合并报告失败: {e}")
            return False

    def _copy_sheet(self, source_ws, target_ws):
        """复制工作表内容和格式"""
        from copy import copy

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy(cell.alignment)

        # 复制合并单元格
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # 复制列宽
        for col_letter, col_dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter].width = col_dim.width

        # 复制行高
        for row_num, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_num].height = row_dim.height


def main():
    """命令行入口"""
    runner = FuturesAnalysisRunner()
    success = runner.run()

    # 合并生成市场概览
    if success:
        runner.merge_analysis_reports()

    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())