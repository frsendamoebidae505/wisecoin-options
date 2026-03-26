#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WiseCoin 期权综合分析脚本。

等价于原 03wisecoin-options-analyze.py，使用新架构模块实现。

功能:
1. 期权综合分析 - 生成 wisecoin-期权排行.xlsx
2. 期权参考数据 - 生成 wisecoin-期权参考.xlsx

Usage:
    python3 -m cli.option_analyzer
    python3 run.py --analyze
"""
import asyncio
import json
import logging
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

import numpy as np
import pandas as pd
from scipy import stats

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from common.logger import StructuredLogger
from common.config import Config
from core.analyzer import (
    OptionAnalyzer,
    PCRAnalyzer,
    MaxPainCalculator,
    OptionScorer,
    UnderlyingAnalyzer,
)
from core.iv_calculator import IVCalculator

# Excel 样式
try:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# TqSDK 导入
try:
    from tqsdk import TqApi, TqAuth
    TQSDK_AVAILABLE = True
except ImportError:
    TQSDK_AVAILABLE = False
    TqApi = None
    TqAuth = None

# 设置日志
logger = StructuredLogger("option_analyzer")

# 默认文件路径
DEFAULT_OPTION_QUOTE_FILE = "wisecoin-期权行情.csv"  # 改用CSV格式
DEFAULT_OPTION_QUOTE_FILE_XLSX = "wisecoin-期权行情.xlsx"  # 兼容旧格式
DEFAULT_FUTURE_QUOTE_FILE = "wisecoin-期货行情.xlsx"
DEFAULT_PARAM_FILE = "wisecoin-symbol-params.json"
DEFAULT_OUTPUT_FILE = "wisecoin-期权排行.xlsx"
DEFAULT_REFERENCE_FILE = "wisecoin-期权参考.xlsx"


class OptionAnalysisRunner:
    """
    期权分析执行器。

    封装期权综合分析流程，生成多维度分析报告。
    """

    def __init__(
        self,
        option_quote_file: str = DEFAULT_OPTION_QUOTE_FILE,
        future_quote_file: str = DEFAULT_FUTURE_QUOTE_FILE,
        param_file: str = DEFAULT_PARAM_FILE,
        output_file: str = DEFAULT_OUTPUT_FILE,
        reference_file: str = DEFAULT_REFERENCE_FILE,
    ):
        """
        初始化分析执行器。

        Args:
            option_quote_file: 期权行情文件路径（支持CSV和XLSX格式）
            future_quote_file: 期货行情文件路径
            param_file: 品种参数文件路径
            output_file: 排行输出文件路径
            reference_file: 参考数据输出文件路径
        """
        # 自动检测文件格式（优先CSV）
        if not os.path.exists(option_quote_file):
            if option_quote_file.endswith('.csv'):
                xlsx_file = option_quote_file.replace('.csv', '.xlsx')
                if os.path.exists(xlsx_file):
                    option_quote_file = xlsx_file
            elif option_quote_file.endswith('.xlsx'):
                csv_file = option_quote_file.replace('.xlsx', '.csv')
                if os.path.exists(csv_file):
                    option_quote_file = csv_file

        self.option_quote_file = option_quote_file
        self.future_quote_file = future_quote_file
        self.param_file = param_file
        self.output_file = output_file
        self.reference_file = reference_file

        # 初始化分析器
        self.option_analyzer = OptionAnalyzer()
        self.pcr_analyzer = PCRAnalyzer()
        self.max_pain_calculator = MaxPainCalculator()
        self.scorer = OptionScorer()
        self.underlying_analyzer = UnderlyingAnalyzer()
        self.iv_calculator = IVCalculator(risk_free_rate=0.015)  # 1.5% 无风险利率

        # 加载品种参数
        self.symbol_params = self._load_symbol_params()

        # 缓存数据
        self._future_prices: Dict[str, float] = {}
        self._future_multipliers: Dict[str, int] = {}
        self._future_names: Dict[str, str] = {}
        self._margin_ratios: Dict[str, float] = {}

    def _load_symbol_params(self) -> Dict:
        """加载品种参数配置"""
        if not os.path.exists(self.param_file):
            logger.warning(f"品种参数文件不存在: {self.param_file}")
            return {}

        try:
            with open(self.param_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载品种参数失败: {e}")
            return {}

    def _get_margin_ratio(self, product_code: str) -> float:
        """获取品种保证金率"""
        if product_code in self._margin_ratios:
            return self._margin_ratios[product_code]

        # 从配置中查找
        for exchange, products in self.symbol_params.items():
            if exchange.startswith('_'):
                continue
            if not isinstance(products, dict):
                continue
            for p_code, p_info in products.items():
                if isinstance(p_info, dict) and 'margin_ratio' in p_info:
                    self._margin_ratios[p_code.upper()] = p_info['margin_ratio']

        return self._margin_ratios.get(product_code.upper(), 0.10)

    def _get_historical_volatility(self, underlyings: List[str]) -> Dict[str, Dict[str, float]]:
        """
        获取标的历史波动率 (HV5, HV20, HV60)。

        Args:
            underlyings: 标的合约列表

        Returns:
            {underlying: {'hv5': float, 'hv20': float, 'hv60': float}}
        """
        hv_map = {}

        def calculate_hv(klines_df, window):
            """计算指定窗口期的历史波动率"""
            if klines_df is None or len(klines_df) < window + 1:
                return np.nan
            try:
                close = klines_df['close'].values
                log_returns = np.log(close[1:] / close[:-1])
                if len(log_returns) >= window:
                    recent_returns = log_returns[-window:]
                    hv = np.std(recent_returns, ddof=1) * np.sqrt(252)
                    return hv if hv > 0 else np.nan
                return np.nan
            except Exception:
                return np.nan

        # 尝试使用 TqSDK 获取历史波动率
        if TQSDK_AVAILABLE and TqApi and TqAuth:
            api = None
            try:
                logger.info(f"获取 {len(underlyings)} 个标的的历史波动率...")
                api = TqApi(auth=TqAuth('huaying', 'bonze13'))

                for und in underlyings:
                    try:
                        klines = api.get_kline_serial(und, 24*60*60, 65)
                        quote = api.get_quote(und)

                        hv5 = calculate_hv(klines, 5)
                        hv20 = calculate_hv(klines, 20)
                        hv60 = calculate_hv(klines, 60)

                        # 使用 tafunc 作为备选
                        if pd.isna(hv5) or hv5 <= 0:
                            try:
                                from tqsdk.tafunc import get_his_volatility
                                hv5 = get_his_volatility(klines, quote)
                            except Exception:
                                pass

                        # 设置默认值
                        if pd.isna(hv5) or hv5 <= 0:
                            hv5 = 0.25
                        if pd.isna(hv20) or hv20 <= 0:
                            hv20 = hv5
                        if pd.isna(hv60) or hv60 <= 0:
                            hv60 = hv20

                        hv_map[und] = {'hv5': hv5, 'hv20': hv20, 'hv60': hv60}

                    except Exception as e:
                        logger.debug(f"获取 {und} 历史波动率失败: {e}")
                        hv_map[und] = {'hv5': 0.25, 'hv20': 0.25, 'hv60': 0.25}

                api.close()
                logger.info("历史波动率获取完成")

            except Exception as e:
                logger.warning(f"连接 TQSDK 失败，使用默认波动率: {e}")
                if api:
                    try:
                        api.close()
                    except Exception:
                        pass

        # 对于没有获取到的标的，使用默认值
        default_hv = {'hv5': 0.25, 'hv20': 0.25, 'hv60': 0.25}
        for und in underlyings:
            if und not in hv_map:
                hv_map[und] = default_hv

        return hv_map

    def _load_future_prices(self) -> bool:
        """加载期货价格数据"""
        if not os.path.exists(self.future_quote_file):
            logger.warning(f"期货行情文件不存在: {self.future_quote_file}")
            return False

        try:
            xls = pd.ExcelFile(self.future_quote_file)
            df = pd.read_excel(xls, sheet_name='Summary')

            id_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
            self._future_prices = df.set_index(id_col)['last_price'].to_dict()

            if 'volume_multiple' in df.columns:
                self._future_multipliers = df.set_index(id_col)['volume_multiple'].to_dict()

            if 'instrument_name' in df.columns:
                self._future_names = df.set_index(id_col)['instrument_name'].to_dict()

            logger.info(f"加载 {len(self._future_prices)} 个标的期货价格")
            return True
        except Exception as e:
            logger.error(f"加载期货行情失败: {e}")
            return False

    def _load_option_data(self) -> pd.DataFrame:
        """加载期权行情数据（支持CSV和XLSX格式）"""
        # 优先检查CSV格式，兼容旧版XLSX格式
        option_file = self.option_quote_file
        if not os.path.exists(option_file):
            if option_file.endswith('.csv'):
                # 尝试XLSX格式
                xlsx_file = option_file.replace('.csv', '.xlsx')
                if os.path.exists(xlsx_file):
                    option_file = xlsx_file
            elif option_file.endswith('.xlsx'):
                # 尝试CSV格式
                csv_file = option_file.replace('.xlsx', '.csv')
                if os.path.exists(csv_file):
                    option_file = csv_file

        if not os.path.exists(option_file):
            logger.error(f"期权行情文件不存在: {self.option_quote_file}")
            return pd.DataFrame()

        try:
            all_data = []

            if option_file.endswith('.csv'):
                # CSV格式：直接读取单个文件
                df = pd.read_csv(option_file)
                if not df.empty:
                    # 如果有_sheet列，保留；否则添加默认值
                    if '_sheet' not in df.columns:
                        df['_sheet'] = 'All'
                    all_data.append(df)
            else:
                # XLSX格式：读取多个sheet
                xls = pd.ExcelFile(option_file)
                for sheet_name in xls.sheet_names:
                    if sheet_name in ["Summary", "Progress", "Summary_Stats"]:
                        continue
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    if df.empty:
                        continue
                    df['_sheet'] = sheet_name
                    all_data.append(df)

            if not all_data:
                return pd.DataFrame()

            options_df = pd.concat(all_data, ignore_index=True)
            logger.info(f"共加载 {len(options_df)} 个期权合约 (从 {option_file})")
            return options_df
        except Exception as e:
            logger.error(f"加载期权行情失败: {e}")
            return pd.DataFrame()

    def _standardize_option_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """标准化期权数据"""
        std_data = []

        for _, row in df.iterrows():
            try:
                # 期权类型
                opt_type = str(row.get('option_class', row.get('call_or_put', ''))).upper()
                if 'CALL' in opt_type or opt_type == 'C':
                    opt_type = 'CALL'
                elif 'PUT' in opt_type or opt_type == 'P':
                    opt_type = 'PUT'
                else:
                    continue

                # 标的合约
                underlying = row.get('underlying_symbol', row.get('product', 'Unknown'))
                if pd.isna(underlying) or underlying == 'Unknown':
                    continue

                # 价格数据
                last_price = row.get('last_price', row.get('close', row.get('pre_close', 0))) or 0
                if pd.isna(last_price) or last_price <= 0:
                    last_price = row.get('pre_close', 0) or 0

                multiplier = row.get('volume_multiple', 1) or 1
                oi = row.get('open_interest', row.get('close_oi', 0)) or 0
                vol = row.get('volume', 0) or 0
                strike = row.get('strike_price', 0) or 0
                expire_days = row.get('expire_rest_days', 30) or 30
                bid = row.get('bid_price1', 0) or 0
                ask = row.get('ask_price1', 0) or 0
                pre_oi = row.get('pre_open_interest', 0) or 0
                symbol = row.get('instrument_id', row.get('symbol', ''))

                # 标的价格
                und_price = self._future_prices.get(underlying, 0)
                if und_price <= 0:
                    und_price = row.get('underlying_price', 0) or 0

                # 计算衍生指标
                chendian = (oi * last_price * multiplier) / 1e8
                chengjiao = (vol * last_price * multiplier) / 1e8
                oi_change = oi - pre_oi if pre_oi > 0 else 0
                turnover = (vol / oi * 100) if oi > 0 else 0
                spread = ask - bid if ask > 0 and bid > 0 else 0
                spread_pct = (spread / last_price * 100) if last_price > 0 else 0

                # 虚实幅度
                if und_price > 0 and strike > 0:
                    if opt_type == 'CALL':
                        moneyness = (und_price - strike) / und_price * 100
                    else:
                        moneyness = (strike - und_price) / und_price * 100
                else:
                    moneyness = 0

                # 内在价值与时间价值
                if und_price > 0:
                    if opt_type == 'CALL':
                        intrinsic = max(0, und_price - strike)
                    else:
                        intrinsic = max(0, strike - und_price)
                    time_value = max(0, last_price - intrinsic)
                else:
                    intrinsic = 0
                    time_value = last_price

                std_data.append({
                    'symbol': symbol,
                    'underlying': underlying,
                    'opt_type': opt_type,
                    'strike': strike,
                    'last_price': last_price,
                    'bid': bid,
                    'ask': ask,
                    'volume': vol,
                    'open_interest': oi,
                    'pre_oi': pre_oi,
                    'oi_change': oi_change,
                    'multiplier': multiplier,
                    'expire_days': expire_days,
                    'und_price': und_price,
                    'chendian': chendian,
                    'chengjiao': chengjiao,
                    'turnover': turnover,
                    'spread': spread,
                    'spread_pct': spread_pct,
                    'moneyness': moneyness,
                    'intrinsic': intrinsic,
                    'time_value': time_value,
                })
            except Exception:
                continue

        return pd.DataFrame(std_data)

    def analyze_by_underlying(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        按标的合约进行综合分析。

        Args:
            df: 标准化的期权数据

        Returns:
            标的分析结果 DataFrame
        """
        results = []

        for underlying, group in df.groupby('underlying'):
            try:
                calls = group[group['opt_type'] == 'CALL']
                puts = group[group['opt_type'] == 'PUT']

                if calls.empty or puts.empty:
                    continue

                # 基础统计
                total_oi = group['open_interest'].sum()
                total_volume = group['volume'].sum()
                call_oi = calls['open_interest'].sum()
                put_oi = puts['open_interest'].sum()
                call_oi_change = calls['oi_change'].sum()
                put_oi_change = puts['oi_change'].sum()

                # 资金计算 - 直接汇总已计算好的沉淀资金和成交资金
                und_price = group['und_price'].iloc[0] if len(group) > 0 else 0
                multiplier = group['multiplier'].iloc[0] if len(group) > 0 else 1

                total_chendian = group['chendian'].sum()
                total_chengjiao = group['chengjiao'].sum()

                # PCR 分析
                pcr_oi = put_oi / call_oi if call_oi > 0 else 1.0
                pcr_vol = puts['volume'].sum() / calls['volume'].sum() if calls['volume'].sum() > 0 else 1.0

                # 最大痛点
                max_pain = self._calculate_max_pain(group, und_price)

                # 情绪分析
                sentiment = self.pcr_analyzer.calculate_sentiment(pcr_oi)

                # 交易类型分类
                trading_type = self._classify_trading_type(call_oi_change, put_oi_change, pcr_oi)

                results.append({
                    '标的合约': underlying,
                    '期货现价': round(und_price, 2),
                    '总持仓量': int(total_oi),
                    '总成交量': int(total_volume),
                    'CALL持仓': int(call_oi),
                    'PUT持仓': int(put_oi),
                    'CALL变化': int(call_oi_change),
                    'PUT变化': int(put_oi_change),
                    'PCR(持仓)': round(pcr_oi, 4),
                    'PCR(成交)': round(pcr_vol, 4),
                    '最大痛点': round(max_pain, 2),
                    '痛点距离%': round((max_pain - und_price) / und_price * 100, 2) if und_price > 0 else 0,
                    '沉淀资金(亿)': round(total_chendian, 4),
                    '成交资金(亿)': round(total_chengjiao, 4),
                    '情绪倾向': sentiment,
                    '交易类型': trading_type,
                })
            except Exception as e:
                logger.debug(f"分析 {underlying} 时出错: {e}")
                continue

        return pd.DataFrame(results)

    def _calculate_max_pain(self, group: pd.DataFrame, und_price: float) -> float:
        """计算最大痛点"""
        if und_price <= 0:
            return 0

        strikes = group['strike'].unique()
        if len(strikes) == 0:
            return 0

        min_pain = float('inf')
        max_pain_strike = und_price

        for strike in strikes:
            try:
                calls = group[(group['opt_type'] == 'CALL') & (group['strike'] == strike)]
                puts = group[(group['opt_type'] == 'PUT') & (group['strike'] == strike)]

                call_oi = calls['open_interest'].sum() if not calls.empty else 0
                put_oi = puts['open_interest'].sum() if not puts.empty else 0

                # 计算卖方损失
                call_pain = max(0, und_price - strike) * call_oi
                put_pain = max(0, strike - und_price) * put_oi
                total_pain = call_pain + put_pain

                if total_pain < min_pain:
                    min_pain = total_pain
                    max_pain_strike = strike
            except Exception:
                continue

        return max_pain_strike

    def _classify_trading_type(self, call_change: float, put_change: float, pcr: float) -> str:
        """分类交易类型"""
        if call_change > 0 and put_change > 0:
            if 0.8 <= pcr <= 1.2:
                return '波动率型'
            elif pcr < 0.8:
                return '偏多混合'
            else:
                return '偏空混合'
        elif call_change > put_change:
            return '方向型看多'
        elif put_change > call_change:
            return '方向型看空'
        else:
            return '中性'

    def generate_ranking_report(self, analysis_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        生成多维度排名报告。

        Args:
            analysis_df: 标的分析结果

        Returns:
            各分页的 DataFrame 字典
        """
        sheets = {}

        # 1. 综合排行
        if not analysis_df.empty:
            ranking = analysis_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            ranking.insert(0, '排名', range(1, len(ranking) + 1))
            sheets['期权排行'] = ranking

            # 2. PCR 排行
            pcr_ranking = analysis_df.sort_values('PCR(持仓)').reset_index(drop=True)
            pcr_ranking.insert(0, '排名', range(1, len(pcr_ranking) + 1))
            sheets['期权PCR'] = pcr_ranking[['排名', '标的合约', 'PCR(持仓)', 'PCR(成交)', '情绪倾向']]

            # 3. 最大痛点排行
            pain_ranking = analysis_df.sort_values('痛点距离%', key=abs).reset_index(drop=True)
            pain_ranking.insert(0, '排名', range(1, len(pain_ranking) + 1))
            sheets['期权痛点'] = pain_ranking[['排名', '标的合约', '期货现价', '最大痛点', '痛点距离%']]

            # 4. 资金排行
            capital_ranking = analysis_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            capital_ranking.insert(0, '排名', range(1, len(capital_ranking) + 1))
            sheets['期权资金'] = capital_ranking[['排名', '标的合约', '沉淀资金(亿)', '成交资金(亿)', '总持仓量']]

            # 5. 方向型期权
            directional = analysis_df[analysis_df['交易类型'].str.contains('方向')].reset_index(drop=True)
            if not directional.empty:
                directional.insert(0, '排名', range(1, len(directional) + 1))
                sheets['方向型期权'] = directional

            # 6. 波动率型期权
            volatility = analysis_df[analysis_df['交易类型'].str.contains('波动')].reset_index(drop=True)
            if not volatility.empty:
                volatility.insert(0, '排名', range(1, len(volatility) + 1))
                sheets['波动率型期权'] = volatility

            # 7. 市场概览
            summary = self._generate_market_summary(analysis_df)
            sheets['期权市场'] = summary

        return sheets

    def _generate_market_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """生成市场概览"""
        if df.empty:
            return pd.DataFrame()

        summary_data = [
            {'指标': '标的数量', '数值': len(df), '说明': '有期权数据的标的合约数'},
            {'指标': '总持仓量', '数值': int(df['总持仓量'].sum()), '说明': '所有期权持仓量合计'},
            {'指标': '总成交量', '数值': int(df['总成交量'].sum()), '说明': '所有期权成交量合计'},
            {'指标': '总沉淀资金(亿)', '数值': round(df['沉淀资金(亿)'].sum(), 2), '说明': '所有期权沉淀资金合计'},
            {'指标': '平均PCR', '数值': round(df['PCR(持仓)'].mean(), 4), '说明': '持仓PCR平均值'},
            {'指标': '看多标的数', '数值': len(df[df['情绪倾向'] > 0]), '说明': '情绪倾向看多的标的'},
            {'指标': '看空标的数', '数值': len(df[df['情绪倾向'] < 0]), '说明': '情绪倾向看空的标的'},
        ]

        return pd.DataFrame(summary_data)

    def save_to_excel(self, sheets: Dict[str, pd.DataFrame], output_file: str) -> bool:
        """
        保存到 Excel 文件。

        Args:
            sheets: 各分页数据
            output_file: 输出文件路径

        Returns:
            是否成功
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    if df.empty:
                        continue
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    if OPENPYXL_AVAILABLE:
                        ws = writer.sheets[sheet_name]
                        self._apply_sheet_formatting(ws, df, sheet_name)

            logger.info(f"保存分析报告: {output_file}")
            return True
        except Exception as e:
            logger.error(f"保存文件失败: {e}")
            return False

    def _calculate_display_width(self, s) -> int:
        """计算字符串的显示宽度（中文占2单位，英文占1单位）"""
        if s is None or pd.isna(s):
            return 0
        s = str(s)
        width = 0
        for char in s:
            if ord(char) > 127:  # 中文字符或全角符号
                width += 2
            else:
                width += 1
        return width

    def _apply_sheet_formatting(self, ws, df, sheet_name):
        """应用工作表格式"""
        # 颜色定义
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        purple_fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")

        # 表头样式 - 深蓝色
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 自动列宽
        for idx, col in enumerate(df.columns):
            header_width = self._calculate_display_width(col)
            content_width = df[col].astype(str).map(self._calculate_display_width).max() if len(df) > 0 else 0
            max_len = max(header_width, content_width) + 2
            ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)

        # 条件格式化
        self._apply_conditional_formatting(ws, df, sheet_name, green_fill, red_fill, yellow_fill, blue_fill, purple_fill)

    def _apply_conditional_formatting(self, ws, df, sheet_name, green_fill, red_fill, yellow_fill, blue_fill, purple_fill):
        """应用条件格式化"""
        # 正值绿色，负值红色的列
        red_green_cols = ['持仓变化', 'CALL持仓变化', 'PUT持仓变化', '痛点距离%', '情绪倾向']
        # 评分类 - 高分绿色
        score_cols = ['综合评分', '流动性评分', '活跃度评分', '类型置信度']
        # PCR类 - 低值绿色高值红色
        pcr_cols = ['PCR(持仓)', 'PCR(成交)', 'PCR(资金)']
        # 交易类型列
        trading_type_cols = ['交易类型', '类型细分']

        cols = list(df.columns)

        for r_idx, row in enumerate(df.itertuples(), start=2):
            for col_name in cols:
                if col_name not in df.columns:
                    continue
                col_idx = cols.index(col_name) + 1

                # 安全获取值
                val = None
                try:
                    val = getattr(row, col_name.replace(' ', '_').replace('(', '').replace(')', '').replace('%', ''), None)
                except:
                    pass

                if val is None:
                    continue

                cell = ws.cell(row=r_idx, column=col_idx)

                # 数值类条件格式
                try:
                    num_val = float(val)

                    if col_name in red_green_cols:
                        if num_val > 0:
                            cell.fill = green_fill
                        elif num_val < 0:
                            cell.fill = red_fill
                    elif col_name in score_cols:
                        if num_val >= 70:
                            cell.fill = green_fill
                        elif num_val >= 40:
                            cell.fill = yellow_fill
                        elif num_val < 40:
                            cell.fill = red_fill
                    elif col_name in pcr_cols:
                        if num_val < 0.8:
                            cell.fill = green_fill
                        elif num_val > 1.2:
                            cell.fill = red_fill
                        else:
                            cell.fill = yellow_fill
                except (ValueError, TypeError):
                    pass

                # 字符串类条件格式
                str_val = str(val)
                if col_name in trading_type_cols:
                    if '方向' in str_val or '看多' in str_val or '看空' in str_val:
                        cell.fill = red_fill
                    elif '波动率' in str_val or '跨式' in str_val:
                        cell.fill = blue_fill
                    elif '混合' in str_val:
                        cell.fill = purple_fill

    def generate_option_reference(self, raw_df: pd.DataFrame) -> bool:
        """
        生成期权参考数据（包含 IV 和 Greeks）。

        Args:
            raw_df: 原始期权数据

        Returns:
            是否成功
        """
        logger.info("开始生成期权参考数据...")

        if raw_df.empty:
            logger.warning("没有期权数据，无法生成参考数据")
            return False

        all_ref_data = []

        for _, row in raw_df.iterrows():
            try:
                symbol = row.get('instrument_id', row.get('symbol', ''))
                underlying = row.get('underlying_symbol', '')
                strike = row.get('strike_price', 0) or 0
                opt_type = str(row.get('option_class', row.get('call_or_put', ''))).upper()

                if 'CALL' in opt_type or opt_type == 'C':
                    opt_type = 'CALL'
                elif 'PUT' in opt_type or opt_type == 'P':
                    opt_type = 'PUT'
                else:
                    continue

                # 期权价格
                opt_price = row.get('last_price', 0) or row.get('pre_close', 0) or 0
                multiplier = row.get('volume_multiple', 1) or 1
                expire_days = row.get('expire_rest_days', 1) or 1
                expire_date = row.get('expire_datetime', '')

                # 标的期货数据
                und_price = self._future_prices.get(underlying, 0)
                if und_price <= 0:
                    continue

                future_multiplier = self._future_multipliers.get(underlying, 1)
                p_name = self._future_names.get(underlying, '')

                # 品种识别
                if '.' in underlying:
                    p_code = re.sub(r'[^a-zA-Z]', '', underlying.split('.')[1])
                else:
                    p_code = re.sub(r'[^a-zA-Z]', '', underlying)

                p_id = p_code.upper()
                margin_ratio = self._get_margin_ratio(p_id) * 100  # 转为百分比

                # 虚实幅度计算
                if opt_type == 'CALL':
                    intrinsic_degree = (und_price - strike) / und_price * 100
                    otm_val = max((strike - und_price) * multiplier, 0)
                    intrinsic_val = max(0, und_price - strike)
                    premium_rate = (strike + opt_price - und_price) / und_price * 100
                else:
                    intrinsic_degree = (strike - und_price) / und_price * 100
                    otm_val = max((und_price - strike) * multiplier, 0)
                    intrinsic_val = max(0, strike - und_price)
                    premium_rate = (und_price + opt_price - strike) / und_price * 100

                # 虚实档位分类
                if intrinsic_degree > 20:
                    intrinsic_level = "深度实值"
                elif 10 < intrinsic_degree <= 20:
                    intrinsic_level = "中度实值"
                elif -10 <= intrinsic_degree <= 10:
                    intrinsic_level = "平值附近"
                elif -20 <= intrinsic_degree < -10:
                    intrinsic_level = "中度虚值"
                else:
                    intrinsic_level = "深度虚值"

                # 保证金计算
                und_margin = und_price * future_multiplier * (margin_ratio / 100)
                margin = (opt_price * multiplier) + max(und_margin - otm_val / 2, und_margin / 2)

                # 收益率计算
                profit_rate = (opt_price / und_price * 100) if und_price > 0 else 0
                annual_profit_rate = profit_rate / max(expire_days, 1) * 365
                leverage_profit = (opt_price * multiplier / margin * 100) if margin > 0 else 0
                annual_leverage_profit = leverage_profit / max(expire_days, 1) * 365

                # 价值分解
                time_value = max(0, opt_price - intrinsic_val)
                time_ratio = (time_value / opt_price * 100) if opt_price > 0 else 0

                # 到期日格式化
                expire_date_str = ''
                if expire_date:
                    if isinstance(expire_date, str):
                        expire_date_str = expire_date.split(' ')[0]
                    elif isinstance(expire_date, datetime):
                        expire_date_str = expire_date.strftime('%Y-%m-%d')

                # 资金计算
                opt_fee = opt_price * multiplier
                oi = row.get('open_interest', 0) or 0
                pre_oi = row.get('pre_open_interest', 0) or 0
                amount = row.get('amount', 0) or 0

                chendian_wan = (oi * opt_fee / 10000)
                chendian_chg_wan = ((oi - pre_oi) * opt_fee / 10000)
                chengjiao_wan = (amount / 10000)
                zijin_total_wan = chendian_wan + chengjiao_wan

                # 期货杠杆
                futures_leverage = round(100.0 / margin_ratio, 2) if margin_ratio > 0 else 0

                # 交割年月
                exercise_year = row.get('exercise_year', 0)
                exercise_month = row.get('exercise_month', 0)
                delivery_ym = f"{int(exercise_year)}{int(exercise_month):02d}" if exercise_year else ''

                ref_row = {
                    '交易所': row.get('exchange_id', ''),
                    '合约代码': symbol,
                    '合约名称': row.get('instrument_name', ''),
                    '期权类型': opt_type,
                    '标的合约': underlying,
                    '标的品种名称': p_name,
                    '标的现价': round(und_price, 2),
                    '期货合约乘数': future_multiplier,
                    '期货保证金率%': round(margin_ratio, 2),
                    '期货杠杆': futures_leverage,
                    '行权价': strike,
                    '期权价': round(opt_price, 4),
                    '虚实幅度%': round(intrinsic_degree, 2),
                    '虚实档位': intrinsic_level,
                    '内在价值': round(intrinsic_val, 4),
                    '溢价率%': round(premium_rate, 2),
                    '时间价值': round(time_value, 4),
                    '时间占比%': round(time_ratio, 2),
                    '剩余天数': expire_days,
                    '买方期权费': round(opt_fee, 2),
                    '标的期货保证金': round(und_margin, 2),
                    '卖方保证金': round(margin, 2),
                    '收益%': round(profit_rate, 4),
                    '收益年化%': round(annual_profit_rate, 2),
                    '杠杆收益%': round(leverage_profit, 2),
                    '杠杆年化%': round(annual_leverage_profit, 2),
                    '到期日': expire_date_str,
                    '交割年月': delivery_ym,
                    '昨收': row.get('pre_close', 0),
                    '今结': row.get('settlement', 0),
                    '昨结': row.get('pre_settlement', 0),
                    '成交量': row.get('volume', 0),
                    '成交金额': amount,
                    '持仓量': oi,
                    '昨持仓量': pre_oi,
                    '合约乘数': multiplier,
                    '最小跳动': row.get('price_tick', 0),
                    '沉淀资金(万)': round(chendian_wan, 2),
                    '沉淀资金变化(万)': round(chendian_chg_wan, 2),
                    '成交资金(万)': round(chengjiao_wan, 2),
                    '资金合计(万)': round(zijin_total_wan, 2),
                }
                all_ref_data.append(ref_row)
            except Exception as e:
                logger.debug(f"处理期权时出错: {e}")
                continue

        if not all_ref_data:
            logger.warning("没有整理出任何期权参考数据")
            return False

        ref_df = pd.DataFrame(all_ref_data)

        # 按资金合计排序
        ref_df = ref_df.sort_values(by='资金合计(万)', ascending=False).reset_index(drop=True)

        # ============ 计算 IV 和 Greeks ============
        logger.info("计算隐含波动率和 Greeks...")
        start_time = time.time()

        # 获取历史波动率
        underlyings = ref_df['标的合约'].unique().tolist()
        hv_map = self._get_historical_volatility(underlyings)

        # 准备向量化数据
        S = ref_df['标的现价'].values.astype(np.float64)
        P = ref_df['期权价'].values.astype(np.float64)
        K = ref_df['行权价'].values.astype(np.float64)

        # 映射历史波动率
        HV5_array = ref_df['标的合约'].map(lambda x: hv_map.get(x, {}).get('hv5', 0.25)).values.astype(np.float64)
        HV20_array = ref_df['标的合约'].map(lambda x: hv_map.get(x, {}).get('hv20', 0.25)).values.astype(np.float64)
        HV60_array = ref_df['标的合约'].map(lambda x: hv_map.get(x, {}).get('hv60', 0.25)).values.astype(np.float64)

        # 使用 HV20 作为默认波动率
        HV_array = HV20_array

        # 期权类型
        Type = ref_df['期权类型'].values

        # 到期时间 (年化)
        T = ref_df['剩余天数'].values.astype(np.float64) / 365.0
        T = np.maximum(T, 0.0001)

        r = 0.015  # 无风险利率 1.5%

        # 计算理论价格
        bs_price = self.iv_calculator.bs_price(S, K, r, HV_array, T, Type)

        # 计算隐含波动率
        impv = self.iv_calculator.implied_volatility(P, S, K, T, Type, init_sigma=HV_array)
        impv = np.where(np.isnan(impv), 0.0, impv)
        impv = np.maximum(impv, 0.0)
        impv = np.where((impv > 0) & (impv < 0.005), 0.005, impv)

        # 计算 Greeks (使用至少 1% 的波动率)
        impv_for_greeks = np.where(impv < 0.01, HV_array, impv)
        impv_for_greeks = np.maximum(impv_for_greeks, 0.01)

        delta = self.iv_calculator.delta(S, K, r, impv_for_greeks, T, Type)
        gamma = self.iv_calculator.gamma(S, K, r, impv_for_greeks, T)
        theta = self.iv_calculator.theta(S, K, r, impv_for_greeks, T, Type)
        vega = self.iv_calculator.vega(S, K, r, impv_for_greeks, T)
        rho = self.iv_calculator.rho(S, K, r, impv_for_greeks, T, Type)

        # 处理 NaN 和 Inf
        delta = np.nan_to_num(delta, nan=0.0, posinf=1.0, neginf=-1.0)
        gamma = np.nan_to_num(gamma, nan=0.0, posinf=0.0, neginf=0.0)
        theta = np.nan_to_num(theta, nan=0.0, posinf=0.0, neginf=0.0)
        vega = np.nan_to_num(vega, nan=0.0, posinf=0.0, neginf=0.0)
        rho = np.nan_to_num(rho, nan=0.0, posinf=0.0, neginf=0.0)

        # 添加列
        ref_df['理论价格'] = np.round(bs_price, 2)
        ref_df['近期波动率'] = np.round(HV5_array * 100, 2)
        ref_df['HV20'] = np.round(HV20_array * 100, 2)
        ref_df['HV60'] = np.round(HV60_array * 100, 2)
        ref_df['隐含波动率'] = np.round(impv * 100, 2)
        ref_df['Delta'] = np.round(delta, 4)
        ref_df['Gamma'] = np.round(gamma, 6)
        ref_df['Theta'] = np.round(theta / 365.0, 4)
        ref_df['Vega'] = np.round(vega / 100.0, 4)
        ref_df['Rho'] = np.round(rho / 100.0, 4)

        # 买方杠杆 = 标的现价 * 期货合约乘数 / (期权价 * 合约乘数) * Delta
        futures_multiplier_arr = ref_df['期货合约乘数'].values.astype(np.float64)
        option_multiplier_arr = ref_df['合约乘数'].values.astype(np.float64)
        with np.errstate(divide='ignore', invalid='ignore'):
            buyer_leverage = np.round((S * futures_multiplier_arr) / (P * option_multiplier_arr) * delta, 2)
            buyer_leverage = np.where(np.isinf(buyer_leverage) | np.isnan(buyer_leverage), 0, buyer_leverage)

        # 插入买方杠杆列
        fee_idx = ref_df.columns.get_loc('买方期权费') + 1
        ref_df.insert(fee_idx, '买方杠杆', buyer_leverage)

        logger.info(f"IV 和 Greeks 计算完成，耗时: {time.time() - start_time:.2f}秒")

        # 保存到 Excel
        try:
            with pd.ExcelWriter(self.reference_file, engine='openpyxl') as writer:
                ref_df.to_excel(writer, sheet_name='期权参考', index=False)

                if OPENPYXL_AVAILABLE:
                    ws = writer.sheets['期权参考']

                    # 表头样式 - 深蓝色
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)

                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    ws.freeze_panes = 'A2'

                    # 定义样式
                    atm_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    itm_fills = [
                        PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                        PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
                        PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
                    ]
                    otm_fills = [
                        PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"),
                        PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"),
                        PatternFill(start_color="81D4FA", end_color="81D4FA", fill_type="solid"),
                    ]
                    expire_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    expire_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    expire_dark_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

                    cols = list(ref_df.columns)
                    symbol_idx = cols.index('合约代码') + 1
                    strike_idx = cols.index('行权价') + 1
                    degree_idx = cols.index('虚实幅度%') + 1
                    expire_idx = cols.index('到期日') + 1
                    days_idx = cols.index('剩余天数') + 1

                    for r_idx, row_data in enumerate(ref_df.itertuples(), start=2):
                        degree = getattr(row_data, '虚实幅度%', 0)
                        days = getattr(row_data, '剩余天数', 0)

                        # 平值高亮
                        if abs(degree) <= 2.0:
                            ws.cell(row=r_idx, column=symbol_idx).fill = atm_fill

                        # 行权价着色
                        if degree > 0:
                            fill_idx = min(2, int(degree / 5))
                            ws.cell(row=r_idx, column=strike_idx).fill = itm_fills[fill_idx]
                        elif degree < 0:
                            abs_degree = abs(degree)
                            fill_idx = min(2, int(abs_degree / 5))
                            ws.cell(row=r_idx, column=strike_idx).fill = otm_fills[fill_idx]

                        # 到期日提醒
                        if days <= 1:
                            ws.cell(row=r_idx, column=expire_idx).fill = expire_dark_red
                        elif days <= 3:
                            ws.cell(row=r_idx, column=expire_idx).fill = expire_red
                        elif days <= 7:
                            ws.cell(row=r_idx, column=expire_idx).fill = expire_yellow

                    # 自动调整列宽
                    def get_visual_width(s):
                        if pd.isna(s):
                            return 0
                        s = str(s)
                        w = 0
                        for char in s:
                            if ord(char) > 127:
                                w += 2
                            else:
                                w += 1
                        return w

                    for idx, col in enumerate(ref_df.columns):
                        header_w = get_visual_width(col)
                        content_w = ref_df[col].head(500).astype(str).apply(get_visual_width).max() if not ref_df.empty else 0
                        max_w = max(header_w, content_w) + 2
                        ws.column_dimensions[get_column_letter(idx + 1)].width = min(max(max_w, 10), 40)

            logger.info(f"期权参考数据生成完成: {self.reference_file}, 共 {len(ref_df)} 条")
            return True

        except Exception as e:
            logger.error(f"保存期权参考文件失败: {e}")
            return False

    def run(self) -> bool:
        """
        执行完整分析流程。

        Returns:
            是否成功
        """
        logger.info("=" * 60)
        logger.info("开始期权综合分析...")
        logger.info("=" * 60)

        # 1. 加载期货价格
        self._load_future_prices()

        # 2. 加载期权数据
        options_df = self._load_option_data()
        if options_df.empty:
            logger.error("未找到有效期权数据")
            return False

        # 3. 标准化数据
        std_df = self._standardize_option_data(options_df)
        if std_df.empty:
            logger.error("标准化后无有效数据")
            return False
        logger.info(f"标准化完成，有效期权: {len(std_df)} 个")

        # 4. 按标的合约分析
        analysis_df = self.analyze_by_underlying(std_df)
        if analysis_df.empty:
            logger.error("分析结果为空")
            return False
        logger.info(f"分析完成，标的合约: {len(analysis_df)} 个")

        # 5. 生成排名报告
        sheets = self.generate_ranking_report(analysis_df)

        # 6. 保存排行到 Excel
        success = self.save_to_excel(sheets, self.output_file)

        # 7. 生成期权参考数据
        ref_success = self.generate_option_reference(options_df)

        if success:
            logger.info("=" * 60)
            logger.info("期权综合分析完成!")
            logger.info(f"输出文件: {self.output_file}")
            logger.info(f"包含 {len(sheets)} 个分析维度")
            if ref_success:
                logger.info(f"参考数据: {self.reference_file}")
            logger.info("=" * 60)

        return success


def main():
    """命令行入口"""
    runner = OptionAnalysisRunner()
    success = runner.run()
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())