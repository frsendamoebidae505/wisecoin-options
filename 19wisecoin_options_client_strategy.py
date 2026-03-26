#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WiseCoin 期权策略生成器
生成除套利外的所有期权策略建议，包括：
- 日历价差（Calendar Spread）
- 铁鹰式（Iron Condor）
- 蝶式（Butterfly）
- 垂直价差（Vertical Spread）
- 比率价差（Ratio Spread）
- 跨式/宽跨式（Straddle/Strangle）
- 备兑/保护性策略（Covered Call/Protective Put）

基于波动率曲面、Greeks、市场情绪等多维度分析
"""

import pandas as pd
import numpy as np
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

class OptionStrategyGenerator:
    """期权策略生成器"""
    
    def __init__(self):
        self.option_ref_df = None
        self.market_overview_df = None  # 市场概览-货权联动
        self.vol_surface_df = None      # 期权参考-波动率曲面
        
        # 市场概览扩展（可选）
        self.vol_option_overview_df = None  # 市场概览-波动率型期权
        self.pain_overview_df = None        # 市场概览-期权痛点
        
        # 行情数据（按需加载）
        self.option_quotes_xls = None
        self.futures_quotes_xls = None
        self.futures_quotes_df = None
        self.futures_quotes_index = None
        
        # sheet 缓存（避免重复读Excel）
        self._option_quote_sheet_cache: dict[str, pd.DataFrame] = {}
        self._option_quote_row_cache = {}
        self._option_quote_sheet_lookup = {}
        
        # 策略指南（内置数据）
        self.strategy_guide_dict = self._init_strategy_guide()
        
        # 数据路径
        self.data_dir = 'wisecoin_options_client_live_temp'
        if not os.path.exists(self.data_dir):
            self.data_dir = '.'

    def _init_strategy_guide(self):
        """初始化策略指南（内置数据）"""
        strategy_guide = {
            '买入看涨期权': {
                '策略分类': '一、单一维度策略（纯方向/纯波动率）',
                '策略名称': '买入看涨期权',
                '市场三维观点（方向+波动率+时间）': '强烈看涨，波动率中性/看涨，无时间执念',
                '核心构建（期货期权专属）': '买入对应期货主力合约 平值/轻度虚值Call期权，纯权利仓',
                '盈亏特征+核心希腊字母暴露': '盈利：期货上涨无上限；亏损：固定为权利金希腊字母：+Delta、+Gamma、+Vega、-Theta',
                '期货实战开仓时机': '1. 期货技术面突破前高/均线关键压力；2. 基本面利好（减产/库存降）；3. 隐含波动率（IV）低位，权利金性价比高；4. 趋势启动初期',
                '核心损益驱动源': '1. 主驱动：Delta正向变动（期货上涨）；2. 次驱动：Vega上升（波动率抬升）',
                '补充策略要点（实操必看）': '1. 优先选平值期权，Delta≈0.5，涨跌敏感度适中；2. 到期日选1-3个月，平衡时间损耗与趋势空间'
            },
            '买入看跌期权': {
                '策略分类': '一、单一维度策略（纯方向/纯波动率）',
                '策略名称': '买入看跌期权',
                '市场三维观点（方向+波动率+时间）': '强烈看跌，波动率中性/看涨，无时间执念',
                '核心构建（期货期权专属）': '买入对应期货主力合约 平值/轻度虚值Put期权，纯权利仓',
                '盈亏特征+核心希腊字母暴露': '盈利：期货下跌无上限；亏损：固定为权利金希腊字母：-Delta、+Gamma、+Vega、-Theta',
                '期货实战开仓时机': '1. 期货技术面跌破前低/均线关键支撑；2. 基本面利空（增产/需求弱）；3. IV低位，权利金成本低；4. 下跌破位初期',
                '核心损益驱动源': '1. 主驱动：Delta负向变动（期货下跌）；2. 次驱动：Vega上升（波动率抬升）',
                '补充策略要点（实操必看）': '1. 规避深度虚值Put，防止行情下跌但期权不涨；2. 商品期货熊市中，可搭配主力合约换月'
            },
            '买入跨式期权': {
                '策略分类': '一、单一维度策略（纯方向/纯波动率）',
                '策略名称': '买入跨式期权',
                '市场三维观点（方向+波动率+时间）': '方向完全不定，波动率大幅上升，时间偏中期',
                '核心构建（期货期权专属）': '同时买入 同标的/同到期日/平值Call+平值Put，双权利仓',
                '盈亏特征+核心希腊字母暴露': '盈利：期货单边大涨/大跌均无上限；亏损：固定为双份权利金总和希腊字母：≈Delta中性、+Gamma、+Vega、-Theta',
                '期货实战开仓时机': '1. 重大事件前夕（EIA/USDA报告、政策决议）；2. IV处于历史低位，波动率修复预期强；3. 期货长期横盘蓄势，即将突破',
                '核心损益驱动源': '1. 主驱动：Vega大幅上升（波动率拉升）；2. 次驱动：Gamma获利（价格快速波动）',
                '补充策略要点（实操必看）': '1. 宽跨式可替代跨式，买虚值Call+Put，降低权利金成本；2. 事件落地后波动率回落，需及时止盈'
            },
            '卖出裸看涨/裸看跌': {
                '策略分类': '一、单一维度策略（纯方向/纯波动率）',
                '策略名称': '卖出裸看涨/裸看跌（极高风险 新手禁用）',
                '市场三维观点（方向+波动率+时间）': '卖出裸Call：强烈看跌+波动率下降卖出裸Put：强烈看涨+波动率下降',
                '核心构建（期货期权专属）': '无持仓对冲，直接卖出期货对应Call/Put期权，纯义务仓',
                '盈亏特征+核心希腊字母暴露': '盈利：固定为权利金，价格不触碰行权价即赚；亏损：理论无限（期货单边大涨/大跌）希腊字母（裸Call）：-Delta、-Gamma、-Vega、+Theta希腊字母（裸Put）：+Delta、-Gamma、-Vega、+Theta',
                '期货实战开仓时机': '1. IV处于历史极端高位，波动率大概率回落；2. 对期货方向判断100%笃定，无突发风险；3. 期货处于极端高位（裸Call）/低位（裸Put）',
                '核心损益驱动源': '主驱动：Theta时间价值衰减，波动率（Vega）下降辅助获利',
                '补充策略要点（实操必看）': '1. 必须缴纳高额保证金，随期货波动实时追加；2. 严格设置止损，轻仓操作，杜绝扛单'
            },
            '牛市价差': {
                '策略分类': '二、双维度策略（方向+时间/波动率）',
                '策略名称': '牛市价差（看涨型）',
                '市场三维观点（方向+波动率+时间）': '温和看涨，波动率中性/下降，时间偏短期',
                '核心构建（期货期权专属）': '买入1手低行权价Call + 卖出1手同标的/同到期日高行权价Call，成对开仓',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（高行权价-低行权价-净权利金）；亏损：有限（固定为净权利金）希腊字母：+Delta、-Vega、+Theta（净）',
                '期货实战开仓时机': '1. 期货震荡上行，上方有明确阻力位（前高/均线）；2. 不想承担单边看涨高成本，追求高胜率；3. 波动率偏高，卖出高行权Call抵消部分成本',
                '核心损益驱动源': '1. 主驱动：Delta温和正向变动（期货缓涨）；2. 次驱动：Theta时间衰减获利',
                '补充策略要点（实操必看）': '1. 行权价价差适中，不宜过宽/过窄，平衡盈亏比；2. 到期前期货未达高行权价，即可全额获利'
            },
            '熊市价差': {
                '策略分类': '二、双维度策略（方向+时间/波动率）',
                '策略名称': '熊市价差（看跌型）',
                '市场三维观点（方向+波动率+时间）': '温和看跌，波动率中性/下降，时间偏短期',
                '核心构建（期货期权专属）': '买入1手高行权价Put + 卖出1手同标的/同到期日低行权价Put，成对开仓',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（高行权价-低行权价-净权利金）；亏损：有限（固定为净权利金）希腊字母：-Delta、-Vega、+Theta（净）',
                '期货实战开仓时机': '1. 期货震荡下行，下方有明确支撑位（前低/成本线）；2. 规避单边看跌高成本，稳健博弈回调；3. 波动率偏高，卖出低行权Put降低成本',
                '核心损益驱动源': '1. 主驱动：Delta温和负向变动（期货缓跌）；2. 次驱动：Theta时间衰减获利',
                '补充策略要点（实操必看）': '1. 避免深度虚值行权价组合，防止行情下跌不达标；2. 震荡市胜率＞单边市，适合波段操作'
            },
            '备兑看涨期权': {
                '策略分类': '二、双维度策略（方向+时间/波动率）',
                '策略名称': '备兑看涨期权',
                '市场三维观点（方向+波动率+时间）': '持有期货多头，温和看涨/横盘，波动率中性/偏高，赚时间收益',
                '核心构建（期货期权专属）': '持有足额期货多头头寸 + 卖出1手同标的/同到期日虚值Call期权',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（权利金+期货涨至行权价的收益，行权价封顶）；亏损：期货下跌无限-已收权利金（风险部分对冲）希腊字母：+Delta（期货）、-Gamma、-Vega、+Theta',
                '期货实战开仓时机': '1. 已持有期货多头，浮盈小幅盈利，预期短期横盘/缓涨；2. IV偏高，卖出Call权利金收益可观；3. 不想卖出期货，想盘活持仓增厚收益',
                '核心损益驱动源': '主驱动：Theta时间价值衰减，赚取权利金；期货小幅上涨辅助获利',
                '补充策略要点（实操必看）': '1. 优先选虚值1-2档Call，给期货上涨留空间，降低被行权概率；2. 期货大涨可平仓Call，保留多头上行收益'
            },
            '保护性看跌期权': {
                '策略分类': '二、双维度策略（方向+时间/波动率）',
                '策略名称': '保护性看跌期权',
                '市场三维观点（方向+波动率+时间）': '持有期货多头，看涨不变，防范短期回调/黑天鹅，波动率中性',
                '核心构建（期货期权专属）': '持有足额期货多头头寸 + 买入1手同标的/同到期日平值/轻度虚值Put期权',
                '盈亏特征+核心希腊字母暴露': '盈利：期货上涨无上限（收益-权利金成本）；亏损：有限（锁定至Put行权价，亏损=期货持仓亏损+权利金）希腊字母：+Delta（期货）、+Gamma（期权）、+Vega',
                '期货实战开仓时机': '1. 期货多头已有大额浮盈，担心突发利空回调；2. 市场情绪不稳，黑天鹅风险高（地缘/疫情）；3. 愿意支付少量权利金，给持仓买保险',
                '核心损益驱动源': '1. 主驱动：期货上涨Delta获利；2. 次驱动：Vega上升（波动率上涨，Put期权增值）',
                '补充策略要点（实操必看）': '1. 虚值Put成本低，平值Put保护力度强，按需选择；2. 行情平稳后可平仓Put，收回部分权利金'
            },
            '领口策略': {
                '策略分类': '二、双维度策略（方向+时间/波动率）',
                '策略名称': '领口策略',
                '市场三维观点（方向+波动率+时间）': '持有期货多头，完全对冲风险，波动率中性，低成本避险',
                '核心构建（期货期权专属）': '备兑看涨+保护性看跌组合：持有期货多头+卖出虚值Call+买入虚值Put（Call权利金覆盖部分/全部Put成本）',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（Call行权价封顶，扣除Put成本）；亏损：有限（Put行权价封底）希腊字母：≈Delta中性、低Vega、Theta可变',
                '期货实战开仓时机': '1. 期货持仓金额大，需完全锁定上下风险；2. 不想支付高额保护成本，用卖Call权利金补贴买Put；3. 长期持有期货，忽略短期波动',
                '核心损益驱动源': '主驱动：完全对冲多空风险，赚取期货持仓稳定收益',
                '补充策略要点（实操必看）': '1. 尽量做到零成本领口（卖Call权利金=买Put成本）；2. 放弃期货超额上涨收益，换取下行绝对安全'
            },
            '铁鹰式价差': {
                '策略分类': '三、核心三维策略（中性+时间+波动率）',
                '策略名称': '铁鹰式价差',
                '市场三维观点（方向+波动率+时间）': '价格区间震荡，波动率下降，时间偏中期',
                '核心构建（期货期权专属）': '4手期权组合（同标的/同到期日）：卖出1手虚值Call+买入1手更虚值Call；卖出1手虚值Put+买入1手更虚值Put',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（卖出Call+Put权利金总和-买入Call+Put权利金总和）；亏损：有限（价差宽度-净权利金）希腊字母：≈Delta中性、-Gamma、-Vega、+Theta',
                '期货实战开仓时机': '1. 期货有清晰上下支撑/压力区间，短期无趋势；2. IV处于历史高位，预期波动率回落；3. 重大事件落地后，市场回归平静',
                '核心损益驱动源': '1. 主驱动：Theta时间价值衰减；2. 次驱动：Vega下降（波动率回落）',
                '补充策略要点（实操必看）': '1. 区间宽度设置合理，参考期货近期震荡幅度；2. 价格突破区间边缘，及时平仓止损'
            },
            '铁蝶式价差': {
                '策略分类': '三、核心三维策略（中性+时间+波动率）',
                '策略名称': '铁蝶式价差',
                '市场三维观点（方向+波动率+时间）': '价格绝对静止，波动率下降，时间偏短期',
                '核心构建（期货期权专属）': '4手期权组合（同标的/同到期日）：卖出1手平值Call+买入1手虚值Call；卖出1手平值Put+买入1手虚值Put（共用平值行权价）',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（收益高于铁鹰）；亏损：有限（亏损幅度窄于铁鹰）希腊字母：≈Delta中性、-Gamma、-Vega、+Theta',
                '期货实战开仓时机': '1. IV偏高，且坚信期货到期前将围绕平值行权价窄幅震荡；2. 期货无重大事件驱动，交投清淡；3. 对震荡区间判断精准',
                '核心损益驱动源': '1. 主驱动：Theta时间价值快速衰减；2. 次驱动：Vega下降（波动率回落）',
                '补充策略要点（实操必看）': '1. 盈利空间集中，到期价格越贴近平值，收益越高；2. 流动性优先，选主力合约平值期权'
            },
            '比率价差': {
                '策略分类': '三、核心三维策略（中性+时间+波动率）',
                '策略名称': '（卖出）比率价差（看涨/看跌）',
                '市场三维观点（方向+波动率+时间）': '看涨比率：温和看涨+强烈看空波动率看跌比率：温和看跌+强烈看空波动率',
                '核心构建（期货期权专属）': '看涨比率：买入1手平值Call + 卖出2手同标的/同到期日更高行权价虚值Call看跌比率：买入1手平值Put + 卖出2手同标的/同到期日更低行权价虚值Put',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（价格至目标行权价时收益最大）；亏损：一侧风险扩大（看涨比率：期货大跌小亏、大涨超目标价盈利回落）希腊字母：轻微+/-Delta、-Gamma、-Vega、+Theta',
                '期货实战开仓时机': '1. IV处于高位，预期波动率将下降；2. 精准判断期货将缓慢涨/跌至目标价位，且停滞不前；3. 期货无极端行情预期',
                '核心损益驱动源': '1. 主驱动：Theta时间衰减；2. 次驱动：Vega下降（波动率回落）',
                '补充策略要点（实操必看）': '1. 优先Delta中性开仓，降低方向风险；2. 严控仓位，价格突破目标位及时减仓'
            },
            '日历价差': {
                '策略分类': '三、核心三维策略（中性+时间+波动率）',
                '策略名称': '日历价差',
                '市场三维观点（方向+波动率+时间）': '短期平静、长期有趋势，波动率期限结构陡峭，赚时间差',
                '核心构建（期货期权专属）': '卖出1手近月平值期权（Call/Put均可） + 买入1手远月同标的/同行权价期权',
                '盈亏特征+核心希腊字母暴露': '盈利：有限（赚近月时间价值衰减＞远月）；亏损：有限（固定为净权利金）希腊字母：≈Delta中性、-Gamma（近月）、+Vega（远月）、+Theta（净）',
                '期货实战开仓时机': '1. IV期限结构陡峭（近月IV＞远月IV）；2. 预期期货近月到期前，围绕行权价窄幅震荡；3. 期货季度合约切换期，无短期重大事件',
                '核心损益驱动源': '主驱动：近月期权Theta时间价值快速衰减',
                '补充策略要点（实操必看）': '1. 优先选Call构建，商品期货多头氛围下流动性更佳；2. 近月到期前，若期货突破震荡区间，及时平仓'
            },
            '对角价差': {
                '策略分类': '三、核心三维策略（中性+时间+波动率）',
                '策略名称': '对角价差',
                '市场三维观点（方向+波动率+时间）': '短期有方向/区间、长期观点不同，优化风险收益比',
                '核心构建（期货期权专属）': '卖出1手近月期权 + 买入1手远月同标的/不同行权价期权',
                '盈亏特征+核心希腊字母暴露': '盈亏特征：灵活，随行权价选择调整希腊字母：灵活，随构建方式变动',
                '期货实战开仓时机': '1. 对市场多周期观点分歧（如短期看震荡、长期看涨）；2. 想兼顾垂直价差方向性+日历价差时间收益；3. IV适中，无极端波动预期',
                '核心损益驱动源': '驱动源：结合垂直价差Delta方向收益+日历价差时间衰减收益',
                '补充策略要点（实操必看）': '1. 行权价与到期日灵活搭配，适配自身交易观点；2. 定期调整仓位，维持风险收益平衡'
            }
        }
        return strategy_guide
    
    def _resolve_data_file(self, filename: str):
        """优先从 data_dir 获取文件，fallback 到当前目录，并过滤 0B 文件。"""
        candidates = [os.path.join(self.data_dir, filename), filename]
        for p in candidates:
            try:
                if os.path.exists(p) and os.path.getsize(p) > 0:
                    return p
            except Exception:
                continue
        return None

    def _normalize_option_type(self, opt_type):
        if opt_type is None:
            return opt_type
        if pd.isna(opt_type):
            return opt_type
        opt_str = str(opt_type).strip().upper()
        if opt_str in ('CALL', 'C', '认购', '看涨'):
            return 'CALL'
        if opt_str in ('PUT', 'P', '认沽', '看跌'):
            return 'PUT'
        return opt_str
            
    def load_data(self):
        """加载所有必要数据"""
        print("=" * 80)
        print("WiseCoin 期权策略生成器 v2.0")
        print("=" * 80)
        print("\n正在加载数据...")
        
        try:
            # 策略指南已在__init__中初始化
            print(f"✓ 已加载策略指南：{len(self.strategy_guide_dict)} 条策略（内置）")
            
            # 1) 期权参考数据
            option_ref_file = self._resolve_data_file('wisecoin-期权参考.xlsx')
            if not option_ref_file:
                raise FileNotFoundError('未找到 wisecoin-期权参考.xlsx（或文件为0B）')
            
            self.option_ref_df = pd.read_excel(option_ref_file, sheet_name='期权参考')
            if '期权类型' in self.option_ref_df.columns:
                self.option_ref_df['期权类型'] = self.option_ref_df['期权类型'].apply(self._normalize_option_type)
            print(f"✓ 已加载期权参考数据：{len(self.option_ref_df)} 条记录")
            
            # 2) 市场概览数据
            market_file = self._resolve_data_file('wisecoin-市场概览.xlsx')
            if not market_file:
                raise FileNotFoundError('未找到 wisecoin-市场概览.xlsx（或文件为0B）')
            
            self.market_overview_df = pd.read_excel(market_file, sheet_name='货权联动')
            print(f"✓ 已加载市场概览数据(货权联动)：{len(self.market_overview_df)} 条记录")
            
            # 市场概览扩展：波动率型期权、期权痛点
            try:
                self.vol_option_overview_df = pd.read_excel(market_file, sheet_name='波动率型期权')
                print(f"✓ 已加载市场概览(波动率型期权)：{len(self.vol_option_overview_df)} 条记录")
            except Exception:
                self.vol_option_overview_df = None
            
            try:
                self.pain_overview_df = pd.read_excel(market_file, sheet_name='期权痛点')
                print(f"✓ 已加载市场概览(期权痛点)：{len(self.pain_overview_df)} 条记录")
            except Exception:
                self.pain_overview_df = None
            
            # 3) 波动率曲面数据
            try:
                self.vol_surface_df = pd.read_excel(option_ref_file, sheet_name='波动率曲面')
                print(f"✓ 已加载波动率曲面数据：{len(self.vol_surface_df)} 条记录")
            except Exception:
                print("⚠ 未找到波动率曲面数据，将使用默认参数")
                self.vol_surface_df = None
            
            # 4) 期权行情数据（盘口 bid/ask）
            quotes_file = self._resolve_data_file('wisecoin-期权行情.xlsx')
            if quotes_file:
                try:
                    self.option_quotes_xls = pd.ExcelFile(quotes_file)
                    print(f"✓ 已加载期权行情Excel索引：{os.path.basename(quotes_file)}")
                except Exception:
                    self.option_quotes_xls = None
                    print("⚠ 期权行情文件存在但无法读取，将回退使用期权参考价格")
            else:
                self.option_quotes_xls = None
                print("⚠ 未找到可用的期权行情数据，将使用期权参考价格")
            
            # 5) 期货行情数据（标的现价/bid/ask）
            fut_file = self._resolve_data_file('wisecoin-期货行情.xlsx')
            if fut_file:
                try:
                    self.futures_quotes_xls = pd.ExcelFile(fut_file)
                    self.futures_quotes_df = pd.read_excel(self.futures_quotes_xls, sheet_name='Summary')
                    if 'instrument_id' in self.futures_quotes_df.columns:
                        self.futures_quotes_index = self.futures_quotes_df.set_index('instrument_id', drop=False)
                    print(f"✓ 已加载期货行情(Summary)：{len(self.futures_quotes_df)} 条记录")
                except Exception:
                    self.futures_quotes_xls = None
                    self.futures_quotes_df = None
                    self.futures_quotes_index = None
                    print("⚠ 期货行情文件存在但无法读取，将回退使用市场概览/期权参考现价")
            else:
                self.futures_quotes_xls = None
                self.futures_quotes_df = None
                self.futures_quotes_index = None
                print("⚠ 未找到期货行情数据，将回退使用市场概览/期权参考现价")
            
            return True
            
        except Exception as e:
            print(f"✗ 数据加载失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_underlyings(self):
        """获取所有标的合约列表（按资金规模排序）"""
        if self.option_ref_df is None or self.option_ref_df.empty:
            return []
            
        # 按标的合约和交割年月分组，计算资金合计
        grouped = self.option_ref_df.groupby(['标的合约', '交割年月']).agg({
            '资金合计(万)': 'sum'
        }).reset_index()
        
        # 按资金合计降序排序
        grouped = grouped.sort_values('资金合计(万)', ascending=False)
        
        underlyings = []
        for _, row in grouped.iterrows():
            underlying = row['标的合约']
            expiry = str(row['交割年月'])
            capital = row['资金合计(万)']
            underlyings.append({
                'underlying': underlying,
                'expiry': expiry,
                'capital': capital,
                'display': f"{underlying}_{expiry}"
            })
            
        print(f"\n发现 {len(underlyings)} 个标的期权品种")
        return underlyings
    
    def generate_strategies_for_underlying(self, underlying, expiry):
        """为指定标的生成所有策略"""
        print(f"\n{'='*80}")
        print(f"生成策略: {underlying} {expiry}")
        print(f"{'='*80}")
        
        if self.option_ref_df is None or self.option_ref_df.empty:
            print("⚠ 期权参考数据为空，无法生成策略")
            return None
        
        option_ref_df = self.option_ref_df
        
        # 筛选该标的的期权数据
        mask = (option_ref_df['标的合约'] == underlying) & \
               (option_ref_df['交割年月'] == int(expiry))
        options_df = option_ref_df[mask].copy()
        
        if options_df.empty:
            print(f"⚠ {underlying}_{expiry} 无可用期权数据")
            return None
        
        # 获取市场状态（含标的现价、最大痛点、IV/HV、到期信息等）
        market_state = self._get_market_state(underlying, expiry, options_df)
        
        # 获取波动率曲面特征（Skew/期限结构/峰度偏度等）
        vol_features = self._get_volatility_features(underlying)
        
        # 将部分曲面要点写入市场状态，便于Excel顶部概览
        if vol_features:
            market_state['曲面-倾斜方向'] = vol_features.get('Skew', '')
            market_state['曲面-期限结构'] = vol_features.get('Term Structure', '')
            market_state['曲面-IV/RV比率'] = vol_features.get('IV_RV', '')
            market_state['曲面-峰度'] = vol_features.get('Kurtosis', '')
            market_state['曲面-偏度'] = vol_features.get('Skewness', '')
            if vol_features.get('推荐策略'):
                market_state['曲面-推荐策略'] = vol_features.get('推荐策略')
        
        strategies = []
        
        # 0. 基础单边策略（纯方向）
        print("\n分析单边看涨/看跌机会...")
        directional_strategies = self._generate_directional_strategies(options_df, market_state, vol_features)
        strategies.extend(directional_strategies)
        
        # 1. 日历价差策略
        print("分析日历价差机会...")
        calendar_strategies = self._generate_calendar_spreads(options_df, market_state, vol_features)
        strategies.extend(calendar_strategies)
        
        # 2. 铁鹰式策略
        print("分析铁鹰式机会...")
        iron_condor_strategies = self._generate_iron_condors(options_df, market_state, vol_features)
        strategies.extend(iron_condor_strategies)
        
        # 3. 蝶式价差策略
        print("分析蝶式价差机会...")
        butterfly_strategies = self._generate_butterflies(options_df, market_state, vol_features)
        strategies.extend(butterfly_strategies)

        print("分析风险反转/蝶式矩阵/期限结构机会...")
        risk_reversal_strategies = self._generate_risk_reversals(options_df, market_state, vol_features)
        strategies.extend(risk_reversal_strategies)
        butterfly_matrix_strategies = self._generate_butterfly_matrix(options_df, market_state, vol_features)
        strategies.extend(butterfly_matrix_strategies)
        term_structure_strategies = self._generate_term_structure_timing(options_df, market_state, vol_features)
        strategies.extend(term_structure_strategies)
        synthetic_strategies = self._generate_synthetic_positions(options_df, market_state, vol_features)
        strategies.extend(synthetic_strategies)
        
        # 4. 垂直价差策略
        print("分析垂直价差机会...")
        vertical_strategies = self._generate_vertical_spreads(options_df, market_state, vol_features)
        strategies.extend(vertical_strategies)
        
        # 5. 比率价差策略
        print("分析比率价差机会...")
        ratio_strategies = self._generate_ratio_spreads(options_df, market_state, vol_features)
        strategies.extend(ratio_strategies)
        
        # 6. 跨式/宽跨式策略
        print("分析跨式策略机会...")
        straddle_strategies = self._generate_straddles(options_df, market_state, vol_features)
        strategies.extend(straddle_strategies)

        print("分析末日期权单边机会...")
        expiry_strategies = self._generate_expiry_directional_options(options_df, market_state, vol_features)
        strategies.extend(expiry_strategies)

        print("分析卖方/中性策略机会...")
        credit_spread_strategies = self._generate_credit_spreads(options_df, market_state, vol_features)
        strategies.extend(credit_spread_strategies)
        short_premium_strategies = self._generate_short_premium_neutral_strategies(options_df, market_state, vol_features)
        strategies.extend(short_premium_strategies)
        iron_butterfly_strategies = self._generate_iron_butterflies(options_df, market_state, vol_features)
        strategies.extend(iron_butterfly_strategies)
        jade_lizard_strategies = self._generate_jade_lizards(options_df, market_state, vol_features)
        strategies.extend(jade_lizard_strategies)
        
        # 7. 备兑/保护性策略
        print("分析备兑/保护性策略...")
        covered_strategies = self._generate_covered_strategies(options_df, market_state, vol_features)
        strategies.extend(covered_strategies)
        
        # 8. 波动率曲面/波动率交易类策略（更专业的曲面驱动策略）
        print("分析波动率曲面策略机会...")
        surface_strategies = self._generate_vol_surface_strategies(options_df, market_state, vol_features)
        strategies.extend(surface_strategies)
        
        # 9. 末日Gamma Scalping策略（高风险高收益）
        print("分析末日Gamma Scalping机会...")
        gamma_strategies = self._generate_gamma_scalping_strategies(options_df, market_state, vol_features)
        strategies.extend(gamma_strategies)
        
        # 10. 波动率锥区间交易（IV vs HV历史分位）
        print("分析波动率锥区间机会...")
        vol_cone_strategies = self._generate_vol_cone_strategies(options_df, market_state, vol_features)
        strategies.extend(vol_cone_strategies)
        
        print(f"\n共生成 {len(strategies)} 个策略建议")
        
        return {
            'underlying': underlying,
            'expiry': expiry,
            'market_state': market_state,
            'vol_features': vol_features,
            'strategies': strategies
        }
    
    def _auto_adjust_column_width(self, worksheet, max_width=60, min_width=8):
        """自动调整列宽"""
        from openpyxl.utils import get_column_letter
        
        # 计算列宽：中文字符算2，英文字符算1.1
        def get_visual_width(s):
            if not s: return 0
            s = str(s)
            return sum(2.1 if ord(c) > 127 else 1.1 for c in s)

        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            # 跳过空列
            if not any(cell.value for cell in col):
                continue
                
            max_content_len = 0
            # 采样前200行
            for cell in col[:200]:
                if cell.value:
                    content_len = get_visual_width(cell.value)
                    max_content_len = max(max_content_len, content_len)
            
            final_width = max(min_width, min(max_content_len + 2, max_width))
            worksheet.column_dimensions[col_letter].width = final_width

    def _get_market_state(self, underlying, expiry, options_df):
        """获取标的市场状态（更完整）
        
        - 标的现价：优先期货行情/市场概览，其次期权参考
        - 波动率指标：IV、近期波动率、HV20、HV60（取ATM附近的代表值）
        - 最大痛点：优先市场概览，其次根据持仓分布估算
        - 到期信息：到期日、到期时间、剩余天数
        """
        state = {}
        
        row = None
        if self.market_overview_df is not None and not self.market_overview_df.empty:
            mask = self.market_overview_df['标的合约'] == underlying
            if mask.any():
                row = self.market_overview_df[mask].iloc[0]
        
        # 1) 市场概览字段（货权联动）
        if row is not None:
            state['标的合约'] = underlying
            state['交割年月'] = expiry
            state['期货现价(概览)'] = row.get('期货现价', '')
            state['期货状态'] = row.get('期货状态', '')
            state['期货方向'] = row.get('期货方向', '')
            state['期货流向'] = row.get('期货流向', '')
            state['期货沉淀(亿)'] = row.get('期货沉淀(亿)', 0)
            state['期权结构'] = row.get('期权结构', '')
            state['期权情绪'] = row.get('期权情绪', '')
            state['期权PCR'] = row.get('期权PCR', 1.0)
            state['期权沉淀(亿)'] = row.get('期权沉淀(亿)', 0)
            state['沉淀资金合计(亿)'] = row.get('沉淀资金合计(亿)', '')
            state['联动状态'] = row.get('联动状态', '')
            state['市场解读'] = row.get('市场解读', '')
            state['适合策略(概览)'] = row.get('适合策略', '')
            state['不适合策略(概览)'] = row.get('不适合策略', '')
            state['最大痛点(概览)'] = row.get('最大痛点', '')
            state['痛点距离%(概览)'] = row.get('痛点距离%', '')
        else:
            state['标的合约'] = underlying
            state['交割年月'] = expiry
        
        # 2) 标的现价：优先期货行情 -> 概览现价 -> 期权参考
        underlying_price, underlying_quote = self._get_underlying_price(underlying, expiry, options_df, row)
        state['标的现价'] = underlying_price
        if underlying_quote:
            ask = underlying_quote.get('ask_price1', None)
            bid = underlying_quote.get('bid_price1', None)
            state['标的买价'] = bid if bid else ''  # Bid = 买方出价
            state['标的卖价'] = ask if ask else ''  # Ask = 卖方出价
            try:
                if ask is not None and bid is not None and float(ask) > 0 and float(bid) > 0:
                    state['标的中间价'] = (float(ask) + float(bid)) / 2.0
                else:
                    state['标的中间价'] = ''
            except Exception:
                state['标的中间价'] = ''
            state['标的报价时间'] = underlying_quote.get('datetime', '') or ''
        else:
            state['标的买价'] = ''
            state['标的卖价'] = ''
            state['标的中间价'] = ''
            state['标的报价时间'] = ''
        
        # 3) 到期与剩余天数
        expiry_date = None
        try:
            if '到期日' in options_df.columns and not options_df['到期日'].isna().all():
                expiry_date = pd.to_datetime(options_df['到期日'].dropna().iloc[0])
        except Exception:
            expiry_date = None
        
        state['到期日'] = expiry_date.strftime('%Y-%m-%d') if expiry_date is not None else ''
        if expiry_date is not None:
            try:
                state['到期时间'] = expiry_date.replace(hour=15, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                state['到期时间'] = expiry_date.strftime('%Y-%m-%d')
        else:
            state['到期时间'] = ''
        state['剩余天数(最小)'] = int(options_df['剩余天数'].min()) if '剩余天数' in options_df.columns else ''
        state['剩余天数(平均)'] = float(options_df['剩余天数'].mean()) if '剩余天数' in options_df.columns else ''
        
        # 4) 最大痛点：优先市场概览痛点sheet/概览字段，其次估算
        max_pain = None
        if row is not None and pd.notna(row.get('最大痛点', np.nan)):
            max_pain = row.get('最大痛点')
        elif self.pain_overview_df is not None and not self.pain_overview_df.empty:
            m2 = self.pain_overview_df['标的合约'] == underlying
            if m2.any():
                max_pain = self.pain_overview_df[m2].iloc[0].get('最大痛点', None)
        if max_pain is None or (isinstance(max_pain, float) and np.isnan(max_pain)):
            max_pain = self._estimate_max_pain_from_oi(options_df)
        max_pain_num = None
        try:
            v = float(str(max_pain).replace(',', ''))
            if not np.isnan(v):
                max_pain_num = v
        except Exception:
            max_pain_num = None
        state['最大痛点'] = max_pain_num if max_pain_num is not None else max_pain
        if underlying_price and max_pain_num is not None and max_pain_num > 0:
            state['痛点距离%'] = (underlying_price - max_pain_num) / max_pain_num * 100.0
        else:
            state['痛点距离%'] = ''
        
        # 5) 波动率指标（ATM附近）
        vol_metrics = self._calc_atm_vol_metrics(options_df, underlying_price)
        state.update(vol_metrics)
        
        # 6) 波动率型期权（市场概览给出的“类型细分/策略建议”）
        if self.vol_option_overview_df is not None and not self.vol_option_overview_df.empty:
            m3 = self.vol_option_overview_df['标的合约'] == underlying
            if m3.any():
                r3 = self.vol_option_overview_df[m3].iloc[0]
                state['波动率类型细分(概览)'] = r3.get('类型细分', '')
                state['波动率策略建议(概览)'] = r3.get('策略建议', '')
        
        # 7) 股指相关特殊处理：指数类标的（SSE.*）期货沉淀常为0/异常值，避免误导
        if isinstance(underlying, str) and underlying.startswith('SSE.'):
            for k in ['期货沉淀(亿)', '期权沉淀(亿)']:
                try:
                    v = float(state.get(k, 0) or 0)
                    if abs(v) > 1e4:
                        state[k] = 0
                except Exception:
                    state[k] = 0
        
        # 7) 行权方式 (E=欧式, A=美式)
        if options_df is not None and not options_df.empty:
            exercise_type = options_df['exercise_type'].iloc[0] if 'exercise_type' in options_df.columns else 'E'
            state['行权方式'] = '美式' if str(exercise_type).upper() == 'A' else '欧式'
        else:
            state['行权方式'] = '未知'
        
        return state
    
    def _map_underlying_to_futures_candidates(self, underlying: str, expiry: str):
        try:
            if not isinstance(underlying, str) or not underlying:
                return []
            candidates = []
            if '.' in underlying:
                exch, rest = underlying.split('.', 1)
                exch = exch.upper()
                rest = rest.upper()
                if exch == 'SSE':
                    index_map = {'000852': 'IM', '000300': 'IF', '000016': 'IH', '000905': 'IC'}
                    prod = index_map.get(rest, None)
                    if prod and expiry:
                        yymm = str(expiry)[2:]
                        candidates.append(f"CFFEX.{prod}{yymm}")
            candidates.append(underlying)
            return candidates
        except Exception:
            return [underlying]
    
    def _get_underlying_price(self, underlying, expiry, options_df, overview_row=None):
        """获取标的现价（用于ATM/痛点距离/策略区间等计算）。返回 (price, quote_dict_or_None)"""
        # 1) 期货行情优先（包含股指映射）
        qf = None
        for inst in self._map_underlying_to_futures_candidates(underlying, expiry):
            qf = self._get_futures_quote(inst)
            if qf:
                lp = qf.get('last_price', 0) or 0
                try:
                    if float(lp) > 0:
                        return float(lp), qf
                except Exception:
                    pass
        
        # 2) 市场概览（货权联动中的“期货现价”）
        if overview_row is not None:
            px = overview_row.get('期货现价', None)
            if px is not None and not (isinstance(px, float) and np.isnan(px)):
                try:
                    if float(px) > 0:
                        return float(px), qf
                except Exception:
                    pass
        
        # 3) 期权参考（标的现价列）
        if options_df is not None and not options_df.empty and '标的现价' in options_df.columns:
            s = options_df['标的现价'].dropna()
            if not s.empty:
                try:
                    v = float(s.median())
                    if v > 0:
                        return v, qf
                except Exception:
                    pass
        
        return 0.0, qf

    def _calc_atm_vol_metrics(self, options_df, underlying_price):
        """计算ATM附近代表性的波动率指标（用于顶部概览）。"""
        metrics: dict[str, float | None] = {
            '隐含波动率(ATM)': None,
            '近期波动率(ATM)': None,
            'HV20(ATM)': None,
            'HV60(ATM)': None,
            'IV/RV(ATM)': None,
            '隐含波动率': None,
            '近期波动率': None,
            'HV20': None,
            'HV60': None,
            'IV/RV': None
        }
        try:
            if options_df is None or options_df.empty:
                return metrics
            if underlying_price is None or underlying_price <= 0:
                return metrics
            
            df = options_df.copy()
            df['strike_diff'] = (df['行权价'] - underlying_price).abs()
            # 取最接近ATM的一小撮（包含CALL/PUT），再按流动性筛一遍
            df = df.sort_values('strike_diff').head(20)
            if '成交量' in df.columns:
                df = df.sort_values(['成交量', '持仓量'], ascending=False)
            rep = df.iloc[0]
            
            iv = rep.get('隐含波动率', None)
            rv = rep.get('近期波动率', None)
            hv20 = rep.get('HV20', None)
            hv60 = rep.get('HV60', None)
            
            metrics['隐含波动率(ATM)'] = float(iv) if iv is not None and not pd.isna(iv) else None
            metrics['近期波动率(ATM)'] = float(rv) if rv is not None and not pd.isna(rv) else None
            metrics['HV20(ATM)'] = float(hv20) if hv20 is not None and not pd.isna(hv20) else None
            metrics['HV60(ATM)'] = float(hv60) if hv60 is not None and not pd.isna(hv60) else None
            
            iv_atm = metrics.get('隐含波动率(ATM)', None)
            rv_atm = metrics.get('近期波动率(ATM)', None)
            if isinstance(iv_atm, (int, float)) and isinstance(rv_atm, (int, float)) and rv_atm != 0:
                metrics['IV/RV(ATM)'] = float(iv_atm) / float(rv_atm)
            metrics['隐含波动率'] = metrics['隐含波动率(ATM)']
            metrics['近期波动率'] = metrics['近期波动率(ATM)']
            metrics['HV20'] = metrics['HV20(ATM)']
            metrics['HV60'] = metrics['HV60(ATM)']
            metrics['IV/RV'] = metrics['IV/RV(ATM)']
            return metrics
        except Exception:
            return metrics

    def _estimate_max_pain_from_oi(self, options_df):
        """用简单Max Pain思想（基于持仓量近似）估算最大痛点。"""
        try:
            if options_df is None or options_df.empty:
                return None
            if '持仓量' not in options_df.columns:
                return None
            strikes = sorted(options_df['行权价'].dropna().unique())
            if not strikes:
                return None
            
            df = options_df[['行权价', '期权类型', '持仓量', '合约乘数']].copy()
            df['持仓量'] = pd.Series(pd.to_numeric(df['持仓量'], errors='coerce')).fillna(0)
            df['合约乘数'] = pd.Series(pd.to_numeric(df['合约乘数'], errors='coerce')).fillna(1)
            
            pains = {}
            for S in strikes:
                # 以到期时标的价格=某行权价S，计算“期权买方到期价值”之和（越小越痛）
                call_payoff = ((S - df['行权价']).clip(lower=0))
                put_payoff = ((df['行权价'] - S).clip(lower=0))
                payoff = np.where(df['期权类型'] == 'CALL', call_payoff, put_payoff)
                total = float((payoff * df['持仓量'] * df['合约乘数']).sum())
                pains[S] = total
            
            if not pains:
                return None
            return min(pains, key=lambda k: pains[k])
        except Exception:
            return None

    def _infer_option_quote_sheet(self, contract_code: str):
        """从合约代码推断 `wisecoin-期权行情.xlsx` 的sheet名。"""
        try:
            if not contract_code or '.' not in contract_code:
                return None
            exch, rest = contract_code.split('.', 1)
            exch = exch.upper()
            
            # SSE指数类：SSE.000300Cxxxx -> SSE_000300
            m = re.match(r'^(\d{6})', rest)
            if exch == 'SSE' and m:
                return f"{exch}_{m.group(1)}"
            
            if exch == 'CFFEX':
                m3 = re.match(r'^([A-Za-z]+)', rest)
                if m3:
                    prod = m3.group(1).upper()
                    index_opt_map = {'IO': 'SSE_000300', 'HO': 'SSE_000016', 'MO': 'SSE_000852'}
                    if prod in index_opt_map:
                        return index_opt_map[prod]
            
            # 商品期权：SHFE.ag2604Cxxxx -> SHFE_AG
            m2 = re.match(r'^([A-Za-z]+)', rest)
            if not m2:
                return None
            product = m2.group(1).upper()
            return f"{exch}_{product}"
        except Exception:
            return None

    def _get_option_quote(self, contract_code: str):
        """获取期权盘口报价（bid/ask/last/time）。"""
        try:
            if self.option_quotes_xls is None:
                return None
            if not contract_code:
                return None
            cached = self._option_quote_row_cache.get(contract_code, None)
            if cached is not None:
                return cached
            sheet = self._infer_option_quote_sheet(contract_code)
            if sheet:
                q = self._get_option_quote_from_sheet(sheet, contract_code)
                if q:
                    self._option_quote_row_cache[contract_code] = q
                    return q
            for sn in getattr(self.option_quotes_xls, 'sheet_names', []) or []:
                q = self._get_option_quote_from_sheet(sn, contract_code)
                if q:
                    self._option_quote_row_cache[contract_code] = q
                    return q
            self._option_quote_row_cache[contract_code] = None
            return None
        except Exception:
            return None

    def _match_quote_sheet_name(self, sheet: str):
        if not sheet:
            return None
        if sheet in self._option_quote_sheet_lookup:
            return self._option_quote_sheet_lookup[sheet]
        names = getattr(self.option_quotes_xls, 'sheet_names', []) or []
        if sheet in names:
            self._option_quote_sheet_lookup[sheet] = sheet
            return sheet
        s_low = str(sheet).lower()
        for n in names:
            if str(n).lower() == s_low:
                self._option_quote_sheet_lookup[sheet] = n
                return n
        for n in names:
            if s_low in str(n).lower():
                self._option_quote_sheet_lookup[sheet] = n
                return n
        self._option_quote_sheet_lookup[sheet] = None
        return None

    def _normalize_quote_df(self, df: pd.DataFrame):
        if df is None or df.empty:
            return None
        key_col = None
        for c in ['instrument_id', '合约代码', 'symbol', 'instrument']:
            if c in df.columns:
                key_col = c
                break
        if not key_col:
            return None
        out = df.copy()
        out[key_col] = out[key_col].astype(str)
        cols = out.columns
        if 'bid_price1' not in cols:
            for c in ['bid', 'bid1', 'BidPrice1', 'bidPrice1']:
                if c in cols:
                    out['bid_price1'] = out[c]
                    break
        if 'ask_price1' not in cols:
            for c in ['ask', 'ask1', 'AskPrice1', 'askPrice1']:
                if c in cols:
                    out['ask_price1'] = out[c]
                    break
        if 'last_price' not in cols:
            for c in ['last', 'last_price1', 'LastPrice', 'lastPrice']:
                if c in cols:
                    out['last_price'] = out[c]
                    break
        if 'datetime' not in cols:
            for c in ['time', 'update_time', 'quote_time', '更新时间']:
                if c in cols:
                    out['datetime'] = out[c]
                    break
        out = out.set_index(key_col, drop=False)
        return out

    def _get_option_quote_from_sheet(self, sheet: str, contract_code: str):
        try:
            if self.option_quotes_xls is None:
                return None
            actual = self._match_quote_sheet_name(sheet)
            if not actual:
                return None
            df = self._option_quote_sheet_cache.get(actual)
            if df is None:
                parsed = self.option_quotes_xls.parse(actual)
                if isinstance(parsed, dict):
                    return None
                df = self._normalize_quote_df(parsed)
                if df is None:
                    return None
                self._option_quote_sheet_cache[actual] = df
            if contract_code not in df.index:
                return None
            row = df.loc[contract_code]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return {
                'datetime': row.get('datetime', ''),
                'bid_price1': float(row.get('bid_price1', 0) or 0),
                'ask_price1': float(row.get('ask_price1', 0) or 0),
                'last_price': float(row.get('last_price', 0) or 0),
                'open_interest': float(row.get('open_interest', 0) or 0)
            }
        except Exception:
            return None

    def _get_futures_quote(self, contract_code: str):
        """获取期货盘口报价（bid/ask/last/time）。"""
        try:
            if self.futures_quotes_index is None:
                return None
            if contract_code not in self.futures_quotes_index.index:
                return None
            row = self.futures_quotes_index.loc[contract_code]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return {
                'datetime': row.get('datetime', ''),
                'bid_price1': float(row.get('bid_price1', 0) or 0),
                'ask_price1': float(row.get('ask_price1', 0) or 0),
                'last_price': float(row.get('last_price', 0) or 0),
                'settlement': float(row.get('settlement', 0) or 0)
            }
        except Exception:
            return None

    def _get_option_trade_price(self, opt_row, action: str):
        """按买卖方向选择成交参考价：买入用 ask1，卖出用 bid1，缺失则回退 last/参考价。"""
        fallback = float(opt_row.get('期权价', 0) or 0)
        q = self._get_option_quote(opt_row.get('合约代码', ''))
        if not q:
            return fallback, None, None, None, None
        bid = q.get('bid_price1', 0) or 0
        ask = q.get('ask_price1', 0) or 0
        last = q.get('last_price', 0) or 0
        
        if action == '买入':
            px = ask if ask > 0 else (last if last > 0 else fallback)
        else:
            px = bid if bid > 0 else (last if last > 0 else fallback)
        mid = (bid + ask) / 2 if bid > 0 and ask > 0 else None
        return float(px), bid if bid > 0 else None, ask if ask > 0 else None, mid, q.get('datetime', '')
    
    def _make_option_leg(self, opt_row, action: str, qty: int = 1):
        """生成统一格式的期权腿（包含盘口价格与权利金现金流）。"""
        multiplier = float(opt_row.get('合约乘数', 1) or 1)
        trade_px, bid, ask, mid, qt = self._get_option_trade_price(opt_row, action)
        premium = trade_px * multiplier * qty
        cashflow = premium if action == '卖出' else -premium
        
        return {
            '操作': action,
            '合约代码': opt_row.get('合约代码', ''),
            '期权类型': opt_row.get('期权类型', ''),
            '行权价': opt_row.get('行权价', ''),
            '剩余天数': opt_row.get('剩余天数', ''),
            '合约乘数': multiplier,
            '参考价': float(opt_row.get('期权价', 0) or 0),
            '盘口买价': ask,
            '盘口卖价': bid,
            '中间价': mid,
            '建议成交价': trade_px,
            '数量': qty,
            '权利金': cashflow,
            '报价时间': qt
        }

    def _make_underlying_leg(self, underlying: str, expiry: str, action: str, qty: int, fallback_price: float, multiplier: float = 1.0):
        """生成标的腿（用于备兑/保护性/领口）。"""
        qf = None
        for inst in self._map_underlying_to_futures_candidates(underlying, expiry):
            qf = self._get_futures_quote(inst)
            if qf:
                lp = qf.get('last_price', 0) or 0
                try:
                    if float(lp) > 0:
                        break
                except Exception:
                    pass
        qt = qf.get('datetime', '') if qf else ''
        bid = qf.get('bid_price1') if qf else None
        ask = qf.get('ask_price1') if qf else None
        last = qf.get('last_price') if qf else None
        if action in ['买入', '持有/买入']:
            trade_px = ask if ask and ask > 0 else (last if last and last > 0 else fallback_price)
            cashflow = -float(trade_px) * float(multiplier) * qty
        else:
            trade_px = bid if bid and bid > 0 else (last if last and last > 0 else fallback_price)
            cashflow = float(trade_px) * float(multiplier) * qty
        return {
            '操作': action,
            '合约代码': underlying,
            '期权类型': '标的',
            '行权价': '',
            '剩余天数': '',
            '合约乘数': float(multiplier),
            '参考价': fallback_price,
            '盘口买价': ask if ask and ask > 0 else None,
            '盘口卖价': bid if bid and bid > 0 else None,
            '中间价': None,
            '建议成交价': float(trade_px) if trade_px else float(fallback_price),
            '数量': qty,
            '权利金': cashflow,
            '报价时间': qt
        }
    
    def _get_volatility_features(self, underlying):
        """获取波动率特征"""
        if self.vol_surface_df is None or self.vol_surface_df.empty:
            return {}
            
        # 从标的合约提取品种代码（如 SSE.000852 -> 000852，SHFE.rb -> rb）
        if '.' in underlying:
            symbol_code = underlying.split('.')[1].upper()
        else:
            symbol_code = underlying.upper()
        
        # 从波动率曲面中提取该品种信息（使用品种代码匹配）
        mask = self.vol_surface_df['品种代码'].str.upper() == symbol_code
        if not mask.any():
            # 尝试模糊匹配
            mask = self.vol_surface_df['品种代码'].str.upper().str.contains(symbol_code[:2], na=False)
            if not mask.any():
                return {}
            
        row = self.vol_surface_df[mask].iloc[0]
        
        return {
            'Skew': row.get('倾斜方向', 'Flat'),
            'Skew_Slope': row.get('IV倾斜度', ''),
            'Term Structure': row.get('期限结构', 'Normal'),
            'Term_Spread': row.get('期限结构差', ''),
            'Short_IV': row.get('短期IV', ''),
            'Long_IV': row.get('长期IV', ''),
            'IV_RV': row.get('IV/RV比率', ''),
            'Kurtosis': row.get('峰度', 'Medium'),
            'Skewness': row.get('偏度', '~0'),
            '市场情绪': row.get('市场情绪', '窄幅震荡'),
            '推荐策略': row.get('推荐策略', ''),
            '推荐合约': row.get('推荐合约', '')
        }
    
    def _generate_directional_strategies(self, options_df, market_state, vol_features):
        """生成单边看涨/看跌策略（基础方向性策略）"""
        strategies = []
        
        direction = market_state.get('期货方向', '').strip()
        iv_rv = self._safe_float(market_state.get('IV/RV(ATM)', None))
        
        # 获取流动性好的期权
        universe = self._get_underlying_option_universe(options_df, market_state)
        liquid_options = universe[
            (universe['成交量'] >= 50) &
            (universe['持仓量'] >= 100) &
            (universe['剩余天数'] >= 7) &
            (universe['剩余天数'] <= 90)
        ].copy()
        
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(
            liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns 
            and not liquid_options['标的现价'].dropna().empty else 0
        )
        
        # 1. 买入看涨期权（强烈看涨 + IV低位）
        if ('涨' in direction or '偏多' in direction) and (iv_rv is None or iv_rv < 1.2):
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            if not calls.empty:
                # 选择ATM或轻度虚值Call（行权价在现价上方0-5%）
                calls['价差'] = abs(calls['行权价'] - atm_price)
                calls = calls[
                    (calls['行权价'] >= atm_price * 0.98) &
                    (calls['行权价'] <= atm_price * 1.05)
                ].sort_values('价差')
                
                if not calls.empty:
                    best_call = calls.iloc[0]
                    strategy = self._build_long_call_strategy(best_call, market_state, vol_features)
                    if strategy:
                        strategies.append(strategy)
        
        # 2. 买入看跌期权（强烈看跌 + IV低位）
        if ('跌' in direction or '偏空' in direction) and (iv_rv is None or iv_rv < 1.2):
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            if not puts.empty:
                # 选择ATM或轻度虚值Put（行权价在现价下方0-5%）
                puts['价差'] = abs(puts['行权价'] - atm_price)
                puts = puts[
                    (puts['行权价'] >= atm_price * 0.95) &
                    (puts['行权价'] <= atm_price * 1.02)
                ].sort_values('价差')
                
                if not puts.empty:
                    best_put = puts.iloc[0]
                    strategy = self._build_long_put_strategy(best_put, market_state, vol_features)
                    if strategy:
                        strategies.append(strategy)
        
        return strategies
    
    def _build_long_call_strategy(self, call_opt, market_state, vol_features):
        """构建买入看涨期权策略"""
        try:
            # 构建期权腿
            leg = self._make_option_leg(call_opt, '买入', 1)
            
            premium = -float(leg.get('权利金', 0) or 0)  # 买入为负现金流
            strike = float(call_opt.get('行权价', 0))
            multiplier = float(call_opt.get('合约乘数', 1) or 1)
            
            # 盈亏平衡点
            breakeven = strike + (-premium / multiplier)
            
            # 获取CSV指南信息
            guide_info = self.strategy_guide_dict.get('买入看涨期权', {})
            
            strategy = {
                '策略类型': '买入看涨期权（Long Call）',
                '策略描述': '纯做多期货的权利仓，风险有限（仅权利金），收益无限',
                '适用场景': guide_info.get('期货实战开仓时机', '强烈看涨+IV低位+技术面突破'),
                '操作建议': [leg],
                '成本分析': {
                    '权利金成本': -premium,
                    '最大亏损': -premium,
                    '最大盈利': '无限',
                    '盈亏比': '无限'
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '需上涨幅度': f'{((breakeven / strike - 1) * 100):.2f}%'
                },
                '风险提示': [
                    '亏损固定为权利金，但可能损失100%权利金',
                    '需要标的大幅上涨才能盈利（克服时间价值损耗）',
                    'Theta时间衰减每天吃掉权利金',
                    '优先选择平值或轻度虚值，Delta约0.5',
                    guide_info.get('补充策略要点（实操必看）', '到期日选1-3个月平衡时间与空间'),
                    f'当前市场状态：{market_state.get("期货方向", "")}，IV/RV={market_state.get("IV/RV(ATM)", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{call_opt.get("Delta", 0):.4f}（正向暴露）',
                    'Gamma': f'{call_opt.get("Gamma", 0):.6f}（凸性加速）',
                    'Theta': f'{call_opt.get("Theta", 0):.4f}（时间损耗）',
                    'Vega': f'{call_opt.get("Vega", 0):.4f}（波动率敏感）'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建买入看涨策略失败: {e}")
            return None
    
    def _build_long_put_strategy(self, put_opt, market_state, vol_features):
        """构建买入看跌期权策略"""
        try:
            # 构建期权腿
            leg = self._make_option_leg(put_opt, '买入', 1)
            
            premium = -float(leg.get('权利金', 0) or 0)  # 买入为负现金流
            strike = float(put_opt.get('行权价', 0))
            multiplier = float(put_opt.get('合约乘数', 1) or 1)
            
            # 盈亏平衡点
            breakeven = strike - (-premium / multiplier)
            
            # 获取CSV指南信息
            guide_info = self.strategy_guide_dict.get('买入看跌期权', {})
            
            strategy = {
                '策略类型': '买入看跌期权（Long Put）',
                '策略描述': '纯做空期货的权利仓，风险有限（仅权利金），收益无限',
                '适用场景': guide_info.get('期货实战开仓时机', '强烈看跌+IV低位+跌破关键支撑'),
                '操作建议': [leg],
                '成本分析': {
                    '权利金成本': -premium,
                    '最大亏损': -premium,
                    '最大盈利': f'{strike * multiplier - (-premium):.2f}（标的跌至0）',
                    '盈亏比': f'{(strike * multiplier / (-premium)):.2f}' if premium < 0 else '无限'
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '需下跌幅度': f'{((1 - breakeven / strike) * 100):.2f}%'
                },
                '风险提示': [
                    '亏损固定为权利金，但可能损失100%权利金',
                    '需要标的大幅下跌才能盈利（克服时间价值损耗）',
                    'Theta时间衰减每天吃掉权利金',
                    '规避深度虚值Put，防止行情下跌但期权不涨',
                    guide_info.get('补充策略要点（实操必看）', '商品期货熊市中可搭配主力合约换月'),
                    f'当前市场状态：{market_state.get("期货方向", "")}，IV/RV={market_state.get("IV/RV(ATM)", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{put_opt.get("Delta", 0):.4f}（负向暴露）',
                    'Gamma': f'{put_opt.get("Gamma", 0):.6f}（凸性加速）',
                    'Theta': f'{put_opt.get("Theta", 0):.4f}（时间损耗）',
                    'Vega': f'{put_opt.get("Vega", 0):.4f}（波动率敏感）'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建买入看跌策略失败: {e}")
            return None
    
    def _generate_calendar_spreads(self, options_df, market_state, vol_features):
        """生成日历价差策略"""
        strategies = []
        
        # 日历价差适合场景：波动率期限结构倒挂或震荡市
        sentiment = vol_features.get('市场情绪', '')
        term_structure = vol_features.get('Term Structure', '')
        
        universe = self._get_underlying_option_universe(options_df, market_state)

        liquid_options = universe[
            (universe['成交量'] >= 50) &
            (universe['持仓量'] >= 100) &
            (universe['剩余天数'] >= 7) &
            (universe['剩余天数'] <= 240)
        ].copy()
        
        if liquid_options.empty:
            liquid_options = universe[
                (universe['成交量'] >= 10) &
                (universe['持仓量'] >= 20) &
                (universe['剩余天数'] >= 3) &
                (universe['剩余天数'] <= 365)
            ].copy()
        if liquid_options.empty:
            return strategies
        
        # 选择平值或轻度虚值期权（使用统一的标的现价）
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        def pick_pair(df):
            if df is None or df.empty:
                return None
            df = df.dropna(subset=['行权价', '剩余天数']).copy()
            if df.empty:
                return None
            df['__strike_diff'] = np.abs(df['行权价'] - float(atm_price))
            df = df.sort_values(by=['__strike_diff'])
            strike_candidates = df['行权价'].dropna().unique().tolist()[:16]
            for strike in strike_candidates:
                g = df[df['行权价'] == strike].copy()
                if len(g) < 2:
                    continue
                g = g.sort_values(by=['剩余天数'])
                near = g[(g['剩余天数'] >= 5) & (g['剩余天数'] <= 45)]
                far = g[(g['剩余天数'] >= 20) & (g['剩余天数'] <= 365)]
                if near.empty or far.empty:
                    continue
                short_term = near.iloc[0]
                long_term = far.iloc[-1]
                if float(long_term.get('剩余天数', 0) or 0) - float(short_term.get('剩余天数', 0) or 0) >= 14:
                    return short_term, long_term
            near = df[(df['剩余天数'] >= 5) & (df['剩余天数'] <= 45)].sort_values(by=['__strike_diff'])
            far = df[(df['剩余天数'] >= 20) & (df['剩余天数'] <= 365)].copy()
            if near.empty or far.empty:
                return None
            short_term = near.iloc[0]
            far['__kdiff'] = np.abs(far['行权价'] - float(short_term.get('行权价', 0) or 0))
            far = far.sort_values(by=['__kdiff', '剩余天数'], ascending=[True, False])
            long_term = far.iloc[0]
            if float(long_term.get('剩余天数', 0) or 0) - float(short_term.get('剩余天数', 0) or 0) < 14:
                return None
            return short_term, long_term

        def pick_pair_by_strike(df, strike):
            if df is None or df.empty:
                return None
            g = df[df['行权价'] == strike].copy()
            if len(g) < 2:
                return None
            g = g.sort_values(by=['剩余天数'])
            near = g[(g['剩余天数'] >= 5) & (g['剩余天数'] <= 45)]
            far = g[(g['剩余天数'] >= 20) & (g['剩余天数'] <= 365)]
            if near.empty or far.empty:
                return None
            short_term = near.iloc[0]
            long_term = far.iloc[-1]
            if float(long_term.get('剩余天数', 0) or 0) - float(short_term.get('剩余天数', 0) or 0) < 14:
                return None
            return short_term, long_term

        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        picked = pick_pair(calls)
        if picked:
            short_term, long_term = picked
            strategy = self._build_calendar_spread_strategy(short_term, long_term, 'CALL', market_state, vol_features)
            if strategy:
                strategies.append(strategy)

        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        picked = pick_pair(puts)
        if picked:
            short_term, long_term = picked
            strategy = self._build_calendar_spread_strategy(short_term, long_term, 'PUT', market_state, vol_features)
            if strategy:
                strategies.append(strategy)

        if not calls.empty and not puts.empty:
            call_strikes = set([float(x) for x in calls['行权价'].dropna().unique().tolist()])
            put_strikes = set([float(x) for x in puts['行权价'].dropna().unique().tolist()])
            common = list(call_strikes.intersection(put_strikes))
            common = sorted(common, key=lambda k: abs(k - float(atm_price)))
            for strike in common[:8]:
                call_pair = pick_pair_by_strike(calls, strike)
                put_pair = pick_pair_by_strike(puts, strike)
                if call_pair and put_pair:
                    strategy = self._build_calendar_straddle_strategy(
                        call_pair[0], call_pair[1], put_pair[0], put_pair[1],
                        market_state, vol_features
                    )
                    if strategy:
                        strategies.append(strategy)
                    break
        
        return strategies
    
    def _build_calendar_spread_strategy(self, short_opt, long_opt, opt_type, market_state, vol_features):
        """构建日历价差策略详情"""
        try:
            leg_short = self._make_option_leg(short_opt, '卖出', 1)
            leg_long = self._make_option_leg(long_opt, '买入', 1)
            legs = [leg_short, leg_long]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))  # >0收入 <0支出
            net_debit = -net_cashflow
            
            # 最大收益：短期到期价值归零时，长期仍有价值
            # 简化估算：短期Theta衰减 * 剩余天数
            max_profit = abs(short_opt.get('Theta', 0)) * short_opt['剩余天数'] * short_opt['合约乘数']
            max_loss = net_debit
            
            # 盈亏平衡点
            strike_short = float(short_opt.get('行权价', 0) or 0)
            strike_long = float(long_opt.get('行权价', 0) or 0)
            breakeven_lower = strike_short * 0.95
            breakeven_upper = strike_short * 1.05

            is_diagonal = abs(strike_short - strike_long) > max(1e-9, abs(strike_short) * 1e-6)
            strategy_type = f'{opt_type}对角日历' if is_diagonal else f'{opt_type}日历价差'
            strategy_desc = (
                f'卖出短期{strike_short:.2f}+买入长期{strike_long:.2f} {opt_type}期权，兼具时间差与价差'
                if is_diagonal
                else f'卖出短期+买入长期同行权价{opt_type}期权，赚取时间价值衰减差'
            )
            
            strategy = {
                '策略类型': strategy_type,
                '策略描述': strategy_desc,
                '适用场景': f'{vol_features.get("市场情绪", "")}，波动率{vol_features.get("Term Structure", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '最佳区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '适合震荡市或期限结构倒挂时使用',
                    '标的价格大幅波动会影响收益',
                    '需在短期期权到期前评估是否调整',
                    f'当前市场状态：{market_state.get("联动状态", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(short_opt.get("Delta", 0) - long_opt.get("Delta", 0)):.4f}',
                    'Gamma': f'{(short_opt.get("Gamma", 0) - long_opt.get("Gamma", 0)):.6f}',
                    'Theta': f'{(short_opt.get("Theta", 0) - long_opt.get("Theta", 0)):.4f}',
                    'Vega': f'{(short_opt.get("Vega", 0) - long_opt.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建日历价差策略失败: {e}")
            return None

    def _build_calendar_straddle_strategy(self, short_call, long_call, short_put, long_put, market_state, vol_features):
        try:
            legs = [
                self._make_option_leg(short_call, '卖出', 1),
                self._make_option_leg(long_call, '买入', 1),
                self._make_option_leg(short_put, '卖出', 1),
                self._make_option_leg(long_put, '买入', 1)
            ]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            theta_call = abs(float(short_call.get('Theta', 0) or 0)) * float(short_call.get('剩余天数', 0) or 0) * float(short_call.get('合约乘数', 1) or 1)
            theta_put = abs(float(short_put.get('Theta', 0) or 0)) * float(short_put.get('剩余天数', 0) or 0) * float(short_put.get('合约乘数', 1) or 1)
            max_profit = theta_call + theta_put
            max_loss = max(net_debit, 0)
            strike = float(short_call.get('行权价', 0) or short_put.get('行权价', 0) or 0)
            breakeven_lower = strike * 0.95
            breakeven_upper = strike * 1.05
            strategy = {
                '策略类型': '日历跨式',
                '策略描述': '卖出近月CALL+PUT、买入远月CALL+PUT，同一行权价的期限结构策略',
                '适用场景': f'{vol_features.get("市场情绪", "")}，期限结构{vol_features.get("Term Structure", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '最佳区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '适合震荡且期限结构倒挂或偏陡时使用',
                    '标的价格大幅波动会影响收益',
                    '需在短期期权到期前评估是否调整',
                    f'当前市场状态：{market_state.get("联动状态", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(short_call.get("Delta", 0) or 0) + float(short_put.get("Delta", 0) or 0) - float(long_call.get("Delta", 0) or 0) - float(long_put.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(short_call.get("Gamma", 0) or 0) + float(short_put.get("Gamma", 0) or 0) - float(long_call.get("Gamma", 0) or 0) - float(long_put.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(short_call.get("Theta", 0) or 0) + float(short_put.get("Theta", 0) or 0) - float(long_call.get("Theta", 0) or 0) - float(long_put.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(short_call.get("Vega", 0) or 0) + float(short_put.get("Vega", 0) or 0) - float(long_call.get("Vega", 0) or 0) - float(long_put.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception:
            return None
    
    def _generate_iron_condors(self, options_df, market_state, vol_features):
        """生成铁鹰式策略
        
        铁鹰式 = 卖出价外看涨 + 买入更价外看涨 + 卖出价外看跌 + 买入更价外看跌
        适合震荡市，赚取时间价值衰减
        """
        strategies = []
        
        # 铁鹰式适合场景：窄幅震荡、震荡筑底、震荡冲高
        sentiment = market_state.get('期权情绪', '')
        if '狂热' in sentiment or '恐慌' in sentiment:
            return strategies  # 不适合单边行情
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 100) &
            (options_df['持仓量'] >= 200) &
            (options_df['剩余天数'] >= 30) &
            (options_df['剩余天数'] <= 60)
        ].copy()
        
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        # 分离看涨和看跌期权
        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        
        if len(calls) < 2 or len(puts) < 2:
            return strategies
        
        # 按行权价排序
        calls = calls.sort_values('行权价')
        puts = puts.sort_values('行权价', ascending=False)
        
        # 选择价外期权（虚值档位0.9-1.1）
        otm_calls = calls[calls['行权价'] > atm_price * 1.02].copy()
        otm_puts = puts[puts['行权价'] < atm_price * 0.98].copy()
        
        if len(otm_calls) < 2 or len(otm_puts) < 2:
            return strategies
        
        # 构建铁鹰式组合
        # 看涨侧：卖出较近价外 + 买入更远价外
        sell_call = otm_calls.iloc[0]  # 卖出较低行权价
        buy_call = otm_calls.iloc[min(1, len(otm_calls)-1)]  # 买入较高行权价
        
        # 看跌侧：卖出较近价外 + 买入更远价外
        sell_put = otm_puts.iloc[0]  # 卖出较高行权价
        buy_put = otm_puts.iloc[min(1, len(otm_puts)-1)]  # 买入较低行权价
        
        # 确保有足够的行权价间距
        call_spread = buy_call['行权价'] - sell_call['行权价']
        put_spread = sell_put['行权价'] - buy_put['行权价']
        
        if call_spread < atm_price * 0.03 or put_spread < atm_price * 0.03:
            return strategies
        
        strategy = self._build_iron_condor_strategy(
            sell_call, buy_call, sell_put, buy_put, 
            market_state, vol_features
        )
        
        if strategy:
            strategies.append(strategy)
        
        return strategies
    
    def _build_iron_condor_strategy(self, sell_call, buy_call, sell_put, buy_put, 
                                     market_state, vol_features):
        """构建铁鹰式策略详情"""
        try:
            multiplier = float(sell_call.get('合约乘数', 1) or 1)
            legs = [
                self._make_option_leg(sell_call, '卖出', 1),
                self._make_option_leg(buy_call, '买入', 1),
                self._make_option_leg(sell_put, '卖出', 1),
                self._make_option_leg(buy_put, '买入', 1)
            ]
            net_credit = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            
            # 最大收益 = 净收入
            max_profit = net_credit
            
            # 最大损失 = 单侧价差 - 净收入
            call_spread = (buy_call['行权价'] - sell_call['行权价']) * multiplier
            put_spread = (sell_put['行权价'] - buy_put['行权价']) * multiplier
            max_loss = max(call_spread, put_spread) - net_credit
            
            # 盈亏平衡点
            breakeven_upper = sell_call['行权价'] + (net_credit / multiplier)
            breakeven_lower = sell_put['行权价'] - (net_credit / multiplier)
            
            # 盈利区间
            profit_zone_lower = sell_put['行权价']
            profit_zone_upper = sell_call['行权价']
            
            strategy = {
                '策略类型': '铁鹰式（Iron Condor）',
                '策略描述': '卖出价外看涨+看跌期权，买入更价外期权保护，赚取时间价值衰减',
                '适用场景': f'{vol_features.get("市场情绪", "窄幅震荡")}，预期标的在区间内震荡',
                '操作建议': legs,
                '成本分析': {
                    '净权利金收入': net_credit,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0,
                    '盈利区间': f'{profit_zone_lower:.2f} - {profit_zone_upper:.2f}'
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '盈利区间宽度': profit_zone_upper - profit_zone_lower
                },
                '风险提示': [
                    '适合震荡市，预期标的价格在盈利区间内波动',
                    '标的突破盈利区间将产生亏损（受保护腿限制）',
                    '最佳收益在到期日标的在盈利区间内',
                    '时间价值衰减有利于该策略',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '建议初学者使用，风险可控'
                ],
                '希腊字母风险': {
                    'Delta': f'{(sell_call.get("Delta", 0) - buy_call.get("Delta", 0) + sell_put.get("Delta", 0) - buy_put.get("Delta", 0)):.4f}',
                    'Gamma': f'{(sell_call.get("Gamma", 0) - buy_call.get("Gamma", 0) + sell_put.get("Gamma", 0) - buy_put.get("Gamma", 0)):.6f}',
                    'Theta': f'{(sell_call.get("Theta", 0) - buy_call.get("Theta", 0) + sell_put.get("Theta", 0) - buy_put.get("Theta", 0)):.4f}',
                    'Vega': f'{(sell_call.get("Vega", 0) - buy_call.get("Vega", 0) + sell_put.get("Vega", 0) - buy_put.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建铁鹰式策略失败: {e}")
            return None
    
    def _generate_butterflies(self, options_df, market_state, vol_features):
        """生成蝶式价差策略
        
        蝶式价差 = 买入1个低行权价 + 卖出2个中间行权价 + 买入1个高行权价
        适合预期标的价格在到期时接近中间行权价
        """
        strategies = []
        
        # 蝶式适合场景：窄幅震荡，预期价格不会大幅波动
        sentiment = vol_features.get('市场情绪', '')
        if '狂热' in sentiment or '恐慌' in sentiment:
            return strategies
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 50) &
            (options_df['持仓量'] >= 100) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 90)
        ].copy()
        
        if liquid_options.empty:
            liquid_options = options_df[
                (options_df['成交量'] >= 20) &
                (options_df['持仓量'] >= 50) &
                (options_df['剩余天数'] >= 14) &
                (options_df['剩余天数'] <= 120)
            ].copy()
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        # 看涨蝶式
        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        if len(calls) >= 3:
            calls = calls.sort_values('行权价')
            
            # 选择接近平值的三个行权价
            calls['strike_diff'] = np.abs(calls['行权价'] - atm_price)
            calls = calls.sort_values('strike_diff')
            
            if len(calls) >= 3:
                # 获取不同行权价的期权
                strikes = sorted(calls['行权价'].unique())
                if len(strikes) >= 3:
                    # 选择等间距的三个行权价
                    low_strike = strikes[0]
                    mid_strike = strikes[len(strikes)//2]
                    high_strike = strikes[-1]
                    
                    # 确保间距相等或接近
                    spread1 = mid_strike - low_strike
                    spread2 = high_strike - mid_strike
                    
                    if abs(spread1 - spread2) / atm_price < 0.02:  # 间距差异小于2%
                        low_opt = calls[calls['行权价'] == low_strike].iloc[0]
                        mid_opt = calls[calls['行权价'] == mid_strike].iloc[0]
                        high_opt = calls[calls['行权价'] == high_strike].iloc[0]
                        
                        strategy = self._build_butterfly_strategy(
                            low_opt, mid_opt, high_opt, 'CALL',
                            market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        # 看跌蝶式
        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        if len(puts) >= 3:
            puts = puts.sort_values('行权价')
            puts['strike_diff'] = np.abs(puts['行权价'] - atm_price)
            puts = puts.sort_values('strike_diff')
            
            if len(puts) >= 3:
                strikes = sorted(puts['行权价'].unique())
                if len(strikes) >= 3:
                    low_strike = strikes[0]
                    mid_strike = strikes[len(strikes)//2]
                    high_strike = strikes[-1]
                    
                    spread1 = mid_strike - low_strike
                    spread2 = high_strike - mid_strike
                    
                    if abs(spread1 - spread2) / atm_price < 0.02:
                        low_opt = puts[puts['行权价'] == low_strike].iloc[0]
                        mid_opt = puts[puts['行权价'] == mid_strike].iloc[0]
                        high_opt = puts[puts['行权价'] == high_strike].iloc[0]
                        
                        strategy = self._build_butterfly_strategy(
                            low_opt, mid_opt, high_opt, 'PUT',
                            market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        return strategies

    def _generate_butterfly_matrix(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            liquid_options = options_df[
                (options_df['成交量'] >= 80) &
                (options_df['持仓量'] >= 150) &
                (options_df['剩余天数'] >= 20) &
                (options_df['剩余天数'] <= 70)
            ].copy()
            if liquid_options.empty:
                return strategies
            atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
            if atm_price <= 0:
                return strategies
            candidate_strategies = []
            for opt_type in ['CALL', 'PUT']:
                df = liquid_options[liquid_options['期权类型'] == opt_type].copy()
                if df.empty:
                    continue
                strikes_raw = pd.Series(pd.to_numeric(df['行权价'], errors='coerce')).dropna().unique()
                if len(strikes_raw) < 5:
                    continue
                strikes = sorted([float(x) for x in strikes_raw])
                strike_arr = np.array(strikes)
                center_idx = int(np.abs(strike_arr - float(atm_price)).argmin())
                lo = max(center_idx - 4, 0)
                hi = min(center_idx + 4, len(strikes) - 1)
                for i in range(lo, hi + 1):
                    for step in [1, 2]:
                        if i - step < 0 or i + step >= len(strikes):
                            continue
                        low_strike = strikes[i - step]
                        mid_strike = strikes[i]
                        high_strike = strikes[i + step]
                        if abs((mid_strike - low_strike) - (high_strike - mid_strike)) / atm_price > 0.012:
                            continue
                        low_opt = df[df['行权价'] == low_strike]
                        mid_opt = df[df['行权价'] == mid_strike]
                        high_opt = df[df['行权价'] == high_strike]
                        if low_opt.empty or mid_opt.empty or high_opt.empty:
                            continue
                        strategy = self._build_butterfly_strategy(
                            low_opt.iloc[0], mid_opt.iloc[0], high_opt.iloc[0], opt_type,
                            market_state, vol_features
                        )
                        if not strategy:
                            continue
                        score = self._score_butterfly_strategy(strategy, atm_price)
                        strategy['策略标签'] = '蝶式矩阵筛选'
                        strategy['策略评分'] = score
                        strategy['策略要点'] = {
                            '矩阵得分': score,
                            '中枢行权价': mid_strike,
                            '对称间距': mid_strike - low_strike
                        }
                        candidate_strategies.append(strategy)
            if not candidate_strategies:
                return strategies
            candidate_strategies.sort(key=lambda x: float(x.get('策略评分', 0) or 0), reverse=True)
            return candidate_strategies[:6]
        except Exception:
            return strategies

    def _generate_risk_reversals(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            liquid_options = options_df[
                (options_df['成交量'] >= 50) &
                (options_df['持仓量'] >= 120) &
                (options_df['剩余天数'] >= 20) &
                (options_df['剩余天数'] <= 90)
            ].copy()
            if liquid_options.empty:
                return strategies
            atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
            if atm_price <= 0:
                return strategies
            skew_label = str(vol_features.get('Skew', '') or '')
            direction = str(market_state.get('期货方向', '') or '')
            sentiment = str(vol_features.get('市场情绪', '') or '')
            bullish_allowed = ('上涨' in direction) or ('偏多' in sentiment) or ('Put' in skew_label) or ('左' in skew_label) or ('负' in skew_label)
            bearish_allowed = ('下跌' in direction) or ('偏空' in sentiment) or ('Call' in skew_label) or ('右' in skew_label)
            if bullish_allowed:
                call_buy = self._pick_option_by_delta(liquid_options, 'CALL', 0.25) or self._pick_nearest_option(liquid_options, 'CALL', atm_price * 1.05)
                put_sell = self._pick_option_by_delta(liquid_options, 'PUT', -0.25) or self._pick_nearest_option(liquid_options, 'PUT', atm_price * 0.95)
                if call_buy is not None and put_sell is not None:
                    strategy = self._build_risk_reversal_strategy(call_buy, put_sell, market_state, vol_features, '看涨')
                    if strategy:
                        strategies.append(strategy)
            if bearish_allowed:
                put_buy = self._pick_option_by_delta(liquid_options, 'PUT', -0.25) or self._pick_nearest_option(liquid_options, 'PUT', atm_price * 0.95)
                call_sell = self._pick_option_by_delta(liquid_options, 'CALL', 0.25) or self._pick_nearest_option(liquid_options, 'CALL', atm_price * 1.05)
                if put_buy is not None and call_sell is not None:
                    strategy = self._build_risk_reversal_strategy(put_buy, call_sell, market_state, vol_features, '看跌')
                    if strategy:
                        strategies.append(strategy)
            return strategies
        except Exception:
            return strategies

    def _generate_term_structure_timing(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            universe = self._get_underlying_option_universe(options_df, market_state)
            liquid_options = universe[
                (universe['成交量'] >= 60) &
                (universe['持仓量'] >= 120) &
                (universe['剩余天数'] >= 15) &
                (universe['剩余天数'] <= 240)
            ].copy()
            if liquid_options.empty:
                return strategies
            atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
            if atm_price <= 0:
                return strategies
            term_structure = str(vol_features.get('Term Structure', '') or '')
            sentiment = str(vol_features.get('市场情绪', '') or '')
            short_iv = self._safe_float(vol_features.get('Short_IV', None))
            long_iv = self._safe_float(vol_features.get('Long_IV', None))
            term_spread = self._safe_float(vol_features.get('Term_Spread', None))
            spread = None
            if short_iv is not None and long_iv is not None:
                spread = short_iv - long_iv
            elif term_spread is not None:
                spread = term_spread
            signal = None
            if ('倒挂' in term_structure) or ('Backward' in term_structure) or (spread is not None and spread > 0):
                signal = '卖近买远'
            elif ('正常' in term_structure) or ('Contango' in term_structure) or (spread is not None and spread < 0):
                if '波动' in sentiment or '事件' in sentiment or '震荡' in sentiment:
                    signal = '买近卖远'
            if signal is None:
                return strategies
            direction = str(market_state.get('期货方向', '') or '')
            opt_type = 'CALL' if '上涨' in direction else ('PUT' if '下跌' in direction else 'CALL')
            df = liquid_options[liquid_options['期权类型'] == opt_type].copy()
            if df.empty:
                return strategies
            df['strike_diff'] = np.abs(df['行权价'] - atm_price)
            target_strike = df.sort_values('strike_diff').iloc[0]['行权价']
            same_strike = df[df['行权价'] == target_strike].copy()
            if len(same_strike) < 2:
                return strategies
            same_strike = same_strike.sort_values('剩余天数')
            short_opt = same_strike.iloc[0]
            long_opt = same_strike.iloc[-1]
            if long_opt['剩余天数'] - short_opt['剩余天数'] < 30:
                return strategies
            strategy = self._build_term_structure_calendar(short_opt, long_opt, opt_type, market_state, vol_features, signal)
            if strategy:
                strategies.append(strategy)
            return strategies
        except Exception:
            return strategies
    
    def _build_butterfly_strategy(self, low_opt, mid_opt, high_opt, opt_type,
                                   market_state, vol_features):
        """构建蝶式价差策略详情"""
        try:
            multiplier = float(low_opt.get('合约乘数', 1) or 1)
            leg_low = self._make_option_leg(low_opt, '买入', 1)
            leg_mid = self._make_option_leg(mid_opt, '卖出', 2)
            leg_high = self._make_option_leg(high_opt, '买入', 1)
            legs = [leg_low, leg_mid, leg_high]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            
            # 最大收益：标的价格=中间行权价时
            strike_spread = (mid_opt['行权价'] - low_opt['行权价']) * multiplier
            max_profit = strike_spread - abs(net_debit)
            
            # 最大损失 = 净成本
            max_loss = abs(net_debit)
            
            # 盈亏平衡点
            if net_debit < 0:  # 净收入
                breakeven_lower = low_opt['行权价'] - (net_debit / multiplier)
                breakeven_upper = high_opt['行权价'] + (net_debit / multiplier)
            else:  # 净支出
                breakeven_lower = low_opt['行权价'] + (net_debit / multiplier)
                breakeven_upper = high_opt['行权价'] - (net_debit / multiplier)
            
            strategy = {
                '策略类型': f'{opt_type}蝶式价差',
                '策略描述': '买入1个低行权价+卖出2个中间行权价+买入1个高行权价，适合窄幅震荡',
                '适用场景': f'窄幅震荡，预期价格接近{mid_opt["行权价"]:.2f}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0,
                    '最佳价位': mid_opt['行权价']
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '盈利区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '适合窄幅震荡市场，预期标的价格接近中间行权价',
                    '标的价格偏离中间行权价过多将减少收益',
                    '最大收益有限，最大亏损为净成本',
                    '需要较准确的价格预测',
                    f'当前市场状态：{market_state.get("联动状态", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(low_opt.get("Delta", 0) - 2*mid_opt.get("Delta", 0) + high_opt.get("Delta", 0)):.4f}',
                    'Gamma': f'{(low_opt.get("Gamma", 0) - 2*mid_opt.get("Gamma", 0) + high_opt.get("Gamma", 0)):.6f}',
                    'Theta': f'{(low_opt.get("Theta", 0) - 2*mid_opt.get("Theta", 0) + high_opt.get("Theta", 0)):.4f}',
                    'Vega': f'{(low_opt.get("Vega", 0) - 2*mid_opt.get("Vega", 0) + high_opt.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建蝶式价差策略失败: {e}")
            return None

    def _build_risk_reversal_strategy(self, opt_buy, opt_sell, market_state, vol_features, direction: str):
        try:
            multiplier = float(opt_buy.get('合约乘数', 1) or 1)
            leg_buy = self._make_option_leg(opt_buy, '买入', 1)
            leg_sell = self._make_option_leg(opt_sell, '卖出', 1)
            legs = [leg_buy, leg_sell]
            k_buy = float(opt_buy.get('行权价', 0) or 0)
            k_sell = float(opt_sell.get('行权价', 0) or 0)
            if direction == '看涨':
                desc = '买入价外CALL+卖出价外PUT，借助偏斜降低成本做多'
                scene = f'偏斜结构PUT更贵、且方向偏多；IV/RV={vol_features.get("IV_RV", "")}；期限结构={vol_features.get("Term Structure", "")}'
                max_profit = '理论无限（上行）'
                max_loss = '下行风险较大（PUT卖方）'
            else:
                desc = '买入价外PUT+卖出价外CALL，借助偏斜降低成本做空'
                scene = f'偏斜结构明显且方向偏空；IV/RV={vol_features.get("IV_RV", "")}；期限结构={vol_features.get("Term Structure", "")}'
                max_profit = '理论无限（下行）'
                max_loss = '上行风险较大（CALL卖方）'
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            premium = -net_cashflow
            if direction == '看涨':
                breakeven_lower = k_sell - (premium / multiplier)
                breakeven_upper = k_buy + (premium / multiplier)
            else:
                breakeven_lower = k_buy - (premium / multiplier)
                breakeven_upper = k_sell + (premium / multiplier)
            strategy = {
                '策略类型': f'{direction}风险反转（Risk Reversal）',
                '策略描述': desc,
                '适用场景': scene,
                '策略标签': '风险反转/偏斜交易',
                '择时依据': {
                    '偏斜方向': vol_features.get('Skew', ''),
                    'IV/RV': vol_features.get('IV_RV', ''),
                    '期限结构': vol_features.get('Term Structure', '')
                },
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(premium, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper
                },
                '风险提示': [
                    '风险反转等价于带偏斜定价的方向性敞口',
                    '卖方腿带来保证金占用与尾部风险',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(opt_buy.get("Delta", 0) or 0) - float(opt_sell.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(opt_buy.get("Gamma", 0) or 0) - float(opt_sell.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(opt_buy.get("Theta", 0) or 0) - float(opt_sell.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(opt_buy.get("Vega", 0) or 0) - float(opt_sell.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建风险反转策略失败: {e}")
            return None

    def _build_term_structure_calendar(self, short_opt, long_opt, opt_type, market_state, vol_features, signal: str):
        try:
            if signal == '卖近买远':
                leg_short = self._make_option_leg(short_opt, '卖出', 1)
                leg_long = self._make_option_leg(long_opt, '买入', 1)
                desc = '卖出短期+买入长期同档期权，做空近端高IV与时间价值'
                scene = '期限结构倒挂或短期IV偏高时使用'
            else:
                leg_short = self._make_option_leg(short_opt, '买入', 1)
                leg_long = self._make_option_leg(long_opt, '卖出', 1)
                desc = '买入短期+卖出长期同档期权，做多近端IV上行与事件波动'
                scene = '期限结构正常且近端IV偏低、事件驱动时使用'
            legs = [leg_short, leg_long]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            strike = float(short_opt.get('行权价', 0) or 0)
            breakeven_lower = strike * 0.95
            breakeven_upper = strike * 1.05
            strategy = {
                '策略类型': f'{opt_type}期限结构择时',
                '策略描述': desc,
                '适用场景': f'{scene}；期限结构={vol_features.get("Term Structure", "")}',
                '策略标签': '期限结构择时',
                '择时依据': {
                    '期限结构': vol_features.get('Term Structure', ''),
                    '短期IV': vol_features.get('Short_IV', ''),
                    '长期IV': vol_features.get('Long_IV', ''),
                    '结构差': vol_features.get('Term_Spread', '')
                },
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': '依赖期限结构与波动率变化',
                    '最大亏损': abs(net_debit)
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '最佳区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '期限结构变化与波动率均值回归假设可能失效',
                    '近端事件冲击会放大盈亏波动',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(short_opt.get("Delta", 0) or 0) - float(long_opt.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(short_opt.get("Gamma", 0) or 0) - float(long_opt.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(short_opt.get("Theta", 0) or 0) - float(long_opt.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(short_opt.get("Vega", 0) or 0) - float(long_opt.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建期限结构择时策略失败: {e}")
            return None
    
    def _generate_vertical_spreads(self, options_df, market_state, vol_features):
        """生成垂直价差策略
        
        看涨垂直价差（Bull Call Spread）= 买入低行权价CALL + 卖出高行权价CALL
        看跌垂直价差（Bear Put Spread）= 买入高行权价PUT + 卖出低行权价PUT
        适合有明确方向预期的市场
        """
        strategies = []
        
        # 根据市场方向选择策略
        direction = market_state.get('期货方向', '')
        sentiment = market_state.get('期权情绪', '')
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 50) &
            (options_df['持仓量'] >= 100) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 90)
        ].copy()
        
        if liquid_options.empty:
            liquid_options = options_df[
                (options_df['成交量'] >= 20) &
                (options_df['持仓量'] >= 50) &
                (options_df['剩余天数'] >= 14) &
                (options_df['剩余天数'] <= 120)
            ].copy()
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        # 看涨垂直价差（适合看涨市场）
        if '上涨' in direction or '看涨' in sentiment:
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            if len(calls) >= 2:
                calls = calls.sort_values('行权价')
                
                # 选择平值和价外期权
                atm_calls = calls[calls['行权价'] <= atm_price * 1.05]
                otm_calls = calls[calls['行权价'] > atm_price * 1.05]
                
                if len(atm_calls) > 0 and len(otm_calls) > 0:
                    buy_call = atm_calls.iloc[-1]  # 买入接近平值
                    sell_call = otm_calls.iloc[0]   # 卖出价外
                    
                    # 确保有足够的行权价间距
                    if sell_call['行权价'] - buy_call['行权价'] >= atm_price * 0.02:
                        strategy = self._build_bull_call_spread(
                            buy_call, sell_call, market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        # 看跌垂直价差（适合看跌市场）
        if '下跌' in direction or '看跌' in sentiment:
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            if len(puts) >= 2:
                puts = puts.sort_values('行权价', ascending=False)
                
                # 选择平值和价外期权
                atm_puts = puts[puts['行权价'] >= atm_price * 0.95]
                otm_puts = puts[puts['行权价'] < atm_price * 0.95]
                
                if len(atm_puts) > 0 and len(otm_puts) > 0:
                    buy_put = atm_puts.iloc[-1]   # 买入接近平值
                    sell_put = otm_puts.iloc[0]   # 卖出价外
                    
                    if buy_put['行权价'] - sell_put['行权价'] >= atm_price * 0.02:
                        strategy = self._build_bear_put_spread(
                            buy_put, sell_put, market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        return strategies
    
    def _build_bull_call_spread(self, buy_call, sell_call, market_state, vol_features):
        """构建看涨垂直价差策略"""
        try:
            multiplier = float(buy_call.get('合约乘数', 1) or 1)
            leg_buy = self._make_option_leg(buy_call, '买入', 1)
            leg_sell = self._make_option_leg(sell_call, '卖出', 1)
            legs = [leg_buy, leg_sell]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            
            # 最大收益 = 行权价差 - 净成本
            strike_spread = (sell_call['行权价'] - buy_call['行权价']) * multiplier
            max_profit = strike_spread - net_debit
            
            # 最大损失 = 净成本
            max_loss = net_debit
            
            # 盈亏平衡点
            breakeven = buy_call['行权价'] + (net_debit / multiplier)
            
            strategy = {
                '策略类型': '看涨垂直价差（Bull Call Spread）',
                '策略描述': '买入低行权价看涨期权+卖出高行权价看涨期权，适合温和看涨',
                '适用场景': f'{market_state.get("期货方向", "看涨")}，预期上涨但涨幅有限',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '最大盈利点': sell_call['行权价'],
                    '说明': f'标的价格>{breakeven:.2f}开始盈利'
                },
                '风险提示': [
                    '适合温和看涨行情，预期涨幅在一定范围内',
                    '标的价格超过高行权价后无额外收益',
                    '相比单独买入看涨期权，成本更低但收益受限',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '建议在有明确上涨信号时使用'
                ],
                '希腊字母风险': {
                    'Delta': f'{(buy_call.get("Delta", 0) - sell_call.get("Delta", 0)):.4f}',
                    'Gamma': f'{(buy_call.get("Gamma", 0) - sell_call.get("Gamma", 0)):.6f}',
                    'Theta': f'{(buy_call.get("Theta", 0) - sell_call.get("Theta", 0)):.4f}',
                    'Vega': f'{(buy_call.get("Vega", 0) - sell_call.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建看涨垂直价差策略失败: {e}")
            return None
    
    def _build_bear_put_spread(self, buy_put, sell_put, market_state, vol_features):
        """构建看跌垂直价差策略"""
        try:
            multiplier = float(buy_put.get('合约乘数', 1) or 1)
            leg_buy = self._make_option_leg(buy_put, '买入', 1)
            leg_sell = self._make_option_leg(sell_put, '卖出', 1)
            legs = [leg_buy, leg_sell]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            
            # 最大收益 = 行权价差 - 净成本
            strike_spread = (buy_put['行权价'] - sell_put['行权价']) * multiplier
            max_profit = strike_spread - net_debit
            
            # 最大损失 = 净成本
            max_loss = net_debit
            
            # 盈亏平衡点
            breakeven = buy_put['行权价'] - (net_debit / multiplier)
            
            strategy = {
                '策略类型': '看跌垂直价差（Bear Put Spread）',
                '策略描述': '买入高行权价看跌期权+卖出低行权价看跌期权，适合温和看跌',
                '适用场景': f'{market_state.get("期货方向", "看跌")}，预期下跌但跌幅有限',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '最大盈利点': sell_put['行权价'],
                    '说明': f'标的价格<{breakeven:.2f}开始盈利'
                },
                '风险提示': [
                    '适合温和看跌行情，预期跌幅在一定范围内',
                    '标的价格低于低行权价后无额外收益',
                    '相比单独买入看跌期权，成本更低但收益受限',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '建议在有明确下跌信号时使用'
                ],
                '希腊字母风险': {
                    'Delta': f'{(buy_put.get("Delta", 0) - sell_put.get("Delta", 0)):.4f}',
                    'Gamma': f'{(buy_put.get("Gamma", 0) - sell_put.get("Gamma", 0)):.6f}',
                    'Theta': f'{(buy_put.get("Theta", 0) - sell_put.get("Theta", 0)):.4f}',
                    'Vega': f'{(buy_put.get("Vega", 0) - sell_put.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建看跌垂直价差策略失败: {e}")
            return None
    
    def _generate_ratio_spreads(self, options_df, market_state, vol_features):
        """生成比率价差策略
        
        看涨比率价差 = 买入1个低行权价CALL + 卖出N个高行权价CALL (N>1)
        看跌比率价差 = 买入1个高行权价PUT + 卖出N个低行权价PUT (N>1)
        适合有方向预期但认为涨/跌幅有限的市场
        """
        strategies = []
        
        # 根据市场方向选择策略
        direction = market_state.get('期货方向', '')
        sentiment = market_state.get('期权情绪', '')
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 50) &
            (options_df['持仓量'] >= 100) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 60)
        ].copy()
        
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        skew_label = str(vol_features.get('Skew', '') or '')
        bullish_ok = ('上涨' in direction) or ('看涨' in sentiment) or ('偏多' in sentiment) or ('Put' in skew_label) or ('左' in skew_label) or ('负' in skew_label)
        bearish_ok = ('下跌' in direction) or ('看跌' in sentiment) or ('偏空' in sentiment) or ('Call' in skew_label) or ('右' in skew_label)

        if not bullish_ok and not bearish_ok:
            bullish_ok = True
            bearish_ok = True

        # 看涨比率价差（适合温和看涨）
        if bullish_ok:
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            if len(calls) >= 2:
                calls = calls.sort_values('行权价')
                
                buy_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.0) if hasattr(self, '_pick_nearest_option') else None
                sell_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.06) if hasattr(self, '_pick_nearest_option') else None
                if buy_call is None:
                    atm_calls = calls[calls['行权价'] <= atm_price * 1.03]
                    buy_call = atm_calls.iloc[-1] if len(atm_calls) > 0 else calls.iloc[0]
                if sell_call is None:
                    otm_calls = calls[calls['行权价'] > atm_price * 1.04]
                    sell_call = otm_calls.iloc[0] if len(otm_calls) > 0 else calls.iloc[-1]
                if buy_call is not None and sell_call is not None:
                    if float(buy_call.get('行权价', 0) or 0) >= float(sell_call.get('行权价', 0) or 0):
                        sell_call = calls[calls['行权价'] > float(buy_call.get('行权价', 0) or 0)]
                        sell_call = sell_call.iloc[0] if not sell_call.empty else None
                    if sell_call is not None:
                        strategy = self._build_ratio_spread_strategy(
                            buy_call, sell_call, 2, 'CALL',
                            market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        # 看跌比率价差（适合温和看跌）
        if bearish_ok:
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            if len(puts) >= 2:
                puts = puts.sort_values('行权价', ascending=False)
                
                buy_put = self._pick_nearest_option(puts, 'PUT', atm_price * 1.0) if hasattr(self, '_pick_nearest_option') else None
                sell_put = self._pick_nearest_option(puts, 'PUT', atm_price * 0.94) if hasattr(self, '_pick_nearest_option') else None
                if buy_put is None:
                    atm_puts = puts[puts['行权价'] >= atm_price * 0.97]
                    buy_put = atm_puts.iloc[-1] if len(atm_puts) > 0 else puts.iloc[0]
                if sell_put is None:
                    otm_puts = puts[puts['行权价'] < atm_price * 0.96]
                    sell_put = otm_puts.iloc[0] if len(otm_puts) > 0 else puts.iloc[-1]
                if buy_put is not None and sell_put is not None:
                    if float(buy_put.get('行权价', 0) or 0) <= float(sell_put.get('行权价', 0) or 0):
                        sell_put = puts[puts['行权价'] < float(buy_put.get('行权价', 0) or 0)]
                        sell_put = sell_put.iloc[0] if not sell_put.empty else None
                    if sell_put is not None:
                        strategy = self._build_ratio_spread_strategy(
                            buy_put, sell_put, 2, 'PUT',
                            market_state, vol_features
                        )
                        if strategy:
                            strategies.append(strategy)
        
        return strategies
    
    def _build_ratio_spread_strategy(self, buy_opt, sell_opt, ratio, opt_type,
                                      market_state, vol_features):
        """构建比率价差策略"""
        try:
            multiplier = float(buy_opt.get('合约乘数', 1) or 1)
            leg_buy = self._make_option_leg(buy_opt, '买入', 1)
            leg_sell = self._make_option_leg(sell_opt, '卖出', int(ratio))
            legs = [leg_buy, leg_sell]
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            net_debit = -net_cashflow
            
            # 最大收益点：标的价格=卖出行权价
            if opt_type == 'CALL':
                max_profit_price = sell_opt['行权价']
                max_profit = (sell_opt['行权价'] - buy_opt['行权价']) * multiplier - net_debit
                
                # 盈亏平衡点
                breakeven_lower = buy_opt['行权价'] + (net_debit / multiplier) if net_debit > 0 else buy_opt['行权价']
                # 上方盈亏平衡点计算较复杂
                breakeven_upper = sell_opt['行权价'] + max_profit / (multiplier * (ratio - 1))
                
            else:  # PUT
                max_profit_price = sell_opt['行权价']
                max_profit = (buy_opt['行权价'] - sell_opt['行权价']) * multiplier - net_debit
                
                breakeven_upper = buy_opt['行权价'] - (net_debit / multiplier) if net_debit > 0 else buy_opt['行权价']
                breakeven_lower = sell_opt['行权价'] - max_profit / (multiplier * (ratio - 1))
            
            # 最大损失（如果标的价格大幅偏离）
            max_loss = net_debit if net_debit > 0 else 0  # 简化计算
            
            strategy = {
                '策略类型': f'{opt_type}比率价差（{ratio}:1）',
                '策略描述': f'买入1个{"低" if opt_type == "CALL" else "高"}行权价{opt_type}+卖出{ratio}个{"高" if opt_type == "CALL" else "低"}行权价{opt_type}',
                '适用场景': f'{market_state.get("期货方向", "")}，预期{"涨" if opt_type == "CALL" else "跌"}幅有限',
                '操作建议': legs,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow,
                    '净权利金支出': max(net_debit, 0),
                    '最大盈利': max_profit,
                    '最大盈利点': max_profit_price,
                    '最大亏损': '潜在无限' if net_debit < 0 else max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else '不确定'
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '最佳区间': f'{min(breakeven_lower, breakeven_upper):.2f} - {max_profit_price:.2f}'
                },
                '风险提示': [
                    f'适合温和{"看涨" if opt_type == "CALL" else "看跌"}，预期{"涨" if opt_type == "CALL" else "跌"}幅有限',
                    f'标的价格超过卖出行权价过多将产生亏损',
                    '卖出期权数量多于买入，有潜在无限风险',
                    '需要密切监控，及时止损或调整',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '建议有经验者使用，风险较高'
                ],
                '希腊字母风险': {
                    'Delta': f'{(buy_opt.get("Delta", 0) - ratio * sell_opt.get("Delta", 0)):.4f}',
                    'Gamma': f'{(buy_opt.get("Gamma", 0) - ratio * sell_opt.get("Gamma", 0)):.6f}',
                    'Theta': f'{(buy_opt.get("Theta", 0) - ratio * sell_opt.get("Theta", 0)):.4f}',
                    'Vega': f'{(buy_opt.get("Vega", 0) - ratio * sell_opt.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建比率价差策略失败: {e}")
            return None
    
    def _generate_straddles(self, options_df, market_state, vol_features):
        """生成跨式/宽跨式策略
        
        跨式（Straddle）= 买入相同行权价的CALL+PUT
        宽跨式（Strangle）= 买入不同行权价的价外CALL+PUT
        适合预期大幅波动但方向不明的市场
        """
        strategies = []
        
        # 跨式/宽跨式适合场景：预期大幅波动、波动率偏低
        sentiment = vol_features.get('市场情绪', '')
        skew = vol_features.get('Skew', '')
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 100) &
            (options_df['持仓量'] >= 200) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 60)
        ].copy()
        
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        # 1. 买入跨式（Straddle）
        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        
        if not calls.empty and not puts.empty:
            # 选择平值期权
            calls['strike_diff'] = np.abs(calls['行权价'] - atm_price)
            puts['strike_diff'] = np.abs(puts['行权价'] - atm_price)
            
            atm_call = calls.loc[calls['strike_diff'].idxmin()]
            atm_put = puts.loc[puts['strike_diff'].idxmin()]
            
            # 如果是同一个行权价，构建跨式
            if abs(atm_call['行权价'] - atm_put['行权价']) < atm_price * 0.01:
                strategy = self._build_straddle_strategy(
                    atm_call, atm_put, '买入', market_state, vol_features
                )
                if strategy:
                    strategies.append(strategy)
        
        # 2. 买入宽跨式（Strangle）
        if not calls.empty and not puts.empty:
            # 选择价外期权
            otm_calls = calls[calls['行权价'] > atm_price * 1.03].copy()
            otm_puts = puts[puts['行权价'] < atm_price * 0.97].copy()
            
            if not otm_calls.empty and not otm_puts.empty:
                # 选择虚值档位较近的期权
                buy_call = otm_calls.sort_values('行权价').iloc[0]
                buy_put = otm_puts.sort_values('行权价', ascending=False).iloc[0]
                
                strategy = self._build_strangle_strategy(
                    buy_call, buy_put, '买入', market_state, vol_features
                )
                if strategy:
                    strategies.append(strategy)
        
        # 3. 卖出宽跨式（Strangle）- 适合震荡市
        if '震荡' in sentiment:
            if not calls.empty and not puts.empty:
                otm_calls = calls[calls['行权价'] > atm_price * 1.05].copy()
                otm_puts = puts[puts['行权价'] < atm_price * 0.95].copy()
                
                if not otm_calls.empty and not otm_puts.empty:
                    sell_call = otm_calls.sort_values('行权价').iloc[0]
                    sell_put = otm_puts.sort_values('行权价', ascending=False).iloc[0]
                    
                    strategy = self._build_strangle_strategy(
                        sell_call, sell_put, '卖出', market_state, vol_features
                    )
                    if strategy:
                        strategies.append(strategy)
        
        return strategies

    def _generate_expiry_directional_options(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            short_term = options_df[
                (options_df['剩余天数'] >= 1) &
                (options_df['剩余天数'] <= 7)
            ].copy()
            if short_term.empty:
                short_term = options_df[
                    (options_df['剩余天数'] >= 1) &
                    (options_df['剩余天数'] <= 10)
                ].copy()
            if short_term.empty:
                return strategies
            liquid_options = short_term[
                (short_term['成交量'] >= 10) &
                (short_term['持仓量'] >= 20)
            ].copy()
            if liquid_options.empty:
                liquid_options = short_term.copy()
            atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
            direction = str(market_state.get('期货方向', '') or '')
            sentiment = str(market_state.get('期权情绪', '') or '')
            bullish = ('上涨' in direction) or ('看涨' in sentiment)
            bearish = ('下跌' in direction) or ('看跌' in sentiment)
            if not bullish and not bearish:
                bullish = True
                bearish = True
            if bullish:
                calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
                if not calls.empty:
                    call_opt = self._pick_option_by_delta(calls, 'CALL', 0.35) if hasattr(self, '_pick_option_by_delta') else None
                    if call_opt is None:
                        call_opt = self._pick_nearest_option(calls, 'CALL', atm_price * 1.02)
                    if call_opt is not None:
                        strategy = self._build_expiry_option_strategy(call_opt, market_state, vol_features, 'CALL')
                        if strategy:
                            strategies.append(strategy)
            if bearish:
                puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
                if not puts.empty:
                    put_opt = self._pick_option_by_delta(puts, 'PUT', -0.35) if hasattr(self, '_pick_option_by_delta') else None
                    if put_opt is None:
                        put_opt = self._pick_nearest_option(puts, 'PUT', atm_price * 0.98)
                    if put_opt is not None:
                        strategy = self._build_expiry_option_strategy(put_opt, market_state, vol_features, 'PUT')
                        if strategy:
                            strategies.append(strategy)
            return strategies
        except Exception:
            return strategies

    def _build_expiry_option_strategy(self, opt_row, market_state, vol_features, opt_type: str):
        try:
            leg = self._make_option_leg(opt_row, '买入', 1)
            strike = float(opt_row.get('行权价', 0) or 0)
            multiplier = float(opt_row.get('合约乘数', 1) or 1)
            premium = -float(leg.get('权利金', 0) or 0)
            if opt_type == 'CALL':
                breakeven = strike + premium / multiplier
                desc = '临近到期的单边上行博弈，小成本博取大波动'
                scene = f'短期事件或趋势驱动；IV/RV={vol_features.get("IV_RV", "")}'
                max_profit = '理论无限'
            else:
                breakeven = strike - premium / multiplier
                desc = '临近到期的单边下行博弈，小成本博取大波动'
                scene = f'短期事件或趋势驱动；IV/RV={vol_features.get("IV_RV", "")}'
                max_profit = '理论无限'
            strategy = {
                '策略类型': f'末日期权单边（买入{opt_type}）',
                '策略描述': desc,
                '适用场景': scene,
                '策略标签': '末日期权/事件驱动',
                '操作建议': [leg],
                '成本分析': {
                    '权利金支出': premium,
                    '最大盈利': max_profit,
                    '最大亏损': premium
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven
                },
                '风险提示': [
                    '末日期权时间价值衰减极快，需严格控制仓位',
                    '适合对当日或短期方向有强观点',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(opt_row.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(opt_row.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(opt_row.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(opt_row.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception:
            return None

    def _generate_credit_spreads(self, options_df, market_state, vol_features):
        strategies = []
        liquid_options = options_df[
            (options_df['成交量'] >= 50) &
            (options_df['持仓量'] >= 100) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 60)
        ].copy()
        if liquid_options.empty:
            return strategies
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        direction = market_state.get('期货方向', '')
        sentiment = market_state.get('期权情绪', '')
        iv_rv = None
        try:
            iv_rv = float(str(vol_features.get('IV_RV', '')).replace('%', '').strip())
        except Exception:
            iv_rv = None

        if '上涨' in direction or '震荡' in sentiment or (iv_rv is not None and iv_rv > 1.05):
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            if not puts.empty:
                sell_put = self._pick_nearest_option(puts, 'PUT', atm_price * 0.98)
                buy_put = self._pick_nearest_option(puts, 'PUT', atm_price * 0.94)
                if sell_put is not None and buy_put is not None:
                    if float(sell_put.get('行权价', 0) or 0) > float(buy_put.get('行权价', 0) or 0):
                        strategy = self._build_bull_put_spread(sell_put, buy_put, market_state, vol_features)
                        if strategy:
                            strategies.append(strategy)

        if '下跌' in direction or '震荡' in sentiment or (iv_rv is not None and iv_rv > 1.05):
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            if not calls.empty:
                sell_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.02)
                buy_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.06)
                if sell_call is not None and buy_call is not None:
                    if float(buy_call.get('行权价', 0) or 0) > float(sell_call.get('行权价', 0) or 0):
                        strategy = self._build_bear_call_spread(sell_call, buy_call, market_state, vol_features)
                        if strategy:
                            strategies.append(strategy)

        return strategies

    def _generate_short_premium_neutral_strategies(self, options_df, market_state, vol_features):
        strategies = []
        liquid_options = options_df[
            (options_df['成交量'] >= 100) &
            (options_df['持仓量'] >= 200) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 45)
        ].copy()
        if liquid_options.empty:
            return strategies
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        sentiment = market_state.get('期权情绪', '')
        iv_rv = None
        try:
            iv_rv = float(str(vol_features.get('IV_RV', '')).replace('%', '').strip())
        except Exception:
            iv_rv = None

        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        if calls.empty or puts.empty:
            return strategies
        calls['strike_diff'] = np.abs(calls['行权价'] - atm_price)
        puts['strike_diff'] = np.abs(puts['行权价'] - atm_price)
        atm_call = calls.loc[calls['strike_diff'].idxmin()]
        atm_put = puts.loc[puts['strike_diff'].idxmin()]

        if ('震荡' in sentiment) or (iv_rv is not None and iv_rv > 1.1):
            if abs(atm_call['行权价'] - atm_put['行权价']) < atm_price * 0.01:
                strategy = self._build_straddle_strategy(atm_call, atm_put, '卖出', market_state, vol_features)
                if strategy:
                    strategies.append(strategy)

        return strategies

    def _generate_iron_butterflies(self, options_df, market_state, vol_features):
        strategies = []
        liquid_options = options_df[
            (options_df['成交量'] >= 100) &
            (options_df['持仓量'] >= 200) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 45)
        ].copy()
        if liquid_options.empty:
            return strategies
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        if calls.empty or puts.empty:
            return strategies
        calls['strike_diff'] = np.abs(calls['行权价'] - atm_price)
        puts['strike_diff'] = np.abs(puts['行权价'] - atm_price)
        sell_call = calls.loc[calls['strike_diff'].idxmin()]
        sell_put = puts.loc[puts['strike_diff'].idxmin()]
        buy_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.05)
        buy_put = self._pick_nearest_option(puts, 'PUT', atm_price * 0.95)
        if buy_call is None or buy_put is None:
            return strategies
        if float(buy_call.get('行权价', 0) or 0) <= float(sell_call.get('行权价', 0) or 0):
            return strategies
        if float(sell_put.get('行权价', 0) or 0) <= float(buy_put.get('行权价', 0) or 0):
            return strategies
        strategy = self._build_iron_butterfly_strategy(sell_call, sell_put, buy_call, buy_put, market_state, vol_features)
        if strategy:
            strategies.append(strategy)
        return strategies

    def _generate_jade_lizards(self, options_df, market_state, vol_features):
        strategies = []
        liquid_options = options_df[
            (options_df['成交量'] >= 100) &
            (options_df['持仓量'] >= 200) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 60)
        ].copy()
        if liquid_options.empty:
            return strategies
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        direction = market_state.get('期货方向', '')
        sentiment = market_state.get('期权情绪', '')
        iv_rv = None
        try:
            iv_rv = float(str(vol_features.get('IV_RV', '')).replace('%', '').strip())
        except Exception:
            iv_rv = None
        if not ('上涨' in direction or '震荡' in sentiment or (iv_rv is not None and iv_rv > 1.1)):
            return strategies

        puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
        calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
        if puts.empty or calls.empty:
            return strategies
        sell_put = self._pick_nearest_option(puts, 'PUT', atm_price * 0.95)
        sell_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.03)
        buy_call = self._pick_nearest_option(calls, 'CALL', atm_price * 1.08)
        if sell_put is None or sell_call is None or buy_call is None:
            return strategies
        if float(buy_call.get('行权价', 0) or 0) <= float(sell_call.get('行权价', 0) or 0):
            return strategies
        strategy = self._build_jade_lizard_strategy(sell_put, sell_call, buy_call, market_state, vol_features)
        if strategy:
            strategies.append(strategy)
        return strategies
    
    def _build_straddle_strategy(self, call_opt, put_opt, action, market_state, vol_features):
        """构建跨式策略（使用盘口价格，输出更专业）"""
        try:
            multiplier = float(call_opt.get('合约乘数', 1) or 1)
            strike = float(call_opt.get('行权价', 0) or 0)
            
            leg_call = self._make_option_leg(call_opt, action, 1)
            leg_put = self._make_option_leg(put_opt, action, 1)
            legs = [leg_call, leg_put]
            
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))  # >0 收入，<0 支出
            if action == '买入':
                total_premium = -net_cashflow
                breakeven_upper = strike + (total_premium / multiplier)
                breakeven_lower = strike - (total_premium / multiplier)
                desc = '买入相同行权价的看涨和看跌期权（做多波动率），适合预期大幅波动'
                适用场景 = f'预期标的大幅波动但方向不明；IV相对RV偏低/事件驱动时优先；IV/RV={vol_features.get("IV_RV", "")}；期限结构={vol_features.get("Term Structure", "")}'
                cost_key = '权利金支出'
                max_profit = '无限'
                max_loss = total_premium
                pnl_ratio = '不确定'
            else:
                total_premium = net_cashflow
                breakeven_upper = strike + (total_premium / multiplier)
                breakeven_lower = strike - (total_premium / multiplier)
                desc = '卖出相同行权价的看涨和看跌期权（做空波动率），适合窄幅震荡'
                适用场景 = f'预期标的窄幅震荡；IV相对RV偏高且有明显回落预期；IV/RV={vol_features.get("IV_RV", "")}；期限结构={vol_features.get("Term Structure", "")}'
                cost_key = '权利金收入'
                max_profit = total_premium
                max_loss = '理论无限'
                pnl_ratio = '不确定'
            
            base_delta = float(call_opt.get('Delta', 0) or 0) + float(put_opt.get('Delta', 0) or 0)
            base_gamma = float(call_opt.get('Gamma', 0) or 0) + float(put_opt.get('Gamma', 0) or 0)
            base_theta = float(call_opt.get('Theta', 0) or 0) + float(put_opt.get('Theta', 0) or 0)
            base_vega = float(call_opt.get('Vega', 0) or 0) + float(put_opt.get('Vega', 0) or 0)
            sign = 1 if action == '买入' else -1
            
            strategy = {
                '策略类型': f'{action}跨式（Straddle）',
                '策略描述': desc,
                '适用场景': 适用场景,
                '操作建议': legs,
                '成本分析': {
                    cost_key: total_premium,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': pnl_ratio
                },
                '盈亏平衡点': {
                    '上限': breakeven_upper,
                    '下限': breakeven_lower,
                    '盈利区间': f'<{breakeven_lower:.2f} 或 >{breakeven_upper:.2f}' if action == '买入'
                               else f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}',
                    '买入：主要风险是Theta消耗与事件后IV回落；卖出：主要风险是Gamma冲击与IV飙升',
                    '建议用盘口 bid/ask 中间价做估值，用限价单分批成交减少滑点',
                    f'当前市场状态：{market_state.get("联动状态", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(sign * base_delta):.4f}',
                    'Gamma': f'{(sign * base_gamma):.6f}',
                    'Theta': f'{(sign * base_theta):.4f}',
                    'Vega': f'{(sign * base_vega):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建跨式策略失败: {e}")
            return None

    def _build_strangle_strategy(self, call_opt, put_opt, action, market_state, vol_features):
        """构建宽跨式策略（使用盘口价格）"""
        try:
            multiplier = float(call_opt.get('合约乘数', 1) or 1)
            
            leg_call = self._make_option_leg(call_opt, action, 1)
            leg_put = self._make_option_leg(put_opt, action, 1)
            legs = [leg_call, leg_put]
            
            net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            if action == '买入':
                total_premium = -net_cashflow
                breakeven_upper = float(call_opt.get('行权价', 0) or 0) + (total_premium / multiplier)
                breakeven_lower = float(put_opt.get('行权价', 0) or 0) - (total_premium / multiplier)
                desc = '买入价外看涨+看跌（做多波动率），成本低于跨式，但需要更大波动'
                适用场景 = f'预期大幅波动/重大事件；预算有限；或希望提高容错区间；IV/RV={vol_features.get("IV_RV", "")}'
                cost_key = '权利金支出'
                max_profit = '无限'
                max_loss = total_premium
            else:
                total_premium = net_cashflow
                breakeven_upper = float(call_opt.get('行权价', 0) or 0) + (total_premium / multiplier)
                breakeven_lower = float(put_opt.get('行权价', 0) or 0) - (total_premium / multiplier)
                desc = '卖出价外看涨+看跌（做空波动率），收入高于铁鹰式但风险更大（无保护腿）'
                适用场景 = f'预期较宽区间震荡；IV偏高且回落概率大；需要严格风控；IV/RV={vol_features.get("IV_RV", "")}'
                cost_key = '权利金收入'
                max_profit = total_premium
                max_loss = '理论无限'
            
            base_delta = float(call_opt.get('Delta', 0) or 0) + float(put_opt.get('Delta', 0) or 0)
            base_gamma = float(call_opt.get('Gamma', 0) or 0) + float(put_opt.get('Gamma', 0) or 0)
            base_theta = float(call_opt.get('Theta', 0) or 0) + float(put_opt.get('Theta', 0) or 0)
            base_vega = float(call_opt.get('Vega', 0) or 0) + float(put_opt.get('Vega', 0) or 0)
            sign = 1 if action == '买入' else -1
            
            strategy = {
                '策略类型': f'{action}宽跨式（Strangle）',
                '策略描述': desc,
                '适用场景': 适用场景,
                '操作建议': legs,
                '成本分析': {
                    cost_key: total_premium,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈利区间宽度': float(call_opt.get('行权价', 0) or 0) - float(put_opt.get('行权价', 0) or 0)
                },
                '盈亏平衡点': {
                    '上限': breakeven_upper,
                    '下限': breakeven_lower,
                    '盈利区间': f'<{breakeven_lower:.2f} 或 >{breakeven_upper:.2f}' if action == '买入'
                               else f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}',
                    '卖出宽跨式无保护腿，建议设置硬止损/预案（可转铁鹰式）',
                    '买入宽跨式的关键是“波动要够大”且“发生要够快”（Theta压力）',
                    f'当前市场状态：{market_state.get("联动状态", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(sign * base_delta):.4f}',
                    'Gamma': f'{(sign * base_gamma):.6f}',
                    'Theta': f'{(sign * base_theta):.4f}',
                    'Vega': f'{(sign * base_vega):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建宽跨式策略失败: {e}")
            return None

    def _build_bull_put_spread(self, sell_put, buy_put, market_state, vol_features):
        try:
            multiplier = float(sell_put.get('合约乘数', 1) or 1)
            legs = [
                self._make_option_leg(sell_put, '卖出', 1),
                self._make_option_leg(buy_put, '买入', 1)
            ]
            net_credit = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            strike_sell = float(sell_put.get('行权价', 0) or 0)
            strike_buy = float(buy_put.get('行权价', 0) or 0)
            max_profit = net_credit
            max_loss = (strike_sell - strike_buy) * multiplier - net_credit
            breakeven = strike_sell - (net_credit / multiplier)
            strategy = {
                '策略类型': '牛市价差（卖出PUT信用价差）',
                '策略描述': '卖出价外PUT+买入更价外PUT，收取权利金，限定下行风险',
                '适用场景': f'轻度看涨或震荡，IV偏高回落预期；IV/RV={vol_features.get("IV_RV", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金收入': net_credit,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '下限': breakeven,
                    '盈利区间': f'>{breakeven:.2f}'
                },
                '风险提示': [
                    '标的下跌到卖出PUT以下会产生亏损，亏损被买入PUT限定',
                    '适合波动率偏高时做空波动率与时间价值',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(sell_put.get("Delta", 0) or 0) - float(buy_put.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(sell_put.get("Gamma", 0) or 0) - float(buy_put.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(sell_put.get("Theta", 0) or 0) - float(buy_put.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(sell_put.get("Vega", 0) or 0) - float(buy_put.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建牛市价差失败: {e}")
            return None

    def _build_bear_call_spread(self, sell_call, buy_call, market_state, vol_features):
        try:
            multiplier = float(sell_call.get('合约乘数', 1) or 1)
            legs = [
                self._make_option_leg(sell_call, '卖出', 1),
                self._make_option_leg(buy_call, '买入', 1)
            ]
            net_credit = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            strike_sell = float(sell_call.get('行权价', 0) or 0)
            strike_buy = float(buy_call.get('行权价', 0) or 0)
            max_profit = net_credit
            max_loss = (strike_buy - strike_sell) * multiplier - net_credit
            breakeven = strike_sell + (net_credit / multiplier)
            strategy = {
                '策略类型': '熊市价差（卖出CALL信用价差）',
                '策略描述': '卖出价外CALL+买入更价外CALL，收取权利金，限定上行风险',
                '适用场景': f'轻度看跌或震荡，IV偏高回落预期；IV/RV={vol_features.get("IV_RV", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金收入': net_credit,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '上限': breakeven,
                    '盈利区间': f'<{breakeven:.2f}'
                },
                '风险提示': [
                    '标的上涨到卖出CALL以上会产生亏损，亏损被买入CALL限定',
                    '适合波动率偏高时做空波动率与时间价值',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(sell_call.get("Delta", 0) or 0) - float(buy_call.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(sell_call.get("Gamma", 0) or 0) - float(buy_call.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(sell_call.get("Theta", 0) or 0) - float(buy_call.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(sell_call.get("Vega", 0) or 0) - float(buy_call.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建熊市价差失败: {e}")
            return None

    def _build_iron_butterfly_strategy(self, sell_call, sell_put, buy_call, buy_put, market_state, vol_features):
        try:
            multiplier = float(sell_call.get('合约乘数', 1) or 1)
            legs = [
                self._make_option_leg(sell_call, '卖出', 1),
                self._make_option_leg(sell_put, '卖出', 1),
                self._make_option_leg(buy_call, '买入', 1),
                self._make_option_leg(buy_put, '买入', 1)
            ]
            net_credit = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            strike_body = float(sell_call.get('行权价', 0) or 0)
            wing_call = float(buy_call.get('行权价', 0) or 0)
            wing_put = float(buy_put.get('行权价', 0) or 0)
            upper_width = (wing_call - strike_body) * multiplier
            lower_width = (strike_body - wing_put) * multiplier
            max_profit = net_credit
            max_loss = max(upper_width, lower_width) - net_credit
            breakeven_upper = strike_body + (net_credit / multiplier)
            breakeven_lower = strike_body - (net_credit / multiplier)
            strategy = {
                '策略类型': '铁蝶（Iron Butterfly）',
                '策略描述': '卖出ATM跨式+买入两翼保护，收取权利金，风险有限',
                '适用场景': f'预期窄幅震荡且IV偏高；IV/RV={vol_features.get("IV_RV", "")}；期限结构={vol_features.get("Term Structure", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金收入': net_credit,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '盈利区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '短跨被保护为有限风险，但临近到期Gamma显著放大',
                    '更适合IV高位回落与标的横盘的阶段',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(sell_call.get("Delta", 0) or 0) + float(sell_put.get("Delta", 0) or 0) - float(buy_call.get("Delta", 0) or 0) - float(buy_put.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(sell_call.get("Gamma", 0) or 0) + float(sell_put.get("Gamma", 0) or 0) - float(buy_call.get("Gamma", 0) or 0) - float(buy_put.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(sell_call.get("Theta", 0) or 0) + float(sell_put.get("Theta", 0) or 0) - float(buy_call.get("Theta", 0) or 0) - float(buy_put.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(sell_call.get("Vega", 0) or 0) + float(sell_put.get("Vega", 0) or 0) - float(buy_call.get("Vega", 0) or 0) - float(buy_put.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建铁蝶策略失败: {e}")
            return None

    def _build_jade_lizard_strategy(self, sell_put, sell_call, buy_call, market_state, vol_features):
        try:
            multiplier = float(sell_put.get('合约乘数', 1) or 1)
            legs = [
                self._make_option_leg(sell_put, '卖出', 1),
                self._make_option_leg(sell_call, '卖出', 1),
                self._make_option_leg(buy_call, '买入', 1)
            ]
            net_credit = float(sum([x.get('权利金', 0) or 0 for x in legs]))
            strike_put = float(sell_put.get('行权价', 0) or 0)
            strike_call_sell = float(sell_call.get('行权价', 0) or 0)
            strike_call_buy = float(buy_call.get('行权价', 0) or 0)
            max_profit = net_credit
            max_loss = '下行较大亏损（无PUT保护）'
            breakeven_lower = strike_put - (net_credit / multiplier)
            breakeven_upper = strike_call_sell + (net_credit / multiplier)
            strategy = {
                '策略类型': '玉蜥蜴（Jade Lizard）',
                '策略描述': '卖出价外PUT + 卖出价外CALL并买入更价外CALL保护，偏多中性卖方策略',
                '适用场景': f'温和看涨或震荡；IV偏高；偏斜结构PUT更贵；IV/RV={vol_features.get("IV_RV", "")}',
                '操作建议': legs,
                '成本分析': {
                    '净权利金收入': net_credit,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss
                },
                '盈亏平衡点': {
                    '下限': breakeven_lower,
                    '上限': breakeven_upper,
                    '盈利区间': f'{breakeven_lower:.2f} - {breakeven_upper:.2f}'
                },
                '风险提示': [
                    '上行风险被CALL价差限制，但下行风险较大',
                    '适合对标的下行有信心或具备现货/期货对冲能力',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(sell_put.get("Delta", 0) or 0) + float(sell_call.get("Delta", 0) or 0) - float(buy_call.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(sell_put.get("Gamma", 0) or 0) + float(sell_call.get("Gamma", 0) or 0) - float(buy_call.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(sell_put.get("Theta", 0) or 0) + float(sell_call.get("Theta", 0) or 0) - float(buy_call.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(sell_put.get("Vega", 0) or 0) + float(sell_call.get("Vega", 0) or 0) - float(buy_call.get("Vega", 0) or 0)):.4f}'
                }
            }
            return strategy
        except Exception as e:
            print(f"构建玉蜥蜴策略失败: {e}")
            return None
    
    def _pick_nearest_option(self, options_df, option_type: str, target_strike: float):
        try:
            df = options_df[options_df['期权类型'] == option_type].copy()
            if df.empty:
                return None
            strike_series = pd.Series(pd.to_numeric(df['行权价'], errors='coerce'), index=df.index)
            strikes = strike_series.to_numpy(dtype=float)
            df['__diff'] = np.abs(strikes - float(target_strike))
            df = df.dropna(subset=['__diff']).sort_values('__diff')
            if df.empty:
                return None
            return df.iloc[0]
        except Exception:
            return None

    def _pick_option_by_delta(self, options_df, option_type: str, target_delta: float):
        try:
            if options_df is None or options_df.empty:
                return None
            df = options_df[options_df['期权类型'] == option_type].copy()
            if df.empty or 'Delta' not in df.columns:
                return None
            df['__delta'] = pd.Series(pd.to_numeric(df['Delta'], errors='coerce'))
            df = df.dropna(subset=['__delta'])
            if df.empty:
                return None
            df['__diff'] = (df['__delta'] - float(target_delta)).abs()
            df = df.dropna(subset=['__diff']).sort_values('__diff')
            if df.empty:
                return None
            return df.iloc[0]
        except Exception:
            return None

    def _safe_float(self, v):
        try:
            if v is None:
                return None
            if isinstance(v, str):
                v = v.replace('%', '').strip()
            f = float(v)
            if np.isnan(f):
                return None
            return f
        except Exception:
            return None

    def _score_strategy(self, strategy):
        try:
            if strategy is None:
                return 0
            existing = self._safe_float(strategy.get('策略评分', None))
            if existing is not None:
                return round(existing, 2)
            base = 40.0
            cost = strategy.get('成本分析', {}) or {}
            ratio = self._safe_float(cost.get('盈亏比', None))
            if ratio is not None:
                base += min(ratio, 3.0) / 3.0 * 35.0
            net_cashflow = self._safe_float(cost.get('净权利金现金流(>0收入,<0支出)', None))
            if net_cashflow is None:
                net_credit = self._safe_float(cost.get('净权利金收入', None))
                net_debit = self._safe_float(cost.get('净权利金支出', None))
                if net_credit is not None:
                    net_cashflow = net_credit
                elif net_debit is not None:
                    net_cashflow = -net_debit
            if net_cashflow is not None and net_cashflow > 0:
                base += 10
            max_loss = cost.get('最大亏损', None)
            if isinstance(max_loss, str) and '无限' in max_loss:
                base -= 15
            if isinstance(max_loss, str) and ('较大' in max_loss or '风险' in max_loss):
                base -= 8
            if '末日期权' in str(strategy.get('策略类型', '')):
                base -= 8
            return max(0, min(100, round(base, 2)))
        except Exception:
            return 0

    def _format_strategy_focus(self, strategy):
        try:
            legs = strategy.get('操作建议', []) if strategy else []
            if not legs:
                return ''
            items = []
            for leg in legs:
                action = leg.get('操作', '')
                code = leg.get('合约代码', '')
                qty = leg.get('数量', '')
                px = leg.get('建议成交价', None)
                if px is None or px == '':
                    px = leg.get('参考价', '')
                text = f'{action}{code}*{qty}'
                if px != '' and px is not None:
                    text += f'@{px}'
                items.append(text)
            return '；'.join(items)
        except Exception:
            return ''

    def _format_strategy_points(self, strategy):
        try:
            points = strategy.get('策略要点', None)
            if not points:
                return ''
            if isinstance(points, dict):
                return '；'.join([f'{k}:{v}' for k, v in points.items()])
            return str(points)
        except Exception:
            return ''

    def _normalize_scene(self, scene):
        try:
            if scene is None:
                return ''
            text = str(scene).strip()
            if not text:
                return ''
            parts = [p.strip() for p in re.split('[；;]', text) if p.strip()]
            cleaned_parts = []
            for part in parts:
                subparts = [sp.strip() for sp in re.split('[，,]', part) if sp.strip()]
                cleaned_sub = []
                for sp in subparts:
                    if re.search(r'IV/RV\s*=\s*$', sp):
                        continue
                    if re.search(r'期限结构\s*=\s*$', sp):
                        continue
                    if re.fullmatch(r'波动率', sp):
                        continue
                    cleaned_sub.append(sp)
                if cleaned_sub:
                    cleaned_parts.append('，'.join(cleaned_sub))
            return '；'.join(cleaned_parts)
        except Exception:
            return str(scene) if scene is not None else ''

    def _judge_market_condition(self, market_state):
        """根据市场状态判断市场环境
        
        Returns:
            dict: {
                '描述': '市场环境描述',
                '方向': 'bullish/bearish/neutral',
                '波动率': 'high/low/normal',
                '时间': 'short/mid/long',
                '期限结构': 'normal/inverted/flat'
            }
        """
        try:
            judgment = {}
            
            # 1. 方向判断
            direction = market_state.get('期货方向', '').strip()
            pc_ratio = self._safe_float(market_state.get('PCR', None))
            
            if '大涨' in direction or '强势上涨' in direction:
                judgment['方向'] = 'strong_bullish'
                judgment['方向描述'] = '强烈看涨'
            elif '上涨' in direction or '偏多' in direction:
                judgment['方向'] = 'bullish'
                judgment['方向描述'] = '温和看涨'
            elif '大跌' in direction or '强势下跌' in direction:
                judgment['方向'] = 'strong_bearish'
                judgment['方向描述'] = '强烈看跌'
            elif '下跌' in direction or '偏空' in direction:
                judgment['方向'] = 'bearish'
                judgment['方向描述'] = '温和看跌'
            elif '震荡' in direction or '横盘' in direction:
                judgment['方向'] = 'neutral'
                judgment['方向描述'] = '震荡中性'
            else:
                judgment['方向'] = 'neutral'
                judgment['方向描述'] = '方向不明'
            
            # 2. 波动率判断
            iv_atm = self._safe_float(market_state.get('隐含波动率(ATM)', None))
            rv_atm = self._safe_float(market_state.get('近期波动率(ATM)', None))
            iv_rv = self._safe_float(market_state.get('IV/RV(ATM)', None))
            
            if iv_rv and iv_rv > 1.3:
                judgment['波动率'] = 'high'
                judgment['波动率描述'] = 'IV高估（IV/RV>1.3）'
            elif iv_rv and iv_rv < 0.8:
                judgment['波动率'] = 'low'
                judgment['波动率描述'] = 'IV低估（IV/RV<0.8）'
            else:
                judgment['波动率'] = 'normal'
                judgment['波动率描述'] = 'IV正常'
            
            # 3. 时间判断（剩余天数）
            days = self._safe_float(market_state.get('剩余天数', None))
            if days and days < 7:
                judgment['时间'] = 'expiry'
                judgment['时间描述'] = '末日期权（<7天）'
            elif days and days < 30:
                judgment['时间'] = 'short'
                judgment['时间描述'] = '短期（<30天）'
            elif days and days < 60:
                judgment['时间'] = 'mid'
                judgment['时间描述'] = '中期（30-60天）'
            else:
                judgment['时间'] = 'long'
                judgment['时间描述'] = '长期（>60天）'
            
            # 4. 期限结构判断
            term_structure = market_state.get('曲面-期限结构', '').strip()
            if 'Contango' in term_structure or '远月高' in term_structure:
                judgment['期限结构'] = 'contango'
                judgment['期限结构描述'] = '远月IV高（Contango）'
            elif 'Backwardation' in term_structure or '近月高' in term_structure:
                judgment['期限结构'] = 'backwardation'
                judgment['期限结构描述'] = '近月IV高（Backwardation）'
            else:
                judgment['期限结构'] = 'flat'
                judgment['期限结构描述'] = '期限结构平坦'
            
            # 5. Skew判断
            skew = market_state.get('曲面-倾斜方向', '').strip()
            if 'Put Skew' in skew or 'Put偏斜' in skew:
                judgment['Skew'] = 'put_skew'
                judgment['Skew描述'] = 'Put端溢价（看跌情绪）'
            elif 'Call Skew' in skew or 'Call偏斜' in skew:
                judgment['Skew'] = 'call_skew'
                judgment['Skew描述'] = 'Call端溢价（看涨情绪）'
            else:
                judgment['Skew'] = 'symmetric'
                judgment['Skew描述'] = 'Skew对称'
            
            # 综合描述
            judgment['描述'] = f"{judgment.get('方向描述', '')}；{judgment.get('波动率描述', '')}；{judgment.get('时间描述', '')}"
            
            return judgment
            
        except Exception as e:
            print(f"市场判断失败: {e}")
            return {'描述': '无法判断', '方向': 'unknown', '波动率': 'unknown', '时间': 'unknown'}

    def _recommend_strategies_by_market(self, market_judgment, market_state):
        """根据市场判断推荐策略
        
        Returns:
            dict: {
                '类型': ['策略类型1', '策略类型2'],
                '核心': ['核心策略1', '核心策略2'],
                '风险': '风险等级',
                '时机': '开仓时机描述'
            }
        """
        try:
            recommended = {
                '类型': [],
                '核心': [],
                '风险': '',
                '时机': ''
            }
            
            direction = market_judgment.get('方向', 'neutral')
            volatility = market_judgment.get('波动率', 'normal')
            time_frame = market_judgment.get('时间', 'mid')
            term_structure = market_judgment.get('期限结构', 'flat')
            
            # ========== 一、单一维度策略（纯方向/纯波动率） ==========
            
            # 1. 强烈看涨 -> 买入看涨期权
            if direction == 'strong_bullish':
                if volatility == 'low':
                    recommended['类型'].append('买入看涨期权')
                    recommended['核心'].append('买入ATM/轻虚值Call')
                    recommended['风险'] = '中'
                    recommended['时机'] = '技术面突破+IV低位'
                
                # 备兑开仓（如果已有多头）
                if volatility == 'high':
                    recommended['类型'].append('备兑开仓')
                    recommended['核心'].append('持有期货+卖出Call')
                    recommended['风险'] = '低'
            
            # 2. 强烈看跌 -> 买入看跌期权
            elif direction == 'strong_bearish':
                if volatility == 'low':
                    recommended['类型'].append('买入看跌期权')
                    recommended['核心'].append('买入ATM/轻虚值Put')
                    recommended['风险'] = '中'
                    recommended['时机'] = '跌破关键支撑+IV低位'
            
            # 3. 方向不定 -> 买入跨式/宽跨式
            elif direction == 'neutral':
                if volatility == 'low' and time_frame in ['short', 'mid']:
                    recommended['类型'].append('买入跨式/宽跨式')
                    recommended['核心'].append('买入ATM Call+Put')
                    recommended['风险'] = '中'
                    recommended['时机'] = '重大事件前+IV历史低位'
            
            # ========== 二、双维度策略（方向+时间/波动率） ==========
            
            # 4. 温和看涨 -> 牛市价差
            if direction == 'bullish':
                recommended['类型'].append('牛市价差')
                recommended['核心'].append('买低K Call+卖高K Call')
                recommended['风险'] = '中'
                recommended['时机'] = '震荡上行+有明确阻力'
            
            # 5. 温和看跌 -> 熊市价差
            elif direction == 'bearish':
                recommended['类型'].append('熊市价差')
                recommended['核心'].append('买高K Put+卖低K Put')
                recommended['风险'] = '中'
                recommended['时机'] = '震荡下行+有明确支撑'
            
            # 6. 保护性策略
            if direction in ['bearish', 'strong_bearish']:
                recommended['类型'].append('保护性看跌/领口')
                recommended['核心'].append('持有标的+买Put保护')
                recommended['风险'] = '低-中'
            
            # ========== 三、三维策略（中性+时间+波动率） ==========
            
            # 7. 震荡市+高IV -> 铁鹰式/铁蝶式
            if direction == 'neutral' and volatility == 'high':
                recommended['类型'].append('铁鹰式/铁蝶式')
                recommended['核心'].append('卖出宽跨+买保护')
                recommended['风险'] = '中'
                recommended['时机'] = '清晰震荡区间+IV高位'
            
            # 8. 期限结构陡峭 -> 日历价差
            if term_structure == 'contango':
                recommended['类型'].append('日历价差/对角日历')
                recommended['核心'].append('卖近月买远月')
                recommended['风险'] = '中'
                recommended['时机'] = 'IV期限结构陡峭'
            
            # 9. 比率价差
            if direction == 'bullish' and volatility == 'high':
                recommended['类型'].append('比率价差')
                recommended['核心'].append('买1手Call+卖2手虚值Call')
                recommended['风险'] = '高'
                recommended['时机'] = 'IV高+预期缓涨至目标价'
            
            # ========== 四、特殊策略 ==========
            
            # 10. 末日期权单边
            if time_frame == 'expiry' and direction in ['strong_bullish', 'strong_bearish']:
                recommended['类型'].append('末日期权单边')
                recommended['核心'].append('买入短期单边期权')
                recommended['风险'] = '极高'
                recommended['时机'] = '重大事件前+高Gamma'
            
            # 如果没有推荐，给出默认
            if not recommended['类型']:
                recommended['类型'] = ['观望']
                recommended['核心'] = ['等待更好时机']
                recommended['风险'] = '无'
                recommended['时机'] = '市场信号不明确'
            
            return recommended
            
        except Exception as e:
            print(f"策略推荐失败: {e}")
            return {
                '类型': ['无推荐'],
                '核心': ['系统错误'],
                '风险': '未知',
                '时机': '无法判断'
            }

    def _create_strategy_guide_sheet(self, wb):
        """创建策略指南sheet（使用内置策略数据）"""
        ws = wb.create_sheet('策略指南')
        
        # 定义所有策略（内置 + 补充）
        all_guides = []
        
        # 1. 从内置策略指南加载
        for strategy_name, strategy_data in self.strategy_guide_dict.items():
            guide_item = {
                '策略分类': strategy_data.get('策略分类', ''),
                '策略名称': strategy_data.get('策略名称', strategy_name),
                '市场观点': strategy_data.get('市场三维观点（方向+波动率+时间）', ''),
                '核心构建': strategy_data.get('核心构建（期货期权专属）', ''),
                '盈亏特征': strategy_data.get('盈亏特征+核心希腊字母暴露', ''),
                '开仓时机': strategy_data.get('期货实战开仓时机', ''),
                '损益驱动': strategy_data.get('核心损益驱动源', ''),
                '实操要点': strategy_data.get('补充策略要点（实操必看）', ''),
            }
            all_guides.append(guide_item)
        
        # 2. 补充代码中特有的策略（不在内置指南中的）
        additional_strategies = [
            {'策略分类': '时间价差策略', '策略名称': '日历跨式', '市场观点': '期限结构与波动率回归', 
             '核心构建': '卖近月CALL+PUT，买远月CALL+PUT', '盈亏特征': 'Vega敏感，Delta中性', 
             '开仓时机': 'IV期限结构陡峭', '损益驱动': '近远月IV差异收敛', '实操要点': '需监控IV期限结构变化'},
            {'策略分类': '复杂价差策略', '策略名称': '玉蜥蜴', '市场观点': '偏多卖方结构', 
             '核心构建': '卖PUT+卖CALL买保护CALL', '盈亏特征': '收权利金，上行保护，下行敞口', 
             '开仓时机': '看涨但IV高时', '损益驱动': 'Theta衰减+Vega下降', '实操要点': '必须有强烈看涨信念'},
            {'策略分类': '复杂价差策略', '策略名称': '海鸥', '市场观点': '偏斜交易', 
             '核心构建': '买主腿+卖贵翼+买便宜翼', '盈亏特征': '非对称收益，降低成本', 
             '开仓时机': 'Skew明显时', '损益驱动': 'Delta方向+偏斜差异', '实操要点': '密切监控Skew变化'},
            {'策略分类': '合成策略', '策略名称': '合成多头/空头', '市场观点': '纯方向判断', 
             '核心构建': 'CALL/PUT组合复制标的', '盈亏特征': 'Delta=±1，完全复制标的', 
             '开仓时机': '期货流动性差或保证金高时', '损益驱动': 'Delta方向变动', '实操要点': '监控合成偏差'},
            {'策略分类': '信用价差', '策略名称': '信用价差', '市场观点': '时间价值衰减', 
             '核心构建': '卖近档+买远档保护', '盈亏特征': '收权利金，风险有限', 
             '开仓时机': 'IV偏高时', '损益驱动': 'Theta时间价值衰减', '实操要点': '严格止损，防止价格突破'},
            {'策略分类': '末日期权', '策略名称': '末日期权单边', '市场观点': '事件驱动与高Gamma', 
             '核心构建': '买入短期单边期权', '盈亏特征': '高Gamma，高风险高收益', 
             '开仓时机': '重大事件前', '损益驱动': 'Gamma快速放大', '实操要点': '控制仓位，及时止盈'},
            {'策略分类': '波动率交易', '策略名称': '风险反转', '市场观点': '偏斜定价与方向敞口', 
             '核心构建': '买CALL卖PUT或买PUT卖CALL', '盈亏特征': '方向性暴露，Skew交易', 
             '开仓时机': 'Skew异常时', '损益驱动': 'Delta方向+Skew收敛', '实操要点': '监控Skew变化'},
        ]
        
        # 去重（优先内置）
        builtin_names = {g['策略名称'] for g in all_guides if g['策略名称']}
        for strat in additional_strategies:
            if strat['策略名称'] not in builtin_names:
                all_guides.append(strat)
        
        # 写入Excel
        if all_guides:
            headers = ['策略分类', '策略名称', '市场观点', '核心构建', '盈亏特征', '开仓时机', '损益驱动', '实操要点']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            row = 2
            for item in all_guides:
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 35
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 35
            
            ws.freeze_panes = 'A2'



    def _create_strategy_details_sheet(self, wb, all_strategies):
        ws = wb.create_sheet('策略明细')
        headers = ['标的合约', '到期月份', '策略类型', '策略评分', '操作要点', '盈亏比', '净权利金', '最大亏损', '最大盈利', '适用场景']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row = 2
        for underlying_data in all_strategies:
            underlying = underlying_data.get('underlying', '')
            expiry = str(underlying_data.get('expiry', ''))
            for strategy in underlying_data.get('strategies', []) or []:
                if not strategy:
                    continue
                cost = strategy.get('成本分析', {}) or {}
                net_cashflow = cost.get('净权利金现金流(>0收入,<0支出)', None)
                if net_cashflow is None:
                    net_credit = cost.get('净权利金收入', None)
                    net_debit = cost.get('净权利金支出', None)
                    if net_credit is not None:
                        net_cashflow = net_credit
                    elif net_debit is not None:
                        net_cashflow = -net_debit
                focus = self._format_strategy_focus(strategy)
                points = self._format_strategy_points(strategy)
                op_focus = focus if focus else points
                if focus and points:
                    op_focus = f'{focus}｜{points}'
                row_values = [
                    underlying,
                    expiry,
                    strategy.get('策略类型', ''),
                    strategy.get('策略评分', ''),
                    op_focus,
                    cost.get('盈亏比', ''),
                    net_cashflow if net_cashflow is not None else '',
                    cost.get('最大亏损', ''),
                    cost.get('最大盈利', ''),
                    self._normalize_scene(strategy.get('适用场景', ''))
                ]
                for col_idx, value in enumerate(row_values, 1):
                    ws.cell(row=row, column=col_idx, value=value)
                row += 1
        
        self._auto_adjust_column_width(ws)
        ws.freeze_panes = 'B2'

    def _get_underlying_option_universe(self, options_df, market_state):
        try:
            df = options_df.copy() if options_df is not None else pd.DataFrame()
            underlying = market_state.get('标的合约', None) if market_state else None
            if not underlying and df is not None and not df.empty and '标的合约' in df.columns:
                underlying = df['标的合约'].dropna().iloc[0] if not df['标的合约'].dropna().empty else None
            if self.option_ref_df is not None and not self.option_ref_df.empty and underlying:
                m = self.option_ref_df['标的合约'] == underlying
                if m.any():
                    df2 = self.option_ref_df[m].copy()
                    if not df2.empty:
                        df = df2
            for c in ['成交量', '持仓量', '剩余天数', '行权价', '合约乘数', 'Delta', 'Gamma', 'Theta', 'Vega', '期权价']:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
            if '期权类型' in df.columns:
                s = pd.Series(df['期权类型'], index=df.index, dtype='string')
                df['期权类型'] = s.str.upper()
            if '剩余天数' not in df.columns and '到期日' in df.columns:
                exp = pd.Series(pd.to_datetime(df['到期日'], errors='coerce'), index=df.index)
                today = pd.Timestamp.today().normalize()
                df['剩余天数'] = (exp - today) / np.timedelta64(1, 'D')
            if '剩余天数' in df.columns:
                df['剩余天数'] = pd.to_numeric(df['剩余天数'], errors='coerce')
            return df
        except Exception:
            return options_df.copy() if options_df is not None else pd.DataFrame()

    def _generate_synthetic_positions(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            df = options_df.copy()
            for c in ['成交量', '持仓量', '剩余天数', '行权价', '合约乘数']:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
            if '期权类型' in df.columns:
                df['期权类型'] = df['期权类型'].astype(str).str.upper()

            liquid = df[
                (df['成交量'] >= 50) &
                (df['持仓量'] >= 100) &
                (df['剩余天数'] >= 10) &
                (df['剩余天数'] <= 120)
            ].copy()
            if liquid.empty:
                liquid = df[
                    (df['成交量'] >= 20) &
                    (df['持仓量'] >= 50) &
                    (df['剩余天数'] >= 7) &
                    (df['剩余天数'] <= 180)
                ].copy()
            if liquid.empty:
                return strategies

            atm_price = market_state.get('标的现价', 0) or float(liquid['标的现价'].dropna().iloc[0] if '标的现价' in liquid.columns and not liquid['标的现价'].dropna().empty else 0)
            if atm_price <= 0:
                return strategies

            calls = liquid[liquid['期权类型'] == 'CALL'].dropna(subset=['行权价']).copy()
            puts = liquid[liquid['期权类型'] == 'PUT'].dropna(subset=['行权价']).copy()
            if calls.empty or puts.empty:
                return strategies

            strike_calls = set([float(x) for x in calls['行权价'].dropna().unique().tolist()])
            strike_puts = set([float(x) for x in puts['行权价'].dropna().unique().tolist()])
            common = list(strike_calls.intersection(strike_puts))
            if not common:
                return strategies
            common = sorted(common, key=lambda k: abs(k - float(atm_price)))
            k = common[0]
            call = calls[np.abs(calls['行权价'] - k) < 1e-9].iloc[0]
            put = puts[np.abs(puts['行权价'] - k) < 1e-9].iloc[0]

            multiplier = float(call.get('合约乘数', 1) or 1)

            legs_long = [self._make_option_leg(call, '买入', 1), self._make_option_leg(put, '卖出', 1)]
            net_cashflow_long = float(sum([x.get('权利金', 0) or 0 for x in legs_long]))
            premium_per_unit_long = (-net_cashflow_long) / multiplier
            breakeven_long = float(k) + premium_per_unit_long
            strategies.append({
                '策略类型': '合成多头（Synthetic Long）',
                '策略描述': '买入CALL + 卖出PUT（同一行权价），等价于做多标的/期货敞口',
                '适用场景': f'需要方向多头但希望用期权结构替代；IV/RV={vol_features.get("IV_RV", "")}',
                '策略标签': '合成/方向',
                '操作建议': legs_long,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow_long,
                    '每单位权利金成本': premium_per_unit_long,
                    '最大盈利': '理论无限（上行）',
                    '最大亏损': f'{k:.2f}附近下行大幅亏损（PUT卖方）'
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven_long
                },
                '风险提示': [
                    '本质包含卖出PUT，存在尾部下行风险与保证金占用',
                    '临近到期可能发生行权/指派，需要管理头寸',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(float(call.get("Delta", 0) or 0) - float(put.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(float(call.get("Gamma", 0) or 0) - float(put.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(float(call.get("Theta", 0) or 0) - float(put.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(float(call.get("Vega", 0) or 0) - float(put.get("Vega", 0) or 0)):.4f}'
                }
            })

            legs_short = [self._make_option_leg(call, '卖出', 1), self._make_option_leg(put, '买入', 1)]
            net_cashflow_short = float(sum([x.get('权利金', 0) or 0 for x in legs_short]))
            premium_per_unit_short = (-net_cashflow_short) / multiplier
            breakeven_short = float(k) - premium_per_unit_short
            strategies.append({
                '策略类型': '合成空头（Synthetic Short）',
                '策略描述': '卖出CALL + 买入PUT（同一行权价），等价于做空标的/期货敞口',
                '适用场景': f'需要方向空头但希望用期权结构替代；IV/RV={vol_features.get("IV_RV", "")}',
                '策略标签': '合成/方向',
                '操作建议': legs_short,
                '成本分析': {
                    '净权利金现金流(>0收入,<0支出)': net_cashflow_short,
                    '每单位权利金成本': premium_per_unit_short,
                    '最大盈利': f'{k:.2f}附近下行收益接近线性（理论到0）',
                    '最大亏损': '理论无限（上行，CALL卖方）'
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven_short
                },
                '风险提示': [
                    '本质包含卖出CALL，存在尾部上行风险与保证金占用',
                    '临近到期可能发生行权/指派，需要管理头寸',
                    f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                ],
                '希腊字母风险': {
                    'Delta': f'{(-float(call.get("Delta", 0) or 0) + float(put.get("Delta", 0) or 0)):.4f}',
                    'Gamma': f'{(-float(call.get("Gamma", 0) or 0) + float(put.get("Gamma", 0) or 0)):.6f}',
                    'Theta': f'{(-float(call.get("Theta", 0) or 0) + float(put.get("Theta", 0) or 0)):.4f}',
                    'Vega': f'{(-float(call.get("Vega", 0) or 0) + float(put.get("Vega", 0) or 0)):.4f}'
                }
            })

            return [x for x in strategies if x]
        except Exception:
            return []

    def _score_butterfly_strategy(self, strategy, atm_price: float):
        try:
            max_profit = strategy.get('成本分析', {}).get('最大盈利', None)
            max_loss = strategy.get('成本分析', {}).get('最大亏损', None)
            max_profit = self._safe_float(max_profit)
            max_loss = self._safe_float(max_loss)
            width = None
            breakeven = strategy.get('盈亏平衡点', {})
            lower = self._safe_float(breakeven.get('下限', None))
            upper = self._safe_float(breakeven.get('上限', None))
            if lower is not None and upper is not None:
                width = upper - lower
            score = 0.0
            if max_profit is not None and max_loss is not None and max_loss > 0:
                ratio = max_profit / max_loss
                score += min(ratio, 3.0) / 3.0 * 60.0
            if width is not None and atm_price > 0:
                score += min(width / (atm_price * 0.06), 1.0) * 40.0
            return round(score, 2)
        except Exception:
            return 0.0
    
    def _generate_vol_surface_strategies(self, options_df, market_state, vol_features):
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            
            S = float(market_state.get('标的现价', 0) or 0)
            if S <= 0:
                return strategies
            
            skew_label = str(vol_features.get('Skew', '') or '')
            direction = str(market_state.get('期货方向', '') or '')
            
            bias = 'Bullish'
            if '下跌' in direction:
                bias = 'Bearish'
            
            if ('Put' in skew_label) or ('负' in skew_label) or ('左' in skew_label) or ('smirk' in skew_label.lower()):
                if bias == 'Bullish':
                    k_put_sell = 0.95 * S
                    k_put_buy = 0.90 * S
                    k_call_buy = 1.05 * S
                    put_sell = self._pick_nearest_option(options_df, 'PUT', k_put_sell)
                    put_buy = self._pick_nearest_option(options_df, 'PUT', k_put_buy)
                    call_buy = self._pick_nearest_option(options_df, 'CALL', k_call_buy)
                    if put_sell is not None and put_buy is not None and call_buy is not None:
                        strategies.append(self._build_seagull_strategy(call_buy, put_sell, put_buy, market_state, vol_features))
                else:
                    k_call_sell = 1.05 * S
                    k_call_buy = 1.10 * S
                    k_put_buy = 0.95 * S
                    call_sell = self._pick_nearest_option(options_df, 'CALL', k_call_sell)
                    call_buy = self._pick_nearest_option(options_df, 'CALL', k_call_buy)
                    put_buy = self._pick_nearest_option(options_df, 'PUT', k_put_buy)
                    if call_sell is not None and call_buy is not None and put_buy is not None:
                        strategies.append(self._build_seagull_strategy(put_buy, call_sell, call_buy, market_state, vol_features))
            
            return [x for x in strategies if x]
        except Exception as e:
            print(f"构建波动率曲面策略失败: {e}")
            return []
    
    def _generate_gamma_scalping_strategies(self, options_df, market_state, vol_features):
        """生成末日Gamma Scalping策略
        
        适用于到期前1-3天，利用高Gamma期权在标的震荡中收割收益
        """
        strategies = []
        try:
            if options_df is None or options_df.empty:
                return strategies
            
            # 筛选末日期权（1-3天到期）
            expiry_df = options_df[
                (options_df['剩余天数'] >= 1) &
                (options_df['剩余天数'] <= 3)
            ].copy()
            
            if expiry_df.empty:
                return strategies
            
            S = float(market_state.get('标的现价', 0) or 0)
            if S <= 0:
                return strategies
            
            # 筛选ATM附近高Gamma期权
            if 'Gamma' not in expiry_df.columns:
                return strategies
            
            expiry_df['虚实比例'] = abs(expiry_df['行权价'] - S) / S * 100
            atm_options = expiry_df[expiry_df['虚实比例'] <= 2].copy()  # 虚实幅度2%以内
            
            if atm_options.empty:
                atm_options = expiry_df.nsmallest(4, '虚实比例')
            
            # 选择Gamma最高的CALL和PUT组成跨式
            calls = atm_options[atm_options['期权类型'] == 'CALL']
            puts = atm_options[atm_options['期权类型'] == 'PUT']
            
            if calls.empty or puts.empty:
                return strategies
            
            # 选最接近ATM且Gamma高的
            best_call = None
            best_put = None
            
            for _, row in calls.iterrows():
                if best_call is None or row.get('Gamma', 0) > best_call.get('Gamma', 0):
                    best_call = row
            
            for _, row in puts.iterrows():
                if best_put is None or row.get('Gamma', 0) > best_put.get('Gamma', 0):
                    best_put = row
            
            if best_call is None or best_put is None:
                return strategies
            
            # 构建Gamma Scalping策略
            strategy = self._build_gamma_scalping_strategy(best_call, best_put, market_state, vol_features)
            if strategy:
                strategies.append(strategy)
            
            return strategies
        except Exception as e:
            print(f"构建Gamma Scalping策略失败: {e}")
            return []
    
    def _build_gamma_scalping_strategy(self, call_opt, put_opt, market_state, vol_features):
        """构建Gamma Scalping策略详情"""
        try:
            underlying = market_state.get('标的合约', '')
            expiry = market_state.get('交割年月', '')
            S = float(market_state.get('标的现价', 0) or 0)
            multiplier = float(call_opt.get('合约乘数', 1) or 1)
            
            call_leg = self._make_option_leg(call_opt, '买入', 1)
            put_leg = self._make_option_leg(put_opt, '买入', 1)
            
            call_px = call_leg['建议成交价']
            put_px = put_leg['建议成交价']
            total_premium = (call_px + put_px) * multiplier
            
            call_gamma = float(call_opt.get('Gamma', 0) or 0)
            put_gamma = float(put_opt.get('Gamma', 0) or 0)
            total_gamma = call_gamma + put_gamma
            
            call_delta = float(call_opt.get('Delta', 0.5) or 0.5)
            put_delta = float(put_opt.get('Delta', -0.5) or -0.5)
            net_delta = call_delta + put_delta
            
            days = int(call_opt.get('剩余天数', 1) or 1)
            
            # 估算Gamma收益：假设每天波动1%，Gamma*波动^2/2
            daily_move_pct = 0.015  # 假设1.5%日波动
            daily_move = S * daily_move_pct
            gamma_pnl_per_move = 0.5 * total_gamma * (daily_move ** 2) * multiplier
            
            # 最大亏损 = 权利金
            max_loss = total_premium
            # 盈亏平衡点
            breakeven_up = call_opt['行权价'] + (call_px + put_px)
            breakeven_down = put_opt['行权价'] - (call_px + put_px)
            
            strategy = {
                '策略类型': 'Gamma Scalping',
                '策略描述': f'在标的合约{underlying}到期前利用高Gamma期权进行动态对冲收割波动收益',
                '策略名称': f'末日Gamma收割 {underlying}',
                '标的合约': underlying,
                '交割年月': expiry,
                '到期日': str(call_opt.get('到期日', '')),
                '剩余天数': days,
                '适用场景': '末日期权震荡收割',
                '市场预期': '预期到期前标的价格剧烈震荡',
                '方向': '中性',
                '操作建议': [call_leg, put_leg],
                '成本分析': {
                    '总权利金支出': round(total_premium, 2),
                    '最大亏损': round(max_loss, 2),
                    '组合Gamma': round(total_gamma, 6),
                    '组合Delta': round(net_delta, 4),
                    '盈亏平衡点(上)': round(breakeven_up, 2),
                    '盈亏平衡点(下)': round(breakeven_down, 2),
                    '预估单次1.5%波动Gamma收益': round(gamma_pnl_per_move, 2),
                },
                '执行计划': {
                    '建仓': f"买入{call_opt['合约代码']}+{put_opt['合约代码']}组成跨式",
                    '对冲': f'每当Delta偏离±0.3时，用期货对冲回Delta中性',
                    '平仓': '到期前2小时内平仓或行权',
                },
                '盈亏平衡点': {
                    '上限': round(breakeven_up, 2),
                    '下限': round(breakeven_down, 2)
                },
                '希腊字母风险': {
                    'Delta': round(net_delta, 4),
                    'Gamma': round(total_gamma, 6),
                    'Vega': '高',
                    'Theta': '极高'
                },
                '风险提示': ['末日期权时间价值衰减极快', '若标的不波动将快速亏损', '对冲标的流动性需充足'],
                '风险等级': '高',
                '策略评分': 65.0 if total_gamma > 0.001 else 50.0,
            }
            return strategy
        except Exception as e:
            print(f"构建Gamma Scalping策略失败: {e}")
            return None
    
    def _generate_vol_cone_strategies(self, options_df, market_state, vol_features):
        """生成波动率锥区间交易策略
        
        当IV处于历史极端分位时，交易波动率回归
        """
        strategies = []
        try:
            iv_atm = market_state.get('隐含波动率(ATM)', None)
            hv20 = market_state.get('HV20(ATM)', None)
            hv60 = market_state.get('HV60(ATM)', None)
            
            if iv_atm is None or hv20 is None:
                return strategies
            
            iv_atm = float(iv_atm)
            hv20 = float(hv20)
            hv60 = float(hv60) if hv60 else hv20
            
            # 计算IV相对HV的比率 - 双轨制
            # 1. 判断高估（卖出）：IV vs max(HV20, HV60) -> 确保IV真的很高
            # 2. 判断低估（买入）：IV vs min(HV20, HV60) -> 确保IV真的很低
            ref_hv_high = max(hv20, hv60)
            ref_hv_low = min(hv20, hv60)
            
            iv_hv_ratio_high = iv_atm / ref_hv_high if ref_hv_high > 0 else 1.0 # 用于判断是否高估
            iv_hv_ratio_low = iv_atm / ref_hv_low if ref_hv_low > 0 else 1.0   # 用于判断是否低估
            
            # 波动率锥判断
            underlying = market_state.get('标的合约', '')
            expiry = market_state.get('交割年月', '')
            
            strategy = None
            
            if iv_hv_ratio_high > 1.3:  # IV显著高于高位HV，卖出波动率
                strategy = {
                    '策略类型': '波动率锥-卖出',
                    '策略描述': f'当前IV相对于HV(Max)溢价较高({(iv_hv_ratio_high-1)*100:.0f}%)，建议卖出高估的IV',
                    '策略名称': f'IV高估卖波动 {underlying}',
                    '标的合约': underlying,
                    '交割年月': expiry,
                    '适用场景': '波动率均值回归',
                    '市场预期': f'IV({iv_atm:.1f}%)高于HV(Max)({ref_hv_high:.1f}%) {(iv_hv_ratio_high-1)*100:.0f}%，预期回归',
                    '方向': '做空波动率',
                    '操作建议': ['卖出跨式', '卖出宽跨式', '铁鹰式', '铁蝶式'],
                    '成本分析': {
                        'IV/HV(Max)比': round(iv_hv_ratio_high, 2),
                        'IV': round(iv_atm, 2),
                        'HV20': round(hv20, 2),
                        'HV60': round(hv60, 2),
                        '波动率高估程度': f'{(iv_hv_ratio_high-1)*100:.0f}%',
                    },
                    '风险提示': ['卖出波动率策略风险无限', '需严格止损', '注意保证金水平'],
                    '希腊字母风险': {'Vega': '负暴露', 'Theta': '每日收益'},
                    '风险等级': '高',
                    '策略评分': min(85.0, 60 + (iv_hv_ratio_high - 1) * 50),
                }
            elif iv_hv_ratio_low < 0.75:  # IV显著低于低位HV，买入波动率
                strategy = {
                    '策略类型': '波动率锥-买入',
                    '策略描述': f'当前IV相对于HV(Min)折价明显({(1-iv_hv_ratio_low)*100:.0f}%)，建议买入低估的IV',
                    '策略名称': f'IV低估买波动 {underlying}',
                    '标的合约': underlying,
                    '交割年月': expiry,
                    '适用场景': '波动率均值回归',
                    '市场预期': f'IV({iv_atm:.1f}%)低于HV(Min)({ref_hv_low:.1f}%) {(1-iv_hv_ratio_low)*100:.0f}%，预期回归',
                    '方向': '做多波动率',
                    '操作建议': ['买入跨式', '买入宽跨式', '日历价差'],
                    '成本分析': {
                        'IV/HV(Min)比': round(iv_hv_ratio_low, 2),
                        'IV': round(iv_atm, 2),
                        'HV20': round(hv20, 2),
                        'HV60': round(hv60, 2),
                        '波动率低估程度': f'{(1-iv_hv_ratio_low)*100:.0f}%',
                    },
                    '风险提示': ['买入波动率策略时间价值衰减', '需波动率上升覆盖Theta损失'],
                    '希腊字母风险': {'Vega': '正暴露', 'Theta': '每日支出'},
                    '风险等级': '中',
                    '策略评分': min(80.0, 55 + (1 - iv_hv_ratio) * 80),
                }
            
            if strategy:
                strategies.append(strategy)
            
            return strategies
        except Exception as e:
            print(f"构建波动率锥策略失败: {e}")
            return []
    
    def _calculate_combo_greeks(self, legs):
        """计算组合Greeks
        
        Args:
            legs: 策略腿列表，每腿需包含操作、Delta、Gamma、Theta、Vega、数量
        
        Returns:
            组合Greeks字典
        """
        try:
            combo = {
                'Delta': 0.0,
                'Gamma': 0.0,
                'Theta': 0.0,
                'Vega': 0.0,
            }
            
            for leg in legs:
                action = leg.get('操作', '买入')
                qty = int(leg.get('数量', 1) or 1)
                multiplier = 1 if action == '买入' else -1
                
                delta = float(leg.get('Delta', 0) or 0)
                gamma = float(leg.get('Gamma', 0) or 0)
                theta = float(leg.get('Theta', 0) or 0)
                vega = float(leg.get('Vega', 0) or 0)
                
                combo['Delta'] += delta * qty * multiplier
                combo['Gamma'] += gamma * qty * multiplier
                combo['Theta'] += theta * qty * multiplier
                combo['Vega'] += vega * qty * multiplier
            
            # 生成对冲建议
            hedge_advice = []
            if abs(combo['Delta']) > 0.1:
                direction = '卖出' if combo['Delta'] > 0 else '买入'
                hedge_advice.append(f"{direction}{abs(combo['Delta']):.2f}手标的对冲Delta")
            
            if combo['Gamma'] < -0.01:
                hedge_advice.append("负Gamma暴露，标的大幅波动不利")
            elif combo['Gamma'] > 0.01:
                hedge_advice.append("正Gamma暴露，标的震荡有利")
            
            if combo['Vega'] < -50:
                hedge_advice.append("空Vega，波动率上升不利")
            elif combo['Vega'] > 50:
                hedge_advice.append("多Vega，波动率上升有利")
            
            combo['对冲建议'] = '; '.join(hedge_advice) if hedge_advice else '组合风险中性'
            
            return combo
        except Exception:
            return {'Delta': 0, 'Gamma': 0, 'Theta': 0, 'Vega': 0, '对冲建议': '计算失败'}
    
    def _build_seagull_strategy(self, wing1, body, wing2, market_state, vol_features):
        try:
            is_bullish = str(wing1.get('期权类型', '')).upper() == 'CALL'
            multiplier = float(wing1.get('合约乘数', 1) or 1)
            if is_bullish:
                leg_call = self._make_option_leg(wing1, '买入', 1)
                leg_put_sell = self._make_option_leg(body, '卖出', 1)
                leg_put_buy = self._make_option_leg(wing2, '买入', 1)
                legs = [leg_call, leg_put_sell, leg_put_buy]
                
                net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
                premium_paid = -net_cashflow
                premium_per_unit = premium_paid / multiplier
                
                k_call = float(wing1.get('行权价', 0) or 0)
                k_put_sell = float(body.get('行权价', 0) or 0)
                k_put_buy = float(wing2.get('行权价', 0) or 0)
                
                max_loss = (k_put_sell - k_put_buy) * multiplier + max(premium_paid, 0)
                breakeven_lower = k_put_sell + premium_per_unit
                breakeven_upper = k_call + premium_per_unit
                
                strategy = {
                    '策略类型': '海鸥（Seagull，偏斜交易）',
                    '策略描述': '买入价外CALL + 卖出价外PUT + 买入更价外PUT（用偏斜结构构建低成本上行敞口）',
                    '适用场景': f'下行偏斜较陡（PUT更贵），且方向偏{market_state.get("期货方向", "震荡")}；适合以更低成本做多上行',
                    '操作建议': legs,
                    '成本分析': {
                        '净权利金现金流(>0收入,<0支出)': net_cashflow,
                        '净权利金支出': max(premium_paid, 0),
                        '最大盈利': '理论无限（上行）',
                        '最大亏损': max_loss
                    },
                    '盈亏平衡点': {
                        '下限': breakeven_lower,
                        '上限': breakeven_upper
                    },
                    '风险提示': [
                        '属于偏斜交易：通常用“卖更贵的翼、买更便宜的翼”降低成本',
                        '下行风险被PUT价差限制，但仍需关注跳空与保证金占用',
                        f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                    ],
                    '希腊字母风险': {
                        'Delta': f'{(float(wing1.get("Delta", 0) or 0) - float(body.get("Delta", 0) or 0) + float(wing2.get("Delta", 0) or 0)):.4f}',
                        'Gamma': f'{(float(wing1.get("Gamma", 0) or 0) - float(body.get("Gamma", 0) or 0) + float(wing2.get("Gamma", 0) or 0)):.6f}',
                        'Theta': f'{(float(wing1.get("Theta", 0) or 0) - float(body.get("Theta", 0) or 0) + float(wing2.get("Theta", 0) or 0)):.4f}',
                        'Vega': f'{(float(wing1.get("Vega", 0) or 0) - float(body.get("Vega", 0) or 0) + float(wing2.get("Vega", 0) or 0)):.4f}'
                    }
                }
                return strategy
            else:
                leg_put = self._make_option_leg(wing1, '买入', 1)
                leg_call_sell = self._make_option_leg(body, '卖出', 1)
                leg_call_buy = self._make_option_leg(wing2, '买入', 1)
                legs = [leg_put, leg_call_sell, leg_call_buy]
                
                net_cashflow = float(sum([x.get('权利金', 0) or 0 for x in legs]))
                premium_paid = -net_cashflow
                premium_per_unit = premium_paid / multiplier
                
                k_put = float(wing1.get('行权价', 0) or 0)
                k_call_sell = float(body.get('行权价', 0) or 0)
                k_call_buy = float(wing2.get('行权价', 0) or 0)
                
                max_loss = (k_call_buy - k_call_sell) * multiplier + max(premium_paid, 0)
                breakeven_upper = k_call_sell - premium_per_unit
                breakeven_lower = k_put - premium_per_unit
                
                strategy = {
                    '策略类型': '海鸥（Seagull，偏斜交易）',
                    '策略描述': '买入价外PUT + 卖出价外CALL + 买入更价外CALL（用偏斜结构构建低成本下行敞口）',
                    '适用场景': f'偏斜结构明显，且方向偏{market_state.get("期货方向", "震荡")}；适合以更低成本做空下行',
                    '操作建议': legs,
                    '成本分析': {
                        '净权利金现金流(>0收入,<0支出)': net_cashflow,
                        '净权利金支出': max(premium_paid, 0),
                        '最大盈利': '理论无限（下行）',
                        '最大亏损': max_loss
                    },
                    '盈亏平衡点': {
                        '上限': breakeven_upper,
                        '下限': breakeven_lower
                    },
                    '风险提示': [
                        '属于偏斜交易：通常用“卖更贵的翼、买更便宜的翼”降低成本',
                        '上行风险被CALL价差限制，但仍需关注跳空与保证金占用',
                        f'标的现价：{market_state.get("标的现价", "")}；最大痛点：{market_state.get("最大痛点", "")}'
                    ],
                    '希腊字母风险': {
                        'Delta': f'{(float(wing1.get("Delta", 0) or 0) - float(body.get("Delta", 0) or 0) + float(wing2.get("Delta", 0) or 0)):.4f}',
                        'Gamma': f'{(float(wing1.get("Gamma", 0) or 0) - float(body.get("Gamma", 0) or 0) + float(wing2.get("Gamma", 0) or 0)):.6f}',
                        'Theta': f'{(float(wing1.get("Theta", 0) or 0) - float(body.get("Theta", 0) or 0) + float(wing2.get("Theta", 0) or 0)):.4f}',
                        'Vega': f'{(float(wing1.get("Vega", 0) or 0) - float(body.get("Vega", 0) or 0) + float(wing2.get("Vega", 0) or 0)):.4f}'
                    }
                }
                return strategy
        except Exception as e:
            print(f"构建海鸥策略失败: {e}")
            return None
    
    def _generate_covered_strategies(self, options_df, market_state, vol_features):
        """生成备兑/保护性策略
        
        备兑开仓（Covered Call）= 持有标的 + 卖出CALL
        保护性看跌（Protective Put）= 持有标的 + 买入PUT
        适合已持有标的或计划持有标的的投资者
        """
        strategies = []
        
        direction = market_state.get('期货方向', '')
        sentiment = market_state.get('期权情绪', '')
        
        # 筛选流动性好的期权
        liquid_options = options_df[
            (options_df['成交量'] >= 50) &
            (options_df['持仓量'] >= 100) &
            (options_df['剩余天数'] >= 20) &
            (options_df['剩余天数'] <= 90)
        ].copy()
        
        if liquid_options.empty:
            return strategies
        
        atm_price = market_state.get('标的现价', 0) or float(liquid_options['标的现价'].dropna().iloc[0] if '标的现价' in liquid_options.columns and not liquid_options['标的现价'].dropna().empty else 0)
        
        # 1. 备兑开仓策略（适合持有标的，预期温和上涨或震荡）
        if '震荡' in sentiment or '上涨' in direction:
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            if not calls.empty:
                # 选择价外看涨期权
                otm_calls = calls[calls['行权价'] > atm_price * 1.02].copy()
                
                if not otm_calls.empty:
                    # 选择虚值档位适中的期权
                    otm_calls['strike_diff'] = np.abs(otm_calls['行权价'] - atm_price * 1.05)
                    sell_call = otm_calls.loc[otm_calls['strike_diff'].idxmin()]
                    
                    strategy = self._build_covered_call_strategy(
                        sell_call, atm_price, market_state, vol_features
                    )
                    if strategy:
                        strategies.append(strategy)
        
        # 2. 保护性看跌策略（适合持有标的，担心下跌风险）
        if '下跌' in direction or '恐慌' in sentiment:
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            if not puts.empty:
                # 选择价外或平值看跌期权
                protective_puts = puts[puts['行权价'] >= atm_price * 0.95].copy()
                
                if not protective_puts.empty:
                    # 选择接近平值的期权
                    protective_puts['strike_diff'] = np.abs(protective_puts['行权价'] - atm_price)
                    buy_put = protective_puts.loc[protective_puts['strike_diff'].idxmin()]
                    
                    strategy = self._build_protective_put_strategy(
                        buy_put, atm_price, market_state, vol_features
                    )
                    if strategy:
                        strategies.append(strategy)
        
        # 3. 领口策略（Collar）= 持有标的 + 买入PUT + 卖出CALL
        if '震荡' in sentiment:
            calls = liquid_options[liquid_options['期权类型'] == 'CALL'].copy()
            puts = liquid_options[liquid_options['期权类型'] == 'PUT'].copy()
            
            if not calls.empty and not puts.empty:
                # 选择价外看涨
                otm_calls = calls[calls['行权价'] > atm_price * 1.05].copy()
                # 选择价外看跌
                otm_puts = puts[puts['行权价'] < atm_price * 0.95].copy()
                
                if not otm_calls.empty and not otm_puts.empty:
                    sell_call = otm_calls.sort_values('行权价').iloc[0]
                    buy_put = otm_puts.sort_values('行权价', ascending=False).iloc[0]
                    
                    strategy = self._build_collar_strategy(
                        buy_put, sell_call, atm_price, market_state, vol_features
                    )
                    if strategy:
                        strategies.append(strategy)
        
        return strategies
    
    def _build_covered_call_strategy(self, sell_call, underlying_price, market_state, vol_features):
        """构建备兑开仓策略"""
        try:
            multiplier = float(sell_call.get('合约乘数', 1) or 1)
            underlying = market_state.get('标的合约', '') or sell_call.get('标的合约', '')
            leg_underlying = self._make_underlying_leg(underlying, str(market_state.get('交割年月', '') or ''), '持有/买入', 1, float(underlying_price), multiplier)
            leg_call = self._make_option_leg(sell_call, '卖出', 1)
            legs = [leg_underlying, leg_call]
            
            underlying_trade_px = float(leg_underlying.get('建议成交价', underlying_price) or underlying_price)
            premium_received = float(leg_call.get('权利金', 0) or 0)
            
            # 最大收益：权利金 + (行权价 - 标的买入价)
            max_profit = premium_received + (sell_call['行权价'] - underlying_trade_px) * multiplier
            
            # 最大损失：标的价格跌至零
            max_loss = underlying_trade_px * multiplier - premium_received
            
            # 盈亏平衡点
            breakeven = underlying_trade_px - (premium_received / multiplier)
            
            strategy = {
                '策略类型': '备兑开仓（Covered Call）',
                '策略描述': '持有标的+卖出价外看涨期权，赚取权利金，适合温和上涨或震荡',
                '适用场景': f'{market_state.get("期货方向", "震荡或温和上涨")}，持有标的预期不会大涨',
                '操作建议': legs,
                '成本分析': {
                    '标的成本': underlying_trade_px * multiplier,
                    '收到权利金': premium_received,
                    '净成本': underlying_trade_px * multiplier - premium_received,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '说明': f'标的价格>{breakeven:.2f}开始盈利'
                },
                '风险提示': [
                    '适合持有标的的投资者，通过卖出期权增强收益',
                    '标的价格超过行权价将被行权，失去进一步上涨收益',
                    '标的价格下跌仍有亏损风险，权利金只能部分对冲',
                    '相当于为标的设定了卖出价格上限',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '适合长期持有标的者增强收益'
                ],
                '希腊字母风险': {
                    'Delta': f'{(1.0 - sell_call.get("Delta", 0)):.4f}',
                    'Gamma': f'{(-sell_call.get("Gamma", 0)):.6f}',
                    'Theta': f'{(-sell_call.get("Theta", 0)):.4f}',
                    'Vega': f'{(-sell_call.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建备兑开仓策略失败: {e}")
            return None
    
    def _build_protective_put_strategy(self, buy_put, underlying_price, market_state, vol_features):
        """构建保护性看跌策略"""
        try:
            multiplier = float(buy_put.get('合约乘数', 1) or 1)
            underlying = market_state.get('标的合约', '') or buy_put.get('标的合约', '')
            leg_underlying = self._make_underlying_leg(underlying, str(market_state.get('交割年月', '') or ''), '持有/买入', 1, float(underlying_price), multiplier)
            leg_put = self._make_option_leg(buy_put, '买入', 1)
            legs = [leg_underlying, leg_put]
            
            underlying_trade_px = float(leg_underlying.get('建议成交价', underlying_price) or underlying_price)
            premium_paid = -float(leg_put.get('权利金', 0) or 0)
            
            # 最大收益：标的价格上涨无限
            # 最大损失：(标的买入价 - 行权价) + 权利金
            max_loss = (underlying_trade_px - buy_put['行权价']) * multiplier + premium_paid
            
            # 盈亏平衡点
            breakeven = underlying_trade_px + (premium_paid / multiplier)
            
            strategy = {
                '策略类型': '保护性看跌（Protective Put）',
                '策略描述': '持有标的+买入看跌期权，对冲下跌风险，适合担心下跌但看好长期',
                '适用场景': f'{market_state.get("期货方向", "担心短期下跌")}，持有标的需要保护',
                '操作建议': legs,
                '成本分析': {
                    '标的成本': underlying_trade_px * multiplier,
                    '保护成本': premium_paid,
                    '总成本': underlying_trade_px * multiplier + premium_paid,
                    '最大盈利': '无限（标的上涨）',
                    '最大亏损': max_loss,
                    '保护行权价': buy_put['行权价']
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '说明': f'标的价格>{breakeven:.2f}开始盈利'
                },
                '风险提示': [
                    '适合持有标的但担心短期下跌风险的投资者',
                    '相当于为标的买了保险，保护价格不低于行权价',
                    '保险成本会降低整体收益，需权衡成本效益',
                    '标的价格未下跌，权利金将全部损失',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '适合重要风险事件前建立保护'
                ],
                '希腊字母风险': {
                    'Delta': f'{(1.0 + buy_put.get("Delta", 0)):.4f}',
                    'Gamma': f'{(buy_put.get("Gamma", 0)):.6f}',
                    'Theta': f'{(buy_put.get("Theta", 0)):.4f}',
                    'Vega': f'{(buy_put.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建保护性看跌策略失败: {e}")
            return None
    
    def _build_collar_strategy(self, buy_put, sell_call, underlying_price, market_state, vol_features):
        """构建领口策略"""
        try:
            multiplier = float(sell_call.get('合约乘数', 1) or 1)
            underlying = market_state.get('标的合约', '') or sell_call.get('标的合约', '')
            leg_underlying = self._make_underlying_leg(underlying, str(market_state.get('交割年月', '') or ''), '持有/买入', 1, float(underlying_price), multiplier)
            leg_put = self._make_option_leg(buy_put, '买入', 1)
            leg_call = self._make_option_leg(sell_call, '卖出', 1)
            legs = [leg_underlying, leg_put, leg_call]
            
            underlying_trade_px = float(leg_underlying.get('建议成交价', underlying_price) or underlying_price)
            put_cost = -float(leg_put.get('权利金', 0) or 0)
            call_premium = float(leg_call.get('权利金', 0) or 0)
            net_cost = put_cost - call_premium
            
            # 最大收益：标的涨至卖出行权价
            max_profit = (sell_call['行权价'] - underlying_trade_px) * multiplier - net_cost
            
            # 最大损失：标的跌至买入行权价
            max_loss = (underlying_trade_px - buy_put['行权价']) * multiplier + net_cost
            
            # 盈亏平衡点
            breakeven = underlying_trade_px + (net_cost / multiplier)
            
            strategy = {
                '策略类型': '领口策略（Collar）',
                '策略描述': '持有标的+买入PUT保护+卖出CALL降低成本，锁定盈亏区间',
                '适用场景': f'{market_state.get("期货方向", "震荡")}，需要保护但不想支付高额权利金',
                '操作建议': legs,
                '成本分析': {
                    '标的成本': underlying_trade_px * multiplier,
                    '期权净成本': net_cost,
                    '总成本': underlying_trade_px * multiplier + net_cost,
                    '最大盈利': max_profit,
                    '最大亏损': max_loss,
                    '盈亏比': max_profit / max_loss if max_loss > 0 else 0,
                    '锁定区间': f'{buy_put["行权价"]:.2f} - {sell_call["行权价"]:.2f}'
                },
                '盈亏平衡点': {
                    '盈亏平衡价': breakeven,
                    '下方保护价': buy_put['行权价'],
                    '上方限制价': sell_call['行权价']
                },
                '风险提示': [
                    '同时限制了最大收益和最大亏损',
                    '适合需要保护但预期涨幅有限的情况',
                    '卖出CALL的收入可以降低或抵消买入PUT的成本',
                    '零成本领口：选择合适行权价使净成本为零',
                    f'当前市场状态：{market_state.get("联动状态", "")}',
                    '适合长期持有者在不确定时期使用'
                ],
                '希腊字母风险': {
                    'Delta': f'{(1.0 + buy_put.get("Delta", 0) - sell_call.get("Delta", 0)):.4f}',
                    'Gamma': f'{(buy_put.get("Gamma", 0) - sell_call.get("Gamma", 0)):.6f}',
                    'Theta': f'{(buy_put.get("Theta", 0) - sell_call.get("Theta", 0)):.4f}',
                    'Vega': f'{(buy_put.get("Vega", 0) - sell_call.get("Vega", 0)):.4f}'
                }
            }
            
            return strategy
            
        except Exception as e:
            print(f"构建领口策略失败: {e}")
            return None
    
    def export_to_excel(self, all_strategies, output_file='wisecoin-期权策略.xlsx'):
        """导出策略到Excel"""
        # 确保输出目录存在
        output_dir = 'wisecoin_options_client_live_temp'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 更新输出文件路径
        if not output_file.startswith(output_dir):
            output_file = os.path.join(output_dir, os.path.basename(output_file))
            
        print(f"\n正在导出策略到 {output_file}...")
        
        wb = Workbook()
        active_ws = wb.active
        if active_ws is not None:
            wb.remove(active_ws)

        summary_ws = wb.create_sheet('汇总')
        
        # ============= 第一部分：市场判断与策略推荐 =============
        summary_row = 1
        summary_ws[f'A{summary_row}'] = '市场判断与策略推荐'
        summary_ws[f'A{summary_row}'].font = Font(bold=True, size=14, color='FFFFFF')
        summary_ws[f'A{summary_row}'].fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        summary_ws.merge_cells(f'A{summary_row}:F{summary_row}')
        summary_row += 1
        
        # 表头
        headers_recommend = ['标的', '市场判断', '推荐策略类型', '核心策略', '风险等级', '开仓时机']
        for col_idx, header in enumerate(headers_recommend, 1):
            cell = summary_ws.cell(row=summary_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_row += 1
        
        # 为每个标的生成推荐
        for underlying_data in all_strategies:
            underlying = underlying_data['underlying']
            market_state = underlying_data.get('market_state', {})
            
            # 市场判断逻辑
            market_judgment = self._judge_market_condition(market_state)
            recommended_strategies = self._recommend_strategies_by_market(market_judgment, market_state)
            
            # 写入推荐
            summary_ws.cell(row=summary_row, column=1, value=underlying)
            summary_ws.cell(row=summary_row, column=2, value=market_judgment.get('描述', ''))
            summary_ws.cell(row=summary_row, column=3, value=', '.join(recommended_strategies.get('类型', [])))
            summary_ws.cell(row=summary_row, column=4, value=', '.join(recommended_strategies.get('核心', [])))
            summary_ws.cell(row=summary_row, column=5, value=recommended_strategies.get('风险', ''))
            summary_ws.cell(row=summary_row, column=6, value=recommended_strategies.get('时机', ''))
            
            # 设置对齐
            for col_idx in range(1, 7):
                summary_ws.cell(row=summary_row, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)
            
            summary_row += 1
        
        summary_row += 2  # 空行
        
        # ============= 第二部分：策略统计 =============
        summary_ws.cell(row=summary_row, column=1, value='策略统计')
        summary_ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
        summary_ws.cell(row=summary_row, column=1).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        summary_ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_row += 1
        
        # 策略统计表头
        summary_ws[f'A{summary_row}'] = '策略类型'
        summary_ws[f'B{summary_row}'] = '适用品种'
        summary_ws[f'C{summary_row}'] = '数量'
        summary_ws[f'D{summary_row}'] = '覆盖到期'
        for col in ['A', 'B', 'C', 'D']:
            cell = summary_ws[f'{col}{summary_row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_row += 1
        
        strategy_index = {}
        for underlying_data in all_strategies:
            underlying = underlying_data['underlying']
            expiry = str(underlying_data['expiry'])
            strategies = underlying_data.get('strategies', [])
            if not strategies:
                continue
            for strat in strategies:
                s_type = str(strat.get('策略类型', ''))
                if not s_type:
                    continue
                entry = strategy_index.setdefault(s_type, {'underlyings': set(), 'expiries': set()})
                entry['underlyings'].add(underlying)
                entry['expiries'].add(expiry)
        for s_type in sorted(strategy_index.keys()):
            entry = strategy_index[s_type]
            underlyings_sorted = sorted(entry['underlyings'])
            expiries_sorted = sorted(entry['expiries'])
            summary_ws[f'A{summary_row}'] = s_type
            summary_ws[f'B{summary_row}'] = '，'.join(underlyings_sorted)
            summary_ws[f'C{summary_row}'] = len(underlyings_sorted)
            summary_ws[f'D{summary_row}'] = '，'.join(expiries_sorted)
            summary_row += 1
            
        self._auto_adjust_column_width(summary_ws)
        summary_ws.freeze_panes = 'A3'  # 冻结前3行（标题+表头）

        for underlying_data in all_strategies:
            for strategy in underlying_data.get('strategies', []) or []:
                if strategy is None:
                    continue
                if strategy.get('策略评分') is None:
                    strategy['策略评分'] = self._score_strategy(strategy)
                if strategy.get('策略要点') is None:
                    focus = self._format_strategy_focus(strategy)
                    if focus:
                        strategy['策略要点'] = {'操作要点': focus}

        self._create_strategy_guide_sheet(wb)
        self._create_strategy_details_sheet(wb, all_strategies)

        # 需要过滤的市场状态字段
        excluded_market_keys = {
            '标的中间价', '剩余天数(最小)', 
            '隐含波动率(ATM)', '近期波动率(ATM)', 'HV20(ATM)', 
            'IV/RV(ATM)', 'HV60(ATM)'
        }
        
        for underlying_data in all_strategies:
            underlying = underlying_data['underlying']
            expiry = underlying_data['expiry']
            strategies = underlying_data['strategies']
            market_state = underlying_data['market_state']
            
            if not strategies:
                continue
            
            # 创建工作表
            sheet_name = f"{underlying}_{expiry}"[:31]  # Excel sheet名称限制31字符
            ws = wb.create_sheet(sheet_name)
            
            # 写入标题
            ws.merge_cells('A1:N1')
            title_cell = ws['A1']
            title_cell.value = f'{underlying} {expiry} 期权策略建议'
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入市场状态
            current_row = 3
            ws.merge_cells(f'A{current_row}:N{current_row}')
            ws[f'A{current_row}'] = '市场状态概览'
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            
            current_row += 1
            for key, value in market_state.items():
                if key in excluded_market_keys:
                    continue
                ws[f'A{current_row}'] = key
                ws[f'B{current_row}'] = value
                current_row += 1
            
            # 写入每个策略
            current_row += 2
            for i, strategy in enumerate(strategies, 1):
                # 策略标题
                ws.merge_cells(f'A{current_row}:N{current_row}')
                ws[f'A{current_row}'] = f'策略 {i}: {strategy["策略类型"]}'
                ws[f'A{current_row}'].font = Font(size=12, bold=True, color='FFFFFF')
                ws[f'A{current_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                current_row += 1
                
                # 策略描述
                ws[f'A{current_row}'] = '策略描述'
                ws[f'B{current_row}'] = strategy.get('策略描述', '')
                current_row += 1
                
                ws[f'A{current_row}'] = '适用场景'
                ws[f'B{current_row}'] = self._normalize_scene(strategy.get('适用场景', ''))
                current_row += 1

                if strategy.get('策略标签') is not None:
                    ws[f'A{current_row}'] = '策略标签'
                    ws[f'B{current_row}'] = strategy.get('策略标签', '')
                    current_row += 1

                if strategy.get('策略评分') is not None:
                    ws[f'A{current_row}'] = '策略评分'
                    ws[f'B{current_row}'] = strategy.get('策略评分', '')
                    current_row += 1

                择时依据 = strategy.get('择时依据', None)
                if 择时依据:
                    current_row += 1
                    ws[f'A{current_row}'] = '择时依据'
                    ws[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1
                    if isinstance(择时依据, dict):
                        for key, value in 择时依据.items():
                            ws[f'A{current_row}'] = key
                            ws[f'B{current_row}'] = value
                            current_row += 1
                    else:
                        ws[f'B{current_row}'] = 择时依据
                        current_row += 1

                策略要点 = strategy.get('策略要点', None)
                if 策略要点:
                    current_row += 1
                    ws[f'A{current_row}'] = '策略要点'
                    ws[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1
                    if isinstance(策略要点, dict):
                        for key, value in 策略要点.items():
                            ws[f'A{current_row}'] = key
                            ws[f'B{current_row}'] = value
                            current_row += 1
                    else:
                        ws[f'B{current_row}'] = 策略要点
                        current_row += 1
                
                # 操作建议
                current_row += 1
                ws[f'A{current_row}'] = '操作建议'
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
                # 表头 (已去除中间价)
                headers = ['操作', '合约代码', '期权类型', '行权价', '剩余天数', '合约乘数', '参考价', '盘口买价', '盘口卖价', '建议成交价', '数量', '权利金', '报价时间']
                highlight_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                key_headers = ['建议成交价', '数量', '权利金']
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    if header in key_headers:
                        cell.fill = highlight_fill
                    else:
                        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                current_row += 1
                
                # 操作详情
                ops = strategy.get('操作建议', [])
                if ops and isinstance(ops, list) and len(ops) > 0 and isinstance(ops[0], dict):
                    for op in ops:
                        ws.cell(row=current_row, column=1, value=op.get('操作', ''))
                        ws.cell(row=current_row, column=2, value=op.get('合约代码', ''))
                        ws.cell(row=current_row, column=3, value=op.get('期权类型', ''))
                        ws.cell(row=current_row, column=4, value=op.get('行权价', ''))
                        ws.cell(row=current_row, column=5, value=op.get('剩余天数', ''))
                        ws.cell(row=current_row, column=6, value=op.get('合约乘数', ''))
                        ws.cell(row=current_row, column=7, value=op.get('参考价', ''))
                        ws.cell(row=current_row, column=8, value=op.get('盘口买价', ''))
                        ws.cell(row=current_row, column=9, value=op.get('盘口卖价', ''))
                        # 中间价已去除
                        ws.cell(row=current_row, column=10, value=op.get('建议成交价', ''))
                        ws.cell(row=current_row, column=11, value=op.get('数量', ''))
                        ws.cell(row=current_row, column=12, value=op.get('权利金', ''))
                        ws.cell(row=current_row, column=13, value=op.get('报价时间', ''))
                        current_row += 1
                else:
                    # 如果没有具体的操作建议详情，或者只是字符串列表
                    if ops and isinstance(ops, list):
                        ws[f'B{current_row-1}'] = ", ".join([str(x) for x in ops])
                    current_row += 1
                
                # 成本分析
                current_row += 1
                ws[f'A{current_row}'] = '成本分析'
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
                cost_analysis = strategy.get('成本分析', {})
                if cost_analysis and isinstance(cost_analysis, dict):
                    for key, value in cost_analysis.items():
                        ws[f'A{current_row}'] = key
                        ws[f'B{current_row}'] = f'{value:.2f}' if isinstance(value, (int, float)) else value
                        current_row += 1
                
                # 盈亏平衡点
                current_row += 1
                ws[f'A{current_row}'] = '盈亏平衡点'
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
                breakeven = strategy.get('盈亏平衡点', {})
                if breakeven:
                    if isinstance(breakeven, dict):
                        for key, value in breakeven.items():
                            ws[f'A{current_row}'] = key
                            ws[f'B{current_row}'] = f'{value:.2f}' if isinstance(value, (int, float)) else value
                            current_row += 1
                    else:
                        ws[f'B{current_row}'] = str(breakeven)
                        current_row += 1
                
                # 希腊字母
                current_row += 1
                ws[f'A{current_row}'] = '希腊字母风险'
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
                greeks = strategy.get('希腊字母风险', {})
                if greeks and isinstance(greeks, dict):
                    for key, value in greeks.items():
                        ws[f'A{current_row}'] = key
                        ws[f'B{current_row}'] = value
                        current_row += 1
                
                # 风险提示
                current_row += 1
                ws[f'A{current_row}'] = '风险提示'
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                current_row += 1
                
                risks = strategy.get('风险提示', [])
                if isinstance(risks, list):
                    for risk in risks:
                        ws[f'A{current_row}'] = '•'
                        ws[f'B{current_row}'] = risk
                        current_row += 1
                else:
                    ws[f'B{current_row}'] = str(risks)
                    current_row += 1
                
                current_row += 2  # 策略间空行
            
            # 设置列宽
            # 设置列宽和冻结
            self._auto_adjust_column_width(ws)
            ws.freeze_panes = 'B2'
            
        wb.save(output_file)
        print(f"✓ 策略已导出到 {output_file}")
    
    def run(self):
        """主运行流程"""
        if not self.load_data():
            return False
        
        # 获取所有标的
        underlyings = self.get_underlyings()
        if not underlyings:
            print("✗ 未找到任何标的数据")
            return False
        
        # 生成策略
        all_strategies = []
        for idx, und_info in enumerate(underlyings, 1):
            print(f"\n进度: {idx}/{len(underlyings)}")
            result = self.generate_strategies_for_underlying(
                und_info['underlying'],
                und_info['expiry']
            )
            if result and result['strategies']:
                all_strategies.append(result)
        
        # 导出Excel
        if all_strategies:
            self.export_to_excel(all_strategies)
            print(f"\n{'='*80}")
            print(f"策略生成完成！共处理 {len(all_strategies)} 个标的")
            print(f"{'='*80}")
            return True
        else:
            print("\n✗ 未生成任何有效策略")
            return False


def main():
    """主函数"""
    generator = OptionStrategyGenerator()
    success = generator.run()
    return 0 if success else 1


if __name__ == '__main__':
    import sys
    sys.exit(main())
