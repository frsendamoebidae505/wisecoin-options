"""
期权趋势排名分析交易系统 by playbonze
基于期权的隐含波动率、希腊字母、成交量和资金流向进行排名分析
自动选择期权品种进行交易
"""

import asyncio
import logging
import requests
import json
import pandas as pd
import numpy as np
import math
import decimal
import random
import datetime
import time
from datetime import datetime as dt
import sys
import os
import re
import pytz
import threading
import matplotlib.pyplot as plt
import socket
import warnings
import traceback
from tqsdk import TqApi, TqAuth, TqAccount, TqSim, TqBacktest, TqKq, BacktestFinished, TqChan, TargetPosTask, TqNotify
from tqsdk.ta import ATR, BOLL, MACD, OPTION_GREEKS, OPTION_IMPV, BS_VALUE, OPTION_VALUE
from tqsdk.tafunc import ma, ema, ema2, sma, time_to_datetime, crossup, crossdown, time_to_str, get_his_volatility, get_delta, get_gamma, get_vega, get_theta, get_t
from typing import List, Callable
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# 添加外部模块路径以支持 UnifiedLogger
sys.path.append(os.path.join(os.path.dirname(__file__), "wisecoin-catboost"))
from pb_quant_seektop_common import UnifiedLogger

# 设置统一日志
logger = UnifiedLogger.setup_logger_auto(__file__)

# 为了兼容现有代码，创建别名
logger_print = logger

# 全局文件配置 - 期权版本
TEMP_DIR = "wisecoin_options_client_live_temp"
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)
SYMBOL_EXCEL_FILE = os.path.join(TEMP_DIR, "wisecoin-期权品种.xlsx")

# 期权筛选参数
OPTION_FILTER_CONFIG = {
    'exclude_exchanges': ['SSE', 'SZSE'],  # 排除的交易所（股票期权）
    'max_quote_num': 99999,                  # 需要获取的期权行情个数（默认100，设置较大值如99999则获取全部）
}


async def get_underlying_futures_klines():
    """
    读取 wisecoin-期权行情.xlsx 中的标的，获取其对应的期货K线数据
    并保存为 wisecoin-期货K线.xlsx，每个合约一个分页，包含最近250个日K全字段。
    """
    QUOTE_EXCEL_FILE = os.path.join(TEMP_DIR, "wisecoin-期权行情.xlsx")
    KLINES_EXCEL_FILE = os.path.join(TEMP_DIR, "wisecoin-期货K线.xlsx")
    KLINE_DATA_LENGTH = 250  # 获取最近250个日K
    KLINE_DURATION = 24 * 60 * 60  # 日K线周期（秒）
    
    if not os.path.exists(QUOTE_EXCEL_FILE):
        logger.warning(f"未找到期权行情文件: {QUOTE_EXCEL_FILE}，无法获取期货K线。")
        return

    try:
        logger.info(f"正在读取 {QUOTE_EXCEL_FILE} 以获取标的期货列表...")
        xls = pd.ExcelFile(QUOTE_EXCEL_FILE)
        all_underlyings = set()

        # 遍历所有 Sheet 获取唯一的 underlying_symbol
        for sheet_name in xls.sheet_names:
            if sheet_name in ["Summary", "Progress", "Summary_Stats"]: 
                continue
            df = pd.read_excel(xls, sheet_name=sheet_name)
            if 'underlying_symbol' in df.columns:
                symbols = df['underlying_symbol'].dropna().unique().tolist()
                for s in symbols:
                    if s and isinstance(s, str) and '.' in s:
                        all_underlyings.add(s)
        
        unique_underlyings = sorted(list(all_underlyings))
        
        if not unique_underlyings:
            logger.warning("未在期权行情中发现有效的标的期货合约。")
            return
            
        logger.info(f"获取到 {len(unique_underlyings)} 个标的期货合约，准备获取日K线数据...")
        
        # 存储所有合约的K线数据
        all_klines_data = {}
        success_count = 0
        fail_count = 0
        
        for idx, symbol in enumerate(unique_underlyings):
            sym_str = str(symbol).strip()
            try:
                logger.info(f"[{idx+1}/{len(unique_underlyings)}] 获取 {sym_str} 的日K线数据...")
                
                # 使用 get_kline_serial 获取日K线数据 (参考 wisecoin-klines.py)
                klines = await api.get_kline_serial(sym_str, duration_seconds=KLINE_DURATION, data_length=KLINE_DATA_LENGTH)
                
                if klines is None or klines.empty:
                    logger.warning(f"合约 {sym_str} K线数据为空，跳过")
                    fail_count += 1
                    continue
                
                # 转换为 DataFrame 并添加合约标识
                klines_df = klines.copy()
                klines_df['symbol'] = sym_str
                
                # 提取品种代码用于分组
                if '.' in sym_str:
                    parts = sym_str.split('.')
                    exchange = parts[0]
                    code_match = re.match(r'^[a-zA-Z]+', parts[1])
                    if code_match:
                        product = f"{exchange}.{code_match.group(0)}"
                    else:
                        product = sym_str
                else:
                    product = "Unknown"
                klines_df['product'] = product
                
                # 转换时间戳为可读格式
                if 'datetime' in klines_df.columns:
                    klines_df['datetime_str'] = klines_df['datetime'].apply(
                        lambda x: time_to_str(x) if pd.notna(x) and x > 0 else ''
                    )
                
                all_klines_data[sym_str] = klines_df
                success_count += 1
                
            except asyncio.TimeoutError:
                logger.warning(f"获取合约 {sym_str} K线数据超时，跳过")
                fail_count += 1
            except Exception as e:
                logger.warning(f"获取合约 {sym_str} K线数据失败: {e}")
                fail_count += 1
            
        if not all_klines_data:
            logger.error(f"未能获取到任何期货K线数据。")
            return
            
        # 导出到 Excel，每个合约一个分页
        logger.info(f"正在将期货K线导出到 {KLINES_EXCEL_FILE}...")
        with pd.ExcelWriter(KLINES_EXCEL_FILE, engine='openpyxl') as writer:
            # 1. Summary Sheet - 汇总所有合约的最新K线
            summary_rows = []
            for symbol, klines_df in all_klines_data.items():
                if not klines_df.empty:
                    latest = klines_df.iloc[-1].to_dict()
                    summary_rows.append(latest)
            
            if summary_rows:
                summary_df = pd.DataFrame(summary_rows)
                # 按 product 排序
                if 'product' in summary_df.columns:
                    summary_df = summary_df.sort_values('product')
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # 调整列宽
                ws = writer.sheets['Summary']
                for col_idx, col in enumerate(summary_df.columns, 1):
                    try:
                        max_len = max(
                            summary_df[col].astype(str).map(len).max() if not summary_df[col].empty else 0,
                            len(str(col))
                        ) + 2
                    except:
                        max_len = 15
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)
            
            # 2. 每个合约一个分页
            for symbol, klines_df in all_klines_data.items():
                # 生成 Sheet 名称（Excel 限制31字符）
                sheet_name = symbol.replace('.', '_')[:31]
                
                # 按时间升序排列（最早的在前）
                if 'datetime' in klines_df.columns:
                    klines_df = klines_df.sort_values('datetime', ascending=True)
                
                klines_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 调整列宽
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = 'A2'  # 冻结表头
                
                for col_idx, col in enumerate(klines_df.columns, 1):
                    try:
                        max_len = max(
                            klines_df[col].astype(str).map(len).max() if not klines_df[col].empty else 0,
                            len(str(col))
                        ) + 2
                    except:
                        max_len = 15
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)
                    
        logger.info(f"🚀 期货K线保存完成: {KLINES_EXCEL_FILE}")
        logger.info(f"   成功: {success_count} 个合约，失败: {fail_count} 个合约")
        logger.info(f"   每个合约包含最近 {KLINE_DATA_LENGTH} 个日K线全字段数据")

    except Exception as e:
        logger.error(f"获取期货K线异常: {e}")
        logger.error(traceback.format_exc())

def _save_quotes_to_excel(file_path, sheet_df_map, all_quote_data, current_count, total_count):
    """内部辅助函数：将已获取的行情保存到 Excel"""
    try:
        logger.info(f"  [保存进度] 正在更新 Excel 数据 ({current_count}/{total_count})...")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheet_df_map.items():
                symbol_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
                if symbol_col not in df.columns:
                    continue
                
                # 合并原始字段和行情字段 (仅包含已获取到的 symbol)
                merged_rows = []
                for _, row in df.iterrows():
                    symbol = row[symbol_col]
                    if symbol in all_quote_data:
                        combined_data = row.to_dict()
                        # 行情数据转换
                        q_info = all_quote_data[symbol]
                        # 确保行情数据是字典
                        if not isinstance(q_info, dict):
                            # 如果是从 Excel 加载回来的，本来就是字典；如果是新获取的 Quote 对象转换的，也是字典
                            pass
                        combined_data.update(q_info)
                        merged_rows.append(combined_data)
                
                if not merged_rows:
                    continue
                
                final_df = pd.DataFrame(merged_rows)
                
                # 按照 underlying_symbol 、 strike_price 、 option_class (或 call_or_put) 排序
                sort_priority = ['underlying_symbol', 'strike_price', 'option_class', 'call_or_put']
                available_sort_keys = [k for k in sort_priority if k in final_df.columns]
                
                if available_sort_keys:
                    # 确保 strike_price 是数值类型以便正确排序
                    if 'strike_price' in final_df.columns:
                        final_df['strike_price'] = pd.to_numeric(final_df['strike_price'], errors='coerce')
                    final_df = final_df.sort_values(available_sort_keys)
                
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 自动调整格式
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(final_df.columns):
                    try:
                        max_len = max(final_df[col].astype(str).map(len).max() if not final_df[col].empty else 0, len(str(col))) + 2
                    except:
                        max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)
                
    except Exception as e:
        logger.error(f"保存 Excel 失败: {e}")
    except Exception as e:
        logger.error(f"阶段性保存 Excel 失败: {e}")


# 运行模式
RUN_MODE = 2
if os.path.basename(__file__) in ['pb-quant-test-1.py', 'pb-quant-test-2.py']:
    RUN_MODE = 1
elif os.path.basename(__file__) in ['pb-quant-kq.py']:
    RUN_MODE = 2
elif os.path.basename(__file__) in ['pb-quant-bh.py']:
    RUN_MODE = 4
elif os.path.basename(__file__) in ['pb-quant-jx.py']:
    RUN_MODE = 6

if RUN_MODE == 1:
    backtest_start_dt, backtest_end_dt = (datetime.date.today() + datetime.timedelta(days=-1), datetime.date.today() + datetime.timedelta(days=0))
    acc_sim = TqSim(init_balance=10000000)
    api = TqApi(account=acc_sim, backtest=TqBacktest(start_dt=backtest_start_dt,
                end_dt=backtest_end_dt), debug=False, web_gui=False, auth=TqAuth('playbonze', 'abC!@#123'))
elif RUN_MODE == 2:
    api = TqApi(TqKq(), debug=False, web_gui=False,
                auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 3:
    api = TqApi(TqAccount('simnow', '207302', 'Bonze!0613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 4:
    api = TqApi(TqAccount('B渤海期货', '98908572', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 5:
    api = TqApi(TqAccount('H华安期货', '100919200', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 6:
    api = TqApi(TqAccount('J金信期货', '80016087', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 7:
    api = TqApi(TqAccount('D东吴期货', '526178061', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 8:
    api = TqApi(TqAccount('H宏源期货', '901212925', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))

logger.info('模式%d，期权策略开始运行。' % RUN_MODE)

# 期权策略参数
try:
    # 创建期权分析任务
    async def run_sequence():
        await get_underlying_futures_klines()
        logger.info("所有标的期货K线数据获取完成。")
        api.close() # 关闭API
        sys.exit(0) # 退出脚本

    api.create_task(run_sequence())

    while True:
        api.wait_update()
        
except BacktestFinished:
    api.close()
    print('\n期权回测完成。')
except Exception as e:
    logger_print.info(f'{repr(e)}，line {sys._getframe().f_lineno}。')
    logger_print.info(traceback.format_exc())
