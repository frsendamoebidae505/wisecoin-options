#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
期权套利机会检测系统 (综合版)

整合8种低风险套利策略，所有操作基于真实盘口数据：

【无风险/低风险套利】
1. 平价套利 (Put-Call Parity): C-P=(F-K)*e^(-rT) 偏离
2. 垂直套利 (Vertical Spread): 期权价格单调性违反
3. 日历套利 (Calendar Spread): 时间价值曲线倒挂
4. 蝶式套利 (Butterfly Spread): 期权价格凸性违反
5. 盒式套利 (Box Spread): 买卖权组合锁定无风险利率

【价值回归套利】
6. 深度实值套利 (Deep ITM): 市场价 < 内在价值
7. 时间价值低估 (Time Value Undervalued): 市场价 < BS理论价格
8. 转换/逆转套利 (Conversion/Reversal): 合成头寸与标的价差

数据源（全部基于真实盘口）：
- wisecoin-期权参考.xlsx：期权基础数据、Greeks、理论价格
- wisecoin-期权行情.xlsx：实时盘口数据 (ask_price1, bid_price1, volume, open_interest)
- wisecoin-期货行情.xlsx：标的期货行情 (bid_price1, ask_price1, upper_limit, lower_limit)
- wisecoin-期货行情-无期权.xlsx：金融指数期货行情

输出：wisecoin-期权套利.xlsx
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
import logging
import traceback
from pathlib import Path
import warnings

warnings.filterwarnings('ignore')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ 全局参数 ============
TEMP_DIR = "wisecoin_options_client_live_temp"
if not Path(TEMP_DIR).exists():
    Path(TEMP_DIR).mkdir(parents=True, exist_ok=True)

RISK_FREE_RATE = 0.02  # 无风险利率 2%
# ... (rest of params remains same)
TRANSACTION_COST_RATE = 0.0003  # 交易成本 0.03%
MIN_PROFIT_THRESHOLD = 0.5  # 最小利润阈值 (%)
PARITY_THRESHOLD = 0.5  # 平价套利偏离阈值 (%)
INTRINSIC_TOLERANCE = 0.99  # 内在价值容差 (99%)
THEORETICAL_TOLERANCE = 0.95  # 理论价格容差 (95%)
IV_HV_BUFFER_THRESHOLD = 0.8  # IV低于min(HV)的缓冲阈值 (80%即留出20%空间)

# 流动性和实用性筛选
MIN_OPEN_INTEREST = 100  # 最小持仓量
MIN_VOLUME = 50  # 最小成交量
MAX_BID_ASK_SPREAD_PCT = 15  # 最大买卖价差百分比
MIN_DAYS_TO_EXPIRY = 3  # 最小剩余天数
MAX_DAYS_TO_EXPIRY = 100  # 最大剩余天数
MAX_UNDERVALUE_PCT = 100  # 最大低估比例

# 涨跌停判断容差
LIMIT_PRICE_TOLERANCE = 0.001  # 0.1%


class IntegratedArbitrageDetector:
    """期权套利机会综合检测器"""
    
    def __init__(self, 
                 options_ref_path=os.path.join(TEMP_DIR, 'wisecoin-期权参考.xlsx'),
                 options_quote_path=os.path.join(TEMP_DIR, 'wisecoin-期权行情.xlsx'),
                 futures_quote_path=os.path.join(TEMP_DIR, 'wisecoin-期货行情.xlsx')):
        """初始化"""
        self.options_ref_path = options_ref_path
        self.options_quote_path = options_quote_path
        self.futures_quote_path = futures_quote_path
        
        self.options_ref = None
        self.options_quote = None
        self.futures_quote = None
        self.non_opt_futures = None
        self.all_futures = None
        self.merged_options = None
        
        # 套利机会列表（8种策略）
        self.parity_opportunities = []       # 平价套利
        self.vertical_opportunities = []     # 垂直套利
        self.calendar_opportunities = []     # 日历套利
        self.butterfly_opportunities = []    # 蝶式套利
        self.box_opportunities = []          # 盒式套利
        self.deep_itm_opportunities = []     # 深度实值套利
        self.time_value_opportunities = []   # 时间价值低估
        self.conversion_opportunities = []   # 转换/逆转套利
        
    def load_data(self):
        """加载所有数据源"""
        logger.info("加载数据源...")
        
        # 加载期权参考数据
        try:
            self.options_ref = pd.read_excel(self.options_ref_path, sheet_name='期权参考')
            logger.info(f"期权参考数据: {len(self.options_ref)} 条")
        except Exception as e:
            logger.error(f"加载期权参考数据失败: {e}")
            return False
        
        # 加载期权盘口数据（合并所有品种分页）
        try:
            xl = pd.ExcelFile(self.options_quote_path)
            dfs = []
            for sheet in xl.sheet_names:
                df = pd.read_excel(xl, sheet_name=sheet)
                if len(df) > 0:
                    dfs.append(df)
            self.options_quote = pd.concat(dfs, ignore_index=True)
            logger.info(f"期权盘口数据: {len(self.options_quote)} 条")
        except Exception as e:
            logger.error(f"加载期权盘口数据失败: {e}")
            return False
        
        # 加载期货行情数据 (已合并金融期货)
        try:
            self.futures_quote = pd.read_excel(self.futures_quote_path)
            logger.info(f"期货行情数据(含金融): {len(self.futures_quote)} 条")
        except Exception as e:
            logger.error(f"加载期货行情数据失败: {e}")
            return False
        
        # 使用统一期货数据
        self.all_futures = self.futures_quote.copy()
        self.all_futures = self.all_futures.drop_duplicates(subset=['instrument_id'], keep='first')
        logger.info(f"所有期货合并完成: {len(self.all_futures)} 条")
        
        # 合并期权数据
        self._merge_options_data()
        
        return True
    
    def _merge_options_data(self):
        """合并期权参考数据和盘口数据"""
        logger.info("合并期权数据...")
        
        # 从期权参考中提取关键字段
        ref_cols = ['合约代码', '期权类型', '标的合约', '标的品种名称', '标的现价', '行权价', '期权价',
                    '剩余天数', '到期日', '交割年月', '期货合约乘数', '隐含波动率', '近期波动率', 'HV20', 'HV60',
                    'Delta', 'Gamma', 'Theta', 'Vega',
                    '买方期权费', '卖方保证金', '标的期货保证金', '理论价格', '内在价值', '时间价值', '虚实幅度%']
        available_cols = [c for c in ref_cols if c in self.options_ref.columns]
        ref_df = self.options_ref[available_cols].copy()
        
        # 从期权盘口中提取关键字段
        quote_cols = ['instrument_id', 'underlying_symbol', 'strike_price', 'option_class',
                      'ask_price1', 'ask_volume1', 'bid_price1', 'bid_volume1',
                      'last_price', 'volume', 'open_interest', 'volume_multiple',
                      'expire_rest_days', 'exercise_type']  # E=欧式, A=美式
        available_quote_cols = [c for c in quote_cols if c in self.options_quote.columns]
        quote_df = self.options_quote[available_quote_cols].copy()
        
        # 标准化期权代码用于匹配
        ref_df['合约代码_标准'] = ref_df['合约代码'].str.replace(r'^[A-Z]+\.', '', regex=True)
        quote_df['合约代码_标准'] = quote_df['instrument_id'].str.replace(r'^[A-Z]+\.', '', regex=True)
        
        # 合并数据
        self.merged_options = pd.merge(
            ref_df, quote_df,
            on='合约代码_标准',
            how='inner',
            suffixes=('_ref', '_quote')
        )
        
        # 筛选: 剩余天数、持仓量、成交量
        before_filter = len(self.merged_options)
        self.merged_options = self.merged_options[
            (self.merged_options['expire_rest_days'] >= MIN_DAYS_TO_EXPIRY) &
            (self.merged_options['expire_rest_days'] <= MAX_DAYS_TO_EXPIRY) &
            (self.merged_options['open_interest'] >= MIN_OPEN_INTEREST) &
            (self.merged_options['volume'] >= MIN_VOLUME)
        ].copy()
        after_filter = len(self.merged_options)
        logger.info(f"数据筛选: {before_filter} -> {after_filter} 条")
        
        # 标准化期权类型字段（兼容 CALL/PUT/C/P/认购/认沽 等格式）
        if '期权类型' in self.merged_options.columns:
            def normalize_option_type(opt_type):
                if opt_type is None:
                    return opt_type
                opt_str = str(opt_type).strip().upper()
                if opt_str in ('CALL', 'C', '认购', '看涨'):
                    return 'CALL'
                elif opt_str in ('PUT', 'P', '认沽', '看跌'):
                    return 'PUT'
                return opt_str
            self.merged_options['期权类型'] = self.merged_options['期权类型'].apply(normalize_option_type)
        
        # 计算内在价值（如果不存在）
        if '内在价值' not in self.merged_options.columns:
            self.merged_options['内在价值'] = np.where(
                self.merged_options['期权类型'].str.upper().str.contains('CALL|C'),
                np.maximum(0, self.merged_options['标的现价'] - self.merged_options['行权价']),
                np.maximum(0, self.merged_options['行权价'] - self.merged_options['标的现价'])
            )
        
        # 使用盘口价格作为实际交易价格
        self.merged_options['实际买入价'] = self.merged_options['ask_price1'].fillna(self.merged_options['期权价'])
        self.merged_options['实际卖出价'] = self.merged_options['bid_price1'].fillna(self.merged_options['期权价'])
        
        # 计算买卖价差百分比
        self.merged_options['买卖价差%'] = (
            (self.merged_options['ask_price1'].fillna(self.merged_options['期权价']) - 
             self.merged_options['bid_price1'].fillna(self.merged_options['期权价'])) / 
            self.merged_options['期权价'].replace(0, np.nan) * 100
        ).fillna(100)
        
        # 检查有效盘口
        self.merged_options['有效盘口'] = (
            self.merged_options['ask_price1'].notna() & 
            self.merged_options['bid_price1'].notna() &
            (self.merged_options['ask_price1'] > 0) &
            (self.merged_options['bid_price1'] > 0) &
            (self.merged_options['买卖价差%'] <= MAX_BID_ASK_SPREAD_PCT)
        )
        
        logger.info(f"合并后有效期权: {len(self.merged_options)} 条, 有效盘口: {self.merged_options['有效盘口'].sum()} 条")
    
    def _calculate_exec_score(self, row):
        """计算可执行性评分 (0-100)"""
        exec_score = 100
        bid_ask_spread = row.get('买卖价差%', 10)
        exec_score -= min(30, bid_ask_spread * 3)
        exec_score -= max(0, 30 - row.get('open_interest', 0) / 10)
        exec_score -= max(0, 20 - row.get('volume', 0) / 5)
        return max(0, min(100, exec_score))
    
    def _get_futures_price(self, underlying_symbol, delivery_month=None):
        """获取标的期货价格（使用真实盘口中间价）"""
        if self.all_futures is None:
            return None
        
        index_map = {
            'SSE.000852': 'IM', 'SSE.000300': 'IF',
            'SSE.000016': 'IH', 'SSE.000905': 'IC',
        }
        
        symbol = str(underlying_symbol).strip()
        is_index = symbol in index_map
        
        if is_index and delivery_month:
            product_id = index_map[symbol]
            month_str = str(delivery_month)
            if len(month_str) == 6: month_str = month_str[2:]
            matches = self.all_futures[
                (self.all_futures['product_id'] == product_id) & 
                (self.all_futures['instrument_id'].str.contains(month_str, na=False))
            ]
        else:
            # 严格匹配逻辑
            # 1. 尝试完全匹配 instrument_id
            matches = self.all_futures[self.all_futures['instrument_id'] == symbol]
            
            # 2. 如果未找到，尝试匹配 Exchange.Symbol 格式的后缀
            if len(matches) == 0:
                short_symbol = symbol.split('.')[-1]
                # 关键修复：使用 endswith('.symbol') 确保严格匹配
                # 例如：匹配 m2605 时，要求 instrument_id 以 .m2605 结尾
                # 这样 DCE.jm2605 (结尾是 .jm2605) 就不会被误匹配
                matches = self.all_futures[
                    self.all_futures['instrument_id'].str.endswith('.' + short_symbol, na=False)
                ]
        
        if len(matches) > 0:
            row = matches.iloc[0]
            # 优先使用盘口中间价
            if pd.notna(row.get('bid_price1')) and row['bid_price1'] > 0 and pd.notna(row.get('ask_price1')) and row['ask_price1'] > 0:
                return (row['bid_price1'] + row['ask_price1']) / 2
            return row.get('last_price') or row.get('settlement')
        
        return None
    
    def _get_futures_info(self, underlying_symbol, delivery_month=None):
        """获取期货合约详细信息"""
        if self.all_futures is None:
            return None
        
        index_map = {'SSE.000852': 'IM', 'SSE.000300': 'IF', 'SSE.000016': 'IH', 'SSE.000905': 'IC'}
        symbol = str(underlying_symbol).strip()
        
        if symbol in index_map and delivery_month:
            product_id = index_map[symbol]
            month_str = str(delivery_month)[2:] if len(str(delivery_month)) == 6 else str(delivery_month)
            matches = self.all_futures[
                (self.all_futures['product_id'] == product_id) & 
                (self.all_futures['instrument_id'].str.contains(month_str, na=False))
            ]
        else:
            # 严格匹配逻辑
            matches = self.all_futures[self.all_futures['instrument_id'] == symbol]
            if len(matches) == 0:
                short_symbol = symbol.split('.')[-1]
                # 关键修复：使用 endswith('.symbol') 且前面必须有点号
                matches = self.all_futures[
                    self.all_futures['instrument_id'].str.endswith('.' + short_symbol, na=False)
                ]
        
        return matches.iloc[0] if len(matches) > 0 else None
    
    def _check_futures_limit_status(self, futures_row):
        """检查期货是否涨跌停"""
        if futures_row is None:
            return False, None, ''
        
        last_price = futures_row.get('last_price', 0) or 0
        upper_limit = futures_row.get('upper_limit', 0) or 0
        lower_limit = futures_row.get('lower_limit', 0) or 0
        bid_price1 = futures_row.get('bid_price1', 0) or 0
        ask_price1 = futures_row.get('ask_price1', 0) or 0
        
        if last_price <= 0 or upper_limit <= 0 or lower_limit <= 0:
            return False, None, ''
        
        is_limit_up = (
            abs(last_price - upper_limit) / upper_limit <= LIMIT_PRICE_TOLERANCE or
            (ask_price1 <= 0 or ask_price1 >= upper_limit * 0.999)
        )
        
        is_limit_down = (
            abs(last_price - lower_limit) / lower_limit <= LIMIT_PRICE_TOLERANCE or
            (bid_price1 <= 0 or bid_price1 <= lower_limit * 1.001)
        )
        
        if is_limit_up:
            return True, 'UP', f"涨停({last_price}≈{upper_limit})"
        elif is_limit_down:
            return True, 'DOWN', f"跌停({last_price}≈{lower_limit})"
        
        return False, None, ''
    
    def _get_underlying_limit_status(self, underlying_symbol, delivery_month=None):
        """获取标的期货的涨跌停状态"""
        f_info = self._get_futures_info(underlying_symbol, delivery_month)
        if f_info is None:
            return False, None, '', None
        is_limit, limit_type, limit_info = self._check_futures_limit_status(f_info)
        return is_limit, limit_type, limit_info, f_info

    # ========== 1. 平价套利检测 ==========
    def detect_parity_arbitrage(self, threshold=PARITY_THRESHOLD):
        """检测期权平价套利机会（使用真实盘口价格）"""
        logger.info("检测平价套利机会...")
        self.parity_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        calls = valid_options[valid_options['期权类型'] == 'CALL'].copy()
        puts = valid_options[valid_options['期权类型'] == 'PUT'].copy()
        
        for _, call_row in calls.iterrows():
            underlying = call_row['标的合约']
            strike = call_row['行权价']
            days_to_expiry = call_row['剩余天数']
            delivery_month = call_row.get('交割年月')
            
            matching_puts = puts[
                (puts['标的合约'] == underlying) & 
                (puts['行权价'] == strike) &
                (puts['剩余天数'] == days_to_expiry)
            ]
            
            if len(matching_puts) == 0:
                continue
            
            put_row = matching_puts.iloc[0]
            
            futures_price = self._get_futures_price(underlying, delivery_month)
            if futures_price is None or futures_price <= 0:
                continue
            
            f_info = self._get_futures_info(underlying, delivery_month)
            f_margin = f_info.get('标的期货保证金', call_row.get('标的期货保证金', 0)) if f_info is not None else call_row.get('标的期货保证金', 0)
            
            # 检查涨跌停
            is_limit, limit_type, limit_info, _ = self._get_underlying_limit_status(underlying, delivery_month)
            futures_status = '正常' if not is_limit else limit_info
            
            T = days_to_expiry / 365.0
            discount_factor = np.exp(-RISK_FREE_RATE * T)
            theoretical_spread = (futures_price - strike) * discount_factor
            
            # 使用真实盘口价格
            call_buy = call_row['实际买入价']  # Ask1
            call_sell = call_row['实际卖出价']  # Bid1
            put_buy = put_row['实际买入价']    # Ask1
            put_sell = put_row['实际卖出价']    # Bid1
            
            multiplier = call_row.get('volume_multiple') or call_row.get('期货合约乘数', 1)
            
            # 正向套利：卖出认购(Bid) + 买入认沽(Ask) + 做多期货
            actual_spread_forward = call_sell - put_buy
            deviation_forward = (actual_spread_forward - theoretical_spread) / futures_price * 100
            forward_blocked = (limit_type == 'UP')  # 涨停无法做多
            
            if deviation_forward > threshold:
                profit = (actual_spread_forward - theoretical_spread) * multiplier
                capital = call_row.get('卖方保证金', 0) + put_row.get('买方期权费', 0) + f_margin
                if capital <= 0: capital = futures_price * multiplier * 0.15
                ret_pct = (profit / capital * 100) if capital > 0 else 0
                ann_ret = (ret_pct * 365 / days_to_expiry) if days_to_expiry > 0 else 0
                
                exec_score = int((self._calculate_exec_score(call_row) + self._calculate_exec_score(put_row)) / 2)
                if forward_blocked: exec_score = max(0, exec_score - 50)
                
                # 获取行权类型（E=欧式, A=美式）
                exercise_type = call_row.get('exercise_type', 'E')
                is_american = str(exercise_type).upper() == 'A'
                
                # 美式期权降低评分（提前行权风险）
                if is_american:
                    exec_score = max(0, exec_score - 20)
                
                risk_warning = '涨停无法做多期货' if forward_blocked else (
                    '美式期权存在提前行权风险' if is_american else '欧式期权需持有到期'
                )
                
                self.parity_opportunities.append({
                    '套利类型': '正向平价套利',
                    '标的合约': underlying,
                    '标的品种': call_row.get('标的品种名称', ''),
                    '行权方式': '美式' if is_american else '欧式',
                    '期货状态': futures_status,
                    '行权价': strike,
                    '剩余天数': int(days_to_expiry),
                    '期货价格': round(futures_price, 2),
                    '认购卖价(Bid1)': round(call_sell, 2),
                    '认沽买价(Ask1)': round(put_buy, 2),
                    '理论价差': round(theoretical_spread, 2),
                    '实际价差': round(actual_spread_forward, 2),
                    '偏离度%': round(deviation_forward, 2),
                    '预期收益': round(profit, 2),
                    '资金占用': round(capital, 2),
                    '年化收益%': round(ann_ret, 2),
                    '认购代码': call_row['合约代码'],
                    '认沽代码': put_row['合约代码'],
                    '可执行性评分': exec_score,
                    '可执行': '否' if forward_blocked else '是',
                    '交易指令': f"卖出{call_row['合约代码']}@{call_sell:.2f} + 买入{put_row['合约代码']}@{put_buy:.2f} + 做多期货@{futures_price:.2f}",
                    '策略说明': '卖出认购+买入认沽+做多期货，锁定正向平价偏离',
                    '风险提示': risk_warning,
                    '风险等级': '高' if (forward_blocked or is_american) else ('低' if deviation_forward > 2 and exec_score >= 60 else '中')
                })
            
            # 反向套利：买入认购(Ask) + 卖出认沽(Bid) + 做空期货
            actual_spread_reverse = call_buy - put_sell
            deviation_reverse = (theoretical_spread - actual_spread_reverse) / futures_price * 100
            reverse_blocked = (limit_type == 'DOWN')  # 跌停无法做空
            
            if deviation_reverse > threshold:
                profit = (theoretical_spread - actual_spread_reverse) * multiplier
                capital = call_row.get('买方期权费', 0) + put_row.get('卖方保证金', 0) + f_margin
                if capital <= 0: capital = futures_price * multiplier * 0.15
                ret_pct = (profit / capital * 100) if capital > 0 else 0
                ann_ret = (ret_pct * 365 / days_to_expiry) if days_to_expiry > 0 else 0
                
                exec_score = int((self._calculate_exec_score(call_row) + self._calculate_exec_score(put_row)) / 2)
                if reverse_blocked: exec_score = max(0, exec_score - 50)
                
                # 获取行权类型（E=欧式, A=美式）
                exercise_type = call_row.get('exercise_type', 'E')
                is_american = str(exercise_type).upper() == 'A'
                
                # 美式期权降低评分（提前行权风险）
                if is_american:
                    exec_score = max(0, exec_score - 20)
                
                risk_warning = '跌停无法做空期货' if reverse_blocked else (
                    '美式期权存在提前行权风险' if is_american else '欧式期权需持有到期'
                )
                
                self.parity_opportunities.append({
                    '套利类型': '反向平价套利',
                    '标的合约': underlying,
                    '标的品种': call_row.get('标的品种名称', ''),
                    '行权方式': '美式' if is_american else '欧式',
                    '期货状态': futures_status,
                    '行权价': strike,
                    '剩余天数': int(days_to_expiry),
                    '期货价格': round(futures_price, 2),
                    '认购买价(Ask1)': round(call_buy, 2),
                    '认沽卖价(Bid1)': round(put_sell, 2),
                    '理论价差': round(theoretical_spread, 2),
                    '实际价差': round(actual_spread_reverse, 2),
                    '偏离度%': round(deviation_reverse, 2),
                    '预期收益': round(profit, 2),
                    '资金占用': round(capital, 2),
                    '年化收益%': round(ann_ret, 2),
                    '认购代码': call_row['合约代码'],
                    '认沽代码': put_row['合约代码'],
                    '可执行性评分': exec_score,
                    '可执行': '否' if reverse_blocked else '是',
                    '交易指令': f"买入{call_row['合约代码']}@{call_buy:.2f} + 卖出{put_row['合约代码']}@{put_sell:.2f} + 做空期货@{futures_price:.2f}",
                    '策略说明': '买入认购+卖出认沽+做空期货，锁定反向平价偏离',
                    '风险提示': risk_warning,
                    '风险等级': '高' if (reverse_blocked or is_american) else ('低' if deviation_reverse > 2 and exec_score >= 60 else '中')
                })
        
        self.parity_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.parity_opportunities)} 个平价套利机会")
        return self.parity_opportunities

    # ========== 2. 垂直套利检测 ==========
    def detect_vertical_arbitrage(self):
        """检测垂直套利机会（价格单调性违反）"""
        logger.info("检测垂直套利机会...")
        self.vertical_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        
        for (underlying, days), group in valid_options.groupby(['标的合约', '剩余天数']):
            # 看涨期权
            calls = group[group['期权类型'] == 'CALL'].sort_values('行权价')
            for i in range(len(calls) - 1):
                low_k = calls.iloc[i]
                high_k = calls.iloc[i + 1]
                k1, k2 = low_k['行权价'], high_k['行权价']
                c1_buy, c2_sell = low_k['实际买入价'], high_k['实际卖出价']
                
                if c1_buy < c2_sell:  # 价格倒挂
                    arbitrage = c2_sell - c1_buy
                    multiplier = low_k.get('volume_multiple') or low_k.get('期货合约乘数', 1)
                    capital = low_k.get('买方期权费', 0) + high_k.get('卖方保证金', 0)
                    if capital <= 0: capital = c1_buy * multiplier
                    profit = arbitrage * multiplier
                    ret_pct = (profit / capital * 100) if capital > 0 else 0
                    ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                    exec_score = int((self._calculate_exec_score(low_k) + self._calculate_exec_score(high_k)) / 2)
                    
                    exercise_type = low_k.get('exercise_type', 'E')
                    is_american = str(exercise_type).upper() == 'A'
                    
                    self.vertical_opportunities.append({
                        '套利类型': '看涨期权垂直套利',
                        '标的合约': underlying,
                        '标的品种': low_k.get('标的品种名称', ''),
                        '行权方式': '美式' if is_american else '欧式',
                        '剩余天数': int(days),
                        '低行权价': k1, '高行权价': k2,
                        '低K买价(Ask1)': round(c1_buy, 2),
                        '高K卖价(Bid1)': round(c2_sell, 2),
                        '价格倒挂': round(arbitrage, 2),
                        '预期收益': round(profit, 2),
                        '资金占用': round(capital, 2),
                        '年化收益%': round(ann_ret, 2),
                        '低K代码': low_k['合约代码'],
                        '高K代码': high_k['合约代码'],
                        '可执行性评分': exec_score,
                        '可执行': '是',
                        '交易指令': f"买入{low_k['合约代码']}@{c1_buy:.2f} + 卖出{high_k['合约代码']}@{c2_sell:.2f}",
                        '策略说明': '买入低K + 卖出高K，锁定价差倒挂收益',
                        '风险提示': '美式期权存在提前指派风险' if is_american else '正常交易',
                        '风险等级': '低'
                    })
            
            # 看跌期权
            puts = group[group['期权类型'] == 'PUT'].sort_values('行权价')
            for i in range(len(puts) - 1):
                low_k = puts.iloc[i]
                high_k = puts.iloc[i + 1]
                k1, k2 = low_k['行权价'], high_k['行权价']
                p1_sell, p2_buy = low_k['实际卖出价'], high_k['实际买入价']
                
                if p2_buy < p1_sell:  # 价格倒挂
                    arbitrage = p1_sell - p2_buy
                    multiplier = low_k.get('volume_multiple') or low_k.get('期货合约乘数', 1)
                    capital = high_k.get('买方期权费', 0) + low_k.get('卖方保证金', 0)
                    if capital <= 0: capital = p2_buy * multiplier
                    profit = arbitrage * multiplier
                    ret_pct = (profit / capital * 100) if capital > 0 else 0
                    ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                    exec_score = int((self._calculate_exec_score(low_k) + self._calculate_exec_score(high_k)) / 2)
                    
                    exercise_type = low_k.get('exercise_type', 'E')
                    is_american = str(exercise_type).upper() == 'A'
                    
                    self.vertical_opportunities.append({
                        '套利类型': '看跌期权垂直套利',
                        '标的合约': underlying,
                        '标的品种': low_k.get('标的品种名称', ''),
                        '行权方式': '美式' if is_american else '欧式',
                        '剩余天数': int(days),
                        '低行权价': k1, '高行权价': k2,
                        '低K卖价(Bid1)': round(p1_sell, 2),
                        '高K买价(Ask1)': round(p2_buy, 2),
                        '价格倒挂': round(arbitrage, 2),
                        '预期收益': round(profit, 2),
                        '资金占用': round(capital, 2),
                        '年化收益%': round(ann_ret, 2),
                        '低K代码': low_k['合约代码'],
                        '高K代码': high_k['合约代码'],
                        '可执行性评分': exec_score,
                        '可执行': '是',
                        '交易指令': f"买入{high_k['合约代码']}@{p2_buy:.2f} + 卖出{low_k['合约代码']}@{p1_sell:.2f}",
                        '策略说明': '买入高K + 卖出低K，锁定价差倒挂收益',
                        '风险提示': '美式期权存在提前指派风险' if is_american else '正常交易',
                        '风险等级': '低'
                    })
        
        self.vertical_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.vertical_opportunities)} 个垂直套利机会")
        return self.vertical_opportunities

    # ========== 3. 日历套利检测 ==========
    def detect_calendar_arbitrage(self):
        """检测日历套利机会（时间价值曲线倒挂）"""
        logger.info("检测日历套利机会...")
        self.calendar_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        
        for (underlying, strike, opt_type), group in valid_options.groupby(['标的合约', '行权价', '期权类型']):
            if len(group) < 2:
                continue
            
            group = group.sort_values('剩余天数')
            
            for i in range(len(group) - 1):
                near = group.iloc[i]
                far = group.iloc[i + 1]
                
                near_sell = near['实际卖出价']  # 卖出近月(Bid)
                far_buy = far['实际买入价']      # 买入远月(Ask)
                
                if near_sell > far_buy:  # 价格倒挂
                    arbitrage = near_sell - far_buy
                    multiplier = near.get('volume_multiple') or near.get('期货合约乘数', 1)
                    capital = near.get('卖方保证金', 0) + far.get('买方期权费', 0)
                    if capital <= 0: capital = far_buy * multiplier
                    profit = arbitrage * multiplier
                    ret_pct = (profit / capital * 100) if capital > 0 else 0
                    ann_ret = (ret_pct * 365 / near['剩余天数']) if near['剩余天数'] > 0 else 0
                    exec_score = int((self._calculate_exec_score(near) + self._calculate_exec_score(far)) / 2)
                    
                    exercise_type = near.get('exercise_type', 'E')
                    is_american = str(exercise_type).upper() == 'A'
                    
                    self.calendar_opportunities.append({
                        '套利类型': f'{opt_type}日历套利',
                        '标的合约': underlying,
                        '标的品种': near.get('标的品种名称', ''),
                        '行权方式': '美式' if is_american else '欧式',
                        '行权价': strike,
                        '近月到期': int(near['剩余天数']),
                        '远月到期': int(far['剩余天数']),
                        '近月卖价(Bid1)': round(near_sell, 2),
                        '远月买价(Ask1)': round(far_buy, 2),
                        '价格倒挂': round(arbitrage, 2),
                        '预期收益': round(profit, 2),
                        '资金占用': round(capital, 2),
                        '年化收益%': round(ann_ret, 2),
                        '近月代码': near['合约代码'],
                        '远月代码': far['合约代码'],
                        '可执行性评分': exec_score,
                        '可执行': '是',
                        '交易指令': f"卖出{near['合约代码']}@{near_sell:.2f} + 买入{far['合约代码']}@{far_buy:.2f}",
                        '策略说明': '卖出近月+买入远月，锁定时间价值倒挂',
                        '风险提示': '美式期权近月端存在提前指派风险' if is_american else '需注意远月流动性',
                        '风险等级': '低'
                    })
        
        self.calendar_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.calendar_opportunities)} 个日历套利机会")
        return self.calendar_opportunities

    # ========== 4. 蝶式套利检测 ==========
    def detect_butterfly_arbitrage(self):
        """检测蝶式套利机会（凸性违反）"""
        logger.info("检测蝶式套利机会...")
        self.butterfly_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        
        for (underlying, days, opt_type), group in valid_options.groupby(['标的合约', '剩余天数', '期权类型']):
            if len(group) < 3:
                continue
            
            group = group.sort_values('行权价').reset_index(drop=True)
            
            for i in range(len(group) - 2):
                for j in range(i + 1, len(group) - 1):
                    for k in range(j + 1, len(group)):
                        low = group.iloc[i]
                        mid = group.iloc[j]
                        high = group.iloc[k]
                        
                        k1, k2, k3 = low['行权价'], mid['行权价'], high['行权价']
                        
                        # 检查是否近似等间距
                        interval1 = k2 - k1
                        interval2 = k3 - k2
                        if abs(interval1 - interval2) / max(interval1, interval2) > 0.1:
                            continue
                        
                        # 使用真实盘口价格
                        c1_buy = low['实际买入价']   # Ask1
                        c2_sell = mid['实际卖出价']  # Bid1
                        c3_buy = high['实际买入价']  # Ask1
                        
                        net_cost = c1_buy + c3_buy - 2 * c2_sell
                        
                        if net_cost < 0:  # 收到净权利金
                            arbitrage = -net_cost
                            multiplier = low.get('volume_multiple') or low.get('期货合约乘数', 1)
                            capital = low.get('买方期权费', 0) + high.get('买方期权费', 0) + 2 * mid.get('卖方保证金', 0)
                            if capital <= 0: capital = (c1_buy + c3_buy) * multiplier
                            profit = arbitrage * multiplier
                            ret_pct = (profit / capital * 100) if capital > 0 else 0
                            ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                            exec_score = int((self._calculate_exec_score(low) + self._calculate_exec_score(mid) + self._calculate_exec_score(high)) / 3)
                            
                            exercise_type = mid.get('exercise_type', 'E')
                            is_american = str(exercise_type).upper() == 'A'
                            
                            self.butterfly_opportunities.append({
                                '套利类型': f'{opt_type}蝶式套利',
                                '标的合约': underlying,
                                '标的品种': mid.get('标的品种名称', ''),
                                '行权方式': '美式' if is_american else '欧式',
                                '剩余天数': int(days),
                                'K1(低)': k1, 'K2(中)': k2, 'K3(高)': k3,
                                'C1买价(Ask1)': round(c1_buy, 2),
                                'C2卖价(Bid1)': round(c2_sell, 2),
                                'C3买价(Ask1)': round(c3_buy, 2),
                                '净收入': round(arbitrage, 2),
                                '预期收益': round(profit, 2),
                                '资金占用': round(capital, 2),
                                '年化收益%': round(ann_ret, 2),
                                'K1代码': low['合约代码'],
                                'K2代码': mid['合约代码'],
                                'K3代码': high['合约代码'],
                                '可执行性评分': exec_score,
                                '可执行': '是',
                                '交易指令': f"买1手{low['合约代码']}@{c1_buy:.2f} + 卖2手{mid['合约代码']}@{c2_sell:.2f} + 买1手{high['合约代码']}@{c3_buy:.2f}",
                                '策略说明': '买入K1/K3 + 卖出2倍K2，利用凸性违反锁定净收入',
                                '风险提示': '美式期权中间端存在提前指派风险' if is_american else '正常交易',
                                '风险等级': '低'
                            })
        
        self.butterfly_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.butterfly_opportunities)} 个蝶式套利机会")
        return self.butterfly_opportunities

    # ========== 5. 盒式套利检测 ==========
    def detect_box_arbitrage(self):
        """检测盒式套利机会 (Box Spread)"""
        logger.info("检测盒式套利机会...")
        self.box_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        
        for (underlying, days), group in valid_options.groupby(['标的合约', '剩余天数']):
            calls = group[group['期权类型'] == 'CALL']
            puts = group[group['期权类型'] == 'PUT']
            
            strikes = set(calls['行权价'].unique()) & set(puts['行权价'].unique())
            strikes = sorted(list(strikes))
            
            if len(strikes) < 2:
                continue
            
            for i in range(len(strikes) - 1):
                k1, k2 = strikes[i], strikes[i + 1]
                
                c1 = calls[calls['行权价'] == k1]
                c2 = calls[calls['行权价'] == k2]
                p1 = puts[puts['行权价'] == k1]
                p2 = puts[puts['行权价'] == k2]
                
                if len(c1) == 0 or len(c2) == 0 or len(p1) == 0 or len(p2) == 0:
                    continue
                
                c1 = c1.iloc[0]
                c2 = c2.iloc[0]
                p1 = p1.iloc[0]
                p2 = p2.iloc[0]
                
                # 盒式套利：买入牛市价差(买C1卖C2) + 买入熊市价差(买P2卖P1)
                # 理论价值 = (K2-K1) * e^(-rT)
                T = days / 365.0
                theoretical_value = (k2 - k1) * np.exp(-RISK_FREE_RATE * T)
                
                # 实际成本（使用真实盘口）
                box_cost = (c1['实际买入价'] - c2['实际卖出价'] + p2['实际买入价'] - p1['实际卖出价'])
                
                if box_cost > 0 and box_cost < theoretical_value * 0.98:
                    arbitrage = theoretical_value - box_cost
                    multiplier = c1.get('volume_multiple') or c1.get('期货合约乘数', 1)
                    capital = box_cost * multiplier
                    profit = arbitrage * multiplier
                    ret_pct = (profit / capital * 100) if capital > 0 else 0
                    ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                    exec_score = int((self._calculate_exec_score(c1) + self._calculate_exec_score(c2) + 
                                      self._calculate_exec_score(p1) + self._calculate_exec_score(p2)) / 4)
                    
                    exercise_type = c1.get('exercise_type', 'E')
                    is_american = str(exercise_type).upper() == 'A'
                    
                    self.box_opportunities.append({
                        '套利类型': '盒式套利',
                        '标的合约': underlying,
                        '标的品种': c1.get('标的品种名称', ''),
                        '行权方式': '美式' if is_american else '欧式',
                        '剩余天数': int(days),
                        'K1(低)': k1, 'K2(高)': k2,
                        '理论价值': round(theoretical_value, 2),
                        '实际成本': round(box_cost, 2),
                        '套利空间': round(arbitrage, 2),
                        '预期收益': round(profit, 2),
                        '资金占用': round(capital, 2),
                        '年化收益%': round(ann_ret, 2),
                        'C1代码': c1['合约代码'], 'C2代码': c2['合约代码'],
                        'P1代码': p1['合约代码'], 'P2代码': p2['合约代码'],
                        '可执行性评分': exec_score,
                        '可执行': '是',
                        '交易指令': f"买入{c1['合约代码']}@{c1['实际买入价']:.2f} + 卖出{c2['合约代码']}@{c2['实际卖出价']:.2f} + 买入{p2['合约代码']}@{p2['实际买入价']:.2f} + 卖出{p1['合约代码']}@{p1['实际卖出价']:.2f}",
                        '策略说明': '牛市价差+熊市价差组合，到期锁定(K2-K1)无风险收益',
                        '风险提示': '美式期权卖出端存在提前指派风险' if is_american else '正常交易',
                        '风险等级': '低'
                    })
        
        self.box_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.box_opportunities)} 个盒式套利机会")
        return self.box_opportunities

    # ========== 6. 深度实值套利检测 ==========
    def detect_deep_itm_arbitrage(self):
        """检测深度实值套利机会（市场价 < 内在价值）"""
        logger.info("检测深度实值套利机会...")
        self.deep_itm_opportunities = []
        
        if self.merged_options is None:
            return []
        
        df = self.merged_options.copy()
        price_col = '实际买入价' if '实际买入价' in df.columns else '期权价'
        
        basic_filter = (
            (df['剩余天数'] >= MIN_DAYS_TO_EXPIRY) &
            (df['剩余天数'] <= MAX_DAYS_TO_EXPIRY) &
            (df['内在价值'] > 0) &
            (df[price_col] < df['内在价值'] * INTRINSIC_TOLERANCE) &
            (df[price_col] > 0) &
            (df['open_interest'] >= MIN_OPEN_INTEREST) &
            (df['volume'] >= MIN_VOLUME) &
            (df['买卖价差%'] <= MAX_BID_ASK_SPREAD_PCT)
        )
        df = df[basic_filter]
        
        for _, row in df.iterrows():
            intrinsic = row['内在价值']
            market_price = row.get(price_col, row.get('期权价', 0))
            arb_spread = intrinsic - market_price
            arb_pct = (arb_spread / market_price * 100) if market_price > 0 else 0
            
            multiplier = row.get('volume_multiple') or row.get('期货合约乘数', 1)
            net_profit = arb_spread * multiplier
            futures_margin = row['标的现价'] * multiplier * 0.15
            capital = row.get('买方期权费', market_price * multiplier) + futures_margin
            ret_pct = (net_profit / capital * 100) if capital > 0 else 0
            ann_ret = (ret_pct * 365 / row['剩余天数']) if row['剩余天数'] > 0 else 0
            
            is_limit, limit_type, limit_info, _ = self._get_underlying_limit_status(row['标的合约'])
            futures_status = '正常' if not is_limit else limit_info
            
            opt_type = row['期权类型']
            is_call = 'CALL' in str(opt_type).upper() or 'C' == str(opt_type).upper()
            
            is_blocked = False
            if is_call and limit_type == 'DOWN':
                is_blocked = True
            elif not is_call and limit_type == 'UP':
                is_blocked = True
            
            buy_price = row.get('实际买入价', market_price)
            if is_call:
                operation = f"买入{row['合约代码']}@{buy_price:.2f} + 卖出标的{row['标的合约']}@{row['标的现价']:.2f}"
            else:
                operation = f"买入{row['合约代码']}@{buy_price:.2f} + 买入标的{row['标的合约']}@{row['标的现价']:.2f}"
            
            exec_score = int(self._calculate_exec_score(row))
            if is_blocked: exec_score = max(0, exec_score - 50)
            
            # 获取行权类型
            exercise_type = row.get('exercise_type', 'E')
            is_american = str(exercise_type).upper() == 'A'
            
            risk_warning = '涨跌停导致对冲受阻' if is_blocked else (
                '美式期权支持提前行权或市场平仓' if is_american else '欧式期权需持有到期'
            )
            
            self.deep_itm_opportunities.append({
                '套利类型': '深度实值套利',
                '期权类型': opt_type,
                '行权方式': '美式' if is_american else '欧式',
                '合约代码': row['合约代码'],
                '标的合约': row['标的合约'],
                '标的品种': row.get('标的品种名称', ''),
                '期货状态': futures_status,
                '行权价': row['行权价'],
                '标的现价': round(row['标的现价'], 2),
                '期权市价': round(market_price, 2),
                '买入价(Ask1)': round(buy_price, 2),
                '内在价值': round(intrinsic, 2),
                '套利空间': round(arb_spread, 2),
                '套利空间%': round(arb_pct, 2),
                '剩余天数': int(row['剩余天数']),
                '净利润': round(net_profit, 2),
                '资金占用': round(capital, 2),
                '年化收益%': round(ann_ret, 2),
                '持仓量': int(row.get('open_interest', 0)),
                '成交量': int(row.get('volume', 0)),
                '可执行性评分': exec_score,
                '可执行': '否' if is_blocked else '是',
                '操作建议': operation,
                '策略说明': '买入期权+对冲标的，锁定内在价值价差',
                '风险提示': risk_warning,
                '风险等级': '高' if is_blocked else ('低' if (arb_pct > 2 or is_american) and exec_score >= 60 else '中')
            })
        
        self.deep_itm_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['套利空间%']))
        logger.info(f"发现 {len(self.deep_itm_opportunities)} 个深度实值套利机会")
        return self.deep_itm_opportunities

    # ========== 7. 时间价值低估套利检测 ==========
    def detect_time_value_arbitrage(self):
        """检测时间价值低估套利机会（市场价 < BS理论价格）"""
        logger.info("检测时间价值低估套利机会...")
        self.time_value_opportunities = []
        
        if self.merged_options is None or '理论价格' not in self.merged_options.columns:
            return []
        
        df = self.merged_options.copy()
        price_col = '实际买入价' if '实际买入价' in df.columns else '期权价'
        
        basic_filter = (
            (df['剩余天数'] >= MIN_DAYS_TO_EXPIRY) &
            (df['剩余天数'] <= MAX_DAYS_TO_EXPIRY) &
            (df[price_col] >= df['内在价值'] * INTRINSIC_TOLERANCE) &
            (df[price_col] < df['理论价格'] * THEORETICAL_TOLERANCE) &
            (df['理论价格'] > 0) &
            (df[price_col] > 0) &
            (df['open_interest'] >= MIN_OPEN_INTEREST) &
            (df['volume'] >= MIN_VOLUME) &
            (df['买卖价差%'] <= MAX_BID_ASK_SPREAD_PCT)
        )
        df = df[basic_filter]
        
        df['undervalue_pct_temp'] = ((df['理论价格'] - df[price_col]) / df[price_col] * 100).fillna(0)
        df = df[df['undervalue_pct_temp'] <= MAX_UNDERVALUE_PCT].copy()
        
        for _, row in df.iterrows():
            market_price = row.get(price_col, row.get('期权价', 0))
            theoretical_price = row['理论价格']
            undervalue = theoretical_price - market_price
            undervalue_pct = (undervalue / market_price * 100) if market_price > 0 else 0
            
            multiplier = row.get('volume_multiple') or row.get('期货合约乘数', 1)
            capital = row.get('买方期权费', market_price * multiplier)
            expected_profit = undervalue * multiplier
            ret_pct = (expected_profit / capital * 100) if capital > 0 else 0
            ann_ret = (ret_pct * 365 / row['剩余天数']) if row['剩余天数'] > 0 else 0
            
            is_limit, limit_type, limit_info, _ = self._get_underlying_limit_status(row['标的合约'])
            futures_status = '正常' if not is_limit else limit_info
            
            opt_type = row['期权类型']
            is_call = 'CALL' in str(opt_type).upper() or 'C' == str(opt_type).upper()
            
            hedge_blocked = False
            if is_call and limit_type == 'DOWN':
                hedge_blocked = True
            elif not is_call and limit_type == 'UP':
                hedge_blocked = True
            
            hv20 = row.get('HV20', row.get('近期波动率', 0))
            hv60 = row.get('HV60', hv20)
            iv = row.get('隐含波动率', 0)
            # 使用 HV20 和 HV60 的较小值作为参考（保守低估判断）
            ref_hv = min(hv20, hv60)
            iv_hv_ratio = (iv / ref_hv) if ref_hv > 0 else 1.0
            
            # 强化过滤：必须留出适当空间 (IV < min(HV) * Threshold)
            if iv_hv_ratio >= IV_HV_BUFFER_THRESHOLD:
                continue
                
            if iv_hv_ratio < 0.7:  # 空间极度低估
                recommended_strategy = "波动率大幅回归策略"
            elif iv_hv_ratio < IV_HV_BUFFER_THRESHOLD:
                recommended_strategy = "波动率均值回归策略"
            else:
                recommended_strategy = "观望"
            
            exec_score = int(self._calculate_exec_score(row))
            
            exercise_type = row.get('exercise_type', 'E')
            is_american = str(exercise_type).upper() == 'A'
            
            self.time_value_opportunities.append({
                '套利类型': '时间价值低估',
                '期权类型': opt_type,
                '行权方式': '美式' if is_american else '欧式',
                '合约代码': row['合约代码'],
                '标的合约': row['标的合约'],
                '标的品种': row.get('标的品种名称', ''),
                '期货状态': futures_status,
                '行权价': row['行权价'],
                '标的现价': round(row['标的现价'], 2),
                '期权市价': round(market_price, 2),
                '买入价(Ask1)': round(row.get('实际买入价', market_price), 2),
                '理论价格': round(theoretical_price, 2),
                '低估金额': round(undervalue, 2),
                '低估比例%': round(undervalue_pct, 2),
                '剩余天数': int(row['剩余天数']),
                '预期收益': round(expected_profit, 2),
                '资金占用': round(capital, 2),
                '年化收益%': round(ann_ret, 2),
                'IV%': round(iv, 2),
                'HV20%': round(hv20, 2),
                'HV60%': round(hv60, 2),
                'IV/HV比': round(iv_hv_ratio, 2),
                '持仓量': int(row.get('open_interest', 0)),
                '成交量': int(row.get('volume', 0)),
                '可执行性评分': exec_score,
                '对冲可执行': '否' if hedge_blocked else '是',
                '推荐策略': recommended_strategy,
                '操作建议': f"买入{row['合约代码']}@{row['实际买入价']:.2f}",
                '策略说明': f"IV低于HV，市场价低于理论价{undervalue_pct:.1f}%",
                '风险提示': '对冲受阻，建议直接买入' if hedge_blocked else '波动率可能继续下跌',
                '风险等级': '低' if undervalue_pct > 10 and exec_score >= 60 else ('中' if undervalue_pct > 5 else '高')
            })
        
        self.time_value_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['低估比例%']))
        logger.info(f"发现 {len(self.time_value_opportunities)} 个时间价值低估机会")
        return self.time_value_opportunities

    # ========== 8. 转换/逆转套利检测 ==========
    def detect_conversion_arbitrage(self):
        """检测转换/逆转套利机会"""
        logger.info("检测转换/逆转套利机会...")
        self.conversion_opportunities = []
        
        if self.merged_options is None:
            return []
        
        valid_options = self.merged_options[self.merged_options['有效盘口']].copy()
        calls = valid_options[valid_options['期权类型'] == 'CALL']
        puts = valid_options[valid_options['期权类型'] == 'PUT']
        
        for _, call_row in calls.iterrows():
            underlying = call_row['标的合约']
            strike = call_row['行权价']
            days = call_row['剩余天数']
            delivery_month = call_row.get('交割年月')
            
            matching_puts = puts[
                (puts['标的合约'] == underlying) & 
                (puts['行权价'] == strike) &
                (puts['剩余天数'] == days)
            ]
            
            if len(matching_puts) == 0:
                continue
            
            put_row = matching_puts.iloc[0]
            futures_price = self._get_futures_price(underlying, delivery_month)
            if futures_price is None or futures_price <= 0:
                continue
            
            is_limit, limit_type, limit_info, f_info = self._get_underlying_limit_status(underlying, delivery_month)
            futures_status = '正常' if not is_limit else limit_info
            
            T = days / 365.0
            discount = np.exp(-RISK_FREE_RATE * T)
            
            multiplier = call_row.get('volume_multiple') or call_row.get('期货合约乘数', 1)
            
            # 转换套利：买入标的 + 买入看跌(Ask) + 卖出看涨(Bid)
            # 锁定价值 = K * discount
            conversion_cost = futures_price + put_row['实际买入价'] - call_row['实际卖出价']
            conversion_value = strike * discount
            conversion_profit = (conversion_value - conversion_cost) * multiplier
            conversion_blocked = (limit_type == 'UP')
            
            if conversion_profit > 0:
                # 资金占用 = 期货保证金 + 买入Put权利金 + 卖出Call卖方保证金
                futures_margin = futures_price * multiplier * 0.15
                put_premium = put_row.get('买方期权费', 0) or (put_row['实际买入价'] * multiplier)
                call_seller_margin = call_row.get('卖方保证金', 0) or (call_row['实际卖出价'] * multiplier * 0.15)
                capital = futures_margin + put_premium + call_seller_margin
                ret_pct = (conversion_profit / capital * 100) if capital > 0 else 0
                ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                exec_score = int((self._calculate_exec_score(call_row) + self._calculate_exec_score(put_row)) / 2)
                if conversion_blocked: exec_score = max(0, exec_score - 50)
                
                # 获取行权类型（E=欧式, A=美式）
                exercise_type = call_row.get('exercise_type', 'E')
                is_american = str(exercise_type).upper() == 'A'
                if is_american:
                    exec_score = max(0, exec_score - 20)
                
                risk_warning = '涨停无法买入标的' if conversion_blocked else (
                    '美式期权存在提前行权风险' if is_american else '欧式期权需持有到期'
                )
                
                self.conversion_opportunities.append({
                    '套利类型': '转换套利',
                    '标的合约': underlying,
                    '标的品种': call_row.get('标的品种名称', ''),
                    '行权方式': '美式' if is_american else '欧式',
                    '期货状态': futures_status,
                    '行权价': strike,
                    '剩余天数': int(days),
                    '期货价格': round(futures_price, 2),
                    '认购卖价(Bid1)': round(call_row['实际卖出价'], 2),
                    '认沽买价(Ask1)': round(put_row['实际买入价'], 2),
                    '组合成本': round(conversion_cost, 2),
                    '锁定价值': round(conversion_value, 2),
                    '预期收益': round(conversion_profit, 2),
                    '资金占用': round(capital, 2),
                    '年化收益%': round(ann_ret, 2),
                    '认购代码': call_row['合约代码'],
                    '认沽代码': put_row['合约代码'],
                    '可执行性评分': exec_score,
                    '可执行': '否' if conversion_blocked else '是',
                    '交易指令': f"买入期货@{futures_price:.2f} + 买入{put_row['合约代码']}@{put_row['实际买入价']:.2f} + 卖出{call_row['合约代码']}@{call_row['实际卖出价']:.2f}",
                    '策略说明': '买入标的+买入看跌+卖出看涨，锁定K价值',
                    '风险提示': risk_warning,
                    '风险等级': '高' if (conversion_blocked or is_american) else '低'
                })
            
            # 逆转套利：卖出标的 + 卖出看跌(Bid) + 买入看涨(Ask)
            reversal_income = futures_price + put_row['实际卖出价'] - call_row['实际买入价']
            reversal_cost = strike * discount
            reversal_profit = (reversal_income - reversal_cost) * multiplier
            reversal_blocked = (limit_type == 'DOWN')
            
            if reversal_profit > 0:
                # 资金占用 = 期货保证金 + 买入Call权利金 + 卖出Put卖方保证金
                futures_margin = futures_price * multiplier * 0.15
                call_premium = call_row.get('买方期权费', 0) or (call_row['实际买入价'] * multiplier)
                put_seller_margin = put_row.get('卖方保证金', 0) or (put_row['实际卖出价'] * multiplier * 0.15)
                capital = futures_margin + call_premium + put_seller_margin
                ret_pct = (reversal_profit / capital * 100) if capital > 0 else 0
                ann_ret = (ret_pct * 365 / days) if days > 0 else 0
                exec_score = int((self._calculate_exec_score(call_row) + self._calculate_exec_score(put_row)) / 2)
                if reversal_blocked: exec_score = max(0, exec_score - 50)
                
                # 获取行权类型（E=欧式, A=美式）
                exercise_type = call_row.get('exercise_type', 'E')
                is_american = str(exercise_type).upper() == 'A'
                if is_american:
                    exec_score = max(0, exec_score - 20)
                
                risk_warning = '跌停无法卖出标的' if reversal_blocked else (
                    '美式期权存在提前行权风险' if is_american else '欧式期权需持有到期'
                )
                
                self.conversion_opportunities.append({
                    '套利类型': '逆转套利',
                    '标的合约': underlying,
                    '标的品种': call_row.get('标的品种名称', ''),
                    '行权方式': '美式' if is_american else '欧式',
                    '期货状态': futures_status,
                    '行权价': strike,
                    '剩余天数': int(days),
                    '期货价格': round(futures_price, 2),
                    '认购买价(Ask1)': round(call_row['实际买入价'], 2),
                    '认沽卖价(Bid1)': round(put_row['实际卖出价'], 2),
                    '组合收入': round(reversal_income, 2),
                    '到期成本': round(reversal_cost, 2),
                    '预期收益': round(reversal_profit, 2),
                    '资金占用': round(capital, 2),
                    '年化收益%': round(ann_ret, 2),
                    '认购代码': call_row['合约代码'],
                    '认沽代码': put_row['合约代码'],
                    '可执行性评分': exec_score,
                    '可执行': '否' if reversal_blocked else '是',
                    '交易指令': f"卖出期货@{futures_price:.2f} + 卖出{put_row['合约代码']}@{put_row['实际卖出价']:.2f} + 买入{call_row['合约代码']}@{call_row['实际买入价']:.2f}",
                    '策略说明': '卖出标的+卖出看跌+买入看涨，锁定超额收益',
                    '风险提示': risk_warning,
                    '风险等级': '高' if (reversal_blocked or is_american) else '低'
                })
        
        self.conversion_opportunities.sort(key=lambda x: (-x['可执行性评分'], -x['年化收益%']))
        logger.info(f"发现 {len(self.conversion_opportunities)} 个转换/逆转套利机会")
        return self.conversion_opportunities

    def run_all_detections(self):
        """运行所有套利检测"""
        logger.info("=" * 70)
        logger.info("开始全面套利机会扫描 (8种策略)...")
        logger.info("=" * 70)
        
        self.detect_parity_arbitrage()
        self.detect_vertical_arbitrage()
        self.detect_calendar_arbitrage()
        self.detect_butterfly_arbitrage()
        self.detect_box_arbitrage()
        self.detect_deep_itm_arbitrage()
        self.detect_time_value_arbitrage()
        self.detect_conversion_arbitrage()
        
        total = (len(self.parity_opportunities) + len(self.vertical_opportunities) +
                 len(self.calendar_opportunities) + len(self.butterfly_opportunities) +
                 len(self.box_opportunities) + len(self.deep_itm_opportunities) +
                 len(self.time_value_opportunities) + len(self.conversion_opportunities))
        
        logger.info("=" * 70)
        logger.info(f"套利扫描完成，共发现 {total} 个机会")
        logger.info(f"  - 平价套利: {len(self.parity_opportunities)} 个")
        logger.info(f"  - 垂直套利: {len(self.vertical_opportunities)} 个")
        logger.info(f"  - 日历套利: {len(self.calendar_opportunities)} 个")
        logger.info(f"  - 蝶式套利: {len(self.butterfly_opportunities)} 个")
        logger.info(f"  - 盒式套利: {len(self.box_opportunities)} 个")
        logger.info(f"  - 深度实值: {len(self.deep_itm_opportunities)} 个")
        logger.info(f"  - 时间价值低估: {len(self.time_value_opportunities)} 个")
        logger.info(f"  - 转换/逆转套利: {len(self.conversion_opportunities)} 个")
        logger.info("=" * 70)
        
        return total

    def _auto_adjust_column_width(self, worksheet, df, max_width=60, min_width=8):
        """自动调整列宽"""
        from openpyxl.utils import get_column_letter
        for col_idx, column in enumerate(df.columns, 1):
            # 计算列宽：中文字符算2，英文字符算1.1，留出一些padding
            def get_visual_width(s):
                if not s: return 0
                s = str(s)
                return sum(2.1 if ord(c) > 127 else 1.1 for c in s)

            header_len = get_visual_width(column)
            max_content_len = 0
            # 采样前200行以提高效率，但要覆盖头部
            for value in df[column].astype(str).head(200):
                content_len = get_visual_width(value)
                max_content_len = max(max_content_len, content_len)
            
            optimal_width = max(header_len, max_content_len) + 2
            final_width = max(min_width, min(optimal_width, max_width))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = final_width

    def _apply_header_style(self, worksheet):
        """应用表头样式"""
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        highlight_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # 重点字段浅黄背景
        
        # 重点关注的字段关键词
        key_fields = ['年化', '收益率', '套利空间', '低估', '资金', '利润']
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 判断是否为重点字段
            is_key = any(k in str(cell.value) for k in key_fields)
            if is_key:
                cell.fill = highlight_fill
            else:
                cell.fill = header_fill
                
        worksheet.freeze_panes = 'B2' # 冻结首行和首列 (A列和第1行)

    def _create_strategy_guide_sheet(self, writer):
        """创建策略指南sheet"""
        strategy_guide = [
            {'策略类型': '平价套利', '英文名称': 'Put-Call Parity', '理论基础': 'C-P=(F-K)*e^(-rT)', '风险等级': '低', '操作说明': '买卖认购认沽+期货对冲'},
            {'策略类型': '垂直套利', '英文名称': 'Vertical Spread', '理论基础': '期权价格单调性', '风险等级': '低', '操作说明': '买低K卖高K(或相反)'},
            {'策略类型': '日历套利', '英文名称': 'Calendar Spread', '理论基础': '时间价值单调性', '风险等级': '低', '操作说明': '卖近月买远月'},
            {'策略类型': '蝶式套利', '英文名称': 'Butterfly Spread', '理论基础': '期权价格凸性', '风险等级': '低', '操作说明': '买K1+K3,卖2*K2'},
            {'策略类型': '盒式套利', '英文名称': 'Box Spread', '理论基础': '牛熊价差组合', '风险等级': '极低', '操作说明': '牛市价差+熊市价差'},
            {'策略类型': '深度实值套利', '英文名称': 'Deep ITM Arbitrage', '理论基础': '市场价<内在价值', '风险等级': '低', '操作说明': '买入期权+对冲标的'},
            {'策略类型': '时间价值低估', '英文名称': 'Time Value Undervalued', '理论基础': '市场价<理论价格', '风险等级': '中', '操作说明': '买入低估期权'},
            {'策略类型': '转换/逆转套利', '英文名称': 'Conversion/Reversal', '理论基础': '合成头寸定价偏离', '风险等级': '低', '操作说明': '期货+期权组合'},
        ]
        df = pd.DataFrame(strategy_guide)
        df.to_excel(writer, sheet_name='策略指南', index=False)
        return df, strategy_guide

    def save_excel_report(self, output_path='wisecoin-期权套利.xlsx'):
        """保存Excel报告"""
        logger.info(f"保存Excel报告: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 汇总页
            all_opportunities = [
                ('平价套利', self.parity_opportunities),
                ('垂直套利', self.vertical_opportunities),
                ('日历套利', self.calendar_opportunities),
                ('蝶式套利', self.butterfly_opportunities),
                ('盒式套利', self.box_opportunities),
                ('深度实值套利', self.deep_itm_opportunities),
                ('时间价值低估', self.time_value_opportunities),
                ('转换逆转套利', self.conversion_opportunities),
            ]
            
            summary_data = {
                '策略类型': [name for name, _ in all_opportunities],
                '机会数量': [len(opps) for _, opps in all_opportunities],
                '最高评分': [max([x.get('可执行性评分', 0) for x in opps], default=0) for _, opps in all_opportunities],
                '最高年化%': [round(max([x.get('年化收益%', 0) for x in opps], default=0), 2) for _, opps in all_opportunities],
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='套利汇总', index=False)
            
            self._create_strategy_guide_sheet(writer)
            
            for sheet_name, opps in all_opportunities:
                if opps:
                    df = pd.DataFrame(opps)
                    
                    # 应用特定的排序逻辑
                    if sheet_name == '垂直套利':
                        if '年化收益%' in df.columns:
                            df = df.sort_values(by='年化收益%', ascending=False)
                    elif sheet_name == '日历套利':
                        if '年化收益%' in df.columns:
                            df = df.sort_values(by='年化收益%', ascending=False)
                    elif sheet_name == '蝶式套利':
                        if '年化收益%' in df.columns:
                            df = df.sort_values(by='年化收益%', ascending=False)
                    elif sheet_name == '深度实值套利':
                        if '合约代码' in df.columns:
                            df = df.sort_values(by='合约代码', ascending=True)
                    elif sheet_name == '时间价值低估':
                        if '合约代码' in df.columns:
                            df = df.sort_values(by='合约代码', ascending=True)
                    elif sheet_name == '转换逆转套利':
                        sort_cols = []
                        sort_orders = []
                        if '标的品种' in df.columns:
                            sort_cols.append('标的品种')
                            sort_orders.append(True)
                        if '行权价' in df.columns:
                            sort_cols.append('行权价')
                            sort_orders.append(False)
                        if sort_cols:
                            df = df.sort_values(by=sort_cols, ascending=sort_orders)
                    
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            
            # 统一设置样式
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                if sheet_name == '套利汇总':
                    self._auto_adjust_column_width(worksheet, summary_df)
                elif sheet_name == '策略指南':
                     # 重新获取指南数据以计算宽度
                    guide_df = pd.DataFrame([
                        {'策略类型': '平价套利', '英文名称': 'Put-Call Parity', '理论基础': 'C-P=(F-K)*e^(-rT)', '风险等级': '低', '操作说明': '买卖认购认沽+期货对冲'},
                        {'策略类型': '垂直套利', '英文名称': 'Vertical Spread', '理论基础': '期权价格单调性', '风险等级': '低', '操作说明': '买低K卖高K(或相反)'},
                        {'策略类型': '日历套利', '英文名称': 'Calendar Spread', '理论基础': '时间价值单调性', '风险等级': '低', '操作说明': '卖近月买远月'},
                        {'策略类型': '蝶式套利', '英文名称': 'Butterfly Spread', '理论基础': '期权价格凸性', '风险等级': '低', '操作说明': '买K1+K3,卖2*K2'},
                        {'策略类型': '盒式套利', '英文名称': 'Box Spread', '理论基础': '牛熊价差组合', '风险等级': '极低', '操作说明': '牛市价差+熊市价差'},
                        {'策略类型': '深度实值套利', '英文名称': 'Deep ITM Arbitrage', '理论基础': '市场价<内在价值', '风险等级': '低', '操作说明': '买入期权+对冲标的'},
                        {'策略类型': '时间价值低估', '英文名称': 'Time Value Undervalued', '理论基础': '市场价<理论价格', '风险等级': '中', '操作说明': '买入低估期权'},
                        {'策略类型': '转换/逆转套利', '英文名称': 'Conversion/Reversal', '理论基础': '合成头寸定价偏离', '风险等级': '低', '操作说明': '期货+期权组合'},
                    ])
                    self._auto_adjust_column_width(worksheet, guide_df)
                else:
                    target_df = None
                    for name, opps in all_opportunities:
                        if name[:31] == sheet_name and opps:
                            target_df = pd.DataFrame(opps)
                            break
                    if target_df is not None:
                         self._auto_adjust_column_width(worksheet, target_df)
                
                self._apply_header_style(worksheet)
        
        logger.info(f"✅ Excel报告已保存: {output_path}")


def print_top_opportunities(opportunities, title, top_n=5):
    """打印Top机会"""
    if not opportunities:
        return
    print(f"\n{'='*70}")
    print(f"📈 {title} - Top {min(top_n, len(opportunities))}")
    print(f"{'='*70}")
    for i, opp in enumerate(opportunities[:top_n], 1):
        exec_status = opp.get('可执行', '是')
        status_icon = '🔴' if exec_status == '否' else '🟢'
        print(f"\n【{i}】{status_icon} {opp.get('套利类型', '')} | 评分: {opp.get('可执行性评分', '-')}")
        print(f"    标的: {opp.get('标的品种', opp.get('标的合约', ''))} | 剩余: {opp.get('剩余天数', '-')}天")
        print(f"    收益: {opp.get('预期收益', opp.get('净利润', 0)):.2f} | 年化: {opp.get('年化收益%', 0):.2f}%")
        print(f"    期货状态: {opp.get('期货状态', '正常')}")
        if opp.get('交易指令'):
            print(f"    交易指令: {opp['交易指令'][:80]}...")
        elif opp.get('操作建议'):
            print(f"    操作建议: {opp['操作建议'][:80]}...")


def main():
    """主函数"""
    print("=" * 70)
    print("期权套利机会综合检测系统 (8种策略)")
    print("=" * 70)
    print("【无风险套利】平价/垂直/日历/蝶式/盒式/转换逆转")
    print("【价值套利】深度实值/时间价值低估")
    print("所有操作基于真实盘口数据 (Ask1/Bid1)")
    print("=" * 70)
    
    detector = IntegratedArbitrageDetector()
    
    if not detector.load_data():
        logger.error("数据加载失败，退出")
        return
    
    detector.run_all_detections()
    
    # 打印各类Top机会
    print_top_opportunities(detector.parity_opportunities, "平价套利", 3)
    print_top_opportunities(detector.vertical_opportunities, "垂直套利", 3)
    print_top_opportunities(detector.calendar_opportunities, "日历套利", 3)
    print_top_opportunities(detector.butterfly_opportunities, "蝶式套利", 3)
    print_top_opportunities(detector.box_opportunities, "盒式套利", 3)
    print_top_opportunities(detector.deep_itm_opportunities, "深度实值套利", 3)
    print_top_opportunities(detector.time_value_opportunities, "时间价值低估", 3)
    print_top_opportunities(detector.conversion_opportunities, "转换/逆转套利", 3)
    
    detector.save_excel_report(os.path.join(TEMP_DIR, "wisecoin-期权套利.xlsx"))
    
    total = (len(detector.parity_opportunities) + len(detector.vertical_opportunities) +
             len(detector.calendar_opportunities) + len(detector.butterfly_opportunities) +
             len(detector.box_opportunities) + len(detector.deep_itm_opportunities) +
             len(detector.time_value_opportunities) + len(detector.conversion_opportunities))
    
    print("\n" + "=" * 70)
    print("✅ 套利分析完成！")
    print(f"📊 发现套利机会共 {total} 个")
    print(f"📁 Excel报告: {os.path.join(TEMP_DIR, 'wisecoin-期权套利.xlsx')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
