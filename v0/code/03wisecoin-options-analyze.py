"""
期权趋势排名分析交易系统 by playbonze
基于期权的隐含波动率、希腊字母、成交量和资金流向进行排名分析
自动选择期权品种进行交易
"""

import asyncio
import logging
import requests
import json
import pandas as pd
import numpy as np
import math
import decimal
import random
import datetime
import time
from datetime import datetime as dt
import sys
import os
import re
import pytz
import threading
import matplotlib.pyplot as plt
import socket
import warnings
import traceback
from tqsdk import TqApi, TqAuth, TqAccount, TqSim, TqBacktest, TqKq, BacktestFinished, TqChan, TargetPosTask, TqNotify
from tqsdk.ta import ATR, BOLL, MACD, OPTION_GREEKS, OPTION_IMPV, BS_VALUE, OPTION_VALUE
from tqsdk.tafunc import ma, ema, ema2, sma, time_to_datetime, crossup, crossdown, time_to_str, get_his_volatility, get_delta, get_gamma, get_vega, get_theta, get_t
from typing import List, Callable
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# 添加外部模块路径以支持 UnifiedLogger
sys.path.append(os.path.join(os.path.dirname(__file__), "wisecoin-catboost"))
from pb_quant_seektop_common import UnifiedLogger

# 设置统一日志
logger = UnifiedLogger.setup_logger_auto(__file__)

# 为了兼容现有代码，创建别名
logger_print = logger


def generate_option_reference():
    """
    读取期权行情、期货行情和交易费用参考，生成 wisecoin-期权参考.xlsx
    """
    logger.info("开始生成期权参考数据...")
    
    OPTION_QUOTE_FILE = "wisecoin-期权行情.xlsx"
    FUTURE_QUOTE_FILE = "wisecoin-期货行情.xlsx"
    PARAM_JSON_FILE = "wisecoin-symbol-params.json"
    OUTPUT_FILE = "wisecoin-期权参考.xlsx"
    
    if not all(os.path.exists(f) for f in [OPTION_QUOTE_FILE, FUTURE_QUOTE_FILE, PARAM_JSON_FILE]):
        logger.warning("缺少必要的输入文件（行情或参数配置），无法生成期权参考。")
        return
        
    try:
        # 1. 读取期货行情和参数配置
        future_xls = pd.ExcelFile(FUTURE_QUOTE_FILE)
        future_df = pd.read_excel(future_xls, sheet_name='Summary')
        future_prices = future_df.set_index('instrument_id')['last_price'].to_dict()
        # 期货合约代码映射到品种名称 (如 CZCE.AP603 -> 苹果2603)
        future_names = future_df.set_index('instrument_id')['instrument_name'].to_dict()
        # 期货合约代码映射到期货合约乘数
        future_multipliers = future_df.set_index('instrument_id')['volume_multiple'].to_dict()
        
        # 加载保证金率配置
        with open(PARAM_JSON_FILE, 'r', encoding='utf-8') as f:
            symbol_params = json.load(f)
        
        # 建立品种到保证金率的映射
        margin_lookup = {}
        for exchange, products in symbol_params.items():
            if exchange.startswith('_'): continue
            for p_code, p_info in products.items():
                if isinstance(p_info, dict) and 'margin_ratio' in p_info:
                    margin_lookup[p_code.upper()] = p_info['margin_ratio'] * 100 # 转为百分比

        # 2. 遍历期权行情并计算
        option_xls = pd.ExcelFile(OPTION_QUOTE_FILE)
        all_option_data = []
        
        for sheet_name in option_xls.sheet_names:
            if sheet_name in ["Summary", "Progress", "Summary_Stats"]: continue
            df = pd.read_excel(option_xls, sheet_name=sheet_name)
            if df.empty: continue
            
            for _, row in df.iterrows():
                symbol = row.get('instrument_id', row.get('symbol'))
                underlying = row.get('underlying_symbol')
                strike = row.get('strike_price', 0)
                opt_type = str(row.get('option_class', row.get('call_or_put', ''))).upper()
                opt_price = row.get('last_price', 0) or row.get('pre_close', 0) or 0
                multiplier = row.get('volume_multiple', 1)
                expire_days = row.get('expire_rest_days', 0) or 1
                expire_date = row.get('expire_datetime', '')
                
                und_price = future_prices.get(underlying, 0)
                if und_price == 0: continue
                
                # 期货相关数据
                future_multiplier = future_multipliers.get(underlying, 1)
                
                # 品种识别
                # 提取产品代码 (如 CZCE.AP603 -> AP)
                if '.' in underlying:
                    p_code = re.sub(r'[^a-zA-Z]', '', underlying.split('.')[1])
                else:
                    p_code = re.sub(r'[^a-zA-Z]', '', underlying)
                
                p_id = p_code.upper()
                margin_ratio = margin_lookup.get(p_id, 15) # 默认15%
                p_name = future_names.get(underlying, '')
                
                # 1. 虚实幅度计算
                if 'CALL' in opt_type or opt_type == 'C':
                    intrinsic_degree = (und_price - strike) / und_price * 100
                    otm_val = max((strike - und_price) * multiplier, 0) 
                    intrinsic_val = max(0, und_price - strike)
                    premium_rate = (strike + opt_price - und_price) / und_price * 100
                else:
                    intrinsic_degree = (strike - und_price) / und_price * 100
                    otm_val = max((und_price - strike) * multiplier, 0)
                    intrinsic_val = max(0, strike - und_price)
                    premium_rate = (und_price + opt_price - strike) / und_price * 100
                
                # 虚实档位分类
                if intrinsic_degree > 20:
                    intrinsic_level = "深度实值"
                elif 10 < intrinsic_degree <= 20:
                    intrinsic_level = "中度实值"
                elif -10 <= intrinsic_degree <= 10:
                    intrinsic_level = "平值附近"
                elif -20 <= intrinsic_degree < -10:
                    intrinsic_level = "中度虚值"
                else: # < -20
                    intrinsic_level = "深度虚值"
                
                # 2. 保证金计算 (新公式)
                # 标的期货保证金 = 标的现价 * 期货合约乘数 * 期货保证金率% / 100
                und_margin = und_price * future_multiplier * (margin_ratio / 100)
                # 保证金 = 期权价*合约乘数+max(标的期货保证金-虚值额/2,标的期货保证金/2)
                margin = (opt_price * multiplier) + max(und_margin - otm_val / 2, und_margin / 2)
                
                # 3. 收益率计算
                profit_rate = (opt_price / und_price * 100) if und_price > 0 else 0
                annual_profit_rate = profit_rate / max(expire_days, 1) * 365
                # 杠杆收益% = 期权价格*合约乘数/保证金
                leverage_profit = (opt_price * multiplier / margin * 100) if margin > 0 else 0
                # 杠杆年化% = 期权价格*合约乘数/剩余天数x365/保证金
                annual_leverage_profit = leverage_profit / max(expire_days, 1) * 365
                
                # 4. 价值分解
                time_value = max(0, opt_price - intrinsic_val)
                time_ratio = (time_value / opt_price * 100) if opt_price > 0 else 0
                
                # 5. 到期日格式化
                # 优化：只保留年月日
                expire_date_str = time_to_str(expire_date).split(' ')[0] if expire_date else ''
                
                # 6. 资金计算 (万)
                opt_fee = opt_price * multiplier
                oi = row.get('open_interest', 0)
                pre_oi = row.get('pre_open_interest', 0)
                amount = row.get('amount', 0)
                
                chendian_wan = (oi * opt_fee / 10000)
                chendian_chg_wan = ((oi - pre_oi) * opt_fee / 10000)
                chengjiao_wan = (amount / 10000)
                zijin_total_wan = chendian_wan + chengjiao_wan

                # 期货杠杆 = 1 / 期货保证金率%
                futures_leverage = round(100.0 / margin_ratio, 2) if margin_ratio > 0 else 0

                # 整理一行数据
                ref_row = {
                    '交易所': row.get('exchange_id', ''),
                    '合约代码': symbol,
                    '合约名称': row.get('instrument_name'),
                    '期权类型': opt_type,
                    '标的合约': underlying,
                    '标的品种名称': p_name,
                    '标的现价': und_price,
                    '期货合约乘数': future_multiplier,
                    '期货保证金率%': margin_ratio,
                    '期货杠杆': futures_leverage,
                    '行权价': strike,
                    '期权价': opt_price,
                    '虚实幅度%': round(intrinsic_degree, 2),
                    '虚实档位': intrinsic_level,
                    '内在价值': round(intrinsic_val, 2),
                    '溢价率%': round(premium_rate, 2),
                    '时间价值': round(time_value, 2),
                    '时间占比%': round(time_ratio, 2),
                    '剩余天数': expire_days,
                    '买方期权费': opt_fee,
                    '标的期货保证金': round(und_margin, 2),
                    '卖方保证金': round(margin, 2),
                    '收益%': round(profit_rate, 4), 
                    '收益年化%': round(annual_profit_rate, 2),
                    '杠杆收益%': round(leverage_profit, 2),
                    '杠杆年化%': round(annual_leverage_profit, 2),
                    '到期日': expire_date_str,
                    '交割年月': f"{int(row.get('exercise_year', 0))}{int(row.get('exercise_month', 0)):02d}" if row.get('exercise_year') else '',
                    '昨收': row.get('pre_close'),
                    '今结': row.get('settlement'),
                    '昨结': row.get('pre_settlement'),
                    '成交量': row.get('volume'),
                    '成交金额': amount,
                    '持仓量': oi,
                    '昨持仓量': pre_oi,
                    '合约乘数': multiplier,
                    '最小跳动': row.get('price_tick'),
                    '沉淀资金(万)': round(chendian_wan, 2),
                    '沉淀资金变化(万)': round(chendian_chg_wan, 2),
                    '成交资金(万)': round(chengjiao_wan, 2),
                    '资金合计(万)': round(zijin_total_wan, 2)
                }
                all_option_data.append(ref_row)
                
        if not all_option_data:
            logger.warning("没有整理出任何期权参考数据。")
            return
            
        ref_df = pd.DataFrame(all_option_data)
        
        # 按照 资金合计(万) 倒序排序
        if '资金合计(万)' in ref_df.columns:
            ref_df = ref_df.sort_values(by='资金合计(万)', ascending=False).reset_index(drop=True)
            
        # 写入 Excel 并进行样式处理
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            ref_df.to_excel(writer, sheet_name='期权参考', index=False)
            ws = writer.sheets['期权参考']
            ws.freeze_panes = 'A2' # 固定表头
            
            # 定义样式
            atm_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # 淡黄色 (平值)
            
            # 实值 (深色系)
            itm_fills = [
                PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"), # 浅红
                PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"), # 中红
                PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"), # 深红
            ]
            # 虚值 (浅色系)
            otm_fills = [
                PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"), # 浅蓝
                PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"), # 中蓝
                PatternFill(start_color="81D4FA", end_color="81D4FA", fill_type="solid"), # 深蓝
            ]
            
            # 到期日颜色
            expire_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            expire_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            expire_dark_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            
            # 获取列索引
            cols = list(ref_df.columns)
            symbol_idx = cols.index('合约代码') + 1
            strike_idx = cols.index('行权价') + 1
            degree_idx = cols.index('虚实幅度%') + 1
            expire_idx = cols.index('到期日') + 1
            days_idx = cols.index('剩余天数') + 1
            
            # 遍历数据行应用样式
            for r_idx, row_data in enumerate(ref_df.itertuples(), start=2):
                degree = row_data[degree_idx]
                days = row_data[days_idx]
                
                # 1. 平值高亮 (虚实幅度绝对值在 2% 以内视为平值)
                if abs(degree) <= 2.0:
                    ws.cell(row=r_idx, column=symbol_idx).fill = atm_fill
                
                # 2. 行权价着色 (实值/虚值)
                if degree > 0: # 实值
                    fill_idx = min(2, int(degree / 5)) # 每 5% 一个深浅档位
                    ws.cell(row=r_idx, column=strike_idx).fill = itm_fills[fill_idx]
                elif degree < 0: # 虚值
                    abs_degree = abs(degree)
                    fill_idx = min(2, int(abs_degree / 5))
                    ws.cell(row=r_idx, column=strike_idx).fill = otm_fills[fill_idx]
                
                # 3. 到期日提醒
                if days <= 1:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_dark_red
                elif days <= 3:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_red
                elif days <= 7:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_yellow
            
            # 自动调整列宽
            def get_visual_width(s):
                if pd.isna(s): return 0
                s = str(s)
                w = 0
                for char in s:
                    if ord(char) > 127: # 中文字符或多字节字符
                        w += 2
                    else:
                        w += 1
                return w

            for idx, col in enumerate(ref_df.columns):
                # 计算表头和内容的视觉长度
                header_w = get_visual_width(col)
                # 为了性能，只对前500行采样计算内容宽度
                content_w = ref_df[col].head(500).astype(str).apply(get_visual_width).max() if not ref_df.empty else 0
                
                # 取大值，并加上少量余量
                max_w = max(header_w, content_w) + 2
                # 设置宽度限制：最小10，最大40
                ws.column_dimensions[get_column_letter(idx + 1)].width = min(max(max_w, 10), 40)
                
        logger.info(f"🚀 期权参考数据生成完成: {OUTPUT_FILE}, 共记录 {len(ref_df)} 条数据。")
        
    except Exception as e:
        logger.error(f"生成期权参考失败: {e}")
        logger.error(traceback.format_exc())

def _classify_option_trading_type(call_oi_change, put_oi_change, pcr, volume_ratio=1.0):
    """
    期权交易类型分类 - 区分「方向型交易」vs「波动率交易」
    
    【方向型期权判断标准】:
    - CALL 与 PUT 明显单边增仓（一方增仓幅度超过另一方的2倍以上）
    - PCR 极端（<0.5 极度看多 或 >1.5 极度看空）
    
    【波动率型期权判断标准】:
    - CALL & PUT 同时增仓（双向增仓，比例接近 0.5-2.0 之间）
    - 成交放大但 PCR 接近 1（0.8-1.2范围）
    
    返回: (交易类型, 类型细分, 置信度)
    """
    # 归一化变化量
    total_oi_change = abs(call_oi_change) + abs(put_oi_change)
    if total_oi_change == 0:
        return ('未知', '无明显变化', 0)
    
    call_ratio = call_oi_change / total_oi_change if total_oi_change > 0 else 0.5
    put_ratio = put_oi_change / total_oi_change if total_oi_change > 0 else 0.5
    
    # 双向增仓判断
    both_increasing = call_oi_change > 0 and put_oi_change > 0
    both_decreasing = call_oi_change < 0 and put_oi_change < 0
    
    # 单边判断阈值
    single_side_threshold = 0.65  # 单边占比超过65%
    
    # PCR极端值判断
    pcr_extreme_bullish = pcr < 0.5
    pcr_extreme_bearish = pcr > 1.5
    pcr_neutral = 0.8 <= pcr <= 1.2
    
    # 分类逻辑
    if both_increasing and pcr_neutral and volume_ratio >= 1.2:
        # 双向增仓 + PCR中性 + 成交放大 -> 波动率交易
        confidence = min(100, 60 + (1 - abs(1 - pcr)) * 40)
        if abs(call_oi_change - put_oi_change) / max(call_oi_change, put_oi_change, 1) < 0.3:
            return ('波动率型', '跨式/宽跨式建仓', confidence)
        else:
            return ('波动率型', '不对称波动率', confidence * 0.8)
    
    elif pcr_extreme_bullish and call_oi_change > put_oi_change:
        # PCR极端看多 + CALL增仓 -> 方向型看多
        confidence = min(100, 70 + (0.5 - pcr) * 60)
        return ('方向型', '强烈看多', confidence)
    
    elif pcr_extreme_bearish and put_oi_change > call_oi_change:
        # PCR极端看空 + PUT增仓 -> 方向型看空
        confidence = min(100, 70 + (pcr - 1.5) * 40)
        return ('方向型', '强烈看空', confidence)
    
    elif abs(call_ratio) > single_side_threshold and call_oi_change > 0:
        # 单边CALL增仓
        confidence = 50 + call_ratio * 50
        return ('方向型', '看多增仓', confidence)
    
    elif abs(put_ratio) > single_side_threshold and put_oi_change > 0:
        # 单边PUT增仓
        confidence = 50 + put_ratio * 50
        return ('方向型', '看空增仓', confidence)
    
    elif both_decreasing:
        # 双向减仓 -> 波动率收敛或获利了结
        if pcr_neutral:
            return ('波动率型', '跨式/宽跨式平仓', 60)
        else:
            return ('方向型', '获利了结', 50)
    
    elif both_increasing:
        # 双向增仓但PCR不中性
        if pcr < 0.8:
            return ('混合型', '偏多波动率', 50)
        elif pcr > 1.2:
            return ('混合型', '偏空波动率', 50)
        else:
            return ('波动率型', '建仓中', 55)
    
    else:
        # 其他情况
        return ('未知', '信号不明确', 30)


def analyze_option_changes():
    """
    期权综合分析系统 - 生成多维度分析报告到 wisecoin-期权排行.xlsx
    
    【分析维度】:
    1. 综合排行 - 多维度加权综合评分
    2. 资金流向 - 沉淀资金、成交资金排行
    3. P/C Ratio分析 - 持仓量/成交量多空情绪
    4. 最大痛点(Max Pain) - 期权到期时对卖方最有利的价位
    5. 隐含波动率分析 - IV偏斜、IV百分位
    6. 流动性分析 - 成交量/持仓量/换手率
    7. 时间价值分析 - 时间价值衰减速度
    8. Greeks敞口 - Delta/Gamma/Vega净敞口
    9. 套利机会 - Put-Call平价偏离度
    10. 交易类型分类 - 方向型 vs 波动率型期权
    """
    logger.info("🚀 开始期权综合分析...")
    
    QUOTE_EXCEL_FILE = "wisecoin-期权行情.xlsx"
    FUTURE_QUOTE_FILE = "wisecoin-期货行情.xlsx"
    OUTPUT_FILE = "wisecoin-期权排行.xlsx"
    
    if not os.path.exists(QUOTE_EXCEL_FILE):
        logger.warning(f"未找到期权行情文件: {QUOTE_EXCEL_FILE}")
        return
    
    try:
        # ============ 1. 数据加载与预处理 ============
        logger.info("加载期权行情数据...")
        xls = pd.ExcelFile(QUOTE_EXCEL_FILE)
        
        # 加载期货行情（用于计算虚实幅度等）
        future_prices = {}
        if os.path.exists(FUTURE_QUOTE_FILE):
            try:
                future_xls = pd.ExcelFile(FUTURE_QUOTE_FILE)
                future_df = pd.read_excel(future_xls, sheet_name='Summary')
                id_col = 'instrument_id' if 'instrument_id' in future_df.columns else 'symbol'
                future_prices = future_df.set_index(id_col)['last_price'].to_dict()
                logger.info(f"加载 {len(future_prices)} 个标的期货价格")
            except Exception as e:
                logger.warning(f"加载期货行情失败: {e}")
        
        # 收集所有期权数据
        all_options = []
        for sheet_name in xls.sheet_names:
            if sheet_name in ["Summary", "Progress", "Summary_Stats"]: 
                continue
            df = pd.read_excel(xls, sheet_name=sheet_name)
            if df.empty:
                continue
            df['_sheet'] = sheet_name
            all_options.append(df)
        
        if not all_options:
            logger.warning("未找到有效期权数据")
            return
            
        options_df = pd.concat(all_options, ignore_index=True)
        logger.info(f"共加载 {len(options_df)} 个期权合约")
        
        # 字段标准化
        col_mapping = {
            'option_class': ['option_class', 'call_or_put'],
            'open_interest': ['open_interest', 'close_oi'],
            'volume': ['volume'],
            'last_price': ['last_price', 'close', 'pre_close'],
            'volume_multiple': ['volume_multiple'],
            'underlying_symbol': ['underlying_symbol', 'product'],
            'strike_price': ['strike_price'],
            'expire_rest_days': ['expire_rest_days'],
            'bid_price1': ['bid_price1'],
            'ask_price1': ['ask_price1'],
            'pre_open_interest': ['pre_open_interest'],
            'instrument_id': ['instrument_id', 'symbol'],
        }
        
        def get_col(df, candidates):
            for c in candidates:
                if c in df.columns:
                    return c
            return None
        
        # 构建标准化数据
        std_data = []
        for _, row in options_df.iterrows():
            try:
                opt_type = str(row.get(get_col(options_df, col_mapping['option_class']) or '', '')).upper()
                if 'CALL' in opt_type or opt_type == 'C':
                    opt_type = 'CALL'
                elif 'PUT' in opt_type or opt_type == 'P':
                    opt_type = 'PUT'
                else:
                    continue  # 跳过无法识别的类型
                
                underlying = row.get(get_col(options_df, col_mapping['underlying_symbol']), 'Unknown')
                if pd.isna(underlying) or underlying == 'Unknown':
                    continue
                    
                last_price = row.get(get_col(options_df, col_mapping['last_price']), 0) or 0
                if pd.isna(last_price) or last_price <= 0:
                    # 尝试使用 pre_close
                    last_price = row.get('pre_close', 0) or 0
                
                multiplier = row.get(get_col(options_df, col_mapping['volume_multiple']), 1) or 1
                oi = row.get(get_col(options_df, col_mapping['open_interest']), 0) or 0
                vol = row.get(get_col(options_df, col_mapping['volume']), 0) or 0
                strike = row.get(get_col(options_df, col_mapping['strike_price']), 0) or 0
                expire_days = row.get(get_col(options_df, col_mapping['expire_rest_days']), 30) or 30
                bid = row.get(get_col(options_df, col_mapping['bid_price1']), 0) or 0
                ask = row.get(get_col(options_df, col_mapping['ask_price1']), 0) or 0
                pre_oi = row.get(get_col(options_df, col_mapping['pre_open_interest']), 0) or 0
                symbol = row.get(get_col(options_df, col_mapping['instrument_id']), '')
                
                # 标的价格
                und_price = future_prices.get(underlying, 0)
                if und_price <= 0:
                    und_price = row.get('underlying_price', 0) or 0
                
                # 计算衍生指标
                # 沉淀资金（亿元）
                chendian = (oi * last_price * multiplier) / 1e8
                # 成交资金（亿元）
                chengjiao = (vol * last_price * multiplier) / 1e8
                # 持仓变化
                oi_change = oi - pre_oi if pre_oi > 0 else 0
                # 换手率
                turnover = (vol / oi * 100) if oi > 0 else 0
                # 买卖价差
                spread = ask - bid if ask > 0 and bid > 0 else 0
                spread_pct = (spread / last_price * 100) if last_price > 0 else 0
                
                # 虚实幅度
                if und_price > 0 and strike > 0:
                    if opt_type == 'CALL':
                        moneyness = (und_price - strike) / und_price * 100
                    else:
                        moneyness = (strike - und_price) / und_price * 100
                else:
                    moneyness = 0
                
                # 内在价值与时间价值
                if und_price > 0:
                    if opt_type == 'CALL':
                        intrinsic = max(0, und_price - strike)
                    else:
                        intrinsic = max(0, strike - und_price)
                    time_value = max(0, last_price - intrinsic)
                else:
                    intrinsic = 0
                    time_value = last_price
                
                time_value_pct = (time_value / last_price * 100) if last_price > 0 else 0
                
                std_data.append({
                    'symbol': symbol,
                    'underlying': underlying,
                    'opt_type': opt_type,
                    'strike': strike,
                    'last_price': last_price,
                    'bid': bid,
                    'ask': ask,
                    'volume': vol,
                    'open_interest': oi,
                    'pre_oi': pre_oi,
                    'oi_change': oi_change,
                    'multiplier': multiplier,
                    'expire_days': expire_days,
                    'und_price': und_price,
                    'chendian': chendian,
                    'chengjiao': chengjiao,
                    'turnover': turnover,
                    'spread': spread,
                    'spread_pct': spread_pct,
                    'moneyness': moneyness,
                    'intrinsic': intrinsic,
                    'time_value': time_value,
                    'time_value_pct': time_value_pct,
                })
            except Exception as e:
                continue
        
        if not std_data:
            logger.warning("标准化后无有效数据")
            return
            
        df = pd.DataFrame(std_data)
        logger.info(f"标准化完成，有效期权: {len(df)} 个")
        
        # ============ 2. 按标的汇总分析 ============
        logger.info("按标的进行多维度分析...")
        
        underlying_analysis = []
        for underlying, group in df.groupby('underlying'):
            calls = group[group['opt_type'] == 'CALL']
            puts = group[group['opt_type'] == 'PUT']
            
            # 基础统计
            total_oi = group['open_interest'].sum()
            total_vol = group['volume'].sum()
            total_chendian = group['chendian'].sum()
            total_chengjiao = group['chengjiao'].sum()
            
            call_oi = calls['open_interest'].sum()
            put_oi = puts['open_interest'].sum()
            call_vol = calls['volume'].sum()
            put_vol = puts['volume'].sum()
            call_chendian = calls['chendian'].sum()
            put_chendian = puts['chendian'].sum()
            
            # P/C Ratio
            pcr_oi = put_oi / call_oi if call_oi > 0 else 0
            pcr_vol = put_vol / call_vol if call_vol > 0 else 0
            pcr_chendian = put_chendian / call_chendian if call_chendian > 0 else 0
            
            # 标的价格
            und_price = group['und_price'].iloc[0] if group['und_price'].iloc[0] > 0 else 0
            
            # 最大痛点 (Max Pain) 计算
            max_pain = _calculate_max_pain(group, und_price)
            max_pain_distance = ((max_pain - und_price) / und_price * 100) if und_price > 0 and max_pain > 0 else 0
            
            # 持仓变化
            total_oi_change = group['oi_change'].sum()
            call_oi_change = calls['oi_change'].sum()
            put_oi_change = puts['oi_change'].sum()
            
            # 平均换手率
            avg_turnover = group['turnover'].mean()
            
            # 平均价差
            avg_spread_pct = group['spread_pct'].mean()
            
            # 平均剩余天数
            avg_expire_days = group['expire_days'].mean()
            
            # 合约数量
            num_contracts = len(group)
            num_calls = len(calls)
            num_puts = len(puts)
            
            # 流动性评分 (基于成交量和换手率)
            liquidity_score = min(100, (total_vol / 1000 + avg_turnover * 2))
            
            # 活跃度评分 (基于持仓变化)
            activity_score = abs(total_oi_change) / max(total_oi, 1) * 100 if total_oi > 0 else 0
            
            # 情绪倾向 (-100 极度看空, 0 中性, 100 极度看多)
            # PCR < 0.7 看多, PCR > 1.0 看空
            if pcr_oi < 0.5:
                sentiment = 80
            elif pcr_oi < 0.7:
                sentiment = 50
            elif pcr_oi < 0.9:
                sentiment = 20
            elif pcr_oi < 1.0:
                sentiment = 0
            elif pcr_oi < 1.2:
                sentiment = -20
            elif pcr_oi < 1.5:
                sentiment = -50
            else:
                sentiment = -80
            
            # 综合评分
            composite_score = (
                liquidity_score * 0.3 +
                activity_score * 0.2 +
                (100 - avg_spread_pct * 10) * 0.2 +  # 价差越小越好
                min(100, total_chendian * 10) * 0.3  # 资金规模
            )
            
            # 交易类型分类 (方向型 vs 波动率型)
            volume_ratio = total_vol / max(group['volume'].mean(), 1) if not group.empty else 1.0
            trading_type, trading_subtype, type_confidence = _classify_option_trading_type(
                call_oi_change, put_oi_change, pcr_oi, volume_ratio
            )
            
            underlying_analysis.append({
                '标的合约': underlying,
                '标的现价': round(und_price, 2),
                '合约数量': num_contracts,
                'CALL数': num_calls,
                'PUT数': num_puts,
                '总持仓量': int(total_oi),
                '总成交量': int(total_vol),
                '沉淀资金(亿)': round(total_chendian, 4),
                '成交资金(亿)': round(total_chengjiao, 4),
                'CALL持仓': int(call_oi),
                'PUT持仓': int(put_oi),
                'PCR(持仓)': round(pcr_oi, 4),
                'PCR(成交)': round(pcr_vol, 4),
                'PCR(资金)': round(pcr_chendian, 4),
                '持仓变化': int(total_oi_change),
                'CALL持仓变化': int(call_oi_change),
                'PUT持仓变化': int(put_oi_change),
                '最大痛点': round(max_pain, 2),
                '痛点距离%': round(max_pain_distance, 2),
                '平均换手率%': round(avg_turnover, 2),
                '平均价差%': round(avg_spread_pct, 4),
                '平均剩余天数': round(avg_expire_days, 1),
                '流动性评分': round(liquidity_score, 1),
                '活跃度评分': round(activity_score, 1),
                '情绪倾向': int(sentiment),
                '综合评分': round(composite_score, 1),
                '交易类型': trading_type,
                '类型细分': trading_subtype,
                '类型置信度': round(type_confidence, 1),
            })
        
        analysis_df = pd.DataFrame(underlying_analysis)
        logger.info(f"完成 {len(analysis_df)} 个标的分析")
        
        # ============ 3. 生成各维度排行 ============
        logger.info("生成多维度排行表...")
        
        ranking_sheets = {}
        
        # 3.1 期权市场
        market_summary = _generate_market_summary(df, analysis_df)
        ranking_sheets['期权市场'] = market_summary
        
        # 3.2 期权排行
        comprehensive_df = analysis_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
        comprehensive_df.insert(0, '排名', range(1, len(comprehensive_df) + 1))
        ranking_sheets['期权排行'] = comprehensive_df
        
        # 3.3 期权痛点
        maxpain_df = analysis_df[['标的合约', '标的现价', '最大痛点', '痛点距离%',
                                   '总持仓量', 'PCR(持仓)', '平均剩余天数']].copy()
        maxpain_df = maxpain_df[maxpain_df['最大痛点'] > 0]  # 过滤无效数据
        maxpain_df = maxpain_df.sort_values('总持仓量', ascending=False).reset_index(drop=True)
        maxpain_df.insert(0, '排名', range(1, len(maxpain_df) + 1))
        # 增加到期方向预测
        maxpain_df['到期方向'] = maxpain_df['痛点距离%'].apply(
            lambda x: '上涨概率大' if x > 2 else ('下跌概率大' if x < -2 else '震荡')
        )
        ranking_sheets['期权痛点'] = maxpain_df

        # 3.4 期权PCR
        pcr_df = analysis_df[['标的合约', '标的现价', 'CALL持仓', 'PUT持仓', 
                              'PCR(持仓)', 'PCR(成交)', 'PCR(资金)', '情绪倾向']].copy()
        pcr_df['持仓合计'] = pcr_df['CALL持仓'] + pcr_df['PUT持仓']
        pcr_df = pcr_df.sort_values(['持仓合计', '标的合约'], ascending=[False, True]).reset_index(drop=True)
        pcr_df.insert(0, '排名', range(1, len(pcr_df) + 1))
        # 增加情绪解读
        def interpret_pcr(pcr):
            if pcr < 0.5: return '极度看多'
            elif pcr < 0.7: return '看多'
            elif pcr < 0.9: return '偏多'
            elif pcr < 1.1: return '中性'
            elif pcr < 1.3: return '偏空'
            elif pcr < 1.5: return '看空'
            else: return '极度看空'
        pcr_df['情绪解读'] = pcr_df['PCR(持仓)'].apply(interpret_pcr)
        ranking_sheets['期权PCR'] = pcr_df

        # 3.5 期权看多
        bullish_df = analysis_df[analysis_df['PCR(持仓)'] < 0.8].copy()
        bullish_df = bullish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
        if not bullish_df.empty:
            bullish_df.insert(0, '排名', range(1, len(bullish_df) + 1))
            bullish_df = bullish_df[['排名', '标的合约', '标的现价', 'PCR(持仓)', 'PCR(成交)',
                                      'CALL持仓', 'PUT持仓', '沉淀资金(亿)', '情绪倾向']]
            ranking_sheets['期权看多'] = bullish_df
        
        # 3.6 期权看空
        bearish_df = analysis_df[analysis_df['PCR(持仓)'] > 1.2].copy()
        bearish_df = bearish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
        if not bearish_df.empty:
            bearish_df.insert(0, '排名', range(1, len(bearish_df) + 1))
            bearish_df = bearish_df[['排名', '标的合约', '标的现价', 'PCR(持仓)', 'PCR(成交)',
                                      'CALL持仓', 'PUT持仓', '沉淀资金(亿)', '情绪倾向']]
            ranking_sheets['期权看空'] = bearish_df
        
        # 3.7 期权临期
        near_expire_df = analysis_df[analysis_df['平均剩余天数'] <= 14].copy()
        near_expire_df = near_expire_df.sort_values(['平均剩余天数', '总持仓量'], ascending=[True, False]).reset_index(drop=True)
        if not near_expire_df.empty:
            near_expire_df.insert(0, '排名', range(1, len(near_expire_df) + 1))
            near_expire_df = near_expire_df[['排名', '标的合约', '标的现价', '平均剩余天数',
                                              '最大痛点', '痛点距离%', '总持仓量', 'PCR(持仓)']]
            ranking_sheets['期权临期'] = near_expire_df

        # 3.8 期权资金
        capital_df = analysis_df[['标的合约', '标的现价', '沉淀资金(亿)', '成交资金(亿)', 
                                   'CALL持仓', 'PUT持仓', 'PCR(资金)', '持仓变化']].copy()
        capital_df = capital_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
        capital_df.insert(0, '排名', range(1, len(capital_df) + 1))
        # 增加资金流入方向判断
        capital_df['资金方向'] = capital_df['持仓变化'].apply(lambda x: '流入' if x > 0 else ('流出' if x < -0 else '持平'))
        ranking_sheets['期权资金'] = capital_df
        
        # 3.9 期权流动性
        liquidity_df = analysis_df[['标的合约', '总成交量', '总持仓量', '平均换手率%',
                                     '平均价差%', '流动性评分', '合约数量']].copy()
        liquidity_df['活跃合计'] = liquidity_df['总成交量'] + liquidity_df['总持仓量']
        liquidity_df = liquidity_df.sort_values('活跃合计', ascending=False).reset_index(drop=True)
        liquidity_df.insert(0, '排名', range(1, len(liquidity_df) + 1))
        # 流动性等级
        def liquidity_grade(score):
            if score >= 80: return 'A (极佳)'
            elif score >= 60: return 'B (良好)'
            elif score >= 40: return 'C (一般)'
            elif score >= 20: return 'D (较差)'
            else: return 'E (极差)'
        liquidity_df['流动性等级'] = liquidity_df['流动性评分'].apply(liquidity_grade)
        ranking_sheets['期权流动性'] = liquidity_df
        
        # 3.10 期权活跃度
        activity_df = analysis_df[['标的合约', '持仓变化', 'CALL持仓变化', 'PUT持仓变化',
                                    '总成交量', '活跃度评分', 'PCR(持仓)']].copy()
        activity_df = activity_df.sort_values('总成交量', ascending=False).reset_index(drop=True)
        activity_df.insert(0, '排名', range(1, len(activity_df) + 1))
        # 活跃方向
        def activity_direction(row):
            call_chg = row['CALL持仓变化']
            put_chg = row['PUT持仓变化']
            if call_chg > 0 and put_chg > 0:
                return '双向增仓'
            elif call_chg < 0 and put_chg < 0:
                return '双向减仓'
            elif call_chg > put_chg:
                return '看多增仓'
            elif put_chg > call_chg:
                return '看空增仓'
            else:
                return '持平'
        activity_df['活跃方向'] = activity_df.apply(activity_direction, axis=1)
        ranking_sheets['期权活跃度'] = activity_df
        
        # 3.11 方向型期权 - CALL/PUT明显单边增仓, PCR极端
        directional_df = analysis_df[analysis_df['交易类型'] == '方向型'].copy()
        if not directional_df.empty:
            directional_df = directional_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            directional_df.insert(0, '排名', range(1, len(directional_df) + 1))
            directional_df = directional_df[['排名', '标的合约', '标的现价', '类型细分', '类型置信度',
                                              'PCR(持仓)', 'CALL持仓变化', 'PUT持仓变化', 
                                              '沉淀资金(亿)', '情绪倾向', '综合评分']]
            # 策略建议
            def directional_strategy(row):
                if '看多' in row['类型细分']:
                    return '买CALL / 卖PUT'
                elif '看空' in row['类型细分']:
                    return '买PUT / 卖CALL'
                else:
                    return '趋势跟随'
            directional_df['策略建议'] = directional_df.apply(directional_strategy, axis=1)
            ranking_sheets['方向型期权'] = directional_df
        
        # 3.12 波动率型期权 - CALL & PUT同时增仓, PCR接近1
        volatility_df = analysis_df[analysis_df['交易类型'] == '波动率型'].copy()
        if not volatility_df.empty:
            volatility_df = volatility_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            volatility_df.insert(0, '排名', range(1, len(volatility_df) + 1))
            volatility_df = volatility_df[['排名', '标的合约', '标的现价', '类型细分', '类型置信度',
                                            'PCR(持仓)', 'CALL持仓变化', 'PUT持仓变化',
                                            '沉淀资金(亿)', '平均剩余天数', '综合评分']]
            # 策略建议
            def volatility_strategy(row):
                subtype = row['类型细分']
                if '建仓' in subtype:
                    return '跨式/宽跨式 (做多波动率)'
                elif '平仓' in subtype:
                    return '卖跨式 (做空波动率)'
                else:
                    return '波动率套利'
            volatility_df['策略建议'] = volatility_df.apply(volatility_strategy, axis=1)
            ranking_sheets['波动率型期权'] = volatility_df
        
        # ============ 4. 保存到 Excel ============
        logger.info(f"保存分析结果到 {OUTPUT_FILE}...")
        _save_ranking_excel(OUTPUT_FILE, ranking_sheets)
        
        logger.info("="*60)
        logger.info("期权综合分析完成")
        logger.info(f"📊 分析报告: {OUTPUT_FILE}")
        logger.info(f"📈 包含 {len(ranking_sheets)} 个分析维度")
        logger.info(f"📉 覆盖 {len(analysis_df)} 个标的合约")
        logger.info("="*60)
        
    except Exception as e:
        logger.error(f"期权分析失败: {e}")
        logger.error(traceback.format_exc())


def _calculate_max_pain(group, und_price):
    """
    计算最大痛点 (Max Pain)
    
    最大痛点是期权到期时，使所有期权买方损失最大（即卖方收益最大）的标的价格。
    计算方法：遍历所有行权价，计算在该价格到期时所有期权的内在价值总和，
    取内在价值总和最小的行权价即为最大痛点。
    """
    if group.empty or und_price <= 0:
        return 0
    
    strikes = group['strike'].unique()
    strikes = [s for s in strikes if s > 0]
    
    if not strikes:
        return 0
    
    min_pain = float('inf')
    max_pain_strike = 0
    
    for test_price in strikes:
        total_pain = 0
        for _, opt in group.iterrows():
            strike = opt['strike']
            oi = opt['open_interest']
            multiplier = opt['multiplier']
            opt_type = opt['opt_type']
            
            if opt_type == 'CALL':
                # Call内在价值 = max(0, 到期价 - 行权价)
                itm = max(0, test_price - strike)
            else:
                # Put内在价值 = max(0, 行权价 - 到期价)
                itm = max(0, strike - test_price)
            
            # 期权买方的价值 = 持仓量 * 内在价值 * 乘数
            total_pain += oi * itm * multiplier
        
        if total_pain < min_pain:
            min_pain = total_pain
            max_pain_strike = test_price
    
    return max_pain_strike


def _generate_market_summary(df, analysis_df):
    """生成市场概览表"""
    summary_data = []
    
    # 整体统计
    total_contracts = len(df)
    total_underlyings = len(analysis_df)
    total_oi = df['open_interest'].sum()
    total_vol = df['volume'].sum()
    total_chendian = analysis_df['沉淀资金(亿)'].sum()
    total_chengjiao = analysis_df['成交资金(亿)'].sum()
    
    calls = df[df['opt_type'] == 'CALL']
    puts = df[df['opt_type'] == 'PUT']
    
    overall_pcr_oi = puts['open_interest'].sum() / calls['open_interest'].sum() if calls['open_interest'].sum() > 0 else 0
    overall_pcr_vol = puts['volume'].sum() / calls['volume'].sum() if calls['volume'].sum() > 0 else 0
    
    # 情绪分布
    bullish_count = len(analysis_df[analysis_df['PCR(持仓)'] < 0.8])
    neutral_count = len(analysis_df[(analysis_df['PCR(持仓)'] >= 0.8) & (analysis_df['PCR(持仓)'] <= 1.2)])
    bearish_count = len(analysis_df[analysis_df['PCR(持仓)'] > 1.2])
    
    summary_data = [
        {'指标': '期权合约总数', '数值': total_contracts, '说明': '所有交易中的期权合约数量'},
        {'指标': '标的品种数', '数值': total_underlyings, '说明': '有期权交易的标的合约数量'},
        {'指标': '市场总持仓(万手)', '数值': round(total_oi / 10000, 2), '说明': '所有期权持仓量合计'},
        {'指标': '市场总成交(万手)', '数值': round(total_vol / 10000, 2), '说明': '所有期权成交量合计'},
        {'指标': '沉淀资金(亿)', '数值': round(total_chendian, 2), '说明': '持仓量×价格×乘数'},
        {'指标': '成交资金(亿)', '数值': round(total_chengjiao, 2), '说明': '成交量×价格×乘数'},
        {'指标': '整体PCR(持仓)', '数值': round(overall_pcr_oi, 4), '说明': 'PUT/CALL持仓比'},
        {'指标': '整体PCR(成交)', '数值': round(overall_pcr_vol, 4), '说明': 'PUT/CALL成交比'},
        {'指标': '看多品种数', '数值': bullish_count, '说明': 'PCR < 0.8 的品种'},
        {'指标': '中性品种数', '数值': neutral_count, '说明': '0.8 ≤ PCR ≤ 1.2 的品种'},
        {'指标': '看空品种数', '数值': bearish_count, '说明': 'PCR > 1.2 的品种'},
    ]
    
    # 市场情绪解读
    if overall_pcr_oi < 0.7:
        market_sentiment = '市场整体偏多'
    elif overall_pcr_oi < 1.0:
        market_sentiment = '市场情绪中性偏多'
    elif overall_pcr_oi < 1.3:
        market_sentiment = '市场情绪中性偏空'
    else:
        market_sentiment = '市场整体偏空'
    
    summary_data.append({'指标': '市场情绪', '数值': market_sentiment, '说明': f'基于整体PCR={overall_pcr_oi:.2f}'})
    
    # 交易类型统计
    if '交易类型' in analysis_df.columns:
        directional_count = len(analysis_df[analysis_df['交易类型'] == '方向型'])
        volatility_count = len(analysis_df[analysis_df['交易类型'] == '波动率型'])
        mixed_count = len(analysis_df[analysis_df['交易类型'] == '混合型'])
        summary_data.append({'指标': '方向型品种', '数值': directional_count, '说明': 'CALL/PUT单边增仓或PCR极端'})
        summary_data.append({'指标': '波动率型品种', '数值': volatility_count, '说明': '双向增仓且PCR接近中性'})
        summary_data.append({'指标': '混合型品种', '数值': mixed_count, '说明': '特征不明确的品种'})
    
    # 资金Top5
    top5_capital = analysis_df.nlargest(5, '沉淀资金(亿)')['标的合约'].tolist()
    summary_data.append({'指标': '资金TOP5', '数值': ', '.join(top5_capital), '说明': '沉淀资金最大的5个品种'})
    
    # 活跃Top5
    top5_active = analysis_df.nlargest(5, '活跃度评分')['标的合约'].tolist()
    summary_data.append({'指标': '活跃TOP5', '数值': ', '.join(top5_active), '说明': '持仓变化最活跃的5个品种'})
    
    return pd.DataFrame(summary_data)


def _save_ranking_excel(file_path, sheets_dict):
    """保存排行数据到Excel并美化格式"""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_dict.items():
                if df is None or df.empty:
                    continue
                    
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                
                # 表头样式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 自动列宽
                for idx, col in enumerate(df.columns):
                    max_len = max(
                        df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                        len(str(col))
                    ) + 2
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)
                
                # 条件格式化
                _apply_conditional_formatting(ws, df, sheet_name)
                
        logger.info(f"✅ Excel保存成功: {file_path}")
    except Exception as e:
        logger.error(f"保存Excel失败: {e}")
        logger.error(traceback.format_exc())


def _apply_conditional_formatting(ws, df, sheet_name):
    """应用条件格式化"""
    # 正值绿色，负值红色的列
    red_green_cols = ['持仓变化', 'CALL持仓变化', 'PUT持仓变化', '痛点距离%', '情绪倾向']
    # 评分类 - 高分绿色
    score_cols = ['综合评分', '流动性评分', '活跃度评分', '类型置信度']
    # PCR类 - 低值绿色高值红色
    pcr_cols = ['PCR(持仓)', 'PCR(成交)', 'PCR(资金)']
    # 交易类型列
    trading_type_cols = ['交易类型', '类型细分']
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    purple_fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
    
    cols = list(df.columns)
    
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for col_name in cols:
            if col_name not in df.columns:
                continue
            col_idx = cols.index(col_name) + 1
            val = getattr(row, col_name.replace(' ', '_').replace('(', '').replace(')', '').replace('%', ''), None)
            
            if val is None:
                continue
            
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue
            
            cell = ws.cell(row=r_idx, column=col_idx)
            
            if col_name in red_green_cols:
                if val > 0:
                    cell.fill = green_fill
                elif val < 0:
                    cell.fill = red_fill
            elif col_name in score_cols:
                if val >= 70:
                    cell.fill = green_fill
                elif val >= 40:
                    cell.fill = yellow_fill
                elif val < 40:
                    cell.fill = red_fill
            elif col_name in pcr_cols:
                if val < 0.8:
                    cell.fill = green_fill
                elif val > 1.2:
                    cell.fill = red_fill
                else:
                    cell.fill = yellow_fill
            elif col_name in trading_type_cols:
                str_val = str(val)
                if '方向' in str_val or '看多' in str_val or '看空' in str_val:
                    cell.fill = red_fill
                elif '波动率' in str_val or '跨式' in str_val:
                    cell.fill = blue_fill
                elif '混合' in str_val:
                    cell.fill = purple_fill



# 运行模式
RUN_MODE = 2
if os.path.basename(__file__) in ['pb-quant-test-1.py', 'pb-quant-test-2.py']:
    RUN_MODE = 1
elif os.path.basename(__file__) in ['pb-quant-kq.py']:
    RUN_MODE = 2
elif os.path.basename(__file__) in ['pb-quant-bh.py']:
    RUN_MODE = 4
elif os.path.basename(__file__) in ['pb-quant-jx.py']:
    RUN_MODE = 6

if RUN_MODE == 1:
    backtest_start_dt, backtest_end_dt = (datetime.date.today() + datetime.timedelta(days=-1), datetime.date.today() + datetime.timedelta(days=0))
    acc_sim = TqSim(init_balance=10000000)
    api = TqApi(account=acc_sim, backtest=TqBacktest(start_dt=backtest_start_dt,
                end_dt=backtest_end_dt), debug=False, web_gui=False, auth=TqAuth('playbonze', 'abC!@#123'))
elif RUN_MODE == 2:
    api = TqApi(TqKq(), debug=False, web_gui=False,
                auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 3:
    api = TqApi(TqAccount('simnow', '207302', 'Bonze!0613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 4:
    api = TqApi(TqAccount('B渤海期货', '98908572', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 5:
    api = TqApi(TqAccount('H华安期货', '100919200', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 6:
    api = TqApi(TqAccount('J金信期货', '80016087', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('playbonze', 'bonze13'))
elif RUN_MODE == 7:
    api = TqApi(TqAccount('D东吴期货', '526178061', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))
elif RUN_MODE == 8:
    api = TqApi(TqAccount('H宏源期货', '901212925', 'bonze613'), debug=False,
                web_gui=False, auth=TqAuth('huaying', 'bonze13'))

logger.info('模式%d，期权策略开始运行。' % RUN_MODE)

# 期权策略参数
try:
    # 创建期权分析任务
    async def run_sequence():
        analyze_option_changes()
        generate_option_reference()
        logger.info("所有期权分析处理完成。")
        return

    api.create_task(run_sequence())

    while True:
        api.wait_update()
        
except BacktestFinished:
    api.close()
    print('\n期权回测完成。')
except Exception as e:
    logger_print.info(f'{repr(e)}，line {sys._getframe().f_lineno}。')
    logger_print.info(traceback.format_exc())
