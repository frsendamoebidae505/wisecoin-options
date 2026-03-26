import sys
import os
import pandas as pd
import numpy as np
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QComboBox, QPushButton, QGroupBox, 
                             QGridLayout, QSplitter, QMessageBox)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QBrush, QFont
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import griddata
from scipy.stats import kurtosis, skew
import platform
import re
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import io

# 配置matplotlib中文显示
if platform.system() == 'Darwin':
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
elif platform.system() == 'Windows':
    plt.rcParams['font.sans-serif'] = ['SimHei']
else:
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


class OptionTShapeWindow(QMainWindow):
    """期权T型报价窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("WiseCoin 期权")
        self.setGeometry(50, 50, 1800, 1000)
        
        # 数据存储
        self.market_overview_df = None
        self.option_ref_df = None
        self.klines_data = {}  # 期货K线数据 {symbol: DataFrame}
        self.contract_list = []  # 合约列表（标的+交割月）
        
        # 初始化UI
        self.init_ui()
        
        # 加载数据
        self.load_data()
    
    def init_ui(self):
        """初始化UI界面"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # ========== 顶部控制区 ==========
        controls_layout = QHBoxLayout()
        
        controls_layout.addWidget(QLabel("期权合约:"))
        
        # 上一个按钮
        self.prev_button = QPushButton("◀ 上一个")
        self.prev_button.setFixedWidth(80)
        self.prev_button.clicked.connect(self.on_prev_contract)
        controls_layout.addWidget(self.prev_button)
        
        self.contract_combo = QComboBox()
        self.contract_combo.setMinimumWidth(250)
        self.contract_combo.currentTextChanged.connect(self.on_contract_changed)
        controls_layout.addWidget(self.contract_combo)
        
        # 下一个按钮
        self.next_button = QPushButton("下一个 ▶")
        self.next_button.setFixedWidth(80)
        self.next_button.clicked.connect(self.on_next_contract)
        controls_layout.addWidget(self.next_button)
        
        controls_layout.addSpacing(20)
        
        self.refresh_button = QPushButton("🔄 刷新数据")
        self.refresh_button.clicked.connect(self.load_data)
        controls_layout.addWidget(self.refresh_button)
        
        
        self.export_important_button = QPushButton("📊 导出重要期权")
        self.export_important_button.clicked.connect(self.export_important_contracts)
        controls_layout.addWidget(self.export_important_button)
        
        controls_layout.addStretch()
        
        self.status_label = QLabel("状态：未加载")
        self.status_label.setStyleSheet("color: gray; font-weight: bold;")
        controls_layout.addWidget(self.status_label)
        
        main_layout.addLayout(controls_layout)
        
        # ========== 上方统计信息区（紧凑布局）==========
        stats_layout = QHBoxLayout()
        
        # 1. 标的信息（精简版）
        self.underlying_stats_group = QGroupBox("标的信息")
        self.underlying_stats_grid = QGridLayout(self.underlying_stats_group)
        self.underlying_stats_grid.setHorizontalSpacing(10)
        self.underlying_stats_grid.setVerticalSpacing(3)
        self.underlying_stat_labels = {}
        
        underlying_fields = [
            ("标的合约", "标的合约"),
            ("期货现价", "期货现价"),
            ("杠杆涨跌%", "杠杆涨跌%"),
            ("期货沉淀(亿)", "期货沉淀(亿)"),
            ("期货状态", "期货状态"),
            ("期货方向", "期货方向"),
            ("期货流向", "期货流向"),
        ]
        
        for idx, (label_text, key) in enumerate(underlying_fields):
            row = idx % 4
            col = (idx // 4) * 2
            
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet("font-size: 11pt; color: #0066cc;")
            
            self.underlying_stats_grid.addWidget(label, row, col)
            self.underlying_stats_grid.addWidget(value_label, row, col + 1)
            self.underlying_stat_labels[key] = value_label
        
        stats_layout.addWidget(self.underlying_stats_group)
        
        # 2. 期权信息（精简版）
        self.option_stats_group = QGroupBox("期权信息")
        self.option_stats_grid = QGridLayout(self.option_stats_group)
        self.option_stats_grid.setHorizontalSpacing(10)
        self.option_stats_grid.setVerticalSpacing(3)
        self.option_stat_labels = {}
        
        option_fields = [
            ("期权结构", "期权结构"),
            ("期权情绪", "期权情绪"),
            ("期权PCR", "期权PCR"),
            ("期权沉淀(亿)", "期权沉淀(亿)"),
            ("最大痛点", "最大痛点"),
            ("痛点距离%", "痛点距离%"),
            ("联动状态", "联动状态"),
            ("共振评分", "共振评分"),
        ]
        
        for idx, (label_text, key) in enumerate(option_fields):
            row = idx % 4
            col = (idx // 4) * 2
            
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet("font-size: 11pt; color: #cc6600;")
            
            self.option_stats_grid.addWidget(label, row, col)
            self.option_stats_grid.addWidget(value_label, row, col + 1)
            self.option_stat_labels[key] = value_label
        
        stats_layout.addWidget(self.option_stats_group)
        stats_layout.setStretch(0, 1)
        stats_layout.setStretch(1, 1) # 调整比例
        
        main_layout.addLayout(stats_layout)
        
        # ========== 3. 联动分析（核心分析 + 波动率曲面信息）==========
        linkage_group = QGroupBox("联动分析")
        linkage_layout = QGridLayout(linkage_group)
        linkage_layout.setHorizontalSpacing(8)
        linkage_layout.setVerticalSpacing(4)
        self.linkage_labels = {}
        
        # 优化布局：
        # 列0: 评分体系 | 列1: 波动结构 | 列2: 偏度特征 | 列3: 策略建议
        # 市场解读、波曲策略、适合策略、不适合策略放在同一行
        linkage_fields = [
            # 第一行
            ("共振等级", "共振等级", 0, 0, 1, "#990099"),
            ("期限结构", "期限结构", 0, 2, 1, "#0066cc"),  # From Vol Surface
            ("倾斜方向", "倾斜方向", 0, 4, 1, "#006666"),  # From Vol Surface
            ("共振标签", "共振标签", 0, 6, 1, "#666600"),

            # 第二行
            ("联动总分", "联动总分", 1, 0, 1, "#990099"),
            ("IV/RV比率", "IV/RV比率", 1, 2, 1, "#cc6600"), # From Vol Surface
            ("期限结构差", "期限结构差", 1, 4, 1, "#0066cc"), # From Vol Surface
            ("IV倾斜度", "IV倾斜度", 1, 6, 1, "#006666"),  # From Vol Surface
            
            # 第三行
            ("价格评分", "价格评分", 2, 0, 1, "#990099"),
            ("峰度", "峰度", 2, 2, 1, "#cc6600"),        # From Vol Surface
            ("短期IV", "短期IV", 2, 4, 1, "#0066cc"),      # From Vol Surface
            ("虚值认沽IV", "虚值认沽IV均值", 2, 6, 1, "#006666"), # From Vol Surface
            ("沉淀合计", "沉淀资金合计(亿)", 2, 8, 1, "#666600"),

            # 第四行
            ("情绪评分", "情绪评分", 3, 0, 1, "#990099"),
            ("偏度", "偏度", 3, 2, 1, "#cc6600"),        # From Vol Surface
            ("长期IV", "长期IV", 3, 4, 1, "#0066cc"),      # From Vol Surface
            ("虚值认购IV", "虚值认购IV均值", 3, 6, 1, "#006666"), # From Vol Surface
            ("合约数量", "合约数量", 3, 8, 1, "#666600"),    # From Vol Surface

            # 第五行：策略信息（同一行，对齐前四列，留空第五列）
            ("市场解读", "市场解读", 4, 0, 1, "#FF0000"),
            ("波曲策略", "推荐策略", 4, 2, 1, "#FF0000"), # From Vol Surface
            ("适合策略", "适合策略", 4, 4, 1, "#FF0000"),
            ("不适合策略", "不适合策略", 4, 6, 1, "#008800"),
        ]
        
        for label_text, key, row, col, colspan, color in linkage_fields:
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet(f"font-size: 11pt; color: {color}; font-weight: bold;")
            value_label.setWordWrap(True)
            
            linkage_layout.addWidget(label, row, col)
            linkage_layout.addWidget(value_label, row, col + 1, 1, colspan)
            self.linkage_labels[key] = value_label
        
        # 设置列宽比例
        for i in range(10):
            linkage_layout.setColumnStretch(i, 1)
        
        main_layout.addWidget(linkage_group)
        
        # ========== 中间：期权T型报价表（全字段对称显示）==========
        t_group = QGroupBox("T型报价")
        t_layout = QVBoxLayout(t_group)
        t_layout.setContentsMargins(1, 1, 1, 1)
        t_layout.setSpacing(0)
        
        self.option_t_table = QTableWidget()
        
        # 定义所有显示字段（期权参考的全字段）
        # 排除行权价(中心列), 标的合约/交割年月(表头选择), 交易所/期权类型等
        self.all_ref_cols = [
            '合约代码', '期权价', '涨跌幅%', '成交量', '成交金额', 
            '持仓量', '昨持仓量', '合约乘数', '最小跳动', '沉淀资金(万)', 
            '成交资金(万)', '理论价格', '近期波动率', '隐含波动率', 'Delta', 
            'Gamma', 'Theta', 'Vega', 'Rho', '虚实幅度%', '虚实档位', '内在价值', '溢价率%', 
            '时间价值', '时间占比%', '买方期权费', '买方杠杆', '标的期货保证金', '卖方保证金', 
            '收益%', '收益年化%', '杠杆收益%', '杠杆年化%', '期货保证金率%', '期货杠杆'
        ]
        
        # 左右对称显示这些字段
        put_headers = [f"P-{f}" for f in self.all_ref_cols]
        call_headers = [f"C-{f}" for f in self.all_ref_cols]
        headers = put_headers + ['行权价'] + call_headers
        
        self.option_t_table.setColumnCount(len(headers))
        self.option_t_table.setHorizontalHeaderLabels(headers)
        
        self.option_t_table.setStyleSheet("""
            QTableWidget {
                border: 2px solid #999999;
                gridline-color: #cccccc;
                font-size: 10pt;
                background-color: white;
            }
            QTableWidget::item {
                padding: 1px 2px;
                border-right: 1px solid #e0e0e0;
            }
            QHeaderView::section {
                background-color: #4a4a4a;
                color: white;
                padding: 2px 3px;
                border: 1px solid #666666;
                font-weight: bold;
                font-size: 9pt;
            }
            QTableWidget::item:selected {
                background-color: #b3d9ff;
            }
            
            /* 垂直滚动条 - 金黄色风格 */
            QScrollBar:vertical {
                border: none;
                background: #f5f5f5;
                width: 14px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #FFD700, stop:1 #FFB900);
                min-height: 30px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:vertical:hover {
                background: #FFA500;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            
            /* 水平滚动条 - 金黄色风格 */
            QScrollBar:horizontal {
                border: none;
                background: #f5f5f5;
                height: 14px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #FFD700, stop:1 #FFB900);
                min-width: 30px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #FFA500;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
        """)
        
        # 设置列宽
        header = self.option_t_table.horizontalHeader()
        for i in range(len(headers)):
            if i == len(self.all_ref_cols):  # 行权价列
                header.setSectionResizeMode(i, QHeaderView.Fixed)
                self.option_t_table.setColumnWidth(i, 90)
            else:
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        
        self.option_t_table.verticalHeader().setVisible(False)
        self.option_t_table.verticalHeader().setDefaultSectionSize(22)  # 设置紧凑行高
        self.option_t_table.setAlternatingRowColors(True)
        self.option_t_table.setSortingEnabled(False)
        self.option_t_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.option_t_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        t_layout.addWidget(self.option_t_table)
        main_layout.addWidget(t_group, stretch=3)
        
        # ========== 下方：图表区（K线 + 波动率曲面 + 微笑曲线）==========
        charts_splitter = QSplitter(Qt.Horizontal)
        
        # 标的K线图（左边1/3）
        kline_group = QGroupBox("标的K线")
        kline_layout = QVBoxLayout(kline_group)
        kline_layout.setContentsMargins(2, 2, 2, 2)
        self.kline_fig = Figure(figsize=(5, 4), dpi=90)
        self.kline_canvas = FigureCanvas(self.kline_fig)
        kline_layout.addWidget(self.kline_canvas)
        
        # 波动率曲面（中间1/3）
        surface_group = QGroupBox("波动率曲面")
        surface_layout = QVBoxLayout(surface_group)
        surface_layout.setContentsMargins(2, 2, 2, 2)
        self.surface_fig = Figure(figsize=(5, 4), dpi=90)
        self.surface_canvas = FigureCanvas(self.surface_fig)
        surface_layout.addWidget(self.surface_canvas)
        
        # 微笑曲线（右边1/3）
        smile_group = QGroupBox("微笑曲线")
        smile_layout = QVBoxLayout(smile_group)
        smile_layout.setContentsMargins(2, 2, 2, 2)
        self.smile_fig = Figure(figsize=(5, 4), dpi=90)
        self.smile_canvas = FigureCanvas(self.smile_fig)
        smile_layout.addWidget(self.smile_canvas)
        
        charts_splitter.addWidget(kline_group)
        charts_splitter.addWidget(surface_group)
        charts_splitter.addWidget(smile_group)
        charts_splitter.setStretchFactor(0, 1)  # K线 1/3
        charts_splitter.setStretchFactor(1, 1)  # 波动率曲面 1/3
        charts_splitter.setStretchFactor(2, 1)  # 微笑曲线 1/3
        
        main_layout.addWidget(charts_splitter, stretch=2)
    
    def load_data(self):
        """加载Excel数据"""
        try:
            self.status_label.setText("状态：加载中...")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            QApplication.processEvents()
            
            # 读取市场概览
            market_file = 'wisecoin-市场概览.xlsx'
            if not os.path.exists(market_file):
                raise FileNotFoundError(f"未找到文件: {market_file}")
            self.market_overview_df = pd.read_excel(market_file, sheet_name='货权联动')
            
            # 读取期权参考
            option_file = 'wisecoin-期权参考.xlsx'
            if not os.path.exists(option_file):
                raise FileNotFoundError(f"未找到文件: {option_file}")
            self.option_ref_df = pd.read_excel(option_file, sheet_name='期权参考')
            
            # 读取波动率曲面数据 (尝试读取，若不存在则忽略)
            self.vol_surface_df = None
            try:
                # 检查是否存在 '波动率曲面' sheet
                xl = pd.ExcelFile(option_file)
                if '波动率曲面' in xl.sheet_names:
                    self.vol_surface_df = pd.read_excel(option_file, sheet_name='波动率曲面')
            except Exception as e:
                print(f"读取波动率曲面失败 (可能未生成): {e}")
            
            # 读取期货K线数据
            klines_file = 'wisecoin-期货K线.xlsx'
            self.klines_data = {}
            if os.path.exists(klines_file):
                try:
                    klines_xl = pd.ExcelFile(klines_file)
                    for sheet_name in klines_xl.sheet_names:
                        if sheet_name == 'Summary':
                            continue
                        df = pd.read_excel(klines_xl, sheet_name=sheet_name)
                        # 过滤 datetime_str 为空的数据
                        if 'datetime_str' in df.columns:
                            df = df[df['datetime_str'].notna() & (df['datetime_str'] != '')]
                        if not df.empty:
                            # 恢复原始合约代码 (sheet名用_替换了.)
                            symbol = sheet_name.replace('_', '.', 1)
                            self.klines_data[symbol] = df
                    print(f"已加载 {len(self.klines_data)} 个合约的K线数据")
                except Exception as e:
                    print(f"读取期货K线数据失败: {e}")
            
            # 计算涨跌幅
            if '昨收' in self.option_ref_df.columns and '期权价' in self.option_ref_df.columns:
                self.option_ref_df['涨跌幅%'] = (
                    (self.option_ref_df['期权价'] - self.option_ref_df['昨收']) / 
                    self.option_ref_df['昨收'] * 100
                ).round(2)
            
            # 构建合约列表
            self.build_contract_list()
            
            # 更新视图
            if self.contract_combo.count() > 0:
                self.on_contract_changed(self.contract_combo.currentText())
            
            self.status_label.setText("状态：数据已加载")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
            
        except Exception as e:
            error_msg = f"加载数据失败: {str(e)}"
            self.status_label.setText(f"状态：{error_msg}")
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
            QMessageBox.critical(self, "错误", error_msg)
    
    def build_contract_list(self):
        """构建合约列表，按期权沉淀排序"""
        if self.market_overview_df is None or self.option_ref_df is None:
            return
        
        if '标的合约' not in self.option_ref_df.columns or '交割年月' not in self.option_ref_df.columns:
            return
        
        contracts_df = self.option_ref_df[['标的合约', '交割年月']].drop_duplicates()
        
        if '标的合约' in self.market_overview_df.columns and '期权沉淀(亿)' in self.market_overview_df.columns:
            contracts_df = contracts_df.merge(
                self.market_overview_df[['标的合约', '期权沉淀(亿)']],
                on='标的合约',
                how='left'
            )
            contracts_df = contracts_df.sort_values('期权沉淀(亿)', ascending=False, na_position='last')
        else:
            contracts_df = contracts_df.sort_values(['标的合约', '交割年月'])
        
        self.contract_list = []
        for _, row in contracts_df.iterrows():
            underlying = str(row['标的合约'])
            expiry = str(row['交割年月'])
            deposit = row.get('期权沉淀(亿)', None)
            
            display_text = f"{underlying} {expiry}"
            if pd.notna(deposit):
                display_text += f" (沉淀:{deposit:.2f}亿)"
            
            self.contract_list.append({
                'display': display_text,
                'underlying': underlying,
                'expiry': expiry
            })
        
        current = self.contract_combo.currentText()
        self.contract_combo.blockSignals(True)
        self.contract_combo.clear()
        self.contract_combo.addItems([c['display'] for c in self.contract_list])
        
        if current:
            index = self.contract_combo.findText(current)
            if index >= 0:
                self.contract_combo.setCurrentIndex(index)
            elif self.contract_list:
                self.contract_combo.setCurrentIndex(0)
        elif self.contract_list:
            self.contract_combo.setCurrentIndex(0)
        
        self.contract_combo.blockSignals(False)
        
        # 更新导航按钮状态
        self._update_nav_buttons()
    
    def on_contract_changed(self, contract_text):
        """合约变化时"""
        if not contract_text or not self.contract_list:
            return
        
        contract_info = None
        for c in self.contract_list:
            if c['display'] == contract_text:
                contract_info = c
                break
        
        if contract_info:
            self.update_option_t_view(contract_info['underlying'], contract_info['expiry'])
        
        # 更新导航按钮状态
        self._update_nav_buttons()
    
    def on_prev_contract(self):
        """切换到上一个合约"""
        current_idx = self.contract_combo.currentIndex()
        if current_idx > 0:
            self.contract_combo.setCurrentIndex(current_idx - 1)
    
    def on_next_contract(self):
        """切换到下一个合约"""
        current_idx = self.contract_combo.currentIndex()
        if current_idx < self.contract_combo.count() - 1:
            self.contract_combo.setCurrentIndex(current_idx + 1)
    
    def _update_nav_buttons(self):
        """更新导航按钮状态"""
        current_idx = self.contract_combo.currentIndex()
        total = self.contract_combo.count()
        
        self.prev_button.setEnabled(current_idx > 0)
        self.next_button.setEnabled(current_idx < total - 1)
        
        # 更新按钮提示
        if current_idx > 0:
            self.prev_button.setToolTip(self.contract_combo.itemText(current_idx - 1))
        if current_idx < total - 1:
            self.next_button.setToolTip(self.contract_combo.itemText(current_idx + 1))
    
    
    
    def _get_t_shape_data_for_export(self, underlying, expiry):
        """获取T型报价数据用于导出（支持指定交割月份）"""
        if self.option_ref_df is None:
            return None
        
        # 筛选条件
        if expiry and pd.notna(expiry):
            df = self.option_ref_df[
                (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
                (self.option_ref_df['交割年月'].astype(str) == str(expiry))
            ].copy()
        else:
            df = self.option_ref_df[
                self.option_ref_df['标的合约'].astype(str) == str(underlying)
            ].copy()
        
        if df.empty:
            return None
        
        if '行权价' not in df.columns or '期权类型' not in df.columns:
            return None
        
        df['期权类型'] = df['期权类型'].apply(self._normalize_option_type)
        df = df[df['期权类型'].isin(['CALL', 'PUT'])]
        
        strikes = sorted(df['行权价'].dropna().unique().tolist(), reverse=True)
        
        if not strikes:
            return None
        
        # 构建T型数据
        rows = []
        for strike in strikes:
            call_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'CALL')]
            put_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'PUT')]
            
            call = call_rows.sort_values('成交金额', ascending=False).iloc[0] if not call_rows.empty else None
            put = put_rows.sort_values('成交金额', ascending=False).iloc[0] if not put_rows.empty else None
            
            row_data = {}
            
            # PUT 数据
            for field in self.all_ref_cols:
                col_name = f"P-{field}"
                if put is not None and field in put.index:
                    row_data[col_name] = put[field]
                else:
                    row_data[col_name] = None
            
            # 行权价
            row_data['行权价'] = strike
            
            # CALL 数据
            for field in self.all_ref_cols:
                col_name = f"C-{field}"
                if call is not None and field in call.index:
                    row_data[col_name] = call[field]
                else:
                    row_data[col_name] = None
            
            rows.append(row_data)
        
        # 构建列顺序
        columns = [f"P-{f}" for f in self.all_ref_cols] + ['行权价'] + [f"C-{f}" for f in self.all_ref_cols]
        
        return pd.DataFrame(rows, columns=columns)
    
    def update_option_t_view(self, underlying, expiry):
        """更新期权T型视图"""
        if self.market_overview_df is None or self.option_ref_df is None:
            return
        
        self.update_stats(underlying)
        self.update_t_shape_table(underlying, expiry)
        self.update_vol_charts(underlying, expiry)
    
    def update_stats(self, underlying):
        """更新统计信息 (货权联动 + 波动率曲面数据匹配)"""
        if self.market_overview_df is None:
            return
        
        if '标的合约' not in self.market_overview_df.columns:
            return
        
        # 1. 匹配 市场概览/货权联动 数据 (按标的合约精确匹配)
        row_match = self.market_overview_df[
            self.market_overview_df['标的合约'].astype(str) == str(underlying)
        ]
        market_row = row_match.iloc[0] if not row_match.empty else None
        
        # 2. 匹配 波动率曲面 分析数据 (按品种代码匹配)
        vol_row = None
        product_code = self._extract_product_code(underlying)
        
        if self.vol_surface_df is not None and product_code:
            # 宽容匹配：去空格、转大写
            if '品种代码' in self.vol_surface_df.columns:
                vol_match = self.vol_surface_df[
                    self.vol_surface_df['品种代码'].astype(str).str.strip().str.upper() == product_code
                ]
                if not vol_match.empty:
                    vol_row = vol_match.iloc[0]
        
        # 3. 如果Excel中没有波动率曲面汇总数据，则根据当前数据实时计算 (确保信息不为空)
        if vol_row is None and self.option_ref_df is not None and product_code:
            try:
                vol_row = self._calculate_vol_surface_metrics(product_code)
            except Exception as e:
                print(f"实时计算波动率指标失败: {e}")
        
        # 4. 更新 UI 标签
        # A. 标的统计信息
        for key, label in self.underlying_stat_labels.items():
            val = '-'
            if market_row is not None:
                val = market_row.get(key, '-')
            
            # 格式化: 期货沉淀(亿) 保留2位小数
            if key == '期货沉淀(亿)' and pd.notna(val) and val != '-':
                try: val = f"{float(val):.2f}"
                except: pass
                
            label.setText(str(val) if pd.notna(val) else '-')
            
        # B. 期权整体统计信息
        for key, label in self.option_stat_labels.items():
            val = '-'
            if market_row is not None:
                val = market_row.get(key, '-')
            
            # 格式化: 期权PCR, 期权沉淀(亿) 保留2位小数
            if key in ['期权PCR', '期权沉淀(亿)'] and pd.notna(val) and val != '-':
                try: val = f"{float(val):.2f}"
                except: pass
                
            label.setText(str(val) if pd.notna(val) else '-')
            
        # C. 货权联动分析 (含波动率曲面汇总字段)
        for key, label in self.linkage_labels.items():
            val = '-'
            # 逻辑：优先从 market_row 获取，若无则从 vol_row 获取
            if market_row is not None and key in market_row.index:
                val = market_row[key]
            
            if (val == '-' or pd.isna(val)) and vol_row is not None and key in vol_row:
                val = vol_row[key]
            
            # 格式化数值 (希腊字母、百分比等)
            if pd.notna(val) and val != '-':
                if key in ['IV倾斜度', '期限结构差', '峰度', '偏度', 'IV/RV比率', '沉淀资金合计(亿)']:
                    try: val = f"{float(val):.2f}"
                    except: pass
                elif ('IV' in key or '%' in key) and isinstance(val, (int, float)):
                    if val < 1.0 and val > 0: # 可能是小数格式
                        val = f"{val*100:.2f}%"
                    else:
                        val = f"{val:.2f}%"
                
            label.setText(str(val) if pd.notna(val) else '-')
    
    def _calculate_vol_surface_metrics(self, product_code):
        """实时计算波动率曲面指标（当Excel中没有时）"""
        # 筛选该品种的所有期权
        mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
        prod_df = self.option_ref_df[mask].copy()
        
        if len(prod_df) < 10:
            return None
        
        result = {}
        
        # 合约数量和资金
        result['合约数量'] = len(prod_df)
        result['资金合计(万)'] = round(prod_df['资金合计(万)'].sum(), 2) if '资金合计(万)' in prod_df.columns else 0
        
        # IV 统计
        iv_values = prod_df['隐含波动率'].dropna()
        hv_values = prod_df['近期波动率'].dropna()
        
        avg_iv = iv_values.mean() if len(iv_values) > 0 else 0
        avg_hv = hv_values.mean() if len(hv_values) > 0 else 0
        result['IV/RV比率'] = round(avg_iv / avg_hv, 2) if avg_hv > 0 else '-'
        
        # 峰度和偏度
        result['峰度'] = round(kurtosis(iv_values, fisher=True), 2) if len(iv_values) >= 4 else 0
        result['偏度'] = round(skew(iv_values), 2) if len(iv_values) >= 4 else 0
        
        # IV Skew (虚值认沽 vs 虚值认购)
        prod_df['期权类型'] = prod_df['期权类型'].apply(self._normalize_option_type)
        otm_puts = prod_df[(prod_df['期权类型'] == 'PUT') & (prod_df['虚实幅度%'] < -5)]
        otm_calls = prod_df[(prod_df['期权类型'] == 'CALL') & (prod_df['虚实幅度%'] < -5)]
        
        put_iv_mean = otm_puts['隐含波动率'].mean() if len(otm_puts) > 0 else 0
        call_iv_mean = otm_calls['隐含波动率'].mean() if len(otm_calls) > 0 else 0
        iv_skew = put_iv_mean - call_iv_mean
        
        result['虚值认沽IV均值'] = round(put_iv_mean, 2)
        result['虚值认购IV均值'] = round(call_iv_mean, 2)
        result['IV倾斜度'] = round(iv_skew, 2)
        result['倾斜方向'] = '看跌倾斜' if iv_skew > 2 else ('看涨倾斜' if iv_skew < -2 else '平坦')
        
        # 期限结构 (短期 vs 长期)
        short_term = prod_df[prod_df['剩余天数'] <= 30]['隐含波动率'].mean()
        long_term = prod_df[prod_df['剩余天数'] > 60]['隐含波动率'].mean()
        term_diff = short_term - long_term if not pd.isna(short_term) and not pd.isna(long_term) else 0
        
        result['短期IV'] = round(short_term, 2) if not pd.isna(short_term) else 0
        result['长期IV'] = round(long_term, 2) if not pd.isna(long_term) else 0
        result['期限结构差'] = round(term_diff, 2)
        result['期限结构'] = '倒挂' if term_diff > 3 else ('升水' if term_diff < -3 else '平坦')
        
        # 市场情绪分类
        iv_rv_ratio = result['IV/RV比率'] if isinstance(result['IV/RV比率'], (int, float)) else 1
        result['市场情绪'] = self._classify_market_sentiment(iv_skew, term_diff, result['峰度'], iv_rv_ratio)
        
        # 推荐策略
        result['推荐策略'] = self._suggest_strategies(result['市场情绪'], iv_rv_ratio, iv_skew, term_diff)
        
        return result
    
    def _classify_market_sentiment(self, iv_skew, term_structure, iv_kurtosis, iv_rv_ratio):
        """根据波动率特征分类市场情绪"""
        if iv_skew > 3 and term_structure > 3:
            return '恐慌下跌'
        elif iv_skew < -3 and term_structure < -3:
            return '狂热上涨'
        elif abs(iv_skew) <= 3 and abs(term_structure) <= 3:
            if iv_kurtosis < -0.5:
                return '窄幅震荡(瘦尾)'
            else:
                return '窄幅震荡'
        elif iv_skew > 0:
            return '震荡筑底'
        elif iv_skew < 0:
            return '震荡冲高'
        return '中性'
    
    def _suggest_strategies(self, sentiment, iv_rv_ratio, iv_skew, term_diff):
        """根据市场情绪推荐策略"""
        strategies = []
        iv_undervalued = iv_rv_ratio < 0.8 if isinstance(iv_rv_ratio, (int, float)) else False
        iv_overvalued = iv_rv_ratio > 1.2 if isinstance(iv_rv_ratio, (int, float)) else False
        
        if '恐慌' in sentiment:
            strategies.append('认沽多头价差')
            if iv_undervalued:
                strategies.append('买入虚值认沽')
        elif '狂热' in sentiment:
            strategies.append('认购多头价差')
            if iv_undervalued:
                strategies.append('买入虚值认购')
        elif '窄幅' in sentiment:
            if iv_overvalued:
                strategies.append('铁鹰式')
                strategies.append('宽跨式空头')
            else:
                strategies.append('日历价差')
        elif '筑底' in sentiment:
            strategies.append('牛市认沽价差')
        elif '冲高' in sentiment:
            strategies.append('保护性认沽')
        else:
            strategies.append('观望')
        
        return ', '.join(strategies[:3])
    
    def update_t_shape_table(self, underlying, expiry):
        """更新T型表格（全字段对称显示，智能多色染色）"""
        self.option_t_table.setRowCount(0)
        
        if self.option_ref_df is None:
            return
        
        df = self.option_ref_df[
            (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
            (self.option_ref_df['交割年月'].astype(str) == str(expiry))
        ].copy()
        
        if df.empty:
            return
        
        if '行权价' not in df.columns or '期权类型' not in df.columns:
            return
        
        df['期权类型'] = df['期权类型'].apply(self._normalize_option_type)
        df = df[df['期权类型'].isin(['CALL', 'PUT'])]
        
        strikes = sorted(df['行权价'].dropna().unique().tolist(), reverse=True)
        
        if not strikes:
            return
        
        atm_price = df['标的现价'].iloc[0] if '标的现价' in df.columns and not df.empty else None
        
        self.option_t_table.setRowCount(len(strikes))
        
        for row_idx, strike in enumerate(strikes):
            call_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'CALL')]
            put_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'PUT')]
            
            call = call_rows.sort_values('成交金额', ascending=False).iloc[0] if not call_rows.empty else None
            put = put_rows.sort_values('成交金额', ascending=False).iloc[0] if not put_rows.empty else None
            
            is_atm = False
            if atm_price is not None:
                is_atm = abs(strike - atm_price) / atm_price < 0.02
            
            col_idx = 0
            
            # 看跌 Wing (P-...)
            for field in self.all_ref_cols:
                value = self._get_field_value(put, field)
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                
                # 智能染色
                if put is not None:
                    fg_color, bg_color = self._get_cell_color(field, put, is_atm, 'PUT')
                    if fg_color:
                        item.setForeground(QBrush(fg_color))
                    if bg_color:
                        item.setBackground(QBrush(bg_color))
                
                self.option_t_table.setItem(row_idx, col_idx, item)
                col_idx += 1
            
            # 行权价 (Center)
            strike_item = QTableWidgetItem(f"{strike}")
            strike_item.setTextAlignment(Qt.AlignCenter)
            strike_item.setFont(self._get_bold_font())
            if is_atm:
                strike_item.setBackground(QBrush(QColor("#ffff99")))
                strike_item.setForeground(QBrush(QColor("#cc0000")))
            else:
                strike_item.setBackground(QBrush(QColor("#f0f0f0")))
            self.option_t_table.setItem(row_idx, col_idx, strike_item)
            col_idx += 1
            
            # 看涨 Wing (C-...)
            for field in self.all_ref_cols:
                value = self._get_field_value(call, field)
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                
                # 智能染色
                if call is not None:
                    fg_color, bg_color = self._get_cell_color(field, call, is_atm, 'CALL')
                    if fg_color:
                        item.setForeground(QBrush(fg_color))
                    if bg_color:
                        item.setBackground(QBrush(bg_color))
                
                self.option_t_table.setItem(row_idx, col_idx, item)
                col_idx += 1
    
    def _get_cell_color(self, field, row, is_atm, option_type):
        """根据字段类型和值返回前景色和背景色"""
        fg_color = None
        bg_color = None
        
        # ATM 行高亮
        if is_atm:
            bg_color = QColor("#fffacd") if option_type == 'PUT' else QColor("#fff0f5")
        
        try:
            value = row[field] if field in row.index else None
            if pd.isna(value):
                return fg_color, bg_color
            
            # Delta: 渐变色 (绝对值越大颜色越深)
            if field == 'Delta':
                abs_delta = abs(float(value))
                if abs_delta > 0.7:
                    fg_color = QColor("#8B0000")  # 深红 (深度实值)
                elif abs_delta > 0.5:
                    fg_color = QColor("#cc0000")  # 红色
                elif abs_delta > 0.3:
                    fg_color = QColor("#ff6600")  # 橙色
                elif abs_delta < 0.15:
                    fg_color = QColor("#0066cc")  # 蓝色 (深度虚值)
                else:
                    fg_color = QColor("#333333")  # 灰色 (中性)
            
            # 隐含波动率/近期波动率: 高IV警示
            elif field in ['隐含波动率', '近期波动率']:
                v = float(value)
                if v > 50:
                    bg_color = QColor("#ffcccc")  # 浅红 (极高波动)
                    fg_color = QColor("#8B0000")
                elif v > 35:
                    bg_color = QColor("#ffe6cc")  # 浅橙
                    fg_color = QColor("#cc6600")
                elif v < 15:
                    bg_color = QColor("#e6f3ff")  # 浅蓝 (低波动)
                    fg_color = QColor("#0066cc")
            
            # 虚实幅度%: ITM红/OTM蓝 渐变
            elif field == '虚实幅度%':
                v = float(value)
                if v > 10:
                    bg_color = QColor("#ffcccc")  # 深度实值
                    fg_color = QColor("#8B0000")
                elif v > 5:
                    bg_color = QColor("#ffe6e6")
                    fg_color = QColor("#cc0000")
                elif v > 0:
                    fg_color = QColor("#ff6600")
                elif v < -10:
                    bg_color = QColor("#cce6ff")  # 深度虚值
                    fg_color = QColor("#003366")
                elif v < -5:
                    bg_color = QColor("#e6f3ff")
                    fg_color = QColor("#0066cc")
                elif v < 0:
                    fg_color = QColor("#3399ff")
            
            # 溢价率%: 低=绿(便宜), 高=红(贵)
            elif field == '溢价率%':
                v = float(value)
                if v < 2:
                    bg_color = QColor("#ccffcc")  # 绿 (便宜)
                    fg_color = QColor("#006600")
                elif v > 10:
                    bg_color = QColor("#ffcccc")  # 红 (贵)
                    fg_color = QColor("#cc0000")
                elif v > 5:
                    fg_color = QColor("#ff6600")
            
            # 涨跌幅%: 涨红/跌绿
            elif field == '涨跌幅%':
                v = float(value)
                if v > 5:
                    fg_color = QColor("#cc0000")
                    bg_color = QColor("#ffe6e6")
                elif v > 0:
                    fg_color = QColor("#cc0000")
                elif v < -5:
                    fg_color = QColor("#008800")
                    bg_color = QColor("#e6ffe6")
                elif v < 0:
                    fg_color = QColor("#008800")
            
            # 成交量/持仓量: 活跃度高亮
            elif field in ['成交量', '持仓量']:
                v = float(value)
                if v > 10000:
                    bg_color = QColor("#fff3cd")  # 黄色 (高活跃)
                    fg_color = QColor("#856404")
                elif v > 5000:
                    fg_color = QColor("#996600")
            
            # Theta: 负值越大颜色越深
            elif field == 'Theta':
                v = float(value)
                if v < -0.01:
                    fg_color = QColor("#8B0000")  # 深红 (时间损耗大)
                elif v < -0.005:
                    fg_color = QColor("#cc0000")
                elif v < 0:
                    fg_color = QColor("#ff6600")
            
            # Vega: 高敏感度高亮
            elif field == 'Vega':
                v = float(value)
                if v > 0.05:
                    fg_color = QColor("#6600cc")  # 紫色 (高波动敏感)
                elif v > 0.02:
                    fg_color = QColor("#9933ff")
            
            # Gamma: 高Gamma高亮
            elif field == 'Gamma':
                v = float(value)
                if v > 0.001:
                    fg_color = QColor("#cc6600")  # 橙色 (高Gamma)
            
            # 收益%/年化收益%: 收益高低
            elif field in ['收益%', '收益年化%', '杠杆收益%', '杠杆年化%']:
                v = float(value)
                if v > 50:
                    fg_color = QColor("#006600")
                    bg_color = QColor("#ccffcc")
                elif v > 20:
                    fg_color = QColor("#008800")
                elif v < 0:
                    fg_color = QColor("#cc0000")
            
            # 时间价值/时间占比%
            elif field == '时间占比%':
                v = float(value)
                if v > 80:
                    fg_color = QColor("#cc6600")  # 高时间价值占比
                elif v < 20:
                    fg_color = QColor("#0066cc")  # 低时间价值占比
            
            # 买方杠杆
            elif field == '买方杠杆':
                v = float(value)
                if v > 50:
                    fg_color = QColor("#cc0000")
                    bg_color = QColor("#fff0f0")
                elif v > 20:
                    fg_color = QColor("#ff6600")
                elif v > 10:
                    fg_color = QColor("#996600")
            
            # 资金类字段: 大资金高亮
            elif '资金' in field:
                v = float(value)
                if v > 1000:
                    bg_color = QColor("#e6f3ff")
                    fg_color = QColor("#003366")
                elif v > 500:
                    fg_color = QColor("#0066cc")
                    
        except (ValueError, TypeError):
            pass
        
        return fg_color, bg_color
    
    def update_vol_charts(self, underlying, expiry):
        """更新波动率图表和K线图"""
        if self.option_ref_df is None:
            return
        
        # 1. 提取品种代码 (e.g. SHFE.au2602 -> AU)
        product_code = self._extract_product_code(underlying)
        
        # 2. 筛选该品种的所有期权数据 (用于波动率曲面)
        # 需匹配所有属于该品种的标的合约
        if product_code:
            # 临时添加品种代码列进行筛选 (或直接apply)
            # 为了性能，建议在 load_data 时预处理，这里直接apply
            mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
            df_product = self.option_ref_df[mask].copy()
            surface_title = f"{product_code}"
        else:
            # 回退：仅使用当前标的
            df_product = self.option_ref_df[
                self.option_ref_df['标的合约'].astype(str) == str(underlying)
            ].copy()
            surface_title = f"{underlying} 波动率曲面"
        
        # 3. 筛选当前标的和到期日的数据 (用于微笑曲线)
        df_smile = self.option_ref_df[
            (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
            (self.option_ref_df['交割年月'].astype(str) == str(expiry))
        ].copy()
        
        # 4. 绘制K线图
        self.plot_kline_chart(underlying)
        
        # 5. 绘制波动率图表
        self.plot_vol_surface(df_product, surface_title)
        self.plot_vol_smile(df_smile, expiry)

    def _extract_product_code(self, symbol):
        """
        从合约代码提取品种代码 (e.g. SHFE.au2602 -> AU, SSE.000852 -> 000852)
        逻辑与 04/05 脚本保持一致，确保跨表匹配成功
        """
        try:
            s = str(symbol).strip()
            if '.' in s:
                parts = s.split('.')
                code_part = parts[1]
            else:
                code_part = s
            
            # 优先匹配开头的纯字母部分 (如 au2602 -> AU)
            match_letter = re.match(r'^([a-zA-Z]+)', code_part)
            if match_letter:
                return match_letter.group(1).upper()
            
            # 如果开头是数字，则匹配数字部分 (如 000852 -> 000852)
            match_number = re.match(r'^(\d+)', code_part)
            if match_number:
                return match_number.group(1)
                
            return code_part.upper()
        except:
            return ''
    
    def plot_kline_chart(self, underlying):
        """绘制标的K线图（简明清晰大方风格）"""
        self.kline_fig.clear()
        
        # 获取K线数据
        kline_df = self.klines_data.get(underlying)
        
        if kline_df is None or kline_df.empty:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'无K线数据\n{underlying}', 
                    transform=ax.transAxes, ha='center', va='center',
                    fontsize=10, color='gray')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 准备数据
        df = kline_df.copy()
        
        # 确保有必要的列
        required_cols = ['open', 'high', 'low', 'close']
        if not all(col in df.columns for col in required_cols):
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, 'K线数据缺少OHLC字段', 
                    transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 转换数据类型
        for col in required_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 过滤无效数据
        df = df.dropna(subset=required_cols)
        
        if df.empty:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, 'K线数据无效', 
                    transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 取最近60根K线显示
        df = df.tail(120).reset_index(drop=True)
        
        try:
            ax = self.kline_fig.add_subplot(111)
            
            # 计算涨跌
            up = df['close'] >= df['open']
            down = df['close'] < df['open']
            
            # 颜色设置（中国市场风格：红涨绿跌）
            up_color = '#CC0000'      # 红色-上涨
            down_color = '#009900'    # 绿色-下跌
            
            # K线宽度
            width = 0.6
            width2 = 0.08
            
            # 绘制上涨K线
            ax.bar(df.index[up], df['close'][up] - df['open'][up], width, 
                   bottom=df['open'][up], color=up_color, edgecolor=up_color, linewidth=0.5)
            ax.bar(df.index[up], df['high'][up] - df['close'][up], width2, 
                   bottom=df['close'][up], color=up_color, linewidth=0)
            ax.bar(df.index[up], df['open'][up] - df['low'][up], width2, 
                   bottom=df['low'][up], color=up_color, linewidth=0)
            
            # 绘制下跌K线
            ax.bar(df.index[down], df['close'][down] - df['open'][down], width, 
                   bottom=df['open'][down], color=down_color, edgecolor=down_color, linewidth=0.5)
            ax.bar(df.index[down], df['high'][down] - df['open'][down], width2, 
                   bottom=df['open'][down], color=down_color, linewidth=0)
            ax.bar(df.index[down], df['close'][down] - df['low'][down], width2, 
                   bottom=df['low'][down], color=down_color, linewidth=0)
            
            # 添加MA5和MA20均线
            if len(df) >= 5:
                ma5 = df['close'].rolling(window=5).mean()
                ax.plot(df.index, ma5, color='#FF9900', linewidth=1, label='MA5', alpha=0.8)
            if len(df) >= 20:
                ma20 = df['close'].rolling(window=20).mean()
                ax.plot(df.index, ma20, color='#0066CC', linewidth=1, label='MA20', alpha=0.8)
            
            # 设置标题和标签
            contract_name = underlying.split('.')[-1] if '.' in underlying else underlying
            ax.set_title(f'{contract_name} 日K线', fontsize=10, fontweight='bold')
            
            # 简化X轴刻度
            if 'datetime_str' in df.columns:
                # 只显示首尾和中间几个日期
                tick_positions = [0, len(df)//4, len(df)//2, 3*len(df)//4, len(df)-1]
                tick_positions = [p for p in tick_positions if p < len(df)]
                tick_labels = [str(df['datetime_str'].iloc[p])[:10] if pd.notna(df['datetime_str'].iloc[p]) else '' 
                              for p in tick_positions]
                ax.set_xticks(tick_positions)
                ax.set_xticklabels(tick_labels, fontsize=7, rotation=15)
            else:
                ax.set_xticks([])
            
            # 设置Y轴格式
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            ax.tick_params(axis='y', labelsize=8)
            
            # 网格线
            ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
            ax.set_axisbelow(True)
            
            # 图例
            if len(df) >= 5:
                ax.legend(loc='upper left', fontsize=7, framealpha=0.8)
            
            # 边框美化
            for spine in ['top', 'right']:
                ax.spines[spine].set_visible(False)
            ax.spines['left'].set_color('#888888')
            ax.spines['bottom'].set_color('#888888')
            
            # 显示最新价格
            last_close = df['close'].iloc[-1]
            ax.axhline(y=last_close, color='#666666', linestyle=':', linewidth=0.8, alpha=0.6)
            ax.text(len(df)-1, last_close, f' {last_close:,.1f}', 
                    fontsize=7, va='center', ha='left', color='#333333')
            
        except Exception as e:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'绘图错误:\n{str(e)[:30]}', 
                    transform=ax.transAxes, ha='center', va='center', fontsize=9)
            ax.axis('off')
        
        self.kline_fig.tight_layout()
        self.kline_canvas.draw()
    
    def _generate_vol_surface_image(self, df, underlying):
        """生成波动率曲面图片（用于导出）"""
        if df is None or df.empty:
            return None
        
        plot_df = df[(df['隐含波动率'] > 0) & (df['剩余天数'] > 0) & (df['行权价'] > 0)].copy()
        
        if len(plot_df) < 6:
            return None
        
        try:
            fig = Figure(figsize=(8, 6), dpi=100)
            ax = fig.add_subplot(111, projection='3d')
            
            # 标准化行权价
            und_price = plot_df['标的现价'].iloc[0] if '标的现价' in plot_df.columns else 0
            if und_price > 0:
                plot_df['Moneyness'] = (plot_df['行权价'] / und_price - 1) * 100
            else:
                plot_df['Moneyness'] = plot_df['行权价']
            
            x = plot_df['Moneyness'].values
            y = plot_df['剩余天数'].values
            z = plot_df['隐含波动率'].values
            
            # 创建网格
            xi = np.linspace(x.min(), x.max(), 35)
            yi = np.linspace(y.min(), y.max(), 35)
            xi_grid, yi_grid = np.meshgrid(xi, yi)
            
            try:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='linear')
            except:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
            
            if np.any(np.isnan(zi_grid)):
                zi_nearest = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
                zi_grid = np.where(np.isnan(zi_grid), zi_nearest, zi_grid)
            
            # 绘制曲面
            surf = ax.plot_surface(xi_grid, yi_grid, zi_grid, cmap='viridis', 
                                   alpha=0.88, edgecolor='none', linewidth=0, antialiased=True)
            
            cbar = fig.colorbar(surf, shrink=0.55, aspect=12, pad=0.1)
            cbar.set_label('IV (%)', fontsize=8)
            
            ax.scatter(x, y, z, color='red', s=12, alpha=0.6, depthshade=True)
            
            y_range = np.linspace(y.min(), y.max(), 20)
            z_atm = np.interp(y_range, y[np.argsort(y)], z[np.argsort(y)])
            ax.plot([0]*len(y_range), y_range, z_atm, color='white', linewidth=2, alpha=0.8, linestyle='--')
            
            ax.set_xlabel('行权价 (%)', fontsize=9, labelpad=8)
            ax.set_ylabel('剩余期限 (天)', fontsize=9, labelpad=8)
            ax.set_zlabel('隐含波动率 (%)', fontsize=9, labelpad=8)
            
            ax.view_init(elev=22, azim=-60)
            ax.tick_params(axis='both', which='major', labelsize=7)
            
            fig.tight_layout()
            
            # 保存到内存
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            return buf
            
        except Exception as e:
            print(f"生成波动率曲面图失败: {e}")
            return None
    
    def _generate_vol_smile_image(self, df, expiry):
        """生成微笑曲线图片（用于导出）"""
        if df is None or df.empty:
            return None
        
        temp = df.copy()
        temp['期权类型'] = temp['期权类型'].apply(self._normalize_option_type)
        temp = temp[(temp['隐含波动率'] > 0) & (temp['行权价'] > 0)]
        
        if temp.empty:
            return None
        
        try:
            fig = Figure(figsize=(8, 5), dpi=100)
            ax = fig.add_subplot(111)
            
            # 计算Moneyness
            und_price = temp['标的现价'].iloc[0] if '标的现价' in temp.columns and pd.notna(temp['标的现价'].iloc[0]) else None
            if und_price and und_price > 0:
                temp['Moneyness'] = (temp['行权价'] / und_price - 1) * 100
                x_label = '价值度 (%)'
                x_col = 'Moneyness'
            else:
                x_label = '行权价'
                x_col = '行权价'
            
            calls = temp[temp['期权类型'] == 'CALL'].sort_values(x_col)
            puts = temp[temp['期权类型'] == 'PUT'].sort_values(x_col)
            
            # 绘制曲线
            if not calls.empty:
                ax.plot(calls[x_col], calls['隐含波动率'], label='看涨 (Call)', 
                        color='#cc0000', marker='o', markersize=6, linewidth=2, alpha=0.85)
            if not puts.empty:
                ax.plot(puts[x_col], puts['隐含波动率'], label='看跌 (Put)', 
                        color='#008800', marker='s', markersize=6, linewidth=2, alpha=0.85)
            
            # 添加ATM线
            if x_col == 'Moneyness':
                ax.axvline(x=0, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label='ATM')
                xlim = ax.get_xlim()
                ylim = ax.get_ylim()
                ax.axvspan(xlim[0], 0, alpha=0.05, color='green')
                ax.axvspan(0, xlim[1], alpha=0.05, color='red')
                ax.set_xlim(xlim)
                ax.set_ylim(ylim)
            elif und_price:
                ax.axvline(x=und_price, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label=f'ATM ({und_price:.2f})')
            
            ax.set_xlabel(x_label, fontsize=10)
            ax.set_ylabel('隐含波动率 (%)', fontsize=10)
            ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.tick_params(axis='both', which='major', labelsize=9)
            
            fig.tight_layout()
            
            # 保存到内存
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            return buf
            
        except Exception as e:
            print(f"生成微笑曲线图失败: {e}")
            return None
    
    def plot_vol_surface(self, df, underlying):
        """绘制波动率曲面 - X轴:行权价, Y轴:剩余期限, Z轴:隐含波动率"""
        self.surface_fig.clear()
        
        if df is None or df.empty:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.surface_canvas.draw()
            return
        
        plot_df = df[(df['隐含波动率'] > 0) & (df['剩余天数'] > 0) & (df['行权价'] > 0)].copy()
        
        if len(plot_df) < 6:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'数据不足 ({len(plot_df)}点)', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.surface_canvas.draw()
            return

        try:
            ax = self.surface_fig.add_subplot(111, projection='3d')
            
            # 标准化行权价 (使用Moneyness百分比)
            und_price = plot_df['标的现价'].iloc[0] if '标的现价' in plot_df.columns else 0
            if und_price > 0:
                plot_df['Moneyness'] = (plot_df['行权价'] / und_price - 1) * 100
            else:
                plot_df['Moneyness'] = plot_df['行权价']
            
            # X轴：行权价(Moneyness)，Y轴：剩余期限（从近到远），Z轴：IV
            x = plot_df['Moneyness'].values      # X轴：Moneyness (%)
            y = plot_df['剩余天数'].values       # Y轴：剩余天数（从近到远）
            z = plot_df['隐含波动率'].values     # Z轴：隐含波动率
            
            # 创建网格
            xi = np.linspace(x.min(), x.max(), 35)
            yi = np.linspace(y.min(), y.max(), 35)
            xi_grid, yi_grid = np.meshgrid(xi, yi)
            
            try:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='linear')
            except:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
            
            # 填充NaN值
            if np.any(np.isnan(zi_grid)):
                zi_nearest = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
                zi_grid = np.where(np.isnan(zi_grid), zi_nearest, zi_grid)
            
            # 绘制曲面 - 业界标准配色viridis
            surf = ax.plot_surface(xi_grid, yi_grid, zi_grid, cmap='viridis', 
                                   alpha=0.88, edgecolor='none', linewidth=0, antialiased=True)
            
            # 添加颜色条
            cbar = self.surface_fig.colorbar(surf, shrink=0.55, aspect=12, pad=0.1)
            cbar.set_label('IV (%)', fontsize=8)
            
            # 绘制实际数据点
            ax.scatter(x, y, z, color='red', s=12, alpha=0.6, depthshade=True)
            
            # 添加ATM参考线 (Moneyness=0)
            y_range = np.linspace(y.min(), y.max(), 20)
            z_atm = np.interp(y_range, y[np.argsort(y)], z[np.argsort(y)])
            ax.plot([0]*len(y_range), y_range, z_atm, color='white', linewidth=2, alpha=0.8, linestyle='--')
            
            # 设置坐标轴 - 业界标准：Strike × Time × IV
            ax.set_xlabel('行权价 (%)', fontsize=9, labelpad=8)
            ax.set_ylabel('剩余期限 (天)', fontsize=9, labelpad=8)
            ax.set_zlabel('隐含波动率 (%)', fontsize=9, labelpad=8)
            
            # 优化视角：从低行权价看向高行权价，近月在前
            ax.view_init(elev=22, azim=-60)
            
            # 设置刻度
            ax.tick_params(axis='both', which='major', labelsize=7)
            
            # 居中优化：手动调整边距，使其稍微偏左
            self.surface_fig.subplots_adjust(left=0.0, right=0.88, top=0.95, bottom=0.05)
            
        except Exception as e:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'绘图错误: {str(e)[:50]}', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
        
        # self.surface_fig.tight_layout()  # 注释掉 tight_layout 以免覆盖 subplots_adjust
        self.surface_canvas.draw()
    
    def plot_vol_smile(self, df, expiry):
        """绘制微笑曲线 (优化版) - 使用Moneyness标准化，添加ATM标记"""
        self.smile_fig.clear()
        
        if df is None or df.empty:
            ax = self.smile_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.smile_canvas.draw()
            return
        
        temp = df.copy()
        temp['期权类型'] = temp['期权类型'].apply(self._normalize_option_type)
        temp = temp[(temp['隐含波动率'] > 0) & (temp['行权价'] > 0)]
        
        if temp.empty:
            ax = self.smile_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无有效波动率数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.smile_canvas.draw()
            return
        
        # 计算Moneyness
        und_price = temp['标的现价'].iloc[0] if '标的现价' in temp.columns and pd.notna(temp['标的现价'].iloc[0]) else None
        if und_price and und_price > 0:
            temp['Moneyness'] = (temp['行权价'] / und_price - 1) * 100
            x_label = '价值度 (%)'
            x_col = 'Moneyness'
        else:
            x_label = '行权价'
            x_col = '行权价'
        
        calls = temp[temp['期权类型'] == 'CALL'].sort_values(x_col)
        puts = temp[temp['期权类型'] == 'PUT'].sort_values(x_col)
        
        ax = self.smile_fig.add_subplot(111)
        
        # 绘制看涨和看跌曲线
        if not calls.empty:
            ax.plot(calls[x_col], calls['隐含波动率'], label='看涨 (Call)', 
                    color='#cc0000', marker='o', markersize=6, linewidth=2, alpha=0.85)
        if not puts.empty:
            ax.plot(puts[x_col], puts['隐含波动率'], label='看跌 (Put)', 
                    color='#008800', marker='s', markersize=6, linewidth=2, alpha=0.85)
        
        # 添加ATM垂直线
        if x_col == 'Moneyness':
            ax.axvline(x=0, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label='ATM')
        elif und_price:
            ax.axvline(x=und_price, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label=f'ATM ({und_price:.2f})')
        
        # 填充ITM/OTM区域
        if x_col == 'Moneyness':
            xlim = ax.get_xlim()
            ylim = ax.get_ylim()
            ax.axvspan(xlim[0], 0, alpha=0.05, color='green', label='_nolegend_')  # ITM for Put
            ax.axvspan(0, xlim[1], alpha=0.05, color='red', label='_nolegend_')    # ITM for Call
            ax.set_xlim(xlim)
            ax.set_ylim(ylim)
        
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel('隐含波动率 (%)', fontsize=10)
        ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # 添加注释
        if not calls.empty or not puts.empty:
            all_iv = pd.concat([calls['隐含波动率'], puts['隐含波动率']]) if not puts.empty else calls['隐含波动率']
            avg_iv = all_iv.mean()
            ax.axhline(y=avg_iv, color='gray', linestyle=':', linewidth=1, alpha=0.6)
            ax.text(ax.get_xlim()[1], avg_iv, f' 平均IV: {avg_iv:.1f}%', 
                    va='center', ha='left', fontsize=8, color='gray')
        
        self.smile_fig.tight_layout()
        self.smile_canvas.draw()
    
    def _normalize_option_type(self, opt_type):
        if pd.isna(opt_type): return ''
        ot = str(opt_type).upper()
        if 'CALL' in ot or ot == 'C': return 'CALL'
        if 'PUT' in ot or ot == 'P': return 'PUT'
        return ''
    
    def _get_field_value(self, row, field):
        if row is None or field not in row.index: return '-'
        value = row[field]
        if pd.isna(value): return '-'
        if isinstance(value, (int, float)):
            if field == 'Gamma': return f"{value:.6f}"
            if field in ['Delta', 'Theta', 'Vega', 'Rho']: return f"{value:.4f}"
            if field in ['隐含波动率', '近期波动率', '虚实幅度%', '涨跌幅%', '收益%', '溢价率%']: return f"{value:.2f}"
            if field in ['期权价', '理论价格', '内在价值', '时间价值', '标的现价', '昨收', '昨结', '今结']: return f"{value:.1f}"
            if '资金' in field or '成交' in field or '持仓' in field or '合约乘数' in field:
                try: return f"{int(value):,}"
                except: return f"{value}"
            return f"{value:.2f}"
        return str(value)
    
    def _get_bold_font(self):
        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        return font
    
    def _safe_get_value(self, row, field, default='-'):
        """安全地从Series或dict中获取值，处理NaN和None"""
        try:
            if isinstance(row, pd.Series):
                if field in row.index:
                    val = row[field]
                    if pd.notna(val):
                        return val
            elif isinstance(row, dict):
                if field in row:
                    val = row[field]
                    if val is not None and (not isinstance(val, float) or not pd.isna(val)):
                        return val
        except:
            pass
        return default
    
    def _calculate_display_width(self, text):
        """计算字符串的显示宽度（中文字符算2个宽度，英文算1个）"""
        if text is None:
            return 0
        text_str = str(text)
        width = 0
        for char in text_str:
            # 判断是否为中文字符（包括中文标点）
            if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                width += 2
            else:
                width += 1
        return width
    
    
    def _adjust_column_widths_for_mixed_data(self, worksheet, export_df, t_shape_data=None):
        """智能调整列宽（用于混合数据：统计信息+T型报价）"""
        # 获取worksheet的最大列数
        max_col = worksheet.max_column
        
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_width = 0
            
            # 遍历该列的所有单元格，计算最大宽度（考虑中文）
            for row_idx in range(1, min(worksheet.max_row + 1, 500)):  # 限制最多检查500行
                cell = worksheet[f'{col_letter}{row_idx}']
                if cell.value is not None:
                    cell_width = self._calculate_display_width(cell.value)
                    max_width = max(max_width, cell_width)
            
            # 如果有T型报价数据，特别处理这部分的列宽
            if t_shape_data is not None and col_idx <= len(t_shape_data.columns):
                # T型报价的列
                col_name = t_shape_data.columns[col_idx - 1] if col_idx <= len(t_shape_data.columns) else None
                if col_name:
                    header_width = self._calculate_display_width(col_name)
                    data_width = t_shape_data.iloc[:, col_idx - 1].apply(lambda x: self._calculate_display_width(x)).max()
                    max_width = max(max_width, header_width, data_width)
            
            # 设置列宽：最小6，最大40（混合数据可能需要更宽），加1.5个字符的边距
            column_width = min(max(max_width * 1.1 + 1.5, 6), 40)
            worksheet.column_dimensions[col_letter].width = column_width
    
    def export_important_contracts(self):
        """导出期权沉淀≥0.20亿的重要合约，每个合约一个分页包含完整信息"""
        try:
            self.status_label.setText("状态：正在导出重要期权...")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            QApplication.processEvents()
            
            if self.market_overview_df is None or self.option_ref_df is None:
                QMessageBox.warning(self, "警告", "请先加载数据！")
                return
            
            # 1. 筛选沉淀≥0.20亿的合约
            if '期权沉淀(亿)' not in self.market_overview_df.columns:
                QMessageBox.warning(self, "警告", "数据中缺少 '期权沉淀(亿)' 字段！")
                return
            
            important_df = self.market_overview_df[
                self.market_overview_df['期权沉淀(亿)'] >= 0.20
            ].copy()
            
            if important_df.empty:
                QMessageBox.information(self, "提示", "没有找到沉淀≥0.20亿的合约！")
                return
            
            # 按沉淀从高到低排序
            important_df = important_df.sort_values('期权沉淀(亿)', ascending=False)
            
            # 2. 创建Excel写入器
            output_file = 'wisecoin-重要期权.xlsx'
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                
                sheet_count = 0
                # 为每个重要合约的每个交割月份创建一个分页
                for _, market_row in important_df.iterrows():
                    underlying = market_row['标的合约']
                    product_code = self._extract_product_code(underlying)
                    
                    # 获取该标的的所有交割月份
                    contract_expiries = self.option_ref_df[
                        self.option_ref_df['标的合约'].astype(str) == str(underlying)
                    ]['交割年月'].unique() if '交割年月' in self.option_ref_df.columns else [None]
                    
                    # 为每个交割月份创建一个分页
                    for expiry in contract_expiries:
                        # 生成sheet名称：合约代码_交割年月
                        contract_part = underlying.split('.')[-1] if '.' in str(underlying) else str(underlying)
                        if expiry and pd.notna(expiry):
                            sheet_name = f"{contract_part}_{expiry}"[:31]
                        else:
                            sheet_name = contract_part[:31]
                        
                        # 准备该合约+交割月的完整数据
                        export_data = []
                        
                        # === 第一部分：标的统计信息 ===
                        export_data.append(['=== 标的统计信息 ==='])
                        export_data.append(['字段', '数值'])
                        export_data.append(['标的合约', self._safe_get_value(market_row, '标的合约')])
                        export_data.append(['交割年月', expiry if expiry and pd.notna(expiry) else '-'])
                        export_data.append(['期货现价', self._safe_get_value(market_row, '期货现价')])
                        export_data.append(['杠杆涨跌%', self._safe_get_value(market_row, '杠杆涨跌%')])
                        export_data.append(['期货沉淀(亿)', self._safe_get_value(market_row, '期货沉淀(亿)')])
                        export_data.append(['期货状态', self._safe_get_value(market_row, '期货状态')])
                        export_data.append(['期货方向', self._safe_get_value(market_row, '期货方向')])
                        export_data.append(['期货流向', self._safe_get_value(market_row, '期货流向')])
                        export_data.append([])  # 空行
                        
                        # === 第二部分：期权整体统计信息 ===
                        export_data.append(['=== 期权整体统计信息 ==='])
                        export_data.append(['字段', '数值'])
                        export_data.append(['期权结构', self._safe_get_value(market_row, '期权结构')])
                        export_data.append(['期权情绪', self._safe_get_value(market_row, '期权情绪')])
                        export_data.append(['期权PCR', self._safe_get_value(market_row, '期权PCR')])
                        export_data.append(['期权沉淀(亿)', self._safe_get_value(market_row, '期权沉淀(亿)')])
                        export_data.append(['最大痛点', self._safe_get_value(market_row, '最大痛点')])
                        export_data.append(['痛点距离%', self._safe_get_value(market_row, '痛点距离%')])
                        export_data.append(['联动状态', self._safe_get_value(market_row, '联动状态')])
                        export_data.append(['共振评分', self._safe_get_value(market_row, '共振评分')])
                        export_data.append([])  # 空行
                        
                        # === 第三部分：货权联动分析（含波动率曲面） ===
                        export_data.append(['=== 货权联动分析（含波动率曲面） ==='])
                        export_data.append(['字段', '数值'])
                        
                        # 获取波动率曲面信息
                        vol_row = None
                        if self.vol_surface_df is not None and product_code:
                            if '品种代码' in self.vol_surface_df.columns:
                                vol_match = self.vol_surface_df[
                                    self.vol_surface_df['品种代码'].astype(str).str.strip().str.upper() == product_code
                                ]
                                if not vol_match.empty:
                                    vol_row = vol_match.iloc[0]
                        
                        # 如果Excel中没有，尝试实时计算
                        if vol_row is None and product_code:
                            try:
                                vol_row = self._calculate_vol_surface_metrics(product_code)
                            except:
                                vol_row = None
                        
                        # 联动分析字段
                        linkage_fields = [
                            '共振等级', '期限结构', '倾斜方向', '共振标签',
                            '联动总分', 'IV/RV比率', '期限结构差', 'IV倾斜度',
                            '价格评分', '峰度', '短期IV', '虚值认沽IV均值', '沉淀资金合计(亿)',
                            '情绪评分', '偏度', '长期IV', '虚值认购IV均值', '合约数量',
                            '市场解读', '推荐策略', '适合策略', '不适合策略'
                        ]
                        
                        for field in linkage_fields:
                            # 优先从 market_row 获取，然后从 vol_row 获取
                            val = self._safe_get_value(market_row, field, None)
                            if val == '-' or val is None:
                                val = self._safe_get_value(vol_row, field, '-') if vol_row is not None else '-'
                            export_data.append([field, val])
                        
                        export_data.append([])  # 空行
                        
                        # === 第四部分：期权T型报价数据（仅该交割月） ===
                        export_data.append(['=== 期权T型报价数据 ==='])
                        export_data.append([])  # 空行
                        
                        # 获取该标的+该交割月的期权数据
                        if expiry and pd.notna(expiry):
                            contract_options = self.option_ref_df[
                                (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
                                (self.option_ref_df['交割年月'].astype(str) == str(expiry))
                            ].copy()
                        else:
                            contract_options = self.option_ref_df[
                                self.option_ref_df['标的合约'].astype(str) == str(underlying)
                            ].copy()
                        
                        if not contract_options.empty:
                            # 构建T型报价格式（与界面一致）
                            t_shape_data = self._get_t_shape_data_for_export(underlying, expiry)
                            
                            if t_shape_data is not None and not t_shape_data.empty:
                                # 添加表头
                                export_data.append(t_shape_data.columns.tolist())
                                
                                # 添加数据行
                                for _, row in t_shape_data.iterrows():
                                    export_data.append(row.tolist())
                            else:
                                export_data.append(['无T型报价数据'])
                        else:
                            export_data.append(['无期权数据'])
                        
                        export_data.append([])  # 空行
                        
                        # 记录图片插入位置（在T型报价数据之后）
                        chart_insert_row = len(export_data) + 1
                        
                        # 写入该合约+交割月的完整数据到一个Sheet
                        export_df = pd.DataFrame(export_data)
                        export_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                        
                        # 智能调整列宽
                        worksheet = writer.sheets[sheet_name]
                        self._adjust_column_widths_for_mixed_data(worksheet, export_df, t_shape_data if 't_shape_data' in locals() else None)
                        
                        # === 生成并插入图片 ===
                        
                        # 生成波动率曲面图（整个品种）
                        if product_code:
                            mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
                            vol_surface_data = self.option_ref_df[mask].copy()
                            
                            surface_img_buf = self._generate_vol_surface_image(vol_surface_data, product_code)
                            if surface_img_buf:
                                img = OpenpyxlImage(surface_img_buf)
                                img.width = 600
                                img.height = 450
                                worksheet.add_image(img, f'A{chart_insert_row}')
                        
                        # 生成微笑曲线图（仅该交割月）
                        if not contract_options.empty:
                            smile_img_buf = self._generate_vol_smile_image(contract_options, expiry if expiry else underlying)
                            if smile_img_buf:
                                img = OpenpyxlImage(smile_img_buf)
                                img.width = 600
                                img.height = 375
                                worksheet.add_image(img, f'J{chart_insert_row}')
                        
                        sheet_count += 1
            
            self.status_label.setText(f"状态：已导出 {sheet_count} 个重要合约到 {output_file}")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
            
            QMessageBox.information(
                self, 
                "导出成功", 
                f"已成功导出 {sheet_count} 个重要期权合约到:\n{output_file}\n\n"
                f"每个合约一个分页，包含:\n"
                f"• 标的统计信息\n"
                f"• 期权整体统计信息\n"
                f"• 货权联动分析（含波动率曲面）\n"
                f"• 期权T型报价数据\n"
                f"• 波动率曲面图（PNG）\n"
                f"• 微笑曲线图（PNG，按交割月）"
            )
            
        except Exception as e:
            error_msg = f"导出失败: {str(e)}"
            self.status_label.setText(f"状态：{error_msg}")
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
            QMessageBox.critical(self, "错误", error_msg)
            import traceback
            traceback.print_exc()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = OptionTShapeWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
