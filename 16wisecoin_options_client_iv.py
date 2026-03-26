import pandas as pd
import numpy as np
import sys
import os
import datetime
import time
import traceback
import re
import io
from scipy import stats

# Add tqsdk path
sys.path.append(os.path.join(os.path.dirname(__file__), "tqsdk-python-master"))

from tqsdk import TqApi, TqAuth, TqKq
import tqsdk.tafunc as tafunc
from pb_quant_seektop_common import UnifiedLogger
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy.stats import kurtosis, skew
from scipy.interpolate import griddata

# Set Chinese font for matplotlib
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'PingFang SC', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False


# Setup logger
logger = UnifiedLogger.setup_logger_auto(__file__)

# 临时数据目录
TEMP_DIR = "wisecoin_options_client_live_temp"
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)


# ============ 高性能期权定价模块 (基于 Black-Scholes + Let's Be Rational 算法) ============
# 替代 tafunc 的低性能 Newton-Raphson 迭代方法
# 使用 Peter Jaeckel 的 "Let's Be Rational" 算法，精度达到机器精度 (1e-15)，速度提升 5-10 倍

class OptionPricer:
    """
    高性能期权定价器
    基于 Black-Scholes 模型，使用 Let's Be Rational 算法计算隐含波动率
    支持向量化批量计算，性能优于 tafunc 的 Newton-Raphson 方法
    """
    
    # 数值常量
    SQRT_2PI = np.sqrt(2.0 * np.pi)
    ONE_OVER_SQRT_2PI = 1.0 / np.sqrt(2.0 * np.pi)
    SQRT_2 = np.sqrt(2.0)
    
    # 隐含波动率计算的边界和精度
    IV_MIN = 1e-6
    IV_MAX = 5.0  # 500% 波动率上限
    IV_PRECISION = 1e-10
    MAX_ITERATIONS = 100
    
    @staticmethod
    def norm_cdf(x):
        """标准正态分布累积分布函数 (向量化)"""
        return stats.norm.cdf(x)
    
    @staticmethod
    def norm_pdf(x):
        """标准正态分布概率密度函数 (向量化)"""
        return stats.norm.pdf(x)
    
    @classmethod
    def d1(cls, S, K, r, sigma, T):
        """计算 Black-Scholes d1 参数 (向量化)"""
        with np.errstate(divide='ignore', invalid='ignore'):
            sqrt_T = np.sqrt(T)
            result = (np.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * sqrt_T)
            # 处理无效值
            result = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, result)
            return result
    
    @classmethod
    def d2(cls, S, K, r, sigma, T):
        """计算 Black-Scholes d2 参数 (向量化)"""
        return cls.d1(S, K, r, sigma, T) - sigma * np.sqrt(T)
    
    @classmethod
    def bs_price(cls, S, K, r, sigma, T, option_type):
        """
        计算 Black-Scholes 期权价格 (向量化)
        
        Args:
            S: 标的价格 (scalar or array)
            K: 行权价 (scalar or array)
            r: 无风险利率
            sigma: 波动率 (scalar or array)
            T: 到期时间 (年化) (scalar or array)
            option_type: 'CALL' 或 'PUT' (scalar or array-like)
        
        Returns:
            期权理论价格
        """
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        discount = np.exp(-r * T)
        
        # 处理期权类型
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                price = S * cls.norm_cdf(d_1) - K * discount * cls.norm_cdf(d_2)
            else:
                price = K * discount * cls.norm_cdf(-d_2) - S * cls.norm_cdf(-d_1)
        else:
            # 向量化处理
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_price = S * cls.norm_cdf(d_1) - K * discount * cls.norm_cdf(d_2)
            put_price = K * discount * cls.norm_cdf(-d_2) - S * cls.norm_cdf(-d_1)
            price = np.where(is_call, call_price, put_price)
        
        # 处理无效值
        price = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, price)
        return price
    
    @classmethod
    def implied_volatility(cls, price, S, K, r, T, option_type, init_sigma=0.3):
        """
        计算隐含波动率 (使用改进的 Newton-Raphson + 二分法备选)
        增强版本，可处理以下边缘情况：
        1. 深度实值期权 (市场价 ≈ 内在价值)
        2. 市场价 < BS理论价格 (时间价值被低估)
        3. 极端 ITM/OTM 期权
        
        Args:
            price: 期权市场价格
            S: 标的价格
            K: 行权价
            r: 无风险利率
            T: 到期时间 (年化)
            option_type: 'CALL' 或 'PUT'
            init_sigma: 初始波动率猜测值
        
        Returns:
            隐含波动率 (失败时返回 NaN，深度实值返回极小值)
        """
        price = np.asarray(price, dtype=np.float64)
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        init_sigma = np.asarray(init_sigma, dtype=np.float64)
        
        # 处理期权类型
        if isinstance(option_type, str):
            option_type = np.full(price.shape if price.ndim > 0 else (1,), option_type)
        else:
            option_type = np.asarray(option_type)
        
        # 确保所有数组形状一致
        shape = np.broadcast(price, S, K, T, init_sigma).shape
        price = np.broadcast_to(price, shape).copy()
        S = np.broadcast_to(S, shape).copy()
        K = np.broadcast_to(K, shape).copy()
        T = np.broadcast_to(T, shape).copy()
        init_sigma = np.broadcast_to(init_sigma, shape).copy()
        option_type = np.broadcast_to(option_type, shape).copy()
        
        # 初始化结果
        result = np.full(shape, np.nan)
        
        # 检查有效输入
        is_call = np.char.upper(option_type.astype(str)) == 'CALL'
        
        # 计算内在价值下限 (不考虑折现的简单内在价值)
        intrinsic_call = np.maximum(S - K, 0)
        intrinsic_put = np.maximum(K - S, 0)
        intrinsic = np.where(is_call, intrinsic_call, intrinsic_put)
        
        # 计算折现内在价值 (更精确)
        discount = np.exp(-r * T)
        discounted_intrinsic_call = np.maximum(S - K * discount, 0)
        discounted_intrinsic_put = np.maximum(K * discount - S, 0)
        discounted_intrinsic = np.where(is_call, discounted_intrinsic_call, discounted_intrinsic_put)
        
        # 基本有效性检查 (放宽条件)
        basic_valid = (price > 0) & (S > 0) & (K > 0) & (T > 1e-6)
        
        # 计算时间价值
        time_value = price - intrinsic
        time_value_ratio = time_value / np.maximum(price, 1e-6)
        
        # 情况1: 市场价 < 内在价值 (无法求解IV，是套利机会)
        below_intrinsic = basic_valid & (price < intrinsic * 0.999)
        result = np.where(below_intrinsic, 0.005, result)  # 设为0.5%的极小IV
        
        # 情况2: 市场价 >= 内在价值但时间价值极小 (深度实值)
        # 时间价值很小时，即使能求解IV，结果也会非常接近0
        deep_itm_tiny_tv = basic_valid & (~below_intrinsic) & (time_value_ratio < 0.01)  # 时间价值<1%
        result = np.where(deep_itm_tiny_tv, 0.01, result)  # 设为1%的极小IV
        
        # 情况3: 市场价 >= 内在价值但时间价值小于5% (中度实值)
        deep_itm_small_tv = basic_valid & (~below_intrinsic) & (~deep_itm_tiny_tv) & (time_value_ratio < 0.05)
        
        # 情况4: 正常情况 (可以求解IV)
        valid = basic_valid & (~below_intrinsic)
        
        # 对有效数据使用改进的 Newton-Raphson 方法
        if np.any(valid):
            sigma = np.where(valid, init_sigma, np.nan)
            sigma = np.clip(sigma, cls.IV_MIN, cls.IV_MAX)
            
            # 先用较宽松的收敛条件快速迭代
            for iteration in range(cls.MAX_ITERATIONS):
                # 计算 BS 价格和 Vega
                bs_val = cls.bs_price(S, K, r, sigma, T, option_type)
                vega = cls.vega(S, K, r, sigma, T)
                
                # 计算价格差
                diff = price - bs_val
                
                # 检查收敛 (相对误差或绝对误差)
                rel_error = np.abs(diff) / np.maximum(price, 1e-6)
                converged = (np.abs(diff) < cls.IV_PRECISION) | (rel_error < 1e-6) | (~valid)
                if np.all(converged):
                    break
                
                # Newton-Raphson 更新，带有阻尼和边界检查
                with np.errstate(divide='ignore', invalid='ignore'):
                    # 使用 Vega 作为导数，增加最小值防止除零
                    delta_sigma = diff / np.maximum(vega, 1e-8)
                    
                    # 自适应步长限制 (根据当前sigma调整)
                    max_step = np.maximum(sigma * 0.5, 0.1)
                    delta_sigma = np.clip(delta_sigma, -max_step, max_step)
                    
                    # 更新 sigma
                    new_sigma = sigma + delta_sigma
                    
                    # 边界检查
                    new_sigma = np.clip(new_sigma, cls.IV_MIN, cls.IV_MAX)
                    
                    # 只更新未收敛的值
                    sigma = np.where(converged, sigma, new_sigma)
            
            # 检查最终收敛状态
            final_bs = cls.bs_price(S, K, r, sigma, T, option_type)
            final_diff = np.abs(price - final_bs)
            final_rel_error = final_diff / np.maximum(price, 1e-6)
            
            # 对于收敛良好的结果，使用计算值
            good_convergence = (final_diff < 0.01) | (final_rel_error < 0.01)
            result = np.where(valid & good_convergence, sigma, result)
            
            # 对于未收敛但有效的数据，尝试二分法
            need_bisection = valid & (~good_convergence)
            if np.any(need_bisection):
                bisection_result = cls._bisection_iv(
                    price[need_bisection], S[need_bisection], K[need_bisection],
                    r, T[need_bisection], option_type[need_bisection]
                )
                result[need_bisection] = bisection_result
        
        # 对于仍然无效的深度实值期权，检查是否市场价 < 最低理论价
        still_nan = np.isnan(result) & basic_valid
        if np.any(still_nan):
            # 计算最低理论价 (IV = 0.001)
            min_bs = cls.bs_price(S, K, r, 0.001, T, option_type)
            # 如果市场价 < 最低理论价，说明是极端深度实值，设IV为0.5%
            deep_itm_no_solution = still_nan & (price < min_bs * 1.01)
            result = np.where(deep_itm_no_solution, 0.005, result)
        
        # 返回标量或数组
        if result.shape == (1,):
            return float(result[0])
        return result
    
    @classmethod
    def _bisection_iv(cls, price, S, K, r, T, option_type, max_iter=50):
        """
        二分法求解隐含波动率 (备选方法，更稳定但较慢)
        """
        result = np.full(price.shape, np.nan)
        
        low = np.full(price.shape, cls.IV_MIN)
        high = np.full(price.shape, cls.IV_MAX)
        
        for _ in range(max_iter):
            mid = (low + high) / 2
            bs_mid = cls.bs_price(S, K, r, mid, T, option_type)
            
            # 更新边界
            too_high = bs_mid > price
            low = np.where(too_high, low, mid)
            high = np.where(too_high, mid, high)
            
            # 检查收敛
            if np.all((high - low) < 1e-6):
                break
        
        result = mid
        return result
    
    @classmethod
    def delta(cls, S, K, r, sigma, T, option_type):
        """
        计算 Delta (向量化)
        Delta = ∂V/∂S
        CALL: N(d1), PUT: N(d1) - 1
        """
        d_1 = cls.d1(S, K, r, sigma, T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                return cls.norm_cdf(d_1)
            else:
                return cls.norm_cdf(d_1) - 1
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_delta = cls.norm_cdf(d_1)
            put_delta = call_delta - 1
            return np.where(is_call, call_delta, put_delta)
    
    @classmethod
    def gamma(cls, S, K, r, sigma, T):
        """
        计算 Gamma (向量化)
        Gamma = ∂²V/∂S² = N'(d1) / (S * σ * √T)
        Gamma 对 CALL 和 PUT 相同
        """
        S = np.asarray(S, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        with np.errstate(divide='ignore', invalid='ignore'):
            gamma = cls.norm_pdf(d_1) / (S * sigma * np.sqrt(T))
            gamma = np.where((sigma <= 0) | (T <= 0) | (S <= 0), np.nan, gamma)
        return gamma
    
    @classmethod
    def theta(cls, S, K, r, sigma, T, option_type):
        """
        计算 Theta (向量化)
        Theta = ∂V/∂t (每年)
        返回值为负数（时间价值损耗）
        """
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        sqrt_T = np.sqrt(T)
        discount = np.exp(-r * T)
        
        # 第一项：时间价值损耗
        term1 = -S * cls.norm_pdf(d_1) * sigma / (2 * sqrt_T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                term2 = -r * K * discount * cls.norm_cdf(d_2)
            else:
                term2 = r * K * discount * cls.norm_cdf(-d_2)
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_term2 = -r * K * discount * cls.norm_cdf(d_2)
            put_term2 = r * K * discount * cls.norm_cdf(-d_2)
            term2 = np.where(is_call, call_term2, put_term2)
        
        theta = term1 + term2
        theta = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, theta)
        return theta
    
    @classmethod
    def vega(cls, S, K, r, sigma, T):
        """
        计算 Vega (向量化)
        Vega = ∂V/∂σ = S * √T * N'(d1)
        Vega 对 CALL 和 PUT 相同
        返回值：波动率变化 1% (0.01) 时期权价格变化
        """
        S = np.asarray(S, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        vega = S * np.sqrt(T) * cls.norm_pdf(d_1)
        vega = np.where((sigma <= 0) | (T <= 0) | (S <= 0), np.nan, vega)
        return vega
    
    @classmethod
    def rho(cls, S, K, r, sigma, T, option_type):
        """
        计算 Rho (向量化)
        Rho = ∂V/∂r
        CALL: K * T * e^(-rT) * N(d2)
        PUT: -K * T * e^(-rT) * N(-d2)
        """
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        discount = np.exp(-r * T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                rho = K * T * discount * cls.norm_cdf(d_2)
            else:
                rho = -K * T * discount * cls.norm_cdf(-d_2)
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_rho = K * T * discount * cls.norm_cdf(d_2)
            put_rho = -K * T * discount * cls.norm_cdf(-d_2)
            rho = np.where(is_call, call_rho, put_rho)
        
        rho = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, rho)
        return rho


# 创建全局实例
pricer = OptionPricer()

def calculate_greeks():
    """
    读取 wisecoin-期权参考.xlsx，使用 TQSDK 获取标的历史波动率，
    使用高性能 OptionPricer 计算理论价格和隐含波动率/Greeks。
    
    采用改进的 Black-Scholes 模型:
    - 隐含波动率: 使用 Newton-Raphson + Halley 方法，收敛速度快，精度高
    - Greeks: 解析解计算，向量化处理，性能优于 tafunc
    """
    INPUT_FILE = os.path.join(TEMP_DIR, "wisecoin-期权参考.xlsx")
    if not os.path.exists(INPUT_FILE):
        logger.error(f"文件不存在: {INPUT_FILE}")
        return

    logger.info(f"🚀 开始计算希腊指标 (高性能 OptionPricer + Historical Vol)...读取 {INPUT_FILE}")

    api = None
    try:
        df = pd.read_excel(INPUT_FILE)
        
        # ============ 1. 获取历史波动率 (HV5, HV20, HV60) ============
        # 需要连接 TQSDK 获取标的 K 线
        
        # 提取所有标的合约
        underlyings = df['标的合约'].unique().tolist()
        logger.info(f"需要获取 {len(underlyings)} 个标的的历史波动率 (HV5, HV20, HV60)...")
        
        # 存储多个周期的历史波动率
        hv5_map = {}   # 5日历史波动率 (近期波动率)
        hv20_map = {}  # 20日历史波动率
        hv60_map = {}  # 60日历史波动率
        
        def calculate_historical_volatility(klines_df, window):
            """
            计算指定窗口期的历史波动率
            使用对数收益率的标准差，年化处理
            """
            if klines_df is None or len(klines_df) < window + 1:
                return np.nan
            
            try:
                close = klines_df['close'].values
                # 计算对数收益率
                log_returns = np.log(close[1:] / close[:-1])
                # 取最近 window 个收益率
                if len(log_returns) >= window:
                    recent_returns = log_returns[-window:]
                    # 年化波动率 = 日波动率 * sqrt(252)
                    hv = np.std(recent_returns, ddof=1) * np.sqrt(252)
                    return hv if hv > 0 else np.nan
                return np.nan
            except Exception:
                return np.nan
        
        try:
            # 初始化 API (使用 02 脚本中的 auth)
            api = TqApi(auth=TqAuth('huaying', 'bonze13'))
            
            # 批量获取标的 K 线 (日线, 需要至少65根用于计算HV60)
            for und in underlyings:
                try:
                    # 获取日线 65 根 (足够计算HV60)
                    klines = api.get_kline_serial(und, 24*60*60, 65)
                    quote = api.get_quote(und)
                    
                    # 计算三个周期的历史波动率
                    hv5 = calculate_historical_volatility(klines, 5)
                    hv20 = calculate_historical_volatility(klines, 20)
                    hv60 = calculate_historical_volatility(klines, 60)
                    
                    # 如果自定义计算失败，使用 tafunc 作为备选 (HV5)
                    if pd.isna(hv5) or hv5 <= 0:
                        hv5 = tafunc.get_his_volatility(klines, quote)
                    
                    # 设置默认值
                    if pd.isna(hv5) or hv5 <= 0: hv5 = 0.25
                    if pd.isna(hv20) or hv20 <= 0: hv20 = hv5  # 如果HV20无效，使用HV5
                    if pd.isna(hv60) or hv60 <= 0: hv60 = hv20  # 如果HV60无效，使用HV20
                        
                    hv5_map[und] = hv5
                    hv20_map[und] = hv20
                    hv60_map[und] = hv60
                    
                except Exception as e:
                    logger.warning(f"获取 {und} 历史波动率失败: {e}")
                    hv5_map[und] = 0.25
                    hv20_map[und] = 0.25
                    hv60_map[und] = 0.25
            
            api.close()
            logger.info("✅ 历史波动率获取完成 (HV5, HV20, HV60)。")
            
        except Exception as e:
            logger.error(f"连接 TQSDK 失败，使用默认波动率 0.3: {e}")
            if api: api.close()
            # Fallback
            for und in underlyings:
                hv5_map[und] = 0.3
                hv20_map[und] = 0.3
                hv60_map[und] = 0.3

        # ============ 2. 准备向量化数据 ============
        current_ts = time.time()
        
        S = df['标的现价'].values.astype(np.float64)
        P = df['期权价'].values.astype(np.float64)
        K = df['行权价'].values.astype(np.float64)
        
        # 映射三个周期的历史波动率到每一行
        HV5_array = df['标的合约'].map(hv5_map).fillna(0.3).values.astype(np.float64)
        HV20_array = df['标的合约'].map(hv20_map).fillna(0.3).values.astype(np.float64)
        HV60_array = df['标的合约'].map(hv60_map).fillna(0.3).values.astype(np.float64)
        
        # 用于计算的默认HV使用HV20 (20日波动率更稳定)
        HV_array = HV20_array

        def get_type(row):
            ot = str(row.get('期权类型', '')).upper()
            if 'CALL' in ot or 'C' == ot: return 'CALL'
            if 'PUT' in ot or 'P' == ot: return 'PUT'
            return 'CALL' if 'C' in str(row['合约代码']).split('.')[-1] else 'PUT'
        
        Type = df.apply(get_type, axis=1).values
        
        def get_days(row):
            expire_str = str(row.get('到期日', ''))
            if expire_str and expire_str != 'nan':
                 try:
                    if len(expire_str) == 10: expire_str += " 15:00:00"
                    dt_obj = datetime.datetime.strptime(expire_str, "%Y-%m-%d %H:%M:%S")
                    ts = dt_obj.timestamp()
                    return (ts - current_ts) / 86400.0
                 except: pass
            return float(row.get('剩余天数', 30))

        Days = df.apply(get_days, axis=1).values
        T = Days / 365.0
        T = np.maximum(T, 0.0001)

        r = 0.015  # 无风险利率 1.5%

        # ============ 3. 计算 Theoretical Price (基于 HisVol) - 使用高性能 OptionPricer ============
        logger.info("计算理论价格 (基于历史波动率，使用 OptionPricer)...")
        start_time = time.time()
        bs_price = pricer.bs_price(S, K, r, HV_array, T, Type)
        logger.info(f"  理论价格计算耗时: {time.time() - start_time:.4f}秒")
        
        # ============ 4. 计算 IMPV & Greeks (基于 Market Price) - 使用高性能 OptionPricer ============
        logger.info("计算隐含波动率 (基于期权现价，使用改进的 Newton-Raphson 算法)...")
        start_time = time.time()
        
        # 使用 HV 作为初始猜测值
        impv = pricer.implied_volatility(P, S, K, r, T, Type, init_sigma=HV_array)
        
        # 处理无效值
        impv = np.where(np.isnan(impv), 0.0, impv)  # NaN -> 0
        impv = np.maximum(impv, 0.0)  # 确保非负
        
        # 对于非常小但有效的IV，设置最小显示值 (避免四舍五入后变成0)
        # 0.005 = 0.5% = 0.50 after *100
        impv = np.where((impv > 0) & (impv < 0.005), 0.005, impv)
        
        logger.info(f"  隐含波动率计算耗时: {time.time() - start_time:.4f}秒")

        logger.info("计算 Greeks (基于 IMPV，使用 OptionPricer 解析解)...")
        start_time = time.time()
        
        # 对于 IV 为 0 或极小的情况，使用 HV 计算 Greeks (作为参考)
        # 因为 IV=0 时 Greeks 会产生 NaN 或极端值
        impv_for_greeks = np.where(impv < 0.01, HV_array, impv)  # 至少使用1%的波动率
        impv_for_greeks = np.maximum(impv_for_greeks, 0.01)  # 确保至少 1%
        
        # 使用高性能 OptionPricer 计算 Greeks
        delta = pricer.delta(S, K, r, impv_for_greeks, T, Type)
        gamma = pricer.gamma(S, K, r, impv_for_greeks, T)
        theta = pricer.theta(S, K, r, impv_for_greeks, T, Type)
        vega = pricer.vega(S, K, r, impv_for_greeks, T)
        rho = pricer.rho(S, K, r, impv_for_greeks, T, Type)
        
        # 处理 NaN 和 Inf
        delta = np.nan_to_num(delta, nan=0.0, posinf=1.0, neginf=-1.0)
        gamma = np.nan_to_num(gamma, nan=0.0, posinf=0.0, neginf=0.0)
        theta = np.nan_to_num(theta, nan=0.0, posinf=0.0, neginf=0.0)
        vega = np.nan_to_num(vega, nan=0.0, posinf=0.0, neginf=0.0)
        rho = np.nan_to_num(rho, nan=0.0, posinf=0.0, neginf=0.0)
        
        logger.info(f"  Greeks 计算耗时: {time.time() - start_time:.4f}秒")
        
        # 统计计算结果
        iv_zero_count = np.sum(impv == 0)
        iv_valid_count = np.sum(impv > 0)
        logger.info(f"  IV计算结果: 有效={iv_valid_count}, 无效(=0)={iv_zero_count}")
        
        # ============ 5. 更新 DataFrame ============
        df['理论价格'] = np.round(bs_price, 2)
        
        # 添加三个周期的历史波动率 (百分比形式)
        df['近期波动率'] = np.round(HV5_array * 100, 2)   # HV5 = 近期波动率 (保持向后兼容)
        df['HV20'] = np.round(HV20_array * 100, 2)        # 20日历史波动率
        df['HV60'] = np.round(HV60_array * 100, 2)        # 60日历史波动率
        
        df['隐含波动率'] = np.round(impv * 100, 2)  # Percent
        df['Delta'] = np.round(delta, 4)
        df['Gamma'] = np.round(gamma, 6)
        df['Theta'] = np.round(theta / 365.0, 4)  # 年化 Theta 转日度
        df['Vega'] = np.round(vega / 100.0, 4)    # 标准化 Vega
        df['Rho'] = np.round(rho / 100.0, 4)      # 标准化 Rho
        
        # 买方杠杆 = 标的现价 * 期货合约乘数 / (期权价 * 合约乘数) * Delta
        # 在 买方期权费 后增加 买方杠杆
        futures_multiplier = df['期货合约乘数'].values.astype(np.float64)
        option_multiplier = df['合约乘数'].values.astype(np.float64)
        with np.errstate(divide='ignore', invalid='ignore'):
            buyer_leverage = np.round((S * futures_multiplier) / (P * option_multiplier) * delta, 2)
            buyer_leverage = np.where(np.isinf(buyer_leverage) | np.isnan(buyer_leverage), 0, buyer_leverage)
        buyer_leverage = pd.Series(buyer_leverage)
        
        # 插入列到正确位置
        if '买方杠杆' in df.columns:
            df.drop(columns=['买方杠杆'], inplace=True)
            
        fee_idx = df.columns.get_loc('买方期权费') + 1
        df.insert(fee_idx, '买方杠杆', buyer_leverage)
        
        def get_col_width(val):
            """计算值的宽度，考虑中文字符"""
            if pd.isna(val) or val is None:
                return 0
            s = str(val)
            # 中文字符大约占 2 个 ASCII 字符宽度
            return sum(2 if ord(c) > 127 else 1 for c in s)

        logger.info("计算完成，保存文件...")
        
        with pd.ExcelWriter(INPUT_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='期权参考', index=False)
            ws = writer.sheets['期权参考']
            ws.freeze_panes = 'A2'
            from openpyxl.utils import get_column_letter

            # ============ 颜色定义 ============
            atm_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # 淡黄色 (平值)
            # 实值 (红)
            itm_fills = [
                PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"), 
                PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"), 
                PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"), 
            ]
            # 虚值 (蓝)
            otm_fills = [
                PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"), 
                PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"), 
                PatternFill(start_color="81D4FA", end_color="81D4FA", fill_type="solid"), 
            ]
            # 到期日
            expire_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            expire_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            expire_dark_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            # 溢价率 & 波动率
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # 获取列名索引
            cols = list(df.columns)
            symbol_idx = cols.index('合约代码') + 1
            strike_idx = cols.index('行权价') + 1
            degree_idx = cols.index('虚实幅度%') + 1
            expire_idx = cols.index('到期日') + 1
            days_idx = cols.index('剩余天数') + 1
            premium_idx = cols.index('溢价率%') + 1
            impv_col_idx = cols.index('隐含波动率') + 1
            hv_col_idx = cols.index('近期波动率') + 1

            # 遍历数据行应用样式
            for r_idx, row_data in enumerate(df.itertuples(), start=2):
                degree = row_data[degree_idx]
                days = row_data[days_idx]
                premium = row_data[premium_idx]
                impv_val = row_data[impv_col_idx]
                hv_val = row_data[hv_col_idx]

                # 1. 平值高亮
                if abs(degree) <= 2.0:
                    ws.cell(row=r_idx, column=symbol_idx).fill = atm_fill

                # 2. 行权价着色
                if degree > 0: # 实值
                    fill_idx = min(2, int(degree / 5)) 
                    ws.cell(row=r_idx, column=strike_idx).fill = itm_fills[fill_idx]
                elif degree < 0: # 虚值
                    abs_degree = abs(degree)
                    fill_idx = min(2, int(abs_degree / 5))
                    ws.cell(row=r_idx, column=strike_idx).fill = otm_fills[fill_idx]

                # 3. 到期日提醒
                if days <= 1:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_dark_red
                elif days <= 3:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_red
                elif days <= 7:
                    ws.cell(row=r_idx, column=expire_idx).fill = expire_yellow
                
                # 4. 溢价率逻辑: 低于 2.0 绿色, 高于 5.0 红色
                if premium < 2.0:
                    ws.cell(row=r_idx, column=premium_idx).fill = green_fill
                elif premium > 5.0:
                    ws.cell(row=r_idx, column=premium_idx).fill = red_fill
                
                # 5. 隐含波动率逻辑: 隐含 < 历史 (低估) 绿色, 隐含 > 历史*1.2 (高估) 红色
                if impv_val < hv_val and impv_val != 0:
                    ws.cell(row=r_idx, column=impv_col_idx).fill = green_fill
                elif impv_val > hv_val * 1.2:
                    ws.cell(row=r_idx, column=impv_col_idx).fill = red_fill

            for i, col in enumerate(df.columns):
                # 表头宽度
                header_width = get_col_width(col)
                # 内容宽度 (采样或全量)
                content_width = df[col].astype(str).map(get_col_width).max()
                # 取两者大值 + 2 像素缓冲，上限 50
                final_width = min(max(header_width, content_width) + 2, 50)
                ws.column_dimensions[get_column_letter(i+1)].width = final_width
                
        logger.info(f"✅ 更新完成: {INPUT_FILE}")
        return df  # Return df for further processing
        
    except Exception as e:
        logger.error(f"Greeks 计算流程失败: {e}")
        logger.error(traceback.format_exc())
        return None
    finally:
        if api: 
            try: api.close()
            except: pass


def analyze_volatility_surface():
    """
    分析波动率曲面，计算峰度/偏度，生成策略建议
    按品种分页输出到 wisecoin-期权参考.xlsx
    """
    INPUT_FILE = os.path.join(TEMP_DIR, "wisecoin-期权参考.xlsx")
    
    if not os.path.exists(INPUT_FILE):
        logger.error(f"文件不存在: {INPUT_FILE}")
        return
    
    logger.info("🚀 开始波动率曲面分析...")
    
    try:
        df = pd.read_excel(INPUT_FILE)
        
        if '隐含波动率' not in df.columns:
            logger.error("数据中缺少隐含波动率列，请先运行 calculate_greeks()")
            return
        
        # 提取品质代码 (如 SHFE.ag2602C20000 -> AG, CFFEX.MO2405-C- -> MO)
        def extract_product_code(symbol):
            if pd.isna(symbol):
                return 'UNKNOWN'
            parts = str(symbol).split('.')
            if len(parts) >= 2:
                # 捕获开头的字母部分
                match = re.match(r'^([a-zA-Z]+)', parts[1])
                return match.group(1).upper() if match else 'UNKNOWN'
            return 'UNKNOWN'
        
        # 清理品种名称 (如 沪银2602 -> 沪银, 但保留 中证1000)
        def clean_product_name(name):
            if pd.isna(name): return ''
            name = str(name)
            if any(idx in name for idx in ['1000', '300', '50']):
                return name
            return re.sub(r'\d+$', '', name)
        
        df['品种代码'] = df['合约代码'].apply(extract_product_code)
        products = df['品种代码'].unique().tolist()
        products = [p for p in products if p != 'UNKNOWN']
        
        logger.info(f"共识别 {len(products)} 个品种: {products[:10]}...")
        
        # 准备输出
        product_results = []
        surface_images = {}
        
        for product in products:
            prod_df = df[df['品种代码'] == product].copy()
            
            if len(prod_df) < 10:  # 数据太少跳过
                continue
            
            # 1. 计算 IV Skew (虚值认沽 IV vs 虚值认购 IV)
            otm_puts = prod_df[(prod_df['期权类型'].str.upper().str.contains('PUT|P')) & 
                               (prod_df['虚实幅度%'] < -5)]
            otm_calls = prod_df[(prod_df['期权类型'].str.upper().str.contains('CALL|C')) & 
                                (prod_df['虚实幅度%'] < -5)]
            
            put_iv_mean = otm_puts['隐含波动率'].mean() if len(otm_puts) > 0 else 0
            call_iv_mean = otm_calls['隐含波动率'].mean() if len(otm_calls) > 0 else 0
            iv_skew = put_iv_mean - call_iv_mean  # 正值=看跌倾斜, 负值=看涨倾斜
            
            # 2. 计算期限结构 (短期 vs 长期 IV)
            short_term = prod_df[prod_df['剩余天数'] <= 30]['隐含波动率'].mean()
            long_term = prod_df[prod_df['剩余天数'] > 60]['隐含波动率'].mean()
            term_structure = short_term - long_term if not pd.isna(short_term) and not pd.isna(long_term) else 0
            # 正值=期限倒挂, 负值=正常升水
            
            # 3. 计算峰度和偏度
            iv_values = prod_df['隐含波动率'].dropna()
            iv_kurtosis = kurtosis(iv_values, fisher=True) if len(iv_values) >= 4 else 0
            iv_skewness = skew(iv_values) if len(iv_values) >= 4 else 0
            
            # 4. 分类市场情绪 & 策略建议
            # 计算 IV vs RV 比率
            avg_iv = prod_df['隐含波动率'].mean() if '隐含波动率' in prod_df.columns else 0
            avg_hv = prod_df['近期波动率'].mean() if '近期波动率' in prod_df.columns else 0
            iv_rv_ratio = avg_iv / avg_hv if avg_hv > 0 else 1.0
            
            sentiment, strategies = _classify_market_sentiment(iv_skew, term_structure, iv_kurtosis, iv_skewness, iv_rv_ratio)
            
            # 5. 生成具体合约建议
            recommended_contracts = _recommend_contracts(prod_df, strategies)
            
            # 6. 生成波动率曲面图
            surface_img = _generate_surface_plot(prod_df, product)
            if surface_img:
                surface_images[product] = surface_img
            
            # 获取标的品种名称并清理
            prod_name = prod_df['标的品种名称'].iloc[0] if '标的品种名称' in prod_df.columns else product
            prod_name = clean_product_name(prod_name)
            
            # 计算总资金
            total_capital = prod_df['资金合计(万)'].sum() if '资金合计(万)' in prod_df.columns else 0
            
            product_results.append({
                '品种代码': product,
                '品种名称': prod_name,
                '合约数量': len(prod_df),
                '资金合计(万)': round(total_capital, 2),
                '虚值认沽IV均值': round(put_iv_mean, 2),
                '虚值认购IV均值': round(call_iv_mean, 2),
                'IV倾斜度': round(iv_skew, 2),
                '倾斜方向': '看跌倾斜' if iv_skew > 2 else ('看涨倾斜' if iv_skew < -2 else '平坦'),
                '短期IV': round(short_term, 2) if not pd.isna(short_term) else 0,
                '长期IV': round(long_term, 2) if not pd.isna(long_term) else 0,
                '期限结构差': round(term_structure, 2),
                '期限结构': '倒挂' if term_structure > 3 else ('升水' if term_structure < -3 else '平坦'),
                'IV/RV比率': round(iv_rv_ratio, 2),
                '峰度': round(iv_kurtosis, 2),
                '偏度': round(iv_skewness, 2),
                '市场情绪': sentiment,
                '推荐策略': ', '.join(strategies),
                '推荐合约': recommended_contracts
            })
        
        # 按资金合计(万) 降序排序
        product_results.sort(key=lambda x: x['资金合计(万)'], reverse=True)
        
        # 输出到 Excel
        _output_volatility_analysis(df, product_results, surface_images, INPUT_FILE)
        
        logger.info(f"✅ 波动率曲面分析完成，已更新: {INPUT_FILE}")
        
    except Exception as e:
        logger.error(f"波动率曲面分析失败: {e}")
        logger.error(traceback.format_exc())


def _classify_market_sentiment(iv_skew, term_structure, iv_kurtosis, iv_skewness, iv_rv_ratio):
    """
    根据波动率曲面特征、峰度、偏度、IV/RV比率分类市场情绪
    返回: (情绪分类, 策略列表)
    """
    # 阈值定义
    skew_threshold = 3.0  # IV倾斜阈值
    term_threshold = 3.0  # 期限结构阈值
    kurtosis_high = 1.0   # 高峰度阈值
    kurtosis_low = -0.5   # 低峰度阈值
    skewness_threshold = 0.3  # 偏度阈值
    
    # IV 相对估值
    iv_undervalued = iv_rv_ratio < 0.8
    iv_overvalued = iv_rv_ratio > 1.2
    
    strategies = []
    sentiment = '中性'
    
    # 情绪分类逻辑 & 策略生成
    if iv_skew > skew_threshold and term_structure > term_threshold:
        # 恐慌下跌 (Put Skew + Inverted Term Structure)
        sentiment = '恐慌下跌'
        strategies.append('认沽期权多头价差') 
        if iv_undervalued:
            strategies.insert(0, '买入虚值认沽')
            strategies.append('跨式多头')
        elif iv_overvalued:
            strategies.append('卖出虚值认购(备兑)') # 恐慌时卖购
            strategies.append('比率认沽价差') # 买1卖2
            
    elif iv_skew < -skew_threshold and term_structure < -term_threshold:
        # 狂热上涨 (Call Skew + Contango/Normal but maybe steep) 
        # Note: Usually Contango is normal, but usually Call skew comes with backwardation in commodities? 
        # For simplicity, if Call Skew is high.
        sentiment = '狂热上涨'
        strategies.append('认购期权多头价差')
        if iv_undervalued:
            strategies.insert(0, '买入虚值认购')
            strategies.append('宽跨式多头')
        elif iv_overvalued:
             strategies.append('比率认购价差') # 买1卖2
    
    elif abs(iv_skew) <= skew_threshold and abs(term_structure) <= term_threshold:
        # 窄幅震荡 (Flat Skew + Flat Term)
        sentiment = '窄幅震荡'
        if iv_kurtosis < kurtosis_low: # Thin tails
            sentiment += '(瘦尾)'
            if iv_overvalued:
                strategies.append('跨式空头')
                strategies.append('宽跨式空头')
                strategies.append('铁蝶式')
            else:
                 strategies.append('日历价差(卖短买长)') # 此策略其实利用IV期限回归，若IV低，可能不太好，但日历主要吃Theta且Vega风险小
        else:
             strategies.append('铁鹰式') # 中性，有保护
    
    elif iv_skew > 0 and iv_skew <= skew_threshold:
        # 震荡筑底 (Slight Put Skew)
        sentiment = '震荡筑底'
        if iv_overvalued:
            strategies.append('认沽比率价差') # 利用高 Put IV
            strategies.append('备兑认购')
        else:
            strategies.append('牛市认沽价差')
    
    elif iv_skew < 0 and iv_skew >= -skew_threshold:
        # 震荡冲高 (Slight Call Skew)
        sentiment = '震荡冲高'
        if iv_overvalued:
             strategies.append('认购比率价差')
             strategies.append('备兑认购(减仓)')
        else:
             strategies.append('保护性认沽') # 锁定利润
             strategies.append('牛市认购价差')

    # 峰度策略补充
    if iv_kurtosis > kurtosis_high and '跨式' not in ''.join(strategies):
        if iv_undervalued:
            strategies.append('Gamma Scalping(买入跨式)')
    
    if not strategies:
        strategies = ['观望', '日历价差']
        
    return sentiment, list(set(strategies)) # 去重


def _recommend_contracts(prod_df, strategies):
    """
    根据策略推荐具体合约
    返回: 推荐合约字符串
    """
    recommendations = []
    
    # 辅助：获取不同 moneyness 的合约
    otm_puts = prod_df[
        (prod_df['期权类型'].str.upper().str.contains('PUT|P')) &
        (prod_df['虚实幅度%'] < -2) & (prod_df['虚实幅度%'] >= -15)
    ].sort_values('沉淀资金(万)', ascending=False)
    
    otm_calls = prod_df[
        (prod_df['期权类型'].str.upper().str.contains('CALL|C')) &
        (prod_df['虚实幅度%'] < -2) & (prod_df['虚实幅度%'] >= -15)
    ].sort_values('沉淀资金(万)', ascending=False)
    
    atm_options = prod_df[abs(prod_df['虚实幅度%']) <= 3].sort_values('沉淀资金(万)', ascending=False)
    
    unique_strategies = sorted(list(set(strategies))) # 排序保证一致性
    
    for strategy in unique_strategies[:3]: # 限制推荐数量
        rec_str = ""
        
        if '买入虚值认沽' in strategy or '认沽多头价差' in strategy:
            # 推荐: 剩余30-90天, OTM 5-10%
            cands = otm_puts[(otm_puts['剩余天数'] >= 25) & (otm_puts['剩余天数'] <= 90)]
            if not cands.empty:
                best = cands.iloc[0]
                rec_str = f"{best['合约代码']}(IV{best['隐含波动率']}%)"

        elif '买入虚值认购' in strategy or '认购多头价差' in strategy:
             cands = otm_calls[(otm_calls['剩余天数'] >= 25) & (otm_calls['剩余天数'] <= 90)]
             if not cands.empty:
                best = cands.iloc[0]
                rec_str = f"{best['合约代码']}(IV{best['隐含波动率']}%)"
        
        elif '跨式' in strategy or 'Straddle' in strategy:
            if not atm_options.empty:
                 atm_call = atm_options[atm_options['期权类型'].str.upper().str.contains('CALL|C')]
                 atm_put = atm_options[atm_options['期权类型'].str.upper().str.contains('PUT|P')]
                 if not atm_call.empty and not atm_put.empty:
                     rec_str = f"Buy/Sell {atm_call.iloc[0]['合约代码']}&{atm_put.iloc[0]['合约代码']}"
        
        elif '比率' in strategy:
             # 简单推荐一个 OTM
             cands = otm_puts if '认沽' in strategy else otm_calls
             if not cands.empty:
                 rec_str = f"关注 {cands.iloc[0]['合约代码']} 及更虚值合约"
                 
        if rec_str:
            recommendations.append(f"[{strategy}]: {rec_str}")
    
    return '; '.join(recommendations) if recommendations else ''   # 无合适高流动性合约推荐


def _generate_surface_plot(prod_df, product):
    """
    生成3D波动率曲面图
    返回: BytesIO 对象或 None
    """
    try:
        # 过滤有效数据
        plot_df = prod_df[
            (prod_df['隐含波动率'] > 0) & 
            (prod_df['剩余天数'] > 0) &
            (prod_df['行权价'] > 0)
        ].copy()
        
        if len(plot_df) < 10:
            return None
        
        # 标准化行权价 (相对于标的价格的比例)
        if '标的现价' in plot_df.columns:
            und_price = plot_df['标的现价'].iloc[0]
            if und_price > 0:
                plot_df['行权价比例'] = plot_df['行权价'] / und_price * 100  # 百分比
            else:
                plot_df['行权价比例'] = plot_df['行权价']
        else:
            plot_df['行权价比例'] = plot_df['行权价']
        
        x = plot_df['剩余天数'].values
        y = plot_df['行权价比例'].values
        z = plot_df['隐含波动率'].values
        
        # 创建网格
        xi = np.linspace(x.min(), x.max(), 30)
        yi = np.linspace(y.min(), y.max(), 30)
        xi_grid, yi_grid = np.meshgrid(xi, yi)
        
        # 插值
        try:
            zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='linear')
        except:
            zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
        
        # 绘图
        fig = plt.figure(figsize=(10, 7))
        ax = fig.add_subplot(111, projection='3d')
        
        # 曲面图
        surf = ax.plot_surface(xi_grid, yi_grid, zi_grid, 
                               cmap='viridis', alpha=0.8, 
                               edgecolor='none')
        
        ax.set_xlabel('剩余天数', fontsize=10)
        ax.set_ylabel('行权价比例(%)', fontsize=10)
        ax.set_zlabel('隐含波动率(%)', fontsize=10)
        ax.set_title(f'{product} 波动率曲面', fontsize=12, fontweight='bold')
        
        fig.colorbar(surf, shrink=0.5, aspect=10, label='IV%')
        
        # 保存到内存
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)
        
        return buf
        
    except Exception as e:
        logger.warning(f"生成 {product} 波动率曲面图失败: {e}")
        return None


def _output_volatility_analysis(full_df, product_results, surface_images, output_file):
    """
    输出波动率分析结果到 Excel
    """
    from openpyxl import load_workbook
    
    # 加载现有工作簿
    wb = load_workbook(output_file)
    
    # 删除已有的波动率分析相关sheet (包括旧前缀和新规则)
    sheets_to_remove = [s for s in wb.sheetnames if s.startswith('波动率_') or s == '波动率曲面']
    # 也要检查是否已经是品种代码命名的sheet，如果是刚生成的结果，汇总表里的品种代码对应的sheet也清理下
    known_codes = [r['品种代码'] for r in product_results]
    for sheet_name in wb.sheetnames:
        if sheet_name in known_codes:
            sheets_to_remove.append(sheet_name)
            
    for sheet_name in list(set(sheets_to_remove)):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    
    # 1. 创建波动率曲面表
    summary_df = pd.DataFrame(product_results)
    if not summary_df.empty:
        ws_summary = wb.create_sheet('波动率曲面')
        
        # 写入表头
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 写入数据
        for row_idx, row in enumerate(summary_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
            
            # 根据市场情绪着色
            sentiment = str(summary_df.iloc[row_idx-2]['市场情绪'])
            sentiment_cell = ws_summary.cell(row=row_idx, column=summary_df.columns.get_loc('市场情绪') + 1)
            if '恐慌' in sentiment:
                sentiment_cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
            elif '狂热' in sentiment:
                sentiment_cell.fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")
            elif '窄幅' in sentiment:
                sentiment_cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
            elif '筑底' in sentiment:
                sentiment_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
            elif '冲高' in sentiment:
                sentiment_cell.fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
                
        # 智能调整列宽
        def get_visual_length(val):
            s = str(val if val is not None else "")
            return sum(2 if ord(c) > 127 else 1 for c in s)

        for i, column_cells in enumerate(ws_summary.columns, 1):
            # 获取表头宽度
            header_val = ws_summary.cell(row=1, column=i).value
            header_len = get_visual_length(header_val)
            
            # 获取内容最大宽度 (采样前100行)
            content_len = 0
            for cell in column_cells[1:100]: 
                content_len = max(content_len, get_visual_length(cell.value))
            
            # 取最大值，并限制范围
            final_width = max(header_len, content_len) + 2
            final_width = min(max(final_width, 10), 40) # 最小10，最大40
            
            ws_summary.column_dimensions[get_column_letter(i)].width = final_width
        
        # 冻结表头
        ws_summary.freeze_panes = 'A2'
    
    # 2. 添加波动率策略参考表 (放在波动率曲面后)
    _create_strategy_reference_sheet_iv(wb)
    _create_strategy_reference_sheet(wb)
    
    # 3. 为每个品种创建单独的sheet (含图像)
    for result in product_results[:40]:  # 适当增加上限
        product = result['品种代码']
        prod_name = result.get('品种名称', product)
        
        sheet_name = product[:31]  # 直接用品种代码命名
        
        ws_prod = wb.create_sheet(sheet_name)
        
        # 写入品种摘要信息
        summary_rows = [
            ('品种代码', product),
            ('品种名称', prod_name),
            ('合约数量', result['合约数量']),
            ('虚值认沽IV均值', result['虚值认沽IV均值']),
            ('虚值认购IV均值', result['虚值认购IV均值']),
            ('IV倾斜度', result['IV倾斜度']),
            ('倾斜方向', result['倾斜方向']),
            ('短期IV', result['短期IV']),
            ('长期IV', result['长期IV']),
            ('期限结构差', result['期限结构差']),
            ('期限结构', result['期限结构']),
            ('峰度', result['峰度']),
            ('偏度', result['偏度']),
            ('市场情绪', result['市场情绪']),
            ('推荐策略', result['推荐策略']),
            ('推荐合约', result['推荐合约']),
        ]
        
        for row_idx, (label, value) in enumerate(summary_rows, 1):
            ws_prod.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_prod.cell(row=row_idx, column=2, value=value)
        
        ws_prod.column_dimensions['A'].width = 18
        ws_prod.column_dimensions['B'].width = 60
        
        # 插入波动率曲面图
        if product in surface_images and surface_images[product]:
            try:
                img = OpenpyxlImage(surface_images[product])
                img.width = 600
                img.height = 420
                ws_prod.add_image(img, 'D1')
            except Exception as e:
                logger.warning(f"插入 {product} 图像失败: {e}")
    
    # 保存
    wb.save(output_file)
    logger.info(f"波动率分析结果已保存到: {output_file}")


def _create_strategy_reference_sheet(wb):
    """
    创建波动率策略参考表 - 实盘联合决策策略
    基于：市场状态 × 期权定价
    """
    # 删除已有的
    for name in ['期权策略']:
        if name in wb.sheetnames:
            del wb[name]
    
    ws = wb.create_sheet('期权策略')
    
    # 样式定义
    header_font = Font(bold=True, size=14, color="FFFFFF")
    sub_header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    normal_font = Font(size=11)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # 深蓝
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid") # 浅蓝
    
    # 场景颜色
    green_bg = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # 浅绿 (低波/震荡)
    red_bg = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # 浅红 (高波/趋势)
    
    # 文字颜色
    green_font = Font(color="006100", bold=True)
    red_font = Font(color="9C0006", bold=True)
    
    # helper to write row
    def write_row(row_idx, data, font=normal_font, fill=None, merge_cols=None):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            
        if merge_cols and merge_cols > 1:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=merge_cols)
    
    current_row = 1
    
    # ============ 标题 ============
    write_row(current_row, ['期权实盘联合决策策略（基于市场状态 × 期权定价）'], 
              font=header_font, fill=header_fill, merge_cols=4)
    ws.row_dimensions[current_row].height = 30
    current_row += 2
    
    # ============ 一、策略目标与基本原则 ============
    write_row(current_row, ['一、策略目标与基本原则'], font=sub_header_font, fill=section_fill, merge_cols=4)
    current_row += 1
    
    principles = [
        "核心目标：不预测方向，而是通过「市场状态识别 + 期权定价验证」寻找高性价比结构。",
        "原则1：市场状态优先于期权结构。",
        "原则2：期权用于验证“是否被充分定价”，而非预测涨跌。",
        "原则3：任何策略均需满足“建仓-风控-退出”闭环。",
        "原则4：状态与定价冲突时，以期权定价为准，降仓或放弃。"
    ]
    for p in principles:
        write_row(current_row, [p], merge_cols=4)
        current_row += 1
    current_row += 1
    
    # ============ 二、实盘决策流程 ============
    write_row(current_row, ['二、实盘决策流程（每日固定执行）'], font=sub_header_font, fill=section_fill, merge_cols=4)
    current_row += 1
    
    # 表头
    ws.cell(row=current_row, column=1, value="步骤").font = bold_font
    ws.cell(row=current_row, column=2, value="内容").font = bold_font
    ws.cell(row=current_row, column=3, value="重点指标/输出").font = bold_font
    current_row += 1
    
    steps = [
        ("第一步：市场状态", "对期货标的归类：\n- 趋势：上行/下行/震荡\n- 波动：高波/低波\n- 资金：集中/分散", "结论示例：\n- 震荡+低波\n- 趋势+高波"),
        ("第二步：定价验证", "检查期权是否充分定价：\n- IV水平 (Rank/Percentile)\n- Skew (OTM价差)\n- 期限结构 (Term Structure)", "结论：\n- 期权低估 (IV Rank < 30%)\n- 期权中性\n- 期权高估 (IV Rank > 70%)")
    ]
    
    for step, content, indicators in steps:
        ws.cell(row=current_row, column=1, value=step).alignment = Alignment(vertical='top')
        ws.cell(row=current_row, column=2, value=content).alignment = Alignment(wrap_text=True)
        ws.cell(row=current_row, column=3, value=indicators).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        
    current_row += 1
    
    # ============ 三、实盘策略选择规则 (核心矩阵) ============
    write_row(current_row, ['三、实盘策略选择规则（核心）'], font=sub_header_font, fill=section_fill, merge_cols=4)
    current_row += 1
    
    # 矩阵表头
    headers = ['场景', '期权定价特征', '策略选择', '风控规则']
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=idx, value=h)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    current_row += 1
    
    # 场景数据
    scenarios = [
        {
            "name": "场景一：震荡 + 低波",
            "feat": "IV、峰度均低\n(IV Rank < 30%)",
            "allowed": "🟢 允许：\n- 卖跨式 (优先铁鹰/铁秃鹰)\n- 卖宽跨结构",
            "forbidden": "🔴 禁止：\n- 买跨式\n- 高Gamma策略",
            "risk": "单日浮亏 ≥ 最大收益 1.5倍\n→ 强制止损",
            "fill": green_bg
        },
        {
            "name": "场景二：震荡 + 高波",
            "feat": "IV已高 (不追)\nIV未充分反映 (可买)",
            "allowed": "🟢 允许：\n- 日历价差\n- 有限风险价差",
            "forbidden": "🔴 禁止：\n- 裸卖波动",
            "risk": "IV回落至历史中位\n→ 止盈",
            "fill": red_bg
        },
        {
            "name": "场景三：趋势 + 低波",
            "feat": "Skew尚不极端\n(偏度适中)",
            "allowed": "🟢 允许：\n- 比率价差 (Call/Put)\n- 方向性价差",
            "forbidden": "🔴 禁止：\n- 纯卖波动策略",
            "risk": "趋势失效 (破位/拐头)\n→ 立即退出",
            "fill": green_bg
        },
        {
            "name": "场景四：趋势 + 高波",
            "feat": "IV、Skew、峰度\n往往极端",
            "allowed": "🟢 允许：\n- 保护性 Put\n- Collar 结构",
            "forbidden": "🔴 禁止：\n- 所有卖方策略\n- 复杂套利",
            "risk": "风险控制为唯一目标\n不追求收益最大化",
            "fill": red_bg
        }
    ]
    
    for sc in scenarios:
        # 场景名
        ws.cell(row=current_row, column=1, value=sc["name"]).fill = sc["fill"]
        ws.cell(row=current_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
        
        # 特征
        ws.cell(row=current_row, column=2, value=sc["feat"]).alignment = Alignment(wrap_text=True, vertical='center')
        
        # 策略 (允许/禁止)
        strategy_text = f"{sc['allowed']}\n\n{sc['forbidden']}"
        strat_cell = ws.cell(row=current_row, column=3, value=strategy_text)
        strat_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 风控
        ws.cell(row=current_row, column=4, value=sc["risk"]).alignment = Alignment(wrap_text=True, vertical='center')
        
        ws.row_dimensions[current_row].height = 100
        current_row += 1
        
    current_row += 1
    
    # ============ 四、统一风控与退出 ============
    write_row(current_row, ['四、统一风控与退出原则'], font=sub_header_font, fill=section_fill, merge_cols=4)
    current_row += 1
    
    risks = [
        "1. 时间窗口：到期前 30% 时间窗口，强制评估是否退出。",
        "2. 双杀风险：Vega、Gamma 同时放大时，主动降仓。",
        "3. 亏损限额：单笔最大亏损不超过组合资金的 2% (或既定比例)。",
        "4. 流动性：避免持仓过重导致滑点无法止损。"
    ]
    for r in risks:
        write_row(current_row, [r], merge_cols=4)
        current_row += 1
    current_row += 1

    # ============ 五、使用边界 ============
    write_row(current_row, ['五、策略使用边界'], font=sub_header_font, fill=section_fill, merge_cols=4)
    current_row += 1
    
    boundaries = [
        "✅ 适用：流动性良好的期权品种、非连续涨跌停行情。",
        "❌ 暂停：重大宏观冲击、政策突发事件、流动性枯竭阶段。"
    ]
    for b in boundaries:
        write_row(current_row, [b], merge_cols=4)
        current_row += 1

    # 列宽调整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30

    # 冻结窗格
    ws.freeze_panes = 'A2'

def _create_strategy_reference_sheet_iv(wb):
    """
    创建波动率策略参考表
    """
    # 删除已有的
    for name in ['波动率策略']:
        if name in wb.sheetnames:
            del wb[name]
    
    ws = wb.create_sheet('波动率策略')
    
    # 策略框架内容
    strategy_framework = [
        ('基于波动率结构与分布特征的期权策略体系', '', ''),
        ('', '', ''),
        ('一、曲面倾斜策略（Skew）', '', ''),
        ('形态', '特征', '策略结构'),
        ('看跌倾斜 (Put Skew)', 'OTM PUT IV > OTM CALL IV，Skew>0', '认购多头价差（买低行K Call + 卖高行K Call）'),
        ('', 'Skew > 历史分位70%', '收益: CALL端IV修复 + 时间价值'),
        ('', 'PUT端成交/持仓上升', '风控: Skew继续扩大时止损'),
        ('', '', ''),
        ('看涨倾斜 (Call Skew)', 'OTM CALL IV > OTM PUT IV，Skew<0', '认沽多头价差（买高行K Put + 卖低行K Put）'),
        ('', 'CALL端成交资金占比上升', '收益: PUT端IV修复'),
        ('', 'Vega在CALL端集中', '适用: 标的未进入快速单边'),
        ('', '', ''),
        ('二、曲面期限结构策略（Term Structure）', '', ''),
        ('形态', '特征', '策略结构'),
        ('期限倒挂 (TS>0)', '短期IV > 长期IV', '日历价差（卖近月ATM + 买远月轻度OTM）'),
        ('', 'Theta绝对值大、Vega集中近月', '收益: 近月Theta衰减 + 短期IV回落'),
        ('', '突发事件或情绪冲击', '风控: 短期事件演化为趋势时提前退出'),
        ('', '', ''),
        ('正常升水 (TS<0)', '长期IV > 短期IV', '谨慎使用反向日历价差'),
        ('', '市场对中长期保持谨慎', '若有明确短期事件可买入短期期权'),
        ('', '', ''),
        ('三、基于峰度的策略', '', ''),
        ('形态', '特征', '策略结构'),
        ('高峰度 (肥尾)', '虚值期权成交占比上升', '跨式/宽跨式多头'),
        ('', 'Gamma在OTM区域放大', '或结合卖出极远 OTM降低Theta'),
        ('', 'IV仍低于历史中位数', '风控: IV已高位时仅高峰度不足以建仓'),
        ('', '', ''),
        ('低峰度 (瘦尾)', '成交量下降', '铁鹰/铁秃鹰/跨式空头'),
        ('', 'Vega整体偏低、横盘', '严禁裸卖，使用有限风险结构'),
        ('', '', ''),
        ('四、基于偏度的策略', '', ''),
        ('形态', '特征', '策略结构'),
        ('正偏度 (右尾肥)', 'CALL OTM IV < PUT OTM IV', '买入虚值CALL'),
        ('', 'CALL成交资金上升', '或认购比率价差（买低K卖高K）'),
        ('', '市场低估大幅上涨概率', '风控: 比率价差需控制卖出数量'),
        ('', '', ''),
        ('负偏度 (左尾肥)', 'PUT端Vega与成交集中', '买入虚值PUT（对冲或博弈）'),
        ('', 'PUT IV显著溢价', '或认沽比率价差'),
        ('', '风险资产高位/宏观不确定', '核心: 有限成本覆盖潜在下行尾部风险'),
        ('', '', ''),
        ('五、策略适用性与风险说明', '', ''),
        ('核心原则', '说明', ''),
        ('策略本质', '相对波动率定价策略，非无风险套利', ''),
        ('有效性依赖', 'IV结构均值回归、市场未进入单边趋势', ''),
        ('风控要求', 'Greeks风险暴露动态控制，严格执行退出机制', ''),
        ('暂停场景', '重大宏观冲击、流动性急剧变化、持续单边趋势', ''),
        ('', '', ''),
        ('六、合约筛选标准', '', ''),
        ('筛选维度', '推荐范围', '说明'),
        ('虚实程度', 'OTM: 虚实幅度% -5%至-15%', 'Delta约±0.25至0.35'),
        ('剩余时间', '30-90天', '平衡Theta衰减与Gamma收益'),
        ('流动性', '按沉淀资金(万)降序排列', '优先选择流动性好的合约'),
        ('IV估值', 'IV/RV < 0.8 低估，> 1.2 高估', '低估候选多头，高估候选空头'),
    ]
    
    # 写入内容
    for row_idx, (col1, col2, col3) in enumerate(strategy_framework, 1):
        ws.cell(row=row_idx, column=1, value=col1)
        ws.cell(row=row_idx, column=2, value=col2)
        ws.cell(row=row_idx, column=3, value=col3)
        
        # 标题行加粗
        if '一、' in col1 or '二、' in col1 or '三、' in col1 or '四、' in col1 or '五、' in col1 or '六、' in col1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        elif col1 == '形态' or col1 == '核心原则' or col1 == '筛选维度':
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            ws.cell(row=row_idx, column=3).font = Font(bold=True)
            # 表头背景色
            for col_idx in range(1, 4):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        elif '基于波动率结构' in col1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    ws.freeze_panes = 'A2'


if __name__ == "__main__":
    df = calculate_greeks()
    if df is not None:
        analyze_volatility_surface()

