"""
期货期权联动分析系统 by playbonze
基于期货和期权的资金流向、持仓变化、多空情绪进行联动分析
生成 wisecoin-货权联动.xlsx 多维度分析报告

【核心分析维度】:
1. 期货资金流向 - 沉淀资金、成交资金（保证金计算）
2. 期权PCR情绪 - Put/Call Ratio多空判断
3. 杠杆涨跌 - 实际涨跌幅 × 杠杆倍数
4. 货权联动信号 - 期货期权共振/背离检测
5. 最大痛点关联 - 期权Max Pain与期货价格距离
"""

import asyncio
import logging
import json
import pandas as pd
import numpy as np
import datetime
from datetime import datetime as dt
import sys
import os
import re
import traceback
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# 添加外部模块路径以支持 UnifiedLogger
sys.path.append(os.path.join(os.path.dirname(__file__), "wisecoin-catboost"))
from pb_quant_seektop_common import UnifiedLogger

# 设置统一日志
logger = UnifiedLogger.setup_logger_auto(__file__)

# 配置文件路径
SYMBOL_PARAMS_FILE = "wisecoin-symbol-params.json"
FUTURE_QUOTE_FILE = "wisecoin-期货行情.xlsx"
FUTURE_QUOTE_FILE_NO_OPT = "wisecoin-期货行情-无期权.xlsx"
OPTION_QUOTE_FILE = "wisecoin-期权行情.xlsx"
OPTION_RANKING_FILE = "wisecoin-期权排行.xlsx"
OUTPUT_FILE = "wisecoin-货权联动.xlsx"


def load_symbol_params():
    """加载品种参数配置"""
    try:
        if not os.path.exists(SYMBOL_PARAMS_FILE):
            logger.warning(f"配置文件不存在: {SYMBOL_PARAMS_FILE}")
            return {}
        
        with open(SYMBOL_PARAMS_FILE, 'r', encoding='utf-8') as f:
            params = json.load(f)
        
        # 构建扁平化的品种参数字典
        flat_params = {}
        for exchange, symbols in params.items():
            if exchange.startswith('_'):
                continue
            if not isinstance(symbols, dict):
                continue
            for symbol, config in symbols.items():
                if symbol.startswith('_'):
                    continue
                if not isinstance(config, dict):
                    continue
                full_key = f"{exchange}.{symbol}"
                flat_params[full_key] = config
        
        logger.info(f"加载 {len(flat_params)} 个品种参数配置")
        return flat_params
    except Exception as e:
        logger.error(f"加载品种参数失败: {e}")
        return {}


def extract_category_name(categories_field):
    """
    从 categories 字段提取板块名称
    
    Args:
        categories_field: 可能是字符串、列表或None
        
    Returns:
        str: 板块名称，如 '农副'、'软商'、'能化' 等
    """
    if pd.isna(categories_field):
        return '未分类'
    
    try:
        # 如果是字符串，尝试解析为JSON
        if isinstance(categories_field, str):
            import ast
            categories_field = ast.literal_eval(categories_field)
        
        # 如果是列表，取第一个元素
        if isinstance(categories_field, list) and len(categories_field) > 0:
            cat = categories_field[0]
            if isinstance(cat, dict) and 'name' in cat:
                return cat['name']
        
        # 如果是字典
        if isinstance(categories_field, dict) and 'name' in categories_field:
            return categories_field['name']
            
    except Exception as e:
        logger.debug(f"解析 categories 失败: {e}, 原始值: {categories_field}")
    
    return '未分类'


def get_margin_ratio(symbol_params, product_code):
    """获取品种保证金率"""
    # 尝试多种匹配方式
    for key, config in symbol_params.items():
        if product_code.upper() in key.upper():
            return config.get('margin_ratio', 0.10)
    # 默认保证金率10%
    return 0.10


def _classify_futures_trend_state(price_change_pct, oi_change_pct):
    """
    期货趋势状态机 - 将期货分为4类状态
    
    【状态分类】:
    1. 趋势强化: 价格↑ + 持仓↑ (多头主动建仓)
    2. 趋势衰减: 价格↑ + 持仓↓ (空头止损离场)
    3. 空头强化: 价格↓ + 持仓↑ (空头主动建仓)
    4. 空头衰减: 价格↓ + 持仓↓ (多头止损离场)
    
    返回: (状态名称, 趋势方向, 趋势强度0-3)
    """
    # 阈值定义
    price_threshold = 0.5  # 价格变化阈值 ±0.5%
    oi_threshold = 1.0     # 持仓变化阈值 ±1%
    
    # 价格方向
    if price_change_pct > price_threshold:
        price_dir = 'up'
    elif price_change_pct < -price_threshold:
        price_dir = 'down'
    else:
        price_dir = 'flat'
    
    # 持仓方向
    if oi_change_pct > oi_threshold:
        oi_dir = 'up'
    elif oi_change_pct < -oi_threshold:
        oi_dir = 'down'
    else:
        oi_dir = 'flat'
    
    # 状态分类
    if price_dir == 'up' and oi_dir == 'up':
        # 趋势强化: 多头主动进场
        strength = min(3, int((abs(price_change_pct) + abs(oi_change_pct)) / 2))
        return ('多头强化', '多', strength)
    
    elif price_dir == 'up' and oi_dir == 'down':
        # 趋势衰减: 空头平仓推升价格
        strength = min(2, int(abs(price_change_pct) / 1.5))
        return ('多头衰减', '多', strength)
    
    elif price_dir == 'down' and oi_dir == 'up':
        # 空头强化: 空头主动进场
        strength = min(3, int((abs(price_change_pct) + abs(oi_change_pct)) / 2))
        return ('空头强化', '空', strength)
    
    elif price_dir == 'down' and oi_dir == 'down':
        # 空头衰减: 多头平仓压低价格
        strength = min(2, int(abs(price_change_pct) / 1.5))
        return ('空头衰减', '空', strength)
    
    elif price_dir == 'flat':
        if oi_dir == 'up':
            return ('震荡蓄势', '震荡', 1)
        elif oi_dir == 'down':
            return ('震荡减仓', '震荡', 1)
        else:
            return ('盘整', '震荡', 0)
    
    else:
        # 持仓持平但价格有变化
        if price_dir == 'up':
            return ('弱多', '多', 1)
        else:
            return ('弱空', '空', 1)


def _classify_option_fund_structure(pcr, call_oi_change, put_oi_change):
    """
    期权资金结构分类
    
    【分类标准】:
    - 看多: PCR < 0.7 或 CALL单边增仓明显
    - 看空: PCR > 1.3 或 PUT单边增仓明显
    - 波动率: PCR接近1且双向增仓
    
    返回: (结构类型, 方向一致性分数0-3)
    """
    # PCR判断
    if pcr < 0.5:
        pcr_signal = '极度看多'
        pcr_score = 3
    elif pcr < 0.7:
        pcr_signal = '看多'
        pcr_score = 2
    elif pcr < 0.9:
        pcr_signal = '偏多'
        pcr_score = 1
    elif pcr <= 1.1:
        pcr_signal = '中性'
        pcr_score = 0
    elif pcr <= 1.3:
        pcr_signal = '偏空'
        pcr_score = -1
    elif pcr <= 1.5:
        pcr_signal = '看空'
        pcr_score = -2
    else:
        pcr_signal = '极度看空'
        pcr_score = -3
    
    # 增仓方向判断
    both_increasing = call_oi_change > 0 and put_oi_change > 0
    
    if both_increasing and 0.8 <= pcr <= 1.2:
        return ('波动率', 0)
    elif pcr_score >= 2:
        return ('看多', pcr_score)
    elif pcr_score <= -2:
        return ('看空', pcr_score)
    elif pcr_score > 0:
        return ('偏多', pcr_score)
    elif pcr_score < 0:
        return ('偏空', pcr_score)
    else:
        return ('中性', 0)


def _calculate_resonance_score(futures_trend_strength, futures_direction,
                                option_structure, option_score,
                                volatility_match=True):
    """
    期货-期权共振评分系统
    
    【评分维度】:
    1. 期货趋势强度 (0-3分)
    2. 期权方向一致性 (0-3分)
    3. 波动率配合度 (0-2分)
    
    【输出】:
    - 总分 0-8 分
    - ⭐⭐⭐⭐ (7-8): 强共振 - 重点跟踪
    - ⭐⭐⭐ (5-6): 共振
    - ⭐⭐ (3-4): 中性
    - ⭐ (1-2): 弱相关
    - ⚠️ (0或负): 明显背离 - 风险提示
    """
    # 1. 期货趋势强度 (0-3)
    trend_score = min(3, max(0, futures_trend_strength))
    
    # 2. 方向一致性 (0-3)
    direction_match = 0
    if futures_direction == '多':
        if option_structure in ['看多', '偏多']:
            direction_match = abs(option_score)  # 0-3
        elif option_structure in ['看空', '偏空']:
            direction_match = -abs(option_score)  # 背离
    elif futures_direction == '空':
        if option_structure in ['看空', '偏空']:
            direction_match = abs(option_score)
        elif option_structure in ['看多', '偏多']:
            direction_match = -abs(option_score)  # 背离
    elif futures_direction == '震荡':
        if option_structure == '波动率':
            direction_match = 2  # 震荡 + 波动率交易 = 匹配
    
    # 3. 波动率配合度 (0-2)
    vol_score = 2 if volatility_match else 0
    
    # 总分
    total_score = trend_score + direction_match + vol_score
    
    # 等级和标签
    if total_score >= 7:
        grade = '⭐⭐⭐⭐'
        label = '强共振'
    elif total_score >= 5:
        grade = '⭐⭐⭐'
        label = '共振'
    elif total_score >= 3:
        grade = '⭐⭐'
        label = '中性'
    elif total_score >= 1:
        grade = '⭐'
        label = '弱相关'
    else:
        grade = '⭐'
        label = '背离'
    
    return total_score, grade, label


def _calculate_linkage_strength_model(futures_row, opt_data, pcr):
    """
    联动强度模型 (Linkage Strength Model)
    
    综合价格、资金、情绪三个维度计算联动强度总分 (0-100)
    
    【模型权重】:
    1. 价格强度 (30%): 杠杆涨跌幅强度 + 趋势状态得分
    2. 资金强度 (40%): 期货资金流向强度 + 期权资金流向一致性
    3. 情绪强度 (30%): PCR偏离度 + 情绪得分
    """
    # ============ 1. 价格强度 (30分) ============
    # 杠杆涨跌幅得分 (0-20分)
    leverage_change = abs(futures_row.get('杠杆涨跌%', 0) or 0)
    price_score_raw = min(20, leverage_change * 2)
    
    # 趋势状态得分 (0-10分)
    trend_strength = futures_row.get('趋势强度', 0) or 0
    futures_dir = futures_row.get('趋势方向', '震荡')
    trend_score = min(10, trend_strength * 3)
    
    price_total = min(30, price_score_raw + trend_score)
    
    # ============ 2. 资金强度 (40分) ============
    # 期货资金强度 (0-20分)
    fut_signal = abs(futures_row.get('流向信号', 0) or 0)
    fut_fund = abs(futures_row.get('沉淀资金(亿)', 0) or 0)
    # 资金规模加成: 每10亿加1分，上限5分
    fund_bonus = min(5, fut_fund / 10)
    # 信号强度: 2分=15分, 1分=10分
    signal_score = 15 if fut_signal >= 2 else (10 if fut_signal >= 1 else 0)
    
    futures_fund_score = min(20, signal_score + fund_bonus)
    
    # 期权资金一致性 (0-20分)
    opt_fund_direction = 0
    call_oi_chg = opt_data.get('CALL持仓变化', 0)
    put_oi_chg = opt_data.get('PUT持仓变化', 0)
    
    if call_oi_chg > put_oi_chg:
        opt_fund_direction = 1  # 偏多
    elif put_oi_chg > call_oi_chg:
        opt_fund_direction = -1 # 偏空
        
    # 判断方向是否一致
    consistency_score = 0
    if futures_dir == '多' and opt_fund_direction == 1:
        consistency_score = 20
    elif futures_dir == '空' and opt_fund_direction == -1:
        consistency_score = 20
    elif futures_dir == '震荡' and call_oi_chg > 0 and put_oi_chg > 0:
        consistency_score = 15 # 震荡市双向增仓视为资金活跃匹配
        
    capital_total = min(40, futures_fund_score + consistency_score)
    
    # ============ 3. 情绪强度 (30分) ============
    # PCR偏离度得分 (0-15分)
    # 偏离中枢1.0越远，情绪越强烈
    pcr_deviation = abs(pcr - 1.0)
    pcr_score = min(15, pcr_deviation * 30)
    
    # 情绪倾向一致性 (0-15分)
    sentiment_score = 0
    opt_sentiment = opt_data.get('情绪倾向', 0) # -100 to 100
    
    if futures_dir == '多':
        if opt_sentiment > 0:
            sentiment_score = min(15, opt_sentiment / 100 * 15)
        elif pcr < 0.7: # 极度看多
            sentiment_score = 15
            
    elif futures_dir == '空':
        if opt_sentiment < 0:
            sentiment_score = min(15, abs(opt_sentiment) / 100 * 15)
        elif pcr > 1.3: # 极度看空
            sentiment_score = 15
            
    sentiment_total = min(30, pcr_score + sentiment_score)
    
    # ============ 总分汇总 ============
    total_score = price_total + capital_total + sentiment_total
    
    return {
        '联动总分': round(total_score, 1),
        '价格评分': round(price_total, 1),
        '资金评分': round(capital_total, 1),
        '情绪评分': round(sentiment_total, 1)
    }


def _determine_linkage_state(futures_state, futures_dir, option_structure, pcr, vol_sentiment=''):
    """
    四层联动判断框架
    
    【第一层】期货趋势（方向）: 多 / 空 / 震荡
    【第二层】期权资金结构（预期）: 看多 / 看空 / 波动率
    【第三层】期权波动率情绪: 狂热 / 恐慌 / 筑底 / 冲高
    【第四层】一致性或背离
    
    返回: (联动状态标签, 市场解读)
    """
    # 基础解读逻辑
    base_state = ''
    base_interpretation = ''
    
    # 1. 期货多头情形
    if futures_dir == '多':
        if option_structure in ['看多', '偏多']:
            base_state, base_interpretation = ('趋势确认', '期货上涨+CALL主导，趋势延续概率高')
            # 叠加情绪
            if '狂热' in vol_sentiment:
                return ('加速赶顶', f"{base_interpretation} | 波动率狂热，小心过热回落")
        elif option_structure in ['看空', '偏空']:
            base_state, base_interpretation = ('顶部警惕', '期货上涨+PUT增仓，注意对冲或反转信号')
            if '恐慌' in vol_sentiment:
                return ('极度背离', f"{base_interpretation} | 波动率恐慌，反转风险极大")
                
    # 2. 期货空头情形
    elif futures_dir == '空':
        if option_structure in ['看空', '偏空']:
            base_state, base_interpretation = ('空头确认', '期货下跌+PUT主导，空头趋势延续')
            if '恐慌' in vol_sentiment:
                return ('加速赶底', f"{base_interpretation} | 波动率恐慌，小心超跌反弹")
        elif option_structure in ['看多', '偏多']:
            base_state, base_interpretation = ('抄底信号', '期货下跌+CALL增仓，可能有抄底资金或错配')
            if '狂热' in vol_sentiment:
                return ('极度背离', f"{base_interpretation} | 波动率狂热，可能是期权诱多")

    # 3. 期货震荡情形
    elif futures_dir == '震荡':
        if option_structure == '波动率':
            base_state, base_interpretation = ('波动率机会', '期货震荡+双向期权放量，适合波动率策略')
        elif option_structure in ['看多', '偏多']:
            base_state, base_interpretation = ('蓄势待涨', '期货震荡+期权看多，关注突破方向')
        elif option_structure in ['看空', '偏空']:
            base_state, base_interpretation = ('蓄势待跌', '期货震荡+期权看空，关注破位风险')
        else:
            base_state, base_interpretation = ('观望', '期货和期权均无明确方向')
            
        # 叠加情绪
        if '筑底' in vol_sentiment:
            return ('震荡筑底', f"{base_interpretation} | 波动率显示筑底特征")
        elif '冲高' in vol_sentiment:
            return ('震荡冲高', f"{base_interpretation} | 波动率显示冲高特征")
            
    # 如果已有定义则返回
    if base_state:
        # 如果有特殊情绪叠加
        if '恐慌' in vol_sentiment and '确认' in base_state:
            return ('恐慌加速', f"{base_interpretation} | 市场陷入恐慌")
        if '狂热' in vol_sentiment and '确认' in base_state:
            return ('狂热过热', f"{base_interpretation} | 市场情绪狂热")
            
        return (base_state, base_interpretation)
    
    return ('信号不明', '需要进一步观察')


def _suggest_strategy(linkage_state, futures_state, option_structure, resonance_label):
    """
    策略导向建议
    
    基于联动状态和共振评分，给出具体的策略建议
    返回: (适合策略, 不适合策略)
    """
    suitable = []
    unsuitable = []
    
    if linkage_state == '趋势确认':
        suitable = ['趋势跟随', '买入CALL', '卖出PUT']
        unsuitable = ['裸卖CALL', '逆势做空']
    
    elif linkage_state == '空头确认':
        suitable = ['趋势跟随', '买入PUT', '卖出CALL']
        unsuitable = ['裸卖PUT', '逆势做多']
    
    elif linkage_state == '顶部警惕':
        suitable = ['保护性PUT', '领口策略', '减仓观望']
        unsuitable = ['单边追价', '裸卖PUT', '激进加仓']
    
    elif linkage_state == '抄底信号':
        suitable = ['分批建仓', '卖出PUT', '牛市价差']
        unsuitable = ['单边追空', '裸卖CALL']
    
    elif linkage_state == '波动率机会':
        suitable = ['买跨式', '买宽跨式', '比率价差']
        unsuitable = ['裸卖跨式', '单边方向策略']
    
    elif linkage_state == '蓄势待涨':
        suitable = ['轻仓试多', '牛市价差', '卖出PUT']
        unsuitable = ['重仓做空', '裸卖CALL']
    
    elif linkage_state == '蓄势待跌':
        suitable = ['轻仓试空', '熊市价差', '卖出CALL']
        unsuitable = ['重仓做多', '裸卖PUT']
    
    elif resonance_label == '背离':
        suitable = ['观望', '对冲', '减仓']
        unsuitable = ['单边追价', '裸卖期权', '加杠杆']
    
    else:
        suitable = ['观望', '小仓位试探']
        unsuitable = ['重仓单边']
    
    return ' / '.join(suitable), ' / '.join(unsuitable)


def analyze_futures_options_correlation():
    """
    期货期权联动分析主函数
    
    【分析逻辑】:
    1. 加载期货行情和期权排行数据
    2. 计算期货的实际资金（保证金口径）和杠杆涨跌
    3. 读取期权的PCR、最大痛点等指标
    4. 进行货权联动信号检测
    5. 生成多分页分析报告
    """
    logger.info("🚀 开始期货期权联动分析...")
    
    # 加载品种参数
    symbol_params = load_symbol_params()
    if not symbol_params:
        logger.warning("未加载到品种参数，将使用默认保证金率10%")
    
    # ============ 1. 加载期货数据 ============
    futures_df_list = []
    
    # 1.1 加载有期权的期货行情
    if os.path.exists(FUTURE_QUOTE_FILE):
        try:
            future_xls = pd.ExcelFile(FUTURE_QUOTE_FILE)
            if 'Summary' in future_xls.sheet_names:
                df1 = pd.read_excel(future_xls, sheet_name='Summary')
                futures_df_list.append(df1)
                logger.info(f"加载期货行情(有期权): {len(df1)} 个合约")
            else:
                logger.warning(f"{FUTURE_QUOTE_FILE} 中未找到 Summary sheet")
        except Exception as e:
            logger.error(f"加载期货行情(有期权)失败: {e}")
            
    # 1.2 加载无期权的期货行情
    if os.path.exists(FUTURE_QUOTE_FILE_NO_OPT):
        try:
            future_xls_no_opt = pd.ExcelFile(FUTURE_QUOTE_FILE_NO_OPT)
            # 优先使用 Summary sheet，否则使用第一个 sheet
            sheet_name = 'Summary' if 'Summary' in future_xls_no_opt.sheet_names else future_xls_no_opt.sheet_names[0]
            df2 = pd.read_excel(future_xls_no_opt, sheet_name=sheet_name)
            futures_df_list.append(df2)
            logger.info(f"加载期货行情(无期权): {len(df2)} 个合约")
        except Exception as e:
            logger.error(f"加载期货行情(无期权)失败: {e}")
    
    if not futures_df_list:
        logger.warning("未找到有效期货行情数据")
        return
        
    futures_df = pd.concat(futures_df_list, ignore_index=True)
    
    # ============ 2. 加载期权排行数据 ============
    option_analysis = None
    if os.path.exists(OPTION_RANKING_FILE):
        try:
            option_xls = pd.ExcelFile(OPTION_RANKING_FILE)
            if '期权排行' in option_xls.sheet_names:
                option_analysis = pd.read_excel(option_xls, sheet_name='期权排行')
                logger.info(f"加载期权排行: {len(option_analysis)} 个标的")
        except Exception as e:
            logger.warning(f"加载期权排行失败: {e}")
    
    # ============ 3. 计算期货分析指标 ============
    logger.info("计算期货分析指标...")
    
    futures_analysis = []
    for _, row in futures_df.iterrows():
        try:
            symbol = row.get('instrument_id') or row.get('symbol', '')
            if not symbol or pd.isna(symbol):
                continue
            
            # 提取品种代码 (兼容 03wisecoin-options-iv.py 逻辑)
            product_code = ''
            if '.' in str(symbol):
                parts = str(symbol).split('.')
                if len(parts) >= 2:
                    # 获取第二段的字母部分，如 SHFE.ag2602 -> AG
                    code_match = re.match(r'^([a-zA-Z]+)', parts[1])
                    if code_match:
                        product_code = code_match.group(1).upper()
                    else:
                        # 兜底处理类似 SSE.000812 或没有字母的情况
                        # 尝试从第一段提取或保持原始
                        product_code = re.sub(r'[^a-zA-Z]', '', parts[1]).upper()
            
            if not product_code:
                product_code = re.sub(r'[^a-zA-Z]', '', str(symbol)).upper()
            
            # 获取保证金率
            margin_ratio = get_margin_ratio(symbol_params, product_code)
            leverage = 1.0 / margin_ratio if margin_ratio > 0 else 10.0  # 杠杆倍数
            
            # 价格数据
            last_price = row.get('last_price', 0) or row.get('settlement', 0) or 0   # row.get('last_price', 0) or 
            pre_close = row.get('pre_close', 0) or row.get('pre_settlement', 0) or 0
            multiplier = row.get('volume_multiple', 1) or 1
            
            if pd.isna(last_price) or last_price <= 0:
                last_price = pre_close
            if pd.isna(pre_close) or pre_close <= 0:
                continue
            
            # 持仓量和成交量
            open_interest = row.get('open_interest', 0) or 0
            volume = row.get('volume', 0) or 0
            pre_oi = row.get('pre_open_interest', 0) or 0
            
            # ============ 核心计算 ============
            # 1. 实际沉淀资金（保证金口径，亿元）
            # 沉淀资金 = 持仓量 × 价格 × 乘数 × 保证金率
            chendian = (open_interest * last_price * multiplier * margin_ratio) / 1e8
            
            # 2. 实际成交资金（保证金口径，亿元）
            # 成交资金 = 成交量 × 价格 × 乘数 × 保证金率
            chengjiao = (volume * last_price * multiplier * margin_ratio) / 1e8
            
            # 3. 杠杆涨跌（%）
            # 杠杆涨跌 = (收盘价 - 昨收价) / 昨收价 × 100 × 杠杆倍数
            price_change_pct = ((last_price - pre_close) / pre_close) * 100
            leverage_change = price_change_pct * leverage
            
            # 4. 持仓变化
            oi_change = open_interest - pre_oi if pre_oi > 0 else 0
            oi_change_pct = (oi_change / pre_oi * 100) if pre_oi > 0 else 0
            
            # 5. 换手率
            turnover = (volume / open_interest * 100) if open_interest > 0 else 0
            
            # 6. 资金流向判断
            if oi_change > 0 and price_change_pct > 0:
                flow_direction = '增仓上涨'
                flow_signal = 2  # 强多
            elif oi_change > 0 and price_change_pct < 0:
                flow_direction = '增仓下跌'
                flow_signal = -2  # 强空
            elif oi_change < 0 and price_change_pct > 0:
                flow_direction = '减仓上涨'
                flow_signal = 1  # 弱多（多头平仓）
            elif oi_change < 0 and price_change_pct < 0:
                flow_direction = '减仓下跌'
                flow_signal = -1  # 弱空（空头平仓）
            else:
                flow_direction = '持平'
                flow_signal = 0
            
            # 7. 期货趋势状态机分类
            trend_state, trend_dir, trend_strength = _classify_futures_trend_state(
                price_change_pct, oi_change_pct
            )
            
            futures_analysis.append({
                '合约': symbol,
                '品种代码': product_code,
                '现价': round(last_price, 2),
                '昨收': round(pre_close, 2),
                '涨跌%': round(price_change_pct, 2),
                '杠杆涨跌%': round(leverage_change, 2),
                '杠杆倍数': round(leverage, 1),
                '保证金率%': round(margin_ratio * 100, 1),
                '持仓量': int(open_interest),
                '成交量': int(volume),
                '持仓变化': int(oi_change),
                '持仓变化%': round(oi_change_pct, 2),
                '沉淀资金(亿)': round(chendian, 4),
                '成交资金(亿)': round(chengjiao, 4),
                '换手率%': round(turnover, 2),
                '资金流向': flow_direction,
                '流向信号': flow_signal,
                '合约乘数': int(multiplier),
                '趋势状态': trend_state,
                '趋势方向': trend_dir,
                '趋势强度': trend_strength,
            })
        except Exception as e:
            continue
    
    if not futures_analysis:
        logger.warning("未生成有效期货分析数据")
        return
    
    futures_df_analysis = pd.DataFrame(futures_analysis)
    logger.info(f"完成 {len(futures_df_analysis)} 个期货合约分析")
    
    # ============ 4. 货权联动分析 ============
    logger.info("执行货权联动分析...")
    
    correlation_analysis = []
    
    # 加载波动率情绪数据
    OPTION_REFERENCE_FILE = "wisecoin-期权参考.xlsx"
    vol_sentiment_map = {}
    
    # 指数期权品种代码映射（这些是股指期权，不是商品期货期权）
    # 标的合约格式: SSE.000852, SSE.000300, SSE.000016
    INDEX_OPTION_CODES = {
        'MO': '000852',  # 中证1000 -> SSE.000852
        'IO': '000300',  # 沪深300  -> SSE.000300
        'HO': '000016',  # 上证50   -> SSE.000016
    }
    
    if os.path.exists(OPTION_REFERENCE_FILE):
        try:
            # 尝试读取波动率曲面分页
            vol_df = pd.read_excel(OPTION_REFERENCE_FILE, sheet_name='波动率曲面')
            if '品种代码' in vol_df.columns and '市场情绪' in vol_df.columns:
                for _, v_row in vol_df.iterrows():
                    p_code = str(v_row['品种代码']).strip().upper()
                    sent = str(v_row['市场情绪'])
                    if p_code and sent:
                        # 对于指数期权，使用指数代码作为键
                        if p_code in INDEX_OPTION_CODES:
                            index_code = INDEX_OPTION_CODES[p_code]
                            vol_sentiment_map[index_code] = sent
                            logger.debug(f"指数期权映射: {p_code} -> {index_code}, 情绪: {sent}")
                        else:
                            # 普通商品期权，使用品种代码
                            vol_sentiment_map[p_code] = sent
            logger.info(f"加载波动率情绪: {len(vol_sentiment_map)} 个品种（含指数期权映射）")
        except Exception as e:
            logger.warning(f"加载波动率情绪失败: {e}")
    
    if option_analysis is not None and not option_analysis.empty:
        # 建立期权数据查找表（按标的合约）
        option_lookup = {}
        for _, opt_row in option_analysis.iterrows():
            underlying = opt_row.get('标的合约', '')
            if underlying:
                option_lookup[underlying] = opt_row
        
        for _, fut_row in futures_df_analysis.iterrows():
            symbol = fut_row['合约']
            # 注意：这里的品种代码已经是大写纯字母，如 AG
            p_code = fut_row['品种代码']
            
            # 查找对应期权数据
            opt_data = option_lookup.get(symbol, None)
            
            if opt_data is not None:
                # 有对应期权数据，进行联动分析
                pcr = opt_data.get('PCR(持仓)', 1.0)
                max_pain = opt_data.get('最大痛点', 0)
                max_pain_distance = opt_data.get('痛点距离%', 0)
                option_sentiment_score = opt_data.get('情绪倾向', 0)
                option_chendian = opt_data.get('沉淀资金(亿)', 0)
                call_oi_change = opt_data.get('CALL持仓变化', 0)
                put_oi_change = opt_data.get('PUT持仓变化', 0)
                
                # 获取波动率市场情绪（支持指数期权特殊处理）
                # 对于指数期权（如 SSE.000852），直接使用指数代码查找
                # 对于普通商品期权（如 SHFE.ag2602），使用品种代码（AG）查找
                vol_sentiment_text = ''
                if '.' in symbol:
                    exchange, contract = symbol.split('.', 1)
                    # 检查是否为指数期权（交易所是SSE或CFFEX，且合约代码是纯数字）
                    if exchange in ['SSE', 'CFFEX'] and contract[:6].isdigit():
                        # 指数期权: SSE.000852, SSE.000300, SSE.000016 等
                        index_code = contract[:6]  # 提取前6位数字
                        vol_sentiment_text = vol_sentiment_map.get(index_code, '')
                        if vol_sentiment_text:
                            logger.debug(f"指数期权匹配情绪: {symbol} -> {index_code} -> {vol_sentiment_text}")
                    else:
                        # 商品期权: 使用品种代码
                        vol_sentiment_text = vol_sentiment_map.get(p_code, '')
                else:
                    # 无交易所前缀，使用品种代码
                    vol_sentiment_text = vol_sentiment_map.get(p_code, '')
                
                # 期货趋势状态
                futures_state = fut_row['趋势状态']
                futures_dir = fut_row['趋势方向']
                futures_strength = fut_row['趋势强度']
                
                # 期权资金结构分类
                option_structure, option_dir_score = _classify_option_fund_structure(
                    pcr, call_oi_change, put_oi_change
                )
                
                # 三层联动判断 (增加波动率情绪判断)
                linkage_state, market_interpretation = _determine_linkage_state(
                    futures_state, futures_dir, option_structure, pcr, vol_sentiment_text
                )
                
                # 完善市场解读 (已由 _determine_linkage_state 处理，此处可根据需要补充)
                
                # 共振评分
                resonance_score, resonance_grade, resonance_label = _calculate_resonance_score(
                    futures_strength, futures_dir,
                    option_structure, option_dir_score,
                    volatility_match=(option_structure == '波动率' and futures_dir == '震荡')
                )
                
                # 联动强度模型评分
                strength_scores = _calculate_linkage_strength_model(fut_row, opt_data, pcr)
                
                # 策略建议
                suitable_strategy, unsuitable_strategy = _suggest_strategy(
                    linkage_state, futures_state, option_structure, resonance_label
                )
                
                # 期货信号 (兼容旧逻辑)
                fut_signal = fut_row['流向信号']
                opt_signal = -option_sentiment_score / 100 * 2
                
                # 联动类型判断 (兼容旧逻辑)
                if fut_signal * opt_signal > 0:
                    correlation_type = '共振'
                    correlation_strength = abs(fut_signal + opt_signal)
                elif fut_signal * opt_signal < 0:
                    correlation_type = '背离'
                    correlation_strength = abs(fut_signal - opt_signal)
                else:
                    correlation_type = '中性'
                    correlation_strength = 0
                
                # 综合方向判断
                combined_signal = fut_signal + opt_signal
                if combined_signal >= 2:
                    combined_direction = '强烈看多'
                elif combined_signal >= 1:
                    combined_direction = '偏多'
                elif combined_signal <= -2:
                    combined_direction = '强烈看空'
                elif combined_signal <= -1:
                    combined_direction = '偏空'
                else:
                    combined_direction = '中性震荡'
                
                correlation_analysis.append({
                    '标的合约': symbol,
                    '期货现价': fut_row['现价'],
                    '杠杆涨跌%': fut_row['杠杆涨跌%'],
                    '期货状态': futures_state,
                    '期货方向': futures_dir,
                    '期货流向': fut_row['资金流向'],
                    '期货沉淀(亿)': fut_row['沉淀资金(亿)'],
                    '期权结构': option_structure,
                    '期权情绪': vol_sentiment_text if vol_sentiment_text else f"{option_sentiment_score}",
                    '期权PCR': round(pcr, 4),
                    # '原情绪分': option_sentiment, # 保留参考但不在前台显示
                    '期权沉淀(亿)': round(option_chendian, 4),
                    '联动状态': linkage_state,
                    '市场解读': market_interpretation,
                    '共振评分': resonance_score,
                    '联动总分': strength_scores['联动总分'],
                    '价格评分': strength_scores['价格评分'],
                    '资金评分': strength_scores['资金评分'],
                    '情绪评分': strength_scores['情绪评分'],
                    '共振等级': resonance_grade,
                    '共振标签': resonance_label,
                    '最大痛点': round(max_pain, 2),
                    '痛点距离%': round(max_pain_distance, 2),
                    '适合策略': suitable_strategy,
                    '不适合策略': unsuitable_strategy,
                })
    
    # ============ 5. 生成排行表 ============
    logger.info("生成多维度排行表...")
    
    ranking_sheets = {}
    
    # 3.1 期货市场
    market_summary = _generate_futures_market_summary(futures_df_analysis, correlation_analysis)
    ranking_sheets['期货市场'] = market_summary
    
    # 3.2 货权联动
    if correlation_analysis:
        corr_df = pd.DataFrame(correlation_analysis)
        corr_df['沉淀资金合计(亿)'] = corr_df['期货沉淀(亿)'] + corr_df['期权沉淀(亿)']
        corr_df = corr_df.sort_values('沉淀资金合计(亿)', ascending=False).reset_index(drop=True)
        corr_df.insert(0, '排名', range(1, len(corr_df) + 1))
        ranking_sheets['货权联动'] = corr_df
        
        # ============ 新增：按品种维度分析 ============
        logger.info("生成期货品种维度分析...")
        product_analysis_sheets = _generate_product_analysis(futures_df_analysis, corr_df, futures_df)
        ranking_sheets.update(product_analysis_sheets)
        
        # ============ 新增：按板块维度分析 ============
        logger.info("生成期货板块维度分析...")
        sector_analysis_sheets = _generate_sector_analysis(futures_df_analysis, corr_df, futures_df)
        ranking_sheets.update(sector_analysis_sheets)
        
    
    # 3.5 期货排行
    fut_comprehensive = futures_df_analysis.copy()
    # 综合评分 = 资金规模 + 活跃度 + 信号强度
    fut_comprehensive['综合评分'] = (
        fut_comprehensive['沉淀资金(亿)'].rank(pct=True) * 30 +
        fut_comprehensive['成交资金(亿)'].rank(pct=True) * 30 +
        fut_comprehensive['流向信号'].abs().rank(pct=True) * 20 +
        fut_comprehensive['杠杆涨跌%'].abs().rank(pct=True) * 20
    )
    fut_comprehensive = fut_comprehensive.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
    fut_comprehensive.insert(0, '排名', range(1, len(fut_comprehensive) + 1))
    ranking_sheets['期货排行'] = fut_comprehensive
    
    # 3.6 期货涨跌
    leverage_df = futures_df_analysis[['合约', '品种代码', '现价', '涨跌%', '杠杆涨跌%', 
                                         '杠杆倍数', '保证金率%', '资金流向']].copy()
    leverage_df = leverage_df.sort_values('杠杆涨跌%', ascending=False).reset_index(drop=True)
    leverage_df.insert(0, '排名', range(1, len(leverage_df) + 1))
    ranking_sheets['期货涨跌'] = leverage_df
    
    # 3.7 期货看多
    bullish_df = futures_df_analysis[futures_df_analysis['流向信号'] > 0].copy()
    bullish_df = bullish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
    if not bullish_df.empty:
        bullish_df.insert(0, '排名', range(1, len(bullish_df) + 1))
        bullish_df = bullish_df[['排名', '合约', '品种代码', '现价', '杠杆涨跌%', 
                                  '持仓变化', '资金流向', '沉淀资金(亿)', '流向信号']]
        ranking_sheets['期货看多'] = bullish_df
    
    # 3.8 期货看空
    bearish_df = futures_df_analysis[futures_df_analysis['流向信号'] < 0].copy()
    bearish_df = bearish_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
    if not bearish_df.empty:
        bearish_df.insert(0, '排名', range(1, len(bearish_df) + 1))
        bearish_df = bearish_df[['排名', '合约', '品种代码', '现价', '杠杆涨跌%', 
                                  '持仓变化', '资金流向', '沉淀资金(亿)', '流向信号']]
        ranking_sheets['期货看空'] = bearish_df

    # 3.9 期货资金
    capital_df = futures_df_analysis[['合约', '品种代码', '现价', '沉淀资金(亿)', '成交资金(亿)', 
                                        '持仓量', '持仓变化', '资金流向']].copy()
    capital_df['资金合计(亿)'] = capital_df['沉淀资金(亿)'] + capital_df['成交资金(亿)']
    capital_df = capital_df.sort_values('资金合计(亿)', ascending=False).reset_index(drop=True)
    capital_df.insert(0, '排名', range(1, len(capital_df) + 1))
    ranking_sheets['期货资金'] = capital_df
    
    # ============ 6. 保存到 Excel ============
    logger.info(f"保存分析结果到 {OUTPUT_FILE}...")
    _save_correlation_excel(OUTPUT_FILE, ranking_sheets)
    
    logger.info("="*60)
    logger.info("期货期权联动分析完成")
    logger.info(f"📊 分析报告: {OUTPUT_FILE}")
    logger.info(f"📈 包含 {len(ranking_sheets)} 个分析维度")
    logger.info(f"📉 期货合约: {len(futures_df_analysis)} 个")
    if correlation_analysis:
        logger.info(f"🔗 货权联动: {len(correlation_analysis)} 个")
    logger.info("="*60)


def _generate_product_analysis(futures_df_analysis, corr_df, futures_df_raw):
    """
    按期货品种（product_id）维度生成汇总分析（单个分页）
    
    【分析内容】:
    1. 所有品种资金排名（按沉淀资金、成交资金降序）
    2. 品种合约数量统计
    3. 品种平均杠杆涨跌
    4. 品种多空情绪
    5. 品种货权联动情况（如有期权）
    
    Returns:
        dict: {'期货品种': DataFrame}
    """
    product_sheets = {}
    
    try:
        # 从原始期货数据中提取 product_id
        if 'product_id' not in futures_df_analysis.columns:
            # 尝试从 futures_df_raw 合并 product_id
            if 'product_id' in futures_df_raw.columns:
                # 建立合约到 product_id 的映射
                product_map = {}
                for _, row in futures_df_raw.iterrows():
                    symbol = row.get('instrument_id') or row.get('symbol', '')
                    product_id = row.get('product_id', '')
                    if symbol and product_id:
                        product_map[symbol] = product_id
                
                # 添加 product_id 列
                futures_df_analysis['product_id'] = futures_df_analysis['合约'].map(product_map)
            else:
                logger.warning("未找到 product_id 字段，跳过品种维度分析")
                return product_sheets
        
        # 按 product_id 分组统计
        product_data = []
        grouped = futures_df_analysis.groupby('product_id')
        
        for product_id, group in grouped:
            if pd.isna(product_id) or product_id == '':
                continue
            
            # 品种汇总统计
            product_summary = {
                '品种代码': product_id,
                '合约数量': len(group),
                '沉淀资金(亿)': round(group['沉淀资金(亿)'].sum(), 4),
                '成交资金(亿)': round(group['成交资金(亿)'].sum(), 4),
                '平均杠杆涨跌%': round(group['杠杆涨跌%'].mean(), 2),
                '最大杠杆涨跌%': round(group['杠杆涨跌%'].max(), 2),
                '最小杠杆涨跌%': round(group['杠杆涨跌%'].min(), 2),
                '总持仓量': int(group['持仓量'].sum()),
                '总成交量': int(group['成交量'].sum()),
                '看多合约数': len(group[group['流向信号'] > 0]),
                '看空合约数': len(group[group['流向信号'] < 0]),
                '中性合约数': len(group[group['流向信号'] == 0]),
            }
            
            # 多空情绪判断
            bullish_ratio = product_summary['看多合约数'] / product_summary['合约数量']
            if bullish_ratio >= 0.6:
                product_summary['品种情绪'] = '偏多'
            elif bullish_ratio <= 0.4:
                product_summary['品种情绪'] = '偏空'
            else:
                product_summary['品种情绪'] = '中性'
            
            # 检查是否有期权联动数据
            if '标的合约' in corr_df.columns:
                # 尝试匹配货权联动数据
                product_corr = corr_df[corr_df['标的合约'].str.contains(product_id, na=False, case=False)]
                if not product_corr.empty:
                    product_summary['有期权联动'] = '是'
                    product_summary['期权PCR均值'] = round(product_corr['期权PCR'].mean(), 4)
                    product_summary['期权痛点均值'] = round(product_corr['最大痛点'].mean(), 2)
                    product_summary['共振评分均值'] = round(product_corr['共振评分'].mean(), 2)
                else:
                    product_summary['有期权联动'] = '否'
            
            product_data.append(product_summary)
        
        # 创建品种汇总表
        if product_data:
            product_df = pd.DataFrame(product_data)
            # 按沉淀资金(亿)、成交资金(亿)降序排列
            product_df = product_df.sort_values(
                by=['沉淀资金(亿)', '成交资金(亿)'], 
                ascending=False
            ).reset_index(drop=True)
            product_df.insert(0, '排名', range(1, len(product_df) + 1))
            product_sheets['期货品种'] = product_df
            logger.info(f"生成品种维度汇总: {len(product_df)} 个品种")
        
    except Exception as e:
        logger.error(f"生成品种维度分析失败: {e}")
        logger.error(traceback.format_exc())
    
    return product_sheets


def _generate_sector_analysis(futures_df_analysis, corr_df, futures_df_raw):
    """
    按期货板块（categories 的 name）维度生成汇总分析（单个分页）
    
    【板块分类】（基于数据动态提取）:
    - 从 categories 字段自动提取板块分类
    - 如：农副、软商、能化、黑色、有色、贵金属等
    
    【分析内容】:
    1. 板块资金汇总及占比
    2. 板块品种数量统计
    3. 板块平均涨跌
    4. 板块多空情绪
    5. 板块内品种排行 (TOP3 + 完整品种列)
    
    Returns:
        dict: {'期货板块': DataFrame}
    """
    sector_sheets = {}
    
    try:
        # 从原始期货数据中提取 categories
        if 'categories' not in futures_df_raw.columns:
            logger.warning("未找到 categories 字段，跳过板块维度分析")
            return sector_sheets
        
        # 建立合约到板块的映射
        sector_map = {}
        
        for _, row in futures_df_raw.iterrows():
            symbol = row.get('instrument_id') or row.get('symbol', '')
            categories = row.get('categories')
            
            if symbol:
                sector_name = extract_category_name(categories)
                sector_map[symbol] = sector_name
        
        # 添加板块列
        futures_df_analysis['板块'] = futures_df_analysis['合约'].map(sector_map)
        
        # 按板块分组统计
        sector_data = []
        grouped = futures_df_analysis.groupby('板块')
        
        total_capital = futures_df_analysis['沉淀资金(亿)'].sum()
        
        for sector_name, group in grouped:
            if pd.isna(sector_name) or sector_name == '未分类':
                continue
            
            sector_capital = group['沉淀资金(亿)'].sum()
            
            # 板块汇总统计
            sector_summary = {
                '板块名称': sector_name,
                '品种数量': group['品种代码'].nunique(),
                '合约数量': len(group),
                '沉淀资金(亿)': round(sector_capital, 4),
                '成交资金(亿)': round(group['成交资金(亿)'].sum(), 4),
                '资金占比%': round(sector_capital / total_capital * 100, 2) if total_capital > 0 else 0,
                '平均杠杆涨跌%': round(group['杠杆涨跌%'].mean(), 2),
                '最大杠杆涨跌%': round(group['杠杆涨跌%'].max(), 2),
                '最小杠杆涨跌%': round(group['杠杆涨跌%'].min(), 2),
                '总持仓量': int(group['持仓量'].sum()),
                '总成交量': int(group['成交量'].sum()),
                '看多合约数': len(group[group['流向信号'] > 0]),
                '看空合约数': len(group[group['流向信号'] < 0]),
                '中性合约数': len(group[group['流向信号'] == 0]),
            }
            
            # 板块情绪判断
            bullish_ratio = sector_summary['看多合约数'] / sector_summary['合约数量']
            if bullish_ratio >= 0.6:
                sector_summary['板块情绪'] = '偏多'
            elif bullish_ratio <= 0.4:
                sector_summary['板块情绪'] = '偏空'
            else:
                sector_summary['板块情绪'] = '中性'
            
            # 板块内品种排行（按沉淀资金）
            product_ranking = group.groupby('品种代码').agg({
                '沉淀资金(亿)': 'sum',
                '成交资金(亿)': 'sum',
                '杠杆涨跌%': 'mean',
                '持仓量': 'sum',
                '成交量': 'sum',
                '流向信号': lambda x: (x > 0).sum() - (x < 0).sum()  # 多空差值
            }).reset_index()
            
            product_ranking.columns = ['品种代码', '沉淀资金(亿)', '成交资金(亿)', 
                                        '平均杠杆涨跌%', '总持仓量', '总成交量', '多空信号']
            product_ranking = product_ranking.sort_values('沉淀资金(亿)', ascending=False)
            
            # 品种情绪
            product_ranking['品种情绪'] = product_ranking['多空信号'].apply(
                lambda x: '看多' if x > 0 else ('看空' if x < 0 else '中性')
            )
            
            # TOP3品种
            top3_products = product_ranking.head(3)['品种代码'].tolist()
            sector_summary['品种TOP3'] = ' | '.join(top3_products)
            
            # 所有品种（完整列表，用 "/" 分隔）
            all_products = product_ranking['品种代码'].tolist()
            sector_summary['品种'] = ' / '.join(all_products)
            
            sector_data.append(sector_summary)
        
        # 创建板块汇总表
        if sector_data:
            sector_df = pd.DataFrame(sector_data)
            sector_df = sector_df.sort_values('沉淀资金(亿)', ascending=False).reset_index(drop=True)
            sector_df.insert(0, '排名', range(1, len(sector_df) + 1))
            sector_sheets['期货板块'] = sector_df
            logger.info(f"生成板块维度汇总: {len(sector_df)} 个板块")
        
    except Exception as e:
        logger.error(f"生成板块维度分析失败: {e}")
        logger.error(traceback.format_exc())
    
    return sector_sheets


def _generate_futures_market_summary(futures_df, correlation_analysis):
    """生成期货市场概览"""
    summary_data = []
    
    # 整体统计
    total_contracts = len(futures_df)
    total_chendian = futures_df['沉淀资金(亿)'].sum()
    total_chengjiao = futures_df['成交资金(亿)'].sum()
    
    # 涨跌统计
    up_count = len(futures_df[futures_df['杠杆涨跌%'] > 0])
    down_count = len(futures_df[futures_df['杠杆涨跌%'] < 0])
    flat_count = total_contracts - up_count - down_count
    
    avg_leverage_change = futures_df['杠杆涨跌%'].mean()
    max_leverage_up = futures_df['杠杆涨跌%'].max()
    max_leverage_down = futures_df['杠杆涨跌%'].min()
    
    # 资金流向统计
    bullish_count = len(futures_df[futures_df['流向信号'] > 0])
    bearish_count = len(futures_df[futures_df['流向信号'] < 0])
    neutral_count = len(futures_df[futures_df['流向信号'] == 0])
    
    summary_data = [
        {'指标': '期货合约总数', '数值': total_contracts, '说明': '分析的期货合约数量'},
        {'指标': '期货沉淀资金(亿)', '数值': round(total_chendian, 2), '说明': '持仓量×价格×乘数×保证金率'},
        {'指标': '期货成交资金(亿)', '数值': round(total_chengjiao, 2), '说明': '成交量×价格×乘数×保证金率'},
        {'指标': '上涨品种数', '数值': up_count, '说明': '杠杆涨跌>0的品种'},
        {'指标': '下跌品种数', '数值': down_count, '说明': '杠杆涨跌<0的品种'},
        {'指标': '平盘品种数', '数值': flat_count, '说明': '杠杆涨跌=0的品种'},
        {'指标': '平均杠杆涨跌%', '数值': round(avg_leverage_change, 2), '说明': '所有品种杠杆涨跌均值'},
        {'指标': '最大杠杆涨幅%', '数值': round(max_leverage_up, 2), '说明': '单日最大杠杆收益'},
        {'指标': '最大杠杆跌幅%', '数值': round(max_leverage_down, 2), '说明': '单日最大杠杆亏损'},
        {'指标': '做多信号品种', '数值': bullish_count, '说明': '增仓上涨或减仓下跌'},
        {'指标': '做空信号品种', '数值': bearish_count, '说明': '增仓下跌或减仓上涨'},
        {'指标': '中性品种', '数值': neutral_count, '说明': '无明显方向'},
    ]
    
    # 市场情绪判断
    if bullish_count > bearish_count * 1.5:
        market_sentiment = '市场整体偏多'
    elif bearish_count > bullish_count * 1.5:
        market_sentiment = '市场整体偏空'
    else:
        market_sentiment = '市场情绪中性'
    
    summary_data.append({'指标': '期货市场情绪', '数值': market_sentiment, 
                         '说明': f'多头{bullish_count}个 vs 空头{bearish_count}个'})
    
    # 货权联动统计
    if correlation_analysis:
        corr_df = pd.DataFrame(correlation_analysis)
        
        # 新增共振评分统计
        strong_resonance = len(corr_df[corr_df['共振评分'] >= 7]) if '共振评分' in corr_df.columns else 0
        medium_resonance = len(corr_df[(corr_df['共振评分'] >= 5) & (corr_df['共振评分'] < 7)]) if '共振评分' in corr_df.columns else 0
        warning_count = len(corr_df[corr_df['共振标签'] == '背离']) if '共振标签' in corr_df.columns else 0
        
        summary_data.append({'指标': '货权联动品种', '数值': len(correlation_analysis), 
                             '说明': '同时有期货和期权数据的品种'})
        summary_data.append({'指标': '强共振', '数值': strong_resonance, 
                             '说明': '共振评分≥7，重点跟踪'})
        summary_data.append({'指标': '共振', '数值': medium_resonance, 
                             '说明': '共振评分5-6，值得关注'})
        summary_data.append({'指标': '背离警示', '数值': warning_count, 
                             '说明': '期货期权明显背离，风险提示'})
        
        # 趋势状态分布
        if '期货状态' in corr_df.columns:
            trend_strengthen = len(corr_df[corr_df['期货状态'] == '多头强化'])
            trend_weaken = len(corr_df[corr_df['期货状态'] == '多头衰减'])
            short_strengthen = len(corr_df[corr_df['期货状态'] == '空头强化'])
            short_weaken = len(corr_df[corr_df['期货状态'] == '空头衰减'])
            summary_data.append({'指标': '多头强化品种', '数值': trend_strengthen, 
                                 '说明': '价↑+持仓↑，多头主动建仓'})
            summary_data.append({'指标': '多头衰减品种', '数值': trend_weaken, 
                                 '说明': '价↑+持仓↓，空头止损'})
            summary_data.append({'指标': '空头强化品种', '数值': short_strengthen, 
                                 '说明': '价↓+持仓↑，空头主动建仓'})
            summary_data.append({'指标': '空头衰减品种', '数值': short_weaken, 
                                 '说明': '价↓+持仓↓，多头止损'})
    
    # 资金Top5
    top5_capital = futures_df.nlargest(5, '沉淀资金(亿)')['合约'].tolist()
    summary_data.append({'指标': '资金TOP5', '数值': ', '.join(top5_capital), 
                         '说明': '沉淀资金最大的5个品种'})
    
    # 涨幅Top5
    top5_up = futures_df.nlargest(5, '杠杆涨跌%')['合约'].tolist()
    summary_data.append({'指标': '涨幅TOP5', '数值': ', '.join(top5_up), 
                         '说明': '杠杆涨幅最大的5个品种'})
    
    # 跌幅Top5
    top5_down = futures_df.nsmallest(5, '杠杆涨跌%')['合约'].tolist()
    summary_data.append({'指标': '跌幅TOP5', '数值': ', '.join(top5_down), 
                         '说明': '杠杆跌幅最大的5个品种'})
    
    return pd.DataFrame(summary_data)


def _calculate_display_width(s):
    """计算字符串的显示宽度（中文占2单位，英文占1单位）"""
    if s is None or pd.isna(s):
        return 0
    s = str(s)
    width = 0
    for char in s:
        if ord(char) > 127:  # 中文字符或全角符号
            width += 2
        else:
            width += 1
    return width


def _save_correlation_excel(file_path, sheets_dict):
    """保存联动分析数据到Excel并美化格式"""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_dict.items():
                if df is None or df.empty:
                    continue
                    
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                
                # 表头样式 - 深蓝色
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 自动列宽
                for idx, col in enumerate(df.columns):
                    # 取表头和内容的最大宽度
                    header_width = _calculate_display_width(col)
                    content_width = df[col].astype(str).map(_calculate_display_width).max() if len(df) > 0 else 0
                    
                    # 针对特定列微调
                    if str(col) == '数值':
                        padding = 6  # 调宽数值列
                    elif str(col) in ['沉淀资金合计(亿)', '资金合计(亿)']:
                        padding = 0  # 紧凑显示长表头、短数据列
                    elif '百分比' in str(col) or '%' in str(col) or '评分' in str(col):
                        padding = 1
                    else:
                        padding = 2
                        
                    max_len = max(header_width, content_width) + padding
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 40)
                
                # 条件格式化
                _apply_futures_formatting(ws, df, sheet_name)
                
        logger.info(f"✅ Excel保存成功: {file_path}")
    except Exception as e:
        logger.error(f"保存Excel失败: {e}")
        logger.error(traceback.format_exc())


def _apply_futures_formatting(ws, df, sheet_name):
    """应用期货分析的条件格式化"""
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    purple_fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    
    # 涨跌类列
    change_cols = ['涨跌%', '杠杆涨跌%', '持仓变化', '持仓变化%']
    # 信号类列
    signal_cols = ['流向信号', '共振评分']
    # 趋势状态列
    trend_state_cols = ['趋势状态', '期货状态']
    # 联动状态列
    linkage_cols = ['联动状态', '共振标签']
    # 强度分数列
    score_cols = ['联动总分']
    
    cols = list(df.columns)
    
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            try:
                val = getattr(row, col_name.replace(' ', '_').replace('(', '').replace(')', '').replace('%', ''), None)
                if val is None:
                    continue
                    
                cell = ws.cell(row=r_idx, column=col_idx)
                
                # 涨跌着色
                if col_name in change_cols:
                    try:
                        num_val = float(val)
                        if num_val > 0:
                            cell.fill = red_fill  # 上涨红色
                        elif num_val < 0:
                            cell.fill = green_fill  # 下跌绿色
                    except:
                        pass
                
                # 信号着色
                elif col_name in signal_cols:
                    try:
                        num_val = float(val)
                        if col_name == '共振评分':
                            if num_val >= 7:
                                cell.fill = gold_fill  # 强共振 金色
                            elif num_val >= 5:
                                cell.fill = red_fill   # 共振 红色
                            elif num_val >= 3:
                                cell.fill = yellow_fill  # 中性
                            elif num_val < 1:
                                cell.fill = purple_fill  # 背离
                        else:
                            if num_val >= 2:
                                cell.fill = red_fill  # 强多
                            elif num_val >= 1:
                                cell.fill = yellow_fill  # 偏多
                            elif num_val <= -2:
                                cell.fill = green_fill  # 强空
                            elif num_val <= -1:
                                cell.fill = blue_fill  # 偏空
                    except:
                        pass
                
                # 趋势状态着色
                elif col_name in trend_state_cols:
                    str_val = str(val)
                    if '多头强化' in str_val:
                        cell.fill = red_fill  # 多头强化
                    elif '多头衰减' in str_val:
                        cell.fill = yellow_fill  # 多头减弱
                    elif '空头强化' in str_val:
                        cell.fill = green_fill  # 空头强化
                    elif '空头衰减' in str_val:
                        cell.fill = blue_fill  # 空头减弱
                    elif '震荡' in str_val:
                        cell.fill = purple_fill
                
                # 联动状态着色
                elif col_name in linkage_cols:
                    str_val = str(val)
                    if '确认' in str_val or '强共振' in str_val or '加速' in str_val:
                        cell.fill = gold_fill
                    elif '警惕' in str_val or '背离' in str_val:
                        cell.fill = purple_fill
                    elif '机会' in str_val or '共振' in str_val:
                        cell.fill = red_fill
                    elif '信号' in str_val:
                        cell.fill = yellow_fill
                
                
                # 综合方向着色 (已移除相关列，但保留代码逻辑以免报错)
                elif col_name == '综合方向':
                    if '多' in str(val):
                        cell.fill = red_fill
                    elif '空' in str(val):
                        cell.fill = green_fill
                
                # 期权结构着色
                elif col_name == '期权结构':
                    str_val = str(val)
                    if '看多' in str_val:
                        cell.fill = red_fill
                    elif '看空' in str_val:
                        cell.fill = green_fill
                    elif '波动率' in str_val:
                        cell.fill = blue_fill
                
                # 期权情绪着色
                elif col_name == '期权情绪':
                    str_val = str(val)
                    if '狂热' in str_val:
                        cell.fill = red_fill
                    elif '恐慌' in str_val:
                        cell.fill = green_fill
                    elif '筑底' in str_val:
                        cell.fill = yellow_fill
                    elif '冲高' in str_val:
                        cell.fill = blue_fill
                
                # 共振等级着色
                elif col_name == '共振等级':
                    str_val = str(val)
                    if '⭐⭐⭐⭐' in str_val:
                        cell.fill = gold_fill
                    elif '⭐⭐⭐' in str_val:
                        cell.fill = red_fill
                    elif '⚠️' in str_val:
                        cell.fill = purple_fill
                
                # 强度模型评分着色
                elif col_name in score_cols:
                    try:
                        score = float(val)
                        if col_name == '联动总分':
                            if score >= 80: cell.fill = gold_fill
                            elif score >= 60: cell.fill = red_fill
                            elif score >= 40: cell.fill = yellow_fill
                            else: cell.fill = blue_fill
                        else:
                            # 子维度满分不同，简单按比例
                            max_s = 30 if col_name != '资金评分' else 40
                            if score >= max_s * 0.8: cell.fill = red_fill
                            elif score >= max_s * 0.5: cell.fill = yellow_fill
                    except:
                        pass
                
            except Exception:
                continue


def _copy_sheet(source_ws, target_ws):
    """跨工作簿复制工作表内容及格式"""
    # 复制单元格
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # 复制合并单元格
    for merge_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merge_range))
        
    # 动态优化列宽 (针对合并后的数据)
    for col_idx in range(1, target_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = 0
        # 遍历前100行计算最大宽度 (避免超大型表性能问题)
        for row_idx in range(1, min(target_ws.max_row, 100) + 1):
            cell_val = target_ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                width = _calculate_display_width(cell_val)
                # 针对合并后的表头也应用定向逻辑
                header_val = target_ws.cell(row=1, column=col_idx).value
                if str(header_val) == '数值':
                    padding = 6
                elif str(header_val) in ['沉淀资金合计(亿)', '资金合计(亿)']:
                    padding = 0
                else:
                    padding = 2
                
                if width + padding > max_width:
                    max_width = width + padding
        target_ws.column_dimensions[col_letter].width = min(max_width, 40)
        
    # 复制行高
    for row_idx, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = row_dim.height

    # 冻结窗格
    target_ws.freeze_panes = source_ws.freeze_panes


def merge_analysis_reports():
    """合并期权排行和货权联动到市场概览并保持格式不变"""
    logger.info("🎬 开始合并生成 wisecoin-市场概览.xlsx...")
    
    reports = {
        'FUTURE': OUTPUT_FILE,  # wisecoin-货权联动.xlsx
        'OPTION': OPTION_RANKING_FILE # wisecoin-期权排行.xlsx
    }
    TARGET_FILE = "wisecoin-市场概览.xlsx"
    
    sheet_order = [
        '期货市场', '期权市场', '货权联动', '期货品种', '期货板块', 
        '期货排行', '期权排行', '期货涨跌', '期权痛点', '期权PCR', '方向型期权', '波动率型期权',
        '期货看多', '期权看多', '期货看空', '期权看空', '期权临期', 
        '期货资金', '期权资金', '期权流动性', '期权活跃度'
    ]
    
    try:
        # 加载源工作簿
        wbs = {}
        for key, path in reports.items():
            if os.path.exists(path):
                wbs[key] = openpyxl.load_workbook(path)
                logger.info(f"加载源文件: {path}")
            else:
                logger.warning(f"缺失源文件: {path}")
        
        if not wbs:
            logger.error("未找到任何源分析文件，无法合并")
            return
            
        # 创建新工作簿
        new_wb = openpyxl.Workbook()
        # 移除默认生成的 Sheet
        if "Sheet" in new_wb.sheetnames:
            del new_wb["Sheet"]
            
        added_count = 0
        for sheet_name in sheet_order:
            source_ws = None
            # 在两个工作簿中查找该分页
            for key, wb in wbs.items():
                if sheet_name in wb.sheetnames:
                    source_ws = wb[sheet_name]
                    break
            
            if source_ws:
                new_ws = new_wb.create_sheet(title=sheet_name)
                _copy_sheet(source_ws, new_ws)
                added_count += 1
                logger.debug(f"已添加分页: {sheet_name}")
        
        new_wb.save(TARGET_FILE)
        logger.info(f"✅ 市场概览合并成功: {TARGET_FILE}, 共包含 {added_count} 个分页")
        
    except Exception as e:
        logger.error(f"合并报告失败: {e}")
        logger.error(traceback.format_exc())


if __name__ == "__main__":
    logger.info("="*60)
    logger.info("期货期权联动分析系统启动")
    logger.info("="*60)
    
    try:
        analyze_futures_options_correlation()
        # 执行报告合并
        merge_analysis_reports()
    except Exception as e:
        logger.error(f"分析失败: {e}")
        logger.error(traceback.format_exc())
