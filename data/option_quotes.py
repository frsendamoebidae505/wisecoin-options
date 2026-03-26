#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# data/option_quotes.py
"""
期权行情获取模块。

提供期权合约获取、行情获取、期货标的数据获取等功能。
支持断点续传、分批获取、自动重建 API 等特性。

Example:
    >>> from data.option_quotes import OptionQuotesManager
    >>> from data.tqsdk_client import TqSdkClient
    >>>
    >>> async with TqSdkClient(run_mode=2) as client:
    ...     manager = OptionQuotesManager(client)
    ...     await manager.get_all_option_symbols()
"""

import asyncio
import os
import re
from typing import List, Dict, Optional, Set, Any
from pathlib import Path

import pandas as pd
import numpy as np

from common.config import Config
from common.logger import StructuredLogger
from common.excel_io import ExcelWriter, ExcelReader
from common.exceptions import DataFetchError

# TqSDK 导入（延迟导入以支持测试）
try:
    from tqsdk import TqApi
    TQSDK_AVAILABLE = True
except ImportError:
    TQSDK_AVAILABLE = False
    TqApi = None


# 期权筛选配置
OPTION_FILTER_CONFIG = {
    'exclude_exchanges': ['SSE', 'SZSE'],  # 排除的交易所（股票期权）
    'max_quote_num': 99999,                 # 需要获取的期权行情个数
}


class OptionQuotesManager:
    """
    期权行情管理器。

    提供期权合约获取、行情同步、期货行情获取等功能。
    支持断点续传、分批获取、自动重建 API。

    Attributes:
        client: TqSdkClient 实例。
        config: 配置实例。
        logger: 日志器实例。
        output_dir: 输出目录。

    Example:
        >>> async with TqSdkClient(run_mode=2) as client:
        ...     manager = OptionQuotesManager(client)
        ...     await manager.run_all()
    """

    # 默认文件名
    SYMBOL_EXCEL_FILE = "wisecoin-期权品种.xlsx"
    QUOTE_EXCEL_FILE = "wisecoin-期权行情.xlsx"
    FUTURES_EXCEL_FILE = "wisecoin-期货行情.xlsx"
    FUTURES_NO_OPTION_EXCEL_FILE = "wisecoin-期货行情-无期权.xlsx"

    def __init__(
        self,
        client: 'TqSdkClient',
        config: Optional[Config] = None,
        logger: Optional[StructuredLogger] = None,
        output_dir: Optional[str] = None,
    ):
        """
        初始化期权行情管理器。

        Args:
            client: TqSdkClient 实例。
            config: 配置实例（可选）。
            logger: 日志器实例（可选）。
            output_dir: 输出目录（可选，默认当前目录）。
        """
        self.client = client
        self.config = config or Config()
        self.logger = logger or StructuredLogger("option_quotes")
        self.output_dir = Path(output_dir) if output_dir else Path.cwd()

        # 从配置获取参数
        self.batch_size = self.config.data.quote_batch_size
        self.save_interval = self.config.data.save_interval
        self.api_rebuild_interval = self.config.data.api_rebuild_interval

    @property
    def api(self) -> TqApi:
        """获取 TqApi 实例。"""
        return self.client.api

    def _get_output_path(self, filename: str) -> Path:
        """
        获取输出文件的完整路径。

        Args:
            filename: 文件名。

        Returns:
            完整路径。
        """
        return self.output_dir / filename

    async def get_all_option_symbols(self) -> List[str]:
        """
        获取所有期权合约。

        获取全市场期权合约，过滤掉股票期权，
        按产品分类导出到 Excel 文件。

        Returns:
            过滤后的期权合约列表。

        Raises:
            DataFetchError: 数据获取失败。
        """
        try:
            self.logger.info("开始获取所有期权合约...")

            # 获取所有期权合约
            option_list = await self.api.query_quotes(ins_class='OPTION', expired=False)

            # 过滤掉股票期权（只保留商品期权和股指期权）
            filtered_options = []
            for symbol in option_list:
                exchange = symbol.split('.')[0]
                if exchange not in OPTION_FILTER_CONFIG['exclude_exchanges']:
                    filtered_options.append(symbol)

            filtered_options = sorted(filtered_options)
            self.logger.info(f"获取到 {len(filtered_options)} 个期权合约（已过滤股票期权）")

            # 获取详细的合约信息
            self.logger.info(f"正在获取 {len(filtered_options)} 个合约的详细信息...")

            # 分批获取合约信息，避免超时
            batch_size = 500
            all_symbol_info = []
            for i in range(0, len(filtered_options), batch_size):
                batch = filtered_options[i:i + batch_size]
                try:
                    batch_df = await asyncio.wait_for(
                        self.api.query_symbol_info(batch), 60
                    )
                    if not batch_df.empty:
                        all_symbol_info.append(batch_df)
                    self.logger.info(
                        f"已获取 {min(i + batch_size, len(filtered_options))}/"
                        f"{len(filtered_options)} 个合约信息"
                    )
                except Exception as e:
                    self.logger.warning(f"获取批次 {i}-{i+batch_size} 失败: {e}")

            if not all_symbol_info:
                self.logger.error("未能获取到任何合约信息")
                return filtered_options

            symbol_info_df = pd.concat(all_symbol_info)

            # 按产品分组并保存到 Excel
            if not symbol_info_df.empty:
                output_path = self._get_output_path(self.SYMBOL_EXCEL_FILE)
                self.logger.info(f"正在按产品分类导出到 {output_path}...")

                self._save_symbols_by_product(symbol_info_df, output_path)

            return filtered_options

        except Exception as e:
            self.logger.error(f"获取期权品种列表失败: {e}")
            raise DataFetchError(f"获取期权品种列表失败: {e}") from e

    def _save_symbols_by_product(self, symbol_info_df: pd.DataFrame, output_path: Path):
        """
        按产品分类保存合约信息到 Excel。

        Args:
            symbol_info_df: 合约信息 DataFrame。
            output_path: 输出文件路径。
        """
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
            # 确定哪个列包含标的信息
            mapping_col = 'underlying_symbol' if 'underlying_symbol' in symbol_info_df.columns else 'symbol'

            # 增加 'product' 列用于分组
            def get_product(row):
                symbol = row[mapping_col]
                if pd.isna(symbol):
                    symbol = row.get('symbol', 'Unknown')

                # 格式如 SHFE.cu2401 -> SHFE.cu
                if '.' in str(symbol):
                    parts = str(symbol).split('.')
                    exchange = parts[0]
                    code_match = re.match(r'^[a-zA-Z]+', parts[1])
                    if code_match:
                        code = code_match.group(0)
                        return f"{exchange}.{code}"
                return str(symbol)

            symbol_info_df['product'] = symbol_info_df.apply(get_product, axis=1)

            # 按 product 分组
            grouped = symbol_info_df.groupby('product')
            underlying_count = 0

            for product, group in grouped:
                # Excel 工作表名称限制为 31 个字符
                sheet_name = str(product).replace('.', '_')[-31:]

                # 按照行权价和标的排序
                sort_cols = []
                if 'underlying_symbol' in group.columns:
                    sort_cols.append('underlying_symbol')
                if 'strike_price' in group.columns:
                    sort_cols.append('strike_price')

                if sort_cols:
                    group = group.sort_values(sort_cols)

                # 写入 Excel
                group.to_excel(writer, sheet_name=sheet_name, index=False)

                # 优化格式：自动调整列宽
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(group.columns):
                    max_len = max(
                        group[col].astype(str).map(len).max() if not group[col].empty else 0,
                        len(str(col))
                    ) + 2
                    column_letter = get_column_letter(idx + 1)
                    ws.column_dimensions[column_letter].width = min(max_len, 50)

                underlying_count += 1

        self.logger.info(
            f"成功将 {len(symbol_info_df)} 个期权合约按 "
            f"{underlying_count} 个产品分类导出到 {output_path}"
        )

    async def get_option_quotes_from_excel(self) -> Dict[str, Any]:
        """
        从 Excel 中读取期权品种并获取实时行情。

        支持"断点续传"：识别已有的合约，但行情 Excel 中尚未获取数据的部分。

        Returns:
            获取到的行情数据字典。

        Raises:
            DataFetchError: 数据获取失败。
        """
        try:
            symbol_file = self._get_output_path(self.SYMBOL_EXCEL_FILE)
            quote_file = self._get_output_path(self.QUOTE_EXCEL_FILE)

            if not symbol_file.exists():
                self.logger.error(f"未找到期权品种文件: {symbol_file}")
                return {}

            self.logger.info(f"正在从 {symbol_file} 读取期权品种并获取行情...")

            # 读取 Excel 的所有工作表
            xls = pd.ExcelFile(str(symbol_file))

            # 1. 收集所有需要获取行情的 symbols
            all_symbols_to_fetch = []
            sheet_df_map = {}

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                sheet_df_map[sheet_name] = df
                symbol_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
                if symbol_col not in df.columns:
                    continue
                symbols = df[symbol_col].dropna().tolist()
                all_symbols_to_fetch.extend(symbols)

            if not all_symbols_to_fetch:
                self.logger.error("Excel 中没有找到任何期权合约")
                return {}

            # 去重并排序
            unique_symbols = sorted(list(set(all_symbols_to_fetch)))

            # 应用获取个数限制
            max_num = OPTION_FILTER_CONFIG.get('max_quote_num', 100)
            if len(unique_symbols) > max_num:
                self.logger.info(
                    f"期权合约总数 {len(unique_symbols)} 超过限制 {max_num}，"
                    f"仅获取前 {max_num} 个"
                )
                unique_symbols = unique_symbols[:max_num]

            # ---------------- 断点续传逻辑 ----------------
            all_quote_data: Dict[str, Any] = {}
            already_fetched_symbols: Set[str] = set()

            if quote_file.exists():
                try:
                    self.logger.info(f"检测到已存在的行情文件 {quote_file}，尝试加载断点...")
                    existing_xls = pd.ExcelFile(str(quote_file))
                    for s_name in existing_xls.sheet_names:
                        if s_name in ["Summary", "Progress"]:
                            continue
                        existing_df = pd.read_excel(existing_xls, sheet_name=s_name)
                        if existing_df.empty:
                            continue

                        # 确定 symbol 列
                        s_col = None
                        if 'symbol' in existing_df.columns:
                            s_col = 'symbol'
                        elif 'instrument_id' in existing_df.columns:
                            s_col = 'instrument_id'

                        if not s_col:
                            continue

                        # 将已有数据加载到 all_quote_data
                        for _, row in existing_df.iterrows():
                            sym = row[s_col]
                            # 只有当 last_price 有效才视为已获取
                            if not pd.isna(row.get('last_price')):
                                all_quote_data[sym] = row.to_dict()
                                already_fetched_symbols.add(sym)

                    self.logger.info(
                        f"成功恢复断点：已获取 {len(already_fetched_symbols)} 个合约，"
                        f"剩余 {max(0, len(unique_symbols) - len(already_fetched_symbols))} 个待同步"
                    )
                except Exception as ex:
                    self.logger.warning(f"加载断点文件失败，将从头开始获取: {ex}")

            # 过滤掉已经获取过的 symbols
            symbols_to_fetch_now = [s for s in unique_symbols if s not in already_fetched_symbols]

            if not symbols_to_fetch_now:
                self.logger.info("所有计划内的期权行情已全部获取完成，无需续传。")
                return all_quote_data

            self.logger.info(
                f"共需同步 {len(unique_symbols)} 个期权合约，"
                f"当前实际待同步 {len(symbols_to_fetch_now)} 个"
            )

            # 2. 分批订阅并同步行情
            self.logger.info("开始同步行情数据...")

            processed_new_count = 0
            api_restart_count = 0

            for i in range(0, len(symbols_to_fetch_now), self.batch_size):
                batch = symbols_to_fetch_now[i:i + self.batch_size]
                try:
                    # 直接使用 async 版本的 api.get_quote_list(batch) 获取并等待行情
                    quotes = await self.api.get_quote_list(batch)

                    # 处理获取到的行情
                    for q in quotes:
                        symbol = q._path[-1]
                        # 整理并转换 quote 字段
                        q_dict = dict(q)
                        # 价格补偿逻辑
                        if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
                            if not pd.isna(q_dict.get('pre_close')):
                                q_dict['last_price'] = q_dict['pre_close']

                        all_quote_data[symbol] = q_dict
                        processed_new_count += 1
                        already_fetched_symbols.add(symbol)

                        # 定期保存
                        if processed_new_count % self.save_interval == 0:
                            self._save_quotes_to_excel(
                                quote_file, sheet_df_map, all_quote_data,
                                len(already_fetched_symbols), len(unique_symbols)
                            )

                        # 定期重建 api
                        if processed_new_count % self.api_rebuild_interval == 0:
                            self._save_quotes_to_excel(
                                quote_file, sheet_df_map, all_quote_data,
                                len(already_fetched_symbols), len(unique_symbols)
                            )
                            api_restart_count += 1
                            self.logger.info(f"达到重建阈值，重建 API，第 {api_restart_count} 次")
                            self.client.rebuild_api()

                    # 最后一批也进行保存
                    if i + self.batch_size >= len(symbols_to_fetch_now):
                        self._save_quotes_to_excel(
                            quote_file, sheet_df_map, all_quote_data,
                            len(already_fetched_symbols), len(unique_symbols)
                        )

                except Exception as e:
                    self.logger.error(f"同步批次 {i//self.batch_size + 1} 时发生错误: {e}")

                # 进度显示
                current_total = len(already_fetched_symbols)
                self.logger.info(
                    f"  [行情同步] 进度: {current_total}/{len(unique_symbols)} "
                    f"({current_total/len(unique_symbols):.1%})"
                )

            self.logger.info(f"行情数据获取及保存完成: {quote_file}")
            return all_quote_data

        except Exception as e:
            self.logger.error(f"获取期权行情异常: {e}")
            raise DataFetchError(f"获取期权行情异常: {e}") from e

    def _save_quotes_to_excel(
        self,
        file_path: Path,
        sheet_df_map: Dict[str, pd.DataFrame],
        all_quote_data: Dict[str, Any],
        current_count: int,
        total_count: int
    ):
        """
        内部辅助函数：将已获取的行情保存到 Excel。

        Args:
            file_path: 输出文件路径。
            sheet_df_map: 原始 Sheet 数据映射。
            all_quote_data: 已获取的行情数据。
            current_count: 当前已获取数量。
            total_count: 总数量。
        """
        from openpyxl.utils import get_column_letter

        try:
            self.logger.info(f"  [保存进度] 正在更新 Excel 数据 ({current_count}/{total_count})...")

            with pd.ExcelWriter(str(file_path), engine='openpyxl') as writer:
                for sheet_name, df in sheet_df_map.items():
                    symbol_col = 'instrument_id' if 'instrument_id' in df.columns else 'symbol'
                    if symbol_col not in df.columns:
                        continue

                    # 合并原始字段和行情字段 (仅包含已获取到的 symbol)
                    merged_rows = []
                    for _, row in df.iterrows():
                        symbol = row[symbol_col]
                        if symbol in all_quote_data:
                            combined_data = row.to_dict()
                            q_info = all_quote_data[symbol]
                            combined_data.update(q_info)
                            merged_rows.append(combined_data)

                    if not merged_rows:
                        continue

                    final_df = pd.DataFrame(merged_rows)

                    # 按照 underlying_symbol 、 strike_price 、 option_class 排序
                    sort_priority = ['underlying_symbol', 'strike_price', 'option_class', 'call_or_put']
                    available_sort_keys = [k for k in sort_priority if k in final_df.columns]

                    if available_sort_keys:
                        if 'strike_price' in final_df.columns:
                            final_df['strike_price'] = pd.to_numeric(final_df['strike_price'], errors='coerce')
                        final_df = final_df.sort_values(available_sort_keys)

                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # 自动调整格式
                    ws = writer.sheets[sheet_name]
                    for idx, col in enumerate(final_df.columns):
                        try:
                            max_len = max(
                                final_df[col].astype(str).map(len).max() if not final_df[col].empty else 0,
                                len(str(col))
                            ) + 2
                        except Exception:
                            max_len = 20
                        ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)

        except Exception as e:
            self.logger.error(f"保存 Excel 失败: {e}")

    async def get_underlying_futures_quotes(self) -> pd.DataFrame:
        """
        获取期权标的期货行情。

        读取期权行情文件中的标的，获取其对应的期货行情，
        并保存为期货行情文件，字段与期权行情保持一致。

        Returns:
            期货行情 DataFrame。

        Raises:
            DataFetchError: 数据获取失败。
        """
        quote_file = self._get_output_path(self.QUOTE_EXCEL_FILE)
        futures_file = self._get_output_path(self.FUTURES_EXCEL_FILE)

        if not quote_file.exists():
            self.logger.warning(f"未找到期权行情文件: {quote_file}，无法获取期货行情。")
            return pd.DataFrame()

        try:
            self.logger.info(f"正在读取 {quote_file} 以获取标的期货列表和字段模板...")
            xls = pd.ExcelFile(str(quote_file))
            all_underlyings: Set[str] = set()

            # 获取字段模板 (从第一个数据 Sheet 获取)
            template_columns = []
            for sheet_name in xls.sheet_names:
                if sheet_name in ["Summary", "Progress", "Summary_Stats"]:
                    continue
                df_temp = pd.read_excel(xls, sheet_name=sheet_name, nrows=0)
                template_columns = df_temp.columns.tolist()
                break

            if not template_columns:
                self.logger.error("未能从期权行情文件中提取到字段模板")
                return pd.DataFrame()

            # 遍历所有 Sheet 获取唯一的 underlying_symbol
            for sheet_name in xls.sheet_names:
                if sheet_name in ["Summary", "Progress", "Summary_Stats"]:
                    continue
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'underlying_symbol' in df.columns:
                    symbols = df['underlying_symbol'].dropna().unique().tolist()
                    for s in symbols:
                        if s and isinstance(s, str) and '.' in s:
                            all_underlyings.add(s)

            unique_underlyings = sorted(list(all_underlyings))

            if not unique_underlyings:
                self.logger.warning("未在期权行情中发现有效的标的期货合约。")
                return pd.DataFrame()

            self.logger.info(
                f"获取到 {len(unique_underlyings)} 个标的期货合约，"
                f"准备同步详细信息和行情..."
            )

            # 1. 获取详细的合约信息 (query_symbol_info)
            batch_size = 500
            all_symbol_info_dfs = []
            for i in range(0, len(unique_underlyings), batch_size):
                batch = unique_underlyings[i:i + batch_size]
                try:
                    info_df = await asyncio.wait_for(self.api.query_symbol_info(batch), 30)
                    if not info_df.empty:
                        all_symbol_info_dfs.append(info_df)
                except Exception as e:
                    self.logger.warning(f"获取期货合约信息批次 {i} 失败: {e}")

            if not all_symbol_info_dfs:
                self.logger.error("未能获取到任何期货合约详细信息")
                return pd.DataFrame()

            full_info_df = pd.concat(all_symbol_info_dfs)

            # 2. 获取实时行情 (Quote List) - 分批获取
            futures_quotes = {}
            for i in range(0, len(unique_underlyings), self.batch_size):
                batch = unique_underlyings[i:i + self.batch_size]
                try:
                    quotes = await self.api.get_quote_list(batch)
                    for q in quotes:
                        futures_quotes[q._path[-1]] = q
                    self.logger.info(
                        f"已同步期货行情: "
                        f"{min(i + self.batch_size, len(unique_underlyings))}/"
                        f"{len(unique_underlyings)}"
                    )
                except Exception as e:
                    self.logger.warning(f"获取期货行情批次 {i} 失败: {e}")

            futures_final_data = []

            # 建立快速查找字典
            info_lookup = self._build_info_lookup(full_info_df)

            self.logger.info(
                f"合约信息汇总完成: {len(full_info_df)} 条记录, "
                f"匹配字典大小: {len(info_lookup)}"
            )

            if unique_underlyings and info_lookup:
                self.logger.info(
                    f"匹配测试: Excel标的='{unique_underlyings[0]}', "
                    f"字典首个Key='{list(info_lookup.keys())[0]}'"
                )

            for symbol in unique_underlyings:
                sym_str = str(symbol).strip()
                try:
                    row_data = self._process_future_symbol(
                        symbol=sym_str,
                        info_lookup=info_lookup,
                        futures_quotes=futures_quotes,
                        template_columns=template_columns
                    )
                    if row_data:
                        futures_final_data.append(row_data)
                except Exception as e:
                    self.logger.warning(f"处理合约 {symbol} 数据失败: {e}")

            if not futures_final_data:
                self.logger.error(
                    f"未能整理出任何期货行情数据。尝试匹配了 {len(unique_underlyings)} 个标的，"
                    f"匹配结果字典大小: {len(info_lookup)}"
                )
                return pd.DataFrame()

            futures_df = pd.DataFrame(futures_final_data)

            # 按产品分组导出到 Excel
            self._save_futures_quotes(futures_file, futures_df, template_columns)

            return futures_df

        except Exception as e:
            self.logger.error(f"获取期货行情异常: {e}")
            raise DataFetchError(f"获取期货行情异常: {e}") from e

    def _build_info_lookup(self, full_info_df: pd.DataFrame) -> Dict[str, Dict]:
        """
        建立合约信息快速查找字典。

        Args:
            full_info_df: 合约信息 DataFrame。

        Returns:
            {symbol: info_dict} 映射字典。
        """
        info_lookup = {}
        for idx, row in full_info_df.iterrows():
            s_key = None

            # 1. 检查 index 是否是 symbol (带 '.' 的字符串)
            if isinstance(idx, str) and '.' in idx:
                s_key = idx

            # 2. 如果 index 不是，检查是否存在 'symbol' 列
            if not s_key:
                s_col = row.get('symbol')
                if not pd.isna(s_col) and s_col:
                    s_key = str(s_col)

            # 3. 最后手段：拼接 exchange_id 和 instrument_id
            if not s_key:
                inst = row.get('instrument_id')
                exch = row.get('exchange_id')
                if inst and exch:
                    s_key = str(inst) if '.' in str(inst) else f"{exch}.{inst}"

            if s_key:
                info_lookup[str(s_key).strip()] = row.to_dict()

        return info_lookup

    def _process_future_symbol(
        self,
        symbol: str,
        info_lookup: Dict[str, Dict],
        futures_quotes: Dict[str, Any],
        template_columns: List[str]
    ) -> Optional[Dict]:
        """
        处理单个期货合约数据。

        Args:
            symbol: 合约代码。
            info_lookup: 合约信息查找字典。
            futures_quotes: 期货行情字典。
            template_columns: 模板列名列表。

        Returns:
            处理后的数据字典，如果无法处理则返回 None。
        """
        static_info = info_lookup.get(symbol)

        if static_info is None:
            # 尝试不区分大小写匹配
            sym_upper = symbol.upper()
            for k, v in info_lookup.items():
                if k.upper() == sym_upper:
                    static_info = v
                    break

        if static_info is None:
            return None

        # 获取动态行情
        q = futures_quotes.get(symbol)
        if q is None:
            return None

        q_dict = dict(q)

        # 价格补偿逻辑
        if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
            if not pd.isna(q_dict.get('pre_close')):
                q_dict['last_price'] = q_dict['pre_close']

        # 合并静态信息和行情信息
        combined = static_info.copy()
        combined.update(q_dict)

        # 增加 product 字段用于分组
        if '.' in symbol:
            parts = symbol.split('.')
            exchange = parts[0]
            code_match = re.match(r'^[a-zA-Z]+', parts[1])
            if code_match:
                combined['product'] = f"{exchange}.{code_match.group(0)}"
            else:
                combined['product'] = symbol
        else:
            combined['product'] = "Unknown"

        # 按照模板对齐字段
        aligned_row = {}
        for col in template_columns:
            aligned_row[col] = combined.get(col, np.nan)

        # 确保 product 字段存在
        aligned_row['product'] = combined['product']

        return aligned_row

    def _save_futures_quotes(
        self,
        futures_file: Path,
        futures_df: pd.DataFrame,
        template_columns: List[str]
    ):
        """
        保存期货行情到 Excel。

        Args:
            futures_file: 输出文件路径。
            futures_df: 期货行情 DataFrame。
            template_columns: 模板列名列表。
        """
        from openpyxl.utils import get_column_letter

        self.logger.info(f"正在将期货行情导出到 {futures_file}...")

        with pd.ExcelWriter(str(futures_file), engine='openpyxl') as writer:
            # 1. 汇总 Sheet
            summary_cols = [c for c in template_columns if c in futures_df.columns]
            if 'product' not in summary_cols and 'product' in futures_df.columns:
                summary_cols.append('product')

            futures_df[summary_cols].to_excel(writer, sheet_name='Summary', index=False)

            # 2. 分产品 Sheet
            grouped = futures_df.groupby('product')
            for product, group in grouped:
                sheet_name = str(product).replace('.', '_')[-31:]
                group[summary_cols].to_excel(writer, sheet_name=sheet_name, index=False)

                # 自动调整列宽
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(summary_cols):
                    try:
                        max_len = max(
                            group[col].astype(str).map(len).max() if not group[col].empty else 0,
                            len(str(col))
                        ) + 2
                    except Exception:
                        max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)

        self.logger.info(
            f"期货行情保存完成: {futures_file}，"
            f"共记录 {len(futures_df)} 个标的合约"
        )

    async def get_not_underlying_futures_quotes(self) -> pd.DataFrame:
        """
        获取全市场非期权标的的期货行情。

        1. 使用 api.query_quotes(ins_class='FUTURE', expired=False)
        2. 过滤掉期权行情文件中已有的 underlying_symbol
        3. 形成非标的期货行情文件

        Returns:
            非标的期货行情 DataFrame。

        Raises:
            DataFetchError: 数据获取失败。
        """
        option_file = self._get_output_path(self.QUOTE_EXCEL_FILE)
        target_file = self._get_output_path(self.FUTURES_NO_OPTION_EXCEL_FILE)

        if not option_file.exists():
            self.logger.warning(
                f"未找到期权行情文件: {option_file}，"
                f"无法过滤标的，将获取全市场期货。"
            )
            underlying_symbols = set()
            template_columns = []
        else:
            try:
                self.logger.info(
                    f"正在读取 {option_file} 以获取已有的期权标的列表..."
                )
                xls = pd.ExcelFile(str(option_file))
                underlying_symbols: Set[str] = set()

                # 获取字段模板
                for sheet_name in xls.sheet_names:
                    if sheet_name in ["Summary", "Progress", "Summary_Stats"]:
                        continue
                    df_temp = pd.read_excel(xls, sheet_name=sheet_name, nrows=0)
                    template_columns = df_temp.columns.tolist()
                    break

                for sheet_name in xls.sheet_names:
                    if sheet_name in ["Summary", "Progress", "Summary_Stats"]:
                        continue
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    if 'underlying_symbol' in df.columns:
                        symbols = df['underlying_symbol'].dropna().unique().tolist()
                        for s in symbols:
                            if s and isinstance(s, str):
                                underlying_symbols.add(s.strip())

                self.logger.info(f"获取到 {len(underlying_symbols)} 个期权标的合约。")
            except Exception as e:
                self.logger.error(f"读取期权行情文件失败: {e}")
                underlying_symbols = set()
                template_columns = []

        try:
            # 1. 获取全市场期货
            self.logger.info("正在获取全市场活跃期货列表...")
            all_futures = await self.api.query_quotes(ins_class='FUTURE', expired=False)

            # 2. 过滤掉期权标的
            not_underlying_futures = [
                s for s in all_futures if s.strip() not in underlying_symbols
            ]
            self.logger.info(
                f"全市场期货共 {len(all_futures)} 个，"
                f"其中非期权标的期货共 {len(not_underlying_futures)} 个"
            )

            if not not_underlying_futures:
                self.logger.info("未发现任何非期权标的的期货合约。")
                return pd.DataFrame()

            # 3. 获取详细信息 (query_symbol_info)
            batch_size = 500
            all_symbol_info_dfs = []
            for i in range(0, len(not_underlying_futures), batch_size):
                batch = not_underlying_futures[i:i + batch_size]
                try:
                    info_df = await asyncio.wait_for(self.api.query_symbol_info(batch), 30)
                    if not info_df.empty:
                        all_symbol_info_dfs.append(info_df)
                    self.logger.info(
                        f"已获取 {min(i + batch_size, len(not_underlying_futures))}/"
                        f"{len(not_underlying_futures)} 个期货信息"
                    )
                except Exception as e:
                    self.logger.warning(f"获取期货信息批次 {i} 失败: {e}")

            if not all_symbol_info_dfs:
                self.logger.error("未能获取到任何期货合约详细信息")
                return pd.DataFrame()

            full_info_df = pd.concat(all_symbol_info_dfs)

            # 4. 获取实时行情 (Quote)
            self.logger.info("开始同步行情数据 (分批订阅)...")

            futures_final_data = []

            # 建立快速查找字典
            info_lookup = self._build_info_lookup(full_info_df)

            self.logger.info(
                f"合约信息汇总完成: {len(full_info_df)} 条记录, "
                f"匹配字典大小: {len(info_lookup)}"
            )

            if not_underlying_futures and info_lookup:
                self.logger.info(
                    f"匹配测试: 计划获取='{not_underlying_futures[0]}', "
                    f"字典首个Key='{list(info_lookup.keys())[0]}'"
                )

            for symbol in not_underlying_futures:
                sym_str = str(symbol).strip()
                try:
                    static_info = info_lookup.get(sym_str)

                    if static_info is None:
                        # 尝试不区分大小写匹配
                        sym_upper = sym_str.upper()
                        for k, v in info_lookup.items():
                            if k.upper() == sym_upper:
                                static_info = v
                                break

                    if static_info is None:
                        continue

                    q = await self.api.get_quote(symbol)
                    q_dict = dict(q)

                    if pd.isna(q_dict.get('last_price')) or q_dict.get('last_price', 0) <= 0:
                        if not pd.isna(q_dict.get('pre_close')):
                            q_dict['last_price'] = q_dict['pre_close']

                    combined = static_info.copy()
                    combined.update(q_dict)

                    # 增加 product 字段用于分组
                    if '.' in sym_str:
                        parts = sym_str.split('.')
                        exchange = parts[0]
                        code_match = re.match(r'^[a-zA-Z]+', parts[1])
                        if code_match:
                            combined['product'] = f"{exchange}.{code_match.group(0)}"
                        else:
                            combined['product'] = sym_str
                    else:
                        combined['product'] = "Unknown"

                    # 对齐字段
                    aligned_row = {}
                    cols_to_use = (
                        template_columns
                        if (template_columns and len(template_columns) > 0)
                        else list(combined.keys())
                    )
                    for col in cols_to_use:
                        aligned_row[col] = combined.get(col, np.nan)

                    if 'product' not in aligned_row:
                        aligned_row['product'] = combined['product']

                    futures_final_data.append(aligned_row)
                except Exception as e:
                    self.logger.debug(f"处理合约 {symbol} 数据失败: {e}")

            if not futures_final_data:
                self.logger.error(
                    f"未能整理出任何非期权标的期货行情数据。"
                    f"尝试匹配了 {len(not_underlying_futures)} 个合约，"
                    f"匹配字典大小: {len(info_lookup)}"
                )
                if len(info_lookup) > 0:
                    self.logger.error(f"字典前5个Key示例: {list(info_lookup.keys())[:5]}")
                return pd.DataFrame()

            futures_df = pd.DataFrame(futures_final_data)

            # 5. 保存到 Excel
            self._save_non_underlying_futures(target_file, futures_df)

            return futures_df

        except Exception as e:
            self.logger.error(f"获取非标的期货行情异常: {e}")
            raise DataFetchError(f"获取非标的期货行情异常: {e}") from e

    def _save_non_underlying_futures(self, target_file: Path, futures_df: pd.DataFrame):
        """
        保存非标的期货行情到 Excel。

        Args:
            target_file: 输出文件路径。
            futures_df: 期货行情 DataFrame。
        """
        from openpyxl.utils import get_column_letter

        self.logger.info(f"正在将非标的期货行情导出到 {target_file}...")

        with pd.ExcelWriter(str(target_file), engine='openpyxl') as writer:
            # Summary Sheet
            futures_df.to_excel(writer, sheet_name='Summary', index=False)

            # 分产品 Sheet
            grouped = futures_df.groupby('product')
            for product, group in grouped:
                sheet_name = str(product).replace('.', '_')[-31:]
                group.to_excel(writer, sheet_name=sheet_name, index=False)

                # 调整列宽
                ws = writer.sheets[sheet_name]
                for idx, col in enumerate(group.columns):
                    try:
                        max_len = max(
                            group[col].astype(str).map(len).max(),
                            len(str(col))
                        ) + 2
                    except Exception:
                        max_len = 20
                    ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 60)

        self.logger.info(
            f"非标的期货行情保存完成: {target_file}，"
            f"共记录 {len(futures_df)} 个合约。"
        )

    async def run_all(self) -> Dict[str, Any]:
        """
        执行完整的数据获取流程。

        依次执行：
        1. 获取所有期权合约
        2. 获取期权行情
        3. 获取标的期货行情
        4. 获取非标的期货行情

        Returns:
            执行结果统计。
        """
        results = {}

        self.logger.info("开始执行完整数据获取流程...")

        # 1. 获取所有期权合约
        option_symbols = await self.get_all_option_symbols()
        results['option_symbols_count'] = len(option_symbols)

        # 2. 获取期权行情
        option_quotes = await self.get_option_quotes_from_excel()
        results['option_quotes_count'] = len(option_quotes)

        # 3. 获取标的期货行情
        futures_quotes = await self.get_underlying_futures_quotes()
        results['futures_quotes_count'] = len(futures_quotes)

        # 4. 获取非标的期货行情
        non_underlying_futures = await self.get_not_underlying_futures_quotes()
        results['non_underlying_futures_count'] = len(non_underlying_futures)

        self.logger.info("所有期权及标的期货、非标的期货信息处理完成。")
        self.logger.info(f"执行结果: {results}")

        return results


# 便捷函数（兼容旧代码）
async def get_all_option_symbols(client: 'TqSdkClient', **kwargs) -> List[str]:
    """
    获取所有期权合约（便捷函数）。

    Args:
        client: TqSdkClient 实例。
        **kwargs: 其他参数。

    Returns:
        期权合约列表。
    """
    manager = OptionQuotesManager(client, **kwargs)
    return await manager.get_all_option_symbols()


async def get_option_quotes_from_excel(client: 'TqSdkClient', **kwargs) -> Dict[str, Any]:
    """
    从 Excel 获取期权行情（便捷函数）。

    Args:
        client: TqSdkClient 实例。
        **kwargs: 其他参数。

    Returns:
        行情数据字典。
    """
    manager = OptionQuotesManager(client, **kwargs)
    return await manager.get_option_quotes_from_excel()


async def get_underlying_futures_quotes(client: 'TqSdkClient', **kwargs) -> pd.DataFrame:
    """
    获取标的期货行情（便捷函数）。

    Args:
        client: TqSdkClient 实例。
        **kwargs: 其他参数。

    Returns:
        期货行情 DataFrame。
    """
    manager = OptionQuotesManager(client, **kwargs)
    return await manager.get_underlying_futures_quotes()


async def get_not_underlying_futures_quotes(client: 'TqSdkClient', **kwargs) -> pd.DataFrame:
    """
    获取非标的期货行情（便捷函数）。

    Args:
        client: TqSdkClient 实例。
        **kwargs: 其他参数。

    Returns:
        期货行情 DataFrame。
    """
    manager = OptionQuotesManager(client, **kwargs)
    return await manager.get_not_underlying_futures_quotes()


async def main():
    """命令行入口函数。"""
    from data.tqsdk_client import TqSdkClient
    import sys
    
    print("=" * 60)
    print("WiseCoin 期权行情获取模块")
    print("=" * 60)
    
    # 使用 run_mode=2 (模拟账户)
    async with TqSdkClient(run_mode=2) as client:
        manager = OptionQuotesManager(client)
        
        # 获取期权合约
        print("正在获取期权合约列表...")
        symbols = await manager.get_all_option_symbols()
        print(f"获取到 {len(symbols)} 个期权合约")
        
        # 获取行情
        print("正在获取期权行情...")
        quotes = await manager.get_option_quotes_from_excel()
        print(f"行情数据已保存")


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())

