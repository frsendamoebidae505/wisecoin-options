import sys
import time
import os
import pandas as pd
import numpy as np
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QComboBox, QPushButton, QGroupBox, 
                             QGridLayout, QSplitter, QMessageBox, QDialog, QTextEdit,
                             QScrollArea)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor, QBrush, QFont
import datetime
import re
import io
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import griddata
from scipy.stats import kurtosis, skew
import platform
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# 配置matplotlib中文显示
if platform.system() == 'Darwin':
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
elif platform.system() == 'Windows':
    plt.rcParams['font.sans-serif'] = ['SimHei']
else:
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


from scipy import stats

# ============ 高性能期权定价模块 (基于 Black-Scholes + Let's Be Rational 算法) ============
# 移植自 16wisecoin_options_client_iv.py

class OptionPricer:
    """
    高性能期权定价器
    基于 Black-Scholes 模型，使用 Let's Be Rational 算法计算隐含波动率
    支持向量化批量计算
    """
    
    # 数值常量
    SQRT_2PI = np.sqrt(2.0 * np.pi)
    ONE_OVER_SQRT_2PI = 1.0 / np.sqrt(2.0 * np.pi)
    SQRT_2 = np.sqrt(2.0)
    
    # 隐含波动率计算的边界和精度
    IV_MIN = 1e-6
    IV_MAX = 5.0  # 500% 波动率上限
    IV_PRECISION = 1e-10
    MAX_ITERATIONS = 100
    
    @staticmethod
    def norm_cdf(x):
        """标准正态分布累积分布函数 (向量化)"""
        return stats.norm.cdf(x)
    
    @staticmethod
    def norm_pdf(x):
        """标准正态分布概率密度函数 (向量化)"""
        return stats.norm.pdf(x)
    
    @classmethod
    def d1(cls, S, K, r, sigma, T):
        """计算 Black-Scholes d1 参数 (向量化)"""
        with np.errstate(divide='ignore', invalid='ignore'):
            sqrt_T = np.sqrt(T)
            result = (np.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * sqrt_T)
            # 处理无效值
            result = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, result)
            return result
    
    @classmethod
    def d2(cls, S, K, r, sigma, T):
        """计算 Black-Scholes d2 参数 (向量化)"""
        return cls.d1(S, K, r, sigma, T) - sigma * np.sqrt(T)
    
    @classmethod
    def bs_price(cls, S, K, r, sigma, T, option_type):
        """
        计算 Black-Scholes 期权价格 (向量化)
        """
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        
        # 避免 T=0 导致的除零
        T = np.maximum(T, 1e-6)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        discount = np.exp(-r * T)
        
        # 处理期权类型
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                price = S * cls.norm_cdf(d_1) - K * discount * cls.norm_cdf(d_2)
            else:
                price = K * discount * cls.norm_cdf(-d_2) - S * cls.norm_cdf(-d_1)
        else:
            # 向量化处理
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_price = S * cls.norm_cdf(d_1) - K * discount * cls.norm_cdf(d_2)
            put_price = K * discount * cls.norm_cdf(-d_2) - S * cls.norm_cdf(-d_1)
            price = np.where(is_call, call_price, put_price)
        
        # 处理无效值
        price = np.where((sigma <= 0) | (T <= 0) | (S <= 0) | (K <= 0), np.nan, price)
        
        # 买方期权价格不应为负 (虽然BS公式理论上保证非负，但数值计算可能有误差)
        price = np.maximum(price, 0.0)
        
        # 标量返回处理
        if price.shape == ():
            return float(price)
        return price
    
    @classmethod
    def delta(cls, S, K, r, sigma, T, option_type):
        """计算 Delta"""
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        T = np.maximum(T, 1e-6)

        d_1 = cls.d1(S, K, r, sigma, T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                return cls.norm_cdf(d_1)
            else:
                return cls.norm_cdf(d_1) - 1
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_delta = cls.norm_cdf(d_1)
            put_delta = call_delta - 1
            return np.where(is_call, call_delta, put_delta)
    
    @classmethod
    def gamma(cls, S, K, r, sigma, T):
        """计算 Gamma"""
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        T = np.maximum(T, 1e-6)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        with np.errstate(divide='ignore', invalid='ignore'):
            gamma = cls.norm_pdf(d_1) / (S * sigma * np.sqrt(T))
            gamma = np.where((sigma <= 0) | (S <= 0), np.nan, gamma)
        return gamma
    
    @classmethod
    def theta(cls, S, K, r, sigma, T, option_type):
        """计算 Theta (返回值为负数)"""
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        T = np.maximum(T, 1e-6)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        sqrt_T = np.sqrt(T)
        discount = np.exp(-r * T)
        
        # 第一项：时间价值损耗
        term1 = -S * cls.norm_pdf(d_1) * sigma / (2 * sqrt_T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                term2 = -r * K * discount * cls.norm_cdf(d_2)
            else:
                term2 = r * K * discount * cls.norm_cdf(-d_2)
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_term2 = -r * K * discount * cls.norm_cdf(d_2)
            put_term2 = r * K * discount * cls.norm_cdf(-d_2)
            term2 = np.where(is_call, call_term2, put_term2)
        
        theta = term1 + term2
        theta = np.where((sigma <= 0) | (S <= 0) | (K <= 0), np.nan, theta)
        return theta
    
    @classmethod
    def vega(cls, S, K, r, sigma, T):
        """计算 Vega (波动率变化 1% (0.01) 时期权价格变化)"""
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        T = np.maximum(T, 1e-6)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        vega = S * np.sqrt(T) * cls.norm_pdf(d_1)
        vega = np.where((sigma <= 0) | (S <= 0), np.nan, vega)
        return vega / 100.0  # 返回每1%波动的变化 (Wisecoin习惯)

    @classmethod
    def rho(cls, S, K, r, sigma, T, option_type):
        """计算 Rho"""
        S = np.asarray(S, dtype=np.float64)
        K = np.asarray(K, dtype=np.float64)
        sigma = np.asarray(sigma, dtype=np.float64)
        T = np.asarray(T, dtype=np.float64)
        T = np.maximum(T, 1e-6)
        
        d_1 = cls.d1(S, K, r, sigma, T)
        d_2 = d_1 - sigma * np.sqrt(T)
        
        discount = np.exp(-r * T)
        
        if isinstance(option_type, str):
            is_call = option_type.upper() == 'CALL'
            if is_call:
                rho = K * T * discount * cls.norm_cdf(d_2)
            else:
                rho = -K * T * discount * cls.norm_cdf(-d_2)
        else:
            option_type = np.asarray(option_type)
            is_call = np.char.upper(option_type.astype(str)) == 'CALL'
            call_rho = K * T * discount * cls.norm_cdf(d_2)
            put_rho = -K * T * discount * cls.norm_cdf(-d_2)
            rho = np.where(is_call, call_rho, put_rho)
        
        rho = np.where((sigma <= 0) | (S <= 0) | (K <= 0), np.nan, rho)
        return rho / 100.0 # 返回每1%利率的变化


class PriceSimulator:
    """价格路径模拟器 - 几何布朗运动（GBM）"""
    
    @staticmethod
    def simulate_gbm(S0, mu, sigma, T, n_paths=1000, n_steps=30):
        """
        使用几何布朗运动模拟价格路径
        
        参数:
            S0: 初始价格
            mu: 漂移率（年化）
            sigma: 波动率（年化）
            T: 时间长度（年）
            n_paths: 路径数量
            n_steps: 时间步数
        
        返回:
            price_paths: (n_paths, n_steps+1) 价格路径矩阵
            time_points: (n_steps+1,) 时间点数组
        """
        dt = T / n_steps
        time_points = np.linspace(0, T, n_steps + 1)
        
        # 布朗运动增量
        dW = np.random.normal(0, np.sqrt(dt), (n_paths, n_steps))
        
        # 初始化价格路径
        price_paths = np.zeros((n_paths, n_steps + 1))
        price_paths[:, 0] = S0
        
        # 计算价格路径：S(t+dt) = S(t) * exp((mu - 0.5*sigma^2)*dt + sigma*dW)
        for i in range(1, n_steps + 1):
            price_paths[:, i] = price_paths[:, i-1] * np.exp(
                (mu - 0.5 * sigma**2) * dt + sigma * dW[:, i-1]
            )
        
        return price_paths, time_points
    
    @staticmethod
    def calculate_terminal_pnl(price_paths, positions, r=0.015):
        """
        计算各路径在到期时的盈亏
        
        参数:
            price_paths: (n_paths, n_steps+1) 价格路径
            positions: 持仓列表
            r: 无风险利率
        
        返回:
            terminal_pnl: (n_paths,) 各路径的到期盈亏
        """
        n_paths = price_paths.shape[0]
        terminal_prices = price_paths[:, -1]  # 到期价格
        terminal_pnl = np.zeros(n_paths)
        
        # 统计有效持仓数
        valid_positions = 0
        
        for pos in positions:
            if pos.get('is_option'):
                # 期权到期盈亏
                strike = pos.get('strike', 0)
                opt_type = pos.get('option_type', 'CALL')
                qty = pos.get('net_pos', 0)
                mult = pos.get('multiplier', 10)
                avg_open_price = pos.get('avg_open_price', 0)
                
                # 数据验证
                if avg_open_price <= 0:
                    avg_open_price = pos.get('option_price', 0) or 0
                
                if strike <= 0 or avg_open_price <= 0 or mult <= 0 or qty == 0:
                    continue
                
                try:
                    # 到期内在价值
                    if opt_type == 'CALL':
                        intrinsic = np.maximum(terminal_prices - strike, 0)
                    else:
                        intrinsic = np.maximum(strike - terminal_prices, 0)
                    
                    # 计算损益（区分买方/卖方）
                    if qty > 0:  # 买方
                        pos_pnl = (intrinsic - avg_open_price) * qty * mult
                    else:  # 卖方
                        pos_pnl = (avg_open_price - intrinsic) * abs(qty) * mult
                    
                    terminal_pnl += pos_pnl
                    valid_positions += 1
                except Exception as e:
                    continue
            else:
                # 期货盈亏
                qty = pos.get('net_pos', 0)
                mult = pos.get('multiplier', 10)
                entry_price = pos.get('avg_open_price', 0) or pos.get('price', 0)
                
                # 数据验证
                if entry_price <= 0 or mult <= 0 or qty == 0:
                    continue
                
                try:
                    pos_pnl = (terminal_prices - entry_price) * qty * mult
                    terminal_pnl += pos_pnl
                    valid_positions += 1
                except Exception as e:
                    continue
        
        # 如果没有有效持仓，返回全NaN数组
        if valid_positions == 0:
            return np.full(n_paths, np.nan)
        
        return terminal_pnl


class RiskMetrics:
    """专业风险度量指标"""
    
    @staticmethod
    def calculate_var(pnl_array, confidence=0.95):
        """
        计算VaR (Value at Risk)
        
        参数:
            pnl_array: 盈亏数组
            confidence: 置信水平
        
        返回:
            VaR值（负数表示损失）
        """
        if len(pnl_array) == 0:
            return 0
        return np.percentile(pnl_array, (1 - confidence) * 100)
    
    @staticmethod
    def calculate_cvar(pnl_array, confidence=0.95):
        """
        计算CVaR (Conditional VaR / Expected Shortfall)
        
        参数:
            pnl_array: 盈亏数组
            confidence: 置信水平
        
        返回:
            CVaR值（尾部平均损失）
        """
        if len(pnl_array) == 0:
            return 0
        var = RiskMetrics.calculate_var(pnl_array, confidence)
        tail_losses = pnl_array[pnl_array <= var]
        if len(tail_losses) == 0:
            return var
        return tail_losses.mean()
    
    @staticmethod
    def calculate_sharpe_ratio(returns, risk_free_rate=0.015):
        """
        计算Sharpe比率
        
        参数:
            returns: 收益率序列
            risk_free_rate: 无风险利率（年化）
        
        返回:
            Sharpe比率
        """
        if len(returns) == 0 or returns.std() == 0:
            return 0
        excess_return = returns.mean() - risk_free_rate / 252
        return excess_return / returns.std() * np.sqrt(252)
    
    @staticmethod
    def calculate_sortino_ratio(returns, risk_free_rate=0.015):
        """
        计算Sortino比率（只考虑下行风险）
        
        参数:
            returns: 收益率序列
            risk_free_rate: 无风险利率（年化）
        
        返回:
            Sortino比率
        """
        if len(returns) == 0:
            return 0
        
        excess_return = returns.mean() - risk_free_rate / 252
        downside_returns = returns[returns < 0]
        
        if len(downside_returns) == 0 or downside_returns.std() == 0:
            return 0
        
        downside_std = downside_returns.std()
        return excess_return / downside_std * np.sqrt(252)
    
    @staticmethod
    def calculate_max_drawdown(pnl_series):
        """
        计算最大回撤
        
        参数:
            pnl_series: 盈亏序列
        
        返回:
            最大回撤（绝对值）
        """
        if len(pnl_series) == 0:
            return 0
        
        cumulative = np.maximum.accumulate(pnl_series)
        drawdown = pnl_series - cumulative
        return abs(drawdown.min()) if len(drawdown) > 0 else 0


class OptionTShapeWindow(QMainWindow):
    """期权持仓窗口 - Live 实时版本（每3分钟自动加载文件）"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("WiseCoin 期权 [持仓版]")
        self.setGeometry(50, 50, 1800, 1000)
        
        # 数据存储
        self.market_overview_df = None
        self.option_ref_df = None
        self.positions_df = None
        self.all_positions_data = {} # 存储所有策略持仓 {sheet_name: dataframe}
        self.positions_view_df = None
        self.klines_data = {}  # 期货K线数据 {symbol: DataFrame}
        self.contract_list = []  # 合约列表（标的+交割月）
        self.position_underlyings = set()
        
        # 实时数据目录
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.temp_dir = os.path.join(self.script_dir, "wisecoin_options_client_live_temp")
        
        # 自动刷新配置
        self.refresh_interval = 180000 * 3  # 3分钟
        self.countdown_seconds = 180 * 3
        
        # 初始化UI
        self.init_ui()
        
        # 初始加载已有数据
        self.load_data()
        
        # 启动自动刷新定时器
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.auto_refresh_check)
        self.refresh_timer.start(self.refresh_interval)
        
        # 启动倒计时更新定时器
        self.countdown_timer = QTimer(self)
        self.countdown_timer.timeout.connect(self.update_countdown)
        self.countdown_timer.start(1000)  # 每秒更新

    def is_trading_time(self):
        """判断当前是否为交易时间"""
        now = datetime.datetime.now()
        weekday = now.weekday() # 0-6 (Mon-Sun)
        t = now.time()
        
        # 定义交易时段
        # 日盘: 09:00 - 15:00 (宽容一点到 15:15)
        day_start = datetime.time(9, 0)
        day_end = datetime.time(15, 15)
        
        # 夜盘: 21:00 - 02:30 (次日)
        night_start = datetime.time(21, 0)
        night_end_next_day = datetime.time(2, 30)
        
        # 判断逻辑
        is_trading = False
        
        # 周一至周五: 日盘 + 夜盘
        if 0 <= weekday <= 4:
            if day_start <= t <= day_end:
                 is_trading = True
            elif t >= night_start:
                 is_trading = True
            elif t <= night_end_next_day: # 凌晨属于前一交易日的夜盘
                 is_trading = True
                 
        # 周六: 只有凌晨 (周五夜盘延续)
        elif weekday == 5:
            if t <= night_end_next_day:
                is_trading = True
                
        return is_trading

    def auto_refresh_check(self):
        """自动刷新检查 (只在交易时间触发加载数据)"""
        # 重置倒计时
        self.countdown_seconds = 180 * 3
        
        if self.is_trading_time():
            self.load_data()
        else:
            self.status_label.setText("状态：非交易时间，暂停自动加载")
            self.status_label.setStyleSheet("color: gray;")

    def update_countdown(self):
        """更新倒计时显示"""
        if self.countdown_seconds > 0:
            self.countdown_seconds -= 1
        
        mins = self.countdown_seconds // 60
        secs = self.countdown_seconds % 60
        
        status_text = f"下次刷新: {mins}:{secs:02d}"
        if not self.is_trading_time():
            status_text += " (非交易时间)"
            
        self.countdown_label.setText(status_text)
    
    def init_ui(self):
        """初始化UI界面"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # ========== 顶部控制区 ==========
        controls_layout = QHBoxLayout()
        
        controls_layout.addWidget(QLabel("期权合约:"))
        
        # 上一个按钮
        self.prev_button = QPushButton("◀ 上一个")
        self.prev_button.setFixedWidth(80)
        self.prev_button.clicked.connect(self.on_prev_contract)
        controls_layout.addWidget(self.prev_button)
        
        self.contract_combo = QComboBox()
        self.contract_combo.setMinimumWidth(250)
        self.contract_combo.currentTextChanged.connect(self.on_contract_changed)
        controls_layout.addWidget(self.contract_combo)
        
        # 下一个按钮
        self.next_button = QPushButton("下一个 ▶")
        self.next_button.setFixedWidth(80)
        self.next_button.clicked.connect(self.on_next_contract)
        controls_layout.addWidget(self.next_button)
        
        controls_layout.addSpacing(20)
        
        # 策略/持仓选择
        controls_layout.addWidget(QLabel("策略持仓:"))
        self.strategy_combo = QComboBox()
        self.strategy_combo.setMinimumWidth(200)
        self.strategy_combo.currentTextChanged.connect(self.on_strategy_changed)
        controls_layout.addWidget(self.strategy_combo)
        
        controls_layout.addSpacing(20)

        self.refresh_button = QPushButton("🔄 刷新数据")
        self.refresh_button.clicked.connect(self.load_data)
        controls_layout.addWidget(self.refresh_button)
        
        # 持仓分析按钮
        self.analyze_button = QPushButton("📊 持仓分析")
        self.analyze_button.clicked.connect(self.show_position_analysis)
        self.analyze_button.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        controls_layout.addWidget(self.analyze_button)
        
        controls_layout.addStretch()
        
        # 倒计时标签
        self.countdown_label = QLabel("下次刷新: 3:00")
        self.countdown_label.setStyleSheet("color: #666; font-size: 10pt; margin-right: 10px;")
        controls_layout.addWidget(self.countdown_label)
        
        self.status_label = QLabel("状态：未加载")
        self.status_label.setStyleSheet("color: gray; font-weight: bold;")
        controls_layout.addWidget(self.status_label)
        
        main_layout.addLayout(controls_layout)
        
        # ========== 上方统计信息区（紧凑布局）==========
        stats_layout = QHBoxLayout()
        
        # 1. 标的信息（精简版）
        self.underlying_stats_group = QGroupBox("标的信息")
        self.underlying_stats_grid = QGridLayout(self.underlying_stats_group)
        self.underlying_stats_grid.setHorizontalSpacing(10)
        self.underlying_stats_grid.setVerticalSpacing(3)
        self.underlying_stat_labels = {}
        
        underlying_fields = [
            ("标的合约", "标的合约"),
            ("期货现价", "期货现价"),
            ("杠杆涨跌%", "杠杆涨跌%"),
            ("期货沉淀(亿)", "期货沉淀(亿)"),
            ("期货状态", "期货状态"),
            ("期货方向", "期货方向"),
            ("期货流向", "期货流向"),
        ]
        
        for idx, (label_text, key) in enumerate(underlying_fields):
            row = idx % 4
            col = (idx // 4) * 2
            
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet("font-size: 11pt; color: #0066cc;")
            
            self.underlying_stats_grid.addWidget(label, row, col)
            self.underlying_stats_grid.addWidget(value_label, row, col + 1)
            self.underlying_stat_labels[key] = value_label
        
        stats_layout.addWidget(self.underlying_stats_group)
        
        # 2. 期权信息（精简版）
        self.option_stats_group = QGroupBox("期权信息")
        self.option_stats_grid = QGridLayout(self.option_stats_group)
        self.option_stats_grid.setHorizontalSpacing(10)
        self.option_stats_grid.setVerticalSpacing(3)
        self.option_stat_labels = {}
        
        option_fields = [
            ("期权结构", "期权结构"),
            ("期权情绪", "期权情绪"),
            ("期权PCR", "期权PCR"),
            ("期权沉淀(亿)", "期权沉淀(亿)"),
            ("最大痛点", "最大痛点"),
            ("痛点距离%", "痛点距离%"),
            ("联动状态", "联动状态"),
            ("共振评分", "共振评分"),
        ]
        
        for idx, (label_text, key) in enumerate(option_fields):
            row = idx % 4
            col = (idx // 4) * 2
            
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet("font-size: 11pt; color: #cc6600;")
            
            self.option_stats_grid.addWidget(label, row, col)
            self.option_stats_grid.addWidget(value_label, row, col + 1)
            self.option_stat_labels[key] = value_label
        
        stats_layout.addWidget(self.option_stats_group)
        stats_layout.setStretch(0, 1)
        stats_layout.setStretch(1, 1) # 调整比例
        
        main_layout.addLayout(stats_layout)
        
        # ========== 3. 联动分析（核心分析 + 波动率曲面信息）==========
        linkage_group = QGroupBox("联动分析")
        linkage_layout = QGridLayout(linkage_group)
        linkage_layout.setHorizontalSpacing(8)
        linkage_layout.setVerticalSpacing(4)
        self.linkage_labels = {}
        
        # 优化布局：
        # 列0: 评分体系 | 列1: 波动结构 | 列2: 偏度特征 | 列3: 策略建议
        # 市场解读、波曲策略、适合策略、不适合策略放在同一行
        linkage_fields = [
            # 第一行
            ("共振等级", "共振等级", 0, 0, 1, "#990099"),
            ("期限结构", "期限结构", 0, 2, 1, "#0066cc"),  # From Vol Surface
            ("倾斜方向", "倾斜方向", 0, 4, 1, "#006666"),  # From Vol Surface
            ("共振标签", "共振标签", 0, 6, 1, "#666600"),

            # 第二行
            ("联动总分", "联动总分", 1, 0, 1, "#990099"),
            ("IV/RV比率", "IV/RV比率", 1, 2, 1, "#cc6600"), # From Vol Surface
            ("期限结构差", "期限结构差", 1, 4, 1, "#0066cc"), # From Vol Surface
            ("IV倾斜度", "IV倾斜度", 1, 6, 1, "#006666"),  # From Vol Surface
            
            # 第三行
            ("价格评分", "价格评分", 2, 0, 1, "#990099"),
            ("峰度", "峰度", 2, 2, 1, "#cc6600"),        # From Vol Surface
            ("短期IV", "短期IV", 2, 4, 1, "#0066cc"),      # From Vol Surface
            ("虚值认沽IV", "虚值认沽IV均值", 2, 6, 1, "#006666"), # From Vol Surface
            ("沉淀合计", "沉淀资金合计(亿)", 2, 8, 1, "#666600"),

            # 第四行
            ("情绪评分", "情绪评分", 3, 0, 1, "#990099"),
            ("偏度", "偏度", 3, 2, 1, "#cc6600"),        # From Vol Surface
            ("长期IV", "长期IV", 3, 4, 1, "#0066cc"),      # From Vol Surface
            ("虚值认购IV", "虚值认购IV均值", 3, 6, 1, "#006666"), # From Vol Surface
            ("合约数量", "合约数量", 3, 8, 1, "#666600"),    # From Vol Surface

            # 第五行：策略信息（同一行，对齐前四列，留空第五列）
            ("市场解读", "市场解读", 4, 0, 1, "#FF0000"),
            ("波曲策略", "推荐策略", 4, 2, 1, "#FF0000"), # From Vol Surface
            ("适合策略", "适合策略", 4, 4, 1, "#FF0000"),
            ("不适合策略", "不适合策略", 4, 6, 1, "#008800"),
        ]
        
        for label_text, key, row, col, colspan, color in linkage_fields:
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-size: 10pt;")
            value_label = QLabel("-")
            value_label.setStyleSheet(f"font-size: 11pt; color: {color}; font-weight: bold;")
            value_label.setWordWrap(True)
            
            linkage_layout.addWidget(label, row, col)
            linkage_layout.addWidget(value_label, row, col + 1, 1, colspan)
            self.linkage_labels[key] = value_label
        
        # 设置列宽比例
        for i in range(10):
            linkage_layout.setColumnStretch(i, 1)
        
        main_layout.addWidget(linkage_group)
        
        # ========== 中间：持仓表（持仓全字段 + 期权参考全字段）==========
        t_group = QGroupBox("持仓")
        t_layout = QVBoxLayout(t_group)
        t_layout.setContentsMargins(1, 1, 1, 1)
        t_layout.setSpacing(0)
        
        self.option_t_table = QTableWidget()
        self.option_t_table.setColumnCount(0)
        self.option_t_table.setHorizontalHeaderLabels([])
        
        self.option_t_table.setStyleSheet("""
            QTableWidget {
                border: 2px solid #999999;
                gridline-color: #cccccc;
                font-size: 10pt;
                background-color: white;
            }
            QTableWidget::item {
                padding: 1px 2px;
                border-right: 1px solid #e0e0e0;
            }
            QHeaderView::section {
                background-color: #4a4a4a;
                color: white;
                padding: 2px 3px;
                border: 1px solid #666666;
                font-weight: bold;
                font-size: 9pt;
            }
            QTableWidget::item:selected {
                background-color: #b3d9ff;
            }
            
            /* 垂直滚动条 - 金黄色风格 */
            QScrollBar:vertical {
                border: none;
                background: #f5f5f5;
                width: 14px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #FFD700, stop:1 #FFB900);
                min-height: 30px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:vertical:hover {
                background: #FFA500;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            
            /* 水平滚动条 - 金黄色风格 */
            QScrollBar:horizontal {
                border: none;
                background: #f5f5f5;
                height: 14px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #FFD700, stop:1 #FFB900);
                min-width: 30px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #FFA500;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
        """)
        
        # 设置列宽模式（动态列在加载后设置）
        header = self.option_t_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        
        self.option_t_table.verticalHeader().setVisible(False)
        self.option_t_table.verticalHeader().setDefaultSectionSize(22)  # 设置紧凑行高
        self.option_t_table.setAlternatingRowColors(True)
        self.option_t_table.setSortingEnabled(False)
        self.option_t_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.option_t_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        t_layout.addWidget(self.option_t_table)
        main_layout.addWidget(t_group, stretch=3)
        
        # ========== 下方：图表区（K线 + 波动率曲面 + 微笑曲线）==========
        charts_splitter = QSplitter(Qt.Horizontal)
        
        # 标的K线图（左边1/3）
        kline_group = QGroupBox("标的K线")
        kline_layout = QVBoxLayout(kline_group)
        kline_layout.setContentsMargins(2, 2, 2, 2)
        self.kline_fig = Figure(figsize=(5, 4), dpi=90)
        self.kline_canvas = FigureCanvas(self.kline_fig)
        kline_layout.addWidget(self.kline_canvas)
        
        # 波动率曲面（中间1/3）
        surface_group = QGroupBox("波动率曲面")
        surface_layout = QVBoxLayout(surface_group)
        surface_layout.setContentsMargins(2, 2, 2, 2)
        self.surface_fig = Figure(figsize=(5, 4), dpi=90)
        self.surface_canvas = FigureCanvas(self.surface_fig)
        surface_layout.addWidget(self.surface_canvas)
        
        # 微笑曲线（右边1/3）
        smile_group = QGroupBox("微笑曲线")
        smile_layout = QVBoxLayout(smile_group)
        smile_layout.setContentsMargins(2, 2, 2, 2)
        self.smile_fig = Figure(figsize=(5, 4), dpi=90)
        self.smile_canvas = FigureCanvas(self.smile_fig)
        smile_layout.addWidget(self.smile_canvas)
        
        charts_splitter.addWidget(kline_group)
        charts_splitter.addWidget(surface_group)
        charts_splitter.addWidget(smile_group)
        charts_splitter.setStretchFactor(0, 1)  # K线 1/3
        charts_splitter.setStretchFactor(1, 1)  # 波动率曲面 1/3
        charts_splitter.setStretchFactor(2, 1)  # 微笑曲线 1/3
        
        main_layout.addWidget(charts_splitter, stretch=2)
    
    def load_data(self):
        """加载Excel数据"""
        try:
            self.status_label.setText("状态：加载中...")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            QApplication.processEvents()
            
            # 读取市场概览
            market_file = os.path.join(self.temp_dir, 'wisecoin-市场概览.xlsx')
            if not os.path.exists(market_file):
                market_file = 'wisecoin-市场概览.xlsx' # 回退到主目录
            if not os.path.exists(market_file):
                raise FileNotFoundError(f"未找到文件: {market_file}")
            self.market_overview_df = pd.read_excel(market_file, sheet_name='货权联动')
            
            # 读取期权参考
            # 说明：live_temp 目录下的 "wisecoin-期权参考.xlsx" 可能是“精选/裁剪版”，不一定覆盖全部期权合约。
            # 如果它无法覆盖当前持仓里的期权合约，会导致误判“未发现任何期权持仓”。
            option_file = os.path.join(self.temp_dir, 'wisecoin-期权参考.xlsx')
            if not os.path.exists(option_file):
                option_file = 'wisecoin-期权参考.xlsx'  # 回退到主目录
            if not os.path.exists(option_file):
                raise FileNotFoundError(f"未找到文件: {option_file}")
            option_ref_primary = pd.read_excel(option_file, sheet_name='期权参考')

            # 读取持仓数据 (改为读取 wisecoin-模拟持仓.xlsx 的所有分页)
            positions_file = os.path.join(self.temp_dir, 'wisecoin-模拟持仓.xlsx')
            if not os.path.exists(positions_file):
                positions_file = 'wisecoin-模拟持仓.xlsx'
            
            self.all_positions_data = {}
            if os.path.exists(positions_file):
                try:
                    # 读取所有 sheet
                    self.all_positions_data = pd.read_excel(positions_file, sheet_name=None)
                    print(f"已加载 {len(self.all_positions_data)} 个策略持仓分页")
                except Exception as e:
                    print(f"读取模拟持仓文件失败: {e}")
            else:
                print("未找到持仓文件: wisecoin-模拟持仓.xlsx")

            # 更新策略选择下拉框
            current_strategy = self.strategy_combo.currentText()
            self.strategy_combo.blockSignals(True)
            self.strategy_combo.clear()
            
            strategies = list(self.all_positions_data.keys())
            if strategies:
                self.strategy_combo.addItems(strategies)
                # 尝试保持之前的选择
                if current_strategy and current_strategy in strategies:
                    self.strategy_combo.setCurrentText(current_strategy)
                    self.positions_df = self.all_positions_data[current_strategy]
                else:
                    self.strategy_combo.setCurrentIndex(0)
                    self.positions_df = self.all_positions_data[strategies[0]]
            else:
                self.positions_df = pd.DataFrame()
            
            self.strategy_combo.blockSignals(False)

            # 若 primary 期权参考未覆盖持仓中的期权合约，则用主目录的完整期权参考进行补全
            self.option_ref_df = option_ref_primary
            try:
                if not self.positions_df.empty and '合约代码' in option_ref_primary.columns:
                    contract_col = self._resolve_contract_column(self.positions_df)
                    if contract_col:
                        pos_norm = set(self.positions_df[contract_col].apply(self._normalize_contract_code).unique())
                        opt_norm_primary = set(option_ref_primary['合约代码'].apply(self._normalize_contract_code).unique())
                        inter_primary = pos_norm & opt_norm_primary

                        # 粗略识别持仓中“像期权”的合约（用于诊断输出）
                        option_like_cnt = 0
                        try:
                            s = self.positions_df[contract_col].astype(str)
                            option_like_cnt = int(s.str.contains(r'(-C-|-P-|C\d{4,6}$|P\d{4,6}$)', regex=True, na=False).sum())
                        except Exception:
                            option_like_cnt = 0

                        if len(inter_primary) == 0:
                            base_dir = os.path.dirname(os.path.abspath(__file__))
                            fallback_file = os.path.join(base_dir, 'wisecoin-期权参考.xlsx')
                            if os.path.exists(fallback_file) and os.path.abspath(fallback_file) != os.path.abspath(option_file):
                                option_ref_fallback = pd.read_excel(fallback_file, sheet_name='期权参考')
                                if '合约代码' in option_ref_fallback.columns:
                                    opt_norm_fallback = set(option_ref_fallback['合约代码'].apply(self._normalize_contract_code).unique())
                                    inter_fallback = pos_norm & opt_norm_fallback
                                    if len(inter_fallback) > 0:
                                        self.option_ref_df = pd.concat([option_ref_primary, option_ref_fallback], ignore_index=True)
                                        self.option_ref_df = self.option_ref_df.drop_duplicates(subset=['合约代码'], keep='first')
                                        print(f"ℹ️ 期权参考补全：primary 未覆盖持仓期权合约（交集0，疑似期权 {option_like_cnt} 条），已合并主目录期权参考（交集 {len(inter_fallback)}）")
                                    else:
                                        print(f"⚠️ 期权参考可能不匹配当前持仓（primary交集0，fallback交集0，疑似期权 {option_like_cnt} 条）。将继续加载但可能无法映射标的。")
                            else:
                                print(f"⚠️ 期权参考可能不匹配当前持仓（primary交集0，疑似期权 {option_like_cnt} 条）。将继续加载但可能无法映射标的。")
            except Exception as e:
                print(f"读取/补全期权参考时发生异常（将继续）：{e}")
            
            # 准备持仓视图数据
            self._prepare_positions_data()
            
            # 读取波动率曲面数据 (尝试读取，若不存在则忽略)
            self.vol_surface_df = None
            try:
                # 检查是否存在 '波动率曲面' sheet
                xl = pd.ExcelFile(option_file)
                if '波动率曲面' in xl.sheet_names:
                    self.vol_surface_df = pd.read_excel(option_file, sheet_name='波动率曲面')
            except Exception as e:
                print(f"读取波动率曲面失败 (可能未生成): {e}")
            
            # 读取期货K线数据
            klines_file = os.path.join(self.temp_dir, 'wisecoin-期货K线.xlsx')
            if not os.path.exists(klines_file):
                klines_file = 'wisecoin-期货K线.xlsx'
            
            self.klines_data = {}
            if os.path.exists(klines_file):
                try:
                    klines_xl = pd.ExcelFile(klines_file)
                    for sheet_name in klines_xl.sheet_names:
                        if sheet_name == 'Summary':
                            continue
                        df = pd.read_excel(klines_xl, sheet_name=sheet_name)
                        # 过滤 datetime_str 为空的数据
                        if 'datetime_str' in df.columns:
                            df = df[df['datetime_str'].notna() & (df['datetime_str'] != '')]
                        if not df.empty:
                            # 恢复原始合约代码 (sheet名用_替换了.)
                            symbol = sheet_name.replace('_', '.', 1)
                            self.klines_data[symbol] = df
                    print(f"已加载 {len(self.klines_data)} 个合约的K线数据")
                except Exception as e:
                    print(f"读取期货K线数据失败: {e}")
            
            # 计算涨跌幅
            if '昨收' in self.option_ref_df.columns and '期权价' in self.option_ref_df.columns:
                self.option_ref_df['涨跌幅%'] = (
                    (self.option_ref_df['期权价'] - self.option_ref_df['昨收']) / 
                    self.option_ref_df['昨收'] * 100
                ).round(2)
            
            # 构建合约列表
            self.build_contract_list()
            
            # 更新视图
            if self.contract_combo.count() > 0:
                self.on_contract_changed(self.contract_combo.currentText())
            
            self.status_label.setText("状态：数据已加载")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
            
        except Exception as e:
            error_msg = f"加载数据失败: {str(e)}"
            self.status_label.setText(f"状态：{error_msg}")
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
            QMessageBox.critical(self, "错误", error_msg)

    def _normalize_contract_code(self, code):
        if pd.isna(code):
            return ''
        s = str(code).strip()
        s = re.sub(r'\s+', '', s)
        s = re.sub(r'^[A-Z]+\.', '', s) # 移除交易所前缀
        s = s.replace('-', '') # 移除连字符，兼容 ni2602-C-130000 与 ni2602C130000
        return s.upper()
    
    def _resolve_contract_column(self, df):
        if df is None or df.empty:
            return None
        candidates = [
            '合约代码', 'instrument_id', 'symbol', '合约', '合约编号', '合约名称'
        ]
        for cand in candidates:
            for col in df.columns:
                if str(col).strip().lower() == cand.lower():
                    return col
        return None
    
    def _resolve_underlying_column(self, df):
        if df is None or df.empty:
            return None
        candidates = [
            '标的合约', 'underlying_symbol', 'underlying', '标的'
        ]
        for cand in candidates:
            for col in df.columns:
                if str(col).strip().lower() == cand.lower():
                    return col
        return None
    
    def _prepare_positions_data(self):
        """准备持仓数据（过滤非期权相关合约 + 合并期权参考字段）"""
        self.positions_view_df = None
        self.position_underlyings = set()
        
        if self.positions_df is None or self.positions_df.empty:
            self.positions_df = pd.DataFrame()
            self._setup_positions_table(pd.DataFrame())
            return
        
        if self.option_ref_df is None or self.option_ref_df.empty:
            self.positions_view_df = self.positions_df.copy()
            self._setup_positions_table(self.positions_view_df)
            return

        contract_col = self._resolve_contract_column(self.positions_df)
        if not contract_col:
            self.positions_view_df = self.positions_df.copy()
            self._setup_positions_table(self.positions_view_df)
            return

        # 1. 预处理期权参考表
        option_ref = self.option_ref_df.copy()
        option_ref['__norm_code'] = option_ref['合约代码'].apply(self._normalize_contract_code)
        option_ref['__norm_und'] = option_ref['标的合约'].apply(self._normalize_contract_code)
        
        # 2. 预处理持仓表
        positions = self.positions_df.copy()
        positions['__norm_code'] = positions[contract_col].apply(self._normalize_contract_code)
        
        # 3. 识别【有期权持仓】的标的
        # 将持仓与期权参考表匹配，找出哪些持仓是期权
        option_holdings = positions.merge(
            option_ref[['__norm_code', '标的合约']],
            on='__norm_code',
            how='inner'
        )
        
        # 记录所有有持仓的期权对应的标的合约（全称，如 SHFE.ni2602）
        self.position_underlyings = set(option_holdings['标的合约'].dropna().unique())
        
        if not self.position_underlyings:
            print("⚠️ 未发现任何期权持仓，持仓表将为空")
            self.positions_view_df = pd.DataFrame()
            self._setup_positions_table(self.positions_view_df)
            return

        # 4. 构建过滤掩码
        # 【修改】持仓保留的并集规则：
        # 只要账户中持有了某个品种的【期权】，那么该品种在持仓表中的【所有持仓】（包括期货、期权）都保留。
        # 如果账户中某个品种只有【期货】而没有任何【期权】，则该品种的持仓被过滤掉。
        
        # 找出持仓中真正的期权合约（通过与期权参考表匹配）
        option_ref_norm_set = set(option_ref['__norm_code'].unique())
        option_holdings_norms = set(positions[positions['__norm_code'].isin(option_ref_norm_set)]['__norm_code'].unique())
        
        # 找出这些期权对应的标的合约（全称）
        underlyings_with_options = set(option_ref[option_ref['__norm_code'].isin(option_holdings_norms)]['标的合约'].unique())
        
        # 记录这些标的的标准化代码，用于匹配对应的期货持仓
        allowed_und_norms = set(option_ref[option_ref['标的合约'].isin(underlyings_with_options)]['__norm_und'].unique())
        
        # 金融指数映射 (针对指数期货持仓)
        index_map = {'SSE.000852': 'IM', 'SSE.000300': 'IF', 'SSE.000016': 'IH', 'SSE.000905': 'IC'}
        index_prods_with_options = [index_map[u] for u in underlyings_with_options if u in index_map]

        mask = (
            positions['__norm_code'].isin(option_ref_norm_set) & positions['__norm_code'].isin(option_holdings_norms) | # 是持有的期权
            positions['__norm_code'].isin(allowed_und_norms) | # 是对应品种的期货 (标准码匹配)
            positions[contract_col].astype(str).isin(underlyings_with_options) # 是对应品种的期货 (全称匹配)
        )
        
        # 补充金融指数期货匹配
        if index_prods_with_options:
            mask = mask | positions['__norm_code'].str.startswith(tuple(index_prods_with_options))
            
        filtered_positions = positions[mask].copy()
        
        # 更新 self.position_underlyings，仅保留【有期权持仓】的标的
        self.position_underlyings = underlyings_with_options
        
        if filtered_positions.empty:
            print("⚠️ 未发现任何期权持仓，持仓表将为空")
            self.positions_view_df = pd.DataFrame()
            self._setup_positions_table(self.positions_view_df)
            return

        # 5. 合并期权参考信息用于展示
        # 先建立一个 映射表：norm_code -> 标的合约全称
        norm_to_full_und = {}
        for _, row in option_ref.iterrows():
            norm_to_full_und[row['__norm_code']] = row['标的合约']
            if row['__norm_und'] not in norm_to_full_und:
                norm_to_full_und[row['__norm_und']] = row['标的合约']
        
        def resolve_final_underlying(row):
            code = row['__norm_code']
            if code in norm_to_full_und:
                return norm_to_full_und[code]
            for und, prod in index_map.items():
                if code.startswith(prod):
                    return und
            return ''

        filtered_positions['__position_underlying'] = filtered_positions.apply(resolve_final_underlying, axis=1)
        
        # 最终合并
        merged = filtered_positions.merge(
            option_ref.drop(columns=['__norm_und']),
            on='__norm_code',
            how='left',
            suffixes=('', '_期权参考')
        )
        
        # 排序：按标的品种和合约代码排序，便于查看日历策略
        sort_cols = []
        if '__position_underlying' in merged.columns: sort_cols.append('__position_underlying')
        if contract_col in merged.columns: sort_cols.append(contract_col)
        if sort_cols:
            merged = merged.sort_values(sort_cols)
        
        # 整理显示列
        position_cols = [c for c in self.positions_df.columns if not str(c).startswith('__')]
        option_cols = []
        for col in self.option_ref_df.columns:
            if col in position_cols:
                option_cols.append(f"{col}_期权参考")
            elif not str(col).startswith('__'):
                option_cols.append(col)
        
        display_cols = [c for c in position_cols if c in merged.columns]
        display_cols += [c for c in option_cols if c in merged.columns]

        # 过滤掉用户要求移除的列
        removed_base_names = {
            'pos_long_his', 'pos_long_today', 'pos_short_his', 'pos_short_today',
            'volume_long_his', 'volume_long_today', 'volume_long_frozen_his', 'volume_long_frozen_today',
            'volume_short_his', 'volume_short_today', 'volume_short_frozen_his', 'volume_short_frozen_today',
            'open_cost_long', 'open_cost_short', 'position_price_long', 'position_price_short',
            'position_cost_long', 'position_cost_short', 'float_profit_long', 'float_profit_short',
            'position_profit_long', 'position_profit_short', 'margin_long', 'margin_short',
            'market_value_long', 'market_value_short', 'volume_long_yd', 'volume_short_yd',
            '交易所', '合约代码', '合约名称', '今结', '昨结',
            'volume_long', 'volume_short', 'pos'
        }
        
        display_cols = [
            c for c in display_cols 
            if c not in removed_base_names and not any(c == f"{base}_期权参考" for base in removed_base_names)
        ]
        
        self.positions_view_df = merged
        self.position_display_cols = display_cols
        self._setup_positions_table(self.positions_view_df)
    
    def _setup_positions_table(self, df):
        """根据持仓数据设置表格列"""
        if df is None:
            self.option_t_table.setColumnCount(0)
            self.option_t_table.setHorizontalHeaderLabels([])
            return
        
        display_cols = getattr(self, 'position_display_cols', None)
        if not display_cols:
            display_cols = [c for c in df.columns if not str(c).startswith('__')]
            self.position_display_cols = display_cols
        
        self.option_t_table.clear()
        self.option_t_table.setColumnCount(len(display_cols))
        self.option_t_table.setHorizontalHeaderLabels(display_cols)
        
        header = self.option_t_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
    
    def build_contract_list(self):
        """构建合约列表，仅保留持仓中与期权相关的标的"""
        if self.market_overview_df is None or self.option_ref_df is None:
            return
        
        if '标的合约' not in self.option_ref_df.columns or '交割年月' not in self.option_ref_df.columns:
            return
        
        # 严格过滤：仅保留【有期权持仓】的标的
        if not self.position_underlyings:
            self.contract_list = []
            self.contract_combo.blockSignals(True)
            self.contract_combo.clear()
            self.contract_combo.blockSignals(False)
            return

        contracts_df = self.option_ref_df[
            self.option_ref_df['标的合约'].astype(str).isin(self.position_underlyings)
        ][['标的合约', '交割年月']].drop_duplicates()
        
        if '标的合约' in self.market_overview_df.columns and '期权沉淀(亿)' in self.market_overview_df.columns:
            contracts_df = contracts_df.merge(
                self.market_overview_df[['标的合约', '期权沉淀(亿)']],
                on='标的合约',
                how='left'
            )
            contracts_df = contracts_df.sort_values('期权沉淀(亿)', ascending=False, na_position='last')
        else:
            contracts_df = contracts_df.sort_values(['标的合约', '交割年月'])
        
        self.contract_list = []
        for _, row in contracts_df.iterrows():
            underlying = str(row['标的合约'])
            expiry = str(row['交割年月'])
            deposit = row.get('期权沉淀(亿)', None)
            
            display_text = f"{underlying} {expiry}"
            if pd.notna(deposit):
                display_text += f" (沉淀:{deposit:.2f}亿)"
            
            self.contract_list.append({
                'display': display_text,
                'underlying': underlying,
                'expiry': expiry
            })
        
        current = self.contract_combo.currentText()
        self.contract_combo.blockSignals(True)
        self.contract_combo.clear()
        self.contract_combo.addItems([c['display'] for c in self.contract_list])
        
        if current:
            index = self.contract_combo.findText(current)
            if index >= 0:
                self.contract_combo.setCurrentIndex(index)
            elif self.contract_list:
                self.contract_combo.setCurrentIndex(0)
        elif self.contract_list:
            self.contract_combo.setCurrentIndex(0)
        
        self.contract_combo.blockSignals(False)
        
        # 更新导航按钮状态
        self._update_nav_buttons()
    
    def on_strategy_changed(self, strategy_name):
        """策略/持仓切换时"""
        if not strategy_name:
            return
            
        if strategy_name in self.all_positions_data:
            self.positions_df = self.all_positions_data[strategy_name]
            
            # 重新处理数据并刷新界面
            self.status_label.setText(f"状态：切换到策略 {strategy_name}")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            
            # 1. 准备持仓视图数据
            self._prepare_positions_data()
            
            # 2. 重新构建合约列表（仅包含当前策略相关的合约）
            self.build_contract_list()
            
            # 3. 更新视图
            if self.contract_combo.count() > 0:
                self.on_contract_changed(self.contract_combo.currentText())
            else:
                # 如果没有合约列表（例如只有期货持仓），清空视图
                self.update_option_t_view("", "")
            
            self.status_label.setText(f"状态：已加载策略 {strategy_name}")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")

    def on_contract_changed(self, contract_text):
        """合约变化时"""
        if not contract_text or not self.contract_list:
            return
        
        contract_info = None
        for c in self.contract_list:
            if c['display'] == contract_text:
                contract_info = c
                break
        
        if contract_info:
            self.update_option_t_view(contract_info['underlying'], contract_info['expiry'])
        
        # 更新导航按钮状态
        self._update_nav_buttons()
    
    def on_prev_contract(self):
        """切换到上一个合约"""
        current_idx = self.contract_combo.currentIndex()
        if current_idx > 0:
            self.contract_combo.setCurrentIndex(current_idx - 1)
    
    def on_next_contract(self):
        """切换到下一个合约"""
        current_idx = self.contract_combo.currentIndex()
        if current_idx < self.contract_combo.count() - 1:
            self.contract_combo.setCurrentIndex(current_idx + 1)
    
    def _update_nav_buttons(self):
        """更新导航按钮状态"""
        current_idx = self.contract_combo.currentIndex()
        total = self.contract_combo.count()
        
        self.prev_button.setEnabled(current_idx > 0)
        self.next_button.setEnabled(current_idx < total - 1)
        
        # 更新按钮提示
        if current_idx > 0:
            self.prev_button.setToolTip(self.contract_combo.itemText(current_idx - 1))
        if current_idx < total - 1:
            self.next_button.setToolTip(self.contract_combo.itemText(current_idx + 1))
    
    
    
    def update_option_t_view(self, underlying, expiry):
        """更新视图（统计信息 + 持仓表 + 波动率图表）"""
        if self.market_overview_df is None or self.option_ref_df is None:
            return
        
        self.update_stats(underlying)
        self.update_positions_table(underlying, expiry)
        self.update_vol_charts(underlying, expiry)

    def update_positions_table(self, underlying, expiry):
        """更新持仓表（显示全量并集，不随选择器切换而变化）"""
        self.option_t_table.setRowCount(0)
        
        if self.positions_view_df is None or self.positions_view_df.empty:
            return
        
        df = self.positions_view_df.copy()
        
        if df.empty:
            return
        
        display_cols = getattr(self, 'position_display_cols', [])
        if not display_cols:
            display_cols = [c for c in df.columns if not str(c).startswith('__')]
            self.position_display_cols = display_cols
            self._setup_positions_table(df)
        
        self.option_t_table.setRowCount(len(df))
        
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, col in enumerate(display_cols):
                val = row.get(col, '-')
                if pd.isna(val):
                    text = '-'
                elif isinstance(val, (int, float)):
                    text = self._get_field_value(row, col)
                else:
                    text = str(val)
                
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                self.option_t_table.setItem(row_idx, col_idx, item)
    
    def update_stats(self, underlying):
        """更新统计信息 (货权联动 + 波动率曲面数据匹配)"""
        if self.market_overview_df is None:
            return
        
        if '标的合约' not in self.market_overview_df.columns:
            return
        
        # 1. 匹配 市场概览/货权联动 数据 (按标的合约精确匹配)
        row_match = self.market_overview_df[
            self.market_overview_df['标的合约'].astype(str) == str(underlying)
        ]
        market_row = row_match.iloc[0] if not row_match.empty else None
        
        # 2. 匹配 波动率曲面 分析数据 (按品种代码匹配)
        vol_row = None
        product_code = self._extract_product_code(underlying)
        
        if self.vol_surface_df is not None and product_code:
            # 宽容匹配：去空格、转大写
            if '品种代码' in self.vol_surface_df.columns:
                vol_match = self.vol_surface_df[
                    self.vol_surface_df['品种代码'].astype(str).str.strip().str.upper() == product_code
                ]
                if not vol_match.empty:
                    vol_row = vol_match.iloc[0]
        
        # 3. 如果Excel中没有波动率曲面汇总数据，则根据当前数据实时计算 (确保信息不为空)
        if vol_row is None and self.option_ref_df is not None and product_code:
            try:
                vol_row = self._calculate_vol_surface_metrics(product_code)
            except Exception as e:
                print(f"实时计算波动率指标失败: {e}")
        
        # 4. 更新 UI 标签
        # A. 标的统计信息
        for key, label in self.underlying_stat_labels.items():
            val = '-'
            if market_row is not None:
                val = market_row.get(key, '-')
            
            # 格式化: 期货沉淀(亿) 保留2位小数
            if key == '期货沉淀(亿)' and pd.notna(val) and val != '-':
                try: val = f"{float(val):.2f}"
                except: pass
                
            label.setText(str(val) if pd.notna(val) else '-')
            
        # B. 期权整体统计信息
        for key, label in self.option_stat_labels.items():
            val = '-'
            if market_row is not None:
                val = market_row.get(key, '-')
            
            # 格式化: 期权PCR, 期权沉淀(亿) 保留2位小数
            if key in ['期权PCR', '期权沉淀(亿)'] and pd.notna(val) and val != '-':
                try: val = f"{float(val):.2f}"
                except: pass
                
            label.setText(str(val) if pd.notna(val) else '-')
            
        # C. 货权联动分析 (含波动率曲面汇总字段)
        for key, label in self.linkage_labels.items():
            val = '-'
            # 逻辑：优先从 market_row 获取，若无则从 vol_row 获取
            if market_row is not None and key in market_row.index:
                val = market_row[key]
            
            if (val == '-' or pd.isna(val)) and vol_row is not None and key in vol_row:
                val = vol_row[key]
            
            # 格式化数值 (希腊字母、百分比等)
            if pd.notna(val) and val != '-':
                if key in ['IV倾斜度', '期限结构差', '峰度', '偏度', 'IV/RV比率', '沉淀资金合计(亿)']:
                    try: val = f"{float(val):.2f}"
                    except: pass
                elif ('IV' in key or '%' in key) and isinstance(val, (int, float)):
                    if val < 1.0 and val > 0: # 可能是小数格式
                        val = f"{val*100:.2f}%"
                    else:
                        val = f"{val:.2f}%"
                
            label.setText(str(val) if pd.notna(val) else '-')
    
    def _calculate_vol_surface_metrics(self, product_code):
        """实时计算波动率曲面指标（当Excel中没有时）"""
        # 筛选该品种的所有期权
        mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
        prod_df = self.option_ref_df[mask].copy()
        
        if len(prod_df) < 10:
            return None
        
        result = {}
        
        # 合约数量和资金
        result['合约数量'] = len(prod_df)
        result['资金合计(万)'] = round(prod_df['资金合计(万)'].sum(), 2) if '资金合计(万)' in prod_df.columns else 0
        
        # IV 统计
        iv_values = prod_df['隐含波动率'].dropna()
        hv_values = prod_df['近期波动率'].dropna()
        
        avg_iv = iv_values.mean() if len(iv_values) > 0 else 0
        avg_hv = hv_values.mean() if len(hv_values) > 0 else 0
        result['IV/RV比率'] = round(avg_iv / avg_hv, 2) if avg_hv > 0 else '-'
        
        # 峰度和偏度
        result['峰度'] = round(kurtosis(iv_values, fisher=True), 2) if len(iv_values) >= 4 else 0
        result['偏度'] = round(skew(iv_values), 2) if len(iv_values) >= 4 else 0
        
        # IV Skew (虚值认沽 vs 虚值认购)
        prod_df['期权类型'] = prod_df['期权类型'].apply(self._normalize_option_type)
        otm_puts = prod_df[(prod_df['期权类型'] == 'PUT') & (prod_df['虚实幅度%'] < -5)]
        otm_calls = prod_df[(prod_df['期权类型'] == 'CALL') & (prod_df['虚实幅度%'] < -5)]
        
        put_iv_mean = otm_puts['隐含波动率'].mean() if len(otm_puts) > 0 else 0
        call_iv_mean = otm_calls['隐含波动率'].mean() if len(otm_calls) > 0 else 0
        iv_skew = put_iv_mean - call_iv_mean
        
        result['虚值认沽IV均值'] = round(put_iv_mean, 2)
        result['虚值认购IV均值'] = round(call_iv_mean, 2)
        result['IV倾斜度'] = round(iv_skew, 2)
        result['倾斜方向'] = '看跌倾斜' if iv_skew > 2 else ('看涨倾斜' if iv_skew < -2 else '平坦')
        
        # 期限结构 (短期 vs 长期)
        short_term = prod_df[prod_df['剩余天数'] <= 30]['隐含波动率'].mean()
        long_term = prod_df[prod_df['剩余天数'] > 60]['隐含波动率'].mean()
        term_diff = short_term - long_term if not pd.isna(short_term) and not pd.isna(long_term) else 0
        
        result['短期IV'] = round(short_term, 2) if not pd.isna(short_term) else 0
        result['长期IV'] = round(long_term, 2) if not pd.isna(long_term) else 0
        result['期限结构差'] = round(term_diff, 2)
        result['期限结构'] = '倒挂' if term_diff > 3 else ('升水' if term_diff < -3 else '平坦')
        
        # 市场情绪分类
        iv_rv_ratio = result['IV/RV比率'] if isinstance(result['IV/RV比率'], (int, float)) else 1
        result['市场情绪'] = self._classify_market_sentiment(iv_skew, term_diff, result['峰度'], iv_rv_ratio)
        
        # 推荐策略
        result['推荐策略'] = self._suggest_strategies(result['市场情绪'], iv_rv_ratio, iv_skew, term_diff)
        
        return result
    
    def _classify_market_sentiment(self, iv_skew, term_structure, iv_kurtosis, iv_rv_ratio):
        """根据波动率特征分类市场情绪"""
        if iv_skew > 3 and term_structure > 3:
            return '恐慌下跌'
        elif iv_skew < -3 and term_structure < -3:
            return '狂热上涨'
        elif abs(iv_skew) <= 3 and abs(term_structure) <= 3:
            if iv_kurtosis < -0.5:
                return '窄幅震荡(瘦尾)'
            else:
                return '窄幅震荡'
        elif iv_skew > 0:
            return '震荡筑底'
        elif iv_skew < 0:
            return '震荡冲高'
        return '中性'
    
    def _suggest_strategies(self, sentiment, iv_rv_ratio, iv_skew, term_diff):
        """根据市场情绪推荐策略"""
        strategies = []
        iv_undervalued = iv_rv_ratio < 0.8 if isinstance(iv_rv_ratio, (int, float)) else False
        iv_overvalued = iv_rv_ratio > 1.2 if isinstance(iv_rv_ratio, (int, float)) else False
        
        if '恐慌' in sentiment:
            strategies.append('认沽多头价差')
            if iv_undervalued:
                strategies.append('买入虚值认沽')
        elif '狂热' in sentiment:
            strategies.append('认购多头价差')
            if iv_undervalued:
                strategies.append('买入虚值认购')
        elif '窄幅' in sentiment:
            if iv_overvalued:
                strategies.append('铁鹰式')
                strategies.append('宽跨式空头')
            else:
                strategies.append('日历价差')
        elif '筑底' in sentiment:
            strategies.append('牛市认沽价差')
        elif '冲高' in sentiment:
            strategies.append('保护性认沽')
        else:
            strategies.append('观望')
        
        return ', '.join(strategies[:3])
    
    def update_t_shape_table(self, underlying, expiry):
        """更新T型表格（全字段对称显示，智能多色染色）"""
        self.option_t_table.setRowCount(0)
        
        if self.option_ref_df is None:
            return
        
        df = self.option_ref_df[
            (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
            (self.option_ref_df['交割年月'].astype(str) == str(expiry))
        ].copy()
        
        if df.empty:
            return
        
        if '行权价' not in df.columns or '期权类型' not in df.columns:
            return
        
        df['期权类型'] = df['期权类型'].apply(self._normalize_option_type)
        df = df[df['期权类型'].isin(['CALL', 'PUT'])]
        
        strikes = sorted(df['行权价'].dropna().unique().tolist(), reverse=True)
        
        if not strikes:
            return
        
        atm_price = df['标的现价'].iloc[0] if '标的现价' in df.columns and not df.empty else None
        
        self.option_t_table.setRowCount(len(strikes))
        
        for row_idx, strike in enumerate(strikes):
            call_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'CALL')]
            put_rows = df[(df['行权价'] == strike) & (df['期权类型'] == 'PUT')]
            
            call = call_rows.sort_values('成交金额', ascending=False).iloc[0] if not call_rows.empty else None
            put = put_rows.sort_values('成交金额', ascending=False).iloc[0] if not put_rows.empty else None
            
            is_atm = False
            if atm_price is not None:
                is_atm = abs(strike - atm_price) / atm_price < 0.02
            
            col_idx = 0
            
            # 看跌 Wing (P-...)
            for field in self.all_ref_cols:
                value = self._get_field_value(put, field)
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                
                # 智能染色
                if put is not None:
                    fg_color, bg_color = self._get_cell_color(field, put, is_atm, 'PUT')
                    if fg_color:
                        item.setForeground(QBrush(fg_color))
                    if bg_color:
                        item.setBackground(QBrush(bg_color))
                
                self.option_t_table.setItem(row_idx, col_idx, item)
                col_idx += 1
            
            # 行权价 (Center)
            strike_item = QTableWidgetItem(f"{strike}")
            strike_item.setTextAlignment(Qt.AlignCenter)
            strike_item.setFont(self._get_bold_font())
            if is_atm:
                strike_item.setBackground(QBrush(QColor("#ffff99")))
                strike_item.setForeground(QBrush(QColor("#cc0000")))
            else:
                strike_item.setBackground(QBrush(QColor("#f0f0f0")))
            self.option_t_table.setItem(row_idx, col_idx, strike_item)
            col_idx += 1
            
            # 看涨 Wing (C-...)
            for field in self.all_ref_cols:
                value = self._get_field_value(call, field)
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                
                # 智能染色
                if call is not None:
                    fg_color, bg_color = self._get_cell_color(field, call, is_atm, 'CALL')
                    if fg_color:
                        item.setForeground(QBrush(fg_color))
                    if bg_color:
                        item.setBackground(QBrush(bg_color))
                
                self.option_t_table.setItem(row_idx, col_idx, item)
                col_idx += 1
    
    def _get_cell_color(self, field, row, is_atm, option_type):
        """根据字段类型和值返回前景色和背景色"""
        fg_color = None
        bg_color = None
        
        # ATM 行高亮
        if is_atm:
            bg_color = QColor("#fffacd") if option_type == 'PUT' else QColor("#fff0f5")
        
        try:
            value = row[field] if field in row.index else None
            if pd.isna(value):
                return fg_color, bg_color
            
            # Delta: 渐变色 (绝对值越大颜色越深)
            if field == 'Delta':
                abs_delta = abs(float(value))
                if abs_delta > 0.7:
                    fg_color = QColor("#8B0000")  # 深红 (深度实值)
                elif abs_delta > 0.5:
                    fg_color = QColor("#cc0000")  # 红色
                elif abs_delta > 0.3:
                    fg_color = QColor("#ff6600")  # 橙色
                elif abs_delta < 0.15:
                    fg_color = QColor("#0066cc")  # 蓝色 (深度虚值)
                else:
                    fg_color = QColor("#333333")  # 灰色 (中性)
            
            # 隐含波动率/近期波动率: 高IV警示
            elif field in ['隐含波动率', '近期波动率']:
                v = float(value)
                if v > 50:
                    bg_color = QColor("#ffcccc")  # 浅红 (极高波动)
                    fg_color = QColor("#8B0000")
                elif v > 35:
                    bg_color = QColor("#ffe6cc")  # 浅橙
                    fg_color = QColor("#cc6600")
                elif v < 15:
                    bg_color = QColor("#e6f3ff")  # 浅蓝 (低波动)
                    fg_color = QColor("#0066cc")
            
            # 虚实幅度%: ITM红/OTM蓝 渐变
            elif field == '虚实幅度%':
                v = float(value)
                if v > 10:
                    bg_color = QColor("#ffcccc")  # 深度实值
                    fg_color = QColor("#8B0000")
                elif v > 5:
                    bg_color = QColor("#ffe6e6")
                    fg_color = QColor("#cc0000")
                elif v > 0:
                    fg_color = QColor("#ff6600")
                elif v < -10:
                    bg_color = QColor("#cce6ff")  # 深度虚值
                    fg_color = QColor("#003366")
                elif v < -5:
                    bg_color = QColor("#e6f3ff")
                    fg_color = QColor("#0066cc")
                elif v < 0:
                    fg_color = QColor("#3399ff")
            
            # 溢价率%: 低=绿(便宜), 高=红(贵)
            elif field == '溢价率%':
                v = float(value)
                if v < 2:
                    bg_color = QColor("#ccffcc")  # 绿 (便宜)
                    fg_color = QColor("#006600")
                elif v > 10:
                    bg_color = QColor("#ffcccc")  # 红 (贵)
                    fg_color = QColor("#cc0000")
                elif v > 5:
                    fg_color = QColor("#ff6600")
            
            # 涨跌幅%: 涨红/跌绿
            elif field == '涨跌幅%':
                v = float(value)
                if v > 5:
                    fg_color = QColor("#cc0000")
                    bg_color = QColor("#ffe6e6")
                elif v > 0:
                    fg_color = QColor("#cc0000")
                elif v < -5:
                    fg_color = QColor("#008800")
                    bg_color = QColor("#e6ffe6")
                elif v < 0:
                    fg_color = QColor("#008800")
            
            # 成交量/持仓量: 活跃度高亮
            elif field in ['成交量', '持仓量']:
                v = float(value)
                if v > 10000:
                    bg_color = QColor("#fff3cd")  # 黄色 (高活跃)
                    fg_color = QColor("#856404")
                elif v > 5000:
                    fg_color = QColor("#996600")
            
            # Theta: 负值越大颜色越深
            elif field == 'Theta':
                v = float(value)
                if v < -0.01:
                    fg_color = QColor("#8B0000")  # 深红 (时间损耗大)
                elif v < -0.005:
                    fg_color = QColor("#cc0000")
                elif v < 0:
                    fg_color = QColor("#ff6600")
            
            # Vega: 高敏感度高亮
            elif field == 'Vega':
                v = float(value)
                if v > 0.05:
                    fg_color = QColor("#6600cc")  # 紫色 (高波动敏感)
                elif v > 0.02:
                    fg_color = QColor("#9933ff")
            
            # Gamma: 高Gamma高亮
            elif field == 'Gamma':
                v = float(value)
                if v > 0.001:
                    fg_color = QColor("#cc6600")  # 橙色 (高Gamma)
            
            # 收益%/年化收益%: 收益高低
            elif field in ['收益%', '收益年化%', '杠杆收益%', '杠杆年化%']:
                v = float(value)
                if v > 50:
                    fg_color = QColor("#006600")
                    bg_color = QColor("#ccffcc")
                elif v > 20:
                    fg_color = QColor("#008800")
                elif v < 0:
                    fg_color = QColor("#cc0000")
            
            # 时间价值/时间占比%
            elif field == '时间占比%':
                v = float(value)
                if v > 80:
                    fg_color = QColor("#cc6600")  # 高时间价值占比
                elif v < 20:
                    fg_color = QColor("#0066cc")  # 低时间价值占比
            
            # 买方杠杆
            elif field == '买方杠杆':
                v = float(value)
                if v > 50:
                    fg_color = QColor("#cc0000")
                    bg_color = QColor("#fff0f0")
                elif v > 20:
                    fg_color = QColor("#ff6600")
                elif v > 10:
                    fg_color = QColor("#996600")
            
            # 资金类字段: 大资金高亮
            elif '资金' in field:
                v = float(value)
                if v > 1000:
                    bg_color = QColor("#e6f3ff")
                    fg_color = QColor("#003366")
                elif v > 500:
                    fg_color = QColor("#0066cc")
                    
        except (ValueError, TypeError):
            pass
        
        return fg_color, bg_color
    
    def update_vol_charts(self, underlying, expiry):
        """更新波动率图表和K线图"""
        if self.option_ref_df is None:
            return
        
        # 1. 提取品种代码 (e.g. SHFE.au2602 -> AU)
        product_code = self._extract_product_code(underlying)
        
        # 2. 筛选该品种的所有期权数据 (用于波动率曲面)
        # 需匹配所有属于该品种的标的合约
        if product_code:
            # 临时添加品种代码列进行筛选 (或直接apply)
            # 为了性能，建议在 load_data 时预处理，这里直接apply
            mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
            df_product = self.option_ref_df[mask].copy()
            surface_title = f"{product_code}"
        else:
            # 回退：仅使用当前标的
            df_product = self.option_ref_df[
                self.option_ref_df['标的合约'].astype(str) == str(underlying)
            ].copy()
            surface_title = f"{underlying} 波动率曲面"
        
        # 3. 筛选当前标的和到期日的数据 (用于微笑曲线)
        df_smile = self.option_ref_df[
            (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
            (self.option_ref_df['交割年月'].astype(str) == str(expiry))
        ].copy()
        
        # 4. 绘制K线图
        self.plot_kline_chart(underlying)
        
        # 5. 绘制波动率图表
        self.plot_vol_surface(df_product, surface_title)
        self.plot_vol_smile(df_smile, expiry)

    def _extract_product_code(self, symbol):
        """
        从合约代码提取品种代码 (e.g. SHFE.au2602 -> AU, SSE.000852 -> 000852)
        逻辑与 04/05 脚本保持一致，确保跨表匹配成功
        """
        try:
            s = str(symbol).strip()
            if '.' in s:
                parts = s.split('.')
                code_part = parts[1]
            else:
                code_part = s
            
            # 优先匹配开头的纯字母部分 (如 au2602 -> AU)
            match_letter = re.match(r'^([a-zA-Z]+)', code_part)
            if match_letter:
                return match_letter.group(1).upper()
            
            # 如果开头是数字，则匹配数字部分 (如 000852 -> 000852)
            match_number = re.match(r'^(\d+)', code_part)
            if match_number:
                return match_number.group(1)
                
            return code_part.upper()
        except:
            return ''
    
    def plot_kline_chart(self, underlying):
        """绘制标的K线图（简明清晰大方风格）"""
        self.kline_fig.clear()
        
        # 获取K线数据
        kline_df = self.klines_data.get(underlying)
        
        if kline_df is None or kline_df.empty:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'无K线数据\n{underlying}', 
                    transform=ax.transAxes, ha='center', va='center',
                    fontsize=10, color='gray')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 准备数据
        df = kline_df.copy()
        
        # 确保有必要的列
        required_cols = ['open', 'high', 'low', 'close']
        if not all(col in df.columns for col in required_cols):
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, 'K线数据缺少OHLC字段', 
                    transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 转换数据类型
        for col in required_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 过滤无效数据
        df = df.dropna(subset=required_cols)
        
        if df.empty:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, 'K线数据无效', 
                    transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.kline_canvas.draw()
            return
        
        # 取最近60根K线显示
        df = df.tail(120).reset_index(drop=True)
        
        try:
            ax = self.kline_fig.add_subplot(111)
            
            # 计算涨跌
            up = df['close'] >= df['open']
            down = df['close'] < df['open']
            
            # 颜色设置（中国市场风格：红涨绿跌）
            up_color = '#CC0000'      # 红色-上涨
            down_color = '#009900'    # 绿色-下跌
            
            # K线宽度
            width = 0.6
            width2 = 0.08
            
            # 绘制上涨K线
            ax.bar(df.index[up], df['close'][up] - df['open'][up], width, 
                   bottom=df['open'][up], color=up_color, edgecolor=up_color, linewidth=0.5)
            ax.bar(df.index[up], df['high'][up] - df['close'][up], width2, 
                   bottom=df['close'][up], color=up_color, linewidth=0)
            ax.bar(df.index[up], df['open'][up] - df['low'][up], width2, 
                   bottom=df['low'][up], color=up_color, linewidth=0)
            
            # 绘制下跌K线
            ax.bar(df.index[down], df['close'][down] - df['open'][down], width, 
                   bottom=df['open'][down], color=down_color, edgecolor=down_color, linewidth=0.5)
            ax.bar(df.index[down], df['high'][down] - df['open'][down], width2, 
                   bottom=df['open'][down], color=down_color, linewidth=0)
            ax.bar(df.index[down], df['close'][down] - df['low'][down], width2, 
                   bottom=df['low'][down], color=down_color, linewidth=0)
            
            # 添加MA5和MA20均线
            if len(df) >= 5:
                ma5 = df['close'].rolling(window=5).mean()
                ax.plot(df.index, ma5, color='#FF9900', linewidth=1, label='MA5', alpha=0.8)
            if len(df) >= 20:
                ma20 = df['close'].rolling(window=20).mean()
                ax.plot(df.index, ma20, color='#0066CC', linewidth=1, label='MA20', alpha=0.8)
            
            # 设置标题和标签
            contract_name = underlying.split('.')[-1] if '.' in underlying else underlying
            ax.set_title(f'{contract_name} 日K线', fontsize=10, fontweight='bold')
            
            # 简化X轴刻度
            if 'datetime_str' in df.columns:
                # 只显示首尾和中间几个日期
                tick_positions = [0, len(df)//4, len(df)//2, 3*len(df)//4, len(df)-1]
                tick_positions = [p for p in tick_positions if p < len(df)]
                tick_labels = [str(df['datetime_str'].iloc[p])[:10] if pd.notna(df['datetime_str'].iloc[p]) else '' 
                              for p in tick_positions]
                ax.set_xticks(tick_positions)
                ax.set_xticklabels(tick_labels, fontsize=7, rotation=15)
            else:
                ax.set_xticks([])
            
            # 设置Y轴格式
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            ax.tick_params(axis='y', labelsize=8)
            
            # 网格线
            ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
            ax.set_axisbelow(True)
            
            # 图例
            if len(df) >= 5:
                ax.legend(loc='upper left', fontsize=7, framealpha=0.8)
            
            # 边框美化
            for spine in ['top', 'right']:
                ax.spines[spine].set_visible(False)
            ax.spines['left'].set_color('#888888')
            ax.spines['bottom'].set_color('#888888')
            
            # 显示最新价格
            last_close = df['close'].iloc[-1]
            ax.axhline(y=last_close, color='#666666', linestyle=':', linewidth=0.8, alpha=0.6)
            ax.text(len(df)-1, last_close, f' {last_close:,.1f}', 
                    fontsize=7, va='center', ha='left', color='#333333')
            
        except Exception as e:
            ax = self.kline_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'绘图错误:\n{str(e)[:30]}', 
                    transform=ax.transAxes, ha='center', va='center', fontsize=9)
            ax.axis('off')
        
        self.kline_fig.tight_layout()
        self.kline_canvas.draw()
    
    def _generate_vol_surface_image(self, df, underlying):
        """生成波动率曲面图片（用于导出）"""
        if df is None or df.empty:
            return None
        
        plot_df = df[(df['隐含波动率'] > 0) & (df['剩余天数'] > 0) & (df['行权价'] > 0)].copy()
        
        if len(plot_df) < 6:
            return None
        
        try:
            fig = Figure(figsize=(8, 6), dpi=100)
            ax = fig.add_subplot(111, projection='3d')
            
            # 标准化行权价
            und_price = plot_df['标的现价'].iloc[0] if '标的现价' in plot_df.columns else 0
            if und_price > 0:
                plot_df['Moneyness'] = (plot_df['行权价'] / und_price - 1) * 100
            else:
                plot_df['Moneyness'] = plot_df['行权价']
            
            x = plot_df['Moneyness'].values
            y = plot_df['剩余天数'].values
            z = plot_df['隐含波动率'].values
            
            # 创建网格
            xi = np.linspace(x.min(), x.max(), 35)
            yi = np.linspace(y.min(), y.max(), 35)
            xi_grid, yi_grid = np.meshgrid(xi, yi)
            
            try:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='linear')
            except:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
            
            if np.any(np.isnan(zi_grid)):
                zi_nearest = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
                zi_grid = np.where(np.isnan(zi_grid), zi_nearest, zi_grid)
            
            # 绘制曲面
            surf = ax.plot_surface(xi_grid, yi_grid, zi_grid, cmap='viridis', 
                                   alpha=0.88, edgecolor='none', linewidth=0, antialiased=True)
            
            cbar = fig.colorbar(surf, shrink=0.55, aspect=12, pad=0.1)
            cbar.set_label('IV (%)', fontsize=8)
            
            ax.scatter(x, y, z, color='red', s=12, alpha=0.6, depthshade=True)
            
            y_range = np.linspace(y.min(), y.max(), 20)
            z_atm = np.interp(y_range, y[np.argsort(y)], z[np.argsort(y)])
            ax.plot([0]*len(y_range), y_range, z_atm, color='white', linewidth=2, alpha=0.8, linestyle='--')
            
            ax.set_xlabel('行权价 (%)', fontsize=9, labelpad=8)
            ax.set_ylabel('剩余期限 (天)', fontsize=9, labelpad=8)
            ax.set_zlabel('隐含波动率 (%)', fontsize=9, labelpad=8)
            
            ax.view_init(elev=22, azim=-60)
            ax.tick_params(axis='both', which='major', labelsize=7)
            
            fig.tight_layout()
            
            # 保存到内存
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            return buf
            
        except Exception as e:
            print(f"生成波动率曲面图失败: {e}")
            return None
    
    def _generate_vol_smile_image(self, df, expiry):
        """生成微笑曲线图片（用于导出）"""
        if df is None or df.empty:
            return None
        
        temp = df.copy()
        temp['期权类型'] = temp['期权类型'].apply(self._normalize_option_type)
        temp = temp[(temp['隐含波动率'] > 0) & (temp['行权价'] > 0)]
        
        if temp.empty:
            return None
        
        try:
            fig = Figure(figsize=(8, 5), dpi=100)
            ax = fig.add_subplot(111)
            
            # 计算Moneyness
            und_price = temp['标的现价'].iloc[0] if '标的现价' in temp.columns and pd.notna(temp['标的现价'].iloc[0]) else None
            if und_price and und_price > 0:
                temp['Moneyness'] = (temp['行权价'] / und_price - 1) * 100
                x_label = '价值度 (%)'
                x_col = 'Moneyness'
            else:
                x_label = '行权价'
                x_col = '行权价'
            
            calls = temp[temp['期权类型'] == 'CALL'].sort_values(x_col)
            puts = temp[temp['期权类型'] == 'PUT'].sort_values(x_col)
            
            # 绘制曲线
            if not calls.empty:
                ax.plot(calls[x_col], calls['隐含波动率'], label='看涨 (Call)', 
                        color='#cc0000', marker='o', markersize=6, linewidth=2, alpha=0.85)
            if not puts.empty:
                ax.plot(puts[x_col], puts['隐含波动率'], label='看跌 (Put)', 
                        color='#008800', marker='s', markersize=6, linewidth=2, alpha=0.85)
            
            # 添加ATM线
            if x_col == 'Moneyness':
                ax.axvline(x=0, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label='ATM')
                xlim = ax.get_xlim()
                ylim = ax.get_ylim()
                ax.axvspan(xlim[0], 0, alpha=0.05, color='green')
                ax.axvspan(0, xlim[1], alpha=0.05, color='red')
                ax.set_xlim(xlim)
                ax.set_ylim(ylim)
            elif und_price:
                ax.axvline(x=und_price, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label=f'ATM ({und_price:.2f})')
            
            ax.set_xlabel(x_label, fontsize=10)
            ax.set_ylabel('隐含波动率 (%)', fontsize=10)
            ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.tick_params(axis='both', which='major', labelsize=9)
            
            fig.tight_layout()
            
            # 保存到内存
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            return buf
            
        except Exception as e:
            print(f"生成微笑曲线图失败: {e}")
            return None
    
    def plot_vol_surface(self, df, underlying):
        """绘制波动率曲面 - X轴:行权价, Y轴:剩余期限, Z轴:隐含波动率"""
        self.surface_fig.clear()
        
        if df is None or df.empty:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.surface_canvas.draw()
            return
        
        plot_df = df[(df['隐含波动率'] > 0) & (df['剩余天数'] > 0) & (df['行权价'] > 0)].copy()
        
        if len(plot_df) < 6:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'数据不足 ({len(plot_df)}点)', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.surface_canvas.draw()
            return

        try:
            ax = self.surface_fig.add_subplot(111, projection='3d')
            
            # 标准化行权价 (使用Moneyness百分比)
            und_price = plot_df['标的现价'].iloc[0] if '标的现价' in plot_df.columns else 0
            if und_price > 0:
                plot_df['Moneyness'] = (plot_df['行权价'] / und_price - 1) * 100
            else:
                plot_df['Moneyness'] = plot_df['行权价']
            
            # X轴：行权价(Moneyness)，Y轴：剩余期限（从近到远），Z轴：IV
            x = plot_df['Moneyness'].values      # X轴：Moneyness (%)
            y = plot_df['剩余天数'].values       # Y轴：剩余天数（从近到远）
            z = plot_df['隐含波动率'].values     # Z轴：隐含波动率
            
            # 创建网格
            xi = np.linspace(x.min(), x.max(), 35)
            yi = np.linspace(y.min(), y.max(), 35)
            xi_grid, yi_grid = np.meshgrid(xi, yi)
            
            try:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='linear')
            except:
                zi_grid = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
            
            # 填充NaN值
            if np.any(np.isnan(zi_grid)):
                zi_nearest = griddata((x, y), z, (xi_grid, yi_grid), method='nearest')
                zi_grid = np.where(np.isnan(zi_grid), zi_nearest, zi_grid)
            
            # 绘制曲面 - 业界标准配色viridis
            surf = ax.plot_surface(xi_grid, yi_grid, zi_grid, cmap='viridis', 
                                   alpha=0.88, edgecolor='none', linewidth=0, antialiased=True)
            
            # 添加颜色条
            cbar = self.surface_fig.colorbar(surf, shrink=0.55, aspect=12, pad=0.1)
            cbar.set_label('IV (%)', fontsize=8)
            
            # 绘制实际数据点
            ax.scatter(x, y, z, color='red', s=12, alpha=0.6, depthshade=True)
            
            # 添加ATM参考线 (Moneyness=0)
            y_range = np.linspace(y.min(), y.max(), 20)
            z_atm = np.interp(y_range, y[np.argsort(y)], z[np.argsort(y)])
            ax.plot([0]*len(y_range), y_range, z_atm, color='white', linewidth=2, alpha=0.8, linestyle='--')
            
            # 设置坐标轴 - 业界标准：Strike × Time × IV
            ax.set_xlabel('行权价 (%)', fontsize=9, labelpad=8)
            ax.set_ylabel('剩余期限 (天)', fontsize=9, labelpad=8)
            ax.set_zlabel('隐含波动率 (%)', fontsize=9, labelpad=8)
            
            # 优化视角：从低行权价看向高行权价，近月在前
            ax.view_init(elev=22, azim=-60)
            
            # 设置刻度
            ax.tick_params(axis='both', which='major', labelsize=7)
            
            # 居中优化：手动调整边距，使其稍微偏左
            self.surface_fig.subplots_adjust(left=0.0, right=0.88, top=0.95, bottom=0.05)
            
        except Exception as e:
            ax = self.surface_fig.add_subplot(111)
            ax.text(0.5, 0.5, f'绘图错误: {str(e)[:50]}', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
        
        # self.surface_fig.tight_layout()  # 注释掉 tight_layout 以免覆盖 subplots_adjust
        self.surface_canvas.draw()
    
    def plot_vol_smile(self, df, expiry):
        """绘制微笑曲线 (优化版) - 使用Moneyness标准化，添加ATM标记"""
        self.smile_fig.clear()
        
        if df is None or df.empty:
            ax = self.smile_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.smile_canvas.draw()
            return
        
        temp = df.copy()
        temp['期权类型'] = temp['期权类型'].apply(self._normalize_option_type)
        temp = temp[(temp['隐含波动率'] > 0) & (temp['行权价'] > 0)]
        
        if temp.empty:
            ax = self.smile_fig.add_subplot(111)
            ax.text(0.5, 0.5, '无有效波动率数据', transform=ax.transAxes, ha='center', va='center')
            ax.axis('off')
            self.smile_canvas.draw()
            return
        
        # 计算Moneyness
        und_price = temp['标的现价'].iloc[0] if '标的现价' in temp.columns and pd.notna(temp['标的现价'].iloc[0]) else None
        if und_price and und_price > 0:
            temp['Moneyness'] = (temp['行权价'] / und_price - 1) * 100
            x_label = '价值度 (%)'
            x_col = 'Moneyness'
        else:
            x_label = '行权价'
            x_col = '行权价'
        
        calls = temp[temp['期权类型'] == 'CALL'].sort_values(x_col)
        puts = temp[temp['期权类型'] == 'PUT'].sort_values(x_col)
        
        ax = self.smile_fig.add_subplot(111)
        
        # 绘制看涨和看跌曲线
        if not calls.empty:
            ax.plot(calls[x_col], calls['隐含波动率'], label='看涨 (Call)', 
                    color='#cc0000', marker='o', markersize=6, linewidth=2, alpha=0.85)
        if not puts.empty:
            ax.plot(puts[x_col], puts['隐含波动率'], label='看跌 (Put)', 
                    color='#008800', marker='s', markersize=6, linewidth=2, alpha=0.85)
        
        # 添加ATM垂直线
        if x_col == 'Moneyness':
            ax.axvline(x=0, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label='ATM')
        elif und_price:
            ax.axvline(x=und_price, color='blue', linestyle='--', linewidth=1.5, alpha=0.7, label=f'ATM ({und_price:.2f})')
        
        # 填充ITM/OTM区域
        if x_col == 'Moneyness':
            xlim = ax.get_xlim()
            ylim = ax.get_ylim()
            ax.axvspan(xlim[0], 0, alpha=0.05, color='green', label='_nolegend_')  # ITM for Put
            ax.axvspan(0, xlim[1], alpha=0.05, color='red', label='_nolegend_')    # ITM for Call
            ax.set_xlim(xlim)
            ax.set_ylim(ylim)
        
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel('隐含波动率 (%)', fontsize=10)
        ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # 添加注释
        if not calls.empty or not puts.empty:
            all_iv = pd.concat([calls['隐含波动率'], puts['隐含波动率']]) if not puts.empty else calls['隐含波动率']
            avg_iv = all_iv.mean()
            ax.axhline(y=avg_iv, color='gray', linestyle=':', linewidth=1, alpha=0.6)
            ax.text(ax.get_xlim()[1], avg_iv, f' 平均IV: {avg_iv:.1f}%', 
                    va='center', ha='left', fontsize=8, color='gray')
        
        self.smile_fig.tight_layout()
        self.smile_canvas.draw()
    
    def _normalize_option_type(self, opt_type):
        if pd.isna(opt_type): return ''
        ot = str(opt_type).upper()
        if 'CALL' in ot or ot == 'C': return 'CALL'
        if 'PUT' in ot or ot == 'P': return 'PUT'
        return ''
    
    def _get_field_value(self, row, field):
        if row is None or field not in row.index: return '-'
        value = row[field]
        if pd.isna(value): return '-'
        if isinstance(value, (int, float)):
            if field == 'Gamma': return f"{value:.6f}"
            if field in ['Delta', 'Theta', 'Vega', 'Rho']: return f"{value:.4f}"
            if field in ['隐含波动率', '近期波动率', '虚实幅度%', '涨跌幅%', '收益%', '溢价率%']: return f"{value:.2f}"
            if field in ['期权价', '理论价格', '内在价值', '时间价值', '标的现价', '昨收', '昨结', '今结']: return f"{value:.1f}"
            if '资金' in field or '成交' in field or '持仓' in field or '合约乘数' in field or '交割年月' in field or 'user_id' in field:
                try: return f"{int(value)}"
                except: return f"{value}"
            return f"{value:.2f}"
        return str(value)
    
    def _get_bold_font(self):
        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        return font
    
    def _safe_get_value(self, row, field, default='-'):
        """安全地从Series或dict中获取值，处理NaN和None"""
        try:
            if isinstance(row, pd.Series):
                if field in row.index:
                    val = row[field]
                    if pd.notna(val):
                        return val
            elif isinstance(row, dict):
                if field in row:
                    val = row[field]
                    if val is not None and (not isinstance(val, float) or not pd.isna(val)):
                        return val
        except:
            pass
        return default
    
    def _calculate_display_width(self, text):
        """计算字符串的显示宽度（中文字符算2个宽度，英文算1个）"""
        if text is None:
            return 0
        text_str = str(text)
        width = 0
        for char in text_str:
            # 判断是否为中文字符（包括中文标点）
            if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                width += 2
            else:
                width += 1
        return width
    
    
    def _adjust_column_widths_for_mixed_data(self, worksheet, export_df, t_shape_data=None):
        """智能调整列宽（用于混合数据：统计信息+T型报价）"""
        # 获取worksheet的最大列数
        max_col = worksheet.max_column
        
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_width = 0
            
            # 遍历该列的所有单元格，计算最大宽度（考虑中文）
            for row_idx in range(1, min(worksheet.max_row + 1, 500)):  # 限制最多检查500行
                cell = worksheet[f'{col_letter}{row_idx}']
                if cell.value is not None:
                    cell_width = self._calculate_display_width(cell.value)
                    max_width = max(max_width, cell_width)
            
            # 如果有T型报价数据，特别处理这部分的列宽
            if t_shape_data is not None and col_idx <= len(t_shape_data.columns):
                # T型报价的列
                col_name = t_shape_data.columns[col_idx - 1] if col_idx <= len(t_shape_data.columns) else None
                if col_name:
                    header_width = self._calculate_display_width(col_name)
                    data_width = t_shape_data.iloc[:, col_idx - 1].apply(lambda x: self._calculate_display_width(x)).max()
                    max_width = max(max_width, header_width, data_width)
            
            # 设置列宽：最小6，最大40（混合数据可能需要更宽），加1.5个字符的边距
            column_width = min(max(max_width * 1.1 + 1.5, 6), 40)
            worksheet.column_dimensions[col_letter].width = column_width
    
    def export_important_contracts(self):
        """导出期权沉淀≥0.20亿的重要合约，每个合约一个分页包含完整信息"""
        try:
            self.status_label.setText("状态：正在导出重要期权...")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            QApplication.processEvents()
            
            if self.market_overview_df is None or self.option_ref_df is None:
                QMessageBox.warning(self, "警告", "请先加载数据！")
                return
            
            # 1. 筛选沉淀≥0.20亿的合约
            if '期权沉淀(亿)' not in self.market_overview_df.columns:
                QMessageBox.warning(self, "警告", "数据中缺少 '期权沉淀(亿)' 字段！")
                return
            
            important_df = self.market_overview_df[
                self.market_overview_df['期权沉淀(亿)'] >= 0.20
            ].copy()
            
            if important_df.empty:
                QMessageBox.information(self, "提示", "没有找到沉淀≥0.20亿的合约！")
                return
            
            # 按沉淀从高到低排序
            important_df = important_df.sort_values('期权沉淀(亿)', ascending=False)
            
            # 2. 创建Excel写入器
            output_file = 'wisecoin-重要期权.xlsx'
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                
                sheet_count = 0
                # 为每个重要合约的每个交割月份创建一个分页
                for _, market_row in important_df.iterrows():
                    underlying = market_row['标的合约']
                    product_code = self._extract_product_code(underlying)
                    
                    # 获取该标的的所有交割月份
                    contract_expiries = self.option_ref_df[
                        self.option_ref_df['标的合约'].astype(str) == str(underlying)
                    ]['交割年月'].unique() if '交割年月' in self.option_ref_df.columns else [None]
                    
                    # 为每个交割月份创建一个分页
                    for expiry in contract_expiries:
                        # 生成sheet名称：合约代码_交割年月
                        contract_part = underlying.split('.')[-1] if '.' in str(underlying) else str(underlying)
                        if expiry and pd.notna(expiry):
                            sheet_name = f"{contract_part}_{expiry}"[:31]
                        else:
                            sheet_name = contract_part[:31]
                        
                        # 准备该合约+交割月的完整数据
                        export_data = []
                        
                        # === 第一部分：标的统计信息 ===
                        export_data.append(['=== 标的统计信息 ==='])
                        export_data.append(['字段', '数值'])
                        export_data.append(['标的合约', self._safe_get_value(market_row, '标的合约')])
                        export_data.append(['交割年月', expiry if expiry and pd.notna(expiry) else '-'])
                        export_data.append(['期货现价', self._safe_get_value(market_row, '期货现价')])
                        export_data.append(['杠杆涨跌%', self._safe_get_value(market_row, '杠杆涨跌%')])
                        export_data.append(['期货沉淀(亿)', self._safe_get_value(market_row, '期货沉淀(亿)')])
                        export_data.append(['期货状态', self._safe_get_value(market_row, '期货状态')])
                        export_data.append(['期货方向', self._safe_get_value(market_row, '期货方向')])
                        export_data.append(['期货流向', self._safe_get_value(market_row, '期货流向')])
                        export_data.append([])  # 空行
                        
                        # === 第二部分：期权整体统计信息 ===
                        export_data.append(['=== 期权整体统计信息 ==='])
                        export_data.append(['字段', '数值'])
                        export_data.append(['期权结构', self._safe_get_value(market_row, '期权结构')])
                        export_data.append(['期权情绪', self._safe_get_value(market_row, '期权情绪')])
                        export_data.append(['期权PCR', self._safe_get_value(market_row, '期权PCR')])
                        export_data.append(['期权沉淀(亿)', self._safe_get_value(market_row, '期权沉淀(亿)')])
                        export_data.append(['最大痛点', self._safe_get_value(market_row, '最大痛点')])
                        export_data.append(['痛点距离%', self._safe_get_value(market_row, '痛点距离%')])
                        export_data.append(['联动状态', self._safe_get_value(market_row, '联动状态')])
                        export_data.append(['共振评分', self._safe_get_value(market_row, '共振评分')])
                        export_data.append([])  # 空行
                        
                        # === 第三部分：货权联动分析（含波动率曲面） ===
                        export_data.append(['=== 货权联动分析（含波动率曲面） ==='])
                        export_data.append(['字段', '数值'])
                        
                        # 获取波动率曲面信息
                        vol_row = None
                        if self.vol_surface_df is not None and product_code:
                            if '品种代码' in self.vol_surface_df.columns:
                                vol_match = self.vol_surface_df[
                                    self.vol_surface_df['品种代码'].astype(str).str.strip().str.upper() == product_code
                                ]
                                if not vol_match.empty:
                                    vol_row = vol_match.iloc[0]
                        
                        # 如果Excel中没有，尝试实时计算
                        if vol_row is None and product_code:
                            try:
                                vol_row = self._calculate_vol_surface_metrics(product_code)
                            except:
                                vol_row = None
                        
                        # 联动分析字段
                        linkage_fields = [
                            '共振等级', '期限结构', '倾斜方向', '共振标签',
                            '联动总分', 'IV/RV比率', '期限结构差', 'IV倾斜度',
                            '价格评分', '峰度', '短期IV', '虚值认沽IV均值', '沉淀资金合计(亿)',
                            '情绪评分', '偏度', '长期IV', '虚值认购IV均值', '合约数量',
                            '市场解读', '推荐策略', '适合策略', '不适合策略'
                        ]
                        
                        for field in linkage_fields:
                            # 优先从 market_row 获取，然后从 vol_row 获取
                            val = self._safe_get_value(market_row, field, None)
                            if val == '-' or val is None:
                                val = self._safe_get_value(vol_row, field, '-') if vol_row is not None else '-'
                            export_data.append([field, val])
                        
                        export_data.append([])  # 空行
                        
                        # === 第四部分：期权T型报价数据（仅该交割月） ===
                        export_data.append(['=== 期权T型报价数据 ==='])
                        export_data.append([])  # 空行
                        
                        # 获取该标的+该交割月的期权数据
                        if expiry and pd.notna(expiry):
                            contract_options = self.option_ref_df[
                                (self.option_ref_df['标的合约'].astype(str) == str(underlying)) &
                                (self.option_ref_df['交割年月'].astype(str) == str(expiry))
                            ].copy()
                        else:
                            contract_options = self.option_ref_df[
                                self.option_ref_df['标的合约'].astype(str) == str(underlying)
                            ].copy()
                        
                        if not contract_options.empty:
                            # 构建T型报价格式（与界面一致）
                            t_shape_data = self._get_t_shape_data_for_export(underlying, expiry)
                            
                            if t_shape_data is not None and not t_shape_data.empty:
                                # 添加表头
                                export_data.append(t_shape_data.columns.tolist())
                                
                                # 添加数据行
                                for _, row in t_shape_data.iterrows():
                                    export_data.append(row.tolist())
                            else:
                                export_data.append(['无T型报价数据'])
                        else:
                            export_data.append(['无期权数据'])
                        
                        export_data.append([])  # 空行
                        
                        # 记录图片插入位置（在T型报价数据之后）
                        chart_insert_row = len(export_data) + 1
                        
                        # 写入该合约+交割月的完整数据到一个Sheet
                        export_df = pd.DataFrame(export_data)
                        export_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                        
                        # 智能调整列宽
                        worksheet = writer.sheets[sheet_name]
                        self._adjust_column_widths_for_mixed_data(worksheet, export_df, t_shape_data if 't_shape_data' in locals() else None)
                        
                        # === 生成并插入图片 ===
                        
                        # 生成波动率曲面图（整个品种）
                        if product_code:
                            mask = self.option_ref_df['标的合约'].apply(self._extract_product_code) == product_code
                            vol_surface_data = self.option_ref_df[mask].copy()
                            
                            surface_img_buf = self._generate_vol_surface_image(vol_surface_data, product_code)
                            if surface_img_buf:
                                img = OpenpyxlImage(surface_img_buf)
                                img.width = 600
                                img.height = 450
                                worksheet.add_image(img, f'A{chart_insert_row}')
                        
                        # 生成微笑曲线图（仅该交割月）
                        if not contract_options.empty:
                            smile_img_buf = self._generate_vol_smile_image(contract_options, expiry if expiry else underlying)
                            if smile_img_buf:
                                img = OpenpyxlImage(smile_img_buf)
                                img.width = 600
                                img.height = 375
                                worksheet.add_image(img, f'J{chart_insert_row}')
                        
                        sheet_count += 1
            
            self.status_label.setText(f"状态：已导出 {sheet_count} 个重要合约到 {output_file}")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
            
            QMessageBox.information(
                self, 
                "导出成功", 
                f"已成功导出 {sheet_count} 个重要期权合约到:\n{output_file}\n\n"
                f"每个合约一个分页，包含:\n"
                f"• 标的统计信息\n"
                f"• 期权整体统计信息\n"
                f"• 货权联动分析（含波动率曲面）\n"
                f"• 期权T型报价数据\n"
                f"• 波动率曲面图（PNG）\n"
                f"• 微笑曲线图（PNG，按交割月）"
            )
            
        except Exception as e:
            error_msg = f"导出失败: {str(e)}"
            self.status_label.setText(f"状态：{error_msg}")
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
            QMessageBox.critical(self, "错误", error_msg)
            import traceback
            traceback.print_exc()

    def show_position_analysis(self):
        """显示专业持仓分析弹窗"""
        if self.positions_view_df is None or self.positions_view_df.empty:
            QMessageBox.warning(self, "提示", "无持仓数据可分析！")
            return
        
        try:
            # 生成分析数据
            analysis = self._analyze_positions()
            
            # 创建分析弹窗 - 浅色风格
            dialog = QDialog(self)
            dialog.setWindowTitle("WiseCoin 期权持仓分析")
            dialog.setGeometry(80, 50, 1600, 950)
            dialog.setStyleSheet("""
                QDialog { background-color: #f5f7fa; }
                QLabel { color: #2c3e50; }
                QGroupBox { 
                    color: #2c3e50; 
                    font-weight: bold; 
                    border: 1px solid #dce1e8;
                    border-radius: 8px;
                    margin-top: 10px;
                    padding-top: 14px;
                    background-color: #ffffff;
                }
                QGroupBox::title { 
                    subcontrol-origin: margin; 
                    left: 12px; 
                    padding: 0 6px;
                    color: #1a5fb4;
                }
            """)
            
            main_layout = QVBoxLayout(dialog)
            main_layout.setSpacing(10)
            main_layout.setContentsMargins(12, 12, 12, 12)
            
            # ========== 主内容区 (无标题栏) ==========
            content_splitter = QSplitter(Qt.Horizontal)
            
            # ---------- 左侧面板：文字分析 ----------
            left_scroll = QScrollArea()
            left_scroll.setWidgetResizable(True)
            left_scroll.setStyleSheet("""
                QScrollArea { border: none; background-color: transparent; }
                QScrollBar:vertical { background: #e8ecf0; width: 10px; border-radius: 5px; }
                QScrollBar::handle:vertical { background: #1a5fb4; border-radius: 5px; min-height: 30px; }
            """)
            
            left_widget = QWidget()
            left_layout = QVBoxLayout(left_widget)
            left_layout.setSpacing(10)
            
            # 1. 策略识别卡片
            self._add_strategy_card(left_layout, analysis)
            
            # 2. Greeks敞口卡片
            self._add_greeks_card(left_layout, analysis)
            
            # 3. 盈亏平衡点卡片
            self._add_breakeven_card(left_layout, analysis)
            
            # 4. 走势匹配卡片
            self._add_trend_match_card(left_layout, analysis)
            
            # 5. 风险警示卡片
            self._add_risk_warning_card(left_layout, analysis)
            
            # 6. 风控建议卡片
            self._add_risk_control_card(left_layout, analysis)
            
            left_layout.addStretch()
            left_scroll.setWidget(left_widget)
            
            # ---------- 右侧面板：图表 ----------
            right_widget = QWidget()
            right_widget.setStyleSheet("background-color: #f5f7fa;")
            right_layout = QVBoxLayout(right_widget)
            right_layout.setSpacing(10)
            
            # 1. 损益图
            pnl_group = QGroupBox("到期损益图 / Payoff Diagram")
            pnl_layout = QVBoxLayout(pnl_group)
            pnl_fig = Figure(figsize=(8, 3.8), dpi=100, facecolor='#ffffff')
            pnl_canvas = FigureCanvas(pnl_fig)
            self._plot_pnl_diagram_light(pnl_fig, analysis)
            pnl_layout.addWidget(pnl_canvas)
            right_layout.addWidget(pnl_group)
            
            # 2. 盈亏概率与收益分析
            prob_group = QGroupBox("收益概率与风险指标 / Risk Metrics")
            prob_layout = QVBoxLayout(prob_group)
            prob_fig = Figure(figsize=(8, 3.2), dpi=100, facecolor='#ffffff')
            prob_canvas = FigureCanvas(prob_fig)
            self._plot_risk_metrics(prob_fig, analysis)
            prob_layout.addWidget(prob_canvas)
            right_layout.addWidget(prob_group)
            
            # 3. 情景敏感度分析
            sensitivity_group = QGroupBox("情景敏感度分析 / Scenario Analysis")
            sensitivity_layout = QVBoxLayout(sensitivity_group)
            sens_fig = Figure(figsize=(8, 2.8), dpi=100, facecolor='#ffffff')
            sens_canvas = FigureCanvas(sens_fig)
            self._plot_sensitivity_matrix_light(sens_fig, analysis)
            sensitivity_layout.addWidget(sens_canvas)
            right_layout.addWidget(sensitivity_group)
            
            content_splitter.addWidget(left_scroll)
            content_splitter.addWidget(right_widget)
            content_splitter.setStretchFactor(0, 42)
            content_splitter.setStretchFactor(1, 58)
            
            main_layout.addWidget(content_splitter)
            
            # 说明：去除右下角“关闭”按钮（保留窗口右上角系统关闭/快捷键即可）
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "分析错误", f"持仓分析失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def _create_info_card(self, title, content_widgets, accent_color="#1a5fb4"):
        """创建浅色风格的信息卡片"""
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background-color: #ffffff;
                border-radius: 8px;
                border-left: 4px solid {accent_color};
                border: 1px solid #e0e5eb;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(8)
        
        # 标题
        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 12pt; font-weight: bold; color: {accent_color}; border: none; background: transparent;")
        layout.addWidget(title_label)
        
        # 分隔线
        line = QWidget()
        line.setFixedHeight(1)
        line.setStyleSheet(f"background-color: #e0e5eb;")
        layout.addWidget(line)
        
        # 内容
        for widget in content_widgets:
            layout.addWidget(widget)
        
        return card

    def _add_strategy_card(self, parent_layout, analysis):
        """添加策略识别卡片（专业版）"""
        content_widgets = []
        
        if analysis['strategies']:
            # 按策略类型分类显示
            combo_strategies = [s for s in analysis['strategies'] if 'spread' in s['type'] or 
                              'straddle' in s['type'] or 'strangle' in s['type'] or
                              'condor' in s['type'] or 'butterfly' in s['type']]
            single_strategies = [s for s in analysis['strategies'] if s not in combo_strategies]
            
            # 优先显示组合策略
            all_strategies = combo_strategies + single_strategies
            
            for i, s in enumerate(all_strategies):
                # 策略卡片容器 - 浅色风格
                strategy_widget = QWidget()
                strategy_widget.setStyleSheet("""
                    border: none; 
                    background-color: #f8fafc; 
                    border-radius: 6px; 
                    margin: 4px 0;
                    border: 1px solid #e2e8f0;
                """)
                s_layout = QVBoxLayout(strategy_widget)
                s_layout.setContentsMargins(10, 8, 10, 8)
                s_layout.setSpacing(4)
                
                # 标题行
                title_widget = QWidget()
                title_widget.setStyleSheet("border: none; background: transparent;")
                title_layout = QHBoxLayout(title_widget)
                title_layout.setContentsMargins(0, 0, 0, 0)
                title_layout.setSpacing(8)
                
                # 策略图标
                icon_map = {
                    'straddle': '⚡', 'strangle': '🔀', 'spread': '📊',
                    'butterfly': '🦋', 'condor': '🦅', 'ratio': '⚖️',
                    'long_call': '📈', 'short_call': '📉', 
                    'long_put': '📉', 'short_put': '📈'
                }
                icon = '📌'
                for key, ico in icon_map.items():
                    if key in s['type']:
                        icon = ico
                        break
                
                name_label = QLabel(f"{icon} {s['name']}")
                name_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: #1e40af;")
                title_layout.addWidget(name_label)
                title_layout.addStretch()
                
                # 策略类型标签
                type_colors = {
                    'long': '#27ae60', 'short': '#e74c3c', 
                    'bull': '#27ae60', 'bear': '#e74c3c',
                    'neutral': '#3498db'
                }
                type_color = '#9b59b6'
                for key, color in type_colors.items():
                    if key in s['type']:
                        type_color = color
                        break
                
                type_label = QLabel(s['type'])
                type_label.setStyleSheet(f"""
                    font-size: 9pt; 
                    color: white; 
                    background-color: {type_color}; 
                    padding: 2px 8px; 
                    border-radius: 10px;
                """)
                title_layout.addWidget(type_label)
                
                s_layout.addWidget(title_widget)
                
                # 描述
                desc_label = QLabel(s['description'])
                desc_label.setStyleSheet("font-size: 10pt; color: #475569;")
                desc_label.setWordWrap(True)
                s_layout.addWidget(desc_label)
                
                # Greeks特征（如果有）
                if 'greeks_profile' in s:
                    greeks_label = QLabel(f"Greeks: {s['greeks_profile']}")
                    greeks_label.setStyleSheet("font-size: 9pt; color: #64748b; font-style: italic;")
                    greeks_label.setWordWrap(True)
                    s_layout.addWidget(greeks_label)
                
                # 理想场景（如果有）
                if 'ideal_scenario' in s:
                    scenario_label = QLabel(f"📍 理想场景: {s['ideal_scenario']}")
                    scenario_label.setStyleSheet("font-size: 9pt; color: #059669;")
                    scenario_label.setWordWrap(True)
                    s_layout.addWidget(scenario_label)
                
                content_widgets.append(strategy_widget)
            
            # 策略计数
            count_label = QLabel(f"共识别 {len(all_strategies)} 个策略/持仓")
            count_label.setStyleSheet("font-size: 9pt; color: #94a3b8; margin-top: 4px;")
            content_widgets.append(count_label)
        else:
            no_strategy = QLabel("⚠️ 未识别出明确策略组合\n建议检查持仓是否形成完整策略结构")
            no_strategy.setStyleSheet("font-size: 10pt; color: #d97706;")
            no_strategy.setWordWrap(True)
            content_widgets.append(no_strategy)
        
        card = self._create_info_card("策略识别 / Strategy Recognition", content_widgets, "#e65100")
        parent_layout.addWidget(card)

    def _add_greeks_card(self, parent_layout, analysis):
        """添加Greeks敞口卡片（专业版）"""
        content_widgets = []
        greeks = analysis['greeks']
        
        # 创建专业表格
        grid_widget = QWidget()
        grid_widget.setStyleSheet("border: none; background: transparent;")
        grid = QGridLayout(grid_widget)
        grid.setSpacing(6)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setColumnStretch(3, 1)  # 让解读列自适应
        
        # 表头 - 增强可见性 (黑色文字，确保对比度)
        headers = ['指标', '数值', '强度', '风险解读']
        for i, h in enumerate(headers):
            lbl = QLabel(h)
            lbl.setStyleSheet("font-size: 10pt; color: #000000; font-weight: bold; background-color: #f1f5f9; padding: 4px; border-radius: 4px;")
            lbl.setMinimumHeight(24)
            lbl.setAlignment(Qt.AlignCenter if i == 1 else Qt.AlignLeft | Qt.AlignVCenter)
            grid.addWidget(lbl, 0, i)
        
        greek_info = [
            ('Δ Delta', greeks['Delta'], 500, '方向风险', '正值→标的涨盈利，负值→标的跌盈利'),
            ('Γ Gamma', greeks['Gamma'], 100, 'Delta加速度', '正值→波动受益，负值→波动受损'),
            ('Θ Theta', greeks['Theta'], 100, '时间损耗/日', '正值→时间流逝受益，负值→每日损耗'),
            ('ν Vega', greeks['Vega'], 200, 'IV敏感度', '正值→IV涨受益，负值→IV跌受益'),
            ('ρ Rho', greeks['Rho'], 50, '利率敏感度', '正值→利率涨受益，负值→利率跌受益'),
        ]
        
        for i, (name, value, max_val, risk_type, interpret) in enumerate(greek_info, 1):
            # 处理 NaN 值
            if pd.isna(value): value = 0.0
            
            # 名称
            name_lbl = QLabel(name)
            name_lbl.setStyleSheet("font-size: 11pt; color: #1e293b; font-weight: bold;")
            grid.addWidget(name_lbl, i, 0)
            
            # 数值 (带颜色)
            val_color = '#16a34a' if value > 0 else '#dc2626' if value < 0 else '#94a3b8'
            val_lbl = QLabel(f"{value:+.1f}")
            val_lbl.setStyleSheet(f"font-size: 12pt; font-weight: bold; color: {val_color};")
            grid.addWidget(val_lbl, i, 1)
            
            # 强度条 (可视化)
            strength_pct = min(abs(value) / max_val * 100, 100) if max_val > 0 else 0
            if pd.isna(strength_pct): strength_pct = 0
            
            bar_color = '#16a34a' if strength_pct < 50 else '#d97706' if strength_pct < 80 else '#dc2626'
            strength_widget = QWidget()
            strength_widget.setFixedSize(80, 16)
            strength_widget.setStyleSheet(f"""
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
            """)
            
            # 内部填充条
            bar_inner = QWidget(strength_widget)
            bar_width = int(strength_pct * 0.78)  # 78px最大宽度
            bar_inner.setGeometry(1, 1, bar_width, 14)
            bar_inner.setStyleSheet(f"background-color: {bar_color}; border-radius: 3px;")
            
            grid.addWidget(strength_widget, i, 2)
            
            # 风险解读
            interpret_lbl = QLabel(risk_type)
            interpret_lbl.setStyleSheet("font-size: 9pt; color: #64748b;")
            interpret_lbl.setToolTip(interpret)  # 悬停显示详细解读
            grid.addWidget(interpret_lbl, i, 3)
        
        content_widgets.append(grid_widget)
        
        # 综合风险评估
        delta = greeks.get('Delta', 0)
        gamma = greeks.get('Gamma', 0)
        theta = greeks.get('Theta', 0)
        vega = greeks.get('Vega', 0)
        
        # 处理可能的 NaN
        delta = 0 if pd.isna(delta) else delta
        gamma = 0 if pd.isna(gamma) else gamma
        theta = 0 if pd.isna(theta) else theta
        vega = 0 if pd.isna(vega) else vega
        
        # 组合方向判断 - 浅色风格颜色
        if abs(delta) > 200:
            direction = "📈 强看涨" if delta > 0 else "📉 强看跌"
            dir_color = "#16a34a" if delta > 0 else "#dc2626"
        elif abs(delta) > 50:
            direction = "↗️ 温和看涨" if delta > 0 else "↘️ 温和看跌"
            dir_color = "#16a34a" if delta > 0 else "#dc2626"
        else:
            direction = "↔️ 方向中性"
            dir_color = "#1a5fb4"
        
        # 波动率偏好
        if vega > 100:
            vol_pref = "做多波动率"
        elif vega < -100:
            vol_pref = "做空波动率"
        else:
            vol_pref = "波动率中性"
        
        # 时间偏好
        if theta > 0:
            time_pref = "时间友好(正Theta)"
        elif theta < -50:
            time_pref = "时间敌对(日损耗)"
        else:
            time_pref = "时间中性"
        
        summary_widget = QWidget()
        summary_widget.setStyleSheet("border: none; background-color: #f8fafc; border-radius: 6px; margin-top: 8px; border: 1px solid #e2e8f0;")
        summary_layout = QVBoxLayout(summary_widget)
        summary_layout.setContentsMargins(10, 8, 10, 8)
        summary_layout.setSpacing(4)
        
        details_lbl = QLabel(f"特征: {vol_pref} | {time_pref} | Delta暴露{abs(delta):.0f}")
        details_lbl.setStyleSheet("font-size: 9pt; color: #64748b;")
        summary_layout.addWidget(details_lbl)
        
        content_widgets.append(summary_widget)
        
        card = self._create_info_card("希腊字母敞口 / Greeks Exposure", content_widgets, "#1a5fb4")
        parent_layout.addWidget(card)

    def _add_breakeven_card(self, parent_layout, analysis):
        """添加盈亏平衡点卡片（专业版）"""
        content_widgets = []
        
        if analysis['breakeven_points']:
            # 按距离排序
            sorted_bps = sorted(analysis['breakeven_points'], 
                              key=lambda x: abs(x.get('distance_pct', 0)))
            
            # 表头
            header_widget = QWidget()
            header_widget.setStyleSheet("border: none;")
            header_layout = QGridLayout(header_widget)
            header_layout.setContentsMargins(0, 0, 0, 4)
            header_layout.setSpacing(10)
            
            headers = ['合约', 'BE价格', '当前价', '距离', '方向']
            for i, h in enumerate(headers):
                lbl = QLabel(h)
                lbl.setStyleSheet("font-size: 9pt; color: #64748b; font-weight: bold;")
                header_layout.addWidget(lbl, 0, i)
            
            content_widgets.append(header_widget)
            
            # 分隔线
            line = QWidget()
            line.setFixedHeight(1)
            line.setStyleSheet("background-color: #e2e8f0;")
            content_widgets.append(line)
            
            # 数据行
            for bp in sorted_bps[:6]:  # 最多显示6个
                row_widget = QWidget()
                row_widget.setStyleSheet("border: none;")
                row_layout = QGridLayout(row_widget)
                row_layout.setContentsMargins(0, 2, 0, 2)
                row_layout.setSpacing(10)
                
                # 合约（简化显示）
                contract = bp.get('contract', '')
                short_contract = contract.split('.')[-1] if '.' in contract else contract
                contract_lbl = QLabel(short_contract[-15:])  # 最多15字符
                contract_lbl.setStyleSheet("font-size: 9pt; color: #64748b;")
                row_layout.addWidget(contract_lbl, 0, 0)
                
                # BE价格
                be_price = bp.get('breakeven', 0)
                be_lbl = QLabel(f"{be_price:.2f}")
                be_lbl.setStyleSheet("font-size: 11pt; font-weight: bold; color: #7c3aed;")
                row_layout.addWidget(be_lbl, 0, 1)
                
                # 当前价
                current = bp.get('current_price', 0)
                current_lbl = QLabel(f"{current:.2f}")
                current_lbl.setStyleSheet("font-size: 10pt; color: #475569;")
                row_layout.addWidget(current_lbl, 0, 2)
                
                # 距离百分比
                dist_pct = bp.get('distance_pct', 0)
                dist_color = '#16a34a' if abs(dist_pct) < 5 else '#d97706' if abs(dist_pct) < 10 else '#dc2626'
                dist_lbl = QLabel(f"{dist_pct:+.2f}%")
                dist_lbl.setStyleSheet(f"font-size: 10pt; font-weight: bold; color: {dist_color};")
                row_layout.addWidget(dist_lbl, 0, 3)
                
                # 方向指示
                if dist_pct > 0:
                    direction = "↑需涨"
                    dir_color = "#16a34a"
                else:
                    direction = "↓需跌"
                    dir_color = "#dc2626"
                dir_lbl = QLabel(direction)
                dir_lbl.setStyleSheet(f"font-size: 9pt; color: {dir_color};")
                row_layout.addWidget(dir_lbl, 0, 4)
                
                content_widgets.append(row_widget)
            
            # 总结
            if sorted_bps:
                nearest = sorted_bps[0]
                nearest_dist = nearest.get('distance_pct', 0)
                summary_text = f"📍 最近BE点距当前仅 {abs(nearest_dist):.2f}%"
                summary_color = '#16a34a' if abs(nearest_dist) < 3 else '#d97706'
                summary = QLabel(summary_text)
                summary.setStyleSheet(f"font-size: 10pt; color: {summary_color}; margin-top: 6px;")
                content_widgets.append(summary)
        else:
            no_bp = QLabel("⚠️ 数据不足，无法精确计算盈亏平衡点\n请确保期权持仓包含完整的价格信息")
            no_bp.setStyleSheet("font-size: 10pt; color: #64748b;")
            no_bp.setWordWrap(True)
            content_widgets.append(no_bp)
        
        card = self._create_info_card("盈亏平衡点 / Break-Even", content_widgets, "#7c3aed")
        parent_layout.addWidget(card)

    def _add_trend_match_card(self, parent_layout, analysis):
        """添加走势匹配卡片（专业简洁版）"""
        content_widgets = []
        
        trend_text = analysis['trend_match']
        lines = trend_text.split('\n')
        
        # 从分析结果提取方向
        greeks = analysis.get('greeks', {})
        delta = greeks.get('Delta', 0)
        if pd.isna(delta): delta = 0
        
        if delta > 50:
            pos_dir_text = "看涨 ↑"
            pos_color = "#16a34a"
        elif delta < -50:
            pos_dir_text = "看跌 ↓"
            pos_color = "#dc2626"
        else:
            pos_dir_text = "中性 ↔"
            pos_color = "#1a5fb4"
        
        # 从趋势文本中提取市场方向
        market_dir_text = "未知"
        market_color = "#64748b"
        for line in lines:
            if '期货方向' in line:
                if '多' in line or '涨' in line:
                    market_dir_text = "偏多 ↑"
                    market_color = "#16a34a"
                elif '空' in line or '跌' in line:
                    market_dir_text = "偏空 ↓"
                    market_color = "#dc2626"
                else:
                    market_dir_text = "震荡 ↔"
                    market_color = "#d97706"
                break
        
        # 简洁的方向对比行
        direction_lbl = QLabel(f"持仓: <span style='color:{pos_color}; font-weight:bold;'>{pos_dir_text}</span> &nbsp;vs&nbsp; 市场: <span style='color:{market_color}; font-weight:bold;'>{market_dir_text}</span>")
        direction_lbl.setStyleSheet("font-size: 11pt; color: #334155;")
        content_widgets.append(direction_lbl)
        
        # 匹配度结论 - 只提取走势匹配行，避免重复
        match_found = False
        for line in lines:
            # 只匹配包含"走势匹配"的行，而非所有包含"匹配"的行
            if '走势匹配' in line:
                if '✅' in line:
                    text_color = "#16a34a"
                    status = "✅ 匹配"
                elif '❌' in line:
                    text_color = "#dc2626"
                    status = "❌ 不匹配"
                else:
                    text_color = "#d97706"
                    status = "⚠️ 谨慎"
                
                # 提取冒号后的说明文字
                match_text = line.split('走势匹配:')[-1].strip() if '走势匹配:' in line else ''
                match_text = match_text.replace('✅', '').replace('❌', '').replace('⚠️', '').strip()
                
                if match_text:
                    result_lbl = QLabel(f"<span style='font-size:12pt; font-weight:bold;'>{status}</span>")
                    result_lbl.setStyleSheet(f"font-size: 12pt; color: {text_color}; margin-top: 6px;")
                    content_widgets.append(result_lbl)
                    
                    detail_lbl = QLabel(match_text)
                    detail_lbl.setStyleSheet("font-size: 9pt; color: #64748b; margin-top: 2px;")
                    detail_lbl.setWordWrap(True)
                    content_widgets.append(detail_lbl)
                else:
                    result_lbl = QLabel(f"<span style='font-size:12pt; font-weight:bold;'>{status}</span>")
                    result_lbl.setStyleSheet(f"font-size: 12pt; color: {text_color}; margin-top: 6px;")
                    content_widgets.append(result_lbl)
                
                match_found = True
                break
        
        # 其他市场信息
        other_info = []
        for line in lines:
            if '期权情绪' in line or '期货状态' in line:
                clean_line = line.replace('📊', '').replace('📈', '').replace('📉', '').replace('😀', '').strip()
                other_info.append(clean_line)
        
        if other_info:
            info_lbl = QLabel(' | '.join(other_info))
            info_lbl.setStyleSheet("font-size: 9pt; color: #94a3b8; margin-top: 4px;")
            info_lbl.setWordWrap(True)
            content_widgets.append(info_lbl)
        
        card = self._create_info_card("走势匹配度 / Trend Alignment", content_widgets, "#0891b2")
        parent_layout.addWidget(card)

    def _add_risk_warning_card(self, parent_layout, analysis):
        """添加风险警示卡片（简洁版）"""
        content_widgets = []
        
        warnings = analysis['risk_warnings']
        
        # 检查是否有整体评估
        overall_assessment = None
        other_warnings = []
        for w in warnings:
            if '整体风险评估' in w:
                overall_assessment = w
            else:
                other_warnings.append(w)
        
        # 整体风险评估 - 简洁文字显示
        if overall_assessment:
            if '低风险' in overall_assessment:
                text_color = "#16a34a"
                icon = "🟢"
            elif '中等风险' in overall_assessment:
                text_color = "#d97706"
                icon = "🟡"
            else:
                text_color = "#dc2626"
                icon = "🔴"
            
            assessment_text = overall_assessment.split(':')[-1].strip() if ':' in overall_assessment else overall_assessment
            assessment_text = assessment_text.replace('✅', '').replace('⚠️', '').replace('🚨', '').strip()
            
            overall_lbl = QLabel(f"{icon} <b style='color:{text_color};'>{assessment_text}</b>")
            overall_lbl.setStyleSheet("font-size: 12pt; color: #334155;")
            content_widgets.append(overall_lbl)
        
        # 其他警示项 - 简洁列表
        for warning in other_warnings[:5]:  # 最多显示5条
            clean_text = warning.replace('⚠️', '').replace('✅', '').strip()
            
            if '✅' in warning:
                color = "#16a34a"
                bullet = "✓"
            else:
                color = "#dc2626"
                bullet = "•"
            
            w_lbl = QLabel(f"<span style='color:{color};'>{bullet}</span> {clean_text}")
            w_lbl.setStyleSheet(f"font-size: 10pt; color: #475569; margin-left: 4px;")
            w_lbl.setWordWrap(True)
            content_widgets.append(w_lbl)
        
        card = self._create_info_card("风险警示 / Risk Alerts", content_widgets, "#dc2626")
        parent_layout.addWidget(card)

    def _add_risk_control_card(self, parent_layout, analysis):
        """添加风控建议卡片（简洁版）"""
        content_widgets = []
        
        if analysis['risk_controls']:
            for i, ctrl in enumerate(analysis['risk_controls'][:3], 1):  # 最多显示3条
                # 建议标题
                action_lbl = QLabel(f"<b style='color:#059669;'>{i}. {ctrl['action']}</b>")
                action_lbl.setStyleSheet("font-size: 11pt; color: #334155;")
                content_widgets.append(action_lbl)
                
                # 原因
                reason_lbl = QLabel(f"   💡 {ctrl['reason']}")
                reason_lbl.setStyleSheet("font-size: 9pt; color: #64748b;")
                reason_lbl.setWordWrap(True)
                content_widgets.append(reason_lbl)
                
                # 具体操作
                specific_lines = ctrl['specific'].split('\n')
                for line in specific_lines[:2]:  # 每条建议最多2行操作
                    if line.strip():
                        spec_lbl = QLabel(f"   <span style='color:#dc2626;'>➤</span> {line.strip()}")
                        spec_lbl.setStyleSheet("font-size: 10pt; color: #475569;")
                        spec_lbl.setWordWrap(True)
                        content_widgets.append(spec_lbl)
        else:
            ok_lbl = QLabel("✅ 当前持仓风险可控，无需额外风控操作")
            ok_lbl.setStyleSheet("font-size: 11pt; color: #16a34a; font-weight: bold;")
            content_widgets.append(ok_lbl)
        
        card = self._create_info_card("风控建议 / Risk Management", content_widgets, "#059669")
        parent_layout.addWidget(card)

    def _analyze_positions(self):
        """分析持仓数据，返回分析结果"""
        result = {
            'strategies': [],
            'greeks': {'Delta': 0, 'Gamma': 0, 'Theta': 0, 'Vega': 0, 'Rho': 0},
            'positions': [],
            'breakeven_points': [],
            'risk_warnings': [],
            'risk_controls': [],
            'trend_match': ''
        }
        
        df = self.positions_view_df.copy()
        
        # 筛选有效期权持仓（需要有 Delta 等希腊字母）
        option_positions = []
        futures_positions = []
        
        for _, row in df.iterrows():
            pos_info = self._extract_position_info(row)
            if pos_info:
                if pos_info['is_option']:
                    option_positions.append(pos_info)
                else:
                    futures_positions.append(pos_info)
        
        result['positions'] = option_positions + futures_positions
        
        # 计算组合希腊字母
        result['greeks'] = self._calculate_portfolio_greeks(option_positions, futures_positions)
        
        # 识别策略
        result['strategies'] = self._identify_strategies(option_positions, futures_positions)
        
        # 计算盈亏平衡点
        result['breakeven_points'] = self._calculate_breakeven(option_positions)
        
        # 风险分析
        result['risk_warnings'], result['risk_controls'] = self._analyze_risk(
            option_positions, futures_positions, result['greeks']
        )
        
        # 走势匹配分析
        result['trend_match'] = self._analyze_trend_match(option_positions, result['strategies'])
        
        return result

    def _extract_position_info(self, row):
        """从持仓行提取关键信息"""
        try:
            # 判断合约代码列
            contract_code = None
            for col in ['instrument_id', 'symbol', '合约代码_期权参考']:
                if col in row.index and pd.notna(row[col]):
                    contract_code = str(row[col])
                    break
            
            if not contract_code:
                return None
            
            # 判断是否为期权
            is_option = bool(re.search(r'[CP]-?\d+|[CP]\d+', contract_code, re.IGNORECASE))
            
            # 提取持仓方向和数量
            pos_long = row.get('pos_long', 0) or 0
            pos_short = row.get('pos_short', 0) or 0
            net_pos = pos_long - pos_short
            
            if net_pos == 0:
                return None

            # 统一的数值/字段读取工具（支持 _期权参考 后缀）
            def clean_num(v, default=0):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return default
                try:
                    return float(v)
                except:
                    return default

            def get_field(row, *field_names):
                for fname in field_names:
                    if fname in row.index and pd.notna(row[fname]):
                        return row[fname]
                    fname_ref = f"{fname}_期权参考"
                    if fname_ref in row.index and pd.notna(row[fname_ref]):
                        return row[fname_ref]
                return None

            # 关键：推断合约乘数（天勤持仓表常常不带“合约乘数”，但 position_cost 已包含乘数）
            last_price = clean_num(row.get('last_price'), 0)
            multiplier = clean_num(get_field(row, '合约乘数'), 0)
            if multiplier <= 0 and last_price > 0:
                # 用 position_cost / (pos * last_price) 反推乘数（对期货/期权都适用）
                if pos_long and pos_long > 0:
                    pcl = clean_num(row.get('position_cost_long'), 0)
                    if pcl > 0:
                        multiplier = pcl / (pos_long * last_price)
                if (multiplier <= 0) and pos_short and pos_short > 0:
                    pcs = clean_num(row.get('position_cost_short'), 0)
                    if pcs > 0:
                        multiplier = pcs / (pos_short * last_price)

            # 防御：如果仍无法推断，宁可用 1 避免把损益放大 10 倍以上
            if multiplier <= 0 or (not np.isfinite(multiplier)):
                multiplier = 1.0

            # 计算开仓均价（优先级：open_price → open_cost → position_cost → last_price/期权价）
            avg_open_price = 0.0
            try:
                if net_pos > 0 and pos_long > 0:
                    open_price_long = clean_num(row.get('open_price_long'), np.nan)
                    open_cost_long = clean_num(row.get('open_cost_long'), 0)
                    pos_cost_long = clean_num(row.get('position_cost_long'), 0)
                    if np.isfinite(open_price_long) and open_price_long > 0:
                        avg_open_price = open_price_long
                    elif open_cost_long > 0:
                        avg_open_price = open_cost_long / (pos_long * multiplier)
                    elif pos_cost_long > 0:
                        avg_open_price = pos_cost_long / (pos_long * multiplier)
                elif net_pos < 0 and pos_short > 0:
                    open_price_short = clean_num(row.get('open_price_short'), np.nan)
                    open_cost_short = clean_num(row.get('open_cost_short'), 0)
                    pos_cost_short = clean_num(row.get('position_cost_short'), 0)
                    if np.isfinite(open_price_short) and open_price_short > 0:
                        avg_open_price = open_price_short
                    elif open_cost_short > 0:
                        avg_open_price = open_cost_short / (pos_short * multiplier)
                    elif pos_cost_short > 0:
                        avg_open_price = pos_cost_short / (pos_short * multiplier)
            except:
                avg_open_price = 0.0

            # 回退：用期权价（来自期权参考）或 last_price
            if avg_open_price <= 0:
                avg_open_price = clean_num(get_field(row, '期权价'), 0) or last_price

            info = {
                'contract': contract_code,
                'is_option': is_option,
                'avg_open_price': avg_open_price,
                'net_pos': net_pos,
                'direction': '多' if net_pos > 0 else '空',
                'quantity': abs(net_pos),
                'multiplier': multiplier,
            }
            
            if is_option:
                # 提取期权特有信息
                info['option_type'] = 'CALL' if 'C' in contract_code.upper() else 'PUT'

                info['strike'] = clean_num(get_field(row, '行权价'))
                info['expiry_days'] = clean_num(get_field(row, '剩余天数'))
                info['underlying_price'] = clean_num(get_field(row, '标的现价'))
                info['option_price'] = clean_num(get_field(row, '期权价'))
                
                # 期权乘数优先使用期权参考字段；若缺失则沿用上面推断的 multiplier
                mult_ref = clean_num(get_field(row, '合约乘数'), 0)
                if mult_ref > 0 and np.isfinite(mult_ref):
                    info['multiplier'] = mult_ref
                
                raw_delta = get_field(row, 'Delta')
                raw_gamma = get_field(row, 'Gamma')
                raw_theta = get_field(row, 'Theta')
                raw_vega = get_field(row, 'Vega')
                raw_rho = get_field(row, 'Rho')
                
                info['delta'] = clean_num(raw_delta)
                info['gamma'] = clean_num(raw_gamma)
                info['theta'] = clean_num(raw_theta)
                info['vega'] = clean_num(raw_vega)
                info['rho'] = clean_num(raw_rho)
                info['iv'] = clean_num(get_field(row, '隐含波动率'))
                info['moneyness'] = clean_num(get_field(row, '虚实幅度%'))
                
                # 标的合约
                info['underlying'] = get_field(row, '标的合约') or ''
                
                needs_greeks = any(pd.isna(v) for v in [raw_delta, raw_gamma, raw_theta, raw_vega, raw_rho])
                if needs_greeks:
                    S = info['underlying_price']
                    K = info['strike']
                    T = max(info['expiry_days'] / 365.0, 1e-6)
                    iv_raw = info['iv']
                    sigma = iv_raw / 100.0 if iv_raw > 3 else iv_raw
                    if sigma <= 0: sigma = 0.25
                    
                    # 使用高性能定价器补全
                    if pd.isna(raw_delta): info['delta'] = OptionPricer.delta(S, K, 0.015, sigma, T, info['option_type'])
                    if pd.isna(raw_gamma): info['gamma'] = OptionPricer.gamma(S, K, 0.015, sigma, T)
                    if pd.isna(raw_theta): info['theta'] = OptionPricer.theta(S, K, 0.015, sigma, T, info['option_type']) / 365.0
                    if pd.isna(raw_vega): info['vega'] = OptionPricer.vega(S, K, 0.015, sigma, T)
                    if pd.isna(raw_rho): info['rho'] = OptionPricer.rho(S, K, 0.015, sigma, T, info['option_type'])

            else:
                # 期货信息
                info['price'] = clean_num(row.get('last_price')) or clean_num(row.get('期货现价'))
                # 期货乘数：使用上面推断的 multiplier，避免默认10导致损益放大
                info['multiplier'] = multiplier
                info['delta'] = clean_num(net_pos)  # 期货 delta = 持仓量
                info['gamma'] = 0
                info['theta'] = 0
                info['vega'] = 0
            
            return info
            
        except Exception as e:
            return None

    def _calculate_portfolio_greeks(self, option_positions, futures_positions):
        """计算组合希腊字母"""
        greeks = {'Delta': 0, 'Gamma': 0, 'Theta': 0, 'Vega': 0, 'Rho': 0}
        
        for pos in option_positions:
            multiplier = pos.get('multiplier', 10)
            qty = pos['net_pos']
            
            greeks['Delta'] += pos.get('delta', 0) * qty * multiplier
            greeks['Gamma'] += pos.get('gamma', 0) * qty * multiplier
            greeks['Theta'] += pos.get('theta', 0) * qty * multiplier
            greeks['Vega'] += pos.get('vega', 0) * qty * multiplier
            greeks['Rho'] += pos.get('rho', 0) * qty * multiplier
        
        # 期货贡献 Delta
        for pos in futures_positions:
            multiplier = pos.get('multiplier', 10)
            greeks['Delta'] += pos['net_pos'] * multiplier
        
        return greeks

    def _identify_strategies(self, option_positions, futures_positions):
        """识别持仓策略"""
        strategies = []
        
        if not option_positions:
            if futures_positions:
                for fut in futures_positions:
                    direction = "多头" if fut['net_pos'] > 0 else "空头"
                    strategies.append({
                        'name': f"期货{direction}",
                        'type': 'directional',
                        'contracts': [fut['contract']],
                        'description': f"{fut['contract']} {direction} {abs(fut['net_pos'])}手"
                    })
            return strategies
        
        # 按标的分组
        by_underlying = {}
        for pos in option_positions:
            und = pos.get('underlying', 'unknown')
            if und not in by_underlying:
                by_underlying[und] = {'calls': [], 'puts': []}
            
            if pos['option_type'] == 'CALL':
                by_underlying[und]['calls'].append(pos)
            else:
                by_underlying[und]['puts'].append(pos)
        
        for underlying, positions in by_underlying.items():
            calls = positions['calls']
            puts = positions['puts']
            
            # 单腿策略
            for call in calls:
                if call['net_pos'] > 0:
                    strategies.append({
                        'name': '买入看涨期权',
                        'type': 'long_call',
                        'contracts': [call['contract']],
                        'description': f"买{call['quantity']}手 {call['contract']} 行权价{call['strike']}"
                    })
                else:
                    strategies.append({
                        'name': '卖出看涨期权',
                        'type': 'short_call',
                        'contracts': [call['contract']],
                        'description': f"卖{call['quantity']}手 {call['contract']} 行权价{call['strike']}"
                    })
            
            for put in puts:
                if put['net_pos'] > 0:
                    strategies.append({
                        'name': '买入看跌期权',
                        'type': 'long_put',
                        'contracts': [put['contract']],
                        'description': f"买{put['quantity']}手 {put['contract']} 行权价{put['strike']}"
                    })
                else:
                    strategies.append({
                        'name': '卖出看跌期权',
                        'type': 'short_put',
                        'contracts': [put['contract']],
                        'description': f"卖{put['quantity']}手 {put['contract']} 行权价{put['strike']}"
                    })
            
            # 检测组合策略
            combo_strategy = self._detect_combo_strategy(calls, puts, underlying)
            if combo_strategy:
                strategies.append(combo_strategy)
        
        return strategies

    def _detect_combo_strategy(self, calls, puts, underlying):
        """检测组合策略（专业版）"""
        if not calls and not puts:
            return None
        
        long_calls = [c for c in calls if c['net_pos'] > 0]
        long_puts = [p for p in puts if p['net_pos'] > 0]
        short_calls = [c for c in calls if c['net_pos'] < 0]
        short_puts = [p for p in puts if p['net_pos'] < 0]
        
        # 按行权价排序
        long_calls.sort(key=lambda x: x.get('strike', 0))
        short_calls.sort(key=lambda x: x.get('strike', 0))
        long_puts.sort(key=lambda x: x.get('strike', 0))
        short_puts.sort(key=lambda x: x.get('strike', 0))
        
        # ========== 跨式/宽跨式策略 ==========
        if long_calls and long_puts and not short_calls and not short_puts:
            call_strike = long_calls[0]['strike']
            put_strike = long_puts[0]['strike']
            premium = (long_calls[0].get('option_price', 0) + long_puts[0].get('option_price', 0))
            
            if call_strike == put_strike:
                return {
                    'name': '买入跨式 (Long Straddle)',
                    'type': 'long_straddle',
                    'contracts': [long_calls[0]['contract'], long_puts[0]['contract']],
                    'description': f"做多波动率策略，行权价{call_strike:.0f}，需波动>{premium:.1f}才盈利",
                    'greeks_profile': 'Delta中性，正Gamma，正Vega，负Theta',
                    'ideal_scenario': '标的大幅单边运动或IV大涨'
                }
            else:
                return {
                    'name': '买入宽跨式 (Long Strangle)',
                    'type': 'long_strangle',
                    'contracts': [long_calls[0]['contract'], long_puts[0]['contract']],
                    'description': f"做多波动率策略，Call@{call_strike:.0f} Put@{put_strike:.0f}",
                    'greeks_profile': 'Delta中性，正Gamma，正Vega，负Theta',
                    'ideal_scenario': '标的大幅波动突破区间'
                }
        
        # 卖出跨式/宽跨式
        if short_calls and short_puts and not long_calls and not long_puts:
            call_strike = short_calls[0]['strike']
            put_strike = short_puts[0]['strike']
            
            if call_strike == put_strike:
                return {
                    'name': '卖出跨式 (Short Straddle)',
                    'type': 'short_straddle',
                    'contracts': [short_calls[0]['contract'], short_puts[0]['contract']],
                    'description': f"做空波动率，行权价{call_strike:.0f}，预期窄幅震荡",
                    'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                    'ideal_scenario': '标的窄幅震荡且IV下降'
                }
            else:
                return {
                    'name': '卖出宽跨式 (Short Strangle)',
                    'type': 'short_strangle',
                    'contracts': [short_calls[0]['contract'], short_puts[0]['contract']],
                    'description': f"做空波动率，Call@{call_strike:.0f} Put@{put_strike:.0f}",
                    'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                    'ideal_scenario': '标的在区间内震荡'
                }
        
        # ========== 垂直价差策略 ==========
        # 牛市看涨价差
        if long_calls and short_calls and not long_puts and not short_puts:
            long_strike = min(c['strike'] for c in long_calls)
            short_strike = max(c['strike'] for c in short_calls)
            if long_strike < short_strike:
                spread_width = short_strike - long_strike
                return {
                    'name': '牛市看涨价差 (Bull Call Spread)',
                    'type': 'bull_call_spread',
                    'contracts': [long_calls[0]['contract'], short_calls[0]['contract']],
                    'description': f"温和看涨，买{long_strike:.0f}卖{short_strike:.0f}，最大收益={spread_width:.0f}",
                    'greeks_profile': '正Delta，正Gamma(弱)，低Vega，低Theta',
                    'ideal_scenario': '标的温和上涨至短腿行权价'
                }
        
        # 熊市看跌价差
        if long_puts and short_puts and not long_calls and not short_calls:
            long_strike = max(p['strike'] for p in long_puts)
            short_strike = min(p['strike'] for p in short_puts)
            if long_strike > short_strike:
                spread_width = long_strike - short_strike
                return {
                    'name': '熊市看跌价差 (Bear Put Spread)',
                    'type': 'bear_put_spread',
                    'contracts': [long_puts[0]['contract'], short_puts[0]['contract']],
                    'description': f"温和看跌，买{long_strike:.0f}卖{short_strike:.0f}，最大收益={spread_width:.0f}",
                    'greeks_profile': '负Delta，正Gamma(弱)，低Vega，低Theta',
                    'ideal_scenario': '标的温和下跌至短腿行权价'
                }
        
        # 熊市看涨价差 (Credit Call Spread)
        if short_calls and long_calls and not long_puts and not short_puts:
            short_strike = min(c['strike'] for c in short_calls)
            long_strike = max(c['strike'] for c in long_calls)
            if short_strike < long_strike:
                return {
                    'name': '熊市看涨价差 (Bear Call Spread)',
                    'type': 'bear_call_spread',
                    'contracts': [short_calls[0]['contract'], long_calls[0]['contract']],
                    'description': f"温和看跌，卖{short_strike:.0f}买{long_strike:.0f}，收取权利金",
                    'greeks_profile': '负Delta，负Gamma(弱)，负Vega，正Theta',
                    'ideal_scenario': '标的下跌或横盘'
                }
        
        # ========== 蝶式策略 ==========
        if len(calls) >= 3 or len(puts) >= 3:
            # 检测蝶式 (Butterfly)
            all_strikes_calls = sorted(set(c['strike'] for c in calls))
            all_strikes_puts = sorted(set(p['strike'] for p in puts))
            
            # Long Call Butterfly
            if len(all_strikes_calls) >= 3:
                strikes = all_strikes_calls[:3]
                if strikes[1] - strikes[0] == strikes[2] - strikes[1]:  # 等距
                    return {
                        'name': '看涨蝶式 (Call Butterfly)',
                        'type': 'call_butterfly',
                        'contracts': [c['contract'] for c in calls[:3]],
                        'description': f"精准预判策略，中心{strikes[1]:.0f}，翼{strikes[0]:.0f}-{strikes[2]:.0f}",
                        'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                        'ideal_scenario': '标的在中心行权价附近到期'
                    }
            
            if len(all_strikes_puts) >= 3:
                strikes = all_strikes_puts[:3]
                if strikes[1] - strikes[0] == strikes[2] - strikes[1]:
                    return {
                        'name': '看跌蝶式 (Put Butterfly)',
                        'type': 'put_butterfly',
                        'contracts': [p['contract'] for p in puts[:3]],
                        'description': f"精准预判策略，中心{strikes[1]:.0f}，翼{strikes[0]:.0f}-{strikes[2]:.0f}",
                        'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                        'ideal_scenario': '标的在中心行权价附近到期'
                    }
        
        # ========== 铁鹰式 (Iron Condor) ==========
        if short_calls and short_puts and long_calls and long_puts:
            short_put_strike = min(p['strike'] for p in short_puts)
            long_put_strike = min(p['strike'] for p in long_puts) if long_puts else 0
            short_call_strike = max(c['strike'] for c in short_calls)
            long_call_strike = max(c['strike'] for c in long_calls) if long_calls else 0
            
            if long_put_strike < short_put_strike < short_call_strike < long_call_strike:
                return {
                    'name': '铁鹰式 (Iron Condor)',
                    'type': 'iron_condor',
                    'contracts': [c['contract'] for c in calls + puts],
                    'description': f"区间震荡策略，看跌翼{long_put_strike:.0f}-{short_put_strike:.0f}，" +
                                  f"看涨翼{short_call_strike:.0f}-{long_call_strike:.0f}",
                    'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                    'ideal_scenario': '标的在中间区间内到期'
                }
        
        # ========== 铁蝶式 (Iron Butterfly) ==========
        if short_calls and short_puts and long_calls and long_puts:
            short_call_strike = short_calls[0]['strike']
            short_put_strike = short_puts[0]['strike']
            if short_call_strike == short_put_strike:
                return {
                    'name': '铁蝶式 (Iron Butterfly)',
                    'type': 'iron_butterfly',
                    'contracts': [c['contract'] for c in calls + puts],
                    'description': f"精准震荡策略，中心{short_call_strike:.0f}，买入两翼保护",
                    'greeks_profile': 'Delta中性，负Gamma，负Vega，正Theta',
                    'ideal_scenario': '标的精准在中心行权价到期'
                }
        
        # ========== 比率价差 (Ratio Spread) ==========
        if long_calls and short_calls and len(short_calls) > len(long_calls):
            return {
                'name': '比率看涨价差 (Call Ratio Spread)',
                'type': 'call_ratio_spread',
                'contracts': [c['contract'] for c in long_calls + short_calls],
                'description': f"买{len(long_calls)}卖{len(short_calls)}比率策略，温和看涨+收权利金",
                'greeks_profile': '低Delta，负Gamma，负Vega，正Theta',
                'ideal_scenario': '标的温和上涨但不超过卖出行权价'
            }
        
        if long_puts and short_puts and len(short_puts) > len(long_puts):
            return {
                'name': '比率看跌价差 (Put Ratio Spread)',
                'type': 'put_ratio_spread',
                'contracts': [p['contract'] for p in long_puts + short_puts],
                'description': f"买{len(long_puts)}卖{len(short_puts)}比率策略，温和看跌+收权利金",
                'greeks_profile': '低Delta，负Gamma，负Vega，正Theta',
                'ideal_scenario': '标的温和下跌但不低于卖出行权价'
            }
        
        return None

    def _calculate_breakeven(self, option_positions):
        """计算盈亏平衡点"""
        breakeven_points = []
        
        if not option_positions:
            return breakeven_points
        
        # 按标的分组计算
        by_underlying = {}
        for pos in option_positions:
            und = pos.get('underlying', 'unknown')
            if und not in by_underlying:
                by_underlying[und] = []
            by_underlying[und].append(pos)
        
        for underlying, positions in by_underlying.items():
            und_price = positions[0].get('underlying_price', 0) if positions else 0
            if und_price <= 0:
                continue
            
            # 简化计算：对于单腿期权
            for pos in positions:
                strike = pos.get('strike', 0)
                premium = pos.get('option_price', 0)
                opt_type = pos.get('option_type', '')
                direction = pos.get('net_pos', 0)
                
                if strike <= 0 or premium <= 0:
                    continue
                
                if opt_type == 'CALL':
                    if direction > 0:  # 买入看涨
                        be = strike + premium
                        breakeven_points.append({
                            'contract': pos['contract'],
                            'breakeven': be,
                            'current_price': und_price,
                            'distance_pct': (be - und_price) / und_price * 100,
                            'description': f"买入看涨: 盈亏平衡 {be:.2f} (距当前 {(be - und_price) / und_price * 100:+.2f}%)"
                        })
                    else:  # 卖出看涨
                        be = strike + premium
                        breakeven_points.append({
                            'contract': pos['contract'],
                            'breakeven': be,
                            'current_price': und_price,
                            'distance_pct': (be - und_price) / und_price * 100,
                            'description': f"卖出看涨: 盈亏平衡 {be:.2f} (距当前 {(be - und_price) / und_price * 100:+.2f}%)"
                        })
                else:  # PUT
                    if direction > 0:  # 买入看跌
                        be = strike - premium
                        breakeven_points.append({
                            'contract': pos['contract'],
                            'breakeven': be,
                            'current_price': und_price,
                            'distance_pct': (be - und_price) / und_price * 100,
                            'description': f"买入看跌: 盈亏平衡 {be:.2f} (距当前 {(be - und_price) / und_price * 100:+.2f}%)"
                        })
                    else:  # 卖出看跌
                        be = strike - premium
                        breakeven_points.append({
                            'contract': pos['contract'],
                            'breakeven': be,
                            'current_price': und_price,
                            'distance_pct': (be - und_price) / und_price * 100,
                            'description': f"卖出看跌: 盈亏平衡 {be:.2f} (距当前 {(be - und_price) / und_price * 100:+.2f}%)"
                        })
        
        return breakeven_points

    def _analyze_risk(self, option_positions, futures_positions, greeks):
        """专业风险分析与风控建议"""
        warnings = []
        controls = []
        
        # 获取持仓的标的和合约信息用于具体建议
        underlyings = set()
        atm_strikes = {}
        for pos in option_positions:
            und = pos.get('underlying', '')
            if und:
                underlyings.add(und)
                und_price = pos.get('underlying_price', 0)
                if und_price > 0:
                    atm_strikes[und] = und_price
        
        # Delta 风险分析
        delta = greeks.get('Delta', 0)
        if pd.isna(delta): delta = 0
        if abs(delta) > 300:
            direction = "多头" if delta > 0 else "空头"
            severity = "严重" if abs(delta) > 800 else "显著"
            warnings.append(f"⚠️ Delta暴露{severity}: {delta:+.0f} ({direction}偏向)")
            
            # 计算对冲手数
            hedge_lots = int(abs(delta) / 10)  # 假设每手delta为10
            
            if delta > 0:
                # 为每个标的给出具体建议
                specific_actions = []
                for und in underlyings:
                    atm = atm_strikes.get(und, 0)
                    if atm > 0:
                        # 建议买入ATM Put
                        put_strike = round(atm / 100) * 100  # 取整百
                        specific_actions.append(f"买入 {und} 行权价{put_strike} 看跌期权")
                
                controls.append({
                    'action': 'Delta对冲 (降低多头暴露)',
                    'reason': f'组合Delta +{delta:.0f}，标的下跌将导致亏损',
                    'specific': f"方案A: 卖出期货 {hedge_lots}手\n" + 
                               f"方案B: {'; '.join(specific_actions[:2]) if specific_actions else '买入ATM看跌期权'}"
                })
            else:
                specific_actions = []
                for und in underlyings:
                    atm = atm_strikes.get(und, 0)
                    if atm > 0:
                        call_strike = round(atm / 100) * 100
                        specific_actions.append(f"买入 {und} 行权价{call_strike} 看涨期权")
                
                controls.append({
                    'action': 'Delta对冲 (降低空头暴露)',
                    'reason': f'组合Delta {delta:.0f}，标的上涨将导致亏损',
                    'specific': f"方案A: 买入期货 {hedge_lots}手\n" + 
                               f"方案B: {'; '.join(specific_actions[:2]) if specific_actions else '买入ATM看涨期权'}"
                })
        
        # Gamma 风险分析
        gamma = greeks.get('Gamma', 0)
        if pd.isna(gamma): gamma = 0
        if gamma < -30:
            severity = "严重" if gamma < -100 else "显著"
            warnings.append(f"⚠️ 负Gamma风险{severity}: {gamma:.2f}，标的大幅波动将加速亏损")
            
            # 找出卖出的ATM期权
            short_atm_options = []
            for pos in option_positions:
                if pos['net_pos'] < 0 and abs(pos.get('moneyness', 0)) < 10:
                    short_atm_options.append(pos['contract'])
            
            controls.append({
                'action': '降低负Gamma风险 (Gamma Scalping)',
                'reason': f'负Gamma {gamma:.2f}，标的价格波动时亏损会加速累积',
                'specific': f"优先平仓: {', '.join(short_atm_options[:3]) if short_atm_options else 'ATM卖出期权'}\n" +
                           "或买入近月ATM期权增加正Gamma"
            })
        
        # Theta 风险分析
        theta = greeks.get('Theta', 0)
        if pd.isna(theta): theta = 0
        if theta < -50:
            daily_loss = abs(theta)
            weekly_loss = daily_loss * 5
            warnings.append(f"⚠️ 时间价值损耗: 日损{daily_loss:.0f}，周损{weekly_loss:.0f}")
            
            controls.append({
                'action': '时间价值管理 (Theta策略)',
                'reason': f'每日Theta损耗 {theta:.2f}，持仓成本较高',
                'specific': f"方案A: 转为日历价差(Calendar Spread)减少Theta\n" +
                           f"方案B: 卖出远月虚值期权收取时间价值对冲"
            })
        
        # Vega 风险分析
        vega = greeks.get('Vega', 0)
        if pd.isna(vega): vega = 0
        if abs(vega) > 150:
            direction = "正" if vega > 0 else "负"
            risk_scenario = "IV下跌" if vega > 0 else "IV上涨"
            warnings.append(f"⚠️ Vega敞口显著: {vega:+.2f} ({direction}Vega)，{risk_scenario}将带来损失")
            
            if vega > 0:
                controls.append({
                    'action': 'Vega风险管理 (波动率下跌保护)',
                    'reason': f'正Vega {vega:.0f}，若IV回落10%将亏损约 {vega * 10:.0f}',
                    'specific': "方案A: 卖出远月虚值期权对冲Vega\n" +
                               "方案B: 构建比率价差(Ratio Spread)降低Vega暴露"
                })
            else:
                controls.append({
                    'action': 'Vega风险管理 (波动率上涨保护)',
                    'reason': f'负Vega {vega:.0f}，若IV上涨10%将亏损约 {abs(vega) * 10:.0f}',
                    'specific': "方案A: 买入期权增加正Vega\n" +
                               "方案B: 减少卖出期权头寸"
                })
        
        # 临近到期风险 (Pin Risk)
        pin_risk_positions = []
        for pos in option_positions:
            days = pos.get('expiry_days', 100)
            moneyness = abs(pos.get('moneyness', 0))
            if days <= 7 and moneyness < 5:
                pin_risk_positions.append({
                    'contract': pos['contract'],
                    'days': days,
                    'moneyness': moneyness
                })
        
        if pin_risk_positions:
            contracts = [p['contract'] for p in pin_risk_positions]
            warnings.append(f"⚠️ Pin Risk警告: {', '.join(contracts)} 临近到期且接近ATM")
            
            for p in pin_risk_positions[:2]:  # 最多展示2个
                controls.append({
                    'action': f"处理Pin Risk: {p['contract']}",
                    'reason': f"还剩{p['days']}天到期，虚实幅度仅{p['moneyness']:.1f}%，Gamma极高",
                    'specific': f"立即行动: 平仓 {p['contract']} 或滚动至下月"
                })
        
        # 裸卖期权风险 (Naked Short Risk)
        naked_calls = []
        naked_puts = []
        for pos in option_positions:
            if pos['net_pos'] < 0:
                if pos['option_type'] == 'CALL':
                    naked_calls.append(pos)
                else:
                    naked_puts.append(pos)
        
        if naked_calls:
            for pos in naked_calls[:2]:
                strike = pos.get('strike', 0)
                higher_strike = strike * 1.05 if strike > 0 else 0  # 建议买入高5%行权价
                warnings.append(f"⚠️ 裸卖看涨: {pos['contract']}，上行风险无限")
                controls.append({
                    'action': f"保护裸卖看涨: {pos['contract']}",
                    'reason': '裸卖看涨期权上行风险理论无限',
                    'specific': f"方案A: 买入行权价 {higher_strike:.0f} 看涨期权构成垂直价差\n" +
                               f"方案B: 买入{pos['quantity']}手标的期货/现货对冲"
                })
        
        if naked_puts:
            for pos in naked_puts[:2]:
                strike = pos.get('strike', 0)
                lower_strike = strike * 0.95 if strike > 0 else 0
                warnings.append(f"⚠️ 裸卖看跌: {pos['contract']}，下行风险巨大")
                controls.append({
                    'action': f"保护裸卖看跌: {pos['contract']}",
                    'reason': '裸卖看跌期权下行风险可达行权价',
                    'specific': f"方案A: 买入行权价 {lower_strike:.0f} 看跌期权构成垂直价差\n" +
                               f"方案B: 确保保证金充足 (建议>行权价×乘数×150%)"
                })
        
        # 整体风险评估
        risk_score = 0
        
        # 安全读取用于评分的数值
        def safe_abs(v):
            if pd.isna(v): return 0
            return abs(v)

        s_delta = safe_abs(delta)
        s_gamma = gamma if not pd.isna(gamma) else 0
        s_vega = safe_abs(vega)

        if s_delta > 500: risk_score += 2
        elif s_delta > 300: risk_score += 1
        
        if s_gamma < -50: risk_score += 2
        elif s_gamma < -30: risk_score += 1
        
        if s_vega > 200: risk_score += 1
        if naked_calls or naked_puts: risk_score += 2
        if pin_risk_positions: risk_score += 1
        
        if risk_score == 0:
            warnings.insert(0, "✅ 整体风险评估: 低风险 - 当前持仓风险敞口在可控范围内")
        elif risk_score <= 2:
            warnings.insert(0, "⚠️ 整体风险评估: 中等风险 - 建议关注上述风险点")
        else:
            warnings.insert(0, "🚨 整体风险评估: 高风险 - 强烈建议采取风控措施")
        
        return warnings, controls

    def _analyze_trend_match(self, option_positions, strategies):
        """分析策略与市场走势的匹配度"""
        if not option_positions:
            return "无期权持仓，无法分析走势匹配"
        
        # 获取市场状态信息
        market_info = {}
        if self.market_overview_df is not None and len(self.contract_list) > 0:
            current_contract = self.contract_list[self.contract_combo.currentIndex()] if self.contract_combo.currentIndex() >= 0 else None
            if current_contract:
                underlying = current_contract.get('underlying', '')
                match_rows = self.market_overview_df[
                    self.market_overview_df['标的合约'].astype(str) == str(underlying)
                ]
                if not match_rows.empty:
                    market_info = match_rows.iloc[0].to_dict()
        
        # 分析结果
        analysis_parts = []
        
        # 判断持仓方向
        total_delta = sum(p.get('delta', 0) * p.get('net_pos', 0) for p in option_positions)
        
        if total_delta > 50:
            pos_direction = "看涨"
        elif total_delta < -50:
            pos_direction = "看跌"
        else:
            pos_direction = "中性/震荡"
        
        analysis_parts.append(f"📊 持仓方向: {pos_direction}")
        
        # 从市场信息判断走势
        futures_direction = market_info.get('期货方向', '')
        futures_status = market_info.get('期货状态', '')
        option_sentiment = market_info.get('期权情绪', '')
        
        if futures_direction:
            analysis_parts.append(f"📈 期货方向: {futures_direction}")
        if futures_status:
            analysis_parts.append(f"📉 期货状态: {futures_status}")
        if option_sentiment:
            analysis_parts.append(f"😀 期权情绪: {option_sentiment}")
        
        # 匹配度判断
        match_status = "未知"
        if pos_direction == "看涨":
            if '多' in str(futures_direction) or '涨' in str(futures_status):
                match_status = "✅ 匹配 - 持仓方向与市场趋势一致"
            elif '空' in str(futures_direction) or '跌' in str(futures_status):
                match_status = "❌ 不匹配 - 持仓看涨但市场偏空，建议减仓或对冲"
            else:
                match_status = "⚠️ 谨慎 - 市场方向不明，建议控制仓位"
        elif pos_direction == "看跌":
            if '空' in str(futures_direction) or '跌' in str(futures_status):
                match_status = "✅ 匹配 - 持仓方向与市场趋势一致"
            elif '多' in str(futures_direction) or '涨' in str(futures_status):
                match_status = "❌ 不匹配 - 持仓看跌但市场偏多，建议减仓或对冲"
            else:
                match_status = "⚠️ 谨慎 - 市场方向不明，建议控制仓位"
        else:
            if '震荡' in str(futures_status) or '中性' in str(option_sentiment):
                match_status = "✅ 匹配 - 中性持仓适合震荡市场"
            else:
                match_status = "⚠️ 谨慎 - 中性持仓在趋势市场中可能错失机会"
        
        analysis_parts.append(f"\n🎯 走势匹配: {match_status}")
        
        return '\n'.join(analysis_parts)

    def _plot_pnl_diagram_pro(self, fig, analysis):
        """绘制专业到期损益图"""
        ax = fig.add_subplot(111, facecolor='#1a1a2e')
        
        option_positions = [p for p in analysis['positions'] if p.get('is_option')]
        futures_positions = [p for p in analysis['positions'] if not p.get('is_option')]
        
        if not option_positions and not futures_positions:
            ax.text(0.5, 0.5, '无期权持仓数据', ha='center', va='center', 
                   fontsize=12, color='#a0a0a0')
            ax.axis('off')
            return
        
        # 获取标的现价
        und_price = option_positions[0].get('underlying_price', 0)
        if und_price <= 0:
            if futures_positions:
                und_price = futures_positions[0].get('price', 0)
        if und_price <= 0:
            und_price = option_positions[0].get('strike', 10000) if option_positions else 10000
        
        # 价格范围
        price_range = np.linspace(und_price * 0.75, und_price * 1.25, 200)
        
        # 计算到期损益
        total_pnl = np.zeros_like(price_range)
        
        for pos in option_positions:
            strike = pos.get('strike', 0)
            premium = pos.get('avg_open_price', 0) or pos.get('option_price', 0)
            opt_type = pos.get('option_type', 'CALL')
            qty = pos.get('net_pos', 0)
            multiplier = pos.get('multiplier', 10)
            
            if strike <= 0:
                continue
            
            if opt_type == 'CALL':
                payoff = np.maximum(price_range - strike, 0) * qty * multiplier
            else:
                payoff = np.maximum(strike - price_range, 0) * qty * multiplier
            
            cost = premium * qty * multiplier
            total_pnl += payoff - cost
        
        for pos in futures_positions:
            qty = pos.get('net_pos', 0)
            multiplier = pos.get('multiplier', 10)
            base_price = pos.get('price', 0)
            if base_price <= 0:
                base_price = und_price
            total_pnl += (price_range - base_price) * qty * multiplier
        
        # 绘制损益曲线（渐变效果）
        ax.plot(price_range, total_pnl, color='#00d4ff', linewidth=2.5, 
               label='到期损益', zorder=5)
        
        # 填充盈亏区域
        ax.fill_between(price_range, total_pnl, 0, where=(total_pnl > 0),
                        alpha=0.4, color='#dc2626', label='盈利区')
        ax.fill_between(price_range, total_pnl, 0, where=(total_pnl < 0),
                        alpha=0.4, color='#16a34a', label='亏损区')
        
        # 零轴
        ax.axhline(y=0, color='#606080', linestyle='-', linewidth=1.5, alpha=0.8)
        
        # 当前价格线
        ax.axvline(x=und_price, color='#ffd700', linestyle='--', linewidth=2, 
                  label=f'当前: {und_price:.0f}', zorder=6)
        
        # 当前价格对应的盈亏
        current_pnl_idx = np.argmin(np.abs(price_range - und_price))
        current_pnl = total_pnl[current_pnl_idx]
        pnl_color = '#dc2626' if current_pnl >= 0 else '#16a34a'
        ax.scatter([und_price], [current_pnl], color=pnl_color, s=100, zorder=7, edgecolor='white')
        ax.annotate(f'当前P/L: {current_pnl:+,.0f}', 
                   xy=(und_price, current_pnl),
                   xytext=(15, 15), textcoords='offset points',
                   fontsize=10, color=pnl_color, fontweight='bold',
                   arrowprops=dict(arrowstyle='->', color=pnl_color, lw=1.5))
        
        # 标注盈亏平衡点
        for bp in analysis['breakeven_points']:
            be = bp.get('breakeven', 0)
            if price_range[0] < be < price_range[-1]:
                ax.axvline(x=be, color='#f39c12', linestyle='-.', linewidth=1.5, alpha=0.8)
                ax.scatter([be], [0], color='#f39c12', s=80, marker='D', zorder=6)
                ax.annotate(f'BE: {be:.0f}', xy=(be, 0), 
                           xytext=(0, -25), textcoords='offset points',
                           fontsize=9, ha='center', color='#f39c12', fontweight='bold')
        
        # 最大盈亏标注
        max_profit = np.max(total_pnl)
        max_loss = np.min(total_pnl)
        max_profit_price = price_range[np.argmax(total_pnl)]
        max_loss_price = price_range[np.argmin(total_pnl)]
        
        if max_profit > 0:
            ax.annotate(f'最大盈利: {max_profit:+,.0f}', 
                       xy=(max_profit_price, max_profit),
                       xytext=(-60, 10), textcoords='offset points',
                       fontsize=9, color='#27ae60', fontweight='bold')
        
        if max_loss < 0:
            ax.annotate(f'最大亏损: {max_loss:+,.0f}', 
                       xy=(max_loss_price, max_loss),
                       xytext=(-60, -20), textcoords='offset points',
                       fontsize=9, color='#e74c3c', fontweight='bold')
        
        # 样式设置
        ax.set_xlabel('标的价格', fontsize=10, color='#c0c0c0')
        ax.set_ylabel('损益 (P/L)', fontsize=10, color='#c0c0c0')
        ax.tick_params(colors='#a0a0a0')
        ax.spines['bottom'].set_color('#404060')
        ax.spines['left'].set_color('#404060')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.legend(loc='upper left', fontsize=8, facecolor='#1a1a2e', 
                 edgecolor='#404060', labelcolor='#c0c0c0')
        ax.grid(True, alpha=0.2, color='#404060')
        
        fig.tight_layout()

    def _plot_pnl_diagram_light(self, fig, analysis):
        """
        绘制浅色风格到期损益图（优化版：基于真实开仓成本）
        
        核心修复：
        1. 使用 avg_open_price 作为成本基准
        2. 买方：到期损益 = (内在价值 - 成本价) * qty * mult
        3. 卖方：到期损益 = (成本价 - 内在价值) * |qty| * mult
        """
        ax = fig.add_subplot(111, facecolor='#ffffff')
        
        option_positions = [p for p in analysis['positions'] if p.get('is_option')]
        futures_positions = [p for p in analysis['positions'] if not p.get('is_option')]
        
        if not option_positions and not futures_positions:
            ax.text(0.5, 0.5, '无期权持仓数据', ha='center', va='center', 
                   fontsize=12, color='#64748b')
            ax.axis('off')
            return
        
        # 获取标的现价
        und_price = 0
        if option_positions:
            und_price = option_positions[0].get('underlying_price', 0)
        if und_price <= 0 and futures_positions:
            und_price = futures_positions[0].get('price', 0)
        if und_price <= 0:
            und_price = option_positions[0].get('strike', 10000) if option_positions else 10000
        
        # 价格范围：当前价±30%
        price_range = np.linspace(und_price * 0.7, und_price * 1.3, 240)
        
        # 计算到期损益（分解：期权腿/期货腿/总组合）
        pnl_option = np.zeros_like(price_range, dtype=float)
        pnl_futures = np.zeros_like(price_range, dtype=float)
        
        # ========== 期权到期损益 ==========
        for pos in option_positions:
            strike = pos.get('strike', 0)
            opt_type = pos.get('option_type', 'CALL')
            qty = pos.get('net_pos', 0)
            multiplier = pos.get('multiplier', 1) or 1
            
            if strike <= 0 or qty == 0:
                continue
            
            # 使用真实开仓成本价（买方最大亏损=权利金，公式天然满足；这里确保成本价来源正确）
            avg_open_price = pos.get('avg_open_price', 0)
            if avg_open_price <= 0:
                avg_open_price = pos.get('option_price', 0)
            if avg_open_price <= 0:
                continue
            
            # 到期内在价值
            if opt_type == 'CALL':
                intrinsic = np.maximum(price_range - strike, 0)
            else:
                intrinsic = np.maximum(strike - price_range, 0)
            
            # 损益（按张/手）
            if qty > 0:
                pos_pnl = (intrinsic - avg_open_price) * qty * multiplier
            else:
                pos_pnl = (avg_open_price - intrinsic) * abs(qty) * multiplier
            
            pnl_option += pos_pnl
        
        # ========== 期货到期损益 ==========
        for pos in futures_positions:
            qty = pos.get('net_pos', 0)
            multiplier = pos.get('multiplier', 1) or 1
            if qty == 0:
                continue
            
            entry_price = pos.get('avg_open_price', 0)
            if entry_price <= 0:
                entry_price = pos.get('price', 0)
            if entry_price <= 0:
                entry_price = und_price
            
            pnl_futures += (price_range - entry_price) * qty * multiplier
        
        total_pnl = pnl_option + pnl_futures
        
        # ========== 绘制损益曲线 ==========
        # 总组合（更浅）
        ax.plot(price_range, total_pnl, color='#1e40af', linewidth=2.2, alpha=0.85,
               label='总组合', zorder=4)
        # 期权腿（更强调，便于理解“权利金上限”）
        if option_positions:
            ax.plot(price_range, pnl_option, color='#7c3aed', linewidth=2.8,
                   label='期权腿', zorder=5)
        
        # 填充总组合盈亏区域（淡色）
        # 颜色约定：盈利=红色，亏损=绿色
        ax.fill_between(price_range, total_pnl, 0, where=(total_pnl >= 0),
                        alpha=0.12, color='#dc2626')
        ax.fill_between(price_range, total_pnl, 0, where=(total_pnl < 0),
                        alpha=0.12, color='#16a34a')
        
        # 零轴
        ax.axhline(y=0, color='#475569', linestyle='-', linewidth=1.5, alpha=0.7)
        
        # 当前价格线
        ax.axvline(x=und_price, color='#d97706', linestyle='--', linewidth=2, 
                  label=f'当前价: {und_price:.2f}', zorder=6, alpha=0.8)
        
        # 当前价格对应的盈亏（总组合 + 期权腿）
        current_pnl_idx = np.argmin(np.abs(price_range - und_price))
        current_pnl = total_pnl[current_pnl_idx]
        current_pnl_opt = pnl_option[current_pnl_idx] if 'pnl_option' in locals() else 0
        pnl_color = '#dc2626' if current_pnl >= 0 else '#16a34a'
        
        ax.scatter([und_price], [current_pnl], color=pnl_color, s=110, zorder=7,
                  edgecolor='white', linewidth=2)
        
        label = f"当前(总): {current_pnl:+,.0f}"
        if option_positions:
            label += f"\n当前(期权): {current_pnl_opt:+,.0f}"
        
        ax.annotate(label,
                   xy=(und_price, current_pnl),
                   xytext=(18, 18), textcoords='offset points',
                   fontsize=9.5, color=pnl_color, fontweight='bold',
                   bbox=dict(boxstyle='round,pad=0.5', facecolor='white',
                            edgecolor=pnl_color, alpha=0.95),
                   arrowprops=dict(arrowstyle='->', color=pnl_color, lw=2))
        
        # 标注盈亏平衡点
        breakeven_count = 0
        for bp in analysis.get('breakeven_points', []):
            be = bp.get('breakeven', 0)
            if price_range[0] < be < price_range[-1]:
                breakeven_count += 1
                ax.axvline(x=be, color='#7c3aed', linestyle='-.', linewidth=1.5, alpha=0.7)
                ax.scatter([be], [0], color='#7c3aed', s=100, marker='D', zorder=6, 
                          edgecolor='white', linewidth=2)
                y_offset = -30 if breakeven_count % 2 == 1 else 20
                ax.annotate(f'BE: {be:.2f}', xy=(be, 0), 
                           xytext=(0, y_offset), textcoords='offset points',
                           fontsize=9, ha='center', color='#7c3aed', fontweight='bold',
                           bbox=dict(boxstyle='round,pad=0.3', facecolor='white', 
                                    edgecolor='#7c3aed', alpha=0.9))
        
        # 最大盈亏标注
        max_profit = np.max(total_pnl)
        max_loss = np.min(total_pnl)
        
        if max_profit > 100:  # 只有超过100才标注
            max_profit_price = price_range[np.argmax(total_pnl)]
            ax.scatter([max_profit_price], [max_profit], color='#dc2626',
                      s=80, marker='^', zorder=6, edgecolor='white')
            ax.annotate(f'最大盈利: {max_profit:+,.0f}',
                       xy=(max_profit_price, max_profit),
                       xytext=(0, 15), textcoords='offset points',
                       fontsize=9, color='#dc2626', fontweight='bold', ha='center',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                                edgecolor='#dc2626', alpha=0.9))
        
        if max_loss < -100:  # 只有超过-100才标注
            max_loss_price = price_range[np.argmin(total_pnl)]
            ax.scatter([max_loss_price], [max_loss], color='#16a34a',
                      s=80, marker='v', zorder=6, edgecolor='white')
            ax.annotate(f'最大亏损: {max_loss:+,.0f}',
                       xy=(max_loss_price, max_loss),
                       xytext=(0, -25), textcoords='offset points',
                       fontsize=9, color='#16a34a', fontweight='bold', ha='center',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                                edgecolor='#16a34a', alpha=0.9))
        
        # 样式设置 - 专业浅色风格
        ax.set_xlabel('标的价格 (到期时)', fontsize=11, color='#334155', fontweight='bold')
        ax.set_ylabel('组合损益 (元)', fontsize=11, color='#334155', fontweight='bold')
        ax.tick_params(colors='#475569', labelsize=10)
        ax.spines['bottom'].set_color('#cbd5e1')
        ax.spines['left'].set_color('#cbd5e1')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.legend(loc='upper left', fontsize=9, framealpha=0.95,
                 edgecolor='#cbd5e1', fancybox=True)
        ax.grid(True, alpha=0.3, color='#e2e8f0', linestyle='--', axis='both')
        
        # 添加说明文字（明确口径，避免“买方亏损不可能这么大”的误解）
        note = '✓ 成本基于持仓字段推断（open/position_cost）'
        if option_positions:
            note += ' | ✓ 期权买方最大亏损=权利金（期权腿）'
        if futures_positions:
            note += ' | 含期货腿：组合损益不受权利金上限约束'
        
        ax.text(0.98, 0.02,
               note,
               transform=ax.transAxes, fontsize=8, color='#64748b',
               ha='right', va='bottom',
               bbox=dict(boxstyle='round,pad=0.4', facecolor='#f8fafc',
                        edgecolor='#e2e8f0', alpha=0.9))
        
        fig.tight_layout()

    def _plot_risk_metrics(self, fig, analysis):
        """
        收益概率与风险指标 - 专业版（GBM模拟 + 完整风险度量）
        
        改进：
        1. 使用几何布朗运动（GBM）模拟价格路径
        2. 计算到期盈亏分布
        3. 专业风险指标：VaR, CVaR, 盈利概率等
        """
        gs = fig.add_gridspec(1, 3, width_ratios=[1, 1, 1.5], wspace=0.28)
        
        option_positions = [p for p in analysis['positions'] if p.get('is_option')]
        futures_positions = [p for p in analysis['positions'] if not p.get('is_option')]
        greeks = analysis['greeks']
        
        # ========== 获取标的价格和参数 ==========
        und_price = 0
        for pos in option_positions:
            p = pos.get('underlying_price', 0)
            if p and p > 0:
                und_price = p
                break
        if und_price <= 0:
            if futures_positions:
                und_price = futures_positions[0].get('price', 0)
        if und_price <= 0:
            und_price = option_positions[0].get('strike', 10000) if option_positions else 10000
        
        # 计算加权平均到期时间和波动率
        r = 0.015
        weighted_days = []
        weighted_sigmas = []
        weighted_sizes = []
        
        for pos in option_positions:
            qty = abs(pos.get('net_pos', 0) or 0)
            mult = pos.get('multiplier', 1) or 1
            size = qty * mult
            if size <= 0: 
                continue
            days = pos.get('expiry_days', 0)
            if days and days > 0:
                weighted_days.append(days * size)
                weighted_sizes.append(size)
            iv_raw = pos.get('iv', 0) or 0
            sigma = iv_raw / 100.0 if iv_raw > 3 else iv_raw
            if sigma > 0:
                weighted_sigmas.append(sigma * size)
        
        if weighted_sizes:
            avg_days = sum(weighted_days) / sum(weighted_sizes) if weighted_days else 30
        else:
            avg_days = 30
        T = max(avg_days / 365.0, 1e-6)
        
        if weighted_sizes and weighted_sigmas:
            sigma = max(sum(weighted_sigmas) / sum(weighted_sizes), 0.05)
        else:
            sigma = 0.25
        
        # ========== 使用GBM模拟价格路径 ==========
        n_paths = 5000  # 模拟路径数
        n_steps = max(int(avg_days), 1)  # 步数 = 天数
        
        # GBM模拟
        price_paths, time_points = PriceSimulator.simulate_gbm(
            S0=und_price, 
            mu=r,  # 风险中性漂移
            sigma=sigma, 
            T=T, 
            n_paths=n_paths, 
            n_steps=n_steps
        )
        
        # 计算各路径的到期盈亏
        all_positions = option_positions + futures_positions
        terminal_pnl = PriceSimulator.calculate_terminal_pnl(
            price_paths, 
            all_positions, 
            r=r
        )
        
        # ========== 数据验证和清理 ==========
        # 检查是否有有效数据
        valid_pnl = terminal_pnl[np.isfinite(terminal_pnl)]
        
        if len(valid_pnl) == 0 or len(all_positions) == 0:
            # 如果没有有效数据，显示空数据提示
            ax1 = fig.add_subplot(gs[0], facecolor='#ffffff')
            ax1.text(0.5, 0.5, '持仓数据不完整\n无法进行风险分析', 
                    ha='center', va='center', fontsize=12, color='#64748b',
                    transform=ax1.transAxes)
            ax1.axis('off')
            
            ax2 = fig.add_subplot(gs[1], facecolor='#ffffff')
            ax2.text(0.5, 0.5, '请检查:\n1. 持仓成本数据\n2. 开仓价格数据\n3. 行权价数据', 
                    ha='center', va='center', fontsize=10, color='#dc2626',
                    transform=ax2.transAxes)
            ax2.axis('off')
            
            ax3 = fig.add_subplot(gs[2], facecolor='#ffffff')
            ax3.text(0.5, 0.5, '需要完整的持仓数据\n才能进行GBM模拟', 
                    ha='center', va='center', fontsize=12, color='#64748b',
                    transform=ax3.transAxes)
            ax3.axis('off')
            
            fig.tight_layout()
            return
        
        # 使用有效数据
        terminal_pnl = valid_pnl

        # 同时计算“期权腿”单独的到期盈亏分布（更符合期权最佳实践口径）
        terminal_pnl_opt = PriceSimulator.calculate_terminal_pnl(price_paths, option_positions, r=r)
        terminal_pnl_opt = terminal_pnl_opt[np.isfinite(terminal_pnl_opt)]
        if len(terminal_pnl_opt) == 0:
            terminal_pnl_opt = None

        # ========== 计算风险指标（总组合） ==========
        profit_prob = (terminal_pnl > 0).sum() / len(terminal_pnl) * 100
        loss_prob = 100 - profit_prob

        expected_pnl = terminal_pnl.mean()
        var_95 = RiskMetrics.calculate_var(terminal_pnl, confidence=0.95)
        cvar_95 = RiskMetrics.calculate_cvar(terminal_pnl, confidence=0.95)

        # ========== 计算风险指标（期权腿） ==========
        profit_prob_opt = None
        expected_pnl_opt = None
        var_95_opt = None
        cvar_95_opt = None
        if terminal_pnl_opt is not None:
            profit_prob_opt = (terminal_pnl_opt > 0).sum() / len(terminal_pnl_opt) * 100
            expected_pnl_opt = terminal_pnl_opt.mean()
            var_95_opt = RiskMetrics.calculate_var(terminal_pnl_opt, confidence=0.95)
            cvar_95_opt = RiskMetrics.calculate_cvar(terminal_pnl_opt, confidence=0.95)
        
        max_profit = terminal_pnl.max()
        max_loss = terminal_pnl.min()
        
        # 当前盈亏（基于当前价格）
        current_pnl = 0
        total_cost = 0
        
        for pos in option_positions:
            strike = pos.get('strike', 0)
            opt_type = pos.get('option_type', 'CALL')
            qty = pos.get('net_pos', 0) or 0
            mult = pos.get('multiplier', 10) or 10
            avg_open_price = pos.get('avg_open_price', 0)
            
            if avg_open_price <= 0:
                avg_open_price = pos.get('option_price', 0) or 0
            
            if strike <= 0 or avg_open_price <= 0 or mult <= 0 or qty == 0:
                continue
            
            try:
                # 当前期权价值（使用BS模型）
                current_option_price = OptionPricer.bs_price(
                    und_price, strike, r, sigma, T, opt_type
                )
                
                # 检查定价结果是否有效
                if np.isnan(current_option_price) or current_option_price < 0:
                    continue
                
                # 当前盈亏
                if qty > 0:  # 买方
                    current_pnl += (current_option_price - avg_open_price) * qty * mult
                else:  # 卖方
                    current_pnl += (avg_open_price - current_option_price) * abs(qty) * mult
                
                # 总成本
                total_cost += avg_open_price * qty * mult
            except Exception as e:
                continue
        
        for pos in futures_positions:
            qty = pos.get('net_pos', 0) or 0
            mult = pos.get('multiplier', 1) or 1
            entry_price = pos.get('avg_open_price', 0) or pos.get('price', 0)
            
            if entry_price > 0 and mult > 0 and qty != 0:
                try:
                    current_pnl += (und_price - entry_price) * qty * mult
                except Exception as e:
                    continue
        
        # ========== 图1: 盈利概率（饼图）==========
        ax1 = fig.add_subplot(gs[0], facecolor='#ffffff')
        
        # 颜色约定：盈利=红色，亏损=绿色
        colors = ['#dc2626', '#16a34a']
        sizes = [max(profit_prob, 1), max(loss_prob, 1)]  # 避免0值
        
        wedges, texts = ax1.pie(sizes, colors=colors, startangle=90,
                                wedgeprops=dict(width=0.6, edgecolor='white', linewidth=2))
        
        ax1.text(0, 0, f'{profit_prob:.1f}%', ha='center', va='center', fontsize=14,
                color='#dc2626' if profit_prob >= 50 else '#16a34a', fontweight='bold')
        
        # 期权腿盈利概率（如果可得）
        if profit_prob_opt is not None:
            ax1.text(0, -0.25, f'期权腿: {profit_prob_opt:.1f}%', ha='center', va='center',
                    fontsize=9, color='#7c3aed', fontweight='bold')
        
        ax1.set_title('盈利概率（GBM模拟）', fontsize=10, color='#334155', pad=8, fontweight='bold')
        
        # 添加说明
        ax1.text(0, -1.3, f'基于{n_paths:,}条价格路径模拟', 
                ha='center', va='top', fontsize=7, color='#64748b',
                transform=ax1.transData)
        
        # ========== 图2: 关键指标 ==========
        ax2 = fig.add_subplot(gs[1], facecolor='#ffffff')
        ax2.axis('off')
        
        # 判断是否有限/无限
        price_range_low = price_paths[:, -1].min()
        price_range_high = price_paths[:, -1].max()
        
        # 检查在极端价格下的盈亏是否趋于最大值
        extreme_low_pnl = PriceSimulator.calculate_terminal_pnl(
            np.array([[price_range_low * 0.5]]), all_positions, r
        )[0]
        extreme_high_pnl = PriceSimulator.calculate_terminal_pnl(
            np.array([[price_range_high * 2.0]]), all_positions, r
        )[0]
        
        # 如果极端情况下盈亏仍在增长，则为无限
        if abs(extreme_high_pnl - max_profit) > abs(max_profit) * 0.2 and max_profit > 0:
            max_profit_text = "↑无限"
        else:
            max_profit_text = f"+{max_profit:,.0f}"
        
        if abs(extreme_low_pnl - max_loss) > abs(max_loss) * 0.2 and max_loss < 0:
            max_loss_text = "↓无限"
        else:
            max_loss_text = f"{max_loss:,.0f}"
        
        prob_text = f"{profit_prob:.1f}%"
        if profit_prob_opt is not None:
            prob_text += f"（期权{profit_prob_opt:.1f}%）"

        exp_text = f"{expected_pnl:+,.0f}"
        if expected_pnl_opt is not None:
            exp_text += f"（期权{expected_pnl_opt:+,.0f}）"

        var_text = f"{var_95:+,.0f}"
        if var_95_opt is not None:
            var_text += f"（期权{var_95_opt:+,.0f}）"

        cvar_text = f"{cvar_95:+,.0f}"
        if cvar_95_opt is not None:
            cvar_text += f"（期权{cvar_95_opt:+,.0f}）"

        # 颜色约定：盈利=红色，亏损=绿色
        metrics = [
            ('盈利概率', prob_text, '#1a5fb4'),
            ('最大盈利', max_profit_text, '#dc2626'),
            ('最大亏损', max_loss_text, '#16a34a'),
            ('当前盈亏', f"{current_pnl:+,.0f}", '#dc2626' if current_pnl >= 0 else '#16a34a'),
            ('期望收益', exp_text, '#dc2626' if expected_pnl >= 0 else '#16a34a'),
            ('VaR(95%)', var_text, '#16a34a' if var_95 <= 0 else '#dc2626'),
            ('CVaR(95%)', cvar_text, '#16a34a' if cvar_95 <= 0 else '#dc2626'),
        ]
        
        ys = np.linspace(0.90, 0.05, len(metrics))
        for y, (label, value, color) in zip(ys, metrics):
            ax2.text(0.05, y, label, fontsize=9, color='#64748b', 
                    transform=ax2.transAxes, va='center')
            ax2.text(0.95, y, value, fontsize=10, color=color, fontweight='bold',
                    transform=ax2.transAxes, va='center', ha='right')
        
        ax2.set_title('风险指标', fontsize=10, color='#334155', pad=10, fontweight='bold')
        
        # ========== 图3: 盈亏分布直方图 ==========
        ax3 = fig.add_subplot(gs[2], facecolor='#ffffff')
        
        # 计算展示范围：优先展示“主要分布”，避免横坐标被极端尾部拉爆导致柱子挤成一团
        # 说明：风险指标仍使用全量 terminal_pnl；这里只是直方图的显示范围裁剪（winsorize视图）。
        view_range_label = 'P1~P99'
        in_view_ratio = 1.0
        try:
            p1, p99 = np.percentile(terminal_pnl, [1, 99])
            p5, p95 = np.percentile(terminal_pnl, [5, 95])
            p25, p75 = np.percentile(terminal_pnl, [25, 75])
            iqr = float(p75 - p25)

            # 如果尾部特别长，用 P5~P95 聚焦主体；否则用 P1~P99 保留更多信息
            denom = float(p95 - p5)
            tail_ratio = float((p99 - p1) / denom) if denom > 0 else float('inf')
            if np.isfinite(tail_ratio) and tail_ratio > 3.0:
                pnl_min, pnl_max = float(p5), float(p95)
                view_range_label = 'P5~P95'
            else:
                pnl_min, pnl_max = float(p1), float(p99)
                view_range_label = 'P1~P99'

            # 再用 IQR 约束一下，避免极端情况下分位数仍然太宽
            if np.isfinite(iqr) and iqr > 0:
                iqr_min = float(p25 - 3.0 * iqr)
                iqr_max = float(p75 + 3.0 * iqr)
                pnl_min = max(pnl_min, iqr_min)
                pnl_max = min(pnl_max, iqr_max)

            mask_view = (terminal_pnl >= pnl_min) & (terminal_pnl <= pnl_max)
            in_view_ratio = float(np.mean(mask_view)) if len(terminal_pnl) else 1.0
        except Exception:
            pnl_min, pnl_max = float(np.min(terminal_pnl)), float(np.max(terminal_pnl))
            view_range_label = 'MIN~MAX'
            in_view_ratio = 1.0

        if (not np.isfinite(pnl_min)) or (not np.isfinite(pnl_max)) or pnl_min == pnl_max:
            pnl_min, pnl_max = float(np.min(terminal_pnl)), float(np.max(terminal_pnl))
            view_range_label = 'MIN~MAX'

        pad = (pnl_max - pnl_min) * 0.08 if pnl_max > pnl_min else max(abs(pnl_min), abs(pnl_max), 100) * 0.1
        pnl_min -= pad
        pnl_max += pad

        # 绘制盈亏分布（bins 也尽量跟着主体分布走）
        n = len(terminal_pnl)
        n_bins = min(60, max(12, n // 30))
        try:
            # Freedman–Diaconis: bin_width = 2*IQR*n^(-1/3)
            p25, p75 = np.percentile(terminal_pnl, [25, 75])
            iqr = float(p75 - p25)
            if np.isfinite(iqr) and iqr > 0 and n > 10:
                bin_w = 2.0 * iqr / (n ** (1/3))
                if np.isfinite(bin_w) and bin_w > 0:
                    est_bins = int(max(10, min(80, (pnl_max - pnl_min) / bin_w)))
                    n_bins = max(12, min(60, est_bins))
        except Exception:
            pass
        
        try:
            counts, bins, patches = ax3.hist(
                terminal_pnl, 
                bins=n_bins, 
                range=(pnl_min, pnl_max),
                alpha=0.7, 
                edgecolor='white', 
                linewidth=0.5
            )
            
            # 根据盈亏着色（盈利=红色，亏损=绿色）
            for i, patch in enumerate(patches):
                bin_center = (bins[i] + bins[i+1]) / 2
                if bin_center >= 0:
                    patch.set_facecolor('#dc2626')
                else:
                    patch.set_facecolor('#16a34a')

            # 固定x轴范围，避免分布被挤到边缘
            ax3.set_xlim(pnl_min, pnl_max)
        except Exception as e:
            # 如果直方图绘制失败，显示简化版本
            ax3.text(0.5, 0.5, f'盈亏分布\n期望: {expected_pnl:+,.0f}\nVaR: {var_95:+,.0f}', 
                    ha='center', va='center', fontsize=11, color='#334155',
                    transform=ax3.transAxes)
            ax3.axis('off')
            fig.tight_layout()
            return
        
        # 标注重要位置
        ymax = counts.max() if len(counts) > 0 else 1
        
        # 期望值线
        if pnl_min <= expected_pnl <= pnl_max:
            ax3.axvline(x=expected_pnl, color='#1a5fb4', linestyle='--', linewidth=2, 
                       label=f'期望: {expected_pnl:+,.0f}', alpha=0.8)
        
        # VaR线
        if pnl_min <= var_95 <= pnl_max:
            ax3.axvline(x=var_95, color='#dc2626', linestyle='-.', linewidth=2, 
                       label=f'VaR(95%): {var_95:+,.0f}', alpha=0.8)
        
        # 当前盈亏线
        if pnl_min <= current_pnl <= pnl_max:
            ax3.axvline(x=current_pnl, color='#d97706', linestyle='-', linewidth=2.5, 
                       label=f'当前: {current_pnl:+,.0f}', alpha=0.9)
        
        # 零线
        if pnl_min <= 0 <= pnl_max:
            ax3.axvline(x=0, color='#475569', linestyle='-', linewidth=1.5, alpha=0.5)
        
        ax3.set_xlabel('到期盈亏 (元)', fontsize=10, color='#334155', fontweight='bold')
        ax3.set_ylabel('频数', fontsize=10, color='#334155', fontweight='bold')
        ax3.set_title('到期盈亏分布', fontsize=10, color='#334155', pad=8, fontweight='bold')
        ax3.tick_params(colors='#475569', labelsize=9)
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        ax3.spines['bottom'].set_color('#cbd5e1')
        ax3.spines['left'].set_color('#cbd5e1')
        ax3.grid(True, alpha=0.3, color='#e2e8f0', axis='y')
        
        # 添加统计信息
        clipped_pct = max(0.0, min(100.0, (1.0 - in_view_ratio) * 100.0))
        stats_text = (
            f'模拟路径: {n_paths:,}\n'
            f'到期天数: {avg_days:.0f}\n'
            f'波动率: {sigma*100:.1f}%\n'
            f'显示范围: {view_range_label}（截断尾部 {clipped_pct:.1f}%）'
        )
        ax3.text(0.02, 0.98, stats_text, 
                transform=ax3.transAxes, fontsize=7, color='#64748b',
                va='top', ha='left',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='#f8fafc', 
                         edgecolor='#e2e8f0', alpha=0.9))
        
        fig.tight_layout()

    def _plot_sensitivity_matrix_light(self, fig, analysis):
        """
        情景敏感度分析 - 专业版（基于持仓价值变动）
        
        核心逻辑：
        1. 计算当前持仓价值和盈亏
        2. 计算各场景下的持仓价值和盈亏
        3. 敏感度 = 场景盈亏 - 当前盈亏
        4. 买方最大亏损 = 已支付权利金
        """
        ax = fig.add_subplot(111, facecolor='#ffffff')
        
        option_positions = [p for p in analysis['positions'] if p.get('is_option')]
        futures_positions = [p for p in analysis['positions'] if not p.get('is_option')]
        
        # 场景定义
        scenarios = ['价格+5%', '价格-5%', 'IV+5%', 'IV-5%', '时间-7天', '综合压力']
        # 分解口径：期权腿 vs 期货腿（避免误解“买方亏损为什么能超过权利金”）
        impacts_opt = [0.0] * 6
        impacts_fut = [0.0] * 6
        
        r = 0.015  # 无风险利率
        
        # 1. 计算期权部分的影响（基于持仓价值变动）
        for pos in option_positions:
            qty = pos.get('net_pos', 0)
            mult = pos.get('multiplier', 1) or 1
            S = pos.get('underlying_price', 0)
            K = pos.get('strike', 0)
            T = max(pos.get('expiry_days', 0) / 365.0, 0.0001)
            
            # IV转换
            iv_raw = pos.get('iv', 0) or 0
            sigma = iv_raw / 100.0 if iv_raw > 3 else iv_raw
            if sigma <= 0: sigma = 0.25
            otype = pos.get('option_type', 'CALL')
            
            # 获取真实开仓成本
            avg_open_price = pos.get('avg_open_price', 0)
            if avg_open_price <= 0:
                avg_open_price = pos.get('option_price', 0)
                if avg_open_price <= 0:
                    avg_open_price = OptionPricer.bs_price(S, K, r, sigma, T, otype)
            
            # 总权利金（持仓成本）
            total_premium = avg_open_price * abs(qty) * mult
            
            # 当前持仓价值
            current_price = OptionPricer.bs_price(S, K, r, sigma, T, otype)
            current_value = current_price * abs(qty) * mult
            
            # 当前盈亏
            if qty > 0:  # 买方
                current_pnl = current_value - total_premium
            else:  # 卖方
                current_pnl = total_premium - current_value
            
            # 各场景定义
            scenario_params = [
                (S*1.05, sigma, T),           # (1) 价格+5%
                (S*0.95, sigma, T),           # (2) 价格-5%
                (S, sigma+0.05, T),           # (3) IV+5%
                (S, max(0.001, sigma-0.05), T), # (4) IV-5%
                (S, sigma, max(1e-6, T - 7/365.0)), # (5) 时间-7天
                (S*0.95, sigma+0.05, max(1e-6, T - 7/365.0)), # (6) 综合压力
            ]
            
            # 计算各场景的盈亏变动
            for i, (S_new, sigma_new, T_new) in enumerate(scenario_params):
                # 场景下的持仓价值
                scenario_price = OptionPricer.bs_price(S_new, K, r, sigma_new, T_new, otype)
                scenario_value = scenario_price * abs(qty) * mult
                
                # 场景下的盈亏
                if qty > 0:  # 买方
                    scenario_pnl = scenario_value - total_premium
                    # 限制：买方最大亏损 = 权利金
                    scenario_pnl = max(scenario_pnl, -total_premium)
                else:  # 卖方
                    scenario_pnl = total_premium - scenario_value
                    # 限制：卖方最大盈利 = 权利金
                    scenario_pnl = min(scenario_pnl, total_premium)
                
                # 盈亏变动 = 场景盈亏 - 当前盈亏
                delta_pnl = scenario_pnl - current_pnl
                impacts_opt[i] += delta_pnl
        
        # 2. 计算期货部分的影响（Delta线性）
        for pos in futures_positions:
            qty = pos.get('net_pos', 0)
            mult = pos.get('multiplier', 1) or 1
            S = pos.get('price', 0)
            
            if S > 0:
                p_change = S * 0.05 * qty * mult
                impacts_fut[0] += p_change   # Price +5%
                impacts_fut[1] -= p_change   # Price -5%
                impacts_fut[5] -= p_change   # Combined

        # 合计
        impacts_total = [a + b for a, b in zip(impacts_opt, impacts_fut)]

        def _bar_colors(vals):
            cs = []
            for v in vals:
                if v > 0:
                    cs.append('#dc2626')  # 盈利=红色
                elif v < 0:
                    cs.append('#16a34a')  # 亏损=绿色
                else:
                    cs.append('#94a3b8')
            return cs

        # 绘制分组柱状图（期权腿 vs 期货腿），并在上方标注合计
        x_pos = np.arange(len(scenarios))
        width = 0.34

        bars_opt = ax.bar(x_pos - width/2, impacts_opt, width=width,
                          color=_bar_colors(impacts_opt),
                          edgecolor='#e2e8f0', linewidth=1.2, label='期权腿')
        bars_fut = ax.bar(x_pos + width/2, impacts_fut, width=width,
                          color=_bar_colors(impacts_fut),
                          edgecolor='#e2e8f0', linewidth=1.2, label='期货腿')

        # 数值标签（期权/期货）
        def _annotate_bars(bars, values):
            for bar, val in zip(bars, values):
                h = bar.get_height()
                c = '#dc2626' if val >= 0 else '#16a34a'
                off = 6 if h >= 0 else -10
                va = 'bottom' if h >= 0 else 'top'
                ax.annotate(f'{val:+,.0f}',
                            xy=(bar.get_x() + bar.get_width() / 2, h),
                            xytext=(0, off), textcoords='offset points',
                            ha='center', va=va, fontsize=9, fontweight='bold', color=c)

        _annotate_bars(bars_opt, impacts_opt)
        _annotate_bars(bars_fut, impacts_fut)

        # 合计标签
        for i, tot in enumerate(impacts_total):
            top = max(impacts_opt[i], impacts_fut[i], tot)
            ax.annotate(f'合计 {tot:+,.0f}',
                        xy=(x_pos[i], top),
                        xytext=(0, 18 if top >= 0 else -22),
                        textcoords='offset points',
                        ha='center',
                        va='bottom' if top >= 0 else 'top',
                        fontsize=8.5, color='#334155',
                        bbox=dict(boxstyle='round,pad=0.25', facecolor='#f8fafc',
                                  edgecolor='#e2e8f0', alpha=0.9))
        
        # 零轴
        ax.axhline(y=0, color='#1e293b', linewidth=2)
        
        # 样式优化
        ax.set_xticks(x_pos)
        ax.set_xticklabels(scenarios, fontsize=10, color='#334155', rotation=0, ha='center')
        ax.legend(loc='upper left', fontsize=9, framealpha=0.95, edgecolor='#e2e8f0', fancybox=True)
        ax.tick_params(axis='y', colors='#475569', labelsize=10)
        ax.spines['bottom'].set_color('#cbd5e1')
        ax.spines['left'].set_color('#cbd5e1')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.3, color='#e2e8f0', axis='y', linestyle='--')
        ax.set_ylabel('预估盈亏变动 (元)', fontsize=11, color='#475569', fontweight='bold')

        # 口径说明（最佳实践：明确区分期权腿与期货腿）
        ax.text(0.01, 0.98,
                '说明：期权腿按BS重定价，并对买方应用“最大亏损=权利金”；\n期货腿按线性价格变动计算（不受权利金上限约束）。',
                transform=ax.transAxes, fontsize=8, color='#64748b',
                va='top', ha='left',
                bbox=dict(boxstyle='round,pad=0.35', facecolor='#f8fafc',
                          edgecolor='#e2e8f0', alpha=0.9))
        
        fig.tight_layout()

    def _plot_greeks_radar(self, fig, analysis):
        """绘制Greeks雷达图"""
        ax = fig.add_subplot(111, polar=True, facecolor='#1a1a2e')
        
        greeks = analysis['greeks']
        labels = ['Delta', 'Gamma', 'Theta', 'Vega', 'Rho']
        
        # 归一化处理（用于雷达图显示）
        values = []
        max_vals = {'Delta': 500, 'Gamma': 100, 'Theta': 200, 'Vega': 300, 'Rho': 100}
        for label in labels:
            val = greeks.get(label, 0)
            if pd.isna(val): val = 0
            normalized = min(abs(val) / max_vals[label], 1.0) * 100
            values.append(normalized)
        
        # 添加首尾闭合
        values_closed = values + values[:1]
        angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
        angles_closed = angles + angles[:1]
        
        # 绘制雷达图
        ax.plot(angles_closed, values_closed, 'o-', linewidth=2, 
               color='#00d4ff', label='当前敞口')
        ax.fill(angles_closed, values_closed, alpha=0.25, color='#00d4ff')
        
        # 设置标签
        ax.set_xticks(angles)
        greek_labels = []
        for i, label in enumerate(labels):
            val = greeks.get(label, 0)
            sign = '+' if val >= 0 else ''
            greek_labels.append(f'{label}\n({sign}{val:.0f})')
        ax.set_xticklabels(greek_labels, fontsize=9, color='#e0e0e0')
        
        # 样式
        ax.set_ylim(0, 100)
        ax.set_yticks([25, 50, 75, 100])
        ax.set_yticklabels(['25%', '50%', '75%', '100%'], fontsize=7, color='#808080')
        ax.spines['polar'].set_color('#404060')
        ax.grid(True, color='#404060', alpha=0.3)
        
        # 标题
        ax.set_title('风险敞口雷达', fontsize=10, color='#ffd700', pad=10)
        
        fig.tight_layout()

    def _plot_greeks_bar_pro(self, fig, analysis):
        """绘制专业Greeks柱状图"""
        ax = fig.add_subplot(111, facecolor='#1a1a2e')
        
        greeks = analysis['greeks']
        labels = ['Δ Delta', 'Γ Gamma', 'Θ Theta', 'ν Vega', 'ρ Rho']
        keys = ['Delta', 'Gamma', 'Theta', 'Vega', 'Rho']
        values = [greeks.get(k, 0) for k in keys]
        
        # 颜色
        colors = ['#27ae60' if v > 0 else '#e74c3c' if v < 0 else '#606080' for v in values]
        
        # 绘制水平柱状图
        y_pos = np.arange(len(labels))
        bars = ax.barh(y_pos, values, color=colors, height=0.6, 
                      edgecolor='#404060', linewidth=1)
        
        # 添加数值标签
        for i, (bar, val) in enumerate(zip(bars, values)):
            width = bar.get_width()
            label_x = width + (max(abs(v) for v in values) * 0.05) if width >= 0 else width - (max(abs(v) for v in values) * 0.15)
            color = '#27ae60' if val >= 0 else '#e74c3c'
            ax.annotate(f'{val:+.1f}',
                       xy=(width, bar.get_y() + bar.get_height() / 2),
                       xytext=(5 if width >= 0 else -5, 0),
                       textcoords="offset points",
                       ha='left' if width >= 0 else 'right', va='center',
                       fontsize=10, fontweight='bold', color=color)
        
        # 零轴
        ax.axvline(x=0, color='#ffd700', linewidth=1.5)
        
        # 样式
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=10, color='#e0e0e0')
        ax.tick_params(axis='x', colors='#a0a0a0')
        ax.spines['bottom'].set_color('#404060')
        ax.spines['left'].set_color('#404060')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.2, color='#404060', axis='x')
        ax.set_title('Greeks净敞口', fontsize=10, color='#ffd700', pad=8)
        
        fig.tight_layout()

    def _plot_sensitivity_matrix(self, fig, analysis):
        """绘制风险敏感度矩阵"""
        ax = fig.add_subplot(111, facecolor='#1a1a2e')
        
        greeks = analysis['greeks']
        
        # 构建敏感度矩阵
        scenarios = ['价格↑5%', '价格↓5%', 'IV↑10%', 'IV↓10%', '时间-7天']
        impacts = []
        
        delta = greeks.get('Delta', 0)
        gamma = greeks.get('Gamma', 0)
        vega = greeks.get('Vega', 0)
        theta = greeks.get('Theta', 0)
        
        # 处理可能的 NaN
        delta = 0 if pd.isna(delta) else delta
        gamma = 0 if pd.isna(gamma) else gamma
        vega = 0 if pd.isna(vega) else vega
        theta = 0 if pd.isna(theta) else theta
        
        # 估算各场景影响
        option_positions = [p for p in analysis['positions'] if p.get('is_option')]
        avg_price = np.mean([p.get('underlying_price', 10000) for p in option_positions]) if option_positions else 10000
        
        # 价格上涨5%
        price_up = delta * (avg_price * 0.05) + 0.5 * gamma * (avg_price * 0.05) ** 2
        impacts.append(price_up)
        
        # 价格下跌5%
        price_down = delta * (-avg_price * 0.05) + 0.5 * gamma * (-avg_price * 0.05) ** 2
        impacts.append(price_down)
        
        # IV上涨10%
        iv_up = vega * 10
        impacts.append(iv_up)
        
        # IV下跌10%
        iv_down = vega * (-10)
        impacts.append(iv_down)
        
        # 时间流逝7天
        time_decay = theta * 7
        impacts.append(time_decay)
        
        # 颜色映射
        colors = ['#27ae60' if v > 0 else '#e74c3c' if v < 0 else '#606080' for v in impacts]
        
        # 绘制
        x_pos = np.arange(len(scenarios))
        bars = ax.bar(x_pos, impacts, color=colors, width=0.6, 
                     edgecolor='#404060', linewidth=1)
        
        # 数值标签
        for bar, val in zip(bars, impacts):
            height = bar.get_height()
            color = '#27ae60' if val >= 0 else '#e74c3c'
            ax.annotate(f'{val:+,.0f}',
                       xy=(bar.get_x() + bar.get_width() / 2, height),
                       xytext=(0, 5 if height >= 0 else -15),
                       textcoords="offset points",
                       ha='center', va='bottom' if height >= 0 else 'top',
                       fontsize=9, fontweight='bold', color=color)
        
        # 零轴
        ax.axhline(y=0, color='#ffd700', linewidth=1.5)
        
        # 样式
        ax.set_xticks(x_pos)
        ax.set_xticklabels(scenarios, fontsize=9, color='#e0e0e0', rotation=15)
        ax.tick_params(axis='y', colors='#a0a0a0')
        ax.spines['bottom'].set_color('#404060')
        ax.spines['left'].set_color('#404060')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.2, color='#404060', axis='y')
        ax.set_ylabel('预估P/L变化', fontsize=9, color='#c0c0c0')
        
        fig.tight_layout()

    def _plot_pnl_diagram(self, fig, analysis):
        """绘制到期损益图（兼容旧版）"""
        self._plot_pnl_diagram_pro(fig, analysis)

    def _plot_greeks_bar(self, fig, analysis):
        """绘制希腊字母柱状图（兼容旧版）"""
        self._plot_greeks_bar_pro(fig, analysis)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = OptionTShapeWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
